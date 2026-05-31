"""汇总导出：SKU → 单个 xlsx（产品概览+竞品对比+图片索引+市场情报）"""
import json, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

HDR_FONT = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
HDR_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
TITLE_FONT = Font(name='Microsoft YaHei', size=16, bold=True, color='1F3864')
SEC_FONT = Font(name='Microsoft YaHei', size=12, bold=True, color='2F5496')
VAL_FONT = Font(name='Microsoft YaHei', size=10)
ORANGE_FILL = 'FFD966'
NEW_FILL = 'E2EFDA'  # light green tint for "new" indicator in legend
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
WRAP = Alignment(vertical='top', wrap_text=True)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
NEW_BADGE_FILL = PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid')  # orange for new


def build_summary_xlsx(mother_dir: str, sku: str):
    d = Path(mother_dir)

    intel = json.loads((d / 'market_intel.json').read_text(encoding='utf-8'))
    listing = json.loads((d / 'listing_output.json').read_text(encoding='utf-8'))
    img_file = d / '生成图片' / 'image_batch.json'
    images = json.loads(img_file.read_text(encoding='utf-8')) if img_file.exists() else {}
    val_file = d / '生成图片' / 'image_validation.json'
    validation = json.loads(val_file.read_text(encoding='utf-8')) if val_file.exists() else {}

    # Load new competitor IDs for highlighting
    new_ids_file = d / 'new_ids.json'
    new_ids = set(json.loads(new_ids_file.read_text(encoding='utf-8'))) if new_ids_file.exists() else set()

    wb = Workbook()

    # ====== Sheet 1: Product Overview ======
    ws = wb.active
    ws.title = '产品概览'
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 70

    ws.merge_cells('A1:B1')
    ws.cell(1, 1, listing['title']).font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    meta = [
        ('SKU', listing['sku']),
        ('Price', f"${listing['price']} (range: {listing['price_range']})"),
        ('Position', listing['product_info']['position']),
        ('Material', listing['product_info']['material']),
        ('Finish', listing['product_info']['finish']),
        ('Count', str(listing['product_info']['count'])),
        ('Condition', listing['product_info']['condition']),
        ('Warranty', listing['product_info']['warranty']),
        ('Origin', listing['product_info']['country_of_origin']),
        ('Valid Competitors', str(listing['product_info']['valid_competitors'])),
        ('Rule-0 Excluded', str(listing['product_info']['rule0_excluded'])),
    ]
    for i, (k, v) in enumerate(meta, 3):
        ws.cell(i, 1, k).font = Font(name='Microsoft YaHei', size=10, bold=True)
        ws.cell(i, 1).border = THIN
        ws.cell(i, 2, str(v)).font = VAL_FONT
        ws.cell(i, 2).border = THIN

    ws.cell(16, 1, 'Bullet Points').font = SEC_FONT
    for j, b in enumerate(listing['bullets'], 17):
        ws.merge_cells(f'A{j}:B{j}')
        ws.cell(j, 1, f"{j-16}. {b}").font = VAL_FONT
        ws.cell(j, 1).alignment = WRAP
        ws.row_dimensions[j].height = 28

    # ====== Data prep: classify products ======
    classifications = {c['item_id']: c['classification']
                       for c in intel.get('product_classifications', [])}

    # Identify anchor IDs from source_links.json
    anchor_ids = set()
    sl = d / 'source_links.json'
    if sl.exists():
        links_data = json.loads(sl.read_text(encoding='utf-8'))
        import re
        anchor_ids = set()
        for l in links_data.get('links', []):
            if l.get('anchor'):
                url = l.get('url', '')
                m = re.search(r'/itm/(\d+)', url)
                if m:
                    anchor_ids.add(m.group(1))

    # Build excluded map: item_id -> exclude info
    excluded_map = {}
    for ex in intel['anchor_analysis'].get('rule0_excluded', []):
        url = ex.get('competitor_url', '')
        for p in d.glob('product_*.json'):
            pid = p.stem.replace('product_', '')
            if pid in url:
                excluded_map[pid] = ex
                break

    def _val(cls, key, default='-'):
        v = cls.get(key, '')
        return v if v and v != 'Unknown' else default

    def _fmt_vehicle(cls):
        v = cls.get('vehicle')
        if not v or not isinstance(v, dict):
            return '-'
        makes = v.get('makes', [])
        models = v.get('models', [])
        parts = []
        if makes:
            parts.append(', '.join(makes))
        if models:
            parts.append(', '.join(models))
        return ' | '.join(parts) if parts else '-'

    def _write_header(ws, headers, widths):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci, h)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = THIN
        for ci, w in enumerate(widths, 1):
            col_letter = chr(64 + ci) if ci <= 26 else chr(64 + ci // 26) + chr(64 + ci % 26)
            ws.column_dimensions[col_letter].width = w

    def _write_row(ws, ri, vals, align_col=2, fill_color=None):
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, str(v) if v is not None else '-')
            cell.font = VAL_FONT; cell.border = THIN
            if ci == align_col:
                cell.alignment = WRAP
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    # Categorize products
    products = sorted(d.glob('product_*.json'))
    anchors = []
    valid_comps = []
    excluded_comps = []

    for p in products:
        pid = p.stem.replace('product_', '')
        data = json.loads(p.read_text(encoding='utf-8'))
        cls = classifications.get(pid, {})
        entry = (pid, data, cls)

        if pid in anchor_ids:
            anchors.append(entry)
        elif pid in excluded_map:
            excluded_comps.append(entry)
        else:
            valid_comps.append(entry)

    # ====== Sheet 2: Valid Competitors ======
    GREEN_FILL = 'C6EFCE'
    ws_val = wb.create_sheet('Valid竞品')
    headers_val = ['Item ID', 'Title', 'Price', 'Position', 'Finish',
                   'Structure', 'Sub-Type', 'Vehicle', 'Images',
                   'Vision Match', 'Vision Diff', 'Status']
    _write_header(ws_val, headers_val, [16, 60, 12, 18, 14, 14, 18, 40, 10, 12, 35, 10])

    new_count_val = 0
    for ri, (pid, data, cls) in enumerate(valid_comps, 2):
        is_new = pid in new_ids
        if is_new:
            new_count_val += 1
        vr = cls.get('vision_result') or {}
        status = 'New-Valid' if is_new else 'Valid'
        fill = 'F4B183' if is_new else GREEN_FILL
        vals = [pid, (data.get('title', '') or '')[:120], data.get('price', ''),
                _val(cls, 'position'), _val(cls, 'finish'),
                _val(cls, 'structure'), _val(cls, 'sub_type'),
                _fmt_vehicle(cls), data.get('imageCount', '-'),
                'Yes' if vr.get('match') else ('No' if vr else '-'),
                vr.get('diff_type', '-'),
                status]
        _write_row(ws_val, ri, vals, fill_color=fill)
    if not valid_comps:
        ws_val.merge_cells('A2:L2')
        ws_val.cell(2, 1, 'No valid competitors found').font = VAL_FONT
    # Legend row
    legend_row = max(len(valid_comps) + 3, 4)
    ws_val.merge_cells(f'A{legend_row}:J{legend_row}')
    legend_cell = ws_val.cell(legend_row, 1, f'   Orange = 本轮新增 ({new_count_val} new)  |  Green = 已有竞品')
    legend_cell.font = Font(name='Microsoft YaHei', size=9, italic=True, color='666666')
    ws_val.auto_filter.ref = f"A1:L{max(len(valid_comps) + 1, 2)}"

    # ====== Sheet 3: Excluded Competitors ======
    YELLOW_FILL = 'FFEB9C'
    ws_ex = wb.create_sheet('Excluded竞品')
    headers_ex = ['Item ID', 'Title', 'Price', 'Position', 'Finish',
                  'Structure', 'Sub-Type', 'Vehicle', 'Images',
                  'Vision Match', 'Vision Diff',
                  'Status', 'Exclude Reason']
    _write_header(ws_ex, headers_ex, [16, 60, 12, 18, 14, 14, 18, 40, 10, 12, 35, 10, 50])

    new_count_ex = 0
    for ri, (pid, data, cls) in enumerate(excluded_comps, 2):
        is_new = pid in new_ids
        if is_new:
            new_count_ex += 1
        status = 'New-Excluded' if is_new else 'Excluded'
        fill = 'F4B183' if is_new else YELLOW_FILL
        ex_info = excluded_map.get(pid, {})
        reason = (f"{ex_info.get('reason', '未知')}: "
                  f"anchor={ex_info.get('anchor_value', '?')}, "
                  f"competitor={ex_info.get('competitor_value', '?')}") if ex_info else '-'
        vr = cls.get('vision_result') or {}
        vals = [pid, (data.get('title', '') or '')[:120], data.get('price', ''),
                _val(cls, 'position'), _val(cls, 'finish'),
                _val(cls, 'structure'), _val(cls, 'sub_type'),
                _fmt_vehicle(cls), data.get('imageCount', '-'),
                'Yes' if vr.get('match') else ('No' if vr else '-'),
                vr.get('diff_type', '-'),
                status, reason]
        _write_row(ws_ex, ri, vals, fill_color=fill)
    if not excluded_comps:
        ws_ex.merge_cells('A2:M2')
        ws_ex.cell(2, 1, 'No excluded competitors').font = VAL_FONT
    # Legend row
    legend_row_ex = max(len(excluded_comps) + 3, 4)
    ws_ex.merge_cells(f'A{legend_row_ex}:K{legend_row_ex}')
    legend_ex = ws_ex.cell(legend_row_ex, 1, f'   Orange = 本轮新增 ({new_count_ex} new)  |  Yellow = 已有竞品')
    legend_ex.font = Font(name='Microsoft YaHei', size=9, italic=True, color='666666')
    ws_ex.auto_filter.ref = f"A1:K{max(len(excluded_comps) + 1, 2)}"

    # ====== Sheet 4: Anchor Products ======
    BLUE_FILL = 'BDD7EE'
    ws_anchor = wb.create_sheet('锚定产品')
    headers_anchor = ['Item ID', 'Title', 'Price', 'Position', 'Finish',
                      'Structure', 'Sub-Type', 'Vehicle', 'Images']
    _write_header(ws_anchor, headers_anchor, [16, 60, 12, 18, 14, 14, 18, 40, 10])

    for ri, (pid, data, cls) in enumerate(anchors, 2):
        vals = [pid, (data.get('title', '') or '')[:120], data.get('price', ''),
                _val(cls, 'position'), _val(cls, 'finish'),
                _val(cls, 'structure'), _val(cls, 'sub_type'),
                _fmt_vehicle(cls), data.get('imageCount', '-')]
        _write_row(ws_anchor, ri, vals, fill_color=BLUE_FILL)
    if not anchors:
        ws_anchor.merge_cells('A2:I2')
        ws_anchor.cell(2, 1, 'No anchor products found').font = VAL_FONT
    ws_anchor.auto_filter.ref = f"A1:I{max(len(anchors) + 1, 2)}"

    # ====== Sheet 3: Image Index ======
    ws3 = wb.create_sheet('图片索引')
    headers3 = ['#', 'Type', 'File', 'Logo', 'Structure', 'Floating', 'Confidence', 'Status']
    for ci, h in enumerate(headers3, 1):
        c = ws3.cell(1, ci, h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = THIN

    for ci, w in enumerate([6, 18, 35, 10, 12, 10, 12, 10], 1):
        ws3.column_dimensions[chr(64 + ci)].width = w

    # 用 image_batch.json 的顺序
    for ri, (key, info) in enumerate(images.items(), 2):
        v = validation.get(info['label'], {})
        vals = [ri - 1, info['label'], Path(info.get('path', '')).name,
                'Yes' if v.get('logo_detected') else 'No',
                'Yes' if v.get('structure_changed') else 'No',
                'Yes' if v.get('floating_issue') else 'No',
                v.get('confidence', '?'),
                'PASS' if v.get('passed') else 'FAIL']
        for ci, val in enumerate(vals, 1):
            cell = ws3.cell(ri, ci, str(val))
            cell.font = VAL_FONT; cell.border = THIN
            cell.alignment = CENTER if ci != 2 else WRAP

    # ====== Sheet 4: Market Intel ======
    ws4 = wb.create_sheet('市场情报')
    ws4.column_dimensions['A'].width = 24
    ws4.column_dimensions['B'].width = 70
    ws4.merge_cells('A1:B1')
    ws4.cell(1, 1, 'Market Intelligence Summary').font = TITLE_FONT

    mi = [
        ('Links Analyzed', intel['total_links_analyzed']),
        ('Valid Listings', intel['valid_listings']),
        ('Price Range', intel['price_range']),
        ('Average Price', f"${intel['avg_price']}"),
        ('Image Count (min/max/avg)',
         f"{intel['image_counts']['min']} / {intel['image_counts']['max']} / {intel['image_counts']['avg']}"),
        ('', ''),
        ('Anchor Position', intel['anchor_analysis']['anchor_classification']['position']),
        ('Anchor Finish', intel['anchor_analysis']['anchor_classification']['finish']),
        ('Anchor Structure', intel['anchor_analysis']['anchor_classification'].get('structure', '?')),
        ('Valid Competitors', intel['anchor_analysis']['valid_competitors_count']),
        ('Rule-0 Excluded', intel['anchor_analysis']['rule0_excluded_count']),
    ]
    for i, (k, v) in enumerate(mi, 3):
        ws4.cell(i, 1, str(k)).font = Font(name='Microsoft YaHei', size=10, bold=True)
        ws4.cell(i, 1).border = THIN
        ws4.cell(i, 2, str(v)).font = VAL_FONT
        ws4.cell(i, 2).border = THIN

    # Year expansion alerts
    year_alerts = intel['anchor_analysis'].get('year_expanded_alerts', [])
    if year_alerts:
        ws4.cell(len(mi) + 4, 1, 'Year Expansion Alerts').font = SEC_FONT
        for ai, alert in enumerate(year_alerts):
            r = len(mi) + 5 + ai
            ws4.merge_cells(f'A{r}:B{r}')
            ws4.cell(r, 1, f"{alert['note']} | {alert['competitor_year']} vs {alert['anchor_year']}").font = VAL_FONT
            ws4.cell(r, 1).alignment = WRAP
            ws4.row_dimensions[r].height = 22
            r += 1
            ws4.merge_cells(f'A{r}:B{r}')
            url_cell = ws4.cell(r, 1, alert['competitor_url'])
            url_cell.font = Font(name='Microsoft YaHei', size=9, color='0563C1', underline='single')
            url_cell.alignment = WRAP
            ws4.row_dimensions[r].height = 18
        gap = len(year_alerts) * 2 + 1
    else:
        gap = 1

    # Top keywords
    ws4.cell(len(mi) + 4 + gap, 1, 'Top Keywords').font = SEC_FONT
    kws = intel.get('common_keywords', [])[:25]
    kw_text = ', '.join(f"{w}({c})" for w, c in kws)
    r = len(mi) + 5 + gap
    ws4.merge_cells(f'A{r}:B{r}')
    ws4.cell(r, 1, kw_text).font = VAL_FONT
    ws4.cell(r, 1).alignment = WRAP
    ws4.row_dimensions[r].height = 40

    # ═══════ Seller Comparison ═══════
    r = len(mi) + 7 + gap
    ws4.cell(r, 1, 'Seller Comparison').font = SEC_FONT
    r += 1
    # 读取锚定ID
    anchor_ids = set()
    sl = d / 'source_links.json'
    if sl.exists():
        links = json.loads(sl.read_text(encoding='utf-8'))
        # extract anchor IDs from URLs
        anchor_ids = set()
        for l in links.get('links', []):
            if l.get('anchor'):
                m2 = re.search(r'/itm/(\d+)', l.get('url', ''))
                if m2:
                    anchor_ids.add(m2.group(1))

    seller_headers = ['Seller', 'Store URL', 'Feedback', '%', 'Anchor/Competitor']
    for ci, h in enumerate(seller_headers, 1):
        c = ws4.cell(r, ci, h)
        c.font = Font(name='Microsoft YaHei', size=10, bold=True, color='FFFFFF')
        c.fill = HDR_FILL; c.alignment = CENTER; c.border = THIN
    ws4.column_dimensions['C'].width = 25; ws4.column_dimensions['D'].width = 35
    ws4.column_dimensions['E'].width = 12; ws4.column_dimensions['F'].width = 12
    ws4.column_dimensions['G'].width = 18

    r += 1
    for pj in sorted(d.glob('product_*.json')):
        data = json.loads(pj.read_text(encoding='utf-8'))
        pid = pj.stem.replace('product_', '')
        seller_name = data.get('seller', '?')
        seller_url = data.get('sellerUrl', '')
        fb = data.get('sellerFeedback', '?')
        fbpct = data.get('sellerFeedbackPct', '?')
        is_anchor = 'Anchor' if pid in anchor_ids else 'Competitor'

        vals = [seller_name, seller_url, str(fb), str(fbpct), is_anchor]
        for ci, val in enumerate(vals, 1):
            cell = ws4.cell(r, ci, str(val))
            cell.font = VAL_FONT; cell.border = THIN
            if ci == 2 and val:
                cell.font = Font(name='Microsoft YaHei', size=9, color='0563C1', underline='single')
        r += 1

    # ═══════ Compatibility Coverage ═══════
    r += 1
    ws4.cell(r, 1, 'Compatibility Coverage').font = SEC_FONT
    r += 1

    # 统计所有产品覆盖的车型数
    total_vehicles = set()
    anchor_vehicles = set()
    for pj in sorted(d.glob('product_*.json')):
        data = json.loads(pj.read_text(encoding='utf-8'))
        pid = pj.stem.replace('product_', '')
        compat = data.get('compatibility', [])
        if not isinstance(compat, list) or not compat:
            continue
        for row in compat:
            if isinstance(row, dict):
                key = f"{row.get('Year','')}|{row.get('Make','')}|{row.get('Model','')}|{row.get('Trim','')}"
                total_vehicles.add(key)
                if pid in anchor_ids:
                    anchor_vehicles.add(key)

    ws4.merge_cells(f'A{r}:B{r}')
    ws4.cell(r, 1, f'Total unique vehicle-trim combinations: {len(total_vehicles)}').font = VAL_FONT
    r += 1
    ws4.merge_cells(f'A{r}:B{r}')
    ws4.cell(r, 1, f'Anchor-only coverage: {len(anchor_vehicles)} combinations').font = VAL_FONT
    r += 1
    ws4.merge_cells(f'A{r}:B{r}')
    ws4.cell(r, 1, f'Competitor-expanded coverage: {len(total_vehicles) - len(anchor_vehicles)} additional combinations').font = VAL_FONT

    out = d / f'{sku}.xlsx'
    wb.save(str(out))
    print(f'Done: {out.name}')
    print(f'Sheets: 产品概览 | Valid竞品 | Excluded竞品 | 锚定产品 | 图片索引 | 市场情报')


if __name__ == '__main__':
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument('mother_dir')
    p.add_argument('sku')
    args = p.parse_args()
    build_summary_xlsx(args.mother_dir, args.sku)
