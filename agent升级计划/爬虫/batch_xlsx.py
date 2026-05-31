"""单个 product JSON → xlsx（6 Sheet：概览/规格/兼容表/物流/卖家/图片）"""
import json, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

HDR_FONT = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
HDR_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
TITLE_FONT = Font(name='Microsoft YaHei', size=14, bold=True, color='1F3864')
LABEL_FONT = Font(name='Microsoft YaHei', size=10, bold=True)
VAL_FONT = Font(name='Microsoft YaHei', size=10)
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
WRAP = Alignment(vertical='top', wrap_text=True)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _hdr(ws, row, cols):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = THIN


def export_one(json_path: Path, output_path: Path = None):
    data = json.loads(json_path.read_text(encoding='utf-8'))
    wb = Workbook()

    # ═══════ Sheet 1: Product Info ═══════
    ws1 = wb.active
    ws1.title = "Product Info"
    ws1.merge_cells('A1:B1')
    ws1.cell(1, 1, data.get('title', 'N/A')).font = TITLE_FONT
    ws1.row_dimensions[1].height = 30

    info_fields = [
        ('eBay Item ID', 'item_id'), ('URL', 'url'), ('Price', 'price'),
        ('Condition', 'condition'), ('Seller', 'seller'),
        ('Seller Feedback', 'sellerFeedback'), ('Feedback %', 'sellerFeedbackPct'),
        ('Quantity', 'quantity'), ('Sold', 'sold'),
        ('Image Count', 'imageCount'), ('Category ID', 'categoryId'),
        ('Scraped At', 'scraped_at'),
    ]
    for i, (label, key) in enumerate(info_fields, 3):
        ws1.cell(i, 1, label).font = LABEL_FONT
        ws1.cell(i, 1).border = THIN
        val = str(data.get(key, ''))
        ws1.cell(i, 2, val).font = VAL_FONT
        ws1.cell(i, 2).border = THIN
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 80

    # ═══════ Sheet 2: Item Specifics ═══════
    ws2 = wb.create_sheet("Item Specifics")
    specs = data.get('specs', {})
    ws2.cell(1, 1, "Parameter").font = HDR_FONT
    ws2.cell(1, 1).fill = HDR_FILL
    ws2.cell(1, 2, "Value").font = HDR_FONT
    ws2.cell(1, 2).fill = HDR_FILL
    ws2.cell(1, 1).border = THIN; ws2.cell(1, 2).border = THIN
    for i, (k, v) in enumerate(specs.items(), 2):
        ws2.cell(i, 1, k).font = LABEL_FONT; ws2.cell(i, 1).border = THIN
        ws2.cell(i, 2, str(v)[:300]).font = VAL_FONT; ws2.cell(i, 2).border = THIN
        ws2.cell(i, 2).alignment = WRAP
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 80

    # ═══════ Sheet 3: Compatibility ═══════
    compat = data.get('compatibility', [])
    if compat and isinstance(compat[0], dict):
        ws3 = wb.create_sheet("Compatibility")
        headers = list(compat[0].keys())
        for ci, h in enumerate(headers, 1):
            ws3.cell(1, ci, h).font = HDR_FONT
            ws3.cell(1, ci).fill = HDR_FILL
            ws3.cell(1, ci).alignment = CENTER
            ws3.cell(1, ci).border = THIN
        for ri, row in enumerate(compat, 2):
            for ci, h in enumerate(headers, 1):
                ws3.cell(ri, ci, str(row.get(h, ''))).font = VAL_FONT
                ws3.cell(ri, ci).border = THIN
        col_letter = lambda n: chr(64 + n) if n <= 26 else chr(64 + (n-1)//26) + chr(65 + (n-1)%26)
        for ci in range(1, len(headers) + 1):
            ws3.column_dimensions[col_letter(ci)].width = 22
        ws3.auto_filter.ref = f"A1:{col_letter(len(headers))}{len(compat) + 1}"

    # ═══════ Sheet 4: Shipping & Returns ═══════
    ws4 = wb.create_sheet("Shipping & Returns")
    shipping_fields = [
        ('Shipping', 'shipping'),
        ('Returns', 'returns'),
        ('Payment', 'payment'),
    ]
    for i, (label, key) in enumerate(shipping_fields, 1):
        ws4.cell(i, 1, label).font = LABEL_FONT
        ws4.cell(i, 1).border = THIN
        ws4.cell(i, 2, str(data.get(key, ''))).font = VAL_FONT
        ws4.cell(i, 2).border = THIN
        ws4.cell(i, 2).alignment = WRAP
    ws4.column_dimensions['A'].width = 18
    ws4.column_dimensions['B'].width = 80

    # ═══════ Sheet 5: Seller Info ═══════
    ws5 = wb.create_sheet("Seller Info")
    seller_fields = [
        ('Seller Name', 'seller'),
        ('Seller Username', 'sellerUsername'),
        ('Store URL', 'sellerUrl'),
        ('Feedback', 'sellerFeedback'),
        ('Feedback %', 'sellerFeedbackPct'),
        ('Other Items URL', 'sellerItemsUrl'),
    ]
    for i, (label, key) in enumerate(seller_fields, 1):
        ws5.cell(i, 1, label).font = LABEL_FONT
        ws5.cell(i, 1).border = THIN
        val = str(data.get(key, ''))
        ws5.cell(i, 2, val).font = VAL_FONT
        ws5.cell(i, 2).border = THIN
        ws5.cell(i, 2).alignment = WRAP
    ws5.column_dimensions['A'].width = 20
    ws5.column_dimensions['B'].width = 80

    # ═══════ Sheet 6: Images ═══════
    ws6 = wb.create_sheet("Images")
    ws6.cell(1, 1, "#").font = HDR_FONT; ws6.cell(1, 1).fill = HDR_FILL
    ws6.cell(1, 2, "URL").font = HDR_FONT; ws6.cell(1, 2).fill = HDR_FILL
    ws6.cell(1, 1).border = THIN; ws6.cell(1, 2).border = THIN
    for i, img in enumerate(data.get('images', []), 2):
        ws6.cell(i, 1, i - 1).font = VAL_FONT; ws6.cell(i, 1).border = THIN
        ws6.cell(i, 2, img).font = VAL_FONT; ws6.cell(i, 2).border = THIN
    ws6.column_dimensions['A'].width = 6
    ws6.column_dimensions['B'].width = 100

    out = output_path or json_path.with_suffix('.xlsx')
    wb.save(str(out))
    return out


def batch_export(mother_dir: str):
    mother = Path(mother_dir)
    products = sorted(mother.glob("product_*.json"))
    if not products:
        print(f"未找到 product_*.json: {mother}")
        return
    print(f"找到 {len(products)} 个 product JSON")
    for p in products:
        out = export_one(p)
        print(f"  {p.name} -> {out.name}")
    print(f"\n全部完成: {len(products)} 个 xlsx")


if __name__ == '__main__':
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument('mother_dir', help='母文件目录（或单个 product JSON 路径）')
    args = p.parse_args()
    path = Path(args.mother_dir)
    if path.is_file():
        export_one(path)
    else:
        batch_export(args.mother_dir)
