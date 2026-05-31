"""Parse scraped eBay text into structured xlsx with modules."""
import sys, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_ebay_text(text):
    lines = text.strip().split("\n")
    modules = {}

    # --- Product Info ---
    info = {}
    for i, line in enumerate(lines):
        if line.strip().startswith("US $"):
            info["Price"] = line.strip()
            break
    for i, line in enumerate(lines):
        if line.strip().startswith("Condition:"):
            info["Condition"] = lines[i+1].strip() if i+1 < len(lines) else ""
            break
    for i, line in enumerate(lines):
        if line.strip() == "eBay item number:":
            info["eBay Item Number"] = lines[i].strip().replace("eBay item number:", "")
            break
    for i, line in enumerate(lines):
        if "Last updated on" in line:
            info["Last Updated"] = line.strip()
            break
    for i, line in enumerate(lines):
        if line.strip().startswith("Quantity:"):
            info["Quantity"] = line.strip().replace("Quantity:", "")
            break
    # Find title (first bold-looking product line near price)
    for i, line in enumerate(lines):
        if "Upper Lower Door Hinges" in line or "Door Hinge Kit" in line or "Door Hinge Set" in line:
            if len(line) > 20 and not line.startswith("|") and not line.startswith("Skip"):
                info["Title"] = line.strip()
                break
    modules["Product Info"] = info

    # --- Item Specifics ---
    specifics = {}
    spec_section = False
    skip_next = False
    for i, line in enumerate(lines):
        if line.strip() == "Item specifics":
            spec_section = True
            continue
        if spec_section:
            if "Item description from the seller" in line:
                break
            if "about the condition" in line:
                continue
            if line.strip().startswith("about the"):
                continue
            stripped = line.strip()
            # Key-value pattern: key on one line, value on next
            if stripped and not stripped.startswith("New:") and not stripped.startswith("Read more"):
                m = re.match(r"^(\D+)$", stripped)
                # Check if it's a known key
                known_keys = ["Condition", "Country of Origin", "Fitment1", "Fitment2", "Item SKU",
                              "Item Model", "Instructions&Hardwares", "Warranty", "Color",
                              "Manufacturer Part Number", "UPC", "Material", "MPN",
                              "Placement on Vehicle", "Brand", "Shipping Notes", "Type",
                              "Model", "Surface Finish", "Category"]
                if stripped in known_keys and i+1 < len(lines):
                    val = lines[i+1].strip()
                    if val not in known_keys and val:
                        specifics[stripped] = val
    modules["Item Specifics"] = specifics

    # --- Compatibility Table ---
    compat = []
    in_compat = False
    data_rows = 0
    for i, line in enumerate(lines):
        if line.strip() == "Compatibility":
            in_compat = True
            continue
        if in_compat:
            if "This part is compatible with" in line:
                continue
            if "Portions of the information" in line:
                break
            if "Results Pagination" in line:
                break
            stripped = line.strip()
            parts = stripped.split("\t")
            if len(parts) >= 5 and parts[0].isdigit():
                compat.append({
                    "Year": parts[0],
                    "Make": parts[1],
                    "Model": parts[2],
                    "Trim": parts[3],
                    "Engine": parts[4],
                    "Notes": parts[5] if len(parts) > 5 else ""
                })
                data_rows += 1
            if data_rows > 400:
                break
    modules["Compatibility"] = compat

    # --- Shipping & Returns ---
    shipping = {}
    ship_keys = ["Shipping", "Located in", "Delivery", "Returns", "Payments"]
    for i, line in enumerate(lines):
        stripped = line.strip()
        for k in ship_keys:
            if stripped.startswith(k + ":"):
                val = stripped.replace(k + ":", "").strip()
                # If value is empty or just ".", check next line
                if (not val or val in [".", "See details", ". See details"]) and i+1 < len(lines):
                    next_val = lines[i+1].strip()
                    if next_val and not any(next_val.startswith(sk + ":") for sk in ship_keys):
                        val = next_val
                if val and val not in ["See details", ". See details"]:
                    shipping[k] = val
    modules["Shipping & Returns"] = shipping

    # --- Seller Info ---
    seller = {}
    for i, line in enumerate(lines):
        if "positive feedback" in line and "items sold" in line:
            seller["Feedback Summary"] = line.strip()
        if line.strip().startswith("Joined "):
            seller["Member Since"] = line.strip()
    # Detailed seller ratings
    seller_ratings = {}
    rating_keys = ["Accurate description", "Reasonable shipping cost", "Shipping speed", "Communication"]
    for i, line in enumerate(lines):
        if line.strip() in rating_keys and i+1 < len(lines):
            try:
                score = float(lines[i+1].strip())
                seller_ratings[line.strip()] = score
            except ValueError:
                pass
    if seller_ratings:
        seller["Detailed Ratings (12mo avg)"] = seller_ratings
    modules["Seller Info"] = seller

    # --- Product Reviews ---
    reviews = []
    in_reviews = False
    for i, line in enumerate(lines):
        if line.strip() == "Most relevant reviews":
            in_reviews = True
            continue
        if in_reviews:
            if "Back to home page" in line:
                break
            # Review starts with "by username"
            if line.startswith("by "):
                review = {"author": line.replace("by ", "").strip()}
                # Next few lines: date, title, body
                j = i + 1
                while j < len(lines) and j < i+10:
                    nl = lines[j].strip()
                    if nl and not nl.startswith("by ") and not nl.startswith("Verified purchase") and not nl.startswith("See all"):
                        if not review.get("date"):
                            review["date"] = nl
                        elif not review.get("title"):
                            review["title"] = nl
                        elif not review.get("body"):
                            review["body"] = nl
                            break
                    j += 1
                if review.get("body"):
                    reviews.append(review)
                i = j
    modules["Reviews"] = reviews

    return modules


def write_xlsx(modules, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def write_header(ws, headers, col_widths=None):
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            if col_widths and ci <= len(col_widths):
                ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci-1]

    def write_row(ws, row_num, values):
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=ci, value=v)
            cell.alignment = cell_align
            cell.border = thin_border

    # Sheet 1: Product Info
    ws = wb.create_sheet("Product Info")
    write_header(ws, ["Field", "Value"], [20, 60])
    info = modules.get("Product Info", {})
    for ri, (k, v) in enumerate(info.items(), 2):
        write_row(ws, ri, [k, v])

    # Sheet 2: Item Specifics
    ws = wb.create_sheet("Item Specifics")
    write_header(ws, ["Field", "Value"], [28, 55])
    specs = modules.get("Item Specifics", {})
    for ri, (k, v) in enumerate(specs.items(), 2):
        write_row(ws, ri, [k, v])

    # Sheet 3: Compatibility
    ws = wb.create_sheet("Compatibility")
    compat_headers = ["Year", "Make", "Model", "Trim", "Engine", "Notes"]
    compat_widths = [8, 16, 12, 28, 40, 10]
    write_header(ws, compat_headers, compat_widths)
    compat = modules.get("Compatibility", [])
    for ri, row in enumerate(compat, 2):
        write_row(ws, ri, [row.get(h, "") for h in compat_headers])

    # Sheet 4: Shipping & Returns
    ws = wb.create_sheet("Shipping & Returns")
    write_header(ws, ["Field", "Value"], [20, 60])
    shipping = modules.get("Shipping & Returns", {})
    for ri, (k, v) in enumerate(shipping.items(), 2):
        write_row(ws, ri, [k, v])

    # Sheet 5: Seller Info
    ws = wb.create_sheet("Seller Info")
    write_header(ws, ["Field", "Value"], [30, 60])
    seller = modules.get("Seller Info", {})
    for ri, (k, v) in enumerate(seller.items(), 2):
        if isinstance(v, dict):
            sub = "; ".join(f"{sk}: {sv}" for sk, sv in v.items())
            write_row(ws, ri, [k, sub])
        else:
            write_row(ws, ri, [k, str(v)])

    # Sheet 6: Reviews
    ws = wb.create_sheet("Reviews")
    rev_headers = ["Author", "Date", "Title", "Body"]
    rev_widths = [16, 14, 30, 55]
    write_header(ws, rev_headers, rev_widths)
    reviews = modules.get("Reviews", [])
    for ri, r in enumerate(reviews, 2):
        write_row(ws, ri, [r.get("author", ""), r.get("date", ""), r.get("title", ""), r.get("body", "")])

    # Sheet 7: Raw Text (for reference)
    ws = wb.create_sheet("Raw Text")
    ws.column_dimensions["A"].width = 120
    # We don't have the raw text here, skip for now

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    txt_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else txt_path.replace(".txt", ".xlsx")

    with open(txt_path, encoding="utf-8") as f:
        text = f.read()

    modules = parse_ebay_text(text)
    write_xlsx(modules, out_path)

    # Print summary
    for name, data in modules.items():
        if isinstance(data, list):
            print(f"  {name}: {len(data)} rows")
        elif isinstance(data, dict):
            print(f"  {name}: {len(data)} fields")
