"""
共享数据提取器 v2 — 使用与 gen_html_5_14_5_20.py 完全一致的列索引
从 Excel 提取标准化 JSON，供 新品复盘 和 三部周报v1 共同消费

用法:
  python extract_weekly_data.py            # 默认最新周期 5.14-5.20
"""

import json, os, sys
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook
from collections import defaultdict
from datetime import date, datetime

THIS_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_FILE = os.path.join(THIS_DIR, '..', '新品复盘', '周报', '新品检查周源数据和PLP数据.xlsx')
OUTPUT_DIR = os.path.join(THIS_DIR, 'output')

# ===== 列索引 — 与 gen_html_5_14_5_20.py 完全一致 =====
# openpyxl iter_rows(values_only=True) 返回 0-indexed tuple
C = {
    'sale_no': 0, 'sku': 1, 'list_date': 2, 'first_order': 3,
    'analyst': 4, 'category': 5, 'expand_type': 6,
    'sales_curr': 17, 'sales_prev': 16,       # 本期=5.14-5.20 / 上期=5.7-5.13
    'rev_curr': 29, 'rev_prev': 28,
    'rival_curr': 41, 'rival_prev': 40,
    'share_curr': 52, 'share_prev': 51,
    'ord8_curr': 74, 'ord8_prev': 73,
    'freq7_curr': 85, 'freq7_prev': 84,
    'nfreq7_curr': 96, 'nfreq7_prev': 95,
    'mkt_curr': 108, 'mkt_prev': 107,
    'op_curr': 119,
    'plp_curr': 125, 'plg_curr': 130,         # PLP=5.11-5.17 / PLG=5.11-5.17
}

ANALYSTS = ['俞东旭', '张潇', '朱培源', '王偲涵', '章鹏', '胡煜星']
CATEGORIES = ['车门系统', '车身外扩件', '挡泥板', '机盖及附件', '其他', '饰条', '牌照板支架']
EXPAND_TYPES = ['原开品', '拓展品', '组合件']
PERIOD_LABEL = '5.14-5.20'
PREV_PERIOD = '5.7-5.13'
CUTOFF_DATE = date(2026, 5, 20)

def safe_float(v, default=0):
    try: return float(v) if v is not None else default
    except: return default

def get_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None


def extract_all():
    print(f"读取: {SOURCE_FILE}")
    wb = load_workbook(SOURCE_FILE, data_only=True)
    ws_main = wb['四三数据累计']

    # ===== 提取四三累计数据 =====
    rows = []
    for row in ws_main.iter_rows(min_row=2, values_only=True):
        if not row[C['sku']]:
            continue
        r = list(row)
        ld = get_date(r[C['list_date']])
        if not ld:
            continue
        sku_data = {
            'saleNo': str(r[C['sale_no']] or ''),
            'sku': str(r[C['sku']] or ''),
            'listDate': str(ld),
            'firstOrder': str(get_date(r[C['first_order']]) or ''),
            'analyst': str(r[C['analyst']] or '').strip(),
            'category': str(r[C['category']] or '').strip() or '未分类',
            'expandType': str(r[C['expand_type']] or '').strip() or '其他',
            # 本期数据
            'salesCurr': safe_float(r[C['sales_curr']]),
            'salesPrev': safe_float(r[C['sales_prev']]),
            'revenueCurr': safe_float(r[C['rev_curr']]),
            'revenuePrev': safe_float(r[C['rev_prev']]),
            'rivalCurr': safe_float(r[C['rival_curr']]),
            'rivalPrev': safe_float(r[C['rival_prev']]),
            'shareCurr': round(safe_float(r[C['share_curr']]) * 100, 1),       # ×100
            'sharePrev': round(safe_float(r[C['share_prev']]) * 100, 1),
            'ord8Curr': str(r[C['ord8_curr']] or '').strip(),
            'ord8Prev': str(r[C['ord8_prev']] or '').strip(),
            'freq7Curr': str(r[C['freq7_curr']] or '').strip(),
            'freq7Prev': str(r[C['freq7_prev']] or '').strip(),
            'nfreq7Curr': str(r[C['nfreq7_curr']] or '').strip(),
            'nfreq7Prev': str(r[C['nfreq7_prev']] or '').strip(),
            'mktStatus': str(r[C['mkt_curr']] or '').strip(),
            'mktStatusPrev': str(r[C['mkt_prev']] or '').strip(),
            'opJudge': str(r[C['op_curr']] or '').strip(),
            'plpCurr': str(r[C['plp_curr']] or '').strip(),
            'plgCurr': round(safe_float(r[C['plg_curr']]) * 100, 1),           # ×100
        }
        rows.append(sku_data)

    # ===== 过滤截止日期 =====
    rows_curr = [r for r in rows if r['listDate'] <= str(CUTOFF_DATE)]
    rows_prev = [r for r in rows if r['listDate'] <= str(date(2026, 5, 13))]

    # ===== KPI 计算 =====
    total_sku = len(rows_curr)
    total_sales = sum(r['salesCurr'] for r in rows_curr)
    total_rev = sum(r['revenueCurr'] for r in rows_curr)
    has_rival = [r for r in rows_curr if r['rivalCurr'] > 0]
    no_rival = [r for r in rows_curr if r['rivalCurr'] == 0]
    y_count = sum(1 for r in has_rival if r['ord8Curr'] == 'Y')
    n_count = sum(1 for r in has_rival if r['ord8Curr'] == 'N')
    no_sale = sum(1 for r in has_rival if r['ord8Curr'] == '未出单')
    low_share = [r for r in rows_curr if r['shareCurr'] < 75 and r['rivalCurr'] > 0]

    # 分析及时率
    timely = sum(1 for r in rows_curr if r['nfreq7Curr'] != '异常' and r['freq7Curr'] != '异常')
    no8d = sum(1 for r in rows_curr if r['nfreq7Curr'] == '异常')
    over7d = sum(1 for r in rows_curr if r['freq7Curr'] == '异常')

    # 已出单合计
    sold_count = y_count + n_count
    sold_rate = round(sold_count / len(has_rival) * 100, 1) if has_rival else 0
    timely_rate = round(timely / total_sku * 100, 1) if total_sku else 0

    # ===== 维度聚合 =====
    def dim_agg(data, key, fields):
        m = defaultdict(lambda: {f: 0 for f in fields})
        for r in data:
            k = r[key] or '未分类'
            for f in fields:
                m[k][f] += r[f]
        return {k: dict(v) for k, v in sorted(m.items())}

    cat_fields = ['salesCurr', 'salesPrev', 'revenueCurr', 'revenuePrev']
    an_fields = ['salesCurr', 'salesPrev', 'revenueCurr', 'revenuePrev']

    categories = dim_agg(rows_curr, 'category', cat_fields)
    analysts = dim_agg(rows_curr, 'analyst', an_fields)

    # 拓展类型聚合
    expand = {}
    for et in EXPAND_TYPES:
        subset = [r for r in rows_curr if r['expandType'] == et]
        h = [r for r in subset if r['rivalCurr'] > 0]
        expand[et] = {
            'sku': len(subset),
            'hasRival': len(h),
            'salesCurr': sum(r['salesCurr'] for r in subset),
            'revenueCurr': sum(r['revenueCurr'] for r in subset),
            'soldCount': sum(1 for r in h if r['ord8Curr'] in ('Y', 'N')),
            'soldRate': round(sum(1 for r in h if r['ord8Curr'] in ('Y', 'N')) / len(h) * 100, 1) if h else 0,
        }

    # ===== 组装输出 =====
    output = {
        'meta': {
            'period': PERIOD_LABEL,
            'prevPeriod': PREV_PERIOD,
            'cutoffDate': str(CUTOFF_DATE),
            'generatedAt': datetime.now().isoformat(),
            'sourceFile': SOURCE_FILE,
            'totalRows': len(rows),
        },
        'kpi': {
            'totalSku': total_sku,
            'newSkuCurr': sum(1 for r in rows_curr if r['listDate'] > str(date(2026, 5, 13))),
            'newSkuPrev': sum(1 for r in rows_curr if r['listDate'] > str(date(2026, 5, 6)) and r['listDate'] <= str(date(2026, 5, 13))),
            'totalSales': total_sales,
            'totalSalesPrev': sum(r['salesPrev'] for r in rows_prev),
            'totalRevenue': round(total_rev, 2),
            'totalRevenuePrev': round(sum(r['revenuePrev'] for r in rows_prev), 2),
            'hasRivalSku': len(has_rival),
            'noRivalSku': len(no_rival),
            'yCount': y_count,
            'nCount': n_count,
            'noSaleCount': no_sale,
            'soldCount': sold_count,
            'soldRate': sold_rate,
            'lowShareSku': len(low_share),
            'avgShare': round(sum(r['shareCurr'] for r in has_rival) / len(has_rival), 1) if has_rival else 0,
        },
        'timeliness': {
            'timelyCount': timely,
            'noAnalysis8d': no8d,
            'over7d': over7d,
            'total': total_sku,
            'timelyRate': timely_rate,
        },
        'dimensions': {
            'categories': categories,
            'analysts': analysts,
            'expandTypes': expand,
        },
        'cum43Data': rows_curr,
        'lowShareData': low_share,
        'constants': {
            'analysts': ANALYSTS,
            'categories': CATEGORIES,
            'expandTypes': EXPAND_TYPES,
        },
    }

    # ===== 输出 =====
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_name = f'weekly_data_{PERIOD_LABEL.replace(".", "-")}.json'
    out_path = os.path.join(OUTPUT_DIR, out_name)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    latest_path = os.path.join(OUTPUT_DIR, 'latest.json')
    with open(latest_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 5.14-5.20 数据提取完成!")
    print(f"   SKU总数: {total_sku}")
    print(f"   总销量: {total_sales}")
    print(f"   总销售额: ${total_rev:,.2f}")
    print(f"   有对手: {len(has_rival)} | 无对手: {len(no_rival)}")
    print(f"   出单: Y={y_count} N={n_count} 未出单={no_sale}")
    print(f"   出单率: {sold_rate}%")
    print(f"   及时分析率: {timely_rate}%")
    print(f"   低占比: {len(low_share)}个")
    print(f"   输出: {out_path}")


if __name__ == '__main__':
    extract_all()
