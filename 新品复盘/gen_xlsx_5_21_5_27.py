"""
生成5.21-5.27周期新品周报数据表（Excel）
"""
import json, os, sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from collections import defaultdict, Counter

WORKDIR = os.path.dirname(os.path.abspath(__file__)) + '/'
SOURCE_FILE = WORKDIR + '周报/新品检查周源数据和PLP数据.xlsx'
OUTPUT_FILE = WORKDIR + '新品周报数据表_5.21-5.27.xlsx'
HTML_FILE = WORKDIR + '新品板块_5.21-5.27.html'

# ===== 从HTML提取数据 =====
print("读取HTML中的数据...")
with open(HTML_FILE, 'r', encoding='utf-8') as f:
    html = f.read()

import re
data_blocks = {}
for m in re.finditer(r'const (\w+) = (.*?);\n', html, re.DOTALL):
    try:
        data_blocks[m.group(1)] = json.loads(m.group(2))
    except:
        pass
print(f"提取到 {len(data_blocks)} 个数据块")

# ===== 辅助函数 =====
def safe_float(v, default=0):
    try: return float(v) if v else default
    except: return default

def num(v, default=0):
    return safe_float(v, default)

# Styles
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='微软雅黑', size=10)
ODD_FILL = PatternFill('solid', fgColor='EEF2FF')
EVEN_FILL = PatternFill('solid', fgColor='FFFFFF')
TOTAL_FILL = PatternFill('solid', fgColor='E8F0FE')
DATA_FONT = Font(name='微软雅黑', size=9)
BOLD_FONT = Font(bold=True, name='微软雅黑', size=9)
THIN = Side(style='thin')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_header(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col+i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN_BORDER

def write_row(ws, row, data, start_col=1, bold=False, bg=None):
    for i, v in enumerate(data):
        c = ws.cell(row=row, column=start_col+i, value=v)
        c.font = BOLD_FONT if bold else DATA_FONT
        c.alignment = Alignment(horizontal='center' if i > 0 else 'left', vertical='center')
        c.border = THIN_BORDER
        if bg:
            c.fill = PatternFill('solid', fgColor=bg)

wb = openpyxl.Workbook()

# ===== Sheet 1: 总盘概览 =====
print("Sheet 1: 总盘概览...")
ws1 = wb.active
ws1.title = "总盘概览"

t = data_blocks['cum43Stats']
pk = data_blocks['prevWeekKpi']
td = data_blocks['timelinessData']['total']

kpi_data = [
    ['累计SKU数', t['total'], f"上周{pk['prevTotalSku']}"],
    ['本品总销量', int(pk['prevTotalSalesQty']), f"环比{pk['salesQtyChange']}"],
    ['总销售额(USD)', f"${pk['prevTotalRevenue']:,.2f}", f"环比{pk['revenueChange']}"],
    ['新品总市占比', f"{t['totalMarketShare']}%", f"上周{t['totalMarketSharePrev']}%"],
    ['出单率(有对手)', f"{round((t['yCount']+t['nCount'])/t['hasRivalCount']*100,1)}%", f"Y:{t['yCount']} N:{t['nCount']} 未:{t['unCount']}"],
    ['分析及时率', td['timelyRate'], f"环比{td['change']}"],
    ['有对手SKU', t['hasRivalCount'], f"无对手{t['noRivalCount']}"],
    ['无对手已出单', t['noRivalSold'], ''],
    ['无对手未出单', t['noRivalUnsold'], ''],
]

write_header(ws1, 1, ['指标', '本周(5.27)', '备注'])
for i, d in enumerate(kpi_data):
    bg = 'EEF2FF' if i % 2 == 0 else 'FFFFFF'
    write_row(ws1, 2+i, d, bg=bg)

# 出单分布
row = len(kpi_data) + 3
write_header(ws1, row, ['分类', '数量', '说明'])
ord_data = [
    ['有对手已出单(Y+N)', t['yCount']+t['nCount'], f"占有对手SKU的{round((t['yCount']+t['nCount'])/t['hasRivalCount']*100,1)}%"],
    ['有对手未出单', t['unCount'], '含本期不分析'],
    ['无对手已出单', t['noRivalSold'], '历史有出单(ord8 Y/N)'],
    ['无对手未出单', t['noRivalUnsold'], '历史无出单'],
    ['合计', t['total'], f"有对手{t['hasRivalCount']}+无对手{t['noRivalCount']}"],
]
for i, d in enumerate(ord_data):
    write_row(ws1, row+1+i, d, bold=(i==len(ord_data)-1), bg='E8F0FE' if i==len(ord_data)-1 else ('EEF2FF' if i%2==0 else 'FFFFFF'))

ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 40

# ===== Sheet 2: 品线维度 =====
print("Sheet 2: 品线维度...")
ws2 = wb.create_sheet("品线维度")
cat_data = data_blocks['categoryRevenueData']
headers = ['品线', 'SKU数', '新上架', '本周销量', '上周销量', '销量环比', '本周销售额', '上周销售额', '销售额环比', '市占比', '上周市占比', '市占环比']
write_header(ws2, 1, headers)
for i, d in enumerate(cat_data):
    write_row(ws2, 2+i, [
        d['category'], d['curSku'], d['curNewSku'],
        d['curSalesQty'], d['prevSalesQty'], d.get('salesQtyChange', '-'),
        f"${d['curRevenue']:,.2f}", f"${d['prevRevenue']:,.2f}", d.get('revenueChange', '-'),
        f"{d.get('curMarketShare', 0)}%", f"{d.get('prevMarketShare', 0)}%", d.get('marketShareChange', '-')
    ], bg='EEF2FF' if i%2==0 else 'FFFFFF')

# 合计行
n = len(cat_data)
total_sku = sum(d['curSku'] for d in cat_data)
total_new = sum(d['curNewSku'] for d in cat_data)
total_sales_c = sum(d['curSalesQty'] for d in cat_data)
total_sales_p = sum(d['prevSalesQty'] for d in cat_data)
total_rev_c = sum(d['curRevenue'] for d in cat_data)
total_rev_p = sum(d['prevRevenue'] for d in cat_data)
total_share_c = round(total_sales_c / (total_sales_c + sum(num(safe_float(d.get('curMarketShare',0))) for d in cat_data)) * 100, 1) if total_sales_c > 0 else 0
write_row(ws2, 2+n, ['合计', total_sku, total_new, total_sales_c, total_sales_p,
    f"+{round((total_sales_c-total_sales_p)/max(abs(total_sales_p),1)*100,1)}%" if total_sales_p else '-',
    f"${total_rev_c:,.2f}", f"${total_rev_p:,.2f}",
    f"+{round((total_rev_c-total_rev_p)/max(abs(total_rev_p),1)*100,1)}%" if total_rev_p else '-',
    f"{t['totalMarketShare']}%", f"{t['totalMarketSharePrev']}%", pk.get('marketShareChange','-')
], bold=True, bg='E8F0FE')

for col in range(1, len(headers)+1):
    ws2.column_dimensions[get_column_letter(col)].width = 14

# ===== Sheet 3: 分析人维度 =====
print("Sheet 3: 分析人维度...")
ws3 = wb.create_sheet("分析人维度")
an_data = data_blocks['analystRevenueData']
headers3 = ['分析人', 'SKU数', '新上架', '本周销量', '上周销量', '销量环比', '本周销售额', '上周销售额', '销售额环比', '市占比', '上周市占比', '市占环比']
write_header(ws3, 1, headers3)
for i, d in enumerate(an_data):
    write_row(ws3, 2+i, [
        d['analyst'], d['curSku'], d['curNewSku'],
        d['curSalesQty'], d['prevSalesQty'], d.get('salesQtyChange', '-'),
        f"${d['curRevenue']:,.2f}", f"${d['prevRevenue']:,.2f}", d.get('revenueChange', '-'),
        f"{d.get('curMarketShare', 0)}%", f"{d.get('prevMarketShare', 0)}%", d.get('marketShareChange', '-')
    ], bg='EEF2FF' if i%2==0 else 'FFFFFF')

# 合计行
n3 = len(an_data)
total_sku3 = sum(d['curSku'] for d in an_data)
total_new3 = sum(d['curNewSku'] for d in an_data)
total_sales_c3 = sum(d['curSalesQty'] for d in an_data)
total_sales_p3 = sum(d['prevSalesQty'] for d in an_data)
total_rev_c3 = sum(d['curRevenue'] for d in an_data)
total_rev_p3 = sum(d['prevRevenue'] for d in an_data)
write_row(ws3, 2+n3, ['合计', total_sku3, total_new3, total_sales_c3, total_sales_p3,
    f"+{round((total_sales_c3-total_sales_p3)/max(abs(total_sales_p3),1)*100,1)}%" if total_sales_p3 else '-',
    f"${total_rev_c3:,.2f}", f"${total_rev_p3:,.2f}",
    f"+{round((total_rev_c3-total_rev_p3)/max(abs(total_rev_p3),1)*100,1)}%" if total_rev_p3 else '-',
    f"{t['totalMarketShare']}%", f"{t['totalMarketSharePrev']}%", pk.get('marketShareChange','-')
], bold=True, bg='E8F0FE')

for col in range(1, len(headers3)+1):
    ws3.column_dimensions[get_column_letter(col)].width = 14

# ===== Sheet 4: 拓展类型 =====
print("Sheet 4: 拓展类型...")
ws4 = wb.create_sheet("拓展类型")
exp_data = data_blocks['expandTypeData']
headers4 = ['拓展类型', '本周SKU', '上周SKU', '出单SKU', '出单率', '上周出单率', '本周销量', '上周销量', '销量环比', '本周销售额', '上周销售额', '销售额环比']
write_header(ws4, 1, headers4)
for i, d in enumerate(exp_data):
    write_row(ws4, 2+i, [
        d['expandType'], d['curSku'], d['prevSku'],
        d['curSalesSku'], d['curSalesRate'], d['prevSalesRate'],
        d['curSalesQty'], d['prevSalesQty'], d['salesQtyChange'],
        f"${d['curRevenue']:,.2f}", f"${d['prevRevenue']:,.2f}", d['revenueChange']
    ], bg='EEF2FF' if i%2==0 else 'FFFFFF')

# 合计行
n4 = len(exp_data)
total_sku4 = sum(d['curSku'] for d in exp_data)
total_prev4 = sum(d['prevSku'] for d in exp_data)
total_sold4 = sum(d['curSalesSku'] for d in exp_data)
total_sales_c4 = sum(d['curSalesQty'] for d in exp_data)
total_sales_p4 = sum(d['prevSalesQty'] for d in exp_data)
total_rev_c4 = sum(d['curRevenue'] for d in exp_data)
total_rev_p4 = sum(d['prevRevenue'] for d in exp_data)
write_row(ws4, 2+n4, ['合计', total_sku4, total_prev4, total_sold4, '-', '-',
    total_sales_c4, total_sales_p4,
    f"+{round((total_sales_c4-total_sales_p4)/max(abs(total_sales_p4),1)*100,1)}%" if total_sales_p4 else '-',
    f"${total_rev_c4:,.2f}", f"${total_rev_p4:,.2f}",
    f"+{round((total_rev_c4-total_rev_p4)/max(abs(total_rev_p4),1)*100,1)}%" if total_rev_p4 else '-'
], bold=True, bg='E8F0FE')

# ===== Sheet 5: 及时率 =====
print("Sheet 5: 及时率...")
ws5 = wb.create_sheet("分析及时率")
time_data = data_blocks['timelinessData']
headers5 = ['分析人', '本周SKU', '及时分析', '8日未分析', '7日未分析', '及时率', '上周及时率', '变化']
write_header(ws5, 1, headers5)
all_analysts = time_data['analysts'] + [time_data['total']]
for i, d in enumerate(all_analysts):
    is_total = d.get('analyst') == '合计'
    write_row(ws5, 2+i, [
        d['analyst'], d['curSku'], d['timelyCount'],
        d['noAnalysis8dCount'], d['noAnalysis7dCount'],
        d['timelyRate'], d['prevTimelyRate'], d.get('change', '-')
    ], bold=is_total, bg='E8F0FE' if is_total else ('EEF2FF' if i%2==0 else 'FFFFFF'))

# ===== Sheet 6: 低占比新品 =====
print("Sheet 6: 低占比新品...")
ws6 = wb.create_sheet("低占比新品")
ls_data = data_blocks['lowShareData']
headers6 = ['SKU', '上架日期', '分析人', '品类', '本周销量', '销量环比', '本周销售额', '对手量', '市占比', '8日出单', '上期市场状态', '本期运作', '本期市场状态', '广告分类']
write_header(ws6, 1, headers6)
for i, d in enumerate(ls_data):
    write_row(ws6, 2+i, [
        d['SKU'], d['launchDate'], d['analyst'], d['category'],
        d['curSalesQty'], d.get('salesQtyChange','-'), f"${d['curRevenue']:,.2f}",
        d['curRivalQty'], f"{d['curMarketShare']}%", d['cur8dStatus'],
        d.get('prevMarketStatus',''), d.get('curOperation',''), d['curMarketStatus'],
        d.get('adClass','')
    ], bg='EEF2FF' if i%2==0 else 'FFFFFF')

# 合计行
n6 = len(ls_data)
total_ls_sales = sum(d['curSalesQty'] for d in ls_data)
total_ls_rev = sum(d['curRevenue'] for d in ls_data)
total_ls_rival = sum(d['curRivalQty'] for d in ls_data)
write_row(ws6, 2+n6, [f'合计({n6}条)', '', '', '', total_ls_sales, '', f'${total_ls_rev:,.2f}', total_ls_rival, '', '', '', '', '', ''], bold=True, bg='E8F0FE')

# ===== Sheet 7: PLP广告 =====
print("Sheet 7: PLP广告...")
ws7 = wb.create_sheet("PLP广告")
plp = data_blocks['plpTotal']
plp_prev = data_blocks['plpPrevTotal']

# PLP概览
write_header(ws7, 1, ['指标', '本周(5.18-5.24)', '上周(5.11-5.17)'])
plp_kpis = [
    ['广告活动数', plp['campaignCount'], plp_prev['campaignCount']],
    ['投放链接数', plp['linkCount'], plp_prev['linkCount']],
    ['曝光量', plp['impression'], plp_prev['impression']],
    ['点击量', plp['click'], plp_prev['click']],
    ['售出数', plp['sold'], plp_prev['sold']],
    ['广告花费(USD)', f"${plp['cost']:,.2f}", f"${plp_prev['cost']:,.2f}"],
    ['广告销售额(USD)', f"${plp['revenue']:,.2f}", f"${plp_prev['revenue']:,.2f}"],
    ['ROAS', plp['roas'], plp_prev['roas']],
    ['CVR', plp['cvr'], plp_prev['cvr']],
    ['CTR', plp['ctr'], plp_prev['ctr']],
    ['CPC', plp['cpc'], plp_prev['cpc']],
    ['CPA', plp['cpa'], plp_prev['cpa']],
    ['ACOS', plp['acos'], plp_prev['acos']],
    ['ACOAS(去重)', plp['acoas'], plp_prev['acoas']],
]
for i, d in enumerate(plp_kpis):
    write_row(ws7, 2+i, d, bg='EEF2FF' if i%2==0 else 'FFFFFF')

# PLP维度
plp_cats = data_blocks['plpCategories']
plp_ans = data_blocks['plpAnalysts']
plp_exps = data_blocks['plpExpandTypes']

row = len(plp_kpis) + 4
dim_headers = ['维度', '活动数', '链接数', '曝光量', '点击量', '售出数', '花费', '广告销售额', 'ROAS', 'CVR', 'CTR', 'CPC', 'CPA', 'ACOS', 'ACOAS']
write_header(ws7, row, dim_headers)

dim_data = [('品线', plp_cats), ('分析人', plp_ans), ('拓展类型', plp_exps)]
r = row + 1
for dim_type, items in dim_data:
    write_row(ws7, r, [f'--- {dim_type} ---'] + ['']*(len(dim_headers)-1), bold=True, bg='D9E2F3')
    r += 1
    for item in items:
        write_row(ws7, r, [
            item['name'], item['campaignCount'], item['linkCount'],
            item['impression'], item['click'], item['sold'],
            f"${item['cost']:,.2f}", f"${item['revenue']:,.2f}",
            item['roas'], item['cvr'], item['ctr'],
            item.get('cpc','-'), item.get('cpa','-'), item['acos'], item['acoas']
        ])
        r += 1
    # 维度合计
    if items:
        t_camp = sum(it['campaignCount'] for it in items)
        t_link = sum(it['linkCount'] for it in items)
        t_impr = sum(it['impression'] for it in items)
        t_click = sum(it['click'] for it in items)
        t_sold = sum(it['sold'] for it in items)
        t_cost = sum(safe_float(it['cost']) for it in items)
        t_rev = sum(safe_float(it['revenue']) for it in items)
        t_roas = round(t_rev/t_cost, 2) if t_cost else 0
        write_row(ws7, r, [f'{dim_type}合计', t_camp, t_link, t_impr, t_click, t_sold,
            f"${t_cost:,.2f}", f"${t_rev:,.2f}", f"{t_roas:.2f}", '-', '-', '-', '-', '-', '-'], bold=True, bg='E8F0FE')
        r += 1

ws7.column_dimensions['A'].width = 20

# ===== Sheet 8: PLG广告 =====
print("Sheet 8: PLG广告...")
ws8 = wb.create_sheet("PLG广告")
pg = data_blocks['plgStats']

write_header(ws8, 1, ['指标', '数值'])
plg_kpis = [
    ['PLG广告花费(USD)', f"${pg['totalSpend']:,.2f}"],
    ['PLG广告销售额(USD)', f"${pg['totalAdRev']:,.2f}"],
    ['PLG自然周总销售额(USD)', f"${pg['totalNatRev']:,.2f}"],
    ['PLG ACOS', pg['acos']],
    ['PLG ACOAS', pg['acoas']],
    ['', ''],
    ['新品总数', pg['totalNewProducts']],
    ['PLP+PLG同开', pg['plpAndPlgBothCount']],
    ['单链接PLP+PLG同开', pg['singleLinkPlpPlgCount']],
    ['单PLG', pg['plgOnlyCount']],
    ['单PLP', pg['plpOnlyCount']],
    ['无广告', pg['noAdCount']],
    ['单PLG且未出单', pg.get('plpDisabledNoSaleCount', 0)],
]
for i, d in enumerate(plg_kpis):
    write_row(ws8, 2+i, d, bg='EEF2FF' if i%2==0 else 'FFFFFF')

# PLG按分析人
row = len(plg_kpis) + 4
write_header(ws8, row, ['分析人', 'SKU数', 'PLP+PLG', '单链接', '单PLG', '单PLP', '无广告', 'PLP未开未出单', 'PLG花费', 'PLG广告销售额', '自然周销售额', 'ACOS', 'ACOAS'])
for i, d in enumerate(pg['byAnalyst']):
    write_row(ws8, row+1+i, [
        d['analyst'], d['total'], d['plpAndPlgBoth'], d['singleLinkPlpPlg'],
        d['plgOnly'], d['plpOnly'], d['noAd'], d['plpDisabledNoSale'],
        f"${d['plgSpend']:,.2f}", f"${d['plgAdRev']:,.2f}", f"${d['plgNatRev']:,.2f}",
        d['acos'], d['acoas']
    ], bg='EEF2FF' if i%2==0 else 'FFFFFF')

# PLG分析人合计行
n8 = len(pg['byAnalyst'])
t_total = sum(d['total'] for d in pg['byAnalyst'])
t_both = sum(d['plpAndPlgBoth'] for d in pg['byAnalyst'])
t_single = sum(d['singleLinkPlpPlg'] for d in pg['byAnalyst'])
t_plg = sum(d['plgOnly'] for d in pg['byAnalyst'])
t_plp = sum(d['plpOnly'] for d in pg['byAnalyst'])
t_noad = sum(d['noAd'] for d in pg['byAnalyst'])
t_dis = sum(d['plpDisabledNoSale'] for d in pg['byAnalyst'])
t_spend = sum(safe_float(d.get('plgSpend', 0)) for d in pg['byAnalyst'])
t_adrev = sum(safe_float(d.get('plgAdRev', 0)) for d in pg['byAnalyst'])
t_natrev = sum(safe_float(d.get('plgNatRev', 0)) for d in pg['byAnalyst'])
write_row(ws8, row+1+n8, ['合计', t_total, t_both, t_single, t_plg, t_plp, t_noad, t_dis,
    f"${t_spend:,.2f}", f"${t_adrev:,.2f}", f"${t_natrev:,.2f}",
    pg['acos'], pg['acoas']], bold=True, bg='E8F0FE')

# ===== Sheet 9: 四三累计明细 =====
print("Sheet 9: 四三累计明细...")
ws9 = wb.create_sheet("四三累计明细")
cum43 = data_blocks['cum43Data']
headers9 = ['SKU', '上架日期', '首次出单', '分析人', '品类', '拓展类型', '本周销量', '本周销售额', '对手量', '市占比', 'PLG费率', '市场状态', '8日出单', '广告分类']
write_header(ws9, 1, headers9)
for i, d in enumerate(cum43):
    write_row(ws9, 2+i, [
        d['SKU'], d['listDate'], d['firstOrderDate'],
        d['analyst'], d['category'], d['expandType'],
        d['curSalesQty'], f"${d['curRevenue']:,.2f}" if d['curRevenue'] else '$0',
        d['curRivalQty'], f"{d['curMarketShare']}%",
        d.get('plgFee', '0%'), d['curMarketStatus'], d['cur8dStatus'],
        d.get('adClass', '')
    ], bg='EEF2FF' if i%2==0 else 'FFFFFF')

# 合计行
n9 = len(cum43)
total_sales9 = sum(d['curSalesQty'] for d in cum43)
total_rev9 = sum(d['curRevenue'] for d in cum43)
total_rival9 = sum(d['curRivalQty'] for d in cum43)
write_row(ws9, 2+n9, [f'合计({n9}条)', '', '', '', '', '', total_sales9, f'${total_rev9:,.2f}', total_rival9, '', '', '', '', ''], bold=True, bg='E8F0FE')

# ===== Sheet 10: 市场分布 =====
print("Sheet 10: 市场分布...")
ws10 = wb.create_sheet("市场分布")
mkt = data_blocks['mktDistOverall']

write_header(ws10, 1, ['市场状态', '本周数量', '本周占比', '上周数量', '上周占比', '变化'])
for i, d in enumerate(mkt['distribution']):
    if d['curCount'] > 0 or d['prevCount'] > 0:
        write_row(ws10, 2+i, [
            d['status'], d['curCount'], f"{d['curPct']}%",
            d['prevCount'], f"{d['prevPct']}%", (d['change'] >= 0 and '+' or '') + str(d['change'])
        ], bg='EEF2FF' if i%2==0 else 'FFFFFF')

# 市占比分布
row = len([d for d in mkt['distribution'] if d['curCount'] > 0 or d['prevCount'] > 0]) + 4
write_header(ws10, row, ['品线', '总SKU', '高市占比(>=75%)', '中市占比(50-75%)', '低市占比(<50%)'])
st = data_blocks['shareTierOverview']
for i, d in enumerate(st['byCategory']):
    write_row(ws10, row+1+i, [d['category'], d['total'], d['high'], d['mid'], d['low']],
              bg='EEF2FF' if i%2==0 else 'FFFFFF')

# 货值分布
po = data_blocks['priceOverview']
row2 = row + len(st['byCategory']) + 3
write_header(ws10, row2, ['价格区间', 'SKU数', '占比'])
for i, d in enumerate(po['distribution']):
    write_row(ws10, row2+1+i, [d['range'], d['count'], f"{d['pct']}%"],
              bg='EEF2FF' if i%2==0 else 'FFFFFF')

write_row(ws10, row2+len(po['distribution'])+2, [
    f"出单均价: ${po['avgPrice']:.2f} | 中位数: ${po['medianPrice']:.2f} | 有出单SKU: {po['totalWithSales']}个"
], bold=True)

# Save
wb.save(OUTPUT_FILE)
print(f"\n✅ 已保存到: {OUTPUT_FILE}")
print(f"共 10 个Sheet: 总盘概览/品线维度/分析人维度/拓展类型/分析及时率/低占比新品/PLP广告/PLG广告/四三累计明细/市场分布")
