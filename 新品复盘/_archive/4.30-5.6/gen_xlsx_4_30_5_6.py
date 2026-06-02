"""
生成 4.30-5.6 新品周报数据表 XLSX
"""
import openpyxl
import json
import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with open(r'C:\Users\Administrator\Desktop\新品复盘\report_data_4_30_5_6.json', 'r', encoding='utf-8') as f:
    D = json.load(f)

P = D['periods']
K = D['kpi']
cats = D['cats']
analysts = D['analysts']
exps = D['exps']

C_HEADER = '0F3460'
C_P1 = '6C757D'
C_P2 = '667EEA'
C_P3 = '2980B9'
C_P4 = 'C0392B'
C_HB  = 'E07B24'
C_GREEN = '08845A'
C_RED = 'C0392B'
C_WHITE = 'FFFFFF'

def hdr(ws, row, col, val, bg=C_HEADER, fg=C_WHITE, bold=True, align='center'):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, color=fg, name='Arial', size=10)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    return c

def cell(ws, row, col, val, bold=False, color=None, align='center'):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, color=color or '000000', name='Arial', size=9)
    c.alignment = Alignment(horizontal=align, vertical='center')
    return c

def add_border(ws, min_r, max_r, min_c, max_c):
    thin = Side(style='thin', color='DDDDDD')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c):
        for c in row:
            c.border = bd

def chg_sign(v):
    if v is None or v == 0: return '-'
    return f'+{v:.0f}' if v > 0 else f'{v:.0f}'

def chg_sign_pct(v):
    if v is None or v == 0: return '-'
    return f'+{v:.1f}%' if v > 0 else f'{v:.1f}%'

wb = openpyxl.Workbook()

# ========== Sheet1: KPI总览 ==========
ws1 = wb.active
ws1.title = 'KPI总览'
headers = ['指标', P[2], P[3], '环比']
for ci, h in enumerate(headers, 1):
    hdr(ws1, 1, ci, h if ci <= 2 else h, bg=C_P3 if ci==2 else (C_P4 if ci==3 else (C_HB if ci==4 else C_HEADER)))
ws1.row_dimensions[1].height = 26

kpi_data = [
    ('在跟SKU总数', K['total_skus'], K['total_skus'], '-'),
    ('总销量', round(D['qty_list'][-2]), round(D['qty_list'][-1]), chg_sign(K['qty_chg'])),
    ('总销售额(USD)', round(D['amt_list'][-2], 2), round(D['amt_list'][-1], 2), chg_sign(K['amt_chg'])),
    ('8日出单(Y)全量', K['y_all_prev'], K['y_all_new'], chg_sign(K['y_all_new'] - K['y_all_prev'])),
    ('8日外出单(N)全量', K['n_all_prev'], K['n_all_new'], chg_sign(K['n_all_new'] - K['n_all_prev'])),
    ('未出单全量', K['un_all_prev'], K['un_all_new'], chg_sign(K['un_all_new'] - K['un_all_prev'])),
    ('全量出单率', f"{D['rate_all'][-2]}%", f"{D['rate_all'][-1]}%", chg_sign_pct(D['rate_all'][-1] - D['rate_all'][-2])),
    ('有对手出单率', f"{K['rate_prev']}%", f"{K['rate_new']}%", chg_sign_pct(K['rate_chg'])),
    ('有对手SKU数', K['riv_prev'], K['riv_new'], chg_sign(K['riv_new'] - K['riv_prev'])),
    ('平均市占比', f"{K['mzb_prev']}%", f"{K['mzb_new']}%", chg_sign_pct(K['mzb_chg'])),
    ('低占比新品(mzb<75%有竞品)', '-', K['low_cnt'], '-'),
    ('PLP花费(USD)', '-', D['plp']['total']['spend'], '-'),
    ('PLP-ROAS', '-', D['plp']['total']['roas'], '-'),
    ('PLP-ACOS', '-', f"{D['plp']['total']['acos']}%", '-'),
]
for ri, (label, v1, v2, v3) in enumerate(kpi_data, 2):
    cell(ws1, ri, 1, label, bold=True, align='left')
    cell(ws1, ri, 2, v1)
    cell(ws1, ri, 3, v2, bold=True)
    c_color = None
    if v3 != '-':
        try:
            nv = float(str(v3).replace('+','').replace('%','').replace('$',''))
            c_color = C_GREEN if nv > 0 else (C_RED if nv < 0 else '888888')
        except: pass
    cell(ws1, ri, 4, v3, color=c_color)

ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 14
add_border(ws1, 1, len(kpi_data)+1, 1, 4)

# ========== Sheet2: 品类维度 ==========
ws2 = wb.create_sheet('品类维度')
cat_hdrs = ['品类', 'SKU数',
            f'{P[0]}销量', f'{P[1]}销量', f'{P[2]}销量', f'{P[3]}销量', '销量环比',
            f'{P[3]}销售额', f'{P[0]}出单率', f'{P[1]}出单率', f'{P[2]}出单率', f'{P[3]}出单率', '出单率环比']
for ci, h in enumerate(cat_hdrs, 1):
    bg = C_P1 if '4.9' in str(h) else (C_P2 if '4.16' in str(h) else (C_P3 if '4.23' in str(h) else (C_P4 if '4.30' in str(h) else (C_HB if '环比' in str(h) else C_HEADER))))
    hdr(ws2, 1, ci, h, bg=bg)
ws2.row_dimensions[1].height = 26

for ri, cat in enumerate(cats, 2):
    d = D['cat_data'][cat]
    q_chg = d['qtys'][-1] - d['qtys'][-2]
    r_chg = d['rates'][-1] - d['rates'][-2]
    vals = [cat, d['skus'],
            d['qtys'][0], d['qtys'][1], d['qtys'][2], d['qtys'][3], chg_sign(q_chg),
            f"${d['amts'][-1]:.2f}",
            f"{d['rates'][0]}%", f"{d['rates'][1]}%", f"{d['rates'][2]}%", f"{d['rates'][3]}%",
            chg_sign_pct(r_chg)]
    for ci, v in enumerate(vals, 1):
        c_color = None
        if ci == 7:
            c_color = C_GREEN if q_chg > 0 else (C_RED if q_chg < 0 else '888888')
        if ci == 13:
            c_color = C_GREEN if r_chg > 0 else (C_RED if r_chg < 0 else '888888')
        cell(ws2, ri, ci, v, bold=(ci==1), color=c_color, align='left' if ci==1 else 'center')

# 合计行
tr = len(cats) + 2
total_qs = [round(sum(D['cat_data'][c]['qtys'][i] for c in cats)) for i in range(4)]
total_ams_last = round(sum(D['cat_data'][c]['amts'][-1] for c in cats), 2)
tq_chg = total_qs[-1] - total_qs[-2]
for ci, v in enumerate(['合计', sum(D['cat_data'][c]['skus'] for c in cats)] + total_qs + [chg_sign(tq_chg), f'${total_ams_last}', '', '', '', '', ''], 1):
    cell(ws2, tr, ci, v, bold=True, color='000000')

for ci, w in enumerate([18, 8, 10, 10, 10, 10, 10, 14, 10, 10, 10, 10, 10], 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
add_border(ws2, 1, tr, 1, len(cat_hdrs))

# ========== Sheet3: 分析人维度 ==========
ws3 = wb.create_sheet('分析人维度')
ana_hdrs = ['分析人', '负责SKU',
            f'{P[0]}销量', f'{P[1]}销量', f'{P[2]}销量', f'{P[3]}销量', '销量环比',
            f'{P[0]}出单率', f'{P[1]}出单率', f'{P[2]}出单率', f'{P[3]}出单率', '出单率环比']
for ci, h in enumerate(ana_hdrs, 1):
    bg = C_P1 if '4.9' in str(h) else (C_P2 if '4.16' in str(h) else (C_P3 if '4.23' in str(h) else (C_P4 if '4.30' in str(h) else (C_HB if '环比' in str(h) else C_HEADER))))
    hdr(ws3, 1, ci, h, bg=bg)
ws3.row_dimensions[1].height = 26

for ri, ana in enumerate(analysts, 2):
    d = D['ana_data'][ana]
    q_chg = d['qtys'][-1] - d['qtys'][-2]
    r_chg = d['rates'][-1] - d['rates'][-2]
    vals = [ana, d['skus'],
            d['qtys'][0], d['qtys'][1], d['qtys'][2], d['qtys'][3], chg_sign(q_chg),
            f"{d['rates'][0]}%", f"{d['rates'][1]}%", f"{d['rates'][2]}%", f"{d['rates'][3]}%",
            chg_sign_pct(r_chg)]
    for ci, v in enumerate(vals, 1):
        c_color = None
        if ci == 7:
            c_color = C_GREEN if q_chg > 0 else (C_RED if q_chg < 0 else '888888')
        if ci == 12:
            c_color = C_GREEN if r_chg > 0 else (C_RED if r_chg < 0 else '888888')
        cell(ws3, ri, ci, v, bold=(ci==1), color=c_color, align='left' if ci==1 else 'center')

tr3 = len(analysts) + 2
add_border(ws3, 1, tr3, 1, len(ana_hdrs))
for ci, w in enumerate([12, 8, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10], 1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

# ========== Sheet4: 拓展类型 ==========
ws4 = wb.create_sheet('拓展类型')
exp_hdrs = ['拓展类型', 'SKU数',
            f'{P[0]}销量', f'{P[1]}销量', f'{P[2]}销量', f'{P[3]}销量', '销量环比',
            f'{P[0]}出单率', f'{P[1]}出单率', f'{P[2]}出单率', f'{P[3]}出单率', '出单率环比']
for ci, h in enumerate(exp_hdrs, 1):
    bg = C_P1 if '4.9' in str(h) else (C_P2 if '4.16' in str(h) else (C_P3 if '4.23' in str(h) else (C_P4 if '4.30' in str(h) else (C_HB if '环比' in str(h) else C_HEADER))))
    hdr(ws4, 1, ci, h, bg=bg)
ws4.row_dimensions[1].height = 26

ri = 2
for exp in exps:
    d = D['exp_data'][exp]
    if d['skus'] == 0: continue
    q_chg = d['qtys'][-1] - d['qtys'][-2]
    r_chg = d['rates'][-1] - d['rates'][-2]
    vals = [exp, d['skus'],
            d['qtys'][0], d['qtys'][1], d['qtys'][2], d['qtys'][3], chg_sign(q_chg),
            f"{d['rates'][0]}%", f"{d['rates'][1]}%", f"{d['rates'][2]}%", f"{d['rates'][3]}%",
            chg_sign_pct(r_chg)]
    for ci, v in enumerate(vals, 1):
        c_color = None
        if ci == 7:
            c_color = C_GREEN if q_chg > 0 else (C_RED if q_chg < 0 else '888888')
        if ci == 12:
            c_color = C_GREEN if r_chg > 0 else (C_RED if r_chg < 0 else '888888')
        cell(ws4, ri, ci, v, bold=(ci==1), color=c_color, align='left' if ci==1 else 'center')
    ri += 1

add_border(ws4, 1, ri-1, 1, len(exp_hdrs))
for ci, w in enumerate([12, 8, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10], 1):
    ws4.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

# ========== Sheet5: 分析及时率 ==========
ws5 = wb.create_sheet('分析及时率')
timely_hdrs = ['分析情况', P[0], P[1], P[2], P[3], '环比']
for ci, h in enumerate(timely_hdrs, 1):
    hdr(ws5, 1, ci, h)
ws5.row_dimensions[1].height = 26

timely_data = [
    ('及时分析', D['timely_in']),
    ('超7日未分析', D['timely_over7']),
    ('8日内新品无分析', D['timely_new']),
]
for ri, (label, vals) in enumerate(timely_data, 2):
    chg = vals[-1] - vals[-2]
    row_vals = [label] + vals + [chg_sign(chg)]
    for ci, v in enumerate(row_vals, 1):
        c_color = C_GREEN if (ci == 6 and chg > 0) else (C_RED if (ci == 6 and chg < 0) else None)
        cell(ws5, ri, ci, v, bold=(ci==1), color=c_color, align='left' if ci==1 else 'center')

# 合计
total_timely = [D['timely_in'][i] + D['timely_over7'][i] + D['timely_new'][i] for i in range(4)]
tchg = total_timely[-1] - total_timely[-2]
for ci, v in enumerate(['合计'] + total_timely + [chg_sign(tchg)], 1):
    cell(ws5, 5, ci, v, bold=True)

add_border(ws5, 1, 5, 1, 6)
for ci, w in enumerate([20, 12, 12, 12, 12, 10], 1):
    ws5.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

# ========== Sheet6: 新品出单情况 ==========
ws6 = wb.create_sheet('新品出单情况')
order_hdrs = ['出单情况', P[0], P[1], P[2], P[3], '环比']
for ci, h in enumerate(order_hdrs, 1):
    hdr(ws6, 1, ci, h)

order_rows = [
    ('8日出单(Y)', D['order_y']),
    ('8日外出单(N)', D['order_n']),
    ('未出单', D['order_un']),
]
for ri, (label, vals) in enumerate(order_rows, 2):
    chg = vals[-1] - vals[-2]
    row_vals = [label] + vals + [chg_sign(chg)]
    for ci, v in enumerate(row_vals, 1):
        cell(ws6, ri, ci, v, bold=(ci==1), align='left' if ci==1 else 'center')

add_border(ws6, 1, 4, 1, 6)
for ci, w in enumerate([16, 12, 12, 12, 12, 10], 1):
    ws6.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

# ========== Sheet7: 新品未出单 ==========
ws7 = wb.create_sheet('新品未出单')
unorder_hdrs = ['未出单原因', P[0], P[1], P[2], P[3], '环比']
for ci, h in enumerate(unorder_hdrs, 1):
    hdr(ws7, 1, ci, h)

unorder_reasons = ['竞争无优势', '无市场', '站内无价格优势', '站外出单']
for ri, rsn in enumerate(unorder_reasons, 2):
    vals = D['unorder_data'].get(rsn, [0, 0, 0, 0])
    chg = vals[-1] - vals[-2]
    row_vals = [rsn] + vals + [chg_sign(chg)]
    for ci, v in enumerate(row_vals, 1):
        cell(ws7, ri, ci, v, bold=(ci==1), align='left' if ci==1 else 'center')

add_border(ws7, 1, 5, 1, 6)
for ci, w in enumerate([18, 12, 12, 12, 12, 10], 1):
    ws7.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

# ========== Sheet8: PLP复盘 ==========
ws8 = wb.create_sheet('PLP复盘')
plp_hdrs = ['维度', '广告SKU', '曝光量', '点击量', '订单数', '花费(USD)', '销售额(USD)', 'ROAS', 'CVR%', 'CTR%', 'ACOS%']
for ci, h in enumerate(plp_hdrs, 1):
    hdr(ws8, 1, ci, h, bg='8E44AD')
ws8.row_dimensions[1].height = 26

def write_plp_row(ws, row, label, d):
    vals = [label, d.get('skus', '-'), d['imp'], d['clk'], d['ord'],
            round(d['spend'], 2), round(d['sales'], 2),
            d['roas'], f"{d['cvr']}%", f"{d['ctr']:.2f}%", f"{d['acos']}%"]
    for ci, v in enumerate(vals, 1):
        cell(ws, row, ci, v, bold=(ci==1), align='left' if ci==1 else 'center')

ri = 2
hdr(ws8, ri, 1, '── 总计 ──', bg='2C3E50', bold=True)
ri += 1
write_plp_row(ws8, ri, '4.30-5.6合计', D['plp']['total'])
ri += 2

for section_title, data_dict in [('── 按分析人 ──', D['plp']['by_analyst']),
                                   ('── 按品类 ──', D['plp']['by_cat']),
                                   ('── 按拓展类型 ──', D['plp']['by_expn'])]:
    hdr(ws8, ri, 1, section_title, bg='2C3E50', bold=True)
    ri += 1
    for k in sorted(data_dict.keys()):
        write_plp_row(ws8, ri, k, data_dict[k])
        ri += 1
    ri += 1

add_border(ws8, 1, ri-1, 1, len(plp_hdrs))
for ci, w in enumerate([16, 10, 12, 10, 10, 14, 14, 10, 10, 10, 10], 1):
    ws8.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

# ========== Sheet9: PLG费率 ==========
ws9 = wb.create_sheet('PLG费率')
plg_hdrs = ['分析人', '有效SKU数', '平均费率', '最高费率']
for ci, h in enumerate(plg_hdrs, 1):
    hdr(ws9, 1, ci, h, bg='8E44AD')

for ri, (ana, d) in enumerate(sorted(D['plg']['by_analyst'].items()), 2):
    cell(ws9, ri, 1, ana, bold=True, align='left')
    cell(ws9, ri, 2, d['count'])
    cell(ws9, ri, 3, f"{d['avg']}%")
    cell(ws9, ri, 4, f"{d['max']}%")

add_border(ws9, 1, len(D['plg']['by_analyst'])+1, 1, 4)
for ci, w in enumerate([12, 12, 12, 12], 1):
    ws9.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

# ========== Sheet10: 低占比新品明细 ==========
ws10 = wb.create_sheet('低占比新品')
low_hdrs = ['销售编号', 'SKU', '品类', '拓展类型', '上架日期', '首次出单', '分析人',
            f'{P[3]}销量', '5.6对手销量', '5.6市占比', '8日出单', '市场状态', '操作判断']
for ci, h in enumerate(low_hdrs, 1):
    hdr(ws10, 1, ci, h, bg=C_RED)

for ri, r in enumerate(D['low_share_rows'], 2):
    for ci, v in enumerate(r, 1):
        c_color = None
        if ci == 10:
            try:
                mzb_val = float(str(v).replace('%', ''))
                c_color = C_RED if mzb_val < 25 else (C_HB if mzb_val < 50 else '888888')
            except: pass
        cell(ws10, ri, ci, v, color=c_color, align='left' if ci in [2, 6, 11, 12, 13] else 'center')

add_border(ws10, 1, len(D['low_share_rows'])+1, 1, len(low_hdrs))
for ci, w in enumerate([10, 22, 14, 10, 12, 12, 10, 12, 12, 12, 10, 14, 14], 1):
    ws10.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

# ── 保存 ──
path = r'C:\Users\Administrator\Desktop\新品复盘\新品周报数据表_4.30-5.6.xlsx'
wb.save(path)
print(f'XLSX saved: {path}')
