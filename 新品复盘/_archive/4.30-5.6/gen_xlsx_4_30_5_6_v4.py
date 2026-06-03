"""
新品周报 4.30-5.6 汇总XLSX v4（排版优化版）
基于 Skill v2.0 + 全面格式化优化
"""
import openpyxl
import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ═══════════════════ 配色方案 ═══════════════════
C_HDR   = '1a3a5c'   # 深蓝主色（更沉稳）
C_TITLE = '0d2137'   # 标题深色
C_P1    = '546e7a'   # 4.9-4.15 灰蓝
C_P2    = '4a6fa5'   # 4.16-4.22 中蓝
C_P3    = '3a8bb5'   # 4.23-4.29 亮蓝
C_P4    = 'c0392b'   # 4.30-5.6 红色（当期突出）
C_HB    = 'e67e22'   # 环比 橙色
C_GRN   = '1b7a3d'   # 正向绿色
C_RED   = 'c0392b'   # 负向红色
C_WHT   = 'FFFFFF'
C_PLP   = '6c3483'   # PLP紫色
C_ORG   = 'bf360c'   # A板块深橙
C_GRN2  = '2e6b3e'   # B板块深绿
C_LTG   = 'f5f5f5'   # 浅灰（斑马纹）
C_HLT   = 'fff3cd'   # 浅黄高亮

PCOLORS = [C_P1, C_P2, C_P3, C_P4]

# ═══════════════════ 样式常量 ═══════════════════
THIN_BORDER = Border(
    left=Side(style='thin', color='d0d0d0'),
    right=Side(style='thin', color='d0d0d0'),
    top=Side(style='thin', color='d0d0d0'),
    bottom=Side(style='thin', color='d0d0d0'),
)
BOTTOM_BORDER = Border(bottom=Side(style='medium', color='1a3a5c'))

FONT_TITLE  = Font(bold=True, color=C_WHT, name='Microsoft YaHei', size=13)
FONT_HDR    = Font(bold=True, color=C_WHT, name='Microsoft YaHei', size=10)
FONT_SEC    = Font(bold=True, color=C_WHT, name='Microsoft YaHei', size=11)
FONT_BOLD   = Font(bold=True, color='1a1a1a', name='Microsoft YaHei', size=9)
FONT_NORMAL = Font(color='2d2d2d', name='Microsoft YaHei', size=9)
FONT_SMALL  = Font(color='666666', name='Microsoft YaHei', size=8)
FONT_PCT    = Font(color='2d2d2d', name='Microsoft YaHei', size=9)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=False)
ALIGN_LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=False)
ALIGN_WRAP   = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ═══════════════════ 样式函数 ═══════════════════
def set_cell(ws, r, c, val, font=FONT_NORMAL, fill=None, align=ALIGN_CENTER, border=THIN_BORDER):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = font
    if fill: cl.fill = fill
    cl.alignment = align
    if border: cl.border = border
    return cl

def hdr_cell(ws, r, c, val, bg=C_HDR, font=FONT_HDR, align=ALIGN_CENTER):
    return set_cell(ws, r, c, val, font=font,
                    fill=PatternFill('solid', fgColor=bg), align=align)

def title_row(ws, r, c1, cn, text, bg=C_TITLE):
    """合并标题行"""
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=cn)
    set_cell(ws, r, c1, text, font=FONT_TITLE,
             fill=PatternFill('solid', fgColor=bg), align=Alignment(horizontal='left', vertical='center'))

def section_row(ws, r, c1, cn, text, bg=C_HDR):
    """分区标题行"""
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=cn)
    set_cell(ws, r, c1, text, font=FONT_SEC,
             fill=PatternFill('solid', fgColor=bg), align=Alignment(horizontal='left', vertical='center'))

def data_row(ws, r, vals, bold_cols=None, colors=None, aligns=None, highlight=False):
    """写入一行数据"""
    bg = PatternFill('solid', fgColor=C_HLT) if highlight else None
    for ci, v in enumerate(vals, 1):
        font = FONT_BOLD if (bold_cols and ci in bold_cols) else FONT_NORMAL
        if v is not None and isinstance(v, str) and v.startswith('+'):
            font = Font(bold=True, color=C_GRN if v.startswith('+') else C_RED, name='Microsoft YaHei', size=9)
        elif v is not None and isinstance(v, str) and v.startswith('-') and len(v) > 1:
            try: float(v.replace('%','').replace('$',''))
            except: pass
            else: font = Font(color=C_RED, name='Microsoft YaHei', size=9)
        if colors and ci in colors: font = Font(bold=True, color=colors[ci], name='Microsoft YaHei', size=9)
        al = aligns.get(ci, ALIGN_LEFT if ci == 1 else ALIGN_CENTER) if aligns else (ALIGN_LEFT if ci == 1 else ALIGN_CENTER)
        set_cell(ws, r, ci, v, font=font, fill=bg, align=al)

def auto_width(ws, min_widths, max_col, extra=2):
    """设置列宽（最小宽度+余量）"""
    for ci, mw in enumerate(min_widths, 1):
        if ci > max_col: break
        ws.column_dimensions[get_column_letter(ci)].width = mw + extra

def freeze_header(ws):
    ws.freeze_panes = 'A2'

# ═══════════════════ 数据读取（同v3）══════════════════
def n(v, default=0):
    if v is None or v == '' or v == '-': return default
    if isinstance(v, str):
        if any(c in v for c in '#—∑Σ'): return default
        v = v.replace(',','').replace('$','').replace('%','')
    try: return float(v)
    except: return default

def pct(v):
    if v is None or v == '' or v == '-': return None
    if isinstance(v, str):
        if any(c in v for c in '#'): return None
        v = v.replace('%','')
    try: fv = float(v); return round(fv*100,2) if fv<=1.0 else round(fv,2)
    except: return None

def get_date(v):
    if v is None: return None
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    return None

def chg_sign(v, as_pct=False):
    if v is None or v == 0: return '-'
    if as_pct: return f'+{v:.1f}%' if v>0 else f'{v:.1f}%'
    return f'+{v:.0f}' if v>0 else f'{v:.0f}'

def chg_color(v):
    if v is None or v == 0: return '888888'
    return C_GRN if v > 0 else C_RED

SRC = r'c:\Users\Hardy\ai-projects\新品复盘\新品检查周源数据和PLP数据.xlsx'
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_main = wb_src['四三数据累计']
ws_plp  = wb_src['PLP明细']

PERIODS = ['4.9-4.15', '4.16-4.22', '4.23-4.29', '4.30-5.6']
CUTOFFS = [datetime.date(2026,4,15), datetime.date(2026,4,22), datetime.date(2026,4,29), datetime.date(2026,5,6)]

Q_COLS=[12,13,14,15]; AMT_COLS=[22,23,24,25]; RIV_COLS=[32,33,34,35]; MZB_COLS=[41,42,43,44]
ORD_COLS=[59,60,61,62]; FQ_COLS=[68,69,70,71]; NFQ_COLS=[77,78,79,80]
MKT_COLS=[87,88,89,90]; OP_COLS=[96,97,98,99]; PLP_COLS=[100,101,102,103]; PLG_COLS=[104,105,106]

class SKU: pass
skus = []
for row in ws_main.iter_rows(min_row=2, values_only=True):
    if not row[1]: continue
    s = SKU()
    s.sid=row[0]; s.sku=str(row[1]).strip()
    s.list_date=get_date(row[2]); s.first_order=get_date(row[3])
    s.analyst=str(row[4] or '').strip(); s.category=str(row[5] or '').strip()
    s.exp_type=str(row[6] or '').strip()
    s.qtys=[n(row[c]) for c in Q_COLS]; s.amts=[n(row[c]) for c in AMT_COLS]
    s.rivals=[n(row[c]) for c in RIV_COLS]; s.mzbs=[pct(row[c]) for c in MZB_COLS]
    s.ord8s=[str(row[c] or '').strip() for c in ORD_COLS]
    s.freqs=[str(row[c] or '').strip() for c in FQ_COLS]
    s.new_freqs=[str(row[c] or '').strip() for c in NFQ_COLS]
    s.mkts=[str(row[c] or '').strip() for c in MKT_COLS]
    s.ops=[str(row[c] or '').strip() for c in OP_COLS]
    s.plp=[str(row[c] or '').strip() for c in PLP_COLS]
    s.plg=[n(row[c]) for c in PLG_COLS]
    skus.append(s)

def valid_for(s, pi):
    if not s.list_date or s.list_date>CUTOFFS[pi]: return False
    if s.ord8s[pi]=='未上架': return False
    return True
def has_rival(s, pi): return s.rivals[pi]>0
def is_y(s, pi): return s.ord8s[pi]=='Y'
def is_n(s, pi): return s.ord8s[pi]=='N'
def is_unordered(s, pi): return s.ord8s[pi]=='未出单'

# ═══════════════════ 核心计算 ═══════════════════
valid_cnt=[sum(1 for s in skus if valid_for(s,pi)) for pi in range(4)]
total_q=[round(sum(s.qtys[pi] for s in skus if valid_for(s,pi))) for pi in range(4)]
total_a=[round(sum(s.amts[pi] for s in skus if valid_for(s,pi)),2) for pi in range(4)]
riv_cnt=[sum(1 for s in skus if valid_for(s,pi) and has_rival(s,pi)) for pi in range(4)]
noriv_cnt=[sum(1 for s in skus if valid_for(s,pi) and not has_rival(s,pi)) for pi in range(4)]
yr_cnt=[sum(1 for s in skus if valid_for(s,pi) and has_rival(s,pi) and is_y(s,pi)) for pi in range(4)]
nr_cnt=[sum(1 for s in skus if valid_for(s,pi) and has_rival(s,pi) and is_n(s,pi)) for pi in range(4)]
unr_cnt=[sum(1 for s in skus if valid_for(s,pi) and has_rival(s,pi) and is_unordered(s,pi)) for pi in range(4)]
riv_rate=[round((yr_cnt[i]+nr_cnt[i])/riv_cnt[i]*100,1) if riv_cnt[i]>0 else 0 for i in range(4)]
y_all=[sum(1 for s in skus if valid_for(s,pi) and is_y(s,pi)) for pi in range(4)]
n_all=[sum(1 for s in skus if valid_for(s,pi) and is_n(s,pi)) for pi in range(4)]
un_all=[sum(1 for s in skus if valid_for(s,pi) and is_unordered(s,pi)) for pi in range(4)]
rate_all=[]
for pi in range(4):
    tot=y_all[pi]+n_all[pi]+un_all[pi]
    rate_all.append(round((y_all[pi]+n_all[pi])/tot*100,1) if tot>0 else 0)
new_list_curr=sum(1 for s in skus if s.list_date and datetime.date(2026,4,30)<=s.list_date<=datetime.date(2026,5,6))

def avg_mzb(pi):
    vals=[s.mzbs[pi] for s in skus if valid_for(s,pi) and s.mzbs[pi] is not None]
    return round(sum(vals)/len(vals),1) if vals else 0
mzb_list=[avg_mzb(pi) for pi in range(4)]

low_skus=[s for s in skus if valid_for(s,3) and s.mzbs[3] is not None and s.mzbs[3]<75 and has_rival(s,3)]
low_skus.sort(key=lambda s: s.mzbs[3] if s.mzbs[3] else 100)

timely_in=[]; timely_n8=[]; timely_o7=[]
for pi in range(4):
    vs=[s for s in skus if valid_for(s,pi)]
    ti=tn=to=0
    for s in vs:
        if s.new_freqs[pi]=='异常': tn+=1
        elif s.freqs[pi]=='异常': to+=1
        else: ti+=1
    timely_in.append(ti); timely_n8.append(tn); timely_o7.append(to)

timely_total=[timely_in[i]+timely_n8[i]+timely_o7[i] for i in range(4)]
timely_rate=[round(timely_in[i]/timely_total[i]*100,1) if timely_total[i]>0 else 0 for i in range(4)]

timely_by_ana={}
for ana in sorted(set(s.analyst for s in skus if s.analyst)):
    vs=[s for s in skus if s.analyst==ana and valid_for(s,3)]
    ti=sum(1 for s in vs if s.new_freqs[3]!='异常' and s.freqs[3]!='异常')
    tn=sum(1 for s in vs if s.new_freqs[3]=='异常')
    to=sum(1 for s in vs if s.new_freqs[3]!='异常' and s.freqs[3]=='异常')
    tot=len(vs)
    timely_by_ana[ana]={'total':tot,'ti':ti,'tn':tn,'to':to,'rate':round(ti/tot*100,1) if tot>0 else 0}

def build_dim(rows, key_fn, periods=4):
    groups=defaultdict(list)
    for s in rows: groups[key_fn(s)].append(s)
    result={}
    for k,grp in sorted(groups.items()):
        d={'total':len(grp)}
        for pi in range(periods):
            vs=[s for s in grp if valid_for(s,pi)]
            d[f'v{pi}']=len(vs); d[f'q{pi}']=round(sum(s.qtys[pi] for s in vs))
            d[f'a{pi}']=round(sum(s.amts[pi] for s in vs),2)
            hr=[s for s in vs if has_rival(s,pi)]
            d[f'r{pi}']=len(hr)
            yr=sum(1 for s in hr if is_y(s,pi)); nr=sum(1 for s in hr if is_n(s,pi))
            d[f'yr{pi}']=yr; d[f'nr{pi}']=nr
            num=yr+nr; denom=len(hr)
            d[f'rate{pi}']=round(num/denom*100,1) if denom>0 else 0
        result[k]=d
    return result

cats=sorted(set(s.category for s in skus if s.category))
analysts_list=['俞东旭','张潇','朱培源','王偲涵','章鹏','胡煜星']
exp_types=['原开品','拓展品','组合件']
cat_data=build_dim(skus, lambda s: s.category)
ana_data=build_dim(skus, lambda s: s.analyst)
exp_data=build_dim(skus, lambda s: s.exp_type)

# PLP
plp_all=[row for row in ws_plp.iter_rows(min_row=2,values_only=True) if row[0] and '4.30' in str(row[0])]
def clean_plp(row):
    try: return {'sku':row[2],'ana':str(row[8] or ''),'cat':str(row[9] or ''),'exp':str(row[10] or ''),'imp':n(row[11]),'clk':n(row[12]),'ord':n(row[13]),'spend':n(row[14]),'sales':n(row[15])}
    except: return None
plp_d=[r for row in plp_all if str(row[1] or '') not in ('总数据','总计') and row[2] and not str(row[2]).startswith('广告') and (r:=clean_plp(row)) and r['sku']]

def agg_plp(rows,key_fn):
    agg={}
    for r in rows:
        k=key_fn(r)
        if k not in agg: agg[k]={'imp':0,'clk':0,'ord':0,'spend':0,'sales':0,'sku_set':set()}
        a=agg[k]; a['imp']+=r['imp']; a['clk']+=r['clk']; a['ord']+=r['ord']; a['spend']+=r['spend']; a['sales']+=r['sales']; a['sku_set'].add(r['sku'])
    result={}
    for k,a in agg.items():
        roas=round(a['sales']/a['spend'],2) if a['spend']>0 else 0
        cvr=round(a['ord']/a['clk']*100,2) if a['clk']>0 else 0
        ctr=round(a['clk']/a['imp']*100,4) if a['imp']>0 else 0
        acos=round(a['spend']/a['sales']*100,2) if a['sales']>0 else 0
        cpc=round(a['spend']/a['clk'],2) if a['clk']>0 else 0
        cpa=round(a['spend']/a['ord'],2) if a['ord']>0 else 0
        result[k]={'skus':len(a['sku_set']),'imp':int(a['imp']),'clk':int(a['clk']),'ord':int(a['ord']),'spend':round(a['spend'],2),'sales':round(a['sales'],2),'roas':roas,'cvr':cvr,'ctr':ctr,'acos':acos,'cpc':cpc,'cpa':cpa}
    return result

plp_ana=agg_plp(plp_d,lambda r: r['ana']); plp_cat=agg_plp(plp_d,lambda r: r['cat']); plp_exp=agg_plp(plp_d,lambda r: r['exp'])
total_imp=sum(d['imp'] for d in plp_ana.values()); total_clk=sum(d['clk'] for d in plp_ana.values())
total_ord=sum(d['ord'] for d in plp_ana.values()); total_spend=sum(d['spend'] for d in plp_ana.values())
total_sales=sum(d['sales'] for d in plp_ana.values())
plp_total={'skus':len(plp_d),'imp':total_imp,'clk':total_clk,'ord':total_ord,'spend':round(total_spend,2),'sales':round(total_sales,2),'roas':round(total_sales/total_spend,2) if total_spend>0 else 0,'cvr':round(total_ord/total_clk*100,2) if total_clk>0 else 0,'ctr':round(total_clk/total_imp*100,4) if total_imp>0 else 0,'acos':round(total_spend/total_sales*100,2) if total_sales>0 else 0,'cpc':round(total_spend/total_clk,2) if total_clk>0 else 0,'cpa':round(total_spend/total_ord,2) if total_ord>0 else 0}

# 上周期PLP
plp_prev_rows=[row for row in ws_plp.iter_rows(min_row=2,values_only=True) if row[0] and ('4.23' in str(row[0]))]
plp_prev_d=[r for row in plp_prev_rows if str(row[1] or '') not in ('总数据','总计') and row[2] and not str(row[2]).startswith('广告') and (r:=clean_plp(row)) and r['sku']]
if plp_prev_d:
    ip=sum(r['imp'] for r in plp_prev_d); cp=sum(r['clk'] for r in plp_prev_d)
    op=sum(r['ord'] for r in plp_prev_d); sp=sum(r['spend'] for r in plp_prev_d)
    sl=sum(r['sales'] for r in plp_prev_d)
    plp_prev={'skus':len(plp_prev_d),'imp':ip,'clk':cp,'ord':op,'spend':round(sp,2),'sales':round(sl,2),'roas':round(sl/sp,2) if sp>0 else 0,'cvr':round(op/cp*100,2) if cp>0 else 0,'ctr':round(cp/ip*100,4) if ip>0 else 0,'acos':round(sp/sl*100,2) if sl>0 else 0,'cpc':round(sp/cp,2) if cp>0 else 0,'cpa':round(sp/op,2) if op>0 else 0}
else: plp_prev=dict(plp_total)

# PLG
plg_cur=[s for s in skus if valid_for(s,3) and s.plg[2]>0]
plg_ana_d=defaultdict(list); plg_cat_d=defaultdict(list)
for s in plg_cur: plg_ana_d[s.analyst or '未知'].append(s.plg[2]); plg_cat_d[s.category or '未分类'].append(s.plg[2])
plg_ana_agg={k:{'cnt':len(v),'avg':round(sum(v)/len(v)*100,1),'max':round(max(v)*100,1)} for k,v in plg_ana_d.items()}
plg_cat_agg={k:{'cnt':len(v),'avg':round(sum(v)/len(v)*100,1),'max':round(max(v)*100,1)} for k,v in plg_cat_d.items()}

# 未出单 A/B
mkt_has=['竞争无优势','无市场','站内无价格优势','站外出单','正常','#N/A','未知']
mkt_no=['无市场','未知','竞争无优势','#N/A','其他']
def classify(mkt, order_list):
    if not mkt: return '未知'
    for o in order_list:
        if o in mkt: return o
    if '价格' in mkt: return '站内无价格优势'
    if '站外' in mkt: return '站外出单'
    if '竞争' in mkt: return '竞争无优势'
    if '无市场' in mkt: return '无市场'
    return '其他'

has_r_curr=[s for s in skus if valid_for(s,3) and is_unordered(s,3) and has_rival(s,3)]
no_r_curr=[s for s in skus if valid_for(s,3) and is_unordered(s,3) and not has_rival(s,3)]
has_r_prev=[s for s in skus if valid_for(s,2) and is_unordered(s,2) and has_rival(s,2)]
no_r_prev=[s for s in skus if valid_for(s,2) and is_unordered(s,2) and not has_rival(s,2)]

def reason_dist(slist, pi, olist):
    dist={r:0 for r in olist}
    for s in slist:
        m=classify(s.mkts[pi], olist)
        if m in dist: dist[m]+=1
        else: dist['其他']+=1
    return dist
a1c=reason_dist(has_r_curr,3,mkt_has); a1p=reason_dist(has_r_prev,2,mkt_has)
b1c=reason_dist(no_r_curr,3,mkt_no); b1p=reason_dist(no_r_prev,2,mkt_no)

def ana_rsn(slist, pi, olist):
    r={}
    for ana in analysts_list:
        al=[s for s in slist if s.analyst==ana]
        d={r:0 for r in olist}
        for s in al:
            m=classify(s.mkts[pi],olist)
            if m in d: d[m]+=1
            else: d['其他']+=1
        r[ana]={'total':len(al),'dist':d}
    return r

def cat_rsn(slist, pi, olist):
    r={}
    for cat in cats:
        cl=[s for s in slist if s.category==cat]
        d={r:0 for r in olist}
        for s in cl:
            m=classify(s.mkts[pi],olist)
            if m in d: d[m]+=1
            else: d['其他']+=1
        r[cat]={'total':len(cl),'dist':d}
    return r

a2c=ana_rsn(has_r_curr,3,mkt_has); a3c=cat_rsn(has_r_curr,3,mkt_has)
b2c=ana_rsn(no_r_curr,3,mkt_no); b3c=cat_rsn(no_r_curr,3,mkt_no)
wb_src.close()

# ══════════════════════════════════════════════════════════════
# 生成 XLSX（排版优化）
# ══════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ============ Sheet 1: 总体数据 ============
ws1 = wb.active
ws1.title = '总体数据'
NC1 = 6

# 全局标题
title_row(ws1, 1, 1, NC1, f'新品周报 4.30-5.6  总体数据总览')
ws1.row_dimensions[1].height = 32

# ── 一、总体概况 ──
section_row(ws1, 3, 1, NC1, '一、总体概况')
ws1.row_dimensions[3].height = 24
# 表头
hdr_hdrs = ['指标', PERIODS[0], PERIODS[1], PERIODS[2], PERIODS[3], '环比']
for ci, h in enumerate(hdr_hdrs, 1): hdr_cell(ws1, 4, ci, h)
ws1.row_dimensions[4].height = 22

overview = [
    ('累计SKU数（截止日）', valid_cnt, True),
    ('本周新上架SKU', ['-','-','-',new_list_curr], False),
    ('总销量', total_q, True),
    ('总销售额 (USD)', [f'${v:,.2f}' for v in total_a], False),
    ('有对手SKU数', riv_cnt, True),
    ('无对手SKU数', noriv_cnt, True),
]
for ri, (label, vals, is_num) in enumerate(overview, 5):
    row_vals = [label] + list(vals)
    if is_num:
        chg = vals[3]-vals[2] if len(vals)>=4 else 0
        row_vals.append(chg_sign(chg))
    else:
        row_vals.append('-')
    data_row(ws1, ri, row_vals, bold_cols={1})
    ws1.row_dimensions[ri].height = 20

# ── 二、分析及时率 ──
sr2 = 12
section_row(ws1, sr2, 1, NC1, '二、分析及时率')
ws1.row_dimensions[sr2].height = 24
for ci, h in enumerate(hdr_hdrs, 1): hdr_cell(ws1, sr2+1, ci, h)

timely_rows = [
    ('及时分析产品数', timely_in, True),
    ('8日内新品无分析', timely_n8, True),
    ('超7日低占比未分析', timely_o7, True),
    ('统计总数', timely_total, True),
    ('及时分析率', [f'{v}%' for v in timely_rate], False),
]
for ri, (label, vals, is_num) in enumerate(timely_rows, sr2+2):
    row_vals = [label] + list(vals)
    if is_num: row_vals.append(chg_sign(vals[3]-vals[2]))
    else: row_vals.append('-')
    data_row(ws1, ri, row_vals, bold_cols={1}, highlight=(label=='及时分析率'))
    ws1.row_dimensions[ri].height = 20

# ── 三、新品出单情况（有对手口径）──
sr3 = sr2 + 2 + len(timely_rows) + 1
section_row(ws1, sr3, 1, NC1, '三、新品出单情况（有对手口径）')
ws1.row_dimensions[sr3].height = 24
for ci, h in enumerate(hdr_hdrs, 1): hdr_cell(ws1, sr3+1, ci, h)

ord_rows = [
    ('有对手总SKU', riv_cnt, True),
    ('8日内出单（Y）', yr_cnt, True),
    ('8日外出单（N）', nr_cnt, True),
    ('真正未出单', unr_cnt, True),
    ('已出单合计 (Y+N)', [yr_cnt[i]+nr_cnt[i] for i in range(4)], True),
    ('出单率', [f'{v}%' for v in riv_rate], False),
]
for ri, (label, vals, is_num) in enumerate(ord_rows, sr3+2):
    row_vals = [label] + list(vals)
    if is_num: row_vals.append(chg_sign(vals[3]-vals[2]))
    else: row_vals.append('-')
    data_row(ws1, ri, row_vals, bold_cols={1}, highlight=(label=='出单率'))
    ws1.row_dimensions[ri].height = 20

auto_width(ws1, [32, 16, 16, 16, 16, 12], NC1)
freeze_header(ws1)

# ============ Sheet 2: 品线维度 ============
ws2 = wb.create_sheet('品线维度')
NC2 = 2 + 4*5 + 2  # 累计SKU + 4期×(上架SKU+销量+销售额+有对手+出单率) + 2个环比
title_row(ws2, 1, 1, NC2, f'品线维度 - 新品周报 4.30-5.6')
ws2.row_dimensions[1].height = 28

# 表头行
headers2 = ['品线', '累计SKU']
for pi in range(4):
    headers2 += [f'{PERIODS[pi]}\n上架SKU', f'{PERIODS[pi]}\n销量', f'{PERIODS[pi]}\n销售额', f'{PERIODS[pi]}\n有对手', f'{PERIODS[pi]}\n出单率']
headers2 += ['销量环比', '出单率环比']
for ci, h in enumerate(headers2, 1):
    bg = C_HDR
    for pi in range(4):
        if PERIODS[pi] in str(h): bg = PCOLORS[pi]; break
    if '环比' in str(h): bg = C_HB
    hdr_cell(ws2, 2, ci, h, bg=bg)
ws2.row_dimensions[2].height = 36

for ri, cat in enumerate(cats, 3):
    d = cat_data[cat]
    qchg = d['q3']-d['q2']; rchg = d['rate3']-d['rate2']
    vals = [cat, d['total']]
    for pi in range(4):
        vals += [d[f'v{pi}'], d[f'q{pi}'], f'${d[f"a{pi}"]:,.2f}', d[f'r{pi}'], f'{d[f"rate{pi}"]}%']
    vals += [chg_sign(qchg), chg_sign(rchg, as_pct=True)]
    data_row(ws2, ri, vals, bold_cols={1}, highlight=(ri%2==1))
    ws2.row_dimensions[ri].height = 18

# 合计
tr2 = len(cats)+3
ts = ['合计'] + [sum(cat_data[c]['total'] for c in cats)]
for pi in range(4):
    ts += [sum(cat_data[c][f'v{pi}'] for c in cats), round(sum(cat_data[c][f'q{pi}'] for c in cats)), f'${round(sum(cat_data[c][f"a{pi}"] for c in cats),2):,.2f}', sum(cat_data[c][f'r{pi}'] for c in cats), '']
ts += [chg_sign(sum(cat_data[c]['q3'] for c in cats)-sum(cat_data[c]['q2'] for c in cats)), '']
data_row(ws2, tr2, ts, bold_cols=set(range(1,NC2+1)))
ws2.row_dimensions[tr2].height = 20

auto_width(ws2, [14,10]+[10,9,12,8,10]*4+[10,10], NC2)
freeze_header(ws2)

# ============ Sheet 3: 分析人维度 ============
ws3 = wb.create_sheet('分析人维度')
# 简化列: 分析人 + 累计SKU + 4×(上架SKU+销量+出单率) + 2环比 = 1+1+12+2=16
NC3 = 16
title_row(ws3, 1, 1, NC3, f'分析人维度 - 新品周报 4.30-5.6')
ws3.row_dimensions[1].height = 28

headers3 = ['分析人', '累计SKU']
for pi in range(4):
    headers3 += [f'{PERIODS[pi]}\n上架SKU', f'{PERIODS[pi]}\n销量', f'{PERIODS[pi]}\n出单率']
headers3 += ['销量环比', '出单率环比']
for ci, h in enumerate(headers3, 1):
    bg = C_HDR
    for pi in range(4):
        if PERIODS[pi] in str(h): bg = PCOLORS[pi]; break
    if '环比' in str(h): bg = C_HB
    hdr_cell(ws3, 2, ci, h, bg=bg)
ws3.row_dimensions[2].height = 36

for ri, ana in enumerate(analysts_list, 3):
    d = ana_data.get(ana)
    if not d: continue
    qchg = d['q3']-d['q2']; rchg = d['rate3']-d['rate2']
    vals = [ana, d['total']]
    for pi in range(4): vals += [d[f'v{pi}'], d[f'q{pi}'], f'{d[f"rate{pi}"]}%']
    vals += [chg_sign(qchg), chg_sign(rchg, as_pct=True)]
    data_row(ws3, ri, vals, bold_cols={1}, highlight=(ri%2==1))
    ws3.row_dimensions[ri].height = 18

tr3 = len(analysts_list)+3
auto_width(ws3, [12,9]+[9,8,9]*4+[9,9], NC3)
freeze_header(ws3)

# ============ Sheet 4: 拓展类型 ============
ws4 = wb.create_sheet('拓展类型')
NC4 = 16
title_row(ws4, 1, 1, NC4, f'拓展类型维度 - 新品周报 4.30-5.6')
ws4.row_dimensions[1].height = 28

headers4 = ['拓展类型', '累计SKU']
for pi in range(4):
    headers4 += [f'{PERIODS[pi]}\n上架SKU', f'{PERIODS[pi]}\n销量', f'{PERIODS[pi]}\n出单率']
headers4 += ['销量环比', '出单率环比']
for ci, h in enumerate(headers4, 1):
    bg = C_HDR
    for pi in range(4):
        if PERIODS[pi] in str(h): bg = PCOLORS[pi]; break
    if '环比' in str(h): bg = C_HB
    hdr_cell(ws4, 2, ci, h, bg=bg)
ws4.row_dimensions[2].height = 36

ri = 3
for exp in exp_types:
    d = exp_data.get(exp)
    if not d or d['total']==0: continue
    qchg = d['q3']-d['q2']; rchg = d['rate3']-d['rate2']
    vals = [exp, d['total']]
    for pi in range(4): vals += [d[f'v{pi}'], d[f'q{pi}'], f'{d[f"rate{pi}"]}%']
    vals += [chg_sign(qchg), chg_sign(rchg, as_pct=True)]
    data_row(ws4, ri, vals, bold_cols={1})
    ws4.row_dimensions[ri].height = 18
    ri += 1

auto_width(ws4, [12,9]+[9,8,9]*4+[9,9], NC4)
freeze_header(ws4)

# ============ Sheet 5: 分析及时率 ============
ws5 = wb.create_sheet('分析及时率')
NC5 = 8
title_row(ws5, 1, 1, NC5, f'分析及时率 - 新品周报 4.30-5.6')
ws5.row_dimensions[1].height = 28

section_row(ws5, 3, 1, NC5, '汇总')
for ci, h in enumerate(['', '指标']+PERIODS+['环比'], 1): hdr_cell(ws5, 4, ci, h)

t5 = [('及时分析', timely_in), ('8日内新品无分析', timely_n8), ('超7日低占比未分析', timely_o7), ('合计', timely_total), ('及时率', [f'{v}%' for v in timely_rate])]
for ri, (label, vals) in enumerate(t5, 5):
    rv = ['', label]+list(vals)
    if isinstance(vals[0],(int,float)): rv.append(chg_sign(vals[3]-vals[2]))
    else: rv.append('-')
    data_row(ws5, ri, rv, bold_cols={2}, highlight=(label=='及时率'))
    ws5.row_dimensions[ri].height = 18

section_row(ws5, 11, 1, NC5, '按分析人明细（4.30-5.6）')
ahdrs = ['', '', '分析人', '负责SKU', '及时分析', '8日内新品无分析', '超7日未分析', '及时率']
for ci, h in enumerate(ahdrs, 1): hdr_cell(ws5, 12, ci, h, bg=C_P4)

for ri, ana in enumerate(analysts_list, 13):
    d = timely_by_ana.get(ana)
    if not d: continue
    data_row(ws5, ri, ['', '', ana, d['total'], d['ti'], d['tn'], d['to'], f'{d["rate"]}%'], bold_cols={3})
    ws5.row_dimensions[ri].height = 18

auto_width(ws5, [4,24,12,12,12,16,16,10], NC5)
freeze_header(ws5)

# ============ Sheet 6: 低占比新品 (25列) ============
ws6 = wb.create_sheet('低占比新品')
NC6 = 25
title_row(ws6, 1, 1, NC6, f'低占比新品明细（市占比<75% 且 有对手）- 4.30-5.6 共{len(low_skus)}个')
ws6.row_dimensions[1].height = 28

low_hdrs = ['销售编号','SKU','上架日期','分析人','品类','拓展类型',
            '本周销量','销量环比','本周销售额','销售额环比',
            '上期对手','本期对手','对手环比',
            '上期市占','本期市占','市占环比',
            '8日出单','7日频次','7日新品\n频次',
            '上期市场','操作判断','本期市场','开启PLP','PLG费率','']
for ci, h in enumerate(low_hdrs, 1): hdr_cell(ws6, 2, ci, h, bg=C_RED)
ws6.row_dimensions[2].height = 36

for ri, s in enumerate(low_skus, 3):
    in_prev = s.list_date and s.list_date <= CUTOFFS[2]
    qc=s.qtys[3]; qp=s.qtys[2] if in_prev else 0
    ac=s.amts[3]; ap=s.amts[2] if in_prev else 0
    rc=s.rivals[3]; rp=s.rivals[2] if in_prev else 0
    mc=s.mzbs[3]; mp=s.mzbs[2] if in_prev else None
    vals = [
        s.sid, s.sku, str(s.list_date)[:10] if s.list_date else '', s.analyst, s.category, s.exp_type,
        round(qc), chg_sign(qc-qp) if in_prev else '-',
        f'${round(ac,2):,.2f}', chg_sign(ac-ap) if in_prev else '-',
        round(rp) if in_prev else '-', round(rc), chg_sign(rc-rp) if in_prev else '-',
        f'{mp}%' if mp is not None else '-', f'{mc}%' if mc is not None else '-',
        chg_sign((mc or 0)-(mp or 0), as_pct=True) if mp is not None and mc is not None else '-',
        s.ord8s[3] or '-', s.freqs[3] or '-', s.new_freqs[3] or '-',
        s.mkts[2] if in_prev and s.mkts[2] else '-', s.ops[3] or '-', s.mkts[3] or '-',
        s.plp[3] if s.plp[3] else '-', f'{round(s.plg[2]*100,1)}%' if s.plg[2]>0 else '-', ''
    ]
    # 市占比颜色
    clrs = {}
    if mc is not None:
        if mc < 25: clrs[15] = C_RED
        elif mc < 50: clrs[15] = C_ORG
    data_row(ws6, ri, vals, bold_cols={2,15}, colors=clrs, aligns={1:ALIGN_CENTER,2:ALIGN_LEFT,3:ALIGN_CENTER,4:ALIGN_LEFT,5:ALIGN_LEFT,6:ALIGN_LEFT})
    ws6.row_dimensions[ri].height = 18

auto_width(ws6, [9,20,11,9,12,10]+[8,8,10,8]*2+[9,9,8]+[10,10,10]+[9,16,16]+[9,10,10,9,9,6], NC6)
freeze_header(ws6)

# ============ Sheet 7: 新品PLP ============
ws7 = wb.create_sheet('新品PLP')
NC7 = 11
title_row(ws7, 1, 1, NC7, f'新品PLP复盘 - 4.30-5.6')
ws7.row_dimensions[1].height = 28

# 总数据区块
section_row(ws7, 3, 1, NC7, '【总数据】')
plp_hdrs = ['指标','本周期(4.30-5.6)','上周期(4.23-4.29)','环比','','','','','','','']
for ci, h in enumerate(plp_hdrs, 1): hdr_cell(ws7, 4, ci, h)

plp_kpis = [
    ('广告SKU数', plp_total['skus'], plp_prev.get('skus','-')),
    ('曝光量', f'{plp_total["imp"]:,}', f'{plp_prev.get("imp",0):,}'),
    ('点击量', f'{plp_total["clk"]:,}', f'{plp_prev.get("clk",0):,}'),
    ('订单数', plp_total['ord'], plp_prev.get('ord','-')),
    ('花费 (USD)', f'${plp_total["spend"]:,.2f}', f'${plp_prev.get("spend",0):,.2f}'),
    ('销售额 (USD)', f'${plp_total["sales"]:,.2f}', f'${plp_prev.get("sales",0):,.2f}'),
    ('ROAS', plp_total['roas'], plp_prev.get('roas','-')),
    ('CVR', f'{plp_total["cvr"]}%', f'{plp_prev.get("cvr",0)}%'),
    ('CTR', f'{plp_total["ctr"]:.4f}%', f'{plp_prev.get("ctr",0):.4f}%'),
    ('CPC', f'${plp_total.get("cpc",0):,.2f}', f'${plp_prev.get("cpc",0):,.2f}'),
    ('CPA', f'${plp_total.get("cpa",0):,.2f}', f'${plp_prev.get("cpa",0):,.2f}'),
    ('ACOS', f'{plp_total["acos"]}%', f'{plp_prev.get("acos",0)}%'),
]
for ri, (label, cur, prv) in enumerate(plp_kpis, 5):
    try:
        cv=float(str(cur).replace('$','').replace(',','').replace('%',''))
        pv=float(str(prv).replace('$','').replace(',','').replace('%',''))
        ch = cv-pv
        cht = chg_sign(ch, as_pct=('%' in str(cur)))
    except: cht = '-'
    data_row(ws7, ri, [label, cur, prv, cht], bold_cols={1})
    ws7.row_dimensions[ri].height = 18

# 多维度区块
plp_metrics_hdr = ['维度','SKU','曝光量','点击量','订单','花费','销售额','ROAS','CVR','CTR','ACOS']

def write_plp_dim(ws, start_r, title, data_dict, sort_keys):
    section_row(ws, start_r, 1, NC7, title)
    for ci, h in enumerate(plp_metrics_hdr, 1): hdr_cell(ws, start_r+1, ci, h, bg=C_PLP)
    r = start_r+2
    for k in sort_keys:
        d = data_dict.get(k)
        if not d: continue
        vals = [k, d['skus'], f'{d["imp"]:,}', f'{d["clk"]:,}', d['ord'],
                f'${d["spend"]:,.2f}', f'${d["sales"]:,.2f}', d['roas'],
                f'{d["cvr"]}%', f'{d["ctr"]:.4f}%', f'{d["acos"]}%']
        data_row(ws, r, vals, bold_cols={1})
        ws.row_dimensions[r].height = 18
        r += 1
    return r

ri = 18
ri = write_plp_dim(ws7, ri, '【按分析人】', plp_ana, analysts_list) + 1
ri = write_plp_dim(ws7, ri, '【按品线】', plp_cat, cats) + 1
ri = write_plp_dim(ws7, ri, '【按拓展类型】', plp_exp, exp_types)

auto_width(ws7, [16,8,12,10,8,12,12,8,8,10,8], NC7)
freeze_header(ws7)

# ============ Sheet 8: 新品出单情况 ============
ws8 = wb.create_sheet('新品出单情况')
NC8 = 10
title_row(ws8, 1, 1, NC8, f'新品出单情况 - 4.30-5.6')
ws8.row_dimensions[1].height = 28

headers8 = ['出单情况']
for pi in range(4):
    headers8 += [f'{PERIODS[pi]}\n数量', f'{PERIODS[pi]}\n占比']
headers8 += ['环比']
for ci, h in enumerate(headers8, 1):
    bg = C_HDR
    for pi in range(4):
        if PERIODS[pi] in str(h): bg = PCOLORS[pi]; break
    if '环比' in str(h): bg = C_HB
    hdr_cell(ws8, 2, ci, h, bg=bg)
ws8.row_dimensions[2].height = 36

ord_data = [
    ('8日出单 (Y)', [y_all[pi] for pi in range(4)], [round(y_all[pi]/valid_cnt[pi]*100,1) if valid_cnt[pi] else 0 for pi in range(4)]),
    ('8日外出单 (N)', [n_all[pi] for pi in range(4)], [round(n_all[pi]/valid_cnt[pi]*100,1) if valid_cnt[pi] else 0 for pi in range(4)]),
    ('未出单', [un_all[pi] for pi in range(4)], [round(un_all[pi]/valid_cnt[pi]*100,1) if valid_cnt[pi] else 0 for pi in range(4)]),
    ('有效SKU合计', valid_cnt, [100.0]*4),
]
for ri, (label, cnts, pcts) in enumerate(ord_data, 3):
    vals = [label]
    for pi in range(4): vals += [cnts[pi], f'{pcts[pi]}%']
    vals.append(chg_sign(cnts[3]-cnts[2]))
    data_row(ws8, ri, vals, bold_cols={1}, highlight=(ri==6))
    ws8.row_dimensions[ri].height = 18

auto_width(ws8, [18]+[10,10]*4+[10], NC8)
freeze_header(ws8)

# ============ Sheet 9: 新品未出单原因 (A/B双板块) ============
ws9 = wb.create_sheet('新品未出单原因')
NC9 = 2 + max(len(mkt_has), len(mkt_no))

title_row(ws9, 1, 1, NC9, f'新品未出单原因分析 - 4.30-5.6')
ws9.row_dimensions[1].height = 28

# ── A板块 ──
section_row(ws9, 3, 1, NC9, f'A. 有对手未出单新品    本周: {len(has_r_curr)}个    上周: {len(has_r_prev)}个',
            bg=C_ORG)
ws9.row_dimensions[3].height = 24

# A1
section_row(ws9, 4, 1, NC9, 'A1. 原因分布', bg=C_ORG)
a1_headers = ['市场状态','本周SKU','占比','上周SKU','上周占比','变化']
for ci, h in enumerate(a1_headers, 1): hdr_cell(ws9, 5, ci, h, bg=C_ORG)

ri=6
for rsn in mkt_has:
    cur=a1c.get(rsn,0); prv=a1p.get(rsn,0)
    vals=[rsn, cur, f'{round(cur/len(has_r_curr)*100,1) if has_r_curr else 0}%',
          prv, f'{round(prv/len(has_r_prev)*100,1) if has_r_prev else 0}%',
          chg_sign(cur-prv)]
    data_row(ws9, ri, vals, bold_cols={1})
    ws9.row_dimensions[ri].height = 18
    ri+=1

vals=['合计',len(has_r_curr),'100%',len(has_r_prev),'100%',chg_sign(len(has_r_curr)-len(has_r_prev))]
data_row(ws9, ri, vals, bold_cols={1,2,3,4,5,6})
ri+=2

# A2
section_row(ws9, ri, 1, NC9, 'A2. 按分析人', bg=C_ORG); ri+=1
a2h=['分析人']+mkt_has+['未出单总数']
for ci, h in enumerate(a2h, 1): hdr_cell(ws9, ri, ci, h, bg=C_ORG)
ri+=1
for ana in analysts_list:
    d=a2c.get(ana,{'total':0,'dist':{}})
    vals=[ana]+[d['dist'].get(rsn,0) for rsn in mkt_has]+[d['total']]
    data_row(ws9, ri, vals, bold_cols={1})
    ws9.row_dimensions[ri].height = 18
    ri+=1
ri+=1

# A3
section_row(ws9, ri, 1, NC9, 'A3. 按品线', bg=C_ORG); ri+=1
a3h=['品线']+mkt_has+['未出单总数']
for ci, h in enumerate(a3h, 1): hdr_cell(ws9, ri, ci, h, bg=C_ORG)
ri+=1
for cat in cats:
    d=a3c.get(cat,{'total':0,'dist':{}})
    vals=[cat]+[d['dist'].get(rsn,0) for rsn in mkt_has]+[d['total']]
    data_row(ws9, ri, vals, bold_cols={1})
    ws9.row_dimensions[ri].height = 18
    ri+=1
ri+=2

# ── B板块 ──
section_row(ws9, ri, 1, NC9, f'B. 无对手未出单新品    本周: {len(no_r_curr)}个    上周: {len(no_r_prev)}个',
            bg=C_GRN2); ri+=1

# B1
section_row(ws9, ri, 1, NC9, 'B1. 原因分布', bg=C_GRN2); ri+=1
b1h=['市场状态','本周SKU','占比','上周SKU','上周占比','变化']
for ci, h in enumerate(b1h, 1): hdr_cell(ws9, ri, ci, h, bg=C_GRN2); ri+=1

for rsn in mkt_no:
    cur=b1c.get(rsn,0); prv=b1p.get(rsn,0)
    vals=[rsn, cur, f'{round(cur/len(no_r_curr)*100,1) if no_r_curr else 0}%',
          prv, f'{round(prv/len(no_r_prev)*100,1) if no_r_prev else 0}%',
          chg_sign(cur-prv)]
    data_row(ws9, ri, vals, bold_cols={1})
    ws9.row_dimensions[ri].height = 18
    ri+=1

vals=['合计',len(no_r_curr),'100%',len(no_r_prev),'100%',chg_sign(len(no_r_curr)-len(no_r_prev))]
data_row(ws9, ri, vals, bold_cols={1,2,3,4,5,6}); ri+=2

# B2
section_row(ws9, ri, 1, NC9, 'B2. 按分析人', bg=C_GRN2); ri+=1
b2h=['分析人']+mkt_no+['未出单总数']
for ci, h in enumerate(b2h, 1): hdr_cell(ws9, ri, ci, h, bg=C_GRN2)
ri+=1
for ana in analysts_list:
    d=b2c.get(ana,{'total':0,'dist':{}})
    vals=[ana]+[d['dist'].get(rsn,0) for rsn in mkt_no]+[d['total']]
    data_row(ws9, ri, vals, bold_cols={1})
    ws9.row_dimensions[ri].height = 18
    ri+=1
ri+=1

# B3
section_row(ws9, ri, 1, NC9, 'B3. 按品线', bg=C_GRN2); ri+=1
b3h=['品线']+mkt_no+['未出单总数']
for ci, h in enumerate(b3h, 1): hdr_cell(ws9, ri, ci, h, bg=C_GRN2)
ri+=1
for cat in cats:
    d=b3c.get(cat,{'total':0,'dist':{}})
    vals=[cat]+[d['dist'].get(rsn,0) for rsn in mkt_no]+[d['total']]
    data_row(ws9, ri, vals, bold_cols={1})
    ws9.row_dimensions[ri].height = 18
    ri+=1

auto_width(ws9, [18]+[12]*(NC9-1), NC9)
freeze_header(ws9)

# ============ Sheet 10: 新品PLG维度 ============
ws10 = wb.create_sheet('新品PLG维度')
NC10 = 5
title_row(ws10, 1, 1, NC10, f'新品PLG维度分析 - 4.30-5.6')
ws10.row_dimensions[1].height = 28

section_row(ws10, 3, 1, NC10, '按分析人')
hdr_cell(ws10, 4, 1, ''); hdr_cell(ws10, 4, 2, '分析人'); hdr_cell(ws10, 4, 3, '有效SKU数'); hdr_cell(ws10, 4, 4, '平均费率'); hdr_cell(ws10, 4, 5, '最高费率')

ri=5
for ana in analysts_list:
    d=plg_ana_agg.get(ana)
    if not d: continue
    data_row(ws10, ri, ['', ana, d['cnt'], f'{d["avg"]}%', f'{d["max"]}%'], bold_cols={2})
    ws10.row_dimensions[ri].height = 18
    ri+=1

ri+=1
section_row(ws10, ri, 1, NC10, '按品类'); ri+=1
hdr_cell(ws10, ri, 1, ''); hdr_cell(ws10, ri, 2, '品类'); hdr_cell(ws10, ri, 3, '有效SKU数'); hdr_cell(ws10, ri, 4, '平均费率'); hdr_cell(ws10, ri, 5, '最高费率')
ri+=1
for cat in cats:
    d=plg_cat_agg.get(cat)
    if not d: continue
    data_row(ws10, ri, ['', cat, d['cnt'], f'{d["avg"]}%', f'{d["max"]}%'], bold_cols={2})
    ws10.row_dimensions[ri].height = 18
    ri+=1

auto_width(ws10, [3,18,12,12,12], NC10)
freeze_header(ws10)

# ── 保存 ──
OUT = r'c:\Users\Hardy\ai-projects\新品复盘\新品周报数据表_4.30-5.6.xlsx'
wb.save(OUT)
print(f'Saved: {OUT}\nSheets: {wb.sheetnames}')
