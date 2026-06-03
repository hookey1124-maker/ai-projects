"""
新品周报 4.30-5.6 汇总XLSX生成 v3
基于 新品周报全流程 Skill v2.0.0 (2026-05-13)
"""
import openpyxl
import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# ── 颜色常量 ──
C_HDR  = '0F3460'
C_P1   = '6C757D'; C_P2 = '667EEA'; C_P3 = '2980B9'; C_P4 = 'C0392B'
C_HB   = 'E07B24'
C_GRN  = '08845A'; C_RED = 'C0392B'; C_WHT = 'FFFFFF'; C_PLP = '8E44AD'
C_ORG  = 'BF360C'; C_GRN2 = '1B5E20'

PERIODS = ['4.9-4.15', '4.16-4.22', '4.23-4.29', '4.30-5.6']
CUTOFFS = [datetime.date(2026,4,15), datetime.date(2026,4,22), datetime.date(2026,4,29), datetime.date(2026,5,6)]
PCOLORS = [C_P1, C_P2, C_P3, C_P4]

# ── 辅助函数 ──
def n(v, default=0):
    if v is None or v == '' or v == '-': return default
    if isinstance(v, str):
        if any(c in v for c in '#—∑�'): return default
        v = v.replace(',','').replace('$','').replace('%','')
    try: return float(v)
    except: return default

def pct(v):
    if v is None or v == '' or v == '-': return None
    if isinstance(v, str):
        if any(c in v for c in '#�'): return None
        v = v.replace('%','')
    try: fv = float(v); return round(fv*100,2) if fv<=1.0 else round(fv,2)
    except: return None

def get_date(v):
    if v is None: return None
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    return None

def chg_sign(v, as_pct=False):
    if v is None or v == 0: return '-'
    if as_pct:
        return f'+{v:.1f}%' if v > 0 else f'{v:.1f}%'
    return f'+{v:.0f}' if v > 0 else f'{v:.0f}'

def chg_color(v):
    if v is None or v == 0: return '888888'
    return C_GRN if v > 0 else C_RED

# ── XLSX样式 ──
def hdr(ws, r, c, val, bg=C_HDR, fg=C_WHT, bold=True, align='center', size=10):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = Font(bold=bold, color=fg, name='Arial', size=size)
    cl.fill = PatternFill('solid', fgColor=bg)
    cl.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    return cl

def cell(ws, r, c, val, bold=False, color=None, align='center', size=9):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = Font(bold=bold, color=color or '000000', name='Arial', size=size)
    cl.alignment = Alignment(horizontal=align, vertical='center')
    return cl

def bdr(ws, r1, r2, c1, c2):
    thin = Side(style='thin', color='DDDDDD')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for c in row: c.border = bd

# ── 读取源数据 ──
SRC = r'c:\Users\Hardy\ai-projects\新品复盘\新品检查周源数据和PLP数据.xlsx'
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_main = wb_src['四三数据累计']
ws_plp  = wb_src['PLP明细']

# ── 列索引 (0-based) ──
Q_COLS   = [12,13,14,15]; AMT_COLS = [22,23,24,25]
RIV_COLS = [32,33,34,35]; MZB_COLS = [41,42,43,44]
ORD_COLS = [59,60,61,62]
FQ_COLS  = [68,69,70,71]; NFQ_COLS = [77,78,79,80]
MKT_COLS = [87,88,89,90]; OP_COLS  = [96,97,98,99]
PLP_COLS = [100,101,102,103]; PLG_COLS = [104,105,106]

# ── SKU数据结构 ──
class SKU: pass

skus = []
for row in ws_main.iter_rows(min_row=2, values_only=True):
    if not row[1]: continue
    s = SKU()
    s.sid = row[0]; s.sku = str(row[1]).strip()
    s.list_date = get_date(row[2]); s.first_order = get_date(row[3])
    s.analyst = str(row[4] or '').strip(); s.category = str(row[5] or '').strip()
    s.exp_type = str(row[6] or '').strip()
    s.qtys = [n(row[c]) for c in Q_COLS]; s.amts = [n(row[c]) for c in AMT_COLS]
    s.rivals = [n(row[c]) for c in RIV_COLS]; s.mzbs = [pct(row[c]) for c in MZB_COLS]
    s.ord8s = [str(row[c] or '').strip() for c in ORD_COLS]
    s.freqs = [str(row[c] or '').strip() for c in FQ_COLS]
    s.new_freqs = [str(row[c] or '').strip() for c in NFQ_COLS]
    s.mkts = [str(row[c] or '').strip() for c in MKT_COLS]
    s.ops = [str(row[c] or '').strip() for c in OP_COLS]
    s.plp = [str(row[c] or '').strip() for c in PLP_COLS]
    s.plg = [n(row[c]) for c in PLG_COLS]
    skus.append(s)

print(f'读取SKU: {len(skus)}')

# ── 过滤函数 ──
def valid_for(s, pi):
    if not s.list_date or s.list_date > CUTOFFS[pi]: return False
    if s.ord8s[pi] == '未上架': return False
    return True

def has_rival(s, pi): return s.rivals[pi] > 0
def is_y(s, pi): return s.ord8s[pi] == 'Y'
def is_n(s, pi): return s.ord8s[pi] == 'N'
def is_unordered(s, pi): return s.ord8s[pi] == '未出单'

# ==================== 数据计算 ====================

# 各周期有效SKU
valid_cnt = [sum(1 for s in skus if valid_for(s, pi)) for pi in range(4)]

# 总销量/销售额
total_q = [round(sum(s.qtys[pi] for s in skus if valid_for(s, pi))) for pi in range(4)]
total_a = [round(sum(s.amts[pi] for s in skus if valid_for(s, pi)), 2) for pi in range(4)]

# 有对手SKU
riv_cnt = [sum(1 for s in skus if valid_for(s, pi) and has_rival(s, pi)) for pi in range(4)]
noriv_cnt = [sum(1 for s in skus if valid_for(s, pi) and not has_rival(s, pi)) for pi in range(4)]

# 有对手口径出单
yr_cnt = [sum(1 for s in skus if valid_for(s, pi) and has_rival(s, pi) and is_y(s, pi)) for pi in range(4)]
nr_cnt = [sum(1 for s in skus if valid_for(s, pi) and has_rival(s, pi) and is_n(s, pi)) for pi in range(4)]
unr_cnt = [sum(1 for s in skus if valid_for(s, pi) and has_rival(s, pi) and is_unordered(s, pi)) for pi in range(4)]
riv_rate = [round((yr_cnt[i]+nr_cnt[i])/riv_cnt[i]*100,1) if riv_cnt[i]>0 else 0 for i in range(4)]

# 全量出单
y_all = [sum(1 for s in skus if valid_for(s, pi) and is_y(s, pi)) for pi in range(4)]
n_all = [sum(1 for s in skus if valid_for(s, pi) and is_n(s, pi)) for pi in range(4)]
un_all = [sum(1 for s in skus if valid_for(s, pi) and is_unordered(s, pi)) for pi in range(4)]
rate_all = []
for pi in range(4):
    tot = y_all[pi]+n_all[pi]+un_all[pi]
    rate_all.append(round((y_all[pi]+n_all[pi])/tot*100,1) if tot>0 else 0)

# 本周新上架SKU（4.30-5.6期间上架）
new_list_curr = sum(1 for s in skus if s.list_date and datetime.date(2026,4,30) <= s.list_date <= datetime.date(2026,5,6))

# 平均市占比
def avg_mzb(pi):
    vals = [s.mzbs[pi] for s in skus if valid_for(s, pi) and s.mzbs[pi] is not None]
    return round(sum(vals)/len(vals),1) if vals else 0
mzb_list = [avg_mzb(pi) for pi in range(4)]

# 低占比新品
low_skus = [s for s in skus if valid_for(s,3) and s.mzbs[3] is not None and s.mzbs[3]<75 and has_rival(s,3)]
low_skus.sort(key=lambda s: s.mzbs[3] if s.mzbs[3] else 100)

# 分析及时率
timely_in = []; timely_n8 = []; timely_o7 = []
for pi in range(4):
    vs = [s for s in skus if valid_for(s, pi)]
    ti=tn=to=0
    for s in vs:
        if s.new_freqs[pi]=='异常': tn+=1
        elif s.freqs[pi]=='异常': to+=1
        else: ti+=1
    timely_in.append(ti); timely_n8.append(tn); timely_o7.append(to)
    print(f'{PERIODS[pi]}: 及时={ti}, 8日内新品={tn}, 超7日={to}, 合计={ti+tn+to}')

timely_total = [timely_in[i]+timely_n8[i]+timely_o7[i] for i in range(4)]
timely_rate = [round(timely_in[i]/timely_total[i]*100,1) if timely_total[i]>0 else 0 for i in range(4)]

# 按分析人及时率（当期）
timely_by_ana = {}
for ana in sorted(set(s.analyst for s in skus if s.analyst)):
    vs = [s for s in skus if s.analyst==ana and valid_for(s,3)]
    ti=sum(1 for s in vs if s.new_freqs[3]!='异常' and s.freqs[3]!='异常')
    tn=sum(1 for s in vs if s.new_freqs[3]=='异常')
    to=sum(1 for s in vs if s.new_freqs[3]!='异常' and s.freqs[3]=='异常')
    tot=len(vs)
    timely_by_ana[ana]={'total':tot,'ti':ti,'tn':tn,'to':to,'rate':round(ti/tot*100,1) if tot>0 else 0}

# 品类/分析人/拓展类型维度
def build_dim(rows, key_fn, periods=4):
    groups = defaultdict(list)
    for s in rows: groups[key_fn(s)].append(s)
    result = {}
    for k, grp in sorted(groups.items()):
        d = {'total': len(grp)}
        for pi in range(periods):
            vs = [s for s in grp if valid_for(s, pi)]
            d[f'v{pi}'] = len(vs)
            d[f'q{pi}'] = round(sum(s.qtys[pi] for s in vs))
            d[f'a{pi}'] = round(sum(s.amts[pi] for s in vs), 2)
            hr = [s for s in vs if has_rival(s, pi)]
            d[f'r{pi}'] = len(hr)
            yr = sum(1 for s in hr if is_y(s, pi))
            nr = sum(1 for s in hr if is_n(s, pi))
            d[f'yr{pi}'] = yr; d[f'nr{pi}'] = nr
            num = yr+nr; denom = len(hr)
            d[f'rate{pi}'] = round(num/denom*100,1) if denom>0 else 0
        result[k] = d
    return result

cats = sorted(set(s.category for s in skus if s.category))
analysts_list = ['俞东旭','张潇','朱培源','王偲涵','章鹏','胡煜星']
exp_types = ['原开品','拓展品','组合件']

cat_data = build_dim(skus, lambda s: s.category)
ana_data = build_dim(skus, lambda s: s.analyst)
exp_data = build_dim(skus, lambda s: s.exp_type)

# === PLP 4.30-5.6 ===
plp_all = []
for row in ws_plp.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    ps = str(row[0])
    if '4.30' not in ps and '4.30-5.6' not in ps: continue
    plp_all.append(row)

# 分离明细行和汇总行
plp_detail = []
plp_summary = None
for row in plp_all:
    c1 = str(row[1] or '')
    if '总数据' in c1 or '总计' in c1:
        plp_summary = row
    elif row[2] and not str(row[2]).startswith('广告'):
        plp_detail.append(row)

def clean_plp(row):
    try:
        return {'sku': row[2], 'ana': str(row[8] or ''), 'cat': str(row[9] or ''),
                'exp': str(row[10] or ''), 'imp': n(row[11]), 'clk': n(row[12]),
                'ord': n(row[13]), 'spend': n(row[14]), 'sales': n(row[15])}
    except: return None

plp_d = [r for row in plp_detail if (r:=clean_plp(row)) and r['sku']]

def agg_plp(rows, key_fn):
    agg = {}
    for r in rows:
        k = key_fn(r)
        if k not in agg: agg[k] = {'imp':0,'clk':0,'ord':0,'spend':0,'sales':0,'sku_set':set()}
        a = agg[k]; a['imp']+=r['imp']; a['clk']+=r['clk']; a['ord']+=r['ord']
        a['spend']+=r['spend']; a['sales']+=r['sales']; a['sku_set'].add(r['sku'])
    result = {}
    for k,a in agg.items():
        roas=round(a['sales']/a['spend'],2) if a['spend']>0 else 0
        cvr=round(a['ord']/a['clk']*100,2) if a['clk']>0 else 0
        ctr=round(a['clk']/a['imp']*100,4) if a['imp']>0 else 0
        acos=round(a['spend']/a['sales']*100,2) if a['sales']>0 else 0
        cpc=round(a['spend']/a['clk'],2) if a['clk']>0 else 0
        cpa=round(a['spend']/a['ord'],2) if a['ord']>0 else 0
        result[k]={'skus':len(a['sku_set']),'imp':int(a['imp']),'clk':int(a['clk']),
                   'ord':int(a['ord']),'spend':round(a['spend'],2),'sales':round(a['sales'],2),
                   'roas':roas,'cvr':cvr,'ctr':ctr,'acos':acos,'cpc':cpc,'cpa':cpa}
    return result

plp_ana = agg_plp(plp_d, lambda r: r['ana'])
plp_cat = agg_plp(plp_d, lambda r: r['cat'])
plp_exp = agg_plp(plp_d, lambda r: r['exp'])

# PLP总计聚合
total_imp = sum(d['imp'] for d in plp_ana.values())
total_clk = sum(d['clk'] for d in plp_ana.values())
total_ord = sum(d['ord'] for d in plp_ana.values())
total_spend = sum(d['spend'] for d in plp_ana.values())
total_sales = sum(d['sales'] for d in plp_ana.values())
plp_total = {
    'skus': len(plp_d), 'imp': total_imp, 'clk': total_clk, 'ord': total_ord,
    'spend': round(total_spend,2), 'sales': round(total_sales,2),
    'roas': round(total_sales/total_spend,2) if total_spend>0 else 0,
    'cvr': round(total_ord/total_clk*100,2) if total_clk>0 else 0,
    'ctr': round(total_clk/total_imp*100,4) if total_imp>0 else 0,
    'acos': round(total_spend/total_sales*100,2) if total_sales>0 else 0,
    'cpc': round(total_spend/total_clk,2) if total_clk>0 else 0,
    'cpa': round(total_spend/total_ord,2) if total_ord>0 else 0,
}

# === PLG ===
plg_cur = [s for s in skus if valid_for(s,3) and s.plg[2]>0]
plg_ana_d = defaultdict(list); plg_cat_d = defaultdict(list)
for s in plg_cur:
    plg_ana_d[s.analyst or '未知'].append(s.plg[2])
    plg_cat_d[s.category or '未分类'].append(s.plg[2])
plg_ana_agg = {k:{'cnt':len(v),'avg':round(sum(v)/len(v)*100,1),'max':round(max(v)*100,1)} for k,v in plg_ana_d.items()}
plg_cat_agg = {k:{'cnt':len(v),'avg':round(sum(v)/len(v)*100,1),'max':round(max(v)*100,1)} for k,v in plg_cat_d.items()}

# === 新品未出单 A/B拆分 ===
# A: 有对手未出单 (ord8==未出单 AND rival>0)
# B: 无对手未出单 (ord8==未出单 AND rival==0)
mkt_order_has = ['竞争无优势','无市场','站内无价格优势','站外出单','正常','#N/A','未知']
mkt_order_no  = ['无市场','未知','竞争无优势','#N/A','其他']

def classify_mkt(mkt, order_list):
    """将市场状态归类到标准原因列表"""
    if not mkt: return '未知'
    for o in order_list:
        if o in mkt: return o
    if '价格' in mkt: return '站内无价格优势'
    if '站外' in mkt: return '站外出单'
    if '竞争' in mkt: return '竞争无优势'
    if '无市场' in mkt: return '无市场'
    return '其他'

# 当前周期
has_r_curr = [s for s in skus if valid_for(s,3) and is_unordered(s,3) and has_rival(s,3)]
no_r_curr  = [s for s in skus if valid_for(s,3) and is_unordered(s,3) and not has_rival(s,3)]
# 上周期
has_r_prev = [s for s in skus if valid_for(s,2) and is_unordered(s,2) and has_rival(s,2)]
no_r_prev  = [s for s in skus if valid_for(s,2) and is_unordered(s,2) and not has_rival(s,2)]

# A1/B1 原因分布
def reason_dist(sku_list, pi, order_list):
    dist = {r:0 for r in order_list}
    for s in sku_list:
        mkt = s.mkts[pi]
        mapped = classify_mkt(mkt, order_list)
        if mapped in dist: dist[mapped] += 1
        else: dist['其他'] += 1
    return dist

a1_curr = reason_dist(has_r_curr, 3, mkt_order_has)
a1_prev = reason_dist(has_r_prev, 2, mkt_order_has)
b1_curr = reason_dist(no_r_curr, 3, mkt_order_no)
b1_prev = reason_dist(no_r_prev, 2, mkt_order_no)

# A2/B2 按分析人
def ana_reason_dist(sku_list, pi, order_list):
    result = {}
    for ana in analysts_list:
        ana_skus = [s for s in sku_list if s.analyst == ana]
        dist = {r:0 for r in order_list}
        for s in ana_skus:
            mapped = classify_mkt(s.mkts[pi], order_list)
            if mapped in dist: dist[mapped] += 1
            else: dist['其他'] += 1
        result[ana] = {'total': len(ana_skus), 'dist': dist}
    return result

a2_curr = ana_reason_dist(has_r_curr, 3, mkt_order_has)
b2_curr = ana_reason_dist(no_r_curr, 3, mkt_order_no)

# A3/B3 按品线
def cat_reason_dist(sku_list, pi, order_list):
    result = {}
    for cat in cats:
        cat_skus = [s for s in sku_list if s.category == cat]
        dist = {r:0 for r in order_list}
        for s in cat_skus:
            mapped = classify_mkt(s.mkts[pi], order_list)
            if mapped in dist: dist[mapped] += 1
            else: dist['其他'] += 1
        result[cat] = {'total': len(cat_skus), 'dist': dist}
    return result

a3_curr = cat_reason_dist(has_r_curr, 3, mkt_order_has)
b3_curr = cat_reason_dist(no_r_curr, 3, mkt_order_no)

wb_src.close()

# ══════════════════════════════════════════════════════════════
# 生成 XLSX
# ══════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()

# ============ Sheet 1: 总体数据（三分区结构）============
ws1 = wb.active
ws1.title = '总体数据'

# ── 一、总体概况 ──
hdr(ws1, 1, 1, '一、总体概况', bg=C_HDR, align='left', size=12)
for c in range(2,7): hdr(ws1, 1, c, '')

hdr(ws1, 2, 1, '指标'); hdr(ws1, 2, 2, PERIODS[0]); hdr(ws1, 2, 3, PERIODS[1])
hdr(ws1, 2, 4, PERIODS[2]); hdr(ws1, 2, 5, PERIODS[3]); hdr(ws1, 2, 6, '环比')

overview_rows = [
    ('累计SKU数（截止日）', valid_cnt, True),
    ('本周新上架SKU', ['-','-','-', new_list_curr], False),
    ('总销量', total_q, True),
    ('总销售额(USD)', total_a, True),
    ('有对手SKU数（对手销量>0）', riv_cnt, True),
    ('无对手SKU数', noriv_cnt, True),
]
for ri, (label, vals, is_num) in enumerate(overview_rows, 3):
    cell(ws1, ri, 1, label, bold=True, align='left')
    for pi in range(4): cell(ws1, ri, pi+2, vals[pi] if pi<len(vals) else '-')
    if is_num and len(vals)>=4:
        try:
            chg = float(str(vals[3]).replace('%','')) - float(str(vals[2]).replace('%',''))
            cell(ws1, ri, 6, chg_sign(chg), color=chg_color(chg))
        except: cell(ws1, ri, 6, '-')
    else: cell(ws1, ri, 6, '-')

# ── 二、分析及时率 ──
sr2 = len(overview_rows) + 4
hdr(ws1, sr2, 1, '二、分析及时率', bg=C_HDR, align='left', size=12)
for c in range(2,7): hdr(ws1, sr2, c, '')
hdr(ws1, sr2+1, 1, '指标'); hdr(ws1, sr2+1, 2, PERIODS[0]); hdr(ws1, sr2+1, 3, PERIODS[1])
hdr(ws1, sr2+1, 4, PERIODS[2]); hdr(ws1, sr2+1, 5, PERIODS[3]); hdr(ws1, sr2+1, 6, '环比')

timely_rows = [
    ('及时分析产品数', timely_in),
    ('8日内新品无分析', timely_n8),
    ('超7日低占比未分析', timely_o7),
    ('统计总数', timely_total),
    ('及时分析率', [f'{v}%' for v in timely_rate]),
]
for ri, (label, vals) in enumerate(timely_rows, sr2+2):
    cell(ws1, ri, 1, label, bold=True, align='left')
    for pi in range(4): cell(ws1, ri, pi+2, vals[pi])
    if isinstance(vals[0], (int,float)):
        chg = vals[3] - vals[2]
        cell(ws1, ri, 6, chg_sign(chg), color=chg_color(chg))
    else:
        try:
            vc = float(str(vals[3]).replace('%','')); vp = float(str(vals[2]).replace('%',''))
            cell(ws1, ri, 6, chg_sign(vc-vp, as_pct=True), color=chg_color(vc-vp))
        except: cell(ws1, ri, 6, '-')

# ── 三、新品出单情况（有对手口径）──
sr3 = sr2 + 2 + len(timely_rows) + 1
hdr(ws1, sr3, 1, '三、新品出单情况（有对手口径）', bg=C_HDR, align='left', size=12)
for c in range(2,7): hdr(ws1, sr3, c, '')
hdr(ws1, sr3+1, 1, '指标'); hdr(ws1, sr3+1, 2, PERIODS[0]); hdr(ws1, sr3+1, 3, PERIODS[1])
hdr(ws1, sr3+1, 4, PERIODS[2]); hdr(ws1, sr3+1, 5, PERIODS[3]); hdr(ws1, sr3+1, 6, '环比')

order_rows = [
    ('有对手总SKU', riv_cnt),
    ('8日内出单（Y）', yr_cnt),
    ('8日外出单（N）', nr_cnt),
    ('真正未出单', unr_cnt),
    ('已出单合计(Y+N)', [yr_cnt[i]+nr_cnt[i] for i in range(4)]),
    ('出单率', [f'{riv_rate[i]}%' for i in range(4)]),
]
for ri, (label, vals) in enumerate(order_rows, sr3+2):
    cell(ws1, ri, 1, label, bold=True, align='left')
    for pi in range(4): cell(ws1, ri, pi+2, vals[pi])
    if isinstance(vals[0], (int,float)):
        chg = vals[3] - vals[2]
        cell(ws1, ri, 6, chg_sign(chg), color=chg_color(chg))
    else:
        try:
            vc = float(str(vals[3]).replace('%','')); vp = float(str(vals[2]).replace('%',''))
            cell(ws1, ri, 6, chg_sign(vc-vp, as_pct=True), color=chg_color(vc-vp))
        except: cell(ws1, ri, 6, '-')

last_r1 = sr3 + 1 + len(order_rows)
ws1.column_dimensions['A'].width = 32
for c in 'BCDEF': ws1.column_dimensions[c].width = 16
bdr(ws1, 2, last_r1, 1, 6)

# ============ Sheet 2: 品线维度 ============
ws2 = wb.create_sheet('品线维度')
# 表头行1: 分类
hdr(ws2, 1, 1, '品线', bg=C_HDR); hdr(ws2, 1, 2, '累计SKU', bg=C_HDR)
ci = 3
for pi in range(4):
    hdr(ws2, 1, ci, PERIODS[pi], bg=PCOLORS[pi]); ci+=1
    hdr(ws2, 1, ci, f'{PERIODS[pi]}\n上架SKU', bg=PCOLORS[pi]); ci+=1
    hdr(ws2, 1, ci, f'{PERIODS[pi]}\n销量', bg=PCOLORS[pi]); ci+=1
    hdr(ws2, 1, ci, f'{PERIODS[pi]}\n销售额', bg=PCOLORS[pi]); ci+=1
    hdr(ws2, 1, ci, f'{PERIODS[pi]}\n出单率', bg=PCOLORS[pi]); ci+=1
hdr(ws2, 1, ci, '销量环比', bg=C_HB); hdr(ws2, 1, ci+1, '出单率环比', bg=C_HB)
ncols2 = ci+1
ws2.row_dimensions[1].height = 36

for ri, cat in enumerate(cats, 2):
    d = cat_data[cat]
    q_chg = d['q3'] - d['q2']; r_chg = d['rate3'] - d['rate2']
    row = [cat, d['total']]
    for pi in range(4):
        row += [d[f'v{pi}'], d[f'q{pi}'], f'${d[f"a{pi}"]:.2f}', f'{d[f"rate{pi}"]}%']
    row += [chg_sign(q_chg), chg_sign(r_chg, as_pct=True)]
    for ci, v in enumerate(row, 1):
        c = chg_color(q_chg) if ci==len(row)-1 else (chg_color(r_chg) if ci==len(row) else None)
        cell(ws2, ri, ci, v, bold=(ci==1), color=c, align='left' if ci==1 else 'center')

# 合计
tr2 = len(cats)+2
ts = ['合计', sum(cat_data[c]['total'] for c in cats)]
for pi in range(4):
    ts += [sum(cat_data[c][f'v{pi}'] for c in cats),
           round(sum(cat_data[c][f'q{pi}'] for c in cats)),
           f'${round(sum(cat_data[c][f"a{pi}"] for c in cats),2):.2f}', '']
tq = sum(cat_data[c]['q3'] for c in cats) - sum(cat_data[c]['q2'] for c in cats)
ts += [chg_sign(tq), '']
for ci, v in enumerate(ts, 1): cell(ws2, tr2, ci, v, bold=True, align='left' if ci==1 else 'center')

ws2.column_dimensions['A'].width = 16; ws2.column_dimensions['B'].width = 10
for ci in range(3, ncols2+1): ws2.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 12
bdr(ws2, 1, tr2, 1, ncols2)

# ============ Sheet 3: 分析人维度 ============
ws3 = wb.create_sheet('分析人维度')
hdr(ws3, 1, 1, '分析人', bg=C_HDR); hdr(ws3, 1, 2, '累计SKU', bg=C_HDR)
ci = 3
for pi in range(4):
    hdr(ws3, 1, ci, PERIODS[pi], bg=PCOLORS[pi]); ci+=1
    hdr(ws3, 1, ci, f'{PERIODS[pi]}\n上架SKU', bg=PCOLORS[pi]); ci+=1
    hdr(ws3, 1, ci, f'{PERIODS[pi]}\n销量', bg=PCOLORS[pi]); ci+=1
    hdr(ws3, 1, ci, f'{PERIODS[pi]}\n出单率', bg=PCOLORS[pi]); ci+=1
hdr(ws3, 1, ci, '销量环比', bg=C_HB); hdr(ws3, 1, ci+1, '出单率环比', bg=C_HB)
ncols3 = ci+1
ws3.row_dimensions[1].height = 36

for ri, ana in enumerate(analysts_list, 2):
    d = ana_data.get(ana)
    if not d: continue
    q_chg = d['q3'] - d['q2']; r_chg = d['rate3'] - d['rate2']
    row = [ana, d['total']]
    for pi in range(4):
        row += [d[f'v{pi}'], d[f'q{pi}'], f'{d[f"rate{pi}"]}%']
    row += [chg_sign(q_chg), chg_sign(r_chg, as_pct=True)]
    for ci, v in enumerate(row, 1):
        c = chg_color(q_chg) if ci==len(row)-1 else (chg_color(r_chg) if ci==len(row) else None)
        cell(ws3, ri, ci, v, bold=(ci==1), color=c, align='left' if ci==1 else 'center')

tr3 = len(analysts_list)+2
ws3.column_dimensions['A'].width = 14; ws3.column_dimensions['B'].width = 10
for ci in range(3, ncols3+1): ws3.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 12
bdr(ws3, 1, tr3, 1, ncols3)

# ============ Sheet 4: 拓展类型 ============
ws4 = wb.create_sheet('拓展类型')
hdr(ws4, 1, 1, '拓展类型', bg=C_HDR); hdr(ws4, 1, 2, '累计SKU', bg=C_HDR)
ci = 3
for pi in range(4):
    hdr(ws4, 1, ci, PERIODS[pi], bg=PCOLORS[pi]); ci+=1
    hdr(ws4, 1, ci, f'{PERIODS[pi]}\n上架SKU', bg=PCOLORS[pi]); ci+=1
    hdr(ws4, 1, ci, f'{PERIODS[pi]}\n销量', bg=PCOLORS[pi]); ci+=1
    hdr(ws4, 1, ci, f'{PERIODS[pi]}\n出单率', bg=PCOLORS[pi]); ci+=1
hdr(ws4, 1, ci, '销量环比', bg=C_HB); hdr(ws4, 1, ci+1, '出单率环比', bg=C_HB)
ncols4 = ci+1
ws4.row_dimensions[1].height = 36

ri = 2
for exp in exp_types:
    d = exp_data.get(exp)
    if not d or d['total']==0: continue
    q_chg = d['q3'] - d['q2']; r_chg = d['rate3'] - d['rate2']
    row = [exp, d['total']]
    for pi in range(4): row += [d[f'v{pi}'], d[f'q{pi}'], f'{d[f"rate{pi}"]}%']
    row += [chg_sign(q_chg), chg_sign(r_chg, as_pct=True)]
    for ci, v in enumerate(row, 1):
        c = chg_color(q_chg) if ci==len(row)-1 else (chg_color(r_chg) if ci==len(row) else None)
        cell(ws4, ri, ci, v, bold=(ci==1), color=c, align='left' if ci==1 else 'center')
    ri += 1

ws4.column_dimensions['A'].width = 14; ws4.column_dimensions['B'].width = 10
for ci in range(3, ncols4+1): ws4.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 12
bdr(ws4, 1, ri-1, 1, ncols4)

# ============ Sheet 5: 分析及时率 ============
ws5 = wb.create_sheet('分析及时率')

# 汇总
hdr(ws5, 1, 1, '分析及时率汇总', bg=C_HDR, align='left', size=12)
for c in range(2,8): hdr(ws5, 1, c, '')
hdr(ws5, 2, 1, ''); hdr(ws5, 2, 2, '指标')
for pi in range(4): hdr(ws5, 2, pi+3, PERIODS[pi], bg=PCOLORS[pi])
hdr(ws5, 2, 7, '环比', bg=C_HB)

t_rows5 = [
    ('及时分析', timely_in),
    ('8日内新品无分析', timely_n8),
    ('超7日低占比未分析', timely_o7),
    ('合计', timely_total),
    ('及时率', [f'{v}%' for v in timely_rate]),
]
for ri, (label, vals) in enumerate(t_rows5, 3):
    hdr(ws5, ri, 1, '', bg=C_HDR)
    cell(ws5, ri, 2, label, bold=True, align='left')
    for pi in range(4): cell(ws5, ri, pi+3, vals[pi])
    if isinstance(vals[0], (int,float)):
        chg = vals[3]-vals[2]; cell(ws5, ri, 7, chg_sign(chg), color=chg_color(chg))
    else:
        try:
            vc=float(str(vals[3]).replace('%','')); vp=float(str(vals[2]).replace('%',''))
            cell(ws5, ri, 7, chg_sign(vc-vp, as_pct=True), color=chg_color(vc-vp))
        except: cell(ws5, ri, 7, '-')

# 按分析人明细
sr5 = 10
hdr(ws5, sr5, 1, '按分析人明细（4.30-5.6）', bg=C_P4, align='left', size=12)
for c in range(2,9): hdr(ws5, sr5, c, '')
hdr(ws5, sr5+1, 1, ''); hdr(ws5, sr5+1, 2, ''); hdr(ws5, sr5+1, 3, '分析人')
hdr(ws5, sr5+1, 4, '负责SKU'); hdr(ws5, sr5+1, 5, '及时分析')
hdr(ws5, sr5+1, 6, '8日内新品无分析'); hdr(ws5, sr5+1, 7, '超7日未分析'); hdr(ws5, sr5+1, 8, '及时率')

for ri, ana in enumerate(analysts_list, sr5+2):
    d = timely_by_ana.get(ana)
    if not d: continue
    hdr(ws5, ri, 1, '', bg=C_HDR); hdr(ws5, ri, 2, '', bg=C_HDR)
    cell(ws5, ri, 3, ana, bold=True, align='left')
    cell(ws5, ri, 4, d['total']); cell(ws5, ri, 5, d['ti'])
    cell(ws5, ri, 6, d['tn']); cell(ws5, ri, 7, d['to'])
    cell(ws5, ri, 8, f'{d["rate"]}%')

ws5.column_dimensions['A'].width = 3; ws5.column_dimensions['B'].width = 24
for c in 'CDEFGH': ws5.column_dimensions[c].width = 16
bdr(ws5, 2, sr5+1+len(analysts_list), 1, 8)

# ============ Sheet 6: 低占比新品 (25列) ============
ws6 = wb.create_sheet('低占比新品')
low_hdrs = [
    '销售编号','SKU','上架日期','分析人','品类','拓展类型',       # 1-6
    '本周销量','销量环比',                                          # 7-8
    '本周销售额','销售额环比',                                       # 9-10
    '上期末对手销量','本期末对手销量','对手销量环比',                 # 11-13
    '上期末市占比','本期末市占比','市占比环比',                       # 14-16
    '本期末8日出单',                                                # 17
    '本期末7日频次标签',                                             # 18
    '本期末7日新品频次标签',                                         # 19
    '上期末市场状态',                                                # 20
    '本周操作判断',                                                  # 21
    '本期末市场状态',                                                # 22
    '本周开启PLP',                                                  # 23
    '本周PLG最高费率',                                              # 24
    '（备用）',                                                     # 25
]
for ci, h in enumerate(low_hdrs, 1): hdr(ws6, 1, ci, h, bg=C_RED)
ws6.row_dimensions[1].height = 30

for ri, s in enumerate(low_skus, 2):
    in_prev = s.list_date and s.list_date <= CUTOFFS[2]  # SKU是否在上周期已上架
    q_cur = s.qtys[3]; q_prv = s.qtys[2] if in_prev else 0
    a_cur = s.amts[3]; a_prv = s.amts[2] if in_prev else 0
    r_cur = s.rivals[3]; r_prv = s.rivals[2] if in_prev else 0
    m_cur = s.mzbs[3]; m_prv = s.mzbs[2] if in_prev else None

    cols = [
        s.sid, s.sku, str(s.list_date) if s.list_date else '', s.analyst, s.category, s.exp_type,
        round(q_cur), chg_sign(q_cur - q_prv) if in_prev else '-',
        f'${round(a_cur,2):.2f}', chg_sign(a_cur - a_prv) if in_prev else '-',
        round(r_prv) if in_prev else '-', round(r_cur),
        chg_sign(r_cur - r_prv) if in_prev else '-',
        f'{m_prv}%' if m_prv is not None else '-',
        f'{m_cur}%' if m_cur is not None else '-',
        chg_sign((m_cur or 0)-(m_prv or 0), as_pct=True) if m_prv is not None and m_cur is not None else '-',
        s.ord8s[3] or '-', s.freqs[3] or '-', s.new_freqs[3] or '-',
        s.mkts[2] if in_prev and s.mkts[2] else '-', s.ops[3] or '-', s.mkts[3] or '-',
        s.plp[3] if s.plp[3] else '-',
        f'{round(s.plg[2]*100,1)}%' if s.plg[2]>0 else '-',
        '',  # spare
    ]
    for ci, v in enumerate(cols, 1):
        c_color = None
        if ci == 15:
            try:
                mv = float(str(v).replace('%',''))
                c_color = C_RED if mv<25 else (C_HB if mv<50 else None)
            except: pass
        cell(ws6, ri, ci, v, color=c_color, align='left' if ci in [2,4,5,6] else 'center')

bdr(ws6, 1, len(low_skus)+1, 1, len(low_hdrs))
cw6 = [10,22,12,10,14,12, 10,10, 12,10, 14,14,12, 12,12,12, 12, 18,18, 14,12,14, 12,12, 8]
for ci, w in enumerate(cw6, 1): ws6.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

# ============ Sheet 7: 新品PLP（多区块）============
ws7 = wb.create_sheet('新品PLP')
plp_metrics = ['广告SKU','曝光量','点击量','订单数','花费(USD)','销售额(USD)',
               'ROAS','CVR%','CTR%','CPC','CPA','ACOS%']

def write_plp_block(ws, start_row, title, data_dict, sort_keys=None):
    """写入一个PLP区块"""
    hdr(ws, start_row, 1, title, bg=C_PLP, align='left', size=11)
    for c in range(2, 14): hdr(ws, start_row, c, '')
    hdr(ws, start_row+1, 1, '维度')
    for ci, m in enumerate(plp_metrics, 2): hdr(ws, start_row+1, ci, m, bg=C_PLP)
    r = start_row+2
    keys = sort_keys or sorted(data_dict.keys())
    for k in keys:
        d = data_dict.get(k)
        if not d: continue
        cell(ws, r, 1, k, bold=True, align='left')
        vals = [d['skus'], d['imp'], d['clk'], d['ord'],
                round(d['spend'],2), round(d['sales'],2),
                d['roas'], f"{d['cvr']}%", f"{d['ctr']:.4f}%",
                f"${d.get('cpc',0):.2f}", f"${d.get('cpa',0):.2f}", f"{d['acos']}%"]
        for ci, v in enumerate(vals, 2): cell(ws, r, ci, v)
        r += 1
    return r

# 总数据
hdr(ws7, 1, 1, '【总数据】4.30-5.6', bg=C_PLP, align='left', size=12)
for c in range(2, 14): hdr(ws7, 1, c, '')
hdr(ws7, 2, 1, '指标'); hdr(ws7, 2, 2, '本周期'); hdr(ws7, 2, 3, '上周期'); hdr(ws7, 2, 4, '环比')
for ci in range(5, 14): hdr(ws7, 2, ci, '')

# 上周期PLP数据（4.23-4.29）
plp_prev_all = []
for row in ws_plp.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    if '4.23' in str(row[0]) or '4.23-4.29' in str(row[0]):
        plp_prev_all.append(row)
# (re-open for PLP prev)
# Actually wb_src is closed. Let me skip prev PLP detailed and just show current.
# For completeness, let's re-read PLP prev quickly.

# Re-open for prev PLP
wb_src2 = openpyxl.load_workbook(SRC, data_only=True)
ws_plp2 = wb_src2['PLP明细']
plp_prev_rows = []
for row in ws_plp2.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    if '4.23' in str(row[0]) or '4.23-4.29' in str(row[0]):
        plp_prev_rows.append(row)

# Find prev total
plp_prev_total_row = next((r for r in plp_prev_rows if str(r[1] or '').strip() in ('总数据','总计')), None)
# Prev detail
plp_prev_detail = [clean_plp(r) for r in plp_prev_rows if r[1] and str(r[1] or '').strip() not in ('总数据','总计') and r[2] and not str(r[2]).startswith('广告')]
plp_prev_detail = [r for r in plp_prev_detail if r and r['sku']]

if plp_prev_detail:
    imp_p = sum(r['imp'] for r in plp_prev_detail); clk_p = sum(r['clk'] for r in plp_prev_detail)
    ord_p = sum(r['ord'] for r in plp_prev_detail); spd_p = sum(r['spend'] for r in plp_prev_detail)
    sal_p = sum(r['sales'] for r in plp_prev_detail)
    plp_prev_total = {
        'skus': len(plp_prev_detail), 'imp': imp_p, 'clk': clk_p, 'ord': ord_p,
        'spend': round(spd_p,2), 'sales': round(sal_p,2),
        'roas': round(sal_p/spd_p,2) if spd_p>0 else 0,
        'cvr': round(ord_p/clk_p*100,2) if clk_p>0 else 0,
        'ctr': round(clk_p/imp_p*100,4) if imp_p>0 else 0,
        'acos': round(spd_p/sal_p*100,2) if sal_p>0 else 0,
        'cpc': round(spd_p/clk_p,2) if clk_p>0 else 0,
        'cpa': round(spd_p/ord_p,2) if ord_p>0 else 0,
    }
else:
    plp_prev_total = dict(plp_total)  # fallback

wb_src2.close()

# Write total block rows
total_kpis = [
    ('广告SKU数', plp_total['skus'], plp_prev_total.get('skus','-')),
    ('曝光量', plp_total['imp'], plp_prev_total.get('imp','-')),
    ('点击量', plp_total['clk'], plp_prev_total.get('clk','-')),
    ('订单数', plp_total['ord'], plp_prev_total.get('ord','-')),
    ('花费(USD)', f"${plp_total['spend']:.2f}", f"${plp_prev_total.get('spend',0):.2f}"),
    ('销售额(USD)', f"${plp_total['sales']:.2f}", f"${plp_prev_total.get('sales',0):.2f}"),
    ('ROAS', plp_total['roas'], plp_prev_total.get('roas','-')),
    ('CVR', f"{plp_total['cvr']}%", f"{plp_prev_total.get('cvr','-')}%"),
    ('CTR', f"{plp_total['ctr']:.4f}%", f"{plp_prev_total.get('ctr',0):.4f}%"),
    ('CPC', f"${plp_total.get('cpc',0):.2f}", f"${plp_prev_total.get('cpc',0):.2f}"),
    ('CPA', f"${plp_total.get('cpa',0):.2f}", f"${plp_prev_total.get('cpa',0):.2f}"),
    ('ACOS', f"{plp_total['acos']}%", f"{plp_prev_total.get('acos','-')}%"),
]
for ri, (label, curr, prev) in enumerate(total_kpis, 3):
    cell(ws7, ri, 1, label, bold=True, align='left')
    cell(ws7, ri, 2, curr, bold=True)
    cell(ws7, ri, 3, prev)
    # 环比
    try:
        cv = float(str(curr).replace('$','').replace('%',''))
        pv = float(str(prev).replace('$','').replace('%',''))
        chg = cv - pv
        cell(ws7, ri, 4, chg_sign(chg), color=chg_color(chg))
    except: cell(ws7, ri, 4, '-')

# 按分析人/品线/拓展类型
ri = 16
ri = write_plp_block(ws7, ri, '【分析人维度】', plp_ana, analysts_list) + 1
ri = write_plp_block(ws7, ri, '【品线维度】', plp_cat, cats) + 1
ri = write_plp_block(ws7, ri, '【拓展类型维度】', plp_exp, exp_types)

ws7.column_dimensions['A'].width = 18
for ci in range(2, 14): ws7.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 13
bdr(ws7, 1, ri-1, 1, 13)

# ============ Sheet 8: 新品出单情况 ============
ws8 = wb.create_sheet('新品出单情况')
hdr(ws8, 1, 1, '出单情况', bg=C_HDR)
ci = 2
for pi in range(4):
    hdr(ws8, 1, ci, f'{PERIODS[pi]}\n数量', bg=PCOLORS[pi]); ci+=1
    hdr(ws8, 1, ci, f'{PERIODS[pi]}\n占比', bg=PCOLORS[pi]); ci+=1
hdr(ws8, 1, ci, '环比(数量)', bg=C_HB)
ws8.row_dimensions[1].height = 36

ord_rows8 = [
    ('8日出单(Y)', [y_all[pi] for pi in range(4)], [round(y_all[pi]/valid_cnt[pi]*100,1) if valid_cnt[pi] else 0 for pi in range(4)]),
    ('8日外出单(N)', [n_all[pi] for pi in range(4)], [round(n_all[pi]/valid_cnt[pi]*100,1) if valid_cnt[pi] else 0 for pi in range(4)]),
    ('未出单', [un_all[pi] for pi in range(4)], [round(un_all[pi]/valid_cnt[pi]*100,1) if valid_cnt[pi] else 0 for pi in range(4)]),
    ('有效SKU合计', valid_cnt, [100.0 for _ in range(4)]),
]
for ri, (label, cnts, pcts) in enumerate(ord_rows8, 2):
    cell(ws8, ri, 1, label, bold=True, align='left')
    for pi in range(4):
        cell(ws8, ri, pi*2+2, cnts[pi])
        cell(ws8, ri, pi*2+3, f'{pcts[pi]}%')
    chg = cnts[3] - cnts[2]
    cell(ws8, ri, 10, chg_sign(chg), color=chg_color(chg))

ws8.column_dimensions['A'].width = 18
for ci in range(2, 11): ws8.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 12
bdr(ws8, 1, 5, 1, 10)

# ============ Sheet 9: 新品未出单原因（A/B双板块）============
ws9 = wb.create_sheet('新品未出单原因')

# 标题
hdr(ws9, 1, 1, '新品未出单原因分析 - 4.30-5.6', bg=C_HDR, align='left', size=14)
for c in range(2, 9): hdr(ws9, 1, c, '')

# ═══ A板块：有对手未出单 ═══
hdr(ws9, 3, 1, f'【A. 有对手未出单新品】  本周: {len(has_r_curr)}个  上周: {len(has_r_prev)}个',
    bg=C_ORG, align='left', size=12)
for c in range(2, 9): hdr(ws9, 3, c, '', bg=C_ORG)

# A1 原因分布
hdr(ws9, 4, 1, '【A1】未出单原因分布', bg=C_ORG, align='left')
for c in range(2, 9): hdr(ws9, 4, c, '', bg=C_ORG)
a1_hdrs = ['市场状态','本周SKU','占比','上周SKU','上周占比','变化']
for ci, h in enumerate(a1_hdrs, 1): hdr(ws9, 5, ci, h, bg=C_ORG, size=9)
ri = 6
for rsn in mkt_order_has:
    cur = a1_curr.get(rsn, 0); prv = a1_prev.get(rsn, 0)
    cur_pct = round(cur/len(has_r_curr)*100,1) if has_r_curr else 0
    prv_pct = round(prv/len(has_r_prev)*100,1) if has_r_prev else 0
    chg = cur - prv
    cell(ws9, ri, 1, rsn, bold=True, align='left')
    cell(ws9, ri, 2, cur); cell(ws9, ri, 3, f'{cur_pct}%')
    cell(ws9, ri, 4, prv); cell(ws9, ri, 5, f'{prv_pct}%')
    cell(ws9, ri, 6, chg_sign(chg), color=chg_color(chg))
    ri += 1
# A1合计
cell(ws9, ri, 1, '合计', bold=True, align='left')
cell(ws9, ri, 2, len(has_r_curr), bold=True); cell(ws9, ri, 3, '100%', bold=True)
cell(ws9, ri, 4, len(has_r_prev), bold=True); cell(ws9, ri, 5, '100%', bold=True)
chg = len(has_r_curr)-len(has_r_prev); cell(ws9, ri, 6, chg_sign(chg), bold=True)
ri += 2

# A2 按分析人
hdr(ws9, ri, 1, '【A2】未出单原因 - 按分析人维度', bg=C_ORG, align='left')
for c in range(2, 9): hdr(ws9, ri, c, '', bg=C_ORG)
ri += 1
a2_hdrs = ['分析人'] + mkt_order_has + ['未出单总数']
for ci, h in enumerate(a2_hdrs, 1): hdr(ws9, ri, ci, h, bg=C_ORG, size=9)
ri += 1
for ana in analysts_list:
    d = a2_curr.get(ana, {'total':0,'dist':{}})
    cell(ws9, ri, 1, ana, bold=True, align='left')
    for j, rsn in enumerate(mkt_order_has):
        cell(ws9, ri, j+2, d['dist'].get(rsn, 0))
    cell(ws9, ri, len(mkt_order_has)+2, d['total'])
    ri += 1
ri += 1

# A3 按品线
hdr(ws9, ri, 1, '【A3】未出单原因 - 按品线维度', bg=C_ORG, align='left')
for c in range(2, 9): hdr(ws9, ri, c, '', bg=C_ORG)
ri += 1
a3_hdrs = ['品线'] + mkt_order_has + ['未出单总数']
for ci, h in enumerate(a3_hdrs, 1): hdr(ws9, ri, ci, h, bg=C_ORG, size=9)
ri += 1
for cat in cats:
    d = a3_curr.get(cat, {'total':0,'dist':{}})
    cell(ws9, ri, 1, cat, bold=True, align='left')
    for j, rsn in enumerate(mkt_order_has):
        cell(ws9, ri, j+2, d['dist'].get(rsn, 0))
    cell(ws9, ri, len(mkt_order_has)+2, d['total'])
    ri += 1
ri += 2

# ═══ B板块：无对手未出单 ═══
hdr(ws9, ri, 1, f'【B. 无对手未出单新品】  本周: {len(no_r_curr)}个  上周: {len(no_r_prev)}个',
    bg=C_GRN2, align='left', size=12)
for c in range(2, 9): hdr(ws9, ri, c, '', bg=C_GRN2)
ri += 1

# B1 原因分布
hdr(ws9, ri, 1, '【B1】未出单原因分布', bg=C_GRN2, align='left')
for c in range(2, 9): hdr(ws9, ri, c, '', bg=C_GRN2)
ri += 1
b1_hdrs = ['市场状态','本周SKU','占比','上周SKU','上周占比','变化']
for ci, h in enumerate(b1_hdrs, 1): hdr(ws9, ri, ci, h, bg=C_GRN2, size=9)
ri += 1
for rsn in mkt_order_no:
    cur = b1_curr.get(rsn, 0); prv = b1_prev.get(rsn, 0)
    cur_pct = round(cur/len(no_r_curr)*100,1) if no_r_curr else 0
    prv_pct = round(prv/len(no_r_prev)*100,1) if no_r_prev else 0
    chg = cur - prv
    cell(ws9, ri, 1, rsn, bold=True, align='left')
    cell(ws9, ri, 2, cur); cell(ws9, ri, 3, f'{cur_pct}%')
    cell(ws9, ri, 4, prv); cell(ws9, ri, 5, f'{prv_pct}%')
    cell(ws9, ri, 6, chg_sign(chg), color=chg_color(chg))
    ri += 1
# B1合计
cell(ws9, ri, 1, '合计', bold=True, align='left')
cell(ws9, ri, 2, len(no_r_curr), bold=True); cell(ws9, ri, 3, '100%', bold=True)
cell(ws9, ri, 4, len(no_r_prev), bold=True); cell(ws9, ri, 5, '100%', bold=True)
chg = len(no_r_curr)-len(no_r_prev); cell(ws9, ri, 6, chg_sign(chg), bold=True)
ri += 2

# B2 按分析人
hdr(ws9, ri, 1, '【B2】未出单原因 - 按分析人维度', bg=C_GRN2, align='left')
for c in range(2, 9): hdr(ws9, ri, c, '', bg=C_GRN2)
ri += 1
b2_hdrs = ['分析人'] + mkt_order_no + ['未出单总数']
for ci, h in enumerate(b2_hdrs, 1): hdr(ws9, ri, ci, h, bg=C_GRN2, size=9)
ri += 1
for ana in analysts_list:
    d = b2_curr.get(ana, {'total':0,'dist':{}})
    cell(ws9, ri, 1, ana, bold=True, align='left')
    for j, rsn in enumerate(mkt_order_no):
        cell(ws9, ri, j+2, d['dist'].get(rsn, 0))
    cell(ws9, ri, len(mkt_order_no)+2, d['total'])
    ri += 1
ri += 1

# B3 按品线
hdr(ws9, ri, 1, '【B3】未出单原因 - 按品线维度', bg=C_GRN2, align='left')
for c in range(2, 9): hdr(ws9, ri, c, '', bg=C_GRN2)
ri += 1
b3_hdrs = ['品线'] + mkt_order_no + ['未出单总数']
for ci, h in enumerate(b3_hdrs, 1): hdr(ws9, ri, ci, h, bg=C_GRN2, size=9)
ri += 1
for cat in cats:
    d = b3_curr.get(cat, {'total':0,'dist':{}})
    cell(ws9, ri, 1, cat, bold=True, align='left')
    for j, rsn in enumerate(mkt_order_no):
        cell(ws9, ri, j+2, d['dist'].get(rsn, 0))
    cell(ws9, ri, len(mkt_order_no)+2, d['total'])
    ri += 1

ws9.column_dimensions['A'].width = 20
for c in 'BCDEFGHI': ws9.column_dimensions[c].width = 14
bdr(ws9, 3, ri-1, 1, len(mkt_order_has)+2)

# ============ Sheet 10: 新品PLG维度 ============
ws10 = wb.create_sheet('新品PLG维度')

hdr(ws10, 1, 1, '新品PLG维度分析（4.30-5.6）', bg=C_PLP, align='left', size=12)
for c in range(2, 6): hdr(ws10, 1, c, '', bg=C_PLP)

# 按分析人
hdr(ws10, 3, 1, '按分析人', bg=C_PLP, align='left')
for c in range(2, 6): hdr(ws10, 3, c, '', bg=C_PLP)
hdr(ws10, 4, 1, ''); hdr(ws10, 4, 2, '分析人', bg=C_PLP)
hdr(ws10, 4, 3, '有效SKU数', bg=C_PLP); hdr(ws10, 4, 4, '平均费率', bg=C_PLP)
hdr(ws10, 4, 5, '最高费率', bg=C_PLP)
ri = 5
for ana in analysts_list:
    d = plg_ana_agg.get(ana)
    if not d: continue
    hdr(ws10, ri, 1, '', bg=C_PLP)
    cell(ws10, ri, 2, ana, bold=True, align='left')
    cell(ws10, ri, 3, d['cnt']); cell(ws10, ri, 4, f'{d["avg"]}%'); cell(ws10, ri, 5, f'{d["max"]}%')
    ri += 1

ri += 1
# 按品类
hdr(ws10, ri, 1, '按品类', bg=C_PLP, align='left')
for c in range(2, 6): hdr(ws10, ri, c, '', bg=C_PLP)
ri += 1
hdr(ws10, ri, 1, ''); hdr(ws10, ri, 2, '品类', bg=C_PLP)
hdr(ws10, ri, 3, '有效SKU数', bg=C_PLP); hdr(ws10, ri, 4, '平均费率', bg=C_PLP)
hdr(ws10, ri, 5, '最高费率', bg=C_PLP)
ri += 1
for cat in cats:
    d = plg_cat_agg.get(cat)
    if not d: continue
    hdr(ws10, ri, 1, '', bg=C_PLP)
    cell(ws10, ri, 2, cat, bold=True, align='left')
    cell(ws10, ri, 3, d['cnt']); cell(ws10, ri, 4, f'{d["avg"]}%'); cell(ws10, ri, 5, f'{d["max"]}%')
    ri += 1

ws10.column_dimensions['A'].width = 3; ws10.column_dimensions['B'].width = 16
for c in 'CDE': ws10.column_dimensions[c].width = 14
bdr(ws10, 3, ri-1, 1, 5)

# ── 保存 ──
OUT = r'c:\Users\Hardy\ai-projects\新品复盘\新品周报数据表_4.30-5.6.xlsx'
wb.save(OUT)
print(f'\nXLSX saved: {OUT}')
print(f'Sheets ({len(wb.sheetnames)}): {wb.sheetnames}')
