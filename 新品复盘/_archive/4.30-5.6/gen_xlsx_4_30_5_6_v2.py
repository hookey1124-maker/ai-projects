"""
新品周报 4.30-5.6 汇总XLSX生成（修正版）
修正：出单率公式/分析及时率列/Sheet结构/低占比25列
"""
import openpyxl
import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# ── 颜色常量 ──
C_HDR  = '0F3460'
C_P1   = '6C757D'  # 4.9-4.15
C_P2   = '667EEA'  # 4.16-4.22
C_P3   = '2980B9'  # 4.23-4.29
C_P4   = 'C0392B'  # 4.30-5.6
C_HB   = 'E07B24'  # 环比
C_GRN  = '08845A'
C_RED  = 'C0392B'
C_WHT  = 'FFFFFF'
C_PLP  = '8E44AD'

PERIODS = ['4.9-4.15', '4.16-4.22', '4.23-4.29', '4.30-5.6']
PERIOD_CUTOFFS = [
    datetime.date(2026, 4, 15),
    datetime.date(2026, 4, 22),
    datetime.date(2026, 4, 29),
    datetime.date(2026, 5, 6),
]
PERIOD_COLORS = [C_P1, C_P2, C_P3, C_P4]

# ── 辅助函数 ──
def n(v, default=0):
    if v is None or v == '' or v == '-': return default
    if isinstance(v, str):
        if '#' in v or '—' in v or '∑' in v or '�' in v: return default
        v = v.replace(',', '').replace('$', '').replace('%', '')
    try: return float(v)
    except: return default

def pct(v):
    """将市占比转为百分数"""
    if v is None or v == '' or v == '-': return None
    if isinstance(v, str):
        if '#' in v or '�' in v: return None
        v = v.replace('%', '')
    try:
        fv = float(v)
        return round(fv * 100, 2) if fv <= 1.0 else round(fv, 2)
    except: return None

def get_date(v):
    if v is None: return None
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    return None

def chg_sign(v, as_pct=False):
    if v is None or v == 0: return '-'
    if as_pct:
        return f'+{v:.1f}%' if v > 0 else f'{v:.1f}%'
    return f'+{v:.0f}' if v > 0 else f'{v:.0f}'

def chg_color(v):
    if v is None or v == 0: return '888888'
    return C_GRN if v > 0 else C_RED

# ── XLSX 样式工具 ──
def hdr(ws, r, c, val, bg=C_HDR, fg=C_WHT, bold=True, align='center'):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = Font(bold=bold, color=fg, name='Arial', size=10)
    cl.fill = PatternFill('solid', fgColor=bg)
    cl.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    return cl

def cell(ws, r, c, val, bold=False, color=None, align='center'):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = Font(bold=bold, color=color or '000000', name='Arial', size=9)
    cl.alignment = Alignment(horizontal=align, vertical='center')
    return cl

def add_border(ws, min_r, max_r, min_c, max_c):
    thin = Side(style='thin', color='DDDDDD')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c):
        for c in row:
            c.border = bd

# ── 读取源数据 ──
SRC = r'c:\Users\Hardy\ai-projects\新品复盘\新品检查周源数据和PLP数据.xlsx'
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_main = wb_src['四三数据累计']
ws_plp  = wb_src['PLP明细']

all_rows_raw = list(ws_main.iter_rows(min_row=2, values_only=True))

# ── 列索引 ──
# 销量: 12,13,14,15 | 销售额: 22,23,24,25
# 对手销量: 32,33,34,35 | 市占比: 41,42,43,44
# 8日出单: 59,60,61,62
# 7日频次标签(超7日): 68,69,70,71
# 7日新品频次标签(8日内新品): 77,78,79,80
# 市场状态: 87,88,89,90
# 操作判断: 96,97,98,99
# PLP: 100,101,102,103 (only 4 cols from 4.9)
# PLG: 104,105,106 (only 3 cols: 4.16-4.22, 4.23-4.29, 4.30-5.6)

Q_COLS    = [12, 13, 14, 15]
AMT_COLS  = [22, 23, 24, 25]
RIV_COLS  = [32, 33, 34, 35]
MZB_COLS  = [41, 42, 43, 44]
ORD_COLS  = [59, 60, 61, 62]
FQ_COLS   = [68, 69, 70, 71]   # 7日频次标签 → 超7日低占比未分析
NFQ_COLS  = [77, 78, 79, 80]   # 7日新品频次标签 → 8日内新品无分析
MKT_COLS  = [87, 88, 89, 90]
OP_COLS   = [96, 97, 98, 99]
PLP_COLS  = [100, 101, 102, 103]
PLG_COLS  = [104, 105, 106]    # only last 3 periods

# ── 构建结构化数据 ──
class SKU:
    """每个SKU的结构化数据"""
    pass

skus = []
for row in all_rows_raw:
    if not row[1]: continue  # skip empty SKU
    s = SKU()
    s.sid = row[0]          # 销售编号
    s.sku = str(row[1]).strip()
    s.list_date = get_date(row[2])
    s.first_order = get_date(row[3])
    s.analyst = str(row[4]).strip() if row[4] else ''
    s.category = str(row[5]).strip() if row[5] else ''
    s.exp_type = str(row[6]).strip() if row[6] else ''

    # Per-period data
    s.qtys = [n(row[c]) for c in Q_COLS]
    s.amts = [n(row[c]) for c in AMT_COLS]
    s.rivals = [n(row[c]) for c in RIV_COLS]
    s.mzbs = [pct(row[c]) for c in MZB_COLS]
    s.ord8s = [str(row[c]).strip() if row[c] else '' for c in ORD_COLS]
    s.freqs = [str(row[c]).strip() if row[c] else '' for c in FQ_COLS]      # 7日频次
    s.new_freqs = [str(row[c]).strip() if row[c] else '' for c in NFQ_COLS] # 7日新品频次
    s.mkts = [str(row[c]).strip() if row[c] else '' for c in MKT_COLS]
    s.ops = [str(row[c]).strip() if row[c] else '' for c in OP_COLS]
    s.plp = [str(row[c]).strip() if row[c] else '' for c in PLP_COLS]
    s.plg = [n(row[c]) for c in PLG_COLS]
    skus.append(s)

print(f'读取SKU: {len(skus)}')

# ── 过滤函数 ──
def valid_for_period(s, pi):
    """SKU在该周期是否有数据（已上架且上架日期<=周期截止日）"""
    if not s.list_date: return False
    if s.list_date > PERIOD_CUTOFFS[pi]: return False
    # 未上架状态不计入
    if s.ord8s[pi] == '未上架': return False
    return True

def has_rival(s, pi):
    """有对手（对手销量>0）"""
    return s.rivals[pi] > 0

def is_y(s, pi):
    return s.ord8s[pi] == 'Y'

def is_n(s, pi):
    return s.ord8s[pi] == 'N'

def is_ordered(s, pi):
    """已出单 = Y或N"""
    return s.ord8s[pi] in ('Y', 'N')

def is_unordered(s, pi):
    """真正未出单"""
    return s.ord8s[pi] == '未出单'

# ── 计算各维度数据 ──

# === KPI 总览 ===
total_sku_all = len(skus)

# 各周期有效SKU（排除未上架和未来上架）
valid_counts = []
for pi in range(4):
    cnt = sum(1 for s in skus if valid_for_period(s, pi))
    valid_counts.append(cnt)

# 各周期总销量/销售额
total_qtys = [sum(s.qtys[pi] for s in skus if valid_for_period(s, pi)) for pi in range(4)]
total_amts = [round(sum(s.amts[pi] for s in skus if valid_for_period(s, pi)), 2) for pi in range(4)]

# 各周期有对手SKU数
rival_counts = [sum(1 for s in skus if valid_for_period(s, pi) and has_rival(s, pi)) for pi in range(4)]

# 各周期出单统计 (有对手口径)
yr_counts = [sum(1 for s in skus if valid_for_period(s, pi) and has_rival(s, pi) and is_y(s, pi)) for pi in range(4)]
nr_counts = [sum(1 for s in skus if valid_for_period(s, pi) and has_rival(s, pi) and is_n(s, pi)) for pi in range(4)]
unr_counts = [sum(1 for s in skus if valid_for_period(s, pi) and has_rival(s, pi) and is_unordered(s, pi)) for pi in range(4)]

# 有对手出单率 = (Y+N) / 有对手SKU ★修正★
rival_rates = []
for pi in range(4):
    denom = rival_counts[pi]
    num = yr_counts[pi] + nr_counts[pi]
    rival_rates.append(round(num / denom * 100, 1) if denom > 0 else 0)

# 全量出单统计
y_all = [sum(1 for s in skus if valid_for_period(s, pi) and is_y(s, pi)) for pi in range(4)]
n_all = [sum(1 for s in skus if valid_for_period(s, pi) and is_n(s, pi)) for pi in range(4)]
un_all = [sum(1 for s in skus if valid_for_period(s, pi) and is_unordered(s, pi)) for pi in range(4)]

# 全量出单率 = (Y+N) / (Y+N+未出单)
rate_all = []
for pi in range(4):
    denom = y_all[pi] + n_all[pi] + un_all[pi]
    num = y_all[pi] + n_all[pi]
    rate_all.append(round(num / denom * 100, 1) if denom > 0 else 0)

# 平均市占比
def avg_mzb(pi):
    vals = [s.mzbs[pi] for s in skus if valid_for_period(s, pi) and s.mzbs[pi] is not None]
    return round(sum(vals)/len(vals), 1) if vals else 0

mzb_list = [avg_mzb(pi) for pi in range(4)]

# 低占比新品（当前周期：市占比<75% 且 有对手）
low_skus_cur = [s for s in skus if valid_for_period(s, 3) and
                s.mzbs[3] is not None and s.mzbs[3] < 75 and has_rival(s, 3)]
low_skus_cur.sort(key=lambda s: s.mzbs[3] if s.mzbs[3] else 100)

print(f'=== KPI验证 ===')
print(f'有效SKU: {valid_counts}')
print(f'总销量: {total_qtys}')
print(f'有对手SKU: {rival_counts}')
tmp = []
for i in range(4):
    tmp.append(f"{yr_counts[i]}+{nr_counts[i]}/{rival_counts[i]}={rival_rates[i]}%")
print(f'有对手出单 Yr+Nr / 有对手: {tmp}')
print(f'全量Y/N/未: {list(zip(y_all, n_all, un_all))}')
print(f'全量出单率: {rate_all}')
print(f'平均市占比: {mzb_list}')
print(f'低占比新品: {len(low_skus_cur)}')

# === 品类维度 ===
categories = sorted(set(s.category for s in skus if s.category))
cat_data = {}
for cat in categories:
    cat_skus = [s for s in skus if s.category == cat]
    d = {'skus': len(cat_skus)}
    for pi in range(4):
        valid = [s for s in cat_skus if valid_for_period(s, pi)]
        d[f'valid_{pi}'] = len(valid)
        d[f'qty_{pi}'] = round(sum(s.qtys[pi] for s in valid))
        d[f'amt_{pi}'] = round(sum(s.amts[pi] for s in valid), 2)
        has_r = [s for s in valid if has_rival(s, pi)]
        d[f'riv_{pi}'] = len(has_r)
        yr = sum(1 for s in has_r if is_y(s, pi))
        nr = sum(1 for s in has_r if is_n(s, pi))
        d[f'yr_{pi}'] = yr
        d[f'nr_{pi}'] = nr
        num = yr + nr
        denom = len(has_r)
        d[f'rate_{pi}'] = round(num / denom * 100, 1) if denom > 0 else 0
    cat_data[cat] = d

print(f'\n=== 品类出单率验证 ===')
for cat in categories:
    d = cat_data[cat]
    print(f'{cat}: SKU={d["skus"]}, 出单率={[d[f"rate_{pi}"] for pi in range(4)]}')

# === 分析人维度 ===
analysts = sorted(set(s.analyst for s in skus if s.analyst))
ana_data = {}
for ana in analysts:
    ana_skus = [s for s in skus if s.analyst == ana]
    d = {'skus': len(ana_skus)}
    for pi in range(4):
        valid = [s for s in ana_skus if valid_for_period(s, pi)]
        d[f'valid_{pi}'] = len(valid)
        d[f'qty_{pi}'] = round(sum(s.qtys[pi] for s in valid))
        has_r = [s for s in valid if has_rival(s, pi)]
        d[f'riv_{pi}'] = len(has_r)
        yr = sum(1 for s in has_r if is_y(s, pi))
        nr = sum(1 for s in has_r if is_n(s, pi))
        d[f'yr_{pi}'] = yr
        d[f'nr_{pi}'] = nr
        num = yr + nr
        denom = len(has_r)
        d[f'rate_{pi}'] = round(num / denom * 100, 1) if denom > 0 else 0
    ana_data[ana] = d

print(f'\n=== 分析人出单率验证 ===')
for ana in analysts:
    d = ana_data[ana]
    print(f'{ana}: SKU={d["skus"]}, 出单率={[d[f"rate_{pi}"] for pi in range(4)]}')

# === 拓展类型 ===
exp_types = ['原开品', '拓展品', '组合件']
exp_data = {}
for exp in exp_types:
    exp_skus = [s for s in skus if s.exp_type == exp]
    d = {'skus': len(exp_skus)}
    for pi in range(4):
        valid = [s for s in exp_skus if valid_for_period(s, pi)]
        d[f'valid_{pi}'] = len(valid)
        d[f'qty_{pi}'] = round(sum(s.qtys[pi] for s in valid))
        has_r = [s for s in valid if has_rival(s, pi)]
        d[f'riv_{pi}'] = len(has_r)
        yr = sum(1 for s in has_r if is_y(s, pi))
        nr = sum(1 for s in has_r if is_n(s, pi))
        d[f'yr_{pi}'] = yr
        d[f'nr_{pi}'] = nr
        num = yr + nr
        denom = len(has_r)
        d[f'rate_{pi}'] = round(num / denom * 100, 1) if denom > 0 else 0
    exp_data[exp] = d

print(f'\n=== 拓展类型出单率验证 ===')
for exp in exp_types:
    d = exp_data[exp]
    print(f'{exp}: SKU={d["skus"]}, 出单率={[d[f"rate_{pi}"] for pi in range(4)]}')

# === 分析及时率（修正：正确使用两个频次标签列）===
# 8日内新品无分析 = 7日新品频次标签 = "异常" (col 77-80)
# 超7日低占比未分析 = 7日频次标签 = "异常" (col 68-71)
# 及时分析 = 以上两列都不是"异常"
# 未上架不计入

timely_in = []      # 及时分析
timely_new8 = []    # 8日内新品无分析 (new_freq = 异常)
timely_over7 = []   # 超7日低占比未分析 (freq = 异常)

for pi in range(4):
    valid = [s for s in skus if valid_for_period(s, pi)]
    ti = 0
    tn = 0
    to = 0
    for s in valid:
        nf = s.new_freqs[pi]  # 7日新品频次标签
        fq = s.freqs[pi]       # 7日频次标签

        is_nf_abnormal = (nf == '异常')
        is_fq_abnormal = (fq == '异常')

        if is_nf_abnormal:
            tn += 1  # 8日内新品无分析
        elif is_fq_abnormal:
            to += 1  # 超7日低占比未分析
        else:
            ti += 1  # 及时分析

    timely_in.append(ti)
    timely_new8.append(tn)
    timely_over7.append(to)
    total_check = ti + tn + to
    print(f'{PERIODS[pi]} 分析及时率: 及时={ti}, 8日内新品无分析={tn}, 超7日未分析={to}, 合计={total_check}')

# 按分析人的及时率（仅当期）
timely_by_analyst = {}
for ana in analysts:
    valid = [s for s in skus if s.analyst == ana and valid_for_period(s, 3)]
    ti = sum(1 for s in valid if s.new_freqs[3] != '异常' and s.freqs[3] != '异常')
    tn = sum(1 for s in valid if s.new_freqs[3] == '异常')
    to = sum(1 for s in valid if s.new_freqs[3] != '异常' and s.freqs[3] == '异常')
    total = len(valid)
    rate = round(ti / total * 100, 1) if total > 0 else 0
    timely_by_analyst[ana] = {'total': total, 'timely': ti, 'new8': tn, 'over7': to, 'rate': rate}

# === 新品出单情况（全量口径，含占比）===
# 8日出单(Y) / 8日外出单(N) / 未出单
order_summary = []
for pi in range(4):
    valid = [s for s in skus if valid_for_period(s, pi)]
    total = len(valid)
    yc = sum(1 for s in valid if is_y(s, pi))
    nc = sum(1 for s in valid if is_n(s, pi))
    uc = sum(1 for s in valid if is_unordered(s, pi))
    order_summary.append({
        'total': total, 'y': yc, 'n': nc, 'un': uc,
        'y_pct': round(yc/total*100, 1) if total else 0,
        'n_pct': round(nc/total*100, 1) if total else 0,
        'un_pct': round(uc/total*100, 1) if total else 0,
    })

# === 新品未出单（仅统计 ord8 = 未出单 的SKU，按市场状态分布）===
unorder_reasons = ['竞争无优势', '无市场', '站内无价格优势', '站外出单']
unorder_data = {r: [0]*4 for r in unorder_reasons}

for pi in range(4):
    for s in skus:
        if not valid_for_period(s, pi): continue
        if not is_unordered(s, pi): continue
        mkt = s.mkts[pi]
        mapped = '竞争无优势'
        if '无市场' in mkt:
            mapped = '无市场'
        elif '价格' in mkt or ('站内' in mkt and '价格' in mkt):
            mapped = '站内无价格优势'
        elif '站外' in mkt:
            mapped = '站外出单'
        elif '竞争' in mkt:
            mapped = '竞争无优势'
        unorder_data[mapped][pi] += 1

print(f'\n=== 未出单原因 ===')
for rsn in unorder_reasons:
    print(f'{rsn}: {unorder_data[rsn]}')

# === PLP 4.30-5.6 ===
plp_new_rows = []
for row in ws_plp.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    period_str = str(row[0])
    if '4.30' in period_str or '4.30-5.6' in period_str:
        plp_new_rows.append(row)

def clean_plp(row):
    try:
        return {
            'sku': row[2], 'analyst': str(row[8] or ''), 'cat': str(row[9] or ''),
            'expn': str(row[10] or ''), 'imp': n(row[11]), 'clk': n(row[12]),
            'orders': n(row[13]), 'spend': n(row[14]), 'sales': n(row[15]),
        }
    except: return None

plp_details = [r for row in plp_new_rows if row[1] != '总数据' and (r := clean_plp(row)) and r['sku']]

def agg_plp(rows, key_fn):
    agg = {}
    for r in rows:
        k = key_fn(r)
        if k not in agg:
            agg[k] = {'imp': 0, 'clk': 0, 'ord': 0, 'spend': 0, 'sales': 0, 'sku_set': set()}
        a = agg[k]
        a['imp'] += r['imp']; a['clk'] += r['clk']; a['ord'] += r['orders']
        a['spend'] += r['spend']; a['sales'] += r['sales']
        a['sku_set'].add(r['sku'])
    result = {}
    for k, a in agg.items():
        roas = round(a['sales']/a['spend'], 2) if a['spend'] > 0 else 0
        cvr  = round(a['ord']/a['clk']*100, 2) if a['clk'] > 0 else 0
        ctr  = round(a['clk']/a['imp']*100, 4) if a['imp'] > 0 else 0
        acos = round(a['spend']/a['sales']*100, 2) if a['sales'] > 0 else 0
        result[k] = {
            'skus': len(a['sku_set']), 'imp': int(a['imp']), 'clk': int(a['clk']),
            'ord': int(a['ord']), 'spend': round(a['spend'], 2),
            'sales': round(a['sales'], 2), 'roas': roas, 'cvr': cvr, 'ctr': ctr, 'acos': acos
        }
    return result

plp_by_analyst = agg_plp(plp_details, lambda r: r['analyst'])
plp_by_cat = agg_plp(plp_details, lambda r: r['cat'])
plp_by_expn = agg_plp(plp_details, lambda r: r['expn'])

# PLP总计
plp_total_row = next((r for r in plp_new_rows if r[1] == '总数据'), None)
if plp_total_row:
    plp_total = {
        'skus': int(n(plp_total_row[5] or plp_total_row[6])),
        'imp': int(n(plp_total_row[11])), 'clk': int(n(plp_total_row[12])),
        'ord': int(n(plp_total_row[13])), 'spend': round(n(plp_total_row[14]), 2),
        'sales': round(n(plp_total_row[15]), 2), 'roas': round(n(plp_total_row[16]), 2),
        'cvr': round(n(plp_total_row[17])*100, 2), 'ctr': round(n(plp_total_row[18])*100, 4),
        'acos': round(n(plp_total_row[21])*100, 2),
    }
else:
    total_imp = sum(d['imp'] for d in plp_by_analyst.values())
    total_clk = sum(d['clk'] for d in plp_by_analyst.values())
    total_ord = sum(d['ord'] for d in plp_by_analyst.values())
    total_spend = sum(d['spend'] for d in plp_by_analyst.values())
    total_sales = sum(d['sales'] for d in plp_by_analyst.values())
    plp_total = {
        'skus': len(plp_details), 'imp': total_imp, 'clk': total_clk,
        'ord': total_ord, 'spend': round(total_spend, 2), 'sales': round(total_sales, 2),
        'roas': round(total_sales/total_spend, 2) if total_spend > 0 else 0,
        'cvr': round(total_ord/total_clk*100, 2) if total_clk > 0 else 0,
        'ctr': round(total_clk/total_imp*100, 4) if total_imp > 0 else 0,
        'acos': round(total_spend/total_sales*100, 2) if total_sales > 0 else 0,
    }

# === PLG费率 ===
# 本周PLG最高费率 col 106
plg_cur = []
for s in skus:
    if valid_for_period(s, 3) and s.plg[2] > 0:  # plg[2] = 4.30-5.6 col106
        plg_cur.append(s)

plg_by_analyst = defaultdict(list)
plg_by_cat = defaultdict(list)
for s in plg_cur:
    plg_by_analyst[s.analyst or '未知'].append(s.plg[2])
    plg_by_cat[s.category or '未分类'].append(s.plg[2])

plg_ana_agg = {}
for k, vals in plg_by_analyst.items():
    plg_ana_agg[k] = {'count': len(vals), 'avg': round(sum(vals)/len(vals)*100, 1), 'max': round(max(vals)*100, 1)}

plg_cat_agg = {}
for k, vals in plg_by_cat.items():
    plg_cat_agg[k] = {'count': len(vals), 'avg': round(sum(vals)/len(vals)*100, 1), 'max': round(max(vals)*100, 1)}

wb_src.close()

print(f'\n=== PLP ===')
print(f'总计: SKU={plp_total["skus"]}, 花费=${plp_total["spend"]}, ROAS={plp_total["roas"]}, ACOS={plp_total["acos"]}%')
print(f'\n=== PLG ===')
print(f'有PLG费率SKU: {len(plg_cur)}')

# ══════════════════════════════════════════════════════════════
# 生成 XLSX
# ══════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()

# ============ Sheet 1: 总体数据 ============
ws1 = wb.active
ws1.title = '总体数据'
h1 = ['指标', PERIODS[0], PERIODS[1], PERIODS[2], PERIODS[3], '环比']
for ci, h in enumerate(h1, 1):
    bg = PERIOD_COLORS[ci-2] if 2 <= ci <= 5 else (C_HB if ci == 6 else C_HDR)
    hdr(ws1, 1, ci, h, bg=bg)
ws1.row_dimensions[1].height = 26

kpi_rows = [
    ('已上架SKU数', valid_counts, True),
    ('总销量', [round(v) for v in total_qtys], True),
    ('总销售额(USD)', total_amts, True),
    ('有对手SKU数', rival_counts, True),
    ('全量出单率', [f'{v}%' for v in rate_all], False),
    ('有对手出单率(Y+N/有对手)', [f'{v}%' for v in rival_rates], False),
    ('有对手-8日出单(Y)', yr_counts, True),
    ('有对手-8日外出单(N)', nr_counts, True),
    ('有对手-未出单', unr_counts, True),
    ('全量-8日出单(Y)', y_all, True),
    ('全量-8日外出单(N)', n_all, True),
    ('全量-未出单', un_all, True),
    ('平均市占比', [f'{v}%' for v in mzb_list], False),
    ('低占比新品(mzb<75%有竞品)', ['-', '-', '-', len(low_skus_cur)], False),
    ('PLP花费(USD)', ['-', '-', '-', plp_total['spend']], False),
    ('PLP-ROAS', ['-', '-', '-', plp_total['roas']], False),
    ('PLP-ACOS', ['-', '-', '-', f"{plp_total['acos']}%"], False),
]

for ri, (label, vals, is_num) in enumerate(kpi_rows, 2):
    cell(ws1, ri, 1, label, bold=True, align='left')
    for pi in range(4):
        cell(ws1, ri, pi+2, vals[pi] if pi < len(vals) else '-')
    # 环比
    if is_num and len(vals) >= 4:
        try:
            v_cur = float(str(vals[3]).replace('%', '').replace('$', '').replace(',', ''))
            v_prv = float(str(vals[2]).replace('%', '').replace('$', '').replace(',', ''))
            chg = v_cur - v_prv
            c = chg_color(chg)
            cell(ws1, ri, 6, chg_sign(chg), color=c)
        except:
            cell(ws1, ri, 6, '-')
    else:
        cell(ws1, ri, 6, '-')

ws1.column_dimensions['A'].width = 30
for c in 'BCDEF':
    ws1.column_dimensions[c].width = 16
add_border(ws1, 1, len(kpi_rows)+1, 1, 6)

# ============ Sheet 2: 品线维度 ============
ws2 = wb.create_sheet('品线维度')
cat_hdrs = ['品线', '累计SKU',
            f'{PERIODS[0]}\n有效SKU', f'{PERIODS[0]}\n销量', f'{PERIODS[0]}\n销售额', f'{PERIODS[0]}\n出单率',
            f'{PERIODS[1]}\n有效SKU', f'{PERIODS[1]}\n销量', f'{PERIODS[1]}\n销售额', f'{PERIODS[1]}\n出单率',
            f'{PERIODS[2]}\n有效SKU', f'{PERIODS[2]}\n销量', f'{PERIODS[2]}\n销售额', f'{PERIODS[2]}\n出单率',
            f'{PERIODS[3]}\n有效SKU', f'{PERIODS[3]}\n销量', f'{PERIODS[3]}\n销售额', f'{PERIODS[3]}\n出单率',
            '销量环比', '出单率环比']
for ci, h in enumerate(cat_hdrs, 1):
    bg = C_HDR
    if '4.9' in str(h): bg = C_P1
    elif '4.16' in str(h): bg = C_P2
    elif '4.23' in str(h): bg = C_P3
    elif '4.30' in str(h): bg = C_P4
    elif '环比' in str(h): bg = C_HB
    hdr(ws2, 1, ci, h, bg=bg)
ws2.row_dimensions[1].height = 36

for ri, cat in enumerate(categories, 2):
    d = cat_data[cat]
    q_chg = d['qty_3'] - d['qty_2']
    r_chg = d['rate_3'] - d['rate_2']
    vals = [cat, d['skus']]
    for pi in range(4):
        vals += [d[f'valid_{pi}'], d[f'qty_{pi}'], f'${d[f"amt_{pi}"]:.2f}', f'{d[f"rate_{pi}"]}%']
    vals += [chg_sign(q_chg), chg_sign(r_chg, as_pct=True)]
    for ci, v in enumerate(vals, 1):
        c = None
        if ci == len(vals)-1: c = chg_color(q_chg)
        if ci == len(vals): c = chg_color(r_chg)
        cell(ws2, ri, ci, v, bold=(ci==1), color=c, align='left' if ci==1 else 'center')

# 合计行
tr2 = len(categories) + 2
total_vals = ['合计', sum(cat_data[c]['skus'] for c in categories)]
for pi in range(4):
    total_vals += [
        sum(cat_data[c][f'valid_{pi}'] for c in categories),
        round(sum(cat_data[c][f'qty_{pi}'] for c in categories)),
        f'${round(sum(cat_data[c][f"amt_{pi}"] for c in categories), 2):.2f}',
        '',  # 出单率不合计
    ]
tq_chg = sum(cat_data[c]['qty_3'] for c in categories) - sum(cat_data[c]['qty_2'] for c in categories)
total_vals += [chg_sign(tq_chg), '']
for ci, v in enumerate(total_vals, 1):
    cell(ws2, tr2, ci, v, bold=True, align='left' if ci==1 else 'center')

for ci in range(1, len(cat_hdrs)+1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14 if ci > 2 else (16 if ci == 1 else 10)
ws2.column_dimensions['A'].width = 16
add_border(ws2, 1, tr2, 1, len(cat_hdrs))

# ============ Sheet 3: 分析人维度 ============
ws3 = wb.create_sheet('分析人维度')
ana_hdrs = ['分析人', '累计SKU',
            f'{PERIODS[0]}\n有效SKU', f'{PERIODS[0]}\n销量', f'{PERIODS[0]}\n出单率',
            f'{PERIODS[1]}\n有效SKU', f'{PERIODS[1]}\n销量', f'{PERIODS[1]}\n出单率',
            f'{PERIODS[2]}\n有效SKU', f'{PERIODS[2]}\n销量', f'{PERIODS[2]}\n出单率',
            f'{PERIODS[3]}\n有效SKU', f'{PERIODS[3]}\n销量', f'{PERIODS[3]}\n出单率',
            '销量环比', '出单率环比']
for ci, h in enumerate(ana_hdrs, 1):
    bg = C_HDR
    if '4.9' in str(h): bg = C_P1
    elif '4.16' in str(h): bg = C_P2
    elif '4.23' in str(h): bg = C_P3
    elif '4.30' in str(h): bg = C_P4
    elif '环比' in str(h): bg = C_HB
    hdr(ws3, 1, ci, h, bg=bg)
ws3.row_dimensions[1].height = 36

for ri, ana in enumerate(analysts, 2):
    d = ana_data[ana]
    q_chg = d['qty_3'] - d['qty_2']
    r_chg = d['rate_3'] - d['rate_2']
    vals = [ana, d['skus']]
    for pi in range(4):
        vals += [d[f'valid_{pi}'], d[f'qty_{pi}'], f'{d[f"rate_{pi}"]}%']
    vals += [chg_sign(q_chg), chg_sign(r_chg, as_pct=True)]
    for ci, v in enumerate(vals, 1):
        c = None
        if ci == len(vals)-1: c = chg_color(q_chg)
        if ci == len(vals): c = chg_color(r_chg)
        cell(ws3, ri, ci, v, bold=(ci==1), color=c, align='left' if ci==1 else 'center')

tr3 = len(analysts) + 2
add_border(ws3, 1, tr3, 1, len(ana_hdrs))
ws3.column_dimensions['A'].width = 14
for ci in range(2, len(ana_hdrs)+1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 12

# ============ Sheet 4: 拓展类型 ============
ws4 = wb.create_sheet('拓展类型')
exp_hdrs = ['拓展类型', '累计SKU',
            f'{PERIODS[0]}\n有效SKU', f'{PERIODS[0]}\n销量', f'{PERIODS[0]}\n出单率',
            f'{PERIODS[1]}\n有效SKU', f'{PERIODS[1]}\n销量', f'{PERIODS[1]}\n出单率',
            f'{PERIODS[2]}\n有效SKU', f'{PERIODS[2]}\n销量', f'{PERIODS[2]}\n出单率',
            f'{PERIODS[3]}\n有效SKU', f'{PERIODS[3]}\n销量', f'{PERIODS[3]}\n出单率',
            '销量环比', '出单率环比']
for ci, h in enumerate(exp_hdrs, 1):
    bg = C_HDR
    if '4.9' in str(h): bg = C_P1
    elif '4.16' in str(h): bg = C_P2
    elif '4.23' in str(h): bg = C_P3
    elif '4.30' in str(h): bg = C_P4
    elif '环比' in str(h): bg = C_HB
    hdr(ws4, 1, ci, h, bg=bg)
ws4.row_dimensions[1].height = 36

ri = 2
for exp in exp_types:
    d = exp_data[exp]
    if d['skus'] == 0: continue
    q_chg = d['qty_3'] - d['qty_2']
    r_chg = d['rate_3'] - d['rate_2']
    vals = [exp, d['skus']]
    for pi in range(4):
        vals += [d[f'valid_{pi}'], d[f'qty_{pi}'], f'{d[f"rate_{pi}"]}%']
    vals += [chg_sign(q_chg), chg_sign(r_chg, as_pct=True)]
    for ci, v in enumerate(vals, 1):
        c = None
        if ci == len(vals)-1: c = chg_color(q_chg)
        if ci == len(vals): c = chg_color(r_chg)
        cell(ws4, ri, ci, v, bold=(ci==1), color=c, align='left' if ci==1 else 'center')
    ri += 1

add_border(ws4, 1, ri-1, 1, len(exp_hdrs))
ws4.column_dimensions['A'].width = 14
for ci in range(2, len(exp_hdrs)+1):
    ws4.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 12

# ============ Sheet 5: 分析及时率 ============
ws5 = wb.create_sheet('分析及时率')

# 5.1 汇总表
hdr(ws5, 1, 1, '分析及时率汇总', bg=C_HDR, align='left')
hdr(ws5, 1, 2, '', bg=C_HDR)
for pi in range(4):
    hdr(ws5, 1, pi+3, PERIODS[pi], bg=PERIOD_COLORS[pi])
hdr(ws5, 1, 7, '环比', bg=C_HB)

t_rows = [
    ('及时分析', timely_in),
    ('8日内新品无分析', timely_new8),
    ('超7日低占比未分析', timely_over7),
]
for ri, (label, vals) in enumerate(t_rows, 2):
    hdr(ws5, ri, 1, '', bg=C_HDR)
    cell(ws5, ri, 2, label, bold=True, align='left')
    for pi in range(4):
        cell(ws5, ri, pi+3, vals[pi])
    chg = vals[3] - vals[2]
    cell(ws5, ri, 7, chg_sign(chg), color=chg_color(chg))

# 合计行
total_timely = [timely_in[i] + timely_new8[i] + timely_over7[i] for i in range(4)]
hdr(ws5, 5, 1, '', bg=C_HDR)
cell(ws5, 5, 2, '合计', bold=True, align='left')
for pi in range(4):
    cell(ws5, 5, pi+3, total_timely[pi], bold=True)
tchg = total_timely[3] - total_timely[2]
cell(ws5, 5, 7, chg_sign(tchg), bold=True, color=chg_color(tchg))

# 及时率
timely_rate = [round(timely_in[i]/total_timely[i]*100, 1) if total_timely[i] > 0 else 0 for i in range(4)]
hdr(ws5, 6, 1, '', bg=C_HDR)
cell(ws5, 6, 2, '及时率', bold=True, align='left')
for pi in range(4):
    cell(ws5, 6, pi+3, f'{timely_rate[pi]}%')
r_chg = timely_rate[3] - timely_rate[2]
cell(ws5, 6, 7, chg_sign(r_chg, as_pct=True), color=chg_color(r_chg))

# 5.2 按分析人明细（仅当期 4.30-5.6）
hdr(ws5, 8, 1, '按分析人明细（4.30-5.6）', bg=C_P4, align='left')
hdr(ws5, 8, 2, '', bg=C_P4); hdr(ws5, 8, 3, '', bg=C_P4)
hdr(ws5, 8, 4, '负责SKU', bg=C_P4)
hdr(ws5, 8, 5, '及时分析', bg=C_P4)
hdr(ws5, 8, 6, '8日内新品无分析', bg=C_P4)
hdr(ws5, 8, 7, '超7日未分析', bg=C_P4)
hdr(ws5, 8, 8, '及时率', bg=C_P4)

for ri, ana in enumerate(analysts, 9):
    d = timely_by_analyst[ana]
    hdr(ws5, ri, 1, '', bg=C_HDR)
    hdr(ws5, ri, 2, '', bg=C_HDR)
    cell(ws5, ri, 3, ana, bold=True, align='left')
    cell(ws5, ri, 4, d['total'])
    cell(ws5, ri, 5, d['timely'])
    cell(ws5, ri, 6, d['new8'])
    cell(ws5, ri, 7, d['over7'])
    cell(ws5, ri, 8, f'{d["rate"]}%')

add_border(ws5, 1, 9+len(analysts)-1, 1, 8)
ws5.column_dimensions['A'].width = 3
ws5.column_dimensions['B'].width = 22
for c in 'CDEFGH':
    ws5.column_dimensions[c].width = 16

# ============ Sheet 6: 低占比新品 (25列) ============
ws6 = wb.create_sheet('低占比新品')
low_hdrs = [
    '销售编号', 'SKU', '上架日期', '分析人', '品类', '拓展类型',           # 1-6
    '本周销量', '销量环比',                                                 # 7-8
    '本周销售额', '销售额环比',                                             # 9-10
    '上期末对手销量', '本期末对手销量', '对手销量环比',                      # 11-13
    '上期末市占比', '本期末市占比', '市占比环比',                            # 14-16
    '本期末8日出单',                                                       # 17
    '本期末7日频次标签',                                                    # 18
    '本期末7日新品频次标签',                                                # 19
    '上期末市场状态',                                                       # 20
    '本周操作判断',                                                         # 21
    '本期末市场状态',                                                       # 22
    '本周开启PLP',                                                         # 23
    'PLG最高费率',                                                          # 24
]
for ci, h in enumerate(low_hdrs, 1):
    hdr(ws6, 1, ci, h, bg=C_RED)
ws6.row_dimensions[1].height = 30

for ri, s in enumerate(low_skus_cur, 2):
    # Previous period values (4.23-4.29, pi=2)
    qty_prev = s.qtys[2]
    amt_prev = s.amts[2]
    riv_prev = s.rivals[2]
    mzb_prev = s.mzbs[2]
    mkt_prev = s.mkts[2]

    # Current period values (4.30-5.6, pi=3)
    qty_cur = s.qtys[3]
    amt_cur = s.amts[3]
    riv_cur = s.rivals[3]
    mzb_cur = s.mzbs[3]

    qty_chg_val = qty_cur - qty_prev
    amt_chg_val = amt_cur - amt_prev
    riv_chg_val = riv_cur - riv_prev
    mzb_chg_val = (mzb_cur or 0) - (mzb_prev or 0)

    cols = [
        s.sid, s.sku,
        str(s.list_date) if s.list_date else '',
        s.analyst, s.category, s.exp_type,
        round(qty_cur), chg_sign(qty_chg_val),
        f'${round(amt_cur, 2):.2f}', chg_sign(amt_chg_val),
        round(riv_prev), round(riv_cur), chg_sign(riv_chg_val),
        f'{mzb_prev}%' if mzb_prev is not None else '-',
        f'{mzb_cur}%' if mzb_cur is not None else '-',
        chg_sign(mzb_chg_val, as_pct=True) if mzb_prev is not None and mzb_cur is not None else '-',
        s.ord8s[3] or '-',
        s.freqs[3] or '-',
        s.new_freqs[3] or '-',
        mkt_prev if mkt_prev else '-',
        s.ops[3] or '-',
        s.mkts[3] or '-',
        s.plp[3] if s.plp[3] else '-',
        f'{round(s.plg[2]*100, 1)}%' if s.plg[2] > 0 else '-',
    ]
    for ci, v in enumerate(cols, 1):
        c_color = None
        if ci == 16:  # 市占比
            try:
                mzb_v = float(str(v).replace('%', ''))
                c_color = C_RED if mzb_v < 25 else (C_HB if mzb_v < 50 else None)
            except: pass
        cell(ws6, ri, ci, v, color=c_color, align='left' if ci in [2,4,5,6] else 'center')

add_border(ws6, 1, len(low_skus_cur)+1, 1, len(low_hdrs))
col_widths_6 = [10, 22, 12, 12, 10, 14, 10, 12, 10, 12, 14, 14, 12, 12, 12, 12, 12, 16, 18, 14, 12, 14, 12, 14, 12]
for ci, w in enumerate(col_widths_6, 1):
    ws6.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

# ============ Sheet 7: 新品PLP ============
ws7 = wb.create_sheet('新品PLP')
plp_hdrs = ['维度', '广告SKU', '曝光量', '点击量', '订单数', '花费(USD)', '销售额(USD)', 'ROAS', 'CVR%', 'CTR%', 'ACOS%']
for ci, h in enumerate(plp_hdrs, 1):
    hdr(ws7, 1, ci, h, bg=C_PLP)
ws7.row_dimensions[1].height = 26

def write_plp_row(ws, r, label, d):
    vals = [label, d.get('skus', '-'), d['imp'], d['clk'], d['ord'],
            round(d['spend'], 2), round(d['sales'], 2),
            d['roas'], f"{d['cvr']}%", f"{d['ctr']:.4f}%", f"{d['acos']}%"]
    for ci, v in enumerate(vals, 1):
        cell(ws, r, ci, v, bold=(ci==1), align='left' if ci==1 else 'center')

ri = 2
cell(ws7, ri, 1, '── 总计 ──', bold=True, align='left')
for c in range(2, 12):
    cell(ws7, ri, c, '')
ri += 1
write_plp_row(ws7, ri, '4.30-5.6合计', plp_total)
ri += 2

for section_title, data_dict in [('── 按分析人 ──', plp_by_analyst),
                                   ('── 按品类 ──', plp_by_cat),
                                   ('── 按拓展类型 ──', plp_by_expn)]:
    cell(ws7, ri, 1, section_title, bold=True, align='left')
    for c in range(2, 12):
        cell(ws7, ri, c, '')
    ri += 1
    for k in sorted(data_dict.keys()):
        write_plp_row(ws7, ri, k, data_dict[k])
        ri += 1
    ri += 1

add_border(ws7, 1, ri-1, 1, len(plp_hdrs))
col_widths_7 = [18, 10, 12, 10, 10, 14, 14, 10, 10, 10, 10]
for ci, w in enumerate(col_widths_7, 1):
    ws7.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

# ============ Sheet 8: 新品出单情况 ============
ws8 = wb.create_sheet('新品出单情况')
order_hdrs = ['出单情况']
for pi in range(4):
    order_hdrs.append(f'{PERIODS[pi]}\n数量')
    order_hdrs.append(f'{PERIODS[pi]}\n占比')
order_hdrs.append('环比(数量)')
for ci, h in enumerate(order_hdrs, 1):
    bg = C_HDR
    if '4.9' in str(h): bg = C_P1
    elif '4.16' in str(h): bg = C_P2
    elif '4.23' in str(h): bg = C_P3
    elif '4.30' in str(h): bg = C_P4
    elif '环比' in str(h): bg = C_HB
    hdr(ws8, 1, ci, h, bg=bg)
ws8.row_dimensions[1].height = 36

order_rows = [
    ('8日出单(Y)', [order_summary[pi]['y'] for pi in range(4)],
     [order_summary[pi]['y_pct'] for pi in range(4)]),
    ('8日外出单(N)', [order_summary[pi]['n'] for pi in range(4)],
     [order_summary[pi]['n_pct'] for pi in range(4)]),
    ('未出单', [order_summary[pi]['un'] for pi in range(4)],
     [order_summary[pi]['un_pct'] for pi in range(4)]),
    ('有效SKU合计', [order_summary[pi]['total'] for pi in range(4)],
     [100.0 for _ in range(4)]),
]

for ri, (label, counts, pcts) in enumerate(order_rows, 2):
    cell(ws8, ri, 1, label, bold=True, align='left')
    for pi in range(4):
        cell(ws8, ri, pi*2+2, counts[pi])
        cell(ws8, ri, pi*2+3, f'{pcts[pi]}%')
    chg = counts[3] - counts[2]
    cell(ws8, ri, 10, chg_sign(chg), color=chg_color(chg))

add_border(ws8, 1, 5, 1, 10)
ws8.column_dimensions['A'].width = 18
for ci in range(2, 11):
    ws8.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 12

# ============ Sheet 9: 新品未出单原因 ============
ws9 = wb.create_sheet('新品未出单原因')
un_hdrs = ['未出单原因', PERIODS[0], PERIODS[1], PERIODS[2], PERIODS[3], '环比']
for ci, h in enumerate(un_hdrs, 1):
    bg = PERIOD_COLORS[ci-2] if 2 <= ci <= 5 else (C_HB if ci == 6 else C_HDR)
    hdr(ws9, 1, ci, h, bg=bg)

for ri, rsn in enumerate(unorder_reasons, 2):
    vals = unorder_data[rsn]
    chg = vals[3] - vals[2]
    cell(ws9, ri, 1, rsn, bold=True, align='left')
    for pi in range(4):
        cell(ws9, ri, pi+2, vals[pi])
    cell(ws9, ri, 6, chg_sign(chg), color=chg_color(chg))

# 合计
cell(ws9, 6, 1, '合计', bold=True, align='left')
for pi in range(4):
    total_un = sum(unorder_data[r][pi] for r in unorder_reasons)
    cell(ws9, 6, pi+2, total_un, bold=True)
uchg = sum(unorder_data[r][3] for r in unorder_reasons) - sum(unorder_data[r][2] for r in unorder_reasons)
cell(ws9, 6, 6, chg_sign(uchg), bold=True, color=chg_color(uchg))

add_border(ws9, 1, 6, 1, 6)
ws9.column_dimensions['A'].width = 20
for c in 'BCDEF':
    ws9.column_dimensions[c].width = 12

# ============ Sheet 10: 新品PLG维度 ============
ws10 = wb.create_sheet('新品PLG维度')

# 按分析人
hdr(ws10, 1, 1, '按分析人', bg=C_PLP, align='left')
hdr(ws10, 1, 2, '', bg=C_PLP)
hdr(ws10, 1, 3, '有效SKU数', bg=C_PLP)
hdr(ws10, 1, 4, '平均费率', bg=C_PLP)
hdr(ws10, 1, 5, '最高费率', bg=C_PLP)

ri = 2
for ana in sorted(plg_ana_agg.keys()):
    d = plg_ana_agg[ana]
    hdr(ws10, ri, 1, '', bg=C_PLP)
    cell(ws10, ri, 2, ana, bold=True, align='left')
    cell(ws10, ri, 3, d['count'])
    cell(ws10, ri, 4, f'{d["avg"]}%')
    cell(ws10, ri, 5, f'{d["max"]}%')
    ri += 1

ri += 1
# 按品类
hdr(ws10, ri, 1, '按品类', bg=C_PLP, align='left')
hdr(ws10, ri, 2, '', bg=C_PLP)
hdr(ws10, ri, 3, '有效SKU数', bg=C_PLP)
hdr(ws10, ri, 4, '平均费率', bg=C_PLP)
hdr(ws10, ri, 5, '最高费率', bg=C_PLP)
ri += 1

for cat in sorted(plg_cat_agg.keys()):
    d = plg_cat_agg[cat]
    hdr(ws10, ri, 1, '', bg=C_PLP)
    cell(ws10, ri, 2, cat, bold=True, align='left')
    cell(ws10, ri, 3, d['count'])
    cell(ws10, ri, 4, f'{d["avg"]}%')
    cell(ws10, ri, 5, f'{d["max"]}%')
    ri += 1

add_border(ws10, 1, ri-1, 1, 5)
ws10.column_dimensions['A'].width = 3
ws10.column_dimensions['B'].width = 16
for c in 'CDE':
    ws10.column_dimensions[c].width = 14

# ── 保存 ──
OUT = r'c:\Users\Hardy\ai-projects\新品复盘\新品周报数据表_4.30-5.6.xlsx'
wb.save(OUT)
print(f'\n✅ XLSX已保存: {OUT}')
print(f'Sheets: {wb.sheetnames}')
