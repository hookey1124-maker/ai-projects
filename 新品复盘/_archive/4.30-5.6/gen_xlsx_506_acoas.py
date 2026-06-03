"""
生成4.30-5.6周期新品周报数据汇总（含PLP ACOAS）
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict, Counter
from datetime import date, datetime

# ===== 配置 =====
WORKDIR = 'c:/Users/Hardy/ai-projects/新品复盘/'
SOURCE_FILE = 'c:/Users/Hardy/ai-projects/三部周报v1/周报/新品检查周源数据和PLP数据.xlsx'
OUTPUT_FILE = WORKDIR + '新品周报数据_4.30-5.6_含ACOAS.xlsx'

# 列索引（基于四三数据累计 sheet）
C = {
    'sale_no': 0, 'sku': 1, 'list_date': 2, 'first_order': 3,
    'analyst': 4, 'category': 5, 'expand_type': 6,
    'sales_prev': 14, 'sales_curr': 15,
    'rev_prev': 24, 'rev_curr': 25,
    'rival_prev': 34, 'rival_curr': 35,
    'share_prev': 43, 'share_curr': 44,
    'ord8_prev': 61, 'ord8_curr': 62,
    'freq7_prev': 70, 'freq7_curr': 71,
    'nfreq7_prev': 79, 'nfreq7_curr': 80,
    'mkt_prev': 89, 'mkt_curr': 90,
    'op_curr': 99,
    'plp_curr': 103, 'plg_curr': 106,
}

# PLP明细列索引
PC = {
    'period': 0, 'campaign': 1, 'sku': 2, 'list_date': 6,
    'analyst': 8, 'category': 9, 'expand_type': 10,
    'impr': 11, 'click': 12, 'sold': 13, 'cost': 14, 'ad_rev': 15,
}

# ===== 辅助函数 =====
def safe_float(v, default=0):
    try:
        return float(v) if v else default
    except:
        return default

def get_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

def num(v, default=0):
    return safe_float(v, default)

def pct(a, b):
    if not b: return '0%'
    return f"{round(a/b*100, 1)}%"

def ratio(a, b):
    if not b: return '-'
    return f"{round((a-b)/abs(b)*100, 1)}%"

def hdr_style(ws, r, c, val, bg='4472C4', fc='FFFFFF', bold=True, align='center'):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.font = Font(bold=bold, color=fc, name='微软雅黑', size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    return cell

def data_cell(ws, r, c, val, bg='FFFFFF', bold=False, align='center', fc='000000'):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.font = Font(bold=bold, color=fc, name='微软雅黑', size=9)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    return cell

def set_border(ws, start_r, end_r, start_c, end_c):
    thin = Side(style='thin')
    for r in range(start_r, end_r+1):
        for c in range(start_c, end_c+1):
            ws.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ===== 常量 =====
ANALYSTS = ['俞东旭', '张潇', '朱培源', '王偲涵', '章鹏', '胡煜星']
CATEGORIES = ['车门系统', '车身外扩件', '挡泥板', '机盖及附件', '其他', '饰条', '牌照板支架']
EXPAND_TYPES = ['原开品', '拓展品', '组合件']

COLOR_HEADER = '4472C4'
COLOR_ODD = 'EEF2FF'
COLOR_EVEN = 'FFFFFF'
COLOR_TOTAL = 'E8F0FE'
COLOR_GREEN = 'C8E6C9'
COLOR_ORANGE = 'FFE0B2'
COLOR_RED = 'FFCDD2'

# ===== 读取源数据 =====
print("读取源数据...")
wb_src = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
ws_main = wb_src['四三数据累计']
ws_plp = wb_src['PLP明细']

# 读取主数据行
rows_raw = []
for row in ws_main.iter_rows(min_row=2, values_only=True):
    if row[C['sku']]:
        rows_raw.append(list(row))

# 日期过滤
cutoff_curr = date(2026, 5, 6)
cutoff_prev = date(2026, 4, 29)
rows_curr = [r for r in rows_raw if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) <= cutoff_curr]
rows_prev = [r for r in rows_raw if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) <= cutoff_prev]

print(f"当前周期SKU: {len(rows_curr)}, 上周期SKU: {len(rows_prev)}")

# ===== 构建SKU-销售额映射 =====
sku_revenue_curr = {}
sku_revenue_prev = {}
for r in rows_curr:
    sku = str(r[C['sku']]).strip()
    rev = num(r[C['rev_curr']])
    if sku and rev > 0:
        sku_revenue_curr[sku] = rev

for r in rows_prev:
    sku = str(r[C['sku']]).strip()
    rev = num(r[C['rev_curr']])  # 用上周期截止时的销售额
    if sku:
        sku_revenue_prev[sku] = rev

# ===== 读取PLP数据 =====
print("读取PLP数据...")
plp_curr_data = defaultdict(lambda: {'impr': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'count': 0})
plp_prev_data = defaultdict(lambda: {'impr': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'count': 0})

for row in ws_plp.iter_rows(min_row=2, values_only=True):
    period = str(row[PC['period']] or '').strip()
    sku = str(row[PC['sku']] or '').strip()
    if not sku or sku.startswith('广告') or sku.startswith('总数据'):
        continue

    list_d = get_date(row[PC['list_date']])
    if period == '4.30-5.6' and list_d and list_d <= cutoff_curr:
        plp_curr_data[sku]['impr'] += num(row[PC['impr']])
        plp_curr_data[sku]['click'] += num(row[PC['click']])
        plp_curr_data[sku]['sold'] += num(row[PC['sold']])
        plp_curr_data[sku]['cost'] += num(row[PC['cost']])
        plp_curr_data[sku]['ad_rev'] += num(row[PC['ad_rev']])
        plp_curr_data[sku]['count'] += 1
    elif period == '4.23-4.29' and list_d and list_d <= cutoff_prev:
        plp_prev_data[sku]['impr'] += num(row[PC['impr']])
        plp_prev_data[sku]['click'] += num(row[PC['click']])
        plp_prev_data[sku]['sold'] += num(row[PC['sold']])
        plp_prev_data[sku]['cost'] += num(row[PC['cost']])
        plp_prev_data[sku]['ad_rev'] += num(row[PC['ad_rev']])
        plp_prev_data[sku]['count'] += 1

print(f"当前周期PLP SKU: {len(plp_curr_data)}, 上周期PLP SKU: {len(plp_prev_data)}")

# ===== 创建新Workbook =====
wb = openpyxl.Workbook()

# ===== Sheet 1: 总体数据 =====
ws1 = wb.active
ws1.title = '总体数据'
ws1.sheet_properties.tabColor = '4472C4'

r = 1
hdr_style(ws1, r, 1, '新品周报汇总 - 4.30-5.6（含PLP ACOAS）', COLOR_HEADER, 'FFFFFF', True, 'left')
ws1.merge_cells(f'A{r}:H{r}')
ws1.row_dimensions[r].height = 28

r += 2
hdr_style(ws1, r, 1, '一、总体概况', COLOR_HEADER, 'FFFFFF', True)
ws1.merge_cells(f'A{r}:D{r}')

r += 1
for ci, h in enumerate(['指标', '4.30-5.6', '4.23-4.29', '环比'], 1):
    hdr_style(ws1, r, ci, h)

# 计算总体指标
total_sku = len(rows_curr)
new_list = sum(1 for r in rows_curr if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) > cutoff_prev)
total_sales_curr = sum(num(r[C['sales_curr']]) for r in rows_curr)
total_sales_prev = sum(num(r[C['sales_curr']]) for r in rows_prev)
total_rev_curr = sum(num(r[C['rev_curr']]) for r in rows_curr)
total_rev_prev = sum(num(r[C['rev_curr']]) for r in rows_prev)
has_rival_curr = sum(1 for r in rows_curr if num(r[C['rival_curr']]) > 0)
has_rival_prev = sum(1 for r in rows_prev if num(r[C['rival_curr']]) > 0)

rows_data = [
    ('累计SKU数', total_sku, len(rows_prev), ratio(total_sku, len(rows_prev))),
    ('本周新上架SKU', new_list, '-', '-'),
    ('总销量', total_sales_curr, total_sales_prev, ratio(total_sales_curr, total_sales_prev)),
    ('总销售额(USD)', total_rev_curr, total_rev_prev, ratio(total_rev_curr, total_rev_prev)),
    ('有对手SKU数', has_rival_curr, has_rival_prev, '-'),
    ('无对手SKU数', total_sku - has_rival_curr, len(rows_prev) - has_rival_prev, '-'),
]

for i, (name, v1, v2, v3) in enumerate(rows_data):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    data_cell(ws1, r, 1, name, bg, align='left')
    data_cell(ws1, r, 2, v1, bg, align='center')
    data_cell(ws1, r, 3, v2, bg, align='center')
    data_cell(ws1, r, 4, v3, bg, align='center')

# 分析及时率
r += 2
hdr_style(ws1, r, 1, '二、分析及时率', COLOR_HEADER, 'FFFFFF', True)
ws1.merge_cells(f'A{r}:D{r}')

r += 1
for ci, h in enumerate(['指标', '5.6（本周）', '4.29（上周）', '变化'], 1):
    hdr_style(ws1, r, ci, h)

# 及时率计算
def calc_timeliness(rows):
    timely = no_8d = no_7d = 0
    for r in rows:
        freq = str(r[C['freq7_curr']] or '').strip()
        nfreq = str(r[C['nfreq7_curr']] or '').strip()
        if nfreq == '异常': no_8d += 1
        elif freq == '异常': no_7d += 1
        else: timely += 1
    return timely, no_8d, no_7d

timely_curr, no_8d_curr, no_7d_curr = calc_timeliness(rows_curr)
timely_prev, no_8d_prev, no_7d_prev = calc_timeliness(rows_prev)
total_timeliness = timely_curr + no_8d_curr + no_7d_curr
rate_curr = f"{round(timely_curr/total_timeliness*100, 1)}%" if total_timeliness else '0%'

timeliness_data = [
    ('及时分析产品数', timely_curr, timely_prev, timely_curr - timely_prev),
    ('8日内新品无分析', no_8d_curr, no_8d_prev, no_8d_curr - no_8d_prev),
    ('超7日低占比未分析', no_7d_curr, no_7d_prev, no_7d_curr - no_7d_prev),
    ('统计总数', total_timeliness, timely_prev + no_8d_prev + no_7d_prev, '-'),
    ('及时分析率', rate_curr, '-', '-'),
]

for i, (name, v1, v2, v3) in enumerate(timeliness_data):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    data_cell(ws1, r, 1, name, bg, align='left')
    data_cell(ws1, r, 2, v1, bg, align='center')
    data_cell(ws1, r, 3, v2, bg, align='center')
    data_cell(ws1, r, 4, v3, bg, align='center')

# 新品出单情况
r += 2
hdr_style(ws1, r, 1, '三、新品出单情况（有对手口径）', COLOR_HEADER, 'FFFFFF', True)
ws1.merge_cells(f'A{r}:D{r}')

r += 1
for ci, h in enumerate(['指标', '5.6（本周）', '4.29（上周）', '变化'], 1):
    hdr_style(ws1, r, ci, h)

def count_ord8(rows):
    y = n = no = 0
    for r in rows:
        v = str(r[C['ord8_curr']] or '').strip()
        if v == 'Y': y += 1
        elif v == 'N': n += 1
        elif v == '未出单': no += 1
    return y, n, no

y_curr, n_curr, no_curr = count_ord8([r for r in rows_curr if num(r[C['rival_curr']]) > 0])
y_prev, n_prev, no_prev = count_ord8([r for r in rows_prev if num(r[C['rival_curr']]) > 0])
sale_rate_curr = f"{round((y_curr+n_curr)/has_rival_curr*100, 1)}%" if has_rival_curr else '0%'

ord8_data = [
    ('有对手总SKU', has_rival_curr, has_rival_prev, has_rival_curr - has_rival_prev),
    ('8日内出单（Y）', y_curr, y_prev, y_curr - y_prev),
    ('8日外出单（N）', n_curr, n_prev, n_curr - n_prev),
    ('真正未出单', no_curr, no_prev, no_curr - no_prev),
    ('已出单合计(Y+N)', y_curr+n_curr, y_prev+n_prev, (y_curr+n_curr)-(y_prev+n_prev)),
    ('出单率', sale_rate_curr, '-', '-'),
]

for i, (name, v1, v2, v3) in enumerate(ord8_data):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    data_cell(ws1, r, 1, name, bg, align='left')
    data_cell(ws1, r, 2, v1, bg, align='center')
    data_cell(ws1, r, 3, v2, bg, align='center')
    data_cell(ws1, r, 4, v3, bg, align='center')

set_border(ws1, 2, r, 1, 4)
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 12

# ===== Sheet 2: 品线维度 =====
ws2 = wb.create_sheet('品线维度')
ws2.sheet_properties.tabColor = '70AD47'

r = 1
hdr_style(ws2, r, 1, '品线维度 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws2.merge_cells(f'A{r}:J{r}')
ws1.row_dimensions[r].height = 28

r += 1
cols = ['品线', 'SKU数', '本周新上架', '本周销量', '上周销量', '销量环比', '本周销售额', '上周销售额', '销售额环比', '有对手SKU']
for ci, h in enumerate(cols, 1):
    hdr_style(ws2, r, ci, h)

cat_data_curr = defaultdict(lambda: {'sku': 0, 'new': 0, 'sales': 0, 'rev': 0, 'has_rival': 0})
cat_data_prev = defaultdict(lambda: {'sku': 0, 'sales': 0, 'rev': 0, 'has_rival': 0})

for r_data in rows_curr:
    cat = r_data[C['category']] or '其他'
    list_d = get_date(r_data[C['list_date']])
    cat_data_curr[cat]['sku'] += 1
    if list_d and list_d > cutoff_prev: cat_data_curr[cat]['new'] += 1
    cat_data_curr[cat]['sales'] += num(r_data[C['sales_curr']])
    cat_data_curr[cat]['rev'] += num(r_data[C['rev_curr']])
    if num(r_data[C['rival_curr']]) > 0: cat_data_curr[cat]['has_rival'] += 1

for r_data in rows_prev:
    cat = r_data[C['category']] or '其他'
    cat_data_prev[cat]['sku'] += 1
    cat_data_prev[cat]['sales'] += num(r_data[C['sales_curr']])
    cat_data_prev[cat]['rev'] += num(r_data[C['rev_curr']])
    if num(r_data[C['rival_curr']]) > 0: cat_data_prev[cat]['has_rival'] += 1

for i, cat in enumerate(CATEGORIES):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d_curr = cat_data_curr.get(cat, {})
    d_prev = cat_data_prev.get(cat, {})
    data_cell(ws2, r, 1, cat, bg, align='left')
    data_cell(ws2, r, 2, d_curr.get('sku', 0), bg)
    data_cell(ws2, r, 3, d_curr.get('new', 0), bg)
    data_cell(ws2, r, 4, d_curr.get('sales', 0), bg)
    data_cell(ws2, r, 5, d_prev.get('sales', 0), bg)
    data_cell(ws2, r, 6, ratio(d_curr.get('sales', 0), d_prev.get('sales', 0)), bg)
    data_cell(ws2, r, 7, d_curr.get('rev', 0), bg)
    data_cell(ws2, r, 8, d_prev.get('rev', 0), bg)
    data_cell(ws2, r, 9, ratio(d_curr.get('rev', 0), d_prev.get('rev', 0)), bg)
    data_cell(ws2, r, 10, d_curr.get('has_rival', 0), bg)

# 合计行
r += 1
data_cell(ws2, r, 1, '合计', COLOR_TOTAL, bold=True, align='left')
data_cell(ws2, r, 2, total_sku, COLOR_TOTAL, bold=True)
data_cell(ws2, r, 3, new_list, COLOR_TOTAL, bold=True)
data_cell(ws2, r, 4, total_sales_curr, COLOR_TOTAL, bold=True)
data_cell(ws2, r, 5, total_sales_prev, COLOR_TOTAL, bold=True)
data_cell(ws2, r, 6, ratio(total_sales_curr, total_sales_prev), COLOR_TOTAL, bold=True)
data_cell(ws2, r, 7, total_rev_curr, COLOR_TOTAL, bold=True)
data_cell(ws2, r, 8, total_rev_prev, COLOR_TOTAL, bold=True)
data_cell(ws2, r, 9, ratio(total_rev_curr, total_rev_prev), COLOR_TOTAL, bold=True)
data_cell(ws2, r, 10, has_rival_curr, COLOR_TOTAL, bold=True)

set_border(ws2, 2, r, 1, 10)
for ci, w in enumerate([12, 8, 10, 10, 10, 10, 12, 12, 10, 10], 1):
    ws2.column_dimensions[chr(64+ci)].width = w

# ===== Sheet 3: 分析人维度 =====
ws3 = wb.create_sheet('分析人维度')
ws3.sheet_properties.tabColor = 'ED7D31'

r = 1
hdr_style(ws3, r, 1, '分析人维度 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws3.merge_cells(f'A{r}:J{r}')

r += 1
for ci, h in enumerate(cols, 1):
    hdr_style(ws3, r, ci, h)

an_data_curr = defaultdict(lambda: {'sku': 0, 'new': 0, 'sales': 0, 'rev': 0, 'has_rival': 0})
an_data_prev = defaultdict(lambda: {'sku': 0, 'sales': 0, 'rev': 0, 'has_rival': 0})

for r_data in rows_curr:
    an = r_data[C['analyst']] or '未知'
    list_d = get_date(r_data[C['list_date']])
    an_data_curr[an]['sku'] += 1
    if list_d and list_d > cutoff_prev: an_data_curr[an]['new'] += 1
    an_data_curr[an]['sales'] += num(r_data[C['sales_curr']])
    an_data_curr[an]['rev'] += num(r_data[C['rev_curr']])
    if num(r_data[C['rival_curr']]) > 0: an_data_curr[an]['has_rival'] += 1

for r_data in rows_prev:
    an = r_data[C['analyst']] or '未知'
    an_data_prev[an]['sku'] += 1
    an_data_prev[an]['sales'] += num(r_data[C['sales_curr']])
    an_data_prev[an]['rev'] += num(r_data[C['rev_curr']])
    if num(r_data[C['rival_curr']]) > 0: an_data_prev[an]['has_rival'] += 1

for i, an in enumerate(ANALYSTS):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d_curr = an_data_curr.get(an, {})
    d_prev = an_data_prev.get(an, {})
    data_cell(ws3, r, 1, an, bg, align='left')
    data_cell(ws3, r, 2, d_curr.get('sku', 0), bg)
    data_cell(ws3, r, 3, d_curr.get('new', 0), bg)
    data_cell(ws3, r, 4, d_curr.get('sales', 0), bg)
    data_cell(ws3, r, 5, d_prev.get('sales', 0), bg)
    data_cell(ws3, r, 6, ratio(d_curr.get('sales', 0), d_prev.get('sales', 0)), bg)
    data_cell(ws3, r, 7, d_curr.get('rev', 0), bg)
    data_cell(ws3, r, 8, d_prev.get('rev', 0), bg)
    data_cell(ws3, r, 9, ratio(d_curr.get('rev', 0), d_prev.get('rev', 0)), bg)
    data_cell(ws3, r, 10, d_curr.get('has_rival', 0), bg)

r += 1
data_cell(ws3, r, 1, '合计', COLOR_TOTAL, bold=True, align='left')
data_cell(ws3, r, 2, total_sku, COLOR_TOTAL, bold=True)
data_cell(ws3, r, 3, new_list, COLOR_TOTAL, bold=True)
data_cell(ws3, r, 4, total_sales_curr, COLOR_TOTAL, bold=True)
data_cell(ws3, r, 5, total_sales_prev, COLOR_TOTAL, bold=True)
data_cell(ws3, r, 6, ratio(total_sales_curr, total_sales_prev), COLOR_TOTAL, bold=True)
data_cell(ws3, r, 7, total_rev_curr, COLOR_TOTAL, bold=True)
data_cell(ws3, r, 8, total_rev_prev, COLOR_TOTAL, bold=True)
data_cell(ws3, r, 9, ratio(total_rev_curr, total_rev_prev), COLOR_TOTAL, bold=True)
data_cell(ws3, r, 10, has_rival_curr, COLOR_TOTAL, bold=True)

set_border(ws3, 2, r, 1, 10)
for ci, w in enumerate([12, 8, 10, 10, 10, 10, 12, 12, 10, 10], 1):
    ws3.column_dimensions[chr(64+ci)].width = w

# ===== Sheet 4: 拓展类型 =====
ws4 = wb.create_sheet('拓展类型')
ws4.sheet_properties.tabColor = '9E67AB'

r = 1
hdr_style(ws4, r, 1, '拓展类型 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws4.merge_cells(f'A{r}:L{r}')

r += 1
cols4 = ['拓展类型', 'SKU数', '本周新上架', '本周销量', '上周销量', '销量环比', '本周销售额', '上周销售额', '有对手SKU', '有对手出单率', '出单率环比']
for ci, h in enumerate(cols4, 1):
    hdr_style(ws4, r, ci, h)

exp_data_curr = defaultdict(lambda: {'sku': 0, 'new': 0, 'sales': 0, 'rev': 0, 'has_rival': 0, 'y': 0, 'n': 0})
exp_data_prev = defaultdict(lambda: {'sku': 0, 'sales': 0, 'rev': 0, 'has_rival': 0, 'y': 0, 'n': 0})

for r_data in rows_curr:
    exp = r_data[C['expand_type']] or '其他'
    list_d = get_date(r_data[C['list_date']])
    exp_data_curr[exp]['sku'] += 1
    if list_d and list_d > cutoff_prev: exp_data_curr[exp]['new'] += 1
    exp_data_curr[exp]['sales'] += num(r_data[C['sales_curr']])
    exp_data_curr[exp]['rev'] += num(r_data[C['rev_curr']])
    if num(r_data[C['rival_curr']]) > 0:
        exp_data_curr[exp]['has_rival'] += 1
        v = str(r_data[C['ord8_curr']] or '').strip()
        if v == 'Y': exp_data_curr[exp]['y'] += 1
        elif v == 'N': exp_data_curr[exp]['n'] += 1

for r_data in rows_prev:
    exp = r_data[C['expand_type']] or '其他'
    exp_data_prev[exp]['sku'] += 1
    exp_data_prev[exp]['sales'] += num(r_data[C['sales_curr']])
    exp_data_prev[exp]['rev'] += num(r_data[C['rev_curr']])
    if num(r_data[C['rival_curr']]) > 0:
        exp_data_prev[exp]['has_rival'] += 1
        v = str(r_data[C['ord8_curr']] or '').strip()
        if v == 'Y': exp_data_prev[exp]['y'] += 1
        elif v == 'N': exp_data_prev[exp]['n'] += 1

for i, exp in enumerate(EXPAND_TYPES):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d_curr = exp_data_curr.get(exp, {})
    d_prev = exp_data_prev.get(exp, {})
    rate_curr = f"{round((d_curr.get('y',0)+d_curr.get('n',0))/d_curr.get('has_rival',1)*100, 1)}%" if d_curr.get('has_rival') else '0%'
    rate_prev = f"{round((d_prev.get('y',0)+d_prev.get('n',0))/d_prev.get('has_rival',1)*100, 1)}%" if d_prev.get('has_rival') else '0%'
    data_cell(ws4, r, 1, exp, bg, align='left')
    data_cell(ws4, r, 2, d_curr.get('sku', 0), bg)
    data_cell(ws4, r, 3, d_curr.get('new', 0), bg)
    data_cell(ws4, r, 4, d_curr.get('sales', 0), bg)
    data_cell(ws4, r, 5, d_prev.get('sales', 0), bg)
    data_cell(ws4, r, 6, ratio(d_curr.get('sales', 0), d_prev.get('sales', 0)), bg)
    data_cell(ws4, r, 7, d_curr.get('rev', 0), bg)
    data_cell(ws4, r, 8, d_prev.get('rev', 0), bg)
    data_cell(ws4, r, 9, d_curr.get('has_rival', 0), bg)
    data_cell(ws4, r, 10, rate_curr, bg)
    data_cell(ws4, r, 11, '-', bg)

r += 1
data_cell(ws4, r, 1, '合计', COLOR_TOTAL, bold=True, align='left')
data_cell(ws4, r, 2, total_sku, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 3, new_list, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 4, total_sales_curr, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 5, total_sales_prev, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 6, ratio(total_sales_curr, total_sales_prev), COLOR_TOTAL, bold=True)
data_cell(ws4, r, 7, total_rev_curr, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 8, total_rev_prev, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 9, has_rival_curr, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 10, sale_rate_curr, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 11, '-', COLOR_TOTAL, bold=True)

set_border(ws4, 2, r, 1, 11)
for ci, w in enumerate([12, 8, 10, 10, 10, 10, 12, 12, 10, 10, 10], 1):
    ws4.column_dimensions[chr(64+ci)].width = w

# ===== Sheet 5: 分析及时率 =====
ws5 = wb.create_sheet('分析及时率')
ws5.sheet_properties.tabColor = 'FFC000'

r = 1
hdr_style(ws5, r, 1, '分析及时率 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws5.merge_cells(f'A{r}:F{r}')

r += 1
cols5 = ['分析人', '及时分析产品数', '8日内新品无分析', '超7日低占比未分析', '统计总数', '及时分析率']
for ci, h in enumerate(cols5, 1):
    hdr_style(ws5, r, ci, h)

an_time_curr = defaultdict(lambda: {'timely': 0, 'no_8d': 0, 'no_7d': 0})
an_time_prev = defaultdict(lambda: {'timely': 0, 'no_8d': 0, 'no_7d': 0})

for r_data in rows_curr:
    an = r_data[C['analyst']] or '未知'
    freq = str(r_data[C['freq7_curr']] or '').strip()
    nfreq = str(r_data[C['nfreq7_curr']] or '').strip()
    if nfreq == '异常': an_time_curr[an]['no_8d'] += 1
    elif freq == '异常': an_time_curr[an]['no_7d'] += 1
    else: an_time_curr[an]['timely'] += 1

for r_data in rows_prev:
    an = r_data[C['analyst']] or '未知'
    freq = str(r_data[C['freq7_curr']] or '').strip()
    nfreq = str(r_data[C['nfreq7_curr']] or '').strip()
    if nfreq == '异常': an_time_prev[an]['no_8d'] += 1
    elif freq == '异常': an_time_prev[an]['no_7d'] += 1
    else: an_time_prev[an]['timely'] += 1

for i, an in enumerate(ANALYSTS):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = an_time_curr.get(an, {})
    total = d.get('timely', 0) + d.get('no_8d', 0) + d.get('no_7d', 0)
    rate = f"{round(d.get('timely', 0)/total*100, 1)}%" if total else '0%'
    data_cell(ws5, r, 1, an, bg, align='left')
    data_cell(ws5, r, 2, d.get('timely', 0), bg)
    data_cell(ws5, r, 3, d.get('no_8d', 0), bg)
    data_cell(ws5, r, 4, d.get('no_7d', 0), bg)
    data_cell(ws5, r, 5, total, bg)
    data_cell(ws5, r, 6, rate, bg)

r += 1
data_cell(ws5, r, 1, '合计', COLOR_TOTAL, bold=True, align='left')
data_cell(ws5, r, 2, timely_curr, COLOR_TOTAL, bold=True)
data_cell(ws5, r, 3, no_8d_curr, COLOR_TOTAL, bold=True)
data_cell(ws5, r, 4, no_7d_curr, COLOR_TOTAL, bold=True)
data_cell(ws5, r, 5, total_timeliness, COLOR_TOTAL, bold=True)
data_cell(ws5, r, 6, rate_curr, COLOR_TOTAL, bold=True)

set_border(ws5, 2, r, 1, 6)
for ci, w in enumerate([12, 15, 15, 18, 10, 12], 1):
    ws5.column_dimensions[chr(64+ci)].width = w

# ===== Sheet 6: 低占比新品 =====
ws6 = wb.create_sheet('低占比新品')
ws6.sheet_properties.tabColor = 'C55A11'

r = 1
hdr_style(ws6, r, 1, '低占比新品明细 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws6.merge_cells(f'A{r}:Y{r}')

r += 1
cols6 = ['销售编号', 'SKU', '上架日期', '分析人', '品类', '拓展类型',
          '本周销量', '销量环比', '本周销售额', '销售额环比',
          '上期末对手销量', '本期末对手销量', '对手销量环比',
          '上期末市占比', '本期末市占比', '市占比环比',
          '8日出单情况', '7日频次标签', '7日新品频次标签',
          '市场状态', '操作判断', '开启PLP', 'PLG最高费率', 'ACOAS']
for ci, h in enumerate(cols6, 1):
    hdr_style(ws6, r, ci, h)

low_share_skus = []
for r_data in rows_curr:
    share = num(r_data[C['share_curr']])
    rival = num(r_data[C['rival_curr']])
    if share < 0.75 and rival > 0:
        low_share_skus.append(r_data)

low_share_skus.sort(key=lambda x: num(x[C['share_curr']]))

for i, r_data in enumerate(low_share_skus):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    sku = str(r_data[C['sku']]).strip()
    rev_curr = num(r_data[C['rev_curr']])
    rev_prev = num(r_data[C['rev_prev']])
    share_curr = num(r_data[C['share_curr']])
    share_prev = num(r_data[C['share_prev']])
    cost = plp_curr_data.get(sku, {}).get('cost', 0)
    total_rev = sku_revenue_curr.get(sku, 0)
    acoas = cost / total_rev if total_rev > 0 else None

    data_cell(ws6, r, 1, r_data[C['sale_no']], bg)
    data_cell(ws6, r, 2, sku, bg)
    data_cell(ws6, r, 3, get_date(r_data[C['list_date']]), bg)
    data_cell(ws6, r, 4, r_data[C['analyst']], bg)
    data_cell(ws6, r, 5, r_data[C['category']], bg)
    data_cell(ws6, r, 6, r_data[C['expand_type']], bg)
    data_cell(ws6, r, 7, num(r_data[C['sales_curr']]), bg)
    data_cell(ws6, r, 8, ratio(num(r_data[C['sales_curr']]), num(r_data[C['sales_prev']])), bg)
    data_cell(ws6, r, 9, rev_curr, bg)
    data_cell(ws6, r, 10, ratio(rev_curr, rev_prev), bg)
    data_cell(ws6, r, 11, num(r_data[C['rival_prev']]), bg)
    data_cell(ws6, r, 12, num(r_data[C['rival_curr']]), bg)
    data_cell(ws6, r, 13, ratio(num(r_data[C['rival_curr']]), num(r_data[C['rival_prev']])), bg)
    data_cell(ws6, r, 14, share_prev, bg)
    data_cell(ws6, r, 15, share_curr, bg)
    data_cell(ws6, r, 16, ratio(share_curr, share_prev), bg)
    data_cell(ws6, r, 17, r_data[C['ord8_curr']], bg)
    data_cell(ws6, r, 18, r_data[C['freq7_curr']], bg)
    data_cell(ws6, r, 19, r_data[C['nfreq7_curr']], bg)
    data_cell(ws6, r, 20, r_data[C['mkt_curr']], bg)
    data_cell(ws6, r, 21, r_data[C['op_curr']], bg)
    data_cell(ws6, r, 22, r_data[C['plp_curr']], bg)
    data_cell(ws6, r, 23, r_data[C['plg_curr']], bg)
    data_cell(ws6, r, 24, f"{acoas:.4f}" if acoas else 'N/A', bg, fc='FF0000' if acoas and acoas > 0.1 else '000000')

set_border(ws6, 2, r, 1, 24)
for ci in range(1, 25):
    ws6.column_dimensions[chr(64+ci) if ci <= 26 else 'A' + chr(64+ci-26)].width = 10

# ===== Sheet 7: 新品PLP（含ACOAS）=====
ws7 = wb.create_sheet('新品PLP')
ws7.sheet_properties.tabColor = '00B0F0'

r = 1
hdr_style(ws7, r, 1, '新品PLP - 4.30-5.6（含ACOAS）', COLOR_HEADER, 'FFFFFF', True, 'left')
ws7.merge_cells(f'A{r}:P{r}')

# 总数据区块
r += 2
hdr_style(ws7, r, 1, '【总数据】', '00B0F0', 'FFFFFF', True)
ws7.merge_cells(f'A{r}:P{r}')

r += 1
cols7 = ['周期', '广告活动数', '广告链接数', '曝光量', '点击量', '售出数', '广告花费', '广告销售额', '总销售额', 'ROAS', 'CVR', 'CTR', 'CPC', 'CPA', 'ACOS', 'ACOAS']
for ci, h in enumerate(cols7, 1):
    hdr_style(ws7, r, ci, h)

# 计算总数据
total_impr = sum(d['impr'] for d in plp_curr_data.values())
total_click = sum(d['click'] for d in plp_curr_data.values())
total_sold = sum(d['sold'] for d in plp_curr_data.values())
total_cost = sum(d['cost'] for d in plp_curr_data.values())
total_ad_rev = sum(d['ad_rev'] for d in plp_curr_data.values())
total_rev_plp = sum(sku_revenue_curr.values())
total_campaigns = len(plp_curr_data)

roas = total_ad_rev / total_cost if total_cost else 0
cvr = total_sold / total_click if total_click else 0
ctr = total_click / total_impr if total_impr else 0
cpc = total_cost / total_click if total_click else 0
cpa = total_cost / total_sold if total_sold else 0
acos = total_cost / total_ad_rev if total_ad_rev else 0
acoas = total_cost / total_rev_plp if total_rev_plp else 0

r += 1
data_cell(ws7, r, 1, '4.30-5.6', COLOR_ODD)
data_cell(ws7, r, 2, total_campaigns, COLOR_ODD)
data_cell(ws7, r, 3, len(plp_curr_data), COLOR_ODD)
data_cell(ws7, r, 4, total_impr, COLOR_ODD)
data_cell(ws7, r, 5, total_click, COLOR_ODD)
data_cell(ws7, r, 6, total_sold, COLOR_ODD)
data_cell(ws7, r, 7, total_cost, COLOR_ODD)
data_cell(ws7, r, 8, total_ad_rev, COLOR_ODD)
data_cell(ws7, r, 9, total_rev_plp, COLOR_ODD)
data_cell(ws7, r, 10, f"{roas:.2f}", COLOR_ODD)
data_cell(ws7, r, 11, f"{cvr:.4f}", COLOR_ODD)
data_cell(ws7, r, 12, f"{ctr:.4f}", COLOR_ODD)
data_cell(ws7, r, 13, f"{cpc:.2f}", COLOR_ODD)
data_cell(ws7, r, 14, f"{cpa:.2f}", COLOR_ODD)
data_cell(ws7, r, 15, f"{acos:.4f}", COLOR_ODD)
data_cell(ws7, r, 16, f"{acoas:.4f}", COLOR_ODD, fc='FF0000')  # ACOAS高亮

# 上周期数据
total_impr_p = sum(d['impr'] for d in plp_prev_data.values())
total_click_p = sum(d['click'] for d in plp_prev_data.values())
total_sold_p = sum(d['sold'] for d in plp_prev_data.values())
total_cost_p = sum(d['cost'] for d in plp_prev_data.values())
total_ad_rev_p = sum(d['ad_rev'] for d in plp_prev_data.values())

r += 1
data_cell(ws7, r, 1, '4.23-4.29', COLOR_EVEN)
data_cell(ws7, r, 2, len(plp_prev_data), COLOR_EVEN)
data_cell(ws7, r, 3, len(plp_prev_data), COLOR_EVEN)
data_cell(ws7, r, 4, total_impr_p, COLOR_EVEN)
data_cell(ws7, r, 5, total_click_p, COLOR_EVEN)
data_cell(ws7, r, 6, total_sold_p, COLOR_EVEN)
data_cell(ws7, r, 7, total_cost_p, COLOR_EVEN)
data_cell(ws7, r, 8, total_ad_rev_p, COLOR_EVEN)
data_cell(ws7, r, 9, '-', COLOR_EVEN)
data_cell(ws7, r, 10, '-', COLOR_EVEN)
data_cell(ws7, r, 11, '-', COLOR_EVEN)
data_cell(ws7, r, 12, '-', COLOR_EVEN)
data_cell(ws7, r, 13, '-', COLOR_EVEN)
data_cell(ws7, r, 14, '-', COLOR_EVEN)
data_cell(ws7, r, 15, '-', COLOR_EVEN)
data_cell(ws7, r, 16, '-', COLOR_EVEN)

# 环比
r += 1
data_cell(ws7, r, 1, '环比', COLOR_TOTAL, bold=True)
data_cell(ws7, r, 2, len(plp_curr_data) - len(plp_prev_data), COLOR_TOTAL, bold=True)
data_cell(ws7, r, 3, '-', COLOR_TOTAL, bold=True)
data_cell(ws7, r, 4, ratio(total_impr, total_impr_p), COLOR_TOTAL, bold=True)
data_cell(ws7, r, 5, ratio(total_click, total_click_p), COLOR_TOTAL, bold=True)
data_cell(ws7, r, 6, ratio(total_sold, total_sold_p), COLOR_TOTAL, bold=True)
data_cell(ws7, r, 7, ratio(total_cost, total_cost_p), COLOR_TOTAL, bold=True)
data_cell(ws7, r, 8, ratio(total_ad_rev, total_ad_rev_p), COLOR_TOTAL, bold=True)
data_cell(ws7, r, 9, '-', COLOR_TOTAL, bold=True)
data_cell(ws7, r, 10, '-', COLOR_TOTAL, bold=True)
data_cell(ws7, r, 11, '-', COLOR_TOTAL, bold=True)
data_cell(ws7, r, 12, '-', COLOR_TOTAL, bold=True)
data_cell(ws7, r, 13, '-', COLOR_TOTAL, bold=True)
data_cell(ws7, r, 14, '-', COLOR_TOTAL, bold=True)
data_cell(ws7, r, 15, '-', COLOR_TOTAL, bold=True)
data_cell(ws7, r, 16, '-', COLOR_TOTAL, bold=True)

set_border(ws7, 3, r, 1, 16)
for ci, w in enumerate([12, 10, 10, 10, 10, 10, 10, 12, 12, 8, 8, 8, 8, 8, 8, 8], 1):
    ws7.column_dimensions[chr(64+ci)].width = w

# ===== Sheet 8: 新品出单情况 =====
ws8 = wb.create_sheet('新品出单情况')
ws8.sheet_properties.tabColor = '7030A0'

r = 1
hdr_style(ws8, r, 1, '新品出单情况 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws8.merge_cells(f'A{r}:G{r}')

r += 2
hdr_style(ws8, r, 1, '【有对手口径】', '7030A0', 'FFFFFF', True)
ws8.merge_cells(f'A{r}:G{r}')

r += 1
cols8 = ['指标', '本周(5.6)', '上周(4.29)', '变化', '', '', '']
for ci, h in enumerate(cols8, 1):
    hdr_style(ws8, r, ci, h)

r += 1
data_cell(ws8, r, 1, '有对手总SKU', COLOR_ODD, align='left')
data_cell(ws8, r, 2, has_rival_curr, COLOR_ODD)
data_cell(ws8, r, 3, has_rival_prev, COLOR_ODD)
data_cell(ws8, r, 4, has_rival_curr - has_rival_prev, COLOR_ODD)

r += 1
data_cell(ws8, r, 1, '8日内出单（Y）', COLOR_EVEN, align='left')
data_cell(ws8, r, 2, y_curr, COLOR_EVEN)
data_cell(ws8, r, 3, y_prev, COLOR_EVEN)
data_cell(ws8, r, 4, y_curr - y_prev, COLOR_EVEN)

r += 1
data_cell(ws8, r, 1, '8日外出单（N）', COLOR_ODD, align='left')
data_cell(ws8, r, 2, n_curr, COLOR_ODD)
data_cell(ws8, r, 3, n_prev, COLOR_ODD)
data_cell(ws8, r, 4, n_curr - n_prev, COLOR_ODD)

r += 1
data_cell(ws8, r, 1, '真正未出单', COLOR_EVEN, align='left')
data_cell(ws8, r, 2, no_curr, COLOR_EVEN)
data_cell(ws8, r, 3, no_prev, COLOR_EVEN)
data_cell(ws8, r, 4, no_curr - no_prev, COLOR_EVEN)

r += 1
data_cell(ws8, r, 1, '已出单合计(Y+N)', COLOR_ODD, align='left')
data_cell(ws8, r, 2, y_curr + n_curr, COLOR_ODD)
data_cell(ws8, r, 3, y_prev + n_prev, COLOR_ODD)
data_cell(ws8, r, 4, (y_curr + n_curr) - (y_prev + n_prev), COLOR_ODD)

r += 1
data_cell(ws8, r, 1, '出单率', COLOR_TOTAL, bold=True, align='left')
data_cell(ws8, r, 2, sale_rate_curr, COLOR_TOTAL, bold=True)
data_cell(ws8, r, 3, '-', COLOR_TOTAL, bold=True)
data_cell(ws8, r, 4, '-', COLOR_TOTAL, bold=True)

set_border(ws8, 3, r, 1, 4)
ws8.column_dimensions['A'].width = 18
for c in ['B', 'C', 'D']:
    ws8.column_dimensions[c].width = 12

# ===== Sheet 9: 新品未出单原因（双板块）=====
ws9 = wb.create_sheet('新品未出单原因')
ws9.sheet_properties.tabColor = 'FFA726'

r = 1
hdr_style(ws9, r, 1, '新品未出单原因分析 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:H{r}')
ws9.row_dimensions[r].height = 28

r += 1
hdr_style(ws9, r, 1, '【说明】Y=8日内出单（已出单）｜N=8日外出单（已出单但较晚）｜未出单=真正从未出单', 'F3E5F5', '4A148C', False, 'left')
ws9.merge_cells(f'A{r}:F{r}')

# 分离两组数据
has_rival_no_curr = [x for x in rows_curr if str(x[C['ord8_curr']] or '').strip() == '未出单' and num(x[C['rival_curr']]) > 0]
no_rival_no_curr = [x for x in rows_curr if str(x[C['ord8_curr']] or '').strip() == '未出单' and num(x[C['rival_curr']]) == 0]
has_rival_no_prev = [x for x in rows_prev if str(x[C['ord8_curr']] or '').strip() == '未出单' and num(x[C['rival_curr']]) > 0]
no_rival_no_prev = [x for x in rows_prev if str(x[C['ord8_curr']] or '').strip() == '未出单' and num(x[C['rival_curr']]) == 0]

# 市场状态列表
mkt_order_has = ['竞争无优势', '无市场', '站内无价格优势', '站外出单', '正常', '#N/A', '未知']
mkt_order_no = ['无市场', '未知', '竞争无优势', '#N/A', '其他']

# 区块一：有对手未出单新品
r += 2
sep_cell = ws9.cell(row=r, column=1, value='━' * 60)
sep_cell.fill = PatternFill('solid', fgColor='E65100')
sep_cell.font = Font(bold=True, color='FFFFFF', name='微软雅黑', size=10)
ws9.merge_cells(f'A{r}:H{r}')

r += 1
hdr_style(ws9, r, 1, f'【A. 有对手未出单新品】  本周: {len(has_rival_no_curr)}个  上周: {len(has_rival_no_prev)}个', 'BF360C', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:F{r}')
ws9.row_dimensions[r].height = 24

# A1: 未出单原因分布
r += 1
hdr_style(ws9, r, 1, '【A1】未出单原因分布', 'E65100', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:F{r}')
r += 1
for ci, h in enumerate(['市场状态', '本周SKU', '占比', '上周SKU', '上周占比', '变化'], 1):
    hdr_style(ws9, r, ci, h, 'FFCCBC')

mkt_has_curr = Counter(str(x[C['mkt_curr']] or '未知') for x in has_rival_no_curr)
mkt_has_prev = Counter(str(x[C['mkt_curr']] or '未知') for x in has_rival_no_prev)

for i, mkt in enumerate(mkt_order_has):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    cnt_c = mkt_has_curr.get(mkt, 0)
    cnt_p = mkt_has_prev.get(mkt, 0)
    tot_c = len(has_rival_no_curr) or 1
    tot_p = len(has_rival_no_prev) or 1
    data_cell(ws9, r, 1, mkt, bg, align='left')
    data_cell(ws9, r, 2, cnt_c, bg)
    data_cell(ws9, r, 3, f"{pct(cnt_c, tot_c)}%", bg)
    data_cell(ws9, r, 4, cnt_p, bg)
    data_cell(ws9, r, 5, f"{pct(cnt_p, tot_p)}%", bg)
    data_cell(ws9, r, 6, cnt_c - cnt_p, bg)
    r += 1

data_cell(ws9, r, 1, '合计', COLOR_TOTAL, bold=True, align='left')
data_cell(ws9, r, 2, len(has_rival_no_curr), COLOR_TOTAL, bold=True)
data_cell(ws9, r, 3, '100%', COLOR_TOTAL, bold=True)
data_cell(ws9, r, 4, len(has_rival_no_prev), COLOR_TOTAL, bold=True)
data_cell(ws9, r, 5, '100%', COLOR_TOTAL, bold=True)
data_cell(ws9, r, 6, len(has_rival_no_curr) - len(has_rival_no_prev), COLOR_TOTAL, bold=True)
r += 1
set_border(ws9, r - len(mkt_order_has) - 2, r - 1, 1, 6)

# A2: 按分析人维度
r += 1
hdr_style(ws9, r, 1, '【A2】未出单原因 - 按分析人维度', 'E65100', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:I{r}')
r += 1
cols_an = ['分析人'] + mkt_order_has + ['未出单总数']
for ci, h in enumerate(cols_an, 1):
    hdr_style(ws9, r, ci, h, 'FFCCBC')

an_has_data = defaultdict(lambda: {m: 0 for m in mkt_order_has})
for x in has_rival_no_curr:
    an = x[C['analyst']] or '未知'
    m = str(x[C['mkt_curr']] or '未知')
    if m in an_has_data[an]: an_has_data[an][m] += 1

for i, an in enumerate(ANALYSTS):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = an_has_data[an]
    total = sum(d.values())
    data_cell(ws9, r, 1, an, bg, align='left')
    for ci, m in enumerate(mkt_order_has, 2):
        data_cell(ws9, r, ci, d.get(m, 0), bg)
    data_cell(ws9, r, len(cols_an), total, bg)
    r += 1

total_row = ['合计'] + [sum(an_has_data[a].get(m, 0) for a in ANALYSTS) for m in mkt_order_has] + [len(has_rival_no_curr)]
for ci, v in enumerate(total_row, 1):
    data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci == 1 else 'center')
r += 1
set_border(ws9, r - len(ANALYSTS) - 2, r - 1, 1, len(cols_an))

# A3: 按品线维度
r += 1
hdr_style(ws9, r, 1, '【A3】未出单原因 - 按品线维度', 'E65100', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:I{r}')
r += 1
cols_cat = ['品线'] + mkt_order_has + ['未出单总数']
for ci, h in enumerate(cols_cat, 1):
    hdr_style(ws9, r, ci, h, 'FFCCBC')

cat_has_data = defaultdict(lambda: {m: 0 for m in mkt_order_has})
for x in has_rival_no_curr:
    cat = x[C['category']] or '未知'
    m = str(x[C['mkt_curr']] or '未知')
    if m in cat_has_data[cat]: cat_has_data[cat][m] += 1
all_cats_a = sorted(cat_has_data.keys())

for i, cat in enumerate(all_cats_a):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = cat_has_data[cat]
    total = sum(d.values())
    data_cell(ws9, r, 1, cat, bg, align='left')
    for ci, m in enumerate(mkt_order_has, 2):
        data_cell(ws9, r, ci, d.get(m, 0), bg)
    data_cell(ws9, r, len(cols_cat), total, bg)
    r += 1

total_row_a = ['合计'] + [sum(cat_has_data[c].get(m, 0) for c in all_cats_a) for m in mkt_order_has] + [len(has_rival_no_curr)]
for ci, v in enumerate(total_row_a, 1):
    data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci == 1 else 'center')
r += 1
set_border(ws9, r - len(all_cats_a) - 2, r - 1, 1, len(cols_cat))

# 区块二：无对手未出单新品
r += 2
sep_cell = ws9.cell(row=r, column=1, value='━' * 60)
sep_cell.fill = PatternFill('solid', fgColor='1B5E20')
sep_cell.font = Font(bold=True, color='FFFFFF', name='微软雅黑', size=10)
ws9.merge_cells(f'A{r}:H{r}')

r += 1
hdr_style(ws9, r, 1, f'【B. 无对手未出单新品】  本周: {len(no_rival_no_curr)}个  上周: {len(no_rival_no_prev)}个', '2E7D32', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:F{r}')
ws9.row_dimensions[r].height = 24

# B1: 未出单原因分布
r += 1
hdr_style(ws9, r, 1, '【B1】未出单原因分布', '388E3C', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:F{r}')
r += 1
for ci, h in enumerate(['市场状态', '本周SKU', '占比', '上周SKU', '上周占比', '变化'], 1):
    hdr_style(ws9, r, ci, h, 'C8E6C9')

mkt_no_curr = Counter(str(x[C['mkt_curr']] or '未知') for x in no_rival_no_curr)
mkt_no_prev = Counter(str(x[C['mkt_curr']] or '未知') for x in no_rival_no_prev)

for i, mkt in enumerate(mkt_order_no):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    cnt_c = mkt_no_curr.get(mkt, 0)
    cnt_p = mkt_no_prev.get(mkt, 0)
    tot_c = len(no_rival_no_curr) or 1
    tot_p = len(no_rival_no_prev) or 1
    data_cell(ws9, r, 1, mkt, bg, align='left')
    data_cell(ws9, r, 2, cnt_c, bg)
    data_cell(ws9, r, 3, f"{pct(cnt_c, tot_c)}%", bg)
    data_cell(ws9, r, 4, cnt_p, bg)
    data_cell(ws9, r, 5, f"{pct(cnt_p, tot_p)}%", bg)
    data_cell(ws9, r, 6, cnt_c - cnt_p, bg)
    r += 1

data_cell(ws9, r, 1, '合计', COLOR_TOTAL, bold=True, align='left')
data_cell(ws9, r, 2, len(no_rival_no_curr), COLOR_TOTAL, bold=True)
data_cell(ws9, r, 3, '100%', COLOR_TOTAL, bold=True)
data_cell(ws9, r, 4, len(no_rival_no_prev), COLOR_TOTAL, bold=True)
data_cell(ws9, r, 5, '100%', COLOR_TOTAL, bold=True)
data_cell(ws9, r, 6, len(no_rival_no_curr) - len(no_rival_no_prev), COLOR_TOTAL, bold=True)
r += 1
set_border(ws9, r - len(mkt_order_no) - 2, r - 1, 1, 6)

# B2: 按分析人维度
r += 1
hdr_style(ws9, r, 1, '【B2】未出单原因 - 按分析人维度', '388E3C', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:G{r}')
r += 1
cols_an_no = ['分析人'] + mkt_order_no + ['未出单总数']
for ci, h in enumerate(cols_an_no, 1):
    hdr_style(ws9, r, ci, h, 'C8E6C9')

an_no_data = defaultdict(lambda: {m: 0 for m in mkt_order_no})
for x in no_rival_no_curr:
    an = x[C['analyst']] or '未知'
    m = str(x[C['mkt_curr']] or '未知')
    if m in an_no_data[an]: an_no_data[an][m] += 1

for i, an in enumerate(ANALYSTS):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = an_no_data[an]
    total = sum(d.values())
    data_cell(ws9, r, 1, an, bg, align='left')
    for ci, m in enumerate(mkt_order_no, 2):
        data_cell(ws9, r, ci, d.get(m, 0), bg)
    data_cell(ws9, r, len(cols_an_no), total, bg)
    r += 1

total_row_no = ['合计'] + [sum(an_no_data[a].get(m, 0) for a in ANALYSTS) for m in mkt_order_no] + [len(no_rival_no_curr)]
for ci, v in enumerate(total_row_no, 1):
    data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci == 1 else 'center')
r += 1
set_border(ws9, r - len(ANALYSTS) - 2, r - 1, 1, len(cols_an_no))

# B3: 按品线维度
r += 1
hdr_style(ws9, r, 1, '【B3】未出单原因 - 按品线维度', '388E3C', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:G{r}')
r += 1
cols_cat_no = ['品线'] + mkt_order_no + ['未出单总数']
for ci, h in enumerate(cols_cat_no, 1):
    hdr_style(ws9, r, ci, h, 'C8E6C9')

cat_no_data = defaultdict(lambda: {m: 0 for m in mkt_order_no})
for x in no_rival_no_curr:
    cat = x[C['category']] or '未知'
    m = str(x[C['mkt_curr']] or '未知')
    if m in cat_no_data[cat]: cat_no_data[cat][m] += 1
all_cats_no = sorted(cat_no_data.keys())

for i, cat in enumerate(all_cats_no):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = cat_no_data[cat]
    total = sum(d.values())
    data_cell(ws9, r, 1, cat, bg, align='left')
    for ci, m in enumerate(mkt_order_no, 2):
        data_cell(ws9, r, ci, d.get(m, 0), bg)
    data_cell(ws9, r, len(cols_cat_no), total, bg)
    r += 1

total_row_no_c = ['合计'] + [sum(cat_no_data[c].get(m, 0) for c in all_cats_no) for m in mkt_order_no] + [len(no_rival_no_curr)]
for ci, v in enumerate(total_row_no_c, 1):
    data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci == 1 else 'center')

ws9.column_dimensions['A'].width = 18
for ci in range(2, 10):
    ws9.column_dimensions[chr(64+ci)].width = 12

# ===== Sheet 10: 新品PLG维度 =====
ws10 = wb.create_sheet('新品PLG维度')
ws10.sheet_properties.tabColor = '00B050'

r = 1
hdr_style(ws10, r, 1, '新品PLG维度 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws10.merge_cells(f'A{r}:E{r}')

r += 1
hdr_style(ws10, r, 1, '（PLG费率数据待补充）', 'FFF9C4', '000000', False, 'left')
ws10.merge_cells(f'A{r}:E{r}')

# ===== 保存文件 =====
print(f"\n保存文件到: {OUTPUT_FILE}")
wb.save(OUTPUT_FILE)
print("完成！")
