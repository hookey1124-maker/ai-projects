"""
新品周报 4.30-5.6 生成脚本
输出：新品周报_4.30-5.6.html（可视化报告）
"""
import openpyxl
import datetime
import json
import sys
sys.stdout.reconfigure(encoding='utf-8')

PERIOD_NEW  = '4.30-5.6'
PERIOD_PREV = '4.23-4.29'

# ── 辅助函数 ──────────────────────────────────────────────
def n(v, default=0):
    if v is None or v == '' or v == '-': return default
    if isinstance(v, str) and '#' in v: return default
    try: return float(v)
    except: return default

def pct(v):
    if v is None or v == '' or v == '-': return None
    if isinstance(v, str) and '#' in v: return None
    try:
        fv = float(v)
        return round(fv * 100, 2) if fv <= 1.0 else round(fv, 2)
    except: return None

def fmt_pct(v):
    if v is None: return '-'
    return f'{v:.1f}%'

def fmt_chg(v, is_pct=False):
    if v is None: return '-'
    sign = '+' if v > 0 else ''
    if is_pct: return f'{sign}{v:.1f}%'
    return f'{sign}{v:.0f}'

# ── 读取数据 ──────────────────────────────────────────────
wb1 = openpyxl.load_workbook(
    r'c:\Users\Hardy\ai-projects\新品复盘\新品检查周源数据和PLP数据.xlsx',
    data_only=True)

ws_main = wb1[wb1.sheetnames[1]]  # 四三数据累计
ws_plp  = wb1[wb1.sheetnames[0]]  # PLP明细

all_rows = [r for r in ws_main.iter_rows(min_row=2, values_only=True) if r[1]]
print(f'有效SKU行数: {len(all_rows)}')

# ── 列索引 (0-based, 从header验证) ───────────────────────
# 销量: col15=4.30-5.6, col14=4.23-4.29, col13=4.16-4.22, col12=4.9-4.15
# 销售额: col25=4.30-5.6, col24=4.23-4.29, col23=4.16-4.22, col22=4.9-4.15
# 对手销量: col35=5.6, col34=4.29, col33=4.22, col32=4.15
# 市占比: col44=5.6, col43=4.29, col42=4.22, col41=4.15
# 8日出单: col62=5.6, col61=4.29, col60=4.22, col59=4.15
# 市场状态: col90=5.6, col89=4.29, col88=4.22, col87=4.15
# 操作判断: col99=4.30-5.6, col98=4.23-4.29, col97=4.16-4.22, col96=4.9-4.15
# PLP: col103=4.30-5.6
# PLG最高费率: col106=4.30-5.6, col105=4.23-4.29

# ── 各周期销量/销售额list（用于趋势图） ─────────────────
period_labels = ['4.9-4.15', '4.16-4.22', '4.23-4.29', '4.30-5.6']
qty_cols = [12, 13, 14, 15]   # 销量列
amt_cols = [22, 23, 24, 25]   # 销售额列
mzb_cols = [41, 42, 43, 44]   # 市占比列
order_cols = [59, 60, 61, 62] # 8日出单列
rival_cols = [32, 33, 34, 35] # 对手销量列
mkt_cols   = [87, 88, 89, 90] # 市场状态列
op_cols    = [96, 97, 98, 99] # 操作判断列

# ── 1. KPI汇总 ────────────────────────────────────────────
qty_list = [sum(n(r[c]) for r in all_rows) for c in qty_cols]
amt_list = [round(sum(n(r[c]) for r in all_rows), 2) for c in amt_cols]

print(f'\n=== 总销量 ===')
for i, p in enumerate(period_labels):
    print(f'  {p}: {qty_list[i]:.0f}')
print(f'  环比: {qty_list[-1] - qty_list[-2]:+.0f}')

print(f'\n=== 总销售额 ===')
for i, p in enumerate(period_labels):
    print(f'  {p}: ${amt_list[i]:.2f}')
print(f'  环比: {amt_list[-1] - amt_list[-2]:+.2f}')

def order_stats(rows, order_col, rival_col):
    """返回(Y数, N数, 有对手数, Y有对手, N有对手, 未有对手, 有对手出单率)"""
    rival_rows = [r for r in rows if n(r[rival_col]) > 0]
    y = sum(1 for r in rows if r[order_col] == 'Y')
    n_cnt = sum(1 for r in rows if r[order_col] == 'N')
    yr = sum(1 for r in rival_rows if r[order_col] == 'Y')
    nr = sum(1 for r in rival_rows if r[order_col] == 'N')
    unr = sum(1 for r in rival_rows if r[order_col] not in ('Y', 'N'))
    rate = round(yr / len(rival_rows) * 100, 1) if rival_rows else 0
    return y, n_cnt, len(rival_rows), yr, nr, unr, rate

# 各周期出单统计
order_stats_list = [order_stats(all_rows, order_cols[i], rival_cols[i]) for i in range(4)]

print(f'\n=== 8日出单(有对手口径) ===')
for i, p in enumerate(period_labels):
    y, n_cnt, riv, yr, nr, unr, rate = order_stats_list[i]
    print(f'  {p}: Y={yr}, N={nr}, 未={unr}, 共{riv}SKU, 出单率={rate}%')

# 本期
y_new, n_new, riv_new, yr_new, nr_new, unr_new, rate_new = order_stats_list[-1]
y_prev, n_prev, riv_prev, yr_prev, nr_prev, unr_prev, rate_prev = order_stats_list[-2]

# 全量出单
y_all = [sum(1 for r in all_rows if r[order_cols[i]] == 'Y') for i in range(4)]
n_all = [sum(1 for r in all_rows if r[order_cols[i]] == 'N') for i in range(4)]
un_all = [sum(1 for r in all_rows if r[order_cols[i]] not in ('Y', 'N')) for i in range(4)]

print(f'\n=== 全量出单 ===')
for i, p in enumerate(period_labels):
    print(f'  {p}: Y={y_all[i]}, N={n_all[i]}, 未={un_all[i]}, 总SKU={len(all_rows)}')

# 平均市占比
def avg_mzb(rows, col):
    vals = [pct(r[col]) for r in rows if pct(r[col]) is not None]
    return round(sum(vals)/len(vals), 1) if vals else 0, len(vals)

mzb_list = [avg_mzb(all_rows, mzb_cols[i])[0] for i in range(4)]
mzb_new, mzb_cnt = avg_mzb(all_rows, mzb_cols[-1])
mzb_prev = mzb_list[-2]
print(f'\n=== 平均市占比 ===')
for i, p in enumerate(period_labels):
    print(f'  {p}: {mzb_list[i]}%')
print(f'  本期: {mzb_new}% (n={mzb_cnt}), 环比: {mzb_new-mzb_prev:+.1f}%')

# 低占比新品
low_skus = [r for r in all_rows if pct(r[mzb_cols[-1]]) is not None and pct(r[mzb_cols[-1]]) < 75 and n(r[rival_cols[-1]]) > 0]
print(f'\n低占比新品: {len(low_skus)}个')

# ── 2. 品类维度 ────────────────────────────────────────────
cats = sorted(set(r[5] for r in all_rows if r[5]))
cat_data = {}
for cat in cats:
    cr = [r for r in all_rows if r[5] == cat]
    tot = len(cr)
    qtys = [sum(n(r[qty_cols[i]]) for r in cr) for i in range(4)]
    amts = [round(sum(n(r[amt_cols[i]]) for r in cr), 2) for i in range(4)]
    ys = [sum(1 for r in cr if r[order_cols[i]] == 'Y') for i in range(4)]
    rates = [round(ys[i]/tot*100, 1) if tot else 0 for i in range(4)]
    cat_data[cat] = {'skus': tot, 'qtys': qtys, 'amts': amts, 'ys': ys, 'rates': rates}
    print(f'品类 {cat}: SKU={tot}, 销量={qtys[-1]:.0f}(环比{qtys[-1]-qtys[-2]:+.0f}), 出单率={rates[-1]}%(环比{rates[-1]-rates[-2]:+.1f}%)')

# ── 3. 分析人维度 ──────────────────────────────────────────
analysts = sorted(set(r[4] for r in all_rows if r[4]))
ana_data = {}
for ana in analysts:
    ar = [r for r in all_rows if r[4] == ana]
    tot = len(ar)
    qtys = [sum(n(r[qty_cols[i]]) for r in ar) for i in range(4)]
    ys = [sum(1 for r in ar if r[order_cols[i]] == 'Y') for i in range(4)]
    rates = [round(ys[i]/tot*100, 1) if tot else 0 for i in range(4)]
    ana_data[ana] = {'skus': tot, 'qtys': qtys, 'ys': ys, 'rates': rates}
    print(f'分析人 {ana}: SKU={tot}, 销量={qtys[-1]:.0f}(环比{qtys[-1]-qtys[-2]:+.0f}), 出单率={rates[-1]}%(环比{rates[-1]-rates[-2]:+.1f}%)')

# ── 4. 拓展类型 ────────────────────────────────────────────
exps = ['原开品', '拓展品', '组合件']
exp_data = {}
for exp in exps:
    er = [r for r in all_rows if r[6] == exp]
    tot = len(er)
    qtys = [sum(n(r[qty_cols[i]]) for r in er) for i in range(4)]
    ys = [sum(1 for r in er if r[order_cols[i]] == 'Y') for i in range(4)]
    rates = [round(ys[i]/tot*100, 1) if tot else 0 for i in range(4)]
    exp_data[exp] = {'skus': tot, 'qtys': qtys, 'ys': ys, 'rates': rates}
    print(f'拓展类型 {exp}: SKU={tot}, 销量={qtys[-1]:.0f}, 出单率={rates[-1]}%(环比{rates[-1]-rates[-2]:+.1f}%)')

# ── 5. 分析及时率 —— 从新品周报数据.xlsx读取 ─────────────────
wb2 = openpyxl.load_workbook(
    r'c:\Users\Hardy\ai-projects\新品复盘\新品周报数据.xlsx',
    data_only=True)
ws_timely = wb2['分析及时率']
timely_rows_raw = {r[0]: list(r) for r in ws_timely.iter_rows(min_row=2, values_only=True)
                   if r[0] and r[0] not in ('总数据', '当日上架未记录', '减少频次本轮不分析')}

# 列偏移: idx7=4.15(4.9-4.15), idx8=4.22(4.16-4.22)?
# 根据sheet规律：3.4计数idx1=3.12-3.18, idx2=3.19-3.25, idx3=3.26-4.1, idx4=4.2-4.8, idx5=待查, idx6=待查...
# 实际最新列应包含4.30-5.6数据
# 检查最新数据
timely_check = timely_rows_raw.get('及时分析产品', [])
print(f'\n=== 分析及时率原始数据长度: {len(timely_check)} ===')
for i, v in enumerate(timely_check[:15]):
    print(f'  idx{i}: {v}')

# 根据实际reader检查，使用最近4个周期数据
# 如果能读到4.30-5.6对应的列，使用实际值；否则用主表推算
# 主表中检查数据的逻辑：统计各SKU的频次标签
# 简化处理：从主表的7天频次标签推算出单情况

# ── 6. 新品出单情况 —— 从新品周报数据.xlsx读取 ──────────────
ws_order = wb2['新品出单情况']
order_rows_raw = {r[0]: list(r) for r in ws_order.iter_rows(min_row=2, values_only=True)
                  if r[0] and r[0] not in ('总数据', '当日上架未记录')}

# ── 7. 新品未出单 —— 从主表市场状态统计 ────────────────────
# 从主表col90=5.6市场状态 + col62=5.6 8日出单 统计
unorder_periods_data = {}
unorder_reasons = ['竞争无优势', '无市场', '站内无价格优势', '站外出单']
for i in range(4):
    dist = {}
    for r in all_rows:
        od = r[order_cols[i]]
        if od not in ('Y', 'N') and od is not None:
            ms = str(r[mkt_cols[i]]) if r[mkt_cols[i]] else '未知'
            # 归类到标准原因
            mapped = ms
            if ms not in unorder_reasons:
                if '竞争' in ms: mapped = '竞争无优势'
                elif '无市场' in ms: mapped = '无市场'
                elif '价格' in ms or '站内' in ms: mapped = '站内无价格优势'
                elif '站外' in ms: mapped = '站外出单'
                else: mapped = '竞争无优势'
            dist[mapped] = dist.get(mapped, 0) + 1
    unorder_periods_data[i] = {k: dist.get(k, 0) for k in unorder_reasons}
    print(f'{period_labels[i]} 未出单原因: {unorder_periods_data[i]}')

# ── 8. 从新品周报数据.xlsx读取已有维度数据并补齐到最新周期 ──
# 构建趋势图数据（5个周期：4.9-4.15, 4.16-4.22, 4.23-4.29, 4.30-5.6 ...）
# 但新品周报数据可能只有到4.22的数据，需要补4.23-4.29和4.30-5.6

# 分析及时率推算
# 从主表col63-71=7天频次标签推算
frequency_labels = ['异常', '正常', '无数据']
def calc_timely_from_main(rows, freq_col):
    """从主表频次标签推算分析及时率"""
    timely = 0   # 正常 = 及时分析
    over7 = 0    # 异常 = 超7日未分析
    new_no = 0   # 无数据 = 8日内新品无分析（近似）
    for r in rows:
        freq = str(r[freq_col]) if r[freq_col] else ''
        if '正常' in freq:
            timely += 1
        elif '异常' in freq:
            over7 += 1
        else:
            new_no += 1
    return timely, over7, new_no

freq_cols = [63, 64, 65, 66, 67, 68, 69, 70, 71]  # 3.11→5.6 7天频次标签
timely_periods = []
freq_idx_map = [5, 6, 7, 8]  # freq_cols[5]=4.15, [6]=4.22, [7]=4.29, [8]=5.6
for i in range(4):
    freq_i = freq_cols[freq_idx_map[i]]
    t, o, nn = calc_timely_from_main(all_rows, freq_i)
    timely_periods.append({'及时分析': t, '超7日未分析': o, '8日内新品无分析': nn, 'total': t+o+nn})
    print(f'{period_labels[i]} 分析及时率(推算): 及时={t}, 超7日={o}, 新品无分析={nn}, 合计={t+o+nn}')

# 也从新品周报数据读取已有数据做基准
timely_from_wb = {'及时分析产品': [], '超7日未分析产品': [], '8日内新品无分析': []}
for key in ['及时分析产品', '超7日未分析产品', '8日内新品无分析']:
    row_data = timely_rows_raw.get(key, [])
    # 尝试取最近4个可用周期(跳过idx0)
    for idx in [5, 6, 7, 8]:  # 4.15, 4.22, 4.29, 5.6周期
        if idx < len(row_data) and row_data[idx] is not None:
            timely_from_wb[key].append(int(n(row_data[idx])))

print(f'\n新品周报数据-分析及时率已有数据:')
for k, v in timely_from_wb.items():
    print(f'  {k}: {v}')

# ── 9. PLP 4.30-5.6 ──────────────────────────────────────
plp_new_rows = []
plp_prev_rows = []
for row in ws_plp.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    period_str = str(row[0])
    if '4.30' in period_str or '4.30-5.6' in period_str:
        plp_new_rows.append(row)
    elif '4.23' in period_str or '4.23-4.29' in period_str:
        plp_prev_rows.append(row)

print(f'\n=== PLP ===')
print(f'4.30-5.6 PLP行数: {len(plp_new_rows)}')
print(f'4.23-4.29 PLP行数: {len(plp_prev_rows)}')

def clean_plp_row(row):
    try:
        sku = row[2]     # SKU name
        analyst = row[8]
        cat = row[9]
        expn = row[10]
        imp = n(row[11])
        clk = n(row[12])
        orders = n(row[13])
        spend = n(row[14])
        sales = n(row[15])
        return sku, analyst, cat, expn, imp, clk, orders, spend, sales
    except:
        return None

plp_details = []
for row in plp_new_rows:
    if row[1] == '总数据': continue
    r = clean_plp_row(row)
    if r and r[0]: plp_details.append(r)

def agg_plp(rows, key_fn):
    agg = {}
    for r in rows:
        k = key_fn(r)
        if k not in agg:
            agg[k] = [0, 0, 0, 0, 0, set()]
        agg[k][0] += r[4]; agg[k][1] += r[5]; agg[k][2] += r[6]
        agg[k][3] += r[7]; agg[k][4] += r[8]
        agg[k][5].add(r[0])  # track unique SKU names
    result = {}
    for k, v in agg.items():
        roas = round(v[4]/v[3], 2) if v[3] > 0 else 0
        cvr  = round(v[2]/v[1]*100, 2) if v[1] > 0 else 0
        ctr  = round(v[1]/v[0]*100, 4) if v[0] > 0 else 0
        acos = round(v[3]/v[4]*100, 2) if v[4] > 0 else 0
        result[k] = {'skus': len(v[5]), 'imp': int(v[0]), 'clk': int(v[1]), 'ord': int(v[2]),
                     'spend': round(v[3], 2), 'sales': round(v[4], 2),
                     'roas': roas, 'cvr': cvr, 'ctr': ctr, 'acos': acos}
    return result

plp_by_analyst = agg_plp(plp_details, lambda r: r[1])  # key by analyst
plp_by_cat = agg_plp(plp_details, lambda r: r[2])      # key by cat
plp_by_expn = agg_plp(plp_details, lambda r: r[3])     # key by expn

# PLP总计
plp_total_row = next((r for r in plp_new_rows if r[1] == '总数据'), None)
if plp_total_row:
    plp_total = {
        'skus': int(n(plp_total_row[5] or plp_total_row[6])),
        'imp': int(n(plp_total_row[11])),
        'clk': int(n(plp_total_row[12])),
        'ord': int(n(plp_total_row[13])),
        'spend': round(n(plp_total_row[14]), 2),
        'sales': round(n(plp_total_row[15]), 2),
        'roas': round(n(plp_total_row[16]), 2),
        'cvr': round(n(plp_total_row[17])*100, 2),
        'ctr': round(n(plp_total_row[18])*100, 4),
        'acos': round(n(plp_total_row[21])*100, 2),
    }
else:
    # 手动聚合
    total_imp = sum(d['imp'] for d in plp_by_analyst.values())
    total_clk = sum(d['clk'] for d in plp_by_analyst.values())
    total_ord = sum(d['ord'] for d in plp_by_analyst.values())
    total_spend = sum(d['spend'] for d in plp_by_analyst.values())
    total_sales = sum(d['sales'] for d in plp_by_analyst.values())
    plp_total = {
        'skus': len(plp_details),
        'imp': total_imp, 'clk': total_clk, 'ord': total_ord,
        'spend': round(total_spend, 2), 'sales': round(total_sales, 2),
        'roas': round(total_sales/total_spend, 2) if total_spend > 0 else 0,
        'cvr': round(total_ord/total_clk*100, 2) if total_clk > 0 else 0,
        'ctr': round(total_clk/total_imp*100, 4) if total_imp > 0 else 0,
        'acos': round(total_spend/total_sales*100, 2) if total_sales > 0 else 0,
    }

print(f'\nPLP总计: SKU={plp_total["skus"]}, 花费=${plp_total["spend"]}, ROAS={plp_total["roas"]}, ACOS={plp_total["acos"]}%')

# ── 10. PLG费率 ───────────────────────────────────────────
plg_rows_valid = [r for r in all_rows if r[106] is not None and r[106] != '' and str(r[106]) != '-' and '#' not in str(r[106])]
plg_by_analyst_data = {}
for r in plg_rows_valid:
    ana = r[4] or '未知'
    rate_val = n(r[106])
    if ana not in plg_by_analyst_data:
        plg_by_analyst_data[ana] = []
    plg_by_analyst_data[ana].append(rate_val)

plg_by_cat_data = {}
for r in plg_rows_valid:
    cat = r[5] or '未分类'
    rate_val = n(r[106])
    if cat not in plg_by_cat_data:
        plg_by_cat_data[cat] = []
    plg_by_cat_data[cat].append(rate_val)

print(f'\nPLG费率记录: {len(plg_rows_valid)}个SKU')

wb1.close()
wb2.close()

# ── 构建JSON数据用于HTML ──────────────────────────────────
# 全量出单率趋势
rate_all = [round(y_all[i]/(y_all[i]+n_all[i])*100, 1) if (y_all[i]+n_all[i]) > 0 else 0 for i in range(4)]

# 分析及时率数据 - 使用频次标签推算
# 新品周报数据中 超7日未分析产品 和 8日内新品无分析 的row key可能不同
# 使用主表频次标签推算更可靠
timely_trend = {'及时分析': [], '超7日未分析': [], '8日内新品无分析': []}
for i in range(4):
    tp = timely_periods[i]
    timely_trend['及时分析'].append(tp['及时分析'])
    timely_trend['超7日未分析'].append(tp['超7日未分析'])
    timely_trend['8日内新品无分析'].append(tp['8日内新品无分析'])

# 新品出单情况 - 从新品周报数据读取
order_trend_y = []
order_trend_n = []
order_trend_un = []
if '8日出单' in order_rows_raw:
    row_y = order_rows_raw['8日出单']
    row_n = order_rows_raw.get('8日外出单', [])
    row_un = order_rows_raw.get('未出单', [])
    for idx in [5, 6, 7, 8]:
        if idx < len(row_y):
            order_trend_y.append(int(n(row_y[idx])))
        if idx < len(row_n):
            order_trend_n.append(int(n(row_n[idx])))
        if idx < len(row_un):
            order_trend_un.append(int(n(row_un[idx])))

# 如果新品周报数据不够4期，用主表数据补
if len(order_trend_y) < 4:
    order_trend_y = y_all[-4:]
    order_trend_n = n_all[-4:]
    order_trend_un = un_all[-4:]

# 确保都是4个元素
for lst in [order_trend_y, order_trend_n, order_trend_un]:
    while len(lst) < 4:
        lst.insert(0, 0)

js_data = {
    'periods': period_labels,
    'qty_list': [round(v) for v in qty_list],
    'amt_list': [round(v, 2) for v in amt_list],
    'mzb_list': mzb_list,
    'rate_all': rate_all,
    'order_y': order_trend_y,
    'order_n': order_trend_n,
    'order_un': order_trend_un,
    'timely_in': timely_trend['及时分析'],
    'timely_over7': timely_trend['超7日未分析'],
    'timely_new': timely_trend['8日内新品无分析'],
    'unorder_data': {r: [unorder_periods_data[i][r] for i in range(4)] for r in unorder_reasons},
    'kpi': {
        'total_skus': len(all_rows),
        'qty_new': round(qty_list[-1]),
        'qty_prev': round(qty_list[-2]),
        'qty_chg': round(qty_list[-1] - qty_list[-2]),
        'amt_new': round(amt_list[-1], 2),
        'amt_prev': round(amt_list[-2], 2),
        'amt_chg': round(amt_list[-1] - amt_list[-2], 2),
        'rate_new': rate_new,
        'rate_prev': rate_prev,
        'rate_chg': round(rate_new - rate_prev, 1),
        'y_all_new': y_all[-1],
        'n_all_new': n_all[-1],
        'un_all_new': un_all[-1],
        'y_all_prev': y_all[-2],
        'n_all_prev': n_all[-2],
        'un_all_prev': un_all[-2],
        'mzb_new': mzb_new,
        'mzb_prev': mzb_prev,
        'mzb_chg': round(mzb_new - mzb_prev, 1),
        'low_cnt': len(low_skus),
        'riv_new': riv_new,
        'riv_prev': riv_prev,
        'unr_new': unr_new,
        'unr_prev': unr_prev,
    },
    'cats': cats,
    'cat_data': {c: {
        'skus': cat_data[c]['skus'],
        'qtys': cat_data[c]['qtys'],
        'amts': cat_data[c]['amts'],       # 4个周期的销售额列表
        'rates': cat_data[c]['rates'],
    } for c in cats},
    'analysts': analysts,
    'ana_data': {a: {
        'skus': ana_data[a]['skus'],
        'qtys': ana_data[a]['qtys'],
        'rates': ana_data[a]['rates'],
    } for a in analysts},
    'exps': exps,
    'exp_data': {e: {
        'skus': exp_data[e]['skus'],
        'qtys': exp_data[e]['qtys'],
        'rates': exp_data[e]['rates'],
    } for e in exps},
    'plp': {
        'total': plp_total,
        'by_analyst': dict(sorted(plp_by_analyst.items())),
        'by_cat': dict(sorted(plp_by_cat.items())),
        'by_expn': dict(sorted(plp_by_expn.items())),
    },
    'plg': {
        'total': len(plg_rows_valid),
        'by_analyst': {k: {'count': len(v), 'avg': round(sum(v)/len(v)*100, 1), 'max': round(max(v)*100, 1)} for k, v in plg_by_analyst_data.items()},
        'by_cat': {k: {'count': len(v), 'avg': round(sum(v)/len(v)*100, 1), 'max': round(max(v)*100, 1)} for k, v in plg_by_cat_data.items()},
    },
    'low_skus_cnt': len(low_skus),
    'low_share_rows': [
        [r[0], r[1], r[5] or '', r[6] or '',
         str(r[2].date()) if isinstance(r[2], datetime.datetime) else str(r[2]),
         str(r[3].date()) if isinstance(r[3], datetime.datetime) else str(r[3]),
         r[4] or '',
         round(n(r[qty_cols[-1]])), round(n(r[rival_cols[-1]])),
         fmt_pct(pct(r[mzb_cols[-1]])),
         r[order_cols[-1]] or '-', r[mkt_cols[-1]] or '-', r[op_cols[-1]] or '-']
        for r in sorted(low_skus, key=lambda x: pct(x[mzb_cols[-1]]) or 100)
    ],
}

# 保存JSON
with open(r'c:\Users\Hardy\ai-projects\新品复盘\report_data_4_30_5_6.json', 'w', encoding='utf-8') as f:
    json.dump(js_data, f, ensure_ascii=False, indent=2, default=str)
print(f'\n✅ JSON数据已保存')

print(f'\n=== 关键KPI汇总 ===')
print(f'总SKU: {len(all_rows)}')
print(f'销量: {qty_list[-1]:.0f} (环比: {qty_list[-1]-qty_list[-2]:+.0f})')
print(f'销售额: ${amt_list[-1]:.2f} (环比: {amt_list[-1]-amt_list[-2]:+.2f})')
print(f'有对手出单率: {rate_new}% (环比: {rate_new-rate_prev:+.1f}%)')
print(f'全量Y: {y_all[-1]}, N: {n_all[-1]}, 未: {un_all[-1]}')
print(f'全量出单率: {rate_all[-1]}%')
print(f'平均市占比: {mzb_new}% (环比: {mzb_new-mzb_prev:+.1f}%)')
print(f'低占比新品: {len(low_skus)}')
print(f'PLP总花费: ${plp_total["spend"]}, ROAS: {plp_total["roas"]}, ACOS: {plp_total["acos"]}%')
