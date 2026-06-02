#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
新品周报汇总数据生成脚本 - 4.30-5.6周期
输入：新品检查周源数据和PLP数据.xlsx
输出：新品周报数据_4.30-5.6.xlsx
Sheets: 总体数据/品线维度/分析人维度/拓展类型/分析及时率/低占比新品/新品PLP/新品出单情况/新品未出单原因/新品PLG维度
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from collections import defaultdict, Counter

# ===================== 读取数据 =====================
wb_src = openpyxl.load_workbook('新品检查周源数据和PLP数据.xlsx', read_only=True, data_only=True)
ws_main = wb_src['四三数据累计']
ws_plp = wb_src['PLP明细']

# 列索引（0-based）— 4.30-5.6周期
C = {
    'sale_no': 0, 'sku': 1, 'list_date': 2, 'first_order': 3,
    'analyst': 4, 'category': 5, 'expand_type': 6,
    # 销量
    'sales_423_429': 14, 'sales_430_506': 15,
    # 销售额
    'rev_423_429': 24, 'rev_430_506': 25,
    # 对手销量
    'rival_429': 34, 'rival_506': 35,
    # 市占比
    'share_429': 43, 'share_506': 44,
    # 追踪间隔
    'iv_429': 52, 'iv_506': 53,
    # 8日出单情况
    'ord8_429': 61, 'ord8_506': 62,
    # 7日频次标签
    'freq7_429': 70, 'freq7_506': 71,
    # 7日新品频次标签
    'nfreq7_429': 79, 'nfreq7_506': 80,
    # 市场状态
    'mkt_429': 89, 'mkt_506': 90,
    # 操作判断
    'op_423_429': 98, 'op_430_506': 99,
    # PLP/PLG
    'plp_430_506': 103, 'plg_430_506': 106,
}

PC = {
    'period': 0, 'campaign': 1, 'sku': 2, 'list_date': 6,
    'analyst': 8, 'category': 9, 'expand_type': 10,
    'impr': 11, 'click': 12, 'sold': 13, 'cost': 14, 'revenue': 15,
    'roas': 16, 'cvr': 17, 'ctr': 18, 'cpc': 19, 'cpa': 20, 'acos': 21
}

def get_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

def num(v, default=0):
    if v is None: return default
    try: return float(v)
    except: return default

def pct(a, b):
    return round(a / b * 100, 1) if b else 0

def ratio(new, old):
    if old == 0: return None
    return round((new - old) / abs(old) * 100, 1)

def fmt_ratio(v):
    if v is None: return '-'
    return f"{v:+.1f}%"

def fmt_pct(v):
    """将小数转成百分比字符串，例如0.02 -> 2.0%"""
    return f"{round(v * 100, 1)}%" if v else "0%"

# 日期边界
cutoff_506 = date(2026, 5, 6)
cutoff_429 = date(2026, 4, 29)
start_430 = date(2026, 4, 30)

# 读取主数据
rows_all = []
for row in ws_main.iter_rows(min_row=2, values_only=True):
    if not row[C['sku']]: continue
    rows_all.append(row)

rows_506 = [r for r in rows_all if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) <= cutoff_506]
rows_429 = [r for r in rows_all if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) <= cutoff_429]

# SKU信息映射
sku_info = {}
for r in rows_all:
    sku_info[r[C['sku']]] = {
        'list_date': get_date(r[C['list_date']]),
        'analyst': r[C['analyst']],
        'category': r[C['category']],
        'expand_type': r[C['expand_type']],
    }

# ===================== PLP读取函数 =====================
def read_plp(period_str, cutoff_date):
    totals = {'impr': 0, 'click': 0, 'sold': 0, 'cost': 0, 'revenue': 0}
    campaigns = set()
    links_count = 0
    by_analyst = defaultdict(lambda: {'campaigns': set(), 'links': 0, 'impr': 0, 'click': 0, 'sold': 0, 'cost': 0, 'revenue': 0})
    by_cat = defaultdict(lambda: {'campaigns': set(), 'links': 0, 'impr': 0, 'click': 0, 'sold': 0, 'cost': 0, 'revenue': 0})
    by_exp = defaultdict(lambda: {'campaigns': set(), 'links': 0, 'impr': 0, 'click': 0, 'sold': 0, 'cost': 0, 'revenue': 0})

    for row in ws_plp.iter_rows(min_row=2, values_only=True):
        if str(row[PC['period']] or '') != period_str: continue
        sku = row[PC['sku']]
        if not sku or not isinstance(sku, str) or sku.startswith('广告'): continue
        analyst = row[PC['analyst']]
        cat = row[PC['category']]
        if not analyst or '总数据' in str(cat or ''): continue

        camp = row[PC['campaign']]
        list_d = get_date(row[PC['list_date']]) or sku_info.get(sku, {}).get('list_date')
        is_later = list_d and list_d > cutoff_date

        exp = row[PC['expand_type']] or sku_info.get(sku, {}).get('expand_type', '未知')

        if is_later:
            continue

        links_count += 1
        campaigns.add(camp)

        for k in ['impr', 'click', 'sold', 'cost', 'revenue']:
            totals[k] += num(row[PC[k]])

        by_analyst[analyst]['campaigns'].add(camp)
        by_analyst[analyst]['links'] += 1
        for k in ['impr', 'click', 'sold', 'cost', 'revenue']:
            by_analyst[analyst][k] += num(row[PC[k]])

        by_cat[cat]['campaigns'].add(camp)
        by_cat[cat]['links'] += 1
        for k in ['impr', 'click', 'sold', 'cost', 'revenue']:
            by_cat[cat][k] += num(row[PC[k]])

        by_exp[exp]['campaigns'].add(camp)
        by_exp[exp]['links'] += 1
        for k in ['impr', 'click', 'sold', 'cost', 'revenue']:
            by_exp[exp][k] += num(row[PC[k]])

    totals['campaigns'] = len(campaigns)
    totals['links'] = links_count

    def calc(d):
        imp, clk, sld = d['impr'], d['click'], d['sold']
        cst, rev = d['cost'], d['revenue']
        return {
            'campaigns': len(d['campaigns']),
            'links': d['links'],
            'impr': round(imp), 'click': round(clk), 'sold': round(sld),
            'cost': round(cst, 2), 'revenue': round(rev, 2),
            'roas': round(rev / cst, 2) if cst else 0,
            'cvr': round(sld / clk * 100, 2) if clk else 0,
            'ctr': round(clk / imp * 100, 2) if imp else 0,
            'cpc': round(cst / clk, 2) if clk else 0,
            'cpa': round(cst / sld, 2) if sld else 0,
            'acos': round(cst / rev * 100, 2) if rev else 0,
        }

    imp, clk, sld = totals['impr'], totals['click'], totals['sold']
    cst, rev = totals['cost'], totals['revenue']
    totals['roas'] = round(rev / cst, 2) if cst else 0
    totals['cvr'] = round(sld / clk * 100, 2) if clk else 0
    totals['ctr'] = round(clk / imp * 100, 2) if imp else 0
    totals['cpc'] = round(cst / clk, 2) if clk else 0
    totals['cpa'] = round(cst / sld, 2) if sld else 0
    totals['acos'] = round(cst / rev * 100, 2) if rev else 0
    totals['impr'] = round(totals['impr'])
    totals['click'] = round(totals['click'])
    totals['sold'] = round(totals['sold'])

    return {
        'total': totals,
        'by_analyst': {k: calc(v) for k, v in by_analyst.items()},
        'by_cat': {k: calc(v) for k, v in by_cat.items()},
        'by_exp': {k: calc(v) for k, v in by_exp.items()},
    }

plp_506 = read_plp('4.30-5.6', cutoff_506)
plp_429 = read_plp('4.23-4.29', cutoff_429)

# ===================== 样式 =====================
wb = openpyxl.Workbook()
wb.remove(wb.active)

COLOR_HEADER = '4B2E83'
COLOR_SUBHDR = '7B5EA7'
COLOR_TOTAL  = 'EDE7F6'
COLOR_ODD    = 'F8F4FF'
COLOR_EVEN   = 'FFFFFF'

# PLG高亮颜色
COLOR_HL1 = 'FFE0B2'   # 橙色：PLP=Y且PLG不为0
COLOR_HL2 = 'FCE4EC'   # 粉红：PLP=N且首次出单=未出单
COLOR_HL3 = 'E8F5E9'   # 浅绿：PLP=N且PLG不为0

def hdr_style(ws, row, col, value, color=COLOR_HEADER, font_color='FFFFFF', bold=True, align='center'):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill('solid', fgColor=color)
    cell.font = Font(bold=bold, color=font_color, name='微软雅黑', size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    return cell

def data_cell(ws, row, col, value, bg=None, bold=False, align='center', num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    cell.font = Font(bold=bold, name='微软雅黑', size=9)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def set_border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style='thin', color='CCCCCC')
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

ANALYSTS = ['俞东旭', '张潇', '朱培源', '王偲涵', '章鹏', '胡煜星']
CATEGORIES = ['车门系统', '车身外扩件', '挡泥板', '机盖及附件', '牌照板支架', '其他', '饰条']
EXP_TYPES = ['原开品', '拓展品', '组合件']

# ===================== Sheet 1: 总体数据 =====================
ws1 = wb.create_sheet('总体数据')
ws1.sheet_properties.tabColor = '4B2E83'

r = 1
hdr_style(ws1, r, 1, '新品周报汇总 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws1.merge_cells(f'A{r}:L{r}')
ws1.row_dimensions[r].height = 30

r += 1
hdr_style(ws1, r, 1, '一、总体概况', '5C3D99', 'FFFFFF', True, 'left')
ws1.merge_cells(f'A{r}:L{r}')

r += 1
for ci, h in enumerate(['指标', '4.30-5.6', '4.23-4.29', '环比'], 1):
    hdr_style(ws1, r, ci, h, COLOR_SUBHDR)

r += 1
total_sales_506 = sum(num(x[C['sales_430_506']]) for x in rows_506)
total_sales_429 = sum(num(x[C['sales_423_429']]) for x in rows_429)
total_rev_506 = sum(num(x[C['rev_430_506']]) for x in rows_506)
total_rev_429 = sum(num(x[C['rev_423_429']]) for x in rows_429)

overview_data = [
    ('累计SKU数', len(rows_506), len(rows_429), ratio(len(rows_506), len(rows_429))),
    ('本周新上架SKU',
     len([x for x in rows_506 if get_date(x[C['list_date']]) and start_430 <= get_date(x[C['list_date']]) <= cutoff_506]),
     len([x for x in rows_429 if get_date(x[C['list_date']]) and date(2026,4,23) <= get_date(x[C['list_date']]) <= cutoff_429]),
     None),
    ('总销量', int(total_sales_506), int(total_sales_429), ratio(total_sales_506, total_sales_429)),
    ('总销售额(USD)', round(total_rev_506, 2), round(total_rev_429, 2), ratio(total_rev_506, total_rev_429)),
    ('有对手SKU数',
     len([x for x in rows_506 if num(x[C['rival_506']]) > 0]),
     len([x for x in rows_429 if num(x[C['rival_429']]) > 0]), None),
    ('无对手SKU数',
     len([x for x in rows_506 if num(x[C['rival_506']]) == 0]),
     len([x for x in rows_429 if num(x[C['rival_429']]) == 0]), None),
]

for i, (label, v506, v429, rr) in enumerate(overview_data):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    data_cell(ws1, r, 1, label, bg, align='left')
    data_cell(ws1, r, 2, v506, bg)
    data_cell(ws1, r, 3, v429, bg)
    data_cell(ws1, r, 4, fmt_ratio(rr), bg)
    r += 1

set_border(ws1, 3, r - 1, 1, 4)
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 12

# 二、分析及时率
r += 1
hdr_style(ws1, r, 1, '二、分析及时率', '5C3D99', 'FFFFFF', True, 'left')
ws1.merge_cells(f'A{r}:L{r}')

r += 1
for ci, h in enumerate(['指标', '5.6（本周）', '4.29（上周）', '变化'], 1):
    hdr_style(ws1, r, ci, h, COLOR_SUBHDR)

timely_506 = sum(1 for x in rows_506 if str(x[C['freq7_506']] or '') != '异常' and str(x[C['nfreq7_506']] or '') != '异常')
new_no_506 = sum(1 for x in rows_506 if str(x[C['nfreq7_506']] or '') == '异常')
over7_506 = sum(1 for x in rows_506 if str(x[C['freq7_506']] or '') == '异常')

timely_429 = sum(1 for x in rows_429 if str(x[C['freq7_429']] or '') != '异常' and str(x[C['nfreq7_429']] or '') != '异常')
new_no_429 = sum(1 for x in rows_429 if str(x[C['nfreq7_429']] or '') == '异常')
over7_429 = sum(1 for x in rows_429 if str(x[C['freq7_429']] or '') == '异常')

timely_data = [
    ('及时分析产品数', timely_506, timely_429),
    ('8日内新品无分析', new_no_506, new_no_429),
    ('超7日低占比未分析', over7_506, over7_429),
    ('统计总数', len(rows_506), len(rows_429)),
    ('及时分析率', f"{pct(timely_506, len(rows_506))}%", f"{pct(timely_429, len(rows_429))}%"),
]

r += 1
for i, row_data in enumerate(timely_data):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    data_cell(ws1, r, 1, row_data[0], bg, align='left')
    data_cell(ws1, r, 2, row_data[1], bg)
    data_cell(ws1, r, 3, row_data[2], bg)
    if isinstance(row_data[1], int) and isinstance(row_data[2], int):
        data_cell(ws1, r, 4, row_data[1] - row_data[2], bg)
    else:
        data_cell(ws1, r, 4, '-', bg)
    r += 1

set_border(ws1, r - len(timely_data) - 1, r - 1, 1, 4)

# 三、新品出单情况
r += 1
hdr_style(ws1, r, 1, '三、新品出单情况（有对手口径）', '5C3D99', 'FFFFFF', True, 'left')
ws1.merge_cells(f'A{r}:L{r}')

r += 1
for ci, h in enumerate(['指标', '5.6（本周）', '4.29（上周）', '变化'], 1):
    hdr_style(ws1, r, ci, h, COLOR_SUBHDR)

rival_506 = [x for x in rows_506 if num(x[C['rival_506']]) > 0]
rival_429 = [x for x in rows_429 if num(x[C['rival_429']]) > 0]
ord_Y_506 = sum(1 for x in rival_506 if str(x[C['ord8_506']] or '').strip() == 'Y')
ord_N_506 = sum(1 for x in rival_506 if str(x[C['ord8_506']] or '').strip() == 'N')
ord_Y_429 = sum(1 for x in rival_429 if str(x[C['ord8_429']] or '').strip() == 'Y')
ord_N_429 = sum(1 for x in rival_429 if str(x[C['ord8_429']] or '').strip() == 'N')
ord_NO_506 = len(rival_506) - ord_Y_506 - ord_N_506
ord_NO_429 = len(rival_429) - ord_Y_429 - ord_N_429
ord_saled_506 = ord_Y_506 + ord_N_506
ord_saled_429 = ord_Y_429 + ord_N_429

order_data = [
    ('有对手总SKU', len(rival_506), len(rival_429)),
    ('8日内出单（Y）', ord_Y_506, ord_Y_429),
    ('8日外出单（N）', ord_N_506, ord_N_429),
    ('真正未出单', ord_NO_506, ord_NO_429),
    ('已出单合计(Y+N)', ord_saled_506, ord_saled_429),
    ('出单率', f"{pct(ord_saled_506, len(rival_506))}%", f"{pct(ord_saled_429, len(rival_429))}%"),
]

r += 1
for i, row_data in enumerate(order_data):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    data_cell(ws1, r, 1, row_data[0], bg, align='left')
    data_cell(ws1, r, 2, row_data[1], bg)
    data_cell(ws1, r, 3, row_data[2], bg)
    if isinstance(row_data[1], int) and isinstance(row_data[2], int):
        data_cell(ws1, r, 4, row_data[1] - row_data[2], bg)
    else:
        data_cell(ws1, r, 4, '-', bg)
    r += 1

set_border(ws1, r - len(order_data) - 1, r - 1, 1, 4)
print(f"Sheet '总体数据' 完成，共 {r} 行")

# ===================== Sheet 2: 品线维度 =====================
ws2 = wb.create_sheet('品线维度')
ws2.sheet_properties.tabColor = '7B5EA7'

r = 1
hdr_style(ws2, r, 1, '品线维度汇总 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws2.merge_cells(f'A{r}:K{r}')
ws2.row_dimensions[r].height = 28

r += 1
cols_cat = ['品线', '本周SKU', '本周新上架', '本周销量', '上周销量', '销量环比',
            '本周销售额', '上周销售额', '销售额环比', '本周有对手', '上周有对手']
for ci, h in enumerate(cols_cat, 1):
    hdr_style(ws2, r, ci, h, COLOR_SUBHDR)

cats_506 = defaultdict(lambda: {'count': 0, 'sales': 0, 'rev': 0, 'new_listed': 0, 'rival': 0})
cats_429 = defaultdict(lambda: {'count': 0, 'sales': 0, 'rev': 0, 'rival': 0})

for x in rows_506:
    cat = x[C['category']] or '未知'
    cats_506[cat]['count'] += 1
    cats_506[cat]['sales'] += num(x[C['sales_430_506']])
    cats_506[cat]['rev'] += num(x[C['rev_430_506']])
    d = get_date(x[C['list_date']])
    if d and start_430 <= d <= cutoff_506:
        cats_506[cat]['new_listed'] += 1
    if num(x[C['rival_506']]) > 0:
        cats_506[cat]['rival'] += 1

for x in rows_429:
    cat = x[C['category']] or '未知'
    cats_429[cat]['count'] += 1
    cats_429[cat]['sales'] += num(x[C['sales_423_429']])
    cats_429[cat]['rev'] += num(x[C['rev_423_429']])
    if num(x[C['rival_429']]) > 0:
        cats_429[cat]['rival'] += 1

r += 1
all_cats = sorted(set(list(cats_506.keys()) + list(cats_429.keys())))
for i, cat in enumerate(all_cats):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    c5, c4 = cats_506[cat], cats_429[cat]
    row_vals = [cat, c5['count'], c5['new_listed'],
                int(c5['sales']), int(c4['sales']), fmt_ratio(ratio(c5['sales'], c4['sales'])),
                round(c5['rev'], 2), round(c4['rev'], 2), fmt_ratio(ratio(c5['rev'], c4['rev'])),
                c5['rival'], c4['rival']]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws2, r, ci, v, bg, align='left' if ci == 1 else 'center')
    r += 1

totals_row = [
    '合计', len(rows_506),
    len([x for x in rows_506 if get_date(x[C['list_date']]) and start_430 <= get_date(x[C['list_date']]) <= cutoff_506]),
    int(sum(cats_506[c]['sales'] for c in cats_506)),
    int(sum(cats_429[c]['sales'] for c in cats_429)),
    fmt_ratio(ratio(sum(cats_506[c]['sales'] for c in cats_506), sum(cats_429[c]['sales'] for c in cats_429))),
    round(sum(cats_506[c]['rev'] for c in cats_506), 2),
    round(sum(cats_429[c]['rev'] for c in cats_429), 2),
    fmt_ratio(ratio(sum(cats_506[c]['rev'] for c in cats_506), sum(cats_429[c]['rev'] for c in cats_429))),
    sum(cats_506[c]['rival'] for c in cats_506),
    sum(cats_429[c]['rival'] for c in cats_429),
]
for ci, v in enumerate(totals_row, 1):
    data_cell(ws2, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci == 1 else 'center')

set_border(ws2, 2, r, 1, len(cols_cat))
for ci in range(1, len(cols_cat) + 1):
    ws2.column_dimensions[get_column_letter(ci)].width = [14, 9, 10, 9, 9, 8, 13, 13, 9, 10, 10][ci - 1]
print("Sheet '品线维度' 完成")

# ===================== Sheet 3: 分析人维度 =====================
ws3 = wb.create_sheet('分析人维度')
ws3.sheet_properties.tabColor = '9C6FCE'

r = 1
hdr_style(ws3, r, 1, '分析人维度汇总 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws3.merge_cells(f'A{r}:I{r}')
ws3.row_dimensions[r].height = 28

r += 1
cols_an = ['分析人', '本周SKU', '本周新上架', '本周销量', '上周销量', '销量环比',
           '本周销售额', '上周销售额', '销售额环比']
for ci, h in enumerate(cols_an, 1):
    hdr_style(ws3, r, ci, h, COLOR_SUBHDR)

ans_506 = defaultdict(lambda: {'count': 0, 'sales': 0, 'rev': 0, 'new_listed': 0})
ans_429 = defaultdict(lambda: {'count': 0, 'sales': 0, 'rev': 0})

for x in rows_506:
    an = x[C['analyst']] or '未知'
    ans_506[an]['count'] += 1
    ans_506[an]['sales'] += num(x[C['sales_430_506']])
    ans_506[an]['rev'] += num(x[C['rev_430_506']])
    d = get_date(x[C['list_date']])
    if d and start_430 <= d <= cutoff_506:
        ans_506[an]['new_listed'] += 1

for x in rows_429:
    an = x[C['analyst']] or '未知'
    ans_429[an]['count'] += 1
    ans_429[an]['sales'] += num(x[C['sales_423_429']])
    ans_429[an]['rev'] += num(x[C['rev_423_429']])

r += 1
for i, an in enumerate(ANALYSTS):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    a5, a4 = ans_506[an], ans_429[an]
    row_vals = [an, a5['count'], a5['new_listed'],
                int(a5['sales']), int(a4['sales']), fmt_ratio(ratio(a5['sales'], a4['sales'])),
                round(a5['rev'], 2), round(a4['rev'], 2), fmt_ratio(ratio(a5['rev'], a4['rev']))]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws3, r, ci, v, bg, align='left' if ci == 1 else 'center')
    r += 1

tot_row = [
    '合计', len(rows_506),
    len([x for x in rows_506 if get_date(x[C['list_date']]) and start_430 <= get_date(x[C['list_date']]) <= cutoff_506]),
    int(sum(ans_506[a]['sales'] for a in ans_506)),
    int(sum(ans_429[a]['sales'] for a in ans_429)),
    fmt_ratio(ratio(sum(ans_506[a]['sales'] for a in ans_506), sum(ans_429[a]['sales'] for a in ans_429))),
    round(sum(ans_506[a]['rev'] for a in ans_506), 2),
    round(sum(ans_429[a]['rev'] for a in ans_429), 2),
    fmt_ratio(ratio(sum(ans_506[a]['rev'] for a in ans_506), sum(ans_429[a]['rev'] for a in ans_429))),
]
for ci, v in enumerate(tot_row, 1):
    data_cell(ws3, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci == 1 else 'center')

set_border(ws3, 2, r, 1, len(cols_an))
for ci, w in enumerate([12, 9, 10, 9, 9, 8, 13, 13, 9], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w
print("Sheet '分析人维度' 完成")

# ===================== Sheet 4: 拓展类型 =====================
ws4 = wb.create_sheet('拓展类型')
ws4.sheet_properties.tabColor = 'B39DDB'

r = 1
hdr_style(ws4, r, 1, '拓展类型维度汇总 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws4.merge_cells(f'A{r}:N{r}')
ws4.row_dimensions[r].height = 28

r += 1
cols_exp = ['拓展类型', '本周SKU', '上周SKU', '本周出单', '本周出单率', '上周出单', '上周出单率', '出单率环比',
            '本周销量', '上周销量', '销量环比', '本周销售额', '上周销售额', '销售额环比']
for ci, h in enumerate(cols_exp, 1):
    hdr_style(ws4, r, ci, h, COLOR_SUBHDR)

exp_506 = defaultdict(lambda: {'count': 0, 'sales': 0, 'rev': 0, 'ord': 0})
exp_429 = defaultdict(lambda: {'count': 0, 'sales': 0, 'rev': 0, 'ord': 0})

for x in rows_506:
    et = x[C['expand_type']] or '未知'
    exp_506[et]['count'] += 1
    exp_506[et]['sales'] += num(x[C['sales_430_506']])
    exp_506[et]['rev'] += num(x[C['rev_430_506']])
    if str(x[C['ord8_506']] or '') not in ('未出单', '未上架', '', 'None'):
        exp_506[et]['ord'] += 1

for x in rows_429:
    et = x[C['expand_type']] or '未知'
    exp_429[et]['count'] += 1
    exp_429[et]['sales'] += num(x[C['sales_423_429']])
    exp_429[et]['rev'] += num(x[C['rev_423_429']])
    if str(x[C['ord8_429']] or '') not in ('未出单', '未上架', '', 'None'):
        exp_429[et]['ord'] += 1

r += 1
for i, et in enumerate(EXP_TYPES):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    e5, e4 = exp_506[et], exp_429[et]
    r5 = pct(e5['ord'], e5['count'])
    r4 = pct(e4['ord'], e4['count'])
    row_vals = [et, e5['count'], e4['count'],
                e5['ord'], f"{r5}%", e4['ord'], f"{r4}%", fmt_ratio(ratio(r5, r4)),
                int(e5['sales']), int(e4['sales']), fmt_ratio(ratio(e5['sales'], e4['sales'])),
                round(e5['rev'], 2), round(e4['rev'], 2), fmt_ratio(ratio(e5['rev'], e4['rev']))]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws4, r, ci, v, bg, align='left' if ci == 1 else 'center')
    r += 1

set_border(ws4, 2, r - 1, 1, len(cols_exp))
widths = [10, 9, 9, 9, 10, 9, 10, 10, 9, 9, 9, 13, 13, 9]
for ci, w in enumerate(widths, 1):
    ws4.column_dimensions[get_column_letter(ci)].width = w
print("Sheet '拓展类型' 完成")

# ===================== Sheet 5: 分析及时率 =====================
ws5 = wb.create_sheet('分析及时率')
ws5.sheet_properties.tabColor = 'CE93D8'

r = 1
hdr_style(ws5, r, 1, '分析及时率汇总 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws5.merge_cells(f'A{r}:J{r}')
ws5.row_dimensions[r].height = 28

r += 1
cols_t = ['分析人', '截止5.6 SKU', '及时分析', '8日内无分析', '超7日未分析', '及时率',
          '截止4.29 SKU', '上周及时分析', '上周及时率', '变化']
for ci, h in enumerate(cols_t, 1):
    hdr_style(ws5, r, ci, h, COLOR_SUBHDR)

r += 1
an_timely_506 = defaultdict(lambda: {'total': 0, 'timely': 0, 'new_no': 0, 'over7': 0})
an_timely_429 = defaultdict(lambda: {'total': 0, 'timely': 0})

for x in rows_506:
    an = x[C['analyst']] or '未知'
    an_timely_506[an]['total'] += 1
    freq7 = str(x[C['freq7_506']] or '')
    nfreq7 = str(x[C['nfreq7_506']] or '')
    if nfreq7 == '异常':
        an_timely_506[an]['new_no'] += 1
    elif freq7 == '异常':
        an_timely_506[an]['over7'] += 1
    else:
        an_timely_506[an]['timely'] += 1

for x in rows_429:
    an = x[C['analyst']] or '未知'
    an_timely_429[an]['total'] += 1
    freq7 = str(x[C['freq7_429']] or '')
    nfreq7 = str(x[C['nfreq7_429']] or '')
    if freq7 != '异常' and nfreq7 != '异常':
        an_timely_429[an]['timely'] += 1

for i, an in enumerate(ANALYSTS):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    t5 = an_timely_506[an]
    t4 = an_timely_429[an]
    rate5 = pct(t5['timely'], t5['total'])
    rate4 = pct(t4['timely'], t4['total'])
    row_vals = [an, t5['total'], t5['timely'], t5['new_no'], t5['over7'],
                f"{rate5}%", t4['total'], t4['timely'], f"{rate4}%",
                fmt_ratio(ratio(rate5, rate4))]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws5, r, ci, v, bg, align='left' if ci == 1 else 'center')
    r += 1

tot_t5 = {'total': len(rows_506), 'timely': timely_506, 'new_no': new_no_506, 'over7': over7_506}
tot_t4 = {'total': len(rows_429), 'timely': timely_429}
rate_t5 = pct(tot_t5['timely'], tot_t5['total'])
rate_t4 = pct(tot_t4['timely'], tot_t4['total'])
tot_row = ['合计', tot_t5['total'], tot_t5['timely'], tot_t5['new_no'], tot_t5['over7'],
           f"{rate_t5}%", tot_t4['total'], tot_t4['timely'], f"{rate_t4}%",
           fmt_ratio(ratio(rate_t5, rate_t4))]
for ci, v in enumerate(tot_row, 1):
    data_cell(ws5, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci == 1 else 'center')

set_border(ws5, 2, r, 1, len(cols_t))
widths_t = [12, 11, 9, 12, 12, 8, 11, 10, 9, 8]
for ci, w in enumerate(widths_t, 1):
    ws5.column_dimensions[get_column_letter(ci)].width = w
print("Sheet '分析及时率' 完成")

# ===================== Sheet 6: 低占比新品 =====================
ws6 = wb.create_sheet('低占比新品')
ws6.sheet_properties.tabColor = 'F48FB1'

r = 1
hdr_style(ws6, r, 1, '低占比新品明细 - 4.30-5.6（市占比<0.75且有对手）', COLOR_HEADER, 'FFFFFF', True, 'left')
ws6.merge_cells(f'A{r}:W{r}')
ws6.row_dimensions[r].height = 28

r += 1
cols_low = ['销售编号', 'SKU', '上架日期', '分析人', '品类', '拓展类型',
            '4.30-5.6销量', '销量环比',
            '4.30-5.6销售额', '销售额环比',
            '4.29对手销量', '5.6对手销量', '对手销量环比',
            '4.29市占比', '5.6市占比', '市占比环比',
            '5.6 8日出单', '5.6 7日频次标签',
            '4.29市场状态', '4.30-5.6操作判断', '5.6市场状态',
            '4.30-5.6开启PLP', 'PLG最高费率']
for ci, h in enumerate(cols_low, 1):
    hdr_style(ws6, r, ci, h, COLOR_SUBHDR)

low_share_rows = [x for x in rows_506 if num(x[C['rival_506']]) > 0 and num(x[C['share_506']]) < 0.75]

r += 1
for i, x in enumerate(low_share_rows):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    s506 = num(x[C['sales_430_506']])
    s429 = num(x[C['sales_423_429']])
    rev506 = num(x[C['rev_430_506']])
    rev429 = num(x[C['rev_423_429']])
    rival429 = num(x[C['rival_429']])
    rival506 = num(x[C['rival_506']])
    share429 = num(x[C['share_429']])
    share506 = num(x[C['share_506']])
    plg_val = x[C['plg_430_506']]
    plg_str = fmt_pct(float(plg_val)) if plg_val and isinstance(plg_val, (int, float)) else (str(plg_val) if plg_val else '0%')
    row_vals = [
        x[C['sale_no']], x[C['sku']],
        x[C['list_date']].strftime('%Y-%m-%d') if get_date(x[C['list_date']]) else '',
        x[C['analyst']] or '', x[C['category']] or '', x[C['expand_type']] or '',
        int(s506), fmt_ratio(ratio(s506, s429)),
        round(rev506, 2), fmt_ratio(ratio(rev506, rev429)),
        int(rival429), int(rival506), fmt_ratio(ratio(rival506, rival429)),
        f"{round(share429*100,1)}%", f"{round(share506*100,1)}%", fmt_ratio(ratio(share506, share429)),
        x[C['ord8_506']] or '', x[C['freq7_506']] or '',
        x[C['mkt_429']] or '', x[C['op_430_506']] or '', x[C['mkt_506']] or '',
        x[C['plp_430_506']] or '', plg_str,
    ]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws6, r, ci, v, bg, align='left' if ci in (2, 4, 5, 6, 19, 20, 21) else 'center')
    r += 1

set_border(ws6, 2, r - 1, 1, len(cols_low))
lw = [10, 20, 12, 8, 12, 9, 11, 8, 13, 10, 12, 12, 10, 10, 10, 10, 10, 14, 14, 18, 14, 12, 12]
for ci, w in enumerate(lw, 1):
    ws6.column_dimensions[get_column_letter(ci)].width = w
print(f"Sheet '低占比新品' 完成，共 {len(low_share_rows)} 行")

# ===================== Sheet 7: 新品PLP =====================
ws7 = wb.create_sheet('新品PLP')
ws7.sheet_properties.tabColor = 'A5D6A7'

r = 1
hdr_style(ws7, r, 1, '新品PLP汇总 - 4.30-5.6（包含周期内所有PLP活动SKU）', COLOR_HEADER, 'FFFFFF', True, 'left')
ws7.merge_cells(f'A{r}:AB{r}')
ws7.row_dimensions[r].height = 28

r += 1
cols_plp = ['维度', '4.30-5.6 广告活动', '链接', '曝光量', '点击量', '售出数',
            '花费', '销售额', 'ROAS', 'CVR%', 'CTR%', 'CPC', 'CPA', 'ACOS%',
            '上周广告活动', '上周链接', '上周曝光量', '上周点击量', '上周售出数',
            '上周花费', '上周销售额', '上周ROAS', '上周CVR%', '上周CTR%', '上周CPC', '上周CPA', '上周ACOS%']
for ci, h in enumerate(cols_plp, 1):
    hdr_style(ws7, r, ci, h, COLOR_SUBHDR)

def plp_row(ws, r_num, label, d5, d4, bg, bold=False):
    row_vals = [
        label,
        d5['campaigns'], d5['links'], d5['impr'], d5['click'], d5['sold'],
        d5['cost'], d5['revenue'], d5['roas'],
        f"{d5['cvr']}%", f"{d5['ctr']}%", d5['cpc'], d5['cpa'], f"{d5['acos']}%",
        d4['campaigns'], d4['links'], d4['impr'], d4['click'], d4['sold'],
        d4['cost'], d4['revenue'], d4['roas'],
        f"{d4['cvr']}%", f"{d4['ctr']}%", d4['cpc'], d4['cpa'], f"{d4['acos']}%",
    ]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws, r_num, ci, v, bg, bold=bold, align='left' if ci == 1 else 'center')

EMPTY_PLP = {'campaigns': 0, 'links': 0, 'impr': 0, 'click': 0, 'sold': 0, 'cost': 0, 'revenue': 0, 'roas': 0, 'cvr': 0, 'ctr': 0, 'cpc': 0, 'cpa': 0, 'acos': 0}

r += 1
hdr_style(ws7, r, 1, '【总数据】', '5C3D99', 'FFFFFF', True, 'left')
ws7.merge_cells(f'A{r}:B{r}')
r += 1
plp_row(ws7, r, '总计', plp_506['total'], plp_429['total'], COLOR_TOTAL, True)
r += 2

hdr_style(ws7, r, 1, '【分析人维度】', '5C3D99', 'FFFFFF', True, 'left')
ws7.merge_cells(f'A{r}:B{r}')
r += 1
for i, an in enumerate(ANALYSTS):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d5 = plp_506['by_analyst'].get(an, EMPTY_PLP.copy())
    d4 = plp_429['by_analyst'].get(an, EMPTY_PLP.copy())
    plp_row(ws7, r, an, d5, d4, bg)
    r += 1
r += 1

hdr_style(ws7, r, 1, '【品线维度】', '5C3D99', 'FFFFFF', True, 'left')
ws7.merge_cells(f'A{r}:B{r}')
r += 1
for i, cat in enumerate(CATEGORIES):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d5 = plp_506['by_cat'].get(cat, EMPTY_PLP.copy())
    d4 = plp_429['by_cat'].get(cat, EMPTY_PLP.copy())
    plp_row(ws7, r, cat, d5, d4, bg)
    r += 1
r += 1

hdr_style(ws7, r, 1, '【拓展类型维度】', '5C3D99', 'FFFFFF', True, 'left')
ws7.merge_cells(f'A{r}:B{r}')
r += 1
for i, et in enumerate(EXP_TYPES):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d5 = plp_506['by_exp'].get(et, EMPTY_PLP.copy())
    d4 = plp_429['by_exp'].get(et, EMPTY_PLP.copy())
    plp_row(ws7, r, et, d5, d4, bg)
    r += 1

set_border(ws7, 2, r - 1, 1, len(cols_plp))
plp_widths = [14, 12, 8, 10, 8, 8, 9, 11, 7, 7, 7, 7, 7, 8, 12, 8, 10, 8, 8, 9, 11, 7, 7, 7, 7, 7, 8]
for ci, w in enumerate(plp_widths, 1):
    ws7.column_dimensions[get_column_letter(ci)].width = w
print("Sheet '新品PLP' 完成")

# ===================== Sheet 8: 新品出单情况 =====================
ws8 = wb.create_sheet('新品出单情况')
ws8.sheet_properties.tabColor = 'FF7043'

r = 1
hdr_style(ws8, r, 1, '新品出单情况汇总 - 4.30-5.6（有对手口径）', COLOR_HEADER, 'FFFFFF', True, 'left')
ws8.merge_cells(f'A{r}:H{r}')
ws8.row_dimensions[r].height = 28

r += 1
hdr_style(ws8, r, 1, '【总体出单情况】', '5C3D99', 'FFFFFF', True, 'left')
ws8.merge_cells(f'A{r}:F{r}')

r += 1
for ci, h in enumerate(['指标', '5.6本周', '4.29上周', '变化'], 1):
    hdr_style(ws8, r, ci, h, COLOR_SUBHDR)

# 8日外/8日内
within_8d_506 = [x for x in rival_506 if get_date(x[C['list_date']]) and get_date(x[C['list_date']]) >= date(2026, 4, 29)]
outside_8d_506 = [x for x in rival_506 if get_date(x[C['list_date']]) and get_date(x[C['list_date']]) < date(2026, 4, 29)]
within_8d_429 = [x for x in rival_429 if get_date(x[C['list_date']]) and get_date(x[C['list_date']]) >= date(2026, 4, 22)]
outside_8d_429 = [x for x in rival_429 if get_date(x[C['list_date']]) and get_date(x[C['list_date']]) < date(2026, 4, 22)]

out_Y_506 = sum(1 for x in outside_8d_506 if str(x[C['ord8_506']] or '').strip() == 'Y')
out_N_506 = sum(1 for x in outside_8d_506 if str(x[C['ord8_506']] or '').strip() == 'N')
out_Y_429 = sum(1 for x in outside_8d_429 if str(x[C['ord8_429']] or '').strip() == 'Y')
out_N_429 = sum(1 for x in outside_8d_429 if str(x[C['ord8_429']] or '').strip() == 'N')
in_Y_506 = sum(1 for x in within_8d_506 if str(x[C['ord8_506']] or '').strip() == 'Y')
in_N_506 = sum(1 for x in within_8d_506 if str(x[C['ord8_506']] or '').strip() == 'N')
in_Y_429 = sum(1 for x in within_8d_429 if str(x[C['ord8_429']] or '').strip() == 'Y')
in_N_429 = sum(1 for x in within_8d_429 if str(x[C['ord8_429']] or '').strip() == 'N')

r += 1
overview_ords = [
    ('有对手总SKU', len(rival_506), len(rival_429), len(rival_506)-len(rival_429)),
    ('8日内出单（Y）', ord_Y_506, ord_Y_429, ord_Y_506-ord_Y_429),
    ('8日外出单（N）', ord_N_506, ord_N_429, ord_N_506-ord_N_429),
    ('真正未出单', ord_NO_506, ord_NO_429, ord_NO_506-ord_NO_429),
    ('已出单合计(Y+N)', ord_saled_506, ord_saled_429, ord_saled_506-ord_saled_429),
    ('出单率', f"{pct(ord_saled_506, len(rival_506))}%", f"{pct(ord_saled_429, len(rival_429))}%", '-'),
    ('-- 8日外（4.29前上架）--', None, None, None),
    ('8日外SKU', len(outside_8d_506), len(outside_8d_429), len(outside_8d_506)-len(outside_8d_429)),
    ('8日外有销量', out_Y_506, out_Y_429, out_Y_506-out_Y_429),
    ('8日外未出单', out_N_506, out_N_429, out_N_506-out_N_429),
    ('8日外出单率', f"{pct(out_Y_506, len(outside_8d_506))}%", f"{pct(out_Y_429, len(outside_8d_429))}%", '-'),
    ('-- 8日内（4.29起上架）--', None, None, None),
    ('8日内SKU', len(within_8d_506), len(within_8d_429), len(within_8d_506)-len(within_8d_429)),
    ('8日内有销量', in_Y_506, in_Y_429, in_Y_506-in_Y_429),
    ('8日内未出单', in_N_506, in_N_429, in_N_506-in_N_429),
    ('8日内出单率', f"{pct(in_Y_506, len(within_8d_506))}%", f"{pct(in_Y_429, len(within_8d_429))}%", '-'),
]
for i, row_data in enumerate(overview_ords):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    if row_data[0].startswith('--'):
        bg = 'E8E0F0'
    data_cell(ws8, r, 1, row_data[0], bg, align='left')
    for ci, v in enumerate(row_data[1:], 2):
        data_cell(ws8, r, ci, v if v is not None else '-', bg)
    r += 1

set_border(ws8, 3, r - 1, 1, 4)

# 按分析人维度
r += 1
hdr_style(ws8, r, 1, '【按分析人维度】', '5C3D99', 'FFFFFF', True, 'left')
ws8.merge_cells(f'A{r}:G{r}')
r += 1
cols_an_ord = ['分析人', '有对手SKU', '8日内(Y)', '8日外(N)', '真正未出单', '已出单', '出单率', '上周出单率', '环比']
for ci, h in enumerate(cols_an_ord, 1):
    hdr_style(ws8, r, ci, h, COLOR_SUBHDR)

r += 1
an_ord_data = defaultdict(lambda: {'total': 0, 'Y': 0, 'N': 0, 'total_429': 0, 'Y_429': 0, 'N_429': 0})
for x in rival_506:
    an = x[C['analyst']] or '未知'
    an_ord_data[an]['total'] += 1
    if str(x[C['ord8_506']] or '').strip() == 'Y': an_ord_data[an]['Y'] += 1
    elif str(x[C['ord8_506']] or '').strip() == 'N': an_ord_data[an]['N'] += 1
for x in rival_429:
    an = x[C['analyst']] or '未知'
    an_ord_data[an]['total_429'] += 1
    if str(x[C['ord8_429']] or '').strip() == 'Y': an_ord_data[an]['Y_429'] += 1
    elif str(x[C['ord8_429']] or '').strip() == 'N': an_ord_data[an]['N_429'] += 1

for i, an in enumerate(ANALYSTS):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = an_ord_data[an]
    saled = d['Y'] + d['N']
    no = d['total'] - saled
    saled_429 = d['Y_429'] + d['N_429']
    rate5 = pct(saled, d['total'])
    rate4 = pct(saled_429, d['total_429'])
    row_vals = [an, d['total'], d['Y'], d['N'], no, saled, f"{rate5}%", f"{rate4}%", fmt_ratio(ratio(rate5, rate4))]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws8, r, ci, v, bg, align='left' if ci == 1 else 'center')
    r += 1

set_border(ws8, r - len(ANALYSTS) - 1, r - 1, 1, len(cols_an_ord))

# 按品线维度
r += 1
hdr_style(ws8, r, 1, '【按品线维度】', '5C3D99', 'FFFFFF', True, 'left')
ws8.merge_cells(f'A{r}:G{r}')
r += 1
for ci, h in enumerate(cols_an_ord, 1):
    hdr_style(ws8, r, ci, h, COLOR_SUBHDR)

r += 1
cat_ord_data = defaultdict(lambda: {'total': 0, 'Y': 0, 'N': 0, 'total_429': 0, 'Y_429': 0, 'N_429': 0})
for x in rival_506:
    cat = x[C['category']] or '未知'
    cat_ord_data[cat]['total'] += 1
    if str(x[C['ord8_506']] or '').strip() == 'Y': cat_ord_data[cat]['Y'] += 1
    elif str(x[C['ord8_506']] or '').strip() == 'N': cat_ord_data[cat]['N'] += 1
for x in rival_429:
    cat = x[C['category']] or '未知'
    cat_ord_data[cat]['total_429'] += 1
    if str(x[C['ord8_429']] or '').strip() == 'Y': cat_ord_data[cat]['Y_429'] += 1
    elif str(x[C['ord8_429']] or '').strip() == 'N': cat_ord_data[cat]['N_429'] += 1

all_cats_r = sorted(cat_ord_data.keys())
for i, cat in enumerate(all_cats_r):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = cat_ord_data[cat]
    saled = d['Y'] + d['N']
    no = d['total'] - saled
    saled_429 = d['Y_429'] + d['N_429']
    rate5 = pct(saled, d['total'])
    rate4 = pct(saled_429, d['total_429'])
    row_vals = [cat, d['total'], d['Y'], d['N'], no, saled, f"{rate5}%", f"{rate4}%", fmt_ratio(ratio(rate5, rate4))]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws8, r, ci, v, bg, align='left' if ci == 1 else 'center')
    r += 1

set_border(ws8, r - len(all_cats_r) - 1, r - 1, 1, len(cols_an_ord))
for ci, w in enumerate([16, 10, 10, 10, 10, 10, 8, 10, 8], 1):
    ws8.column_dimensions[get_column_letter(ci)].width = w
print("Sheet '新品出单情况' 完成")

# ===================== Sheet 9: 新品未出单原因（分有对手/无对手）=====================
ws9 = wb.create_sheet('新品未出单原因')
ws9.sheet_properties.tabColor = 'FFA726'

r = 1
hdr_style(ws9, r, 1, '新品未出单原因分析 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:H{r}')
ws9.row_dimensions[r].height = 28

r += 1
hdr_style(ws9, r, 1, '【说明】Y=8日内出单（已出单）｜N=8日外出单（已出单但较晚）｜未出单=真正从未出单', 'F3E5F5', '4A148C', False, 'left')
ws9.merge_cells(f'A{r}:F{r}')

# ---- 分离两组数据 ----
# 有对手未出单：ord8==未出单 且 对手销量>0
has_rival_no_506 = [x for x in rows_506 if str(x[C['ord8_506']] or '').strip() == '未出单' and num(x[C['rival_506']]) > 0]
has_rival_no_429 = [x for x in rows_429 if str(x[C['ord8_429']] or '').strip() == '未出单' and num(x[C['rival_429']]) > 0]
# 无对手未出单：ord8==未出单 且 对手销量==0
no_rival_no_506 = [x for x in rows_506 if str(x[C['ord8_506']] or '').strip() == '未出单' and num(x[C['rival_506']]) == 0]
no_rival_no_429 = [x for x in rows_429 if str(x[C['ord8_429']] or '').strip() == '未出单' and num(x[C['rival_429']]) == 0]

# 市场状态列表
mkt_order = ['竞争无优势', '无市场', '站内无价格优势', '站外出单', '正常', '#N/A', '未知']

# ============================================================
# 区块一：有对手未出单新品
# ============================================================
r += 1
sep_cell = ws9.cell(row=r, column=1, value='━' * 60)
sep_cell.fill = PatternFill('solid', fgColor='E65100')
sep_cell.font = Font(bold=True, color='FFFFFF', name='微软雅黑', size=10)
sep_cell.alignment = Alignment(horizontal='left', vertical='center')
ws9.merge_cells(f'A{r}:H{r}')

r += 1
hdr_style(ws9, r, 1, f'【A. 有对手未出单新品】  本周: {len(has_rival_no_506)}个  上周: {len(has_rival_no_429)}个', 'BF360C', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:F{r}')
ws9.row_dimensions[r].height = 24

# A1: 未出单原因分布
r += 1
hdr_style(ws9, r, 1, '【A1】未出单原因分布', 'E65100', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:F{r}')
r += 1
for ci, h in enumerate(['市场状态', '本周SKU', '占比', '上周SKU', '上周占比', '变化'], 1):
    hdr_style(ws9, r, ci, h, 'FFCCBC')
mkt_has_506 = Counter(str(x[C['mkt_506']] or '未知') for x in has_rival_no_506)
mkt_has_429 = Counter(str(x[C['mkt_429']] or '未知') for x in has_rival_no_429)
r += 1
for i, mkt in enumerate(mkt_order):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    cnt506 = mkt_has_506.get(mkt, 0)
    cnt429 = mkt_has_429.get(mkt, 0)
    tot506 = len(has_rival_no_506) or 1
    tot429 = len(has_rival_no_429) or 1
    row_vals = [mkt, cnt506, f"{pct(cnt506, tot506)}%", cnt429, f"{pct(cnt429, tot429)}%", cnt506-cnt429]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws9, r, ci, v, bg, align='left' if ci == 1 else 'center')
    r += 1
row_vals = ['合计', len(has_rival_no_506), '100%', len(has_rival_no_429), '100%', len(has_rival_no_506)-len(has_rival_no_429)]
for ci, v in enumerate(row_vals, 1):
    data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci == 1 else 'center')
r += 1
set_border(ws9, r - len(mkt_order) - 2, r - 1, 1, 6)

# A2: 按分析人维度
r += 1
hdr_style(ws9, r, 1, '【A2】未出单原因 - 按分析人维度', 'E65100', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:H{r}')
r += 1
cols_an = ['分析人'] + mkt_order + ['未出单总数']
for ci, h in enumerate(cols_an, 1):
    hdr_style(ws9, r, ci, h, 'FFCCBC')
an_has_data = defaultdict(lambda: {m: 0 for m in mkt_order})
for x in has_rival_no_506:
    an = x[C['analyst']] or '未知'
    m = str(x[C['mkt_506']] or '未知')
    if m in an_has_data[an]: an_has_data[an][m] += 1
r += 1
for i, an in enumerate(ANALYSTS):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = an_has_data[an]
    total = sum(d.values())
    row_vals = [an] + [d.get(m, 0) for m in mkt_order] + [total]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws9, r, ci, v, bg, align='left' if ci == 1 else 'center')
    r += 1
total_row = ['合计'] + [sum(an_has_data[a].get(m, 0) for a in ANALYSTS) for m in mkt_order] + [sum(sum(an_has_data[a].values()) for a in ANALYSTS)]
for ci, v in enumerate(total_row, 1):
    data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci == 1 else 'center')
r += 1
set_border(ws9, r - len(ANALYSTS) - 2, r - 1, 1, len(cols_an))

# A3: 按品线维度
r += 1
hdr_style(ws9, r, 1, '【A3】未出单原因 - 按品线维度', 'E65100', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:H{r}')
r += 1
cols_cat = ['品线'] + mkt_order + ['未出单总数']
for ci, h in enumerate(cols_cat, 1):
    hdr_style(ws9, r, ci, h, 'FFCCBC')
cat_has_data = defaultdict(lambda: {m: 0 for m in mkt_order})
for x in has_rival_no_506:
    cat = x[C['category']] or '未知'
    m = str(x[C['mkt_506']] or '未知')
    if m in cat_has_data[cat]: cat_has_data[cat][m] += 1
all_cats_a = sorted(cat_has_data.keys())
r += 1
for i, cat in enumerate(all_cats_a):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = cat_has_data[cat]
    total = sum(d.values())
    row_vals = [cat] + [d.get(m, 0) for m in mkt_order] + [total]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws9, r, ci, v, bg, align='left' if ci == 1 else 'center')
    r += 1
total_row_a = ['合计'] + [sum(cat_has_data[c].get(m, 0) for c in all_cats_a) for m in mkt_order] + [len(has_rival_no_506)]
for ci, v in enumerate(total_row_a, 1):
    data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci == 1 else 'center')
r += 1
set_border(ws9, r - len(all_cats_a) - 2, r - 1, 1, len(cols_cat))

# ============================================================
# 区块二：无对手未出单新品
# ============================================================
r += 1
sep_cell2 = ws9.cell(row=r, column=1, value='━' * 60)
sep_cell2.fill = PatternFill('solid', fgColor='1B5E20')
sep_cell2.font = Font(bold=True, color='FFFFFF', name='微软雅黑', size=10)
sep_cell2.alignment = Alignment(horizontal='left', vertical='center')
ws9.merge_cells(f'A{r}:H{r}')

r += 1
hdr_style(ws9, r, 1, f'【B. 无对手未出单新品】  本周: {len(no_rival_no_506)}个  上周: {len(no_rival_no_429)}个', '1B5E20', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:F{r}')
ws9.row_dimensions[r].height = 24

# 无对手新品的市场状态原因
mkt_no_order = ['无市场', '未知', '竞争无优势', '#N/A', '其他']

# B1: 未出单原因分布
r += 1
hdr_style(ws9, r, 1, '【B1】未出单原因分布', '1B5E20', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:F{r}')
r += 1
for ci, h in enumerate(['市场状态', '本周SKU', '占比', '上周SKU', '上周占比', '变化'], 1):
    hdr_style(ws9, r, ci, h, 'C8E6C9')
mkt_no_506 = Counter(str(x[C['mkt_506']] or '未知') for x in no_rival_no_506)
mkt_no_429 = Counter(str(x[C['mkt_429']] or '未知') for x in no_rival_no_429)
r += 1
for i, mkt in enumerate(mkt_no_order):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    cnt506 = mkt_no_506.get(mkt, 0)
    cnt429 = mkt_no_429.get(mkt, 0)
    tot506 = len(no_rival_no_506) or 1
    tot429 = len(no_rival_no_429) or 1
    row_vals = [mkt, cnt506, f"{pct(cnt506, tot506)}%", cnt429, f"{pct(cnt429, tot429)}%", cnt506-cnt429]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws9, r, ci, v, bg, align='left' if ci == 1 else 'center')
    r += 1
row_vals = ['合计', len(no_rival_no_506), '100%', len(no_rival_no_429), '100%', len(no_rival_no_506)-len(no_rival_no_429)]
for ci, v in enumerate(row_vals, 1):
    data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci == 1 else 'center')
r += 1
set_border(ws9, r - len(mkt_no_order) - 2, r - 1, 1, 6)

# B2: 按分析人维度
r += 1
hdr_style(ws9, r, 1, '【B2】未出单原因 - 按分析人维度', '1B5E20', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:H{r}')
r += 1
cols_an_b = ['分析人'] + mkt_no_order + ['未出单总数']
for ci, h in enumerate(cols_an_b, 1):
    hdr_style(ws9, r, ci, h, 'C8E6C9')
an_no_data = defaultdict(lambda: {m: 0 for m in mkt_no_order})
for x in no_rival_no_506:
    an = x[C['analyst']] or '未知'
    m = str(x[C['mkt_506']] or '未知')
    if m in an_no_data[an]: an_no_data[an][m] += 1
r += 1
for i, an in enumerate(ANALYSTS):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = an_no_data[an]
    total = sum(d.values())
    row_vals = [an] + [d.get(m, 0) for m in mkt_no_order] + [total]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws9, r, ci, v, bg, align='left' if ci == 1 else 'center')
    r += 1
total_row_b = ['合计'] + [sum(an_no_data[a].get(m, 0) for a in ANALYSTS) for m in mkt_no_order] + [sum(sum(an_no_data[a].values()) for a in ANALYSTS)]
for ci, v in enumerate(total_row_b, 1):
    data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci == 1 else 'center')
r += 1
set_border(ws9, r - len(ANALYSTS) - 2, r - 1, 1, len(cols_an_b))

# B3: 按品线维度
r += 1
hdr_style(ws9, r, 1, '【B3】未出单原因 - 按品线维度', '1B5E20', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:H{r}')
r += 1
cols_cat_b = ['品线'] + mkt_no_order + ['未出单总数']
for ci, h in enumerate(cols_cat_b, 1):
    hdr_style(ws9, r, ci, h, 'C8E6C9')
cat_no_data = defaultdict(lambda: {m: 0 for m in mkt_no_order})
for x in no_rival_no_506:
    cat = x[C['category']] or '未知'
    m = str(x[C['mkt_506']] or '未知')
    if m in cat_no_data[cat]: cat_no_data[cat][m] += 1
all_cats_b = sorted(cat_no_data.keys())
r += 1
for i, cat in enumerate(all_cats_b):
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = cat_no_data[cat]
    total = sum(d.values())
    row_vals = [cat] + [d.get(m, 0) for m in mkt_no_order] + [total]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws9, r, ci, v, bg, align='left' if ci == 1 else 'center')
    r += 1
total_row_c = ['合计'] + [sum(cat_no_data[c].get(m, 0) for c in all_cats_b) for m in mkt_no_order] + [len(no_rival_no_506)]
for ci, v in enumerate(total_row_c, 1):
    data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci == 1 else 'center')
r += 1
set_border(ws9, r - len(all_cats_b) - 2, r - 1, 1, len(cols_cat_b))

# 列宽
ws9_widths = [14, 10, 8, 10, 8, 8, 8, 12]
for ci, w in enumerate(ws9_widths, 1):
    ws9.column_dimensions[get_column_letter(ci)].width = w
print("Sheet '新品未出单原因' 完成（新结构：有对手/无对手双板块）")


# ===================== Sheet 10: 新品PLG维度 =====================
ws10 = wb.create_sheet('新品PLG维度')
ws10.sheet_properties.tabColor = '26C6DA'

r = 1
hdr_style(ws10, r, 1, '新品PLG维度明细 - 4.30-5.6', COLOR_HEADER, 'FFFFFF', True, 'left')
ws10.merge_cells(f'A{r}:P{r}')
ws10.row_dimensions[r].height = 28

# 图例说明
r += 1
legend_cell1 = ws10.cell(row=r, column=1, value='● 橙色高亮：PLG开启（PLP=Y且PLG费率>0%）')
legend_cell1.fill = PatternFill('solid', fgColor=COLOR_HL1)
legend_cell1.font = Font(name='微软雅黑', size=9)
legend_cell1.alignment = Alignment(horizontal='left', vertical='center')
ws10.merge_cells(f'A{r}:E{r}')

legend_cell2 = ws10.cell(row=r, column=6, value='● 粉红高亮：PLP=N且首次出单=未出单')
legend_cell2.fill = PatternFill('solid', fgColor=COLOR_HL2)
legend_cell2.font = Font(name='微软雅黑', size=9)
legend_cell2.alignment = Alignment(horizontal='left', vertical='center')
ws10.merge_cells(f'F{r}:J{r}')

legend_cell3 = ws10.cell(row=r, column=11, value='● 绿色高亮：PLP=N且PLG费率>0%')
legend_cell3.fill = PatternFill('solid', fgColor=COLOR_HL3)
legend_cell3.font = Font(name='微软雅黑', size=9)
legend_cell3.alignment = Alignment(horizontal='left', vertical='center')
ws10.merge_cells(f'K{r}:N{r}')
ws10.row_dimensions[r].height = 20

r += 1
cols_plg = ['销售编号', 'SKU', '上架日期', '首次出单日期', '分析人', '品类', '拓展类型',
            '5.6 8日出单', '4.30-5.6销量', '4.30-5.6销售额',
            '5.6对手销量', '5.6市占比', '5.6市场状态',
            '4.30-5.6操作判断', '4.30-5.6开启PLP', 'PLG最高费率']
for ci, h in enumerate(cols_plg, 1):
    hdr_style(ws10, r, ci, h, COLOR_SUBHDR)

r += 1
for x in rows_506:
    sku = x[C['sku']]
    plp = str(x[C['plp_430_506']] or '')
    plg_raw = x[C['plg_430_506']]
    plg_val = float(plg_raw) if isinstance(plg_raw, (int, float)) else 0
    fod = str(x[C['first_order']] or '')

    # 判断高亮类型（优先级：HL1 > HL2 > HL3 > 无色）
    if plp == 'Y' and plg_val > 0:
        row_bg = COLOR_HL1
    elif plp == 'N' and fod == '未出单':
        row_bg = COLOR_HL2
    elif plp == 'N' and plg_val > 0:
        row_bg = COLOR_HL3
    else:
        row_bg = None  # 不高亮，用斑马纹

    # 如果不高亮，根据行号奇偶填背景
    if row_bg is None:
        row_bg = COLOR_ODD if (r % 2 == 0) else COLOR_EVEN

    plg_str = fmt_pct(plg_val) if plg_val > 0 else '0%'
    share506 = num(x[C['share_506']])
    fod_disp = x[C['first_order']]
    if isinstance(fod_disp, datetime): fod_disp = fod_disp.strftime('%Y-%m-%d')
    elif fod_disp is None: fod_disp = '未出单'

    row_vals = [
        x[C['sale_no']], sku,
        x[C['list_date']].strftime('%Y-%m-%d') if get_date(x[C['list_date']]) else '',
        fod_disp,
        x[C['analyst']] or '', x[C['category']] or '', x[C['expand_type']] or '',
        x[C['ord8_506']] or '',
        int(num(x[C['sales_430_506']])),
        round(num(x[C['rev_430_506']]), 2),
        int(num(x[C['rival_506']])),
        f"{round(share506 * 100, 1)}%",
        x[C['mkt_506']] or '',
        x[C['op_430_506']] or '',
        plp,
        plg_str,
    ]
    for ci, v in enumerate(row_vals, 1):
        data_cell(ws10, r, ci, v, row_bg, align='left' if ci in (1, 2, 4, 5, 6, 7, 13, 14) else 'center')
    r += 1

set_border(ws10, 3, r - 1, 1, len(cols_plg))
plg_widths = [10, 18, 11, 14, 8, 13, 9, 10, 11, 13, 12, 10, 14, 18, 12, 12]
for ci, w in enumerate(plg_widths, 1):
    ws10.column_dimensions[get_column_letter(ci)].width = w
print(f"Sheet '新品PLG维度' 完成，共 {len(rows_506)} 行数据")

# ===================== 保存 =====================
output_file = '新品周报数据_4.30-5.6.xlsx'
wb.save(output_file)
print(f"\n✅ 文件已保存：{output_file}")
print(f"包含Sheets: {[ws.title for ws in wb.worksheets]}")
