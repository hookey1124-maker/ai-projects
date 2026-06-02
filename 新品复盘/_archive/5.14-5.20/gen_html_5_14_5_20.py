"""
生成5.14-5.20周期新品周报HTML看板（7 Tab：总盘概览/市场分布/品效分析/低占比/广告追踪/四三累计/汇报输出）
"""
import json, os, sys
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook
from collections import defaultdict, Counter
from datetime import date, datetime

WORKDIR = os.path.dirname(os.path.abspath(__file__)) + '/'
SOURCE_FILE = WORKDIR + '周报/新品检查周源数据和PLP数据.xlsx'
OUTPUT_FILE = WORKDIR + '新品板块_5.14-5.20.html'

# ===== 列索引（复用 gen_xlsx 配置）=====
C = {
    'sale_no': 0, 'sku': 1, 'list_date': 2, 'first_order': 3,
    'analyst': 4, 'category': 5, 'expand_type': 6,
    'sales_curr': 17, 'sales_prev': 16,
    'rev_curr': 29, 'rev_prev': 28,
    'rival_curr': 41, 'rival_prev': 40,
    'share_curr': 52, 'share_prev': 51,
    'ord8_curr': 74, 'ord8_prev': 73,
    'freq7_curr': 85, 'freq7_prev': 84,
    'nfreq7_curr': 96, 'nfreq7_prev': 95,
    'mkt_curr': 108, 'mkt_prev': 107,
    'op_curr': 119,
    'plp_curr': 125, 'plg_curr': 130,
}
PC = {
    'period': 0, 'campaign': 1, 'sku': 2, 'id': 3, 'store': 4,
    'plp_start': 5, 'list_date': 6, 'first_order': 7, 'analyst': 8,
    'category': 9, 'expand_type': 10,
    'impr': 11, 'click': 12, 'sold': 13, 'cost': 14, 'ad_rev': 15,
    'total_rev': 16, 'roas': 17, 'cvr': 18, 'ctr': 19,
    'cpc': 20, 'cpa': 21, 'acos': 22, 'acoas': 23, 'plg_enabled': 24,
}

def safe_float(v, default=0):
    try: return float(v) if v else default
    except: return default

def get_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

def num(v, default=0):
    return safe_float(v, default)

def calc_rate(a, b):
    if not b: return 0
    return round(a/b, 4)

def ratio_str(a, b):
    if not b: return '-'
    return f"{round((a-b)/abs(b)*100, 1)}%"

def get_cat(r_data):
    c = str(r_data[5] or '').strip()
    return c if c else '未分类'

def get_an(r_data):
    a = str(r_data[4] or '').strip()
    return a if a else '未知'

def get_exp(r_data):
    e = str(r_data[6] or '').strip()
    return e if e else '其他'

ANALYSTS = ['俞东旭', '张潇', '朱培源', '王偲涵', '章鹏', '胡煜星']
CATEGORIES = ['车门系统', '车身外扩件', '挡泥板', '机盖及附件', '其他', '饰条', '牌照板支架', '未分类']
EXPAND_TYPES = ['原开品', '拓展品', '组合件']
ALL_MKT_STATUSES = ['正常', '竞争无优势', '无市场', '站外出单', '站内无价格优势', '#N/A', '未知', '其他']
PLP_CURR = '5.11-5.17'
PLP_PREV = '5.4-5.10'
cutoff_curr = date(2026, 5, 20)
cutoff_prev = date(2026, 5, 13)

# ===== 读取源数据 =====
print("读取源数据...")
wb_src = load_workbook(SOURCE_FILE, data_only=True)
ws_main = wb_src['四三数据累计']
ws_plp = wb_src['PLP明细']
ws_dept = wb_src['四三销售数据']

rows_raw = []
for row in ws_main.iter_rows(min_row=2, values_only=True):
    if row[C['sku']]:
        rows_raw.append(list(row))

rows_curr = [r for r in rows_raw if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) <= cutoff_curr]
rows_prev = [r for r in rows_raw if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) <= cutoff_prev]

total_sku = len(rows_curr)
total_sales_curr = sum(num(r[C['sales_curr']]) for r in rows_curr)
total_sales_prev = sum(num(r[C['sales_prev']]) for r in rows_prev)
total_rev_curr = sum(num(r[C['rev_curr']]) for r in rows_curr)
total_rev_prev = sum(num(r[C['rev_prev']]) for r in rows_prev)
has_rival_curr = sum(1 for r in rows_curr if num(r[C['rival_curr']]) > 0)
has_rival_prev = sum(1 for r in rows_prev if num(r[C['rival_prev']]) > 0)

dept_total_sales = 0
dept_total_revenue = 0.0
for row in ws_dept.iter_rows(min_row=2, max_row=2, values_only=True):
    dept_total_sales = int(num(row[4])) if row[4] else 0
    dept_total_revenue = num(row[5])

dept_ratio = total_rev_curr / dept_total_revenue * 100 if dept_total_revenue > 0 else 0

# ===== PLP数据 =====
plp_curr_rows = []
plp_plg_enabled = set()
for row in ws_plp.iter_rows(min_row=2, values_only=True):
    period = str(row[PC['period']] or '').strip()
    sku = str(row[PC['sku']] or '').strip()
    if not sku or sku.startswith('广告') or sku.startswith('总数据'):
        continue
    list_d = get_date(row[PC['list_date']])
    if period == PLP_CURR and list_d and list_d <= cutoff_curr:
        plp_curr_rows.append(list(row))
        if str(row[PC['plg_enabled']] or '').strip() == 'Y':
            plp_plg_enabled.add(sku)

sku_campaign_count = Counter()
for row in plp_curr_rows:
    sku = str(row[PC['sku']] or '').strip()
    if sku in plp_plg_enabled:
        sku_campaign_count[sku] += 1

def agg_plp(rows):
    by_sku = defaultdict(lambda: {'impr': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'total_rev': 0, 'campaigns': set()})
    for r in rows:
        sku = str(r[PC['sku']] or '').strip()
        camp = str(r[PC['campaign']] or '').strip()
        by_sku[sku]['impr'] += num(r[PC['impr']])
        by_sku[sku]['click'] += num(r[PC['click']])
        by_sku[sku]['sold'] += num(r[PC['sold']])
        by_sku[sku]['cost'] += num(r[PC['cost']])
        by_sku[sku]['ad_rev'] += num(r[PC['ad_rev']])
        by_sku[sku]['total_rev'] += num(r[PC['total_rev']])
        if camp: by_sku[sku]['campaigns'].add(camp)
    return by_sku

def plp_totals(by_sku):
    t = {'imp': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'total_rev': 0, 'campaigns': 0}
    all_camps = set()
    for sku, d in by_sku.items():
        t['imp'] += d['impr']; t['click'] += d['click']; t['sold'] += d['sold']
        t['cost'] += d['cost']; t['ad_rev'] += d['ad_rev']
        t['total_rev'] += d['total_rev']
        all_camps.update(d['campaigns'])
    t['campaigns'] = len(all_camps)
    t['links'] = len(by_sku)
    t['roas'] = calc_rate(t['ad_rev'], t['cost'])
    t['cvr'] = calc_rate(t['sold'], t['click'])
    t['ctr'] = calc_rate(t['click'], t['imp'])
    t['cpc'] = round(t['cost'] / t['click'], 2) if t['click'] else 0
    t['cpa'] = round(t['cost'] / t['sold'], 2) if t['sold'] else 0
    t['acos'] = calc_rate(t['cost'], t['ad_rev'])
    t['acoas'] = calc_rate(t['cost'], t['total_rev'])
    return t

plp_curr_sku = agg_plp(plp_curr_rows)
plp_t = plp_totals(plp_curr_sku)
sku_rev_curr = {}
for r in rows_curr:
    sku = str(r[C['sku']]).strip()
    if sku: sku_rev_curr[sku] = num(r[C['rev_curr']])

# ACOAS 去重计算
sku_total_rev_dedup = {}
for sku in plp_curr_sku:
    info_rows = [r for r in plp_curr_rows if str(r[PC['sku']] or '').strip() == sku]
    if info_rows:
        sku_total_rev_dedup[sku] = num(info_rows[0][PC['total_rev']])  # 取第一条的总销售额

dedup_total_rev = sum(sku_total_rev_dedup.values())
dedup_acoas = round(plp_t['cost'] / dedup_total_rev, 4) if dedup_total_rev > 0 else 0

# ===== 函数：广告分类 =====
def get_ad_class(r_data):
    sku = str(r_data[C['sku']]).strip()
    plp_on = str(r_data[C['plp_curr']] or '').strip().upper() == 'Y'
    plg_on = num(r_data[C['plg_curr']]) > 0
    in_plg_detail = sku in plp_plg_enabled
    is_unsold = str(r_data[C['ord8_curr']] or '').strip() == '未出单'

    if in_plg_detail and sku_campaign_count.get(sku, 0) == 1:
        return '单链接PLP+PLG同开'
    elif plp_on and plg_on:
        return 'PLP+PLG同开'
    elif plg_on and not plp_on:
        return '单PLG且未出单' if is_unsold else '单PLG'
    elif plp_on and not plg_on:
        return '单PLP'
    return '无广告'

# ===== 构建21个数据块 =====
print("构建数据块...")

# 1. cum43Data
cum43Data = []
for r_data in rows_curr:
    sku = str(r_data[C['sku']]).strip()
    list_d = get_date(r_data[C['list_date']])
    first_d = get_date(r_data[C['first_order']])
    cum43Data.append({
        'saleNo': r_data[C['sale_no']] or '',
        'SKU': sku,
        'listDate': str(list_d)[:10] if list_d else '',
        'firstOrderDate': str(first_d)[:10] if first_d else '未出单',
        'analyst': get_an(r_data),
        'category': get_cat(r_data),
        'expandType': get_exp(r_data),
        'curSalesQty': int(num(r_data[C['sales_curr']])),
        'prevSalesQty': int(num(r_data[C['sales_prev']])),
        'curRevenue': round(num(r_data[C['rev_curr']]), 2),
        'prevRevenue': round(num(r_data[C['rev_prev']]), 2),
        'curRivalQty': int(num(r_data[C['rival_curr']])),
        'prevRivalQty': int(num(r_data[C['rival_prev']])),
        'curMarketShare': round(num(r_data[C['share_curr']]) * 100, 1),
        'prevMarketShare': round(num(r_data[C['share_prev']]) * 100, 1),
        'cur8dStatus': str(r_data[C['ord8_curr']] or '').strip(),
        'curFreq7': str(r_data[C['freq7_curr']] or '').strip(),
        'curNewFreq7': str(r_data[C['nfreq7_curr']] or '').strip(),
        'curMarketStatus': str(r_data[C['mkt_curr']] or '').strip(),
        'prevMarketStatus': str(r_data[C['mkt_prev']] or '').strip(),
        'curOperation': str(r_data[C['op_curr']] or '').strip(),
        'plpEnabled': 'Y' if str(r_data[C['plp_curr']] or '').strip().upper() == 'Y' else 'N',
        'plgFee': f"{round(num(r_data[C['plg_curr']]) * 100, 1)}%",
        'adClass': get_ad_class(r_data),
    })

# 2. cum43Stats — Y/N/未出单 only count among SKUs with rivals (mutually exclusive)
rows_with_rival = [r for r in rows_curr if num(r[C['rival_curr']]) > 0]
y_curr = sum(1 for r in rows_with_rival if str(r[C['ord8_curr']] or '').strip() == 'Y')
n_curr = sum(1 for r in rows_with_rival if str(r[C['ord8_curr']] or '').strip() == 'N')
no_curr = sum(1 for r in rows_with_rival if str(r[C['ord8_curr']] or '').strip() == '未出单')
normal_cnt = sum(1 for r in rows_curr if str(r[C['mkt_curr']] or '').strip() == '正常')
competitive_cnt = sum(1 for r in rows_curr if str(r[C['mkt_curr']] or '').strip() == '竞争无优势')
nomarket_cnt = sum(1 for r in rows_curr if str(r[C['mkt_curr']] or '').strip() == '无市场')
station_out_cnt = sum(1 for r in rows_curr if str(r[C['mkt_curr']] or '').strip() == '站外出单')

cum43Stats = {
    'total': total_sku, 'yCount': y_curr, 'nCount': n_curr, 'unCount': no_curr,
    'normalCount': normal_cnt, 'competitiveCount': competitive_cnt, 'noMarketCount': nomarket_cnt,
    'stationOutCount': station_out_cnt,
    'hasRivalCount': has_rival_curr, 'noRivalCount': total_sku - has_rival_curr,
}

# 3. lowShareData
lowShareData = []
for r_data in rows_curr:
    share = num(r_data[C['share_curr']]) * 100
    rival = num(r_data[C['rival_curr']])
    if share < 75 and rival > 0:
        sku = str(r_data[C['sku']]).strip()
        lowShareData.append({
            'salesCode': str(r_data[C['sale_no']] or ''),
            'SKU': sku,
            'launchDate': str(get_date(r_data[C['list_date']]))[:10] if get_date(r_data[C['list_date']]) else '',
            'analyst': get_an(r_data),
            'category': get_cat(r_data),
            'expandType': get_exp(r_data),
            'curSalesQty': int(num(r_data[C['sales_curr']])),
            'salesQtyChange': ratio_str(num(r_data[C['sales_curr']]), num(r_data[C['sales_prev']])),
            'curRevenue': round(num(r_data[C['rev_curr']]), 2),
            'revenueChange': ratio_str(num(r_data[C['rev_curr']]), num(r_data[C['rev_prev']])),
            'prevRivalQty': int(num(r_data[C['rival_prev']])),
            'curRivalQty': int(num(r_data[C['rival_curr']])),
            'rivalQtyChange': ratio_str(num(r_data[C['rival_curr']]), num(r_data[C['rival_prev']])),
            'prevMarketShare': round(num(r_data[C['share_prev']]) * 100, 1),
            'curMarketShare': round(share, 1),
            'prevMarketStatus': str(r_data[C['mkt_prev']] or '').strip(),
            'curOperation': str(r_data[C['op_curr']] or '').strip(),
            'curMarketStatus': str(r_data[C['mkt_curr']] or '').strip(),
            'cur8dStatus': str(r_data[C['ord8_curr']] or '').strip(),
            'plpEnabled': 'Y' if str(r_data[C['plp_curr']] or '').strip().upper() == 'Y' else 'N',
            'plgFee': f"{round(num(r_data[C['plg_curr']]) * 100, 1)}%",
            'adClass': get_ad_class(r_data),
        })
lowShareData.sort(key=lambda x: x['curMarketShare'])

# 4. expandTypeData
exp_curr = defaultdict(lambda: {'sku': 0, 'sales_qty': 0, 'rev': 0, 'sold_sku': 0, 'has_rival': 0})
exp_prev = defaultdict(lambda: {'sku': 0, 'sales_qty': 0, 'rev': 0, 'sold_sku': 0, 'has_rival': 0})
for r_data in rows_curr:
    exp = get_exp(r_data)
    exp_curr[exp]['sku'] += 1
    exp_curr[exp]['sales_qty'] += num(r_data[C['sales_curr']])
    exp_curr[exp]['rev'] += num(r_data[C['rev_curr']])
    if num(r_data[C['rival_curr']]) > 0:
        exp_curr[exp]['has_rival'] += 1
        if str(r_data[C['ord8_curr']] or '').strip() in ('Y', 'N'):
            exp_curr[exp]['sold_sku'] += 1
for r_data in rows_prev:
    exp = get_exp(r_data)
    exp_prev[exp]['sku'] += 1
    exp_prev[exp]['sales_qty'] += num(r_data[C['sales_prev']])
    exp_prev[exp]['rev'] += num(r_data[C['rev_prev']])
    if num(r_data[C['rival_prev']]) > 0:
        exp_prev[exp]['has_rival'] += 1
        if str(r_data[C['ord8_prev']] or '').strip() in ('Y', 'N'):
            exp_prev[exp]['sold_sku'] += 1

expandTypeData = []
for exp in EXPAND_TYPES:
    if exp in exp_curr or exp in exp_prev:
        d_c = exp_curr.get(exp, {})
        d_p = exp_prev.get(exp, {})
        cur_sold = d_c.get('sold_sku', 0); has_r_c = d_c.get('has_rival', 0) or 1
        prev_sold = d_p.get('sold_sku', 0); has_r_p = d_p.get('has_rival', 0) or 1
        expandTypeData.append({
            'expandType': exp,
            'curSku': d_c.get('sku', 0), 'prevSku': d_p.get('sku', 0),
            'curSalesSku': cur_sold,
            'curSalesRate': f"{round(cur_sold/has_r_c*100, 1)}%",
            'prevSalesSku': prev_sold,
            'prevSalesRate': f"{round(prev_sold/has_r_p*100, 1)}%",
            'curSalesQty': int(d_c.get('sales_qty', 0)), 'prevSalesQty': int(d_p.get('sales_qty', 0)),
            'salesQtyChange': ratio_str(d_c.get('sales_qty', 0), d_p.get('sales_qty', 0)),
            'curRevenue': round(d_c.get('rev', 0), 2), 'prevRevenue': round(d_p.get('rev', 0), 2),
            'revenueChange': ratio_str(d_c.get('rev', 0), d_p.get('rev', 0)),
        })

# 5. timelinessData
an_time = defaultdict(lambda: {'timely': 0, 'no_8d': 0, 'no_7d': 0})
for r_data in rows_curr:
    an = get_an(r_data)
    freq = str(r_data[C['freq7_curr']] or '').strip()
    nfreq = str(r_data[C['nfreq7_curr']] or '').strip()
    if nfreq == '异常': an_time[an]['no_8d'] += 1
    elif freq == '异常': an_time[an]['no_7d'] += 1
    else: an_time[an]['timely'] += 1

an_time_prev = defaultdict(lambda: {'timely': 0, 'no_8d': 0, 'no_7d': 0})
for r_data in rows_prev:
    an = get_an(r_data)
    freq = str(r_data[C['freq7_prev']] or '').strip()
    nfreq = str(r_data[C['nfreq7_prev']] or '').strip()
    if nfreq == '异常': an_time_prev[an]['no_8d'] += 1
    elif freq == '异常': an_time_prev[an]['no_7d'] += 1
    else: an_time_prev[an]['timely'] += 1

ans_ordered = [a for a in ANALYSTS if a in an_time] + [a for a in sorted(an_time.keys()) if a not in ANALYSTS]
timeliness_total_curr = sum(1 for r in rows_curr if str(r[C['nfreq7_curr']] or '').strip() != '异常' and str(r[C['freq7_curr']] or '').strip() != '异常')
timeliness_total_prev = sum(1 for r in rows_prev if str(r[C['nfreq7_prev']] or '').strip() != '异常' and str(r[C['freq7_prev']] or '').strip() != '异常')

timelinessData = {'analysts': [], 'total': {}}
for an in ans_ordered:
    d_c = an_time.get(an, {'timely': 0, 'no_8d': 0, 'no_7d': 0})
    d_p = an_time_prev.get(an, {'timely': 0, 'no_8d': 0, 'no_7d': 0})
    total_c = d_c['timely'] + d_c['no_8d'] + d_c['no_7d']
    total_p = d_p['timely'] + d_p['no_8d'] + d_p['no_7d']
    rate_c = f"{round(d_c['timely']/total_c*100, 1)}%" if total_c else '0%'
    rate_p = f"{round(d_p['timely']/total_p*100, 1)}%" if total_p else '0%'
    timelinessData['analysts'].append({
        'analyst': an, 'curSku': total_c, 'prevSku': total_p,
        'timelyCount': d_c['timely'], 'prevTimelyCount': d_p['timely'],
        'noAnalysis8dCount': d_c['no_8d'], 'noAnalysis7dCount': d_c['no_7d'],
        'timelyRate': rate_c, 'prevTimelyRate': rate_p,
        'change': ratio_str(float(rate_c.replace('%','')), float(rate_p.replace('%',''))) if rate_c != '0%' and rate_p != '0%' else '-',
    })
tc_all = timeliness_total_curr
tp_all = timeliness_total_prev
tr_c = f"{round(tc_all/total_sku*100, 1)}%" if total_sku else '0%'
tr_p = f"{round(tp_all/len(rows_prev)*100, 1)}%" if rows_prev else '0%'
timelinessData['total'] = {
    'analyst': '合计', 'curSku': total_sku, 'prevSku': len(rows_prev),
    'timelyCount': tc_all, 'prevTimelyCount': tp_all,
    'noAnalysis8dCount': sum(1 for r in rows_curr if str(r[C['nfreq7_curr']] or '').strip() == '异常'),
    'noAnalysis7dCount': sum(1 for r in rows_curr if str(r[C['freq7_curr']] or '').strip() == '异常'),
    'timelyRate': tr_c, 'prevTimelyRate': tr_p,
    'change': ratio_str(float(tr_c.replace('%','')), float(tr_p.replace('%',''))) if tr_c != '0%' and tr_p != '0%' else '-',
}

# 6. hasCompetitorUnsold
has_rival_no_curr = [r for r in rows_curr if str(r[C['ord8_curr']] or '').strip() == '未出单' and num(r[C['rival_curr']]) > 0]
has_rival_no_prev = [r for r in rows_prev if str(r[C['ord8_prev']] or '').strip() == '未出单' and num(r[C['rival_prev']]) > 0]
mkt_has_order = ['竞争无优势', '无市场', '站内无价格优势', '站外出单', '正常', '#N/A', '未知']
mkt_no_order = ['无市场', '未知', '竞争无优势', '#N/A', '其他']

def build_unsold_analysis(items, mkt_order, col_mkt):
    reasons = []
    mkt_counter = Counter(str(r[col_mkt] or '未知') for r in items)
    for mkt in mkt_order:
        reasons.append({'name': mkt, 'count': mkt_counter.get(mkt, 0)})
    by_analyst = []
    for an in ans_ordered:
        row = {'analyst': an}
        an_items = [r for r in items if get_an(r) == an]
        an_mkt = Counter(str(r[col_mkt] or '未知') for r in an_items)
        for mkt in mkt_order:
            row[mkt] = an_mkt.get(mkt, 0)
        row['total'] = len(an_items)
        by_analyst.append(row)
    by_category = []
    cat_items_map = defaultdict(list)
    for r in items: cat_items_map[get_cat(r)].append(r)
    for cat in sorted(cat_items_map.keys()):
        row = {'category': cat}
        cat_mkt = Counter(str(r[col_mkt] or '未知') for r in cat_items_map[cat])
        for mkt in mkt_order:
            row[mkt] = cat_mkt.get(mkt, 0)
        row['total'] = len(cat_items_map[cat])
        by_category.append(row)
    return {
        'total': len(items), 'prevTotal': 0, 'reasons': reasons,
        'byAnalyst': by_analyst, 'byCategory': by_category,
    }

hasCompetitorUnsold = build_unsold_analysis(has_rival_no_curr, mkt_has_order, C['mkt_curr'])
hasCompUnsoldPrev = build_unsold_analysis(has_rival_no_prev, mkt_has_order, C['mkt_prev'])
hasCompetitorUnsold['prevTotal'] = len(has_rival_no_prev)
hasCompetitorUnsold['change'] = len(has_rival_no_curr) - len(has_rival_no_prev)

# 7-11. PLP 广告数据
plpTotal = {
    'campaignCount': plp_t['campaigns'], 'linkCount': plp_t['links'],
    'impression': int(plp_t['imp']), 'click': int(plp_t['click']),
    'sold': int(plp_t['sold']), 'cost': round(plp_t['cost'], 2),
    'revenue': round(plp_t['ad_rev'], 2), 'totalRevenue': round(dedup_total_rev, 2),
    'roas': f"{plp_t['roas']:.2f}", 'cvr': f"{plp_t['cvr']*100:.2f}%",
    'ctr': f"{plp_t['ctr']*100:.2f}%", 'cpc': f"${plp_t['cpc']:.2f}",
    'cpa': f"${plp_t['cpa']:.2f}", 'acos': f"{plp_t['acos']*100:.2f}%",
    'acoas': f"{dedup_acoas*100:.2f}%",
}

# PLP by dimension
def plp_dim_data(by_sku, key_fn):
    groups = defaultdict(lambda: {'imp': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'total_rev': 0, 'campaigns': set(), 'skus': set()})
    for sku, d in by_sku.items():
        info = None
        for r in rows_curr:
            if str(r[C['sku']]).strip() == sku:
                info = r; break
        k = key_fn(sku, info) if info else key_fn(sku, None)
        g = groups[k]
        g['imp'] += d['impr']; g['click'] += d['click']; g['sold'] += d['sold']
        g['cost'] += d['cost']; g['ad_rev'] += d['ad_rev']
        g['campaigns'].update(d['campaigns']); g['skus'].add(sku)
    # Dedup total_rev
    for k, g in groups.items():
        dedup_rev = 0
        for sku in g['skus']:
            dedup_rev += sku_total_rev_dedup.get(sku, 0)
        g['total_rev'] = dedup_rev

    result = []
    for k, g in sorted(groups.items()):
        roas_v = calc_rate(g['ad_rev'], g['cost'])
        cvr_v = calc_rate(g['sold'], g['click'])
        ctr_v = calc_rate(g['click'], g['imp'])
        acos_v = calc_rate(g['cost'], g['ad_rev'])
        acoas_v = calc_rate(g['cost'], g['total_rev'])
        result.append({
            'name': k, 'campaignCount': len(g['campaigns']), 'linkCount': len(g['skus']),
            'impression': int(g['imp']), 'click': int(g['click']), 'sold': int(g['sold']),
            'cost': round(g['cost'], 2), 'revenue': round(g['ad_rev'], 2),
            'totalRevenue': round(g['total_rev'], 2),
            'roas': f"{roas_v:.2f}", 'cvr': f"{cvr_v*100:.2f}%",
            'ctr': f"{ctr_v*100:.2f}%", 'cpc': f"${round(g['cost']/g['click'], 2)}" if g['click'] else '-',
            'cpa': f"${round(g['cost']/g['sold'], 2)}" if g['sold'] else '-',
            'acos': f"{acos_v*100:.2f}%", 'acoas': f"{acoas_v*100:.2f}%",
        })
    return result

plpCategories = plp_dim_data(plp_curr_sku, lambda sku, info: get_cat(info) if info else '未分类')
plpExpandTypes = plp_dim_data(plp_curr_sku, lambda sku, info: get_exp(info) if info else '其他')
plpAnalysts = plp_dim_data(plp_curr_sku, lambda sku, info: get_an(info) if info else '未知')

# PLP prev totals
plp_prev_rows_list = []
for row in ws_plp.iter_rows(min_row=2, values_only=True):
    period = str(row[PC['period']] or '').strip()
    sku = str(row[PC['sku']] or '').strip()
    if not sku or sku.startswith('广告') or sku.startswith('总数据'): continue
    list_d = get_date(row[PC['list_date']])
    if period == PLP_PREV and list_d and list_d <= cutoff_prev:
        plp_prev_rows_list.append(list(row))
plp_prev_sku = agg_plp(plp_prev_rows_list)
plp_pt = plp_totals(plp_prev_sku)
plpPrevTotal = {
    'campaignCount': plp_pt['campaigns'], 'linkCount': plp_pt['links'],
    'impression': int(plp_pt['imp']), 'click': int(plp_pt['click']),
    'sold': int(plp_pt['sold']), 'cost': round(plp_pt['cost'], 2),
    'revenue': round(plp_pt['ad_rev'], 2), 'totalRevenue': round(plp_pt['total_rev'], 2),
    'roas': f"{plp_pt['roas']:.2f}", 'cvr': f"{plp_pt['cvr']*100:.2f}%",
    'ctr': f"{plp_pt['ctr']*100:.2f}%", 'cpc': f"${plp_pt['cpc']:.2f}",
    'cpa': f"${plp_pt['cpa']:.2f}", 'acos': f"{plp_pt['acos']*100:.2f}%",
    'acoas': f"{plp_pt['acoas']*100:.2f}%",
}

# 12. unsoldNoCompetitor
no_rival_no_curr = [r for r in rows_curr if str(r[C['ord8_curr']] or '').strip() == '未出单' and num(r[C['rival_curr']]) == 0]
no_rival_no_prev = [r for r in rows_prev if str(r[C['ord8_prev']] or '').strip() == '未出单' and num(r[C['rival_prev']]) == 0]

unsoldNoCompetitor = build_unsold_analysis(no_rival_no_curr, mkt_no_order, C['mkt_curr'])
unsoldNoCompetitor['prevTotal'] = len(no_rival_no_prev)
unsoldNoCompetitor['change'] = len(no_rival_no_curr) - len(no_rival_no_prev)

# 13. prevWeekKpi
prevWeekKpi = {
    'prevTotalSku': len(rows_prev),
    'prevTotalSalesQty': total_sales_prev,
    'prevTotalRevenue': round(total_rev_prev, 2),
    'salesQtyChange': ratio_str(total_sales_curr, total_sales_prev),
    'revenueChange': ratio_str(total_rev_curr, total_rev_prev),
    'skuChange': ratio_str(total_sku, len(rows_prev)),
    'prevTimelyRate': tr_p,
    'prevSoldRate': f"{round((y_curr+n_curr)/has_rival_curr*100, 1)}%" if has_rival_curr else '0%',
    'prevLowShareCount': len(lowShareData),
    'deptRatio': f"{dept_ratio:.1f}%",
    'deptTotalRevenue': round(dept_total_revenue, 2),
}

# 14. plgStats
plg_categories_map = {'PLP+PLG同开': [], '单链接PLP+PLG同开': [], '单PLG': [], '单PLP': [], '单PLG且未出单': [], '无广告': []}
for r_data in rows_curr:
    adc = get_ad_class(r_data)
    if adc in plg_categories_map:
        plg_categories_map[adc].append(r_data)

plgStats = {
    'plpAndPlgBothCount': len(plg_categories_map['PLP+PLG同开']),
    'singleLinkPlpPlgCount': len(plg_categories_map['单链接PLP+PLG同开']),
    'plgOnlyCount': len(plg_categories_map['单PLG']),
    'plpOnlyCount': len(plg_categories_map['单PLP']),
    'noAdCount': len(plg_categories_map['无广告']),
    'plpDisabledNoSaleCount': len(plg_categories_map['单PLG且未出单']),
    'totalNewProducts': total_sku,
    'byAnalyst': [],
}
for an in ans_ordered:
    an_items = [r for r in rows_curr if get_an(r) == an]
    an_ad = Counter(get_ad_class(r) for r in an_items)
    plgStats['byAnalyst'].append({
        'analyst': an, 'total': len(an_items),
        'plpAndPlgBoth': an_ad.get('PLP+PLG同开', 0),
        'singleLinkPlpPlg': an_ad.get('单链接PLP+PLG同开', 0),
        'plgOnly': an_ad.get('单PLG', 0), 'plpOnly': an_ad.get('单PLP', 0),
        'noAd': an_ad.get('无广告', 0), 'plpDisabledNoSale': an_ad.get('单PLG且未出单', 0),
    })

# 15-16. categoryRevenueData / analystRevenueData
categoryRevenueData = []
cat_curr_data = defaultdict(lambda: {'sku': 0, 'new': 0, 'sales': 0, 'rev': 0})
cat_prev_data = defaultdict(lambda: {'sku': 0, 'sales': 0, 'rev': 0})
for r_data in rows_curr:
    cat = get_cat(r_data)
    list_d = get_date(r_data[C['list_date']])
    cat_curr_data[cat]['sku'] += 1
    if list_d and list_d > cutoff_prev: cat_curr_data[cat]['new'] += 1
    cat_curr_data[cat]['sales'] += num(r_data[C['sales_curr']])
    cat_curr_data[cat]['rev'] += num(r_data[C['rev_curr']])
for r_data in rows_prev:
    cat = get_cat(r_data)
    cat_prev_data[cat]['sku'] += 1
    cat_prev_data[cat]['sales'] += num(r_data[C['sales_prev']])
    cat_prev_data[cat]['rev'] += num(r_data[C['rev_prev']])

cats_ordered = [c for c in CATEGORIES if c in cat_curr_data] + [c for c in sorted(cat_curr_data.keys()) if c not in CATEGORIES]
for cat in cats_ordered:
    d_c = cat_curr_data.get(cat, {})
    d_p = cat_prev_data.get(cat, {})
    categoryRevenueData.append({
        'category': cat, 'curSku': d_c.get('sku', 0), 'curNewSku': d_c.get('new', 0),
        'curSalesQty': int(d_c.get('sales', 0)), 'prevSalesQty': int(d_p.get('sales', 0)),
        'salesQtyChange': ratio_str(d_c.get('sales', 0), d_p.get('sales', 0)),
        'curRevenue': round(d_c.get('rev', 0), 2), 'prevRevenue': round(d_p.get('rev', 0), 2),
        'revenueChange': ratio_str(d_c.get('rev', 0), d_p.get('rev', 0)),
    })

analystRevenueData = []
an_curr_data = defaultdict(lambda: {'sku': 0, 'new': 0, 'sales': 0, 'rev': 0})
an_prev_data = defaultdict(lambda: {'sku': 0, 'sales': 0, 'rev': 0})
for r_data in rows_curr:
    an = get_an(r_data)
    list_d = get_date(r_data[C['list_date']])
    an_curr_data[an]['sku'] += 1
    if list_d and list_d > cutoff_prev: an_curr_data[an]['new'] += 1
    an_curr_data[an]['sales'] += num(r_data[C['sales_curr']])
    an_curr_data[an]['rev'] += num(r_data[C['rev_curr']])
for r_data in rows_prev:
    an = get_an(r_data)
    an_prev_data[an]['sku'] += 1
    an_prev_data[an]['sales'] += num(r_data[C['sales_prev']])
    an_prev_data[an]['rev'] += num(r_data[C['rev_prev']])
for an in ans_ordered:
    d_c = an_curr_data.get(an, {})
    d_p = an_prev_data.get(an, {})
    analystRevenueData.append({
        'analyst': an, 'curSku': d_c.get('sku', 0), 'curNewSku': d_c.get('new', 0),
        'curSalesQty': int(d_c.get('sales', 0)), 'prevSalesQty': int(d_p.get('sales', 0)),
        'salesQtyChange': ratio_str(d_c.get('sales', 0), d_p.get('sales', 0)),
        'curRevenue': round(d_c.get('rev', 0), 2), 'prevRevenue': round(d_p.get('rev', 0), 2),
        'revenueChange': ratio_str(d_c.get('rev', 0), d_p.get('rev', 0)),
    })

# 17. plgRecords
plgRecords = []
for r_data in rows_curr:
    sku = str(r_data[C['sku']]).strip()
    list_d = get_date(r_data[C['list_date']])
    first_d = get_date(r_data[C['first_order']])
    ac = get_ad_class(r_data)
    plgRecords.append({
        'salesCode': str(r_data[C['sale_no']] or ''), 'SKU': sku,
        'launchDate': str(list_d)[:10] if list_d else '',
        'firstSaleDate': str(first_d)[:10] if first_d else '未出单',
        'analyst': get_an(r_data), 'category': get_cat(r_data),
        'expandType': get_exp(r_data),
        'cur8dStatus': str(r_data[C['ord8_curr']] or '').strip(),
        'curSalesQty': int(num(r_data[C['sales_curr']])),
        'curRevenue': round(num(r_data[C['rev_curr']]), 2),
        'curRivalQty': int(num(r_data[C['rival_curr']])),
        'curMarketShare': f"{round(num(r_data[C['share_curr']]) * 100, 1)}%",
        'curMarketStatus': str(r_data[C['mkt_curr']] or '').strip(),
        'curOperation': str(r_data[C['op_curr']] or '').strip(),
        'plpEnabled': 'Y' if str(r_data[C['plp_curr']] or '').strip().upper() == 'Y' else 'N',
        'plgFee': f"{round(num(r_data[C['plg_curr']]) * 100, 1)}%",
        'adClass': ac,
    })

# 18-20. plpSummaryData: merge plpCategories, plpExpandTypes, plpAnalysts
plpSummaryData = []
for d in plpCategories:
    d['dimType'] = '品线'; plpSummaryData.append(d)
for d in plpExpandTypes:
    d['dimType'] = '拓展类型'; plpSummaryData.append(d)
for d in plpAnalysts:
    d['dimType'] = '分析人'; plpSummaryData.append(d)

# 21. plpDetailData
plpDetailData = []
for row in plp_curr_rows:
    sku = str(row[PC['sku']] or '').strip()
    info = None
    for r in rows_curr:
        if str(r[C['sku']]).strip() == sku:
            info = r; break
    cost = num(row[PC['cost']]); ad_rev = num(row[PC['ad_rev']])
    total_r = num(row[PC['total_rev']]); impr = num(row[PC['impr']])
    click = num(row[PC['click']]); sold = num(row[PC['sold']])
    plg_y = str(row[PC['plg_enabled']] or '').strip() == 'Y'
    roas_v = calc_rate(ad_rev, cost)
    cvr_v = calc_rate(sold, click)
    ctr_v = calc_rate(click, impr)
    acos_v = calc_rate(cost, ad_rev)
    acoas_v = calc_rate(cost, total_r)
    ad_class = get_ad_class(info) if info else '无广告'
    plpDetailData.append({
        'period': row[PC['period']] or '', 'campaign': row[PC['campaign']] or '',
        'SKU': sku, 'id': str(row[PC['id']] or ''), 'store': row[PC['store']] or '',
        'plpStartDate': str(get_date(row[PC['plp_start']]))[:10] if get_date(row[PC['plp_start']]) else '',
        'listDate': str(get_date(row[PC['list_date']]))[:10] if get_date(row[PC['list_date']]) else '',
        'firstOrderDate': str(get_date(row[PC['first_order']]))[:10] if get_date(row[PC['first_order']]) else '',
        'analyst': get_an(info) if info else (row[PC['analyst']] or ''),
        'category': get_cat(info) if info else (row[PC['category']] or '未分类'),
        'expandType': get_exp(info) if info else (row[PC['expand_type']] or ''),
        'impressions': int(impr), 'clicks': int(click), 'salesQty': int(sold),
        'spend': round(cost, 2), 'adRevenue': round(ad_rev, 2),
        'totalRevenue': round(total_r, 2),
        'roas': roas_v, 'cvr': cvr_v, 'ctr': ctr_v,
        'cpc': round(cost / click, 3) if click else 0,
        'cpa': round(cost / sold, 2) if sold else 0,
        'acos': acos_v, 'acoas': acoas_v,
        'plgEnabled': 'Y' if plg_y else 'N',
        'adClass': ad_class,
    })
plpDetailData.sort(key=lambda x: (x['analyst'], x['SKU']))

# ===== 品效数据（新增）=====
WEEK_SALES_COLS = [7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17]
WEEK_REV_COLS  = [19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29]
WEEK_PERIODS = [
    ('3.5-3.11', 3), ('3.12-3.18', 3), ('3.19-3.25', 3), ('3.26-4.1', 3),
    ('4.2-4.8', 4), ('4.9-4.15', 4), ('4.16-4.22', 4), ('4.23-4.29', 4),
    ('4.30-5.6', 5), ('5.7-5.13', 5), ('5.14-5.20', 5),
]

cohort_data = defaultdict(lambda: {
    'sku_count': 0, 'skus': set(),
    'monthly_sales': defaultdict(float), 'monthly_revenue': defaultdict(float),
    'monthly_sku_sold': defaultdict(set),
})
for r_data in rows_curr:
    list_d = get_date(r_data[C['list_date']])
    if not list_d: continue
    launch_month = list_d.month
    sku = str(r_data[C['sku']]).strip()
    cohort = cohort_data[launch_month]
    cohort['sku_count'] += 1
    cohort['skus'].add(sku)
    for wi, (period, month_num) in enumerate(WEEK_PERIODS):
        sales = num(r_data[WEEK_SALES_COLS[wi]])
        rev = num(r_data[WEEK_REV_COLS[wi]])
        if sales > 0:
            cohort['monthly_sales'][month_num] += sales
            cohort['monthly_sku_sold'][month_num].add(sku)
        if rev > 0:
            cohort['monthly_revenue'][month_num] += rev

launch_months = sorted(cohort_data.keys())
MONTH_LABELS = {1: '1月', 2: '2月', 3: '3月', 4: '4月', 5: '5月'}

pinxiaoData = []
for m in launch_months:
    d = cohort_data[m]
    sc = d['sku_count']
    monthly_rev = []
    monthly_px = []
    for mon in range(1, 6):
        rev = d['monthly_revenue'].get(mon, 0)
        monthly_rev.append(round(rev, 2))
        monthly_px.append(round(rev/sc, 2) if sc > 0 else 0)
    pinxiaoData.append({
        'launchMonth': m, 'skuCount': sc,
        'monthlyRevenue': monthly_rev,
        'monthlyPinxiao': monthly_px,
    })

# 品效 × 品线
cat_cohort = defaultdict(lambda: defaultdict(lambda: {'sku_count': 0, 'latest_month_rev': 0}))
for r_data in rows_curr:
    list_d = get_date(r_data[C['list_date']])
    if not list_d: continue
    cat = get_cat(r_data)
    launch_month = list_d.month
    sku = str(r_data[C['sku']]).strip()
    cat_cohort[cat][launch_month]['sku_count'] += 1
    for wi, (period, month_num) in enumerate(WEEK_PERIODS):
        if month_num == 5:
            cat_cohort[cat][launch_month]['latest_month_rev'] += num(r_data[WEEK_REV_COLS[wi]])

pinxiaoCatData = []
for cat in cats_ordered:
    if cat not in cat_cohort: continue
    px_list = []
    for m in launch_months:
        d = cat_cohort[cat].get(m, {'sku_count': 0, 'latest_month_rev': 0})
        sc = d['sku_count']
        rev = d['latest_month_rev']
        px_list.append(round(rev/sc) if sc > 0 else 0)
    pinxiaoCatData.append({'category': cat, 'pinxiao': px_list})

# 品效 × 拓展类型
exp_cohort = defaultdict(lambda: defaultdict(lambda: {'sku_count': 0, 'latest_month_rev': 0}))
for r_data in rows_curr:
    list_d = get_date(r_data[C['list_date']])
    if not list_d: continue
    exp = get_exp(r_data)
    launch_month = list_d.month
    exp_cohort[exp][launch_month]['sku_count'] += 1
    for wi, (period, month_num) in enumerate(WEEK_PERIODS):
        if month_num == 5:
            exp_cohort[exp][launch_month]['latest_month_rev'] += num(r_data[WEEK_REV_COLS[wi]])

pinxiaoExpData = []
for exp in EXPAND_TYPES:
    if exp not in exp_cohort: continue
    px_list = []
    for m in launch_months:
        d = exp_cohort[exp].get(m, {'sku_count': 0, 'latest_month_rev': 0})
        sc = d['sku_count']
        rev = d['latest_month_rev']
        px_list.append(round(rev/sc) if sc > 0 else 0)
    pinxiaoExpData.append({'expandType': exp, 'pinxiao': px_list})

# ===== 市场分布数据（Tab7）=====
def normalize_mkt(r_data, col_idx):
    v = str(r_data[col_idx] or '').strip()
    return v if v in ALL_MKT_STATUSES else '其他'

mkt_curr_counter = Counter()
mkt_prev_counter = Counter()
for r_data in rows_curr:
    mkt_curr_counter[normalize_mkt(r_data, C['mkt_curr'])] += 1
for r_data in rows_prev:
    mkt_prev_counter[normalize_mkt(r_data, C['mkt_prev'])] += 1

mktDistOverall = {
    'curTotal': total_sku,
    'prevTotal': len(rows_prev),
    'distribution': []
}
for s in ALL_MKT_STATUSES:
    cur_c = mkt_curr_counter.get(s, 0)
    prev_c = mkt_prev_counter.get(s, 0)
    mktDistOverall['distribution'].append({
        'status': s,
        'curCount': cur_c,
        'prevCount': prev_c,
        'curPct': round(cur_c / total_sku * 100, 1) if total_sku else 0,
        'prevPct': round(prev_c / len(rows_prev) * 100, 1) if rows_prev else 0,
        'change': cur_c - prev_c,
    })

mktDistAnalyst = {'statuses': ALL_MKT_STATUSES, 'data': []}
for an in ans_ordered:
    an_curr_rows = [r for r in rows_curr if get_an(r) == an]
    an_prev_rows = [r for r in rows_prev if get_an(r) == an]
    cur_m = Counter(normalize_mkt(r, C['mkt_curr']) for r in an_curr_rows)
    prev_m = Counter(normalize_mkt(r, C['mkt_prev']) for r in an_prev_rows)
    entry = {'analyst': an, 'curTotal': len(an_curr_rows), 'prevTotal': len(an_prev_rows)}
    for s in ALL_MKT_STATUSES:
        entry[s + '_cur'] = cur_m.get(s, 0)
        entry[s + '_prev'] = prev_m.get(s, 0)
    mktDistAnalyst['data'].append(entry)

mktDistCategory = {'statuses': ALL_MKT_STATUSES, 'data': []}
for cat in cats_ordered:
    cat_curr_rows = [r for r in rows_curr if get_cat(r) == cat]
    cat_prev_rows = [r for r in rows_prev if get_cat(r) == cat]
    cur_m = Counter(normalize_mkt(r, C['mkt_curr']) for r in cat_curr_rows)
    prev_m = Counter(normalize_mkt(r, C['mkt_prev']) for r in cat_prev_rows)
    entry = {'category': cat, 'curTotal': len(cat_curr_rows), 'prevTotal': len(cat_prev_rows)}
    for s in ALL_MKT_STATUSES:
        entry[s + '_cur'] = cur_m.get(s, 0)
        entry[s + '_prev'] = prev_m.get(s, 0)
    mktDistCategory['data'].append(entry)


# ===== 货值分布 =====
PRICE_RANGES = [
    ("$0-10", 0, 10), ("$10-20", 10, 20), ("$20-30", 20, 30),
    ("$30-50", 30, 50), ("$50-100", 50, 100), ("$100+", 100, float("inf"))
]
price_list = []
for r_data in rows_curr:
    sales = num(r_data[C["sales_curr"]])
    rev = num(r_data[C["rev_curr"]])
    if sales > 0:
        price_list.append({
            "sku": str(r_data[C["sku"]]).strip(),
            "price": round(rev / sales, 2),
            "analyst": get_an(r_data),
            "category": get_cat(r_data),
        })

priceDist = []
for label, lo, hi in PRICE_RANGES:
    cnt = sum(1 for p in price_list if lo <= p["price"] < hi)
    priceDist.append({"range": label, "count": cnt, "pct": round(cnt/len(price_list)*100,1) if price_list else 0})

avg_price = round(sum(p["price"] for p in price_list)/len(price_list), 2) if price_list else 0
sorted_prices = sorted(p["price"] for p in price_list)
n = len(sorted_prices)
median_price = round(sorted_prices[n//2], 2) if n > 0 else 0

priceByAnalyst = []
for an in ans_ordered:
    an_prices = [p for p in price_list if p["analyst"] == an]
    entry = {"analyst": an, "total": len(an_prices)}
    for label, lo, hi in PRICE_RANGES:
        entry[label] = sum(1 for p in an_prices if lo <= p["price"] < hi)
    priceByAnalyst.append(entry)

priceByCategory = []
for cat in cats_ordered:
    cat_prices = [p for p in price_list if p["category"] == cat]
    entry = {"category": cat, "total": len(cat_prices)}
    for label, lo, hi in PRICE_RANGES:
        entry[label] = sum(1 for p in cat_prices if lo <= p["price"] < hi)
    priceByCategory.append(entry)

priceOverview = {
    "avgPrice": avg_price, "medianPrice": median_price,
    "totalWithSales": len(price_list),
    "priceRanges": [label for label, _, _ in PRICE_RANGES],
    "distribution": priceDist,
    "byAnalyst": priceByAnalyst,
    "byCategory": priceByCategory,
}

# ===== 市占比分布 by 品线 =====
shareTierByCat = []
for cat in cats_ordered:
    cat_rows = [r for r in rows_curr if get_cat(r) == cat]
    high = sum(1 for r in cat_rows if num(r[C["share_curr"]]) * 100 >= 75)
    mid = sum(1 for r in cat_rows if 50 <= num(r[C["share_curr"]]) * 100 < 75)
    low = sum(1 for r in cat_rows if num(r[C["share_curr"]]) * 100 < 50)
    shareTierByCat.append({"category": cat, "high": high, "mid": mid, "low": low, "total": len(cat_rows)})

shareTierOverview = {
    "tiers": ["高市占比(≥75%)", "中市占比(50-75%)", "低市占比(<50%)"],
    "byCategory": shareTierByCat,
}
# ===== 序列化 JS 变量 =====
print("序列化数据...")
data_blocks = {
    'cum43Data': cum43Data, 'cum43Stats': cum43Stats, 'lowShareData': lowShareData,
    'expandTypeData': expandTypeData, 'timelinessData': timelinessData,
    'hasCompetitorUnsold': hasCompetitorUnsold, 'plpTotal': plpTotal,
    'plpPrevTotal': plpPrevTotal, 'plpCategories': plpCategories,
    'plpExpandTypes': plpExpandTypes, 'plpAnalysts': plpAnalysts,
    'unsoldNoCompetitor': unsoldNoCompetitor, 'prevWeekKpi': prevWeekKpi,
    'plgStats': plgStats, 'categoryRevenueData': categoryRevenueData,
    'analystRevenueData': analystRevenueData, 'plgRecords': plgRecords,
    'plpSummaryData': plpSummaryData, 'plpDetailData': plpDetailData,
    'pinxiaoData': pinxiaoData, 'pinxiaoCatData': pinxiaoCatData,
    'pinxiaoExpData': pinxiaoExpData, 'launchMonths': launch_months,
    'mktDistOverall': mktDistOverall, 'mktDistAnalyst': mktDistAnalyst, 'mktDistCategory': mktDistCategory, 'priceOverview': priceOverview, 'shareTierOverview': shareTierOverview,
}

js_lines = []
for key, val in data_blocks.items():
    js_lines.append(f"const {key} = {json.dumps(val, ensure_ascii=False, indent=0)};")

js_vars_block = '\n'.join(js_lines)

print(f"数据块: {len(data_blocks)} 个")
print(f"JS变量总大小: {len(js_vars_block)} 字符")

# ===== HTML 生成 =====
print("生成HTML...")

CSS = r'''
<style>
* { margin: 0; padding: 0; box-sizing: border-box; }
body { font-family: 'Segoe UI', '微软雅黑', sans-serif; background: #f0f2f5; color: #1a1a2e; display: flex; min-height: 100vh; }
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #f1f1f1; }
::-webkit-scrollbar-thumb { background: #ccc; border-radius: 3px; }

.sidebar { width: 230px; min-width: 230px; background: #fff; height: 100vh; position: sticky; top: 0; overflow-y: auto; box-shadow: 2px 0 8px rgba(0,0,0,0.06); z-index: 10; }
.sidebar h2 { font-size: 16px; color: #0f3460; padding: 20px 16px 12px; border-bottom: 2px solid #e8f0fe; }
.sidebar ul { list-style: none; padding: 8px 0; }
.sidebar li a { display: block; padding: 12px 16px; color: #555; text-decoration: none; font-size: 13px; border-left: 3px solid transparent; transition: all 0.2s; }
.sidebar li a:hover, .sidebar li a.active { background: #f0f6ff; color: #0f3460; border-left-color: #0f3460; font-weight: 600; }

.main { flex: 1; padding: 24px; overflow: auto; }
.header { background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%); background-size: 200% 200%; animation: gradientShift 8s ease infinite; color: #fff; padding: 28px 40px; border-radius: 10px; margin-bottom: 20px; }
.header h1 { font-size: 24px; font-weight: 700; letter-spacing: 2px; }
.header p { font-size: 13px; opacity: 0.75; margin-top: 6px; }

.tab-content { display: none; }
.tab-content.active { display: block; animation: fadeIn 0.4s ease-out; }

/* === KPI Cards === */
.kpi-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 12px; margin-bottom: 20px; }
.kpi-card { background: #fff; border-radius: 10px; padding: 16px 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); text-align: center; transition: transform 0.2s, box-shadow 0.2s; animation: fadeInUp 0.5s ease-out forwards; }
.kpi-card:hover { transform: translateY(-2px); box-shadow: 0 4px 16px rgba(0,0,0,0.12); }
.kpi-card .label { font-size: 11px; color: #888; margin-bottom: 6px; }
.kpi-card .val { font-size: 24px; font-weight: 700; }
.kpi-card .hb { font-size: 11px; margin-top: 4px; }
.kpi-card.primary .val { color: #0f3460; }
.kpi-card.success .val { color: #08845a; }
.kpi-card.warning .val { color: #e07b24; }
.kpi-card.danger .val { color: #c0392b; }
.kpi-card.info .val { color: #2980b9; }
.kpi-card.purple .val { color: #8e44ad; }
.kpi-card.animate-pulse { animation: pulse 2s ease-in-out infinite; }

/* === Sections & Charts === */
.section { background: #fff; border-radius: 10px; padding: 20px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); animation: fadeInUp 0.55s ease-out forwards; }
.section h3 { font-size: 15px; color: #0f3460; margin-bottom: 16px; padding-bottom: 8px; border-bottom: 2px solid #e8f0fe; }
.sub-module { margin-bottom: 16px; }
.sub-module h4 { font-size: 13px; color: #0f3460; background: #f5f7ff; padding: 8px 12px; border-left: 3px solid #0f3460; margin-bottom: 10px; }

.chart-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 20px; }
.chart-box { background: #fff; border-radius: 10px; padding: 16px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); animation: fadeInUp 0.6s ease-out forwards; transition: transform 0.25s, box-shadow 0.25s; }
.chart-box:hover { transform: translateY(-3px); box-shadow: 0 6px 20px rgba(0,0,0,0.12); }
.chart-box h4 { font-size: 13px; color: #0f3460; margin-bottom: 12px; }
.chart-box canvas { max-height: 280px; }

/* === Scrollable Tables with Sticky Header/Footer === */
.table-scroll-wrap { max-height: 68vh; overflow-y: auto; overflow-x: auto; border: 1px solid #e0e0e0; border-radius: 6px; }
.table-scroll-wrap thead th { position: sticky; top: 0; z-index: 10; }
.table-scroll-wrap tfoot td { position: sticky; bottom: 0; z-index: 10; background: #f0f6ff; font-weight: 700; }

.data-table { width: 100%; border-collapse: collapse; font-size: 12px; }
.data-table th { background: #0f3460; color: #fff; padding: 8px 10px; font-weight: 600; white-space: nowrap; text-align: center; }
.data-table td { padding: 6px 10px; border-bottom: 1px solid #eee; text-align: center; }
.data-table tr:hover td { background: #f5f7ff; }
.data-table .num { text-align: right; }
.data-table .total-row td { font-weight: 700; background: #e8f0fe; }

.hb-up { color: #08845a; font-weight: 600; }
.hb-down { color: #c0392b; font-weight: 600; }
.hb-flat { color: #888; }

.badge { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 10px; color: #fff; font-weight: 600; }
.badge-green { background: #08845a; }
.badge-orange { background: #e07b24; }
.badge-red { background: #c0392b; }
.badge-blue { background: #2980b9; }
.badge-purple { background: #8e44ad; }
.badge-gray { background: #6c757d; }

.filter-bar { background: #f5f7ff; padding: 12px 16px; border-radius: 8px; margin-bottom: 16px; display: flex; flex-wrap: wrap; gap: 10px; align-items: center; }
.filter-bar-sticky { position: sticky; top: -1px; z-index: 100; background: #f5f7ff; border-radius: 0; padding: 14px 18px; margin-bottom: 14px; display: flex; flex-wrap: wrap; gap: 10px; align-items: center; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }
.filter-bar select, .filter-bar .fg, .filter-bar-sticky .fg { display: flex; align-items: center; gap: 4px; font-size: 12px; }
.filter-bar .fg label, .filter-bar-sticky .fg label { color: #555; white-space: nowrap; }
.filter-bar-sticky select { padding: 6px 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 12px; background: #fff; }
.filter-bar select:focus, .filter-bar-sticky input[type="checkbox"] { margin: 0 4px 0 0; vertical-align: middle; }
.filter-bar-sticky select:focus { border-color: #0f3460; outline: none; }
.filter-bar .reset-btn, .filter-bar-sticky .reset-btn { color: #c0392b; border: 1px solid #c0392b; background: #fff; padding: 6px 14px; border-radius: 6px; cursor: pointer; font-size: 12px; }
.filter-bar .reset-btn:hover, .filter-bar-sticky .reset-btn:hover { background: #c0392b; color: #fff; }
.filter-bar .count, .filter-bar-sticky .count { font-size: 12px; color: #888; margin-left: auto; }

/* === Animations === */
@keyframes fadeInUp { from { opacity: 0; transform: translateY(24px); } to { opacity: 1; transform: translateY(0); } }
@keyframes fadeIn { from { opacity: 0; } to { opacity: 1; } }
@keyframes pulse { 0%,100% { transform: scale(1); } 50% { transform: scale(1.04); } }
@keyframes shimmer { 0% { background-position: -200px 0; } 100% { background-position: calc(200px + 100%) 0; } }
@keyframes slideInLeft { from { opacity: 0; transform: translateX(-30px); } to { opacity: 1; transform: translateX(0); } }
@keyframes slideInRight { from { opacity: 0; transform: translateX(30px); } to { opacity: 1; transform: translateX(0); } }
@keyframes bounceIn { 0% { opacity: 0; transform: scale(0.3); } 50% { opacity: 1; transform: scale(1.05); } 70% { transform: scale(0.9); } 100% { transform: scale(1); } }
@keyframes gradientShift { 0% { background-position: 0% 50%; } 50% { background-position: 100% 50%; } 100% { background-position: 0% 50%; } }
.anim-fade-in-up { animation: fadeInUp 0.6s ease-out forwards; }
.anim-fade-in { animation: fadeIn 0.5s ease-out forwards; }
.anim-slide-left { animation: slideInLeft 0.5s ease-out forwards; }
.anim-slide-right { animation: slideInRight 0.5s ease-out forwards; }
.anim-bounce { animation: bounceIn 0.7s ease-out forwards; }

.report-block { background: #f9fafb; border-radius: 8px; padding: 14px 18px; margin-bottom: 12px; border-left: 4px solid #0f3460; position: relative; }
.report-block h4 { font-size: 13px; color: #0f3460; margin-bottom: 8px; }
.report-block pre { white-space: pre-wrap; font-size: 12px; color: #444; font-family: '微软雅黑', sans-serif; margin: 0; }
.report-block .copy-btn { position: absolute; top: 12px; right: 16px; background: #0f3460; color: #fff; border: none; padding: 4px 12px; border-radius: 4px; cursor: pointer; font-size: 11px; transition: background 0.2s; }
.report-block .copy-btn:hover { background: #16213e; }

.risk-high { background: #fff5f5; border-left-color: #c0392b; }
.risk-high h4 { color: #c0392b; }
.risk-medium { background: #fffdf5; border-left-color: #e07b24; }
.risk-medium h4 { color: #e07b24; }
.risk-low { background: #f5fff5; border-left-color: #08845a; }
.risk-low h4 { color: #08845a; }

.findings-card { background: #f5f9ff; border-radius: 8px; padding: 14px 18px; margin-bottom: 10px; border-left: 4px solid #2980b9; }
.findings-card .title { font-size: 13px; font-weight: 600; color: #0f3460; margin-bottom: 4px; }
.findings-card .desc { font-size: 12px; color: #555; }

.action-card { background: #fdf5ff; border-radius: 8px; padding: 14px 18px; margin-bottom: 10px; border-left: 4px solid #8e44ad; }
.action-card .title { font-size: 13px; font-weight: 600; color: #8e44ad; margin-bottom: 4px; }
.action-card .desc { font-size: 12px; color: #555; }

.pinxiao-table .px-high { color: #08845a; font-weight: 700; }
.pinxiao-table .px-mid { color: #000; }
.pinxiao-table .px-low { color: #c0392b; font-weight: 600; }
.trend-up { color: #08845a; }
.trend-down { color: #c0392b; }
.trend-flat { color: #888; }

@media(max-width:900px) {
  .sidebar { display: none; }
  .chart-grid { grid-template-columns: 1fr; }
  .kpi-grid { grid-template-columns: repeat(2,1fr); }
}
</style>
'''

HTML_BODY = r'''
<div class="sidebar">
  <h2>&#128202; 新品周报 5.14-5.20</h2>
  <ul>
    <li><a href="javascript:void(0)" class="active" onclick="switchTab('tab1',this)">&#128200; 总盘概览</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab2',this)">&#127758; 市场分布</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab3',this)">&#128270; 品效分析</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab4',this)">&#128065; 低占比分析</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab5',this)">&#128176; 广告追踪</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab6',this)">&#128202; 四三累计</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab7',this)">&#128221; 汇报输出</a></li>
  </ul>
</div>
<div class="main">
  <div class="header">
    <h1>&#128202; 新品周报看板 &middot; 5.14 - 5.20 &nbsp;&#127793;</h1>
    <p>数据截至 2026-05-20 &nbsp;|&nbsp; 三部新品 &nbsp;|&nbsp; 品效分析 + PLP/PLG广告追踪</p>
  </div>

  <div class="tab-content active" id="tab1">
    <div class="kpi-grid" id="t1-kpi"></div>
    <div class="chart-grid">
      <div class="chart-box"><h4>&#128200; 出单分布</h4><canvas id="chart-ord8"></canvas></div>
      <div class="chart-box"><h4>&#128200; 品线销量对比</h4><canvas id="chart-cat-sales"></canvas></div>
      <div class="chart-box"><h4>&#128101; 分析人销量对比</h4><canvas id="chart-an-sales"></canvas></div>
      <div class="chart-box"><h4>&#9200; 分析人及时率对比</h4><canvas id="chart-an-timely"></canvas></div>
      <div class="chart-box"><h4>&#128176; 品线销售额对比</h4><canvas id="chart-cat-rev"></canvas></div>
      <div class="chart-box"><h4>&#128101; 分析人销售额对比</h4><canvas id="chart-an-rev"></canvas></div>
    </div>
    <div class="section"><h3>&#128200; 新品出单情况</h3><div id="t1-ord8"></div></div>
    <div class="section"><h3>&#128269; 多维度分析</h3>
      <div class="sub-module"><h4>品线维度</h4><div id="t1-cat-table"></div></div>
      <div class="sub-module"><h4>分析人维度</h4><div id="t1-an-table"></div></div>
    </div>
    <div class="section"><h3>&#128202; 拓展类型 & 及时率</h3>
      <div class="sub-module"><h4>拓展类型</h4><div id="t1-exp-table"></div></div>
      <div class="sub-module"><h4>分析及时率</h4><div id="t1-time-table"></div></div>
    </div>
  </div>

  <div class="tab-content" id="tab4">
    <div class="kpi-grid" id="t4-kpi"></div>
    <div class="section"><h3>&#128308; A. 有对手未出单新品</h3>
      <div class="sub-module"><h4>按分析人</h4><div id="t4-has-an"></div></div>
      <div class="sub-module"><h4>按品线</h4><div id="t4-has-cat"></div></div>
    </div>
    <div class="section"><h3>&#128993; B. 无对手未出单新品</h3>
      <div class="sub-module"><h4>按分析人</h4><div id="t4-no-an"></div></div>
      <div class="sub-module"><h4>按品线</h4><div id="t4-no-cat"></div></div>
    </div>
    <div class="section"><h3>&#128065; 低占比新品明细</h3><div class="filter-bar-sticky" id="t4-lowshare-filters"></div><div id="t4-lowshare-table"></div></div>
  </div>

  <div class="tab-content" id="tab5">
    <div class="kpi-grid" id="t5-plp-kpi"></div>
    <div class="section"><h3>&#128176; PLP 核心指标</h3><div class="kpi-grid" id="t5-plp-core"></div></div>
    <div class="section"><h3>&#128269; PLP 维度分析</h3>
      <div class="sub-module"><h4>按分析人</h4><div id="t5-plp-an"></div></div>
      <div class="sub-module"><h4>按品线</h4><div id="t5-plp-cat"></div></div>
      <div class="sub-module"><h4>按拓展类型</h4><div id="t5-plp-exp"></div></div>
    </div>
    <div class="section"><h3>&#128200; PLG 费率统计</h3><div id="t5-plg"></div></div>
    <div class="section"><h3>&#128221; PLP 广告明细</h3><div id="t5-plp-detail"></div></div>
  </div>

  <div class="tab-content" id="tab6">
    <div class="kpi-grid" id="t6-kpi"></div>
    <div class="filter-bar-sticky" id="t6-filters"></div>
    <div class="section" style="padding-top:12px">
      <div id="t6-table"></div>
    </div>
  </div>

  <div class="tab-content" id="tab3">
    <div class="section"><h3>&#128202; 品效Cohort分析（按上架月份）</h3><div id="t3-cohort"></div></div>
    <div class="chart-grid">
      <div class="chart-box"><h4>&#128200; 月度销售额趋势</h4><canvas id="chart-px-rev"></canvas></div>
      <div class="chart-box"><h4>&#128200; 品效趋势（$/SKU）</h4><canvas id="chart-px-trend"></canvas></div>
    </div>
    <div class="section"><h3>&#128269; 品效 x 品线</h3><div id="t3-cat"></div></div>
    <div class="section"><h3>&#128269; 品效 x 拓展类型</h3><div id="t3-exp"></div></div>
  </div>

  <div class="tab-content" id="tab7">
    <div class="kpi-grid" id="t7-kpi"></div>
    <div class="section"><h3>&#9888;&#65039; 风险预警</h3><div id="t7-risk"></div></div>
    <div class="section"><h3>&#128270; 本周期主要发现</h3><div id="t7-findings"></div></div>
    <div class="section"><h3>&#127919; 下周重点动作</h3><div id="t7-actions"></div></div>
    <div class="section"><h3>&#128203; 可复制周报文案</h3><div id="t7-report"></div></div>
  </div>
  <div class="tab-content" id="tab2">
    <div class="kpi-grid" id="t2-kpi"></div>

    <div class="section"><h3>&#127758; 市场状态分布</h3>
      <div class="chart-grid">
        <div class="chart-box"><h4>&#128200; 本周市场状态占比</h4><canvas id="chart-mkt-ring" style="max-height:320px"></canvas></div>
        <div class="chart-box"><h4>&#128200; 本周vs上周各状态SKU数</h4><canvas id="chart-mkt-bar"></canvas></div>
      </div>
    </div>
    <div class="section"><h3>&#128203; 市场状态明细</h3><div id="t2-mkt-table"></div></div>

    <div class="section"><h3>&#128176; 货值分布（销售额/销量）</h3>
      <div class="chart-grid">
        <div class="chart-box"><h4>&#128200; 价格区间SKU数分布</h4><canvas id="chart-price-dist"></canvas></div>
        <div class="chart-box"><h4>&#128200; 按分析人-各价格区间SKU数</h4><canvas id="chart-price-an"></canvas></div>
      </div>
    </div>
    <div class="section"><h3>&#128203; 货值明细</h3><div id="t2-price-table"></div></div>

    <div class="section"><h3>&#127919; 市占比分布（品线 x 高中低）</h3>
      <div class="chart-grid">
        <div class="chart-box"><h4>&#128200; 各品线高中低市占比SKU分布</h4><canvas id="chart-share-tier"></canvas></div>
      </div>
    </div>
    <div class="section"><h3>&#128203; 市占比分布明细</h3><div id="t2-share-tier-table"></div></div>
  </div>
</div>
'''

JS_CODE = r'''
function fmtNum(n) { return n == null || n === '' ? '-' : Number(n).toLocaleString('zh-CN'); }
function fmtMoney(n) { return n == null || n === '' ? '-' : '$' + Number(n).toFixed(2); }
function pct(v) { return v == null || v === '' ? '-' : v; }
function acoasPct(v) {
  if (v == null || v === '' || v === 0) return '0.00%';
  if (typeof v === 'string') { var n = parseFloat(v); return isNaN(n) ? v : n.toFixed(2) + '%'; }
  return (v * 100).toFixed(2) + '%';
}
function hbSign(v) {
  if (typeof v !== 'string') return v;
  if (v === '-' || v === '0%' || v === '0.0%' || v === '+0%') return '<span class="hb-flat">-</span>';
  if (v.startsWith('+')) return '<span class="hb-up">' + v + '</span>';
  if (v.startsWith('-')) return '<span class="hb-down">' + v + '</span>';
  return '<span class="hb-flat">' + v + '</span>';
}
function badge(s, cls) { return '<span class="badge ' + cls + '">' + s + '</span>'; }
function badgeStatus(v) {
  if (v === '竞争无优势') return badge('竞争弱', 'badge-orange');
  if (v === '无市场') return badge('无市场', 'badge-red');
  if (v === '正常') return badge('正常', 'badge-blue');
  if (v === '站外出单') return badge('站外出单', 'badge-purple');
  if (v === '站内无价格优势') return badge('无价优', 'badge-orange');
  if (v === '未知') return badge('未知', 'badge-gray');
  if (v === '其他') return badge('其他', 'badge-gray');
  return v;
}
function badge8d(v) {
  if (v === 'Y') return badge('Y', 'badge-green');
  if (v === 'N') return badge('N', 'badge-orange');
  return badge(v, 'badge-red');
}
function badgeAdClass(v) {
  if (!v) return '';
  if (v.indexOf('单链接') === 0) return '<span style="color:#c0392b;font-weight:600">' + v + '</span>';
  if (v === '单PLG且未出单') return '<span style="color:#c0392b;font-weight:600">' + v + '</span>';
  if (v === 'PLP+PLG同开') return '<span style="color:#8e44ad;font-weight:600">' + v + '</span>';
  if (v === '单PLG') return '<span style="color:#e07b24;">' + v + '</span>';
  if (v === '单PLP') return '<span style="color:#2980b9;">' + v + '</span>';
  return v;
}

function switchTab(tabId, el) {
  document.querySelectorAll('.tab-content').forEach(function(t) { t.classList.remove('active'); });
  document.getElementById(tabId).classList.add('active');
  document.querySelectorAll('.sidebar a').forEach(function(a) { a.classList.remove('active'); });
  if (el) el.classList.add('active');
  if (tabId === 'tab1' && !window._charts1Init) { initCharts1(); }
  if (tabId === 'tab3' && !window._charts3Init) { initCharts3(); }
  if (tabId === 'tab2' && !window._charts2Init) { initCharts2(); }
}

(function() {
  var t = cum43Stats;
  var pk = prevWeekKpi;
  var saleRate = t.hasRivalCount ? (t.yCount + t.nCount) / t.hasRivalCount * 100 : 0;

  document.getElementById('t1-kpi').innerHTML =
    '<div class="kpi-card primary"><div class="label">累计SKU数</div><div class="val">' + t.total + '</div><div class="hb">' + hbSign(pk.skuChange) + ' 上周' + pk.prevTotalSku + '</div></div>' +
    '<div class="kpi-card success"><div class="label">总销量</div><div class="val">' + fmtNum(pk.prevTotalSalesQty) + '</div><div class="hb">' + hbSign(pk.salesQtyChange) + '</div></div>' +
    '<div class="kpi-card purple"><div class="label">总销售额</div><div class="val">' + fmtMoney(pk.prevTotalRevenue) + '</div><div class="hb">' + hbSign(pk.revenueChange) + '</div></div>' +
    '<div class="kpi-card info"><div class="label">新品销售占比</div><div class="val">' + pk.deptRatio + '</div><div class="hb">部门总销售额 ' + fmtMoney(pk.deptTotalRevenue) + '</div></div>' +
    '<div class="kpi-card success"><div class="label">出单率</div><div class="val">' + saleRate.toFixed(1) + '%</div><div class="hb">Y:' + t.yCount + ' N:' + t.nCount + ' 未:' + t.unCount + '</div></div>' +
    '<div class="kpi-card info"><div class="label">分析及时率</div><div class="val">' + timelinessData.total.timelyRate + '</div><div class="hb">' + hbSign(timelinessData.total.change) + '</div></div>';

  var ordHtml = '<table class="data-table"><thead><tr><th>指标</th><th>本周(5.20)</th></tr></thead><tbody>' +
    '<tr><td>有对手新品总SKU</td><td>' + t.hasRivalCount + '</td></tr>' +
    '<tr><td>8日内出单(Y)</td><td>' + t.yCount + '</td></tr>' +
    '<tr><td>8日外出单(N)</td><td>' + t.nCount + '</td></tr>' +
    '<tr><td>真正未出单</td><td>' + t.unCount + '</td></tr>' +
    '<tr class="total-row"><td>已出单合计(Y+N)</td><td>' + (t.yCount+t.nCount) + '</td></tr>' +
    '<tr class="total-row"><td>出单率</td><td>' + saleRate.toFixed(1) + '%</td></tr>' +
    '</tbody></table>';
  document.getElementById('t1-ord8').innerHTML = ordHtml;

  var catHtml = '<table class="data-table"><thead><tr><th>品线</th><th>SKU数</th><th>本周新上架</th><th>本周销量</th><th>上周销量</th><th>销量环比</th><th>本周销售额</th><th>上周销售额</th><th>销售额环比</th></tr></thead><tbody>';
  categoryRevenueData.forEach(function(d) {
    catHtml += '<tr><td>' + d.category + '</td><td>' + d.curSku + '</td><td>' + d.curNewSku + '</td><td>' + fmtNum(d.curSalesQty) + '</td><td>' + fmtNum(d.prevSalesQty) + '</td><td>' + hbSign(d.salesQtyChange) + '</td><td>' + fmtMoney(d.curRevenue) + '</td><td>' + fmtMoney(d.prevRevenue) + '</td><td>' + hbSign(d.revenueChange) + '</td></tr>';
  });
  catHtml += '</tbody></table>';
  document.getElementById('t1-cat-table').innerHTML = catHtml;

  var anHtml = '<table class="data-table"><thead><tr><th>分析人</th><th>SKU数</th><th>本周新上架</th><th>本周销量</th><th>上周销量</th><th>销量环比</th><th>本周销售额</th><th>上周销售额</th><th>销售额环比</th></tr></thead><tbody>';
  analystRevenueData.forEach(function(d) {
    anHtml += '<tr><td>' + d.analyst + '</td><td>' + d.curSku + '</td><td>' + d.curNewSku + '</td><td>' + fmtNum(d.curSalesQty) + '</td><td>' + fmtNum(d.prevSalesQty) + '</td><td>' + hbSign(d.salesQtyChange) + '</td><td>' + fmtMoney(d.curRevenue) + '</td><td>' + fmtMoney(d.prevRevenue) + '</td><td>' + hbSign(d.revenueChange) + '</td></tr>';
  });
  anHtml += '</tbody></table>';
  document.getElementById('t1-an-table').innerHTML = anHtml;

  var expHtml = '<table class="data-table"><thead><tr><th>拓展类型</th><th>本周SKU</th><th>上周SKU</th><th>出单SKU</th><th>出单率</th><th>上周出单率</th><th>本周销量</th><th>上周销量</th><th>销量环比</th><th>本周销售额</th><th>上周销售额</th><th>销售额环比</th></tr></thead><tbody>';
  expandTypeData.forEach(function(d) {
    expHtml += '<tr><td>' + d.expandType + '</td><td>' + d.curSku + '</td><td>' + d.prevSku + '</td><td>' + d.curSalesSku + '</td><td>' + d.curSalesRate + '</td><td>' + d.prevSalesRate + '</td><td>' + fmtNum(d.curSalesQty) + '</td><td>' + fmtNum(d.prevSalesQty) + '</td><td>' + hbSign(d.salesQtyChange) + '</td><td>' + fmtMoney(d.curRevenue) + '</td><td>' + fmtMoney(d.prevRevenue) + '</td><td>' + hbSign(d.revenueChange) + '</td></tr>';
  });
  expHtml += '</tbody></table>';
  document.getElementById('t1-exp-table').innerHTML = expHtml;

  var timeHtml = '<table class="data-table"><thead><tr><th>分析人</th><th>本周SKU</th><th>及时分析</th><th>8日未分析</th><th>7日未分析</th><th>及时率</th><th>上周及时率</th><th>变化</th></tr></thead><tbody>';
  timelinessData.analysts.forEach(function(d) {
    timeHtml += '<tr><td>' + d.analyst + '</td><td>' + d.curSku + '</td><td>' + d.timelyCount + '</td><td>' + d.noAnalysis8dCount + '</td><td>' + d.noAnalysis7dCount + '</td><td>' + d.timelyRate + '</td><td>' + d.prevTimelyRate + '</td><td>' + hbSign(d.change) + '</td></tr>';
  });
  var td = timelinessData.total;
  timeHtml += '<tr class="total-row"><td>' + td.analyst + '</td><td>' + td.curSku + '</td><td>' + td.timelyCount + '</td><td>' + td.noAnalysis8dCount + '</td><td>' + td.noAnalysis7dCount + '</td><td>' + td.timelyRate + '</td><td>' + td.prevTimelyRate + '</td><td>' + hbSign(td.change) + '</td></tr>';
  timeHtml += '</tbody></table>';
  document.getElementById('t1-time-table').innerHTML = timeHtml;
})();

window._charts1Init = false;
function initCharts1() {
  if (window._charts1Init) return;
  window._charts1Init = true;
  var t = cum43Stats;

  new Chart(document.getElementById('chart-ord8'), {
    type: 'doughnut',
    data: {
      labels: ['有对手已出单', '有对手未出单', '无对手SKU'],
      datasets: [{ data: [t.yCount + t.nCount, t.unCount, t.total - t.hasRivalCount], backgroundColor: ['#08845a', '#e07b24', '#c0392b'] }]
    },
    options: { responsive: true, plugins: { legend: { position: 'bottom' } } }
  });

  var catLabels = categoryRevenueData.map(function(d) { return d.category; });
  new Chart(document.getElementById('chart-cat-sales'), {
    type: 'bar', data: { labels: catLabels, datasets: [
      { label: '本周销量', data: categoryRevenueData.map(function(d){return d.curSalesQty;}), backgroundColor: '#0f3460' },
      { label: '上周销量', data: categoryRevenueData.map(function(d){return d.prevSalesQty;}), backgroundColor: '#ccc' }
    ]},
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true } } }
  });
  new Chart(document.getElementById('chart-cat-rev'), {
    type: 'bar', data: { labels: catLabels, datasets: [
      { label: '本周销售额', data: categoryRevenueData.map(function(d){return d.curRevenue;}), backgroundColor: '#8e44ad' },
      { label: '上周销售额', data: categoryRevenueData.map(function(d){return d.prevRevenue;}), backgroundColor: '#ddd' }
    ]},
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true } } }
  });

  var anLabels = analystRevenueData.map(function(d) { return d.analyst; });
  new Chart(document.getElementById('chart-an-sales'), {
    type: 'bar', data: { labels: anLabels, datasets: [
      { label: '本周销量', data: analystRevenueData.map(function(d){return d.curSalesQty;}), backgroundColor: '#0f3460' },
      { label: '上周销量', data: analystRevenueData.map(function(d){return d.prevSalesQty;}), backgroundColor: '#ccc' }
    ]},
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true } } }
  });
  new Chart(document.getElementById('chart-an-rev'), {
    type: 'bar', data: { labels: anLabels, datasets: [
      { label: '本周销售额', data: analystRevenueData.map(function(d){return d.curRevenue;}), backgroundColor: '#8e44ad' },
      { label: '上周销售额', data: analystRevenueData.map(function(d){return d.prevRevenue;}), backgroundColor: '#ddd' }
    ]},
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true } } }
  });
  new Chart(document.getElementById('chart-an-timely'), {
    type: 'bar', data: { labels: anLabels, datasets: [
      { label: '及时率(%)', data: timelinessData.analysts.map(function(d){return parseFloat(d.timelyRate)||0;}), backgroundColor: '#2980b9' }
    ]},
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true, max: 100 } } }
  });
}
setTimeout(initCharts1, 100);

(function() {
  var hcu = hasCompetitorUnsold;
  var unc = unsoldNoCompetitor;
  var t = cum43Stats;

  document.getElementById('t4-kpi').innerHTML =
    '<div class="kpi-card warning"><div class="label">有对手未出单</div><div class="val">' + hcu.total + '</div><div class="hb">上周 ' + hcu.prevTotal + ' | ' + (hcu.change >= 0 ? '+' : '') + hcu.change + '</div></div>' +
    '<div class="kpi-card danger"><div class="label">无对手未出单</div><div class="val">' + unc.total + '</div><div class="hb">上周 ' + unc.prevTotal + ' | ' + (unc.change >= 0 ? '+' : '') + unc.change + '</div></div>' +
    '<div class="kpi-card info"><div class="label">有对手SKU总数</div><div class="val">' + t.hasRivalCount + '</div></div>' +
    '<div class="kpi-card primary"><div class="label">有对手未出单占比</div><div class="val">' + (t.unCount/t.total*100).toFixed(1) + '%</div><div class="hb">' + t.unCount + ' / ' + t.total + '</div></div>';

  var hasMkts = ['竞争无优势', '无市场', '站内无价格优势', '站外出单', '正常', '#N/A', '未知'];
  var hasAnHtml = '<table class="data-table"><thead><tr><th>分析人</th>';
  hasMkts.forEach(function(m) { hasAnHtml += '<th>' + m + '</th>'; });
  hasAnHtml += '<th>未出单总数</th></tr></thead><tbody>';
  hcu.byAnalyst.forEach(function(d) {
    hasAnHtml += '<tr><td>' + d.analyst + '</td>';
    hasMkts.forEach(function(m) { hasAnHtml += '<td>' + (d[m] || 0) + '</td>'; });
    hasAnHtml += '<td><b>' + d.total + '</b></td></tr>';
  });
  hasAnHtml += '</tbody></table>';
  document.getElementById('t4-has-an').innerHTML = hasAnHtml;

  var hasCatHtml = '<table class="data-table"><thead><tr><th>品线</th>';
  hasMkts.forEach(function(m) { hasCatHtml += '<th>' + m + '</th>'; });
  hasCatHtml += '<th>未出单总数</th></tr></thead><tbody>';
  hcu.byCategory.forEach(function(d) {
    hasCatHtml += '<tr><td>' + d.category + '</td>';
    hasMkts.forEach(function(m) { hasCatHtml += '<td>' + (d[m] || 0) + '</td>'; });
    hasCatHtml += '<td><b>' + d.total + '</b></td></tr>';
  });
  hasCatHtml += '</tbody></table>';
  document.getElementById('t4-has-cat').innerHTML = hasCatHtml;

  var noMkts = ['无市场', '未知', '竞争无优势', '#N/A', '其他'];
  var noAnHtml = '<table class="data-table"><thead><tr><th>分析人</th>';
  noMkts.forEach(function(m) { noAnHtml += '<th>' + m + '</th>'; });
  noAnHtml += '<th>未出单总数</th></tr></thead><tbody>';
  unc.byAnalyst.forEach(function(d) {
    noAnHtml += '<tr><td>' + d.analyst + '</td>';
    noMkts.forEach(function(m) { noAnHtml += '<td>' + (d[m] || 0) + '</td>'; });
    noAnHtml += '<td><b>' + d.total + '</b></td></tr>';
  });
  noAnHtml += '</tbody></table>';
  document.getElementById('t4-no-an').innerHTML = noAnHtml;

  var noCatHtml = '<table class="data-table"><thead><tr><th>品线</th>';
  noMkts.forEach(function(m) { noCatHtml += '<th>' + m + '</th>'; });
  noCatHtml += '<th>未出单总数</th></tr></thead><tbody>';
  unc.byCategory.forEach(function(d) {
    noCatHtml += '<tr><td>' + d.category + '</td>';
    noMkts.forEach(function(m) { noCatHtml += '<td>' + (d[m] || 0) + '</td>'; });
    noCatHtml += '<td><b>' + d.total + '</b></td></tr>';
  });
  noCatHtml += '</tbody></table>';
  document.getElementById('t4-no-cat').innerHTML = noCatHtml;

  // Build lowShare filter bar
  var ls8dOpts = ['', 'Y', 'N', '未出单'];
  var ls8dNames = ['全部', 'Y', 'N', '未出单'];
  var lsOpVals = {};
  lowShareData.forEach(function(d) { var v = d.curOperation || '-'; lsOpVals[v] = 1; });
  var lsOpKeys = Object.keys(lsOpVals).sort();
  var lsAdVals = {};
  lowShareData.forEach(function(d) { lsAdVals[d.adClass] = 1; });
  var lsAdKeys = Object.keys(lsAdVals).sort();

  // Toggle checkbox dropdown
  window.toggleLsOpDropdown = function() {
    var dd = document.getElementById('ls-op-drop');
    dd.style.display = dd.style.display === 'none' ? 'block' : 'none';
  };
  // Close dropdown when clicking outside
  document.addEventListener('click', function(e) {
    var dd = document.getElementById('ls-op-drop');
    var lbl = document.getElementById('ls-op-label');
    if (dd && lbl && !lbl.contains(e.target) && !dd.contains(e.target)) {
      dd.style.display = 'none';
    }
  });

  var lsFilHtml = '<span class="fg"><label>8日出单</label><select id="ls-f-8d" onchange="renderLowShareTable()">';
  for (var i = 0; i < ls8dOpts.length; i++) lsFilHtml += '<option value="' + ls8dOpts[i] + '">' + ls8dNames[i] + '</option>';
  lsFilHtml += '</select></span>';
  lsFilHtml += '<span class="fg" style="position:relative">';
  lsFilHtml += '<label id="ls-op-label" style="cursor:pointer" onclick="toggleLsOpDropdown()">本期运作判断 &#9662;</label>';
  lsFilHtml += '<div id="ls-op-drop" style="display:none;position:absolute;top:100%;left:0;z-index:200;background:#fff;border:1px solid #ddd;border-radius:6px;padding:6px 0;max-height:220px;overflow-y:auto;min-width:150px;box-shadow:0 4px 12px rgba(0,0,0,0.12);">';
  lsFilHtml += '<label style="display:block;padding:4px 12px;font-size:12px;cursor:pointer;white-space:nowrap;"><input type="checkbox" id="ls-op-all" onchange="var c=this.checked;var boxes=document.querySelectorAll(&quot;.ls-op-cb&quot;);boxes.forEach(function(b){b.checked=c;});renderLowShareTable();" checked> <b>全选</b></label>';
  lsFilHtml += '<hr style="margin:4px 0;border-color:#eee">';
  lsOpKeys.forEach(function(v) {
    lsFilHtml += '<label style="display:block;padding:4px 12px;font-size:12px;cursor:pointer;white-space:nowrap;"><input type="checkbox" class="ls-op-cb" value="' + v + '" onchange="renderLowShareTable()" checked> ' + v + '</label>';
  });
  lsFilHtml += '</div></span>';
  lsFilHtml += '<span class="fg"><label>广告分类</label><select id="ls-f-ad" onchange="renderLowShareTable()"><option value="">全部</option>';
  lsAdKeys.forEach(function(v) { lsFilHtml += '<option value="' + v + '">' + v + '</option>'; });
  lsFilHtml += '</select></span>';
  lsFilHtml += '<button class="reset-btn" onclick="resetLowShareFilters()">重置筛选</button>';
  lsFilHtml += '<span class="count" id="t4-ls-count"></span>';
  document.getElementById('t4-lowshare-filters').innerHTML = lsFilHtml;

  window.resetLowShareFilters = function() {
    document.getElementById("ls-f-8d").value = "";
    document.querySelectorAll(".ls-op-cb").forEach(function(b) { b.checked = true; });
    document.getElementById("ls-op-all").checked = true;
    document.getElementById("ls-f-ad").value = "";
    window.renderLowShareTable();
  };

  window.renderLowShareTable = function() {
    var f8d = document.getElementById('ls-f-8d').value;
    var fOpChecked = document.querySelectorAll('.ls-op-cb:checked');
    var fOpAll = fOpChecked.length === document.querySelectorAll('.ls-op-cb').length;
    var fAd = document.getElementById('ls-f-ad').value;
    var filtered = lowShareData.filter(function(d) {
      if (f8d && d.cur8dStatus !== f8d) return false;
      if (!fOpAll && !Array.from(fOpChecked).some(function(cb) { return cb.value === (d.curOperation || '-'); })) return false;
      if (fAd && d.adClass !== fAd) return false;
      return true;
    });
    var lsTotalSales = 0, lsTotalRev = 0;
    var h = '<div class="table-scroll-wrap"><table class="data-table"><thead><tr><th>SKU</th><th>上架日期</th><th>分析人</th><th>品类</th><th>本周销量</th><th>销量环比</th><th>本周销售额</th><th>对手量</th><th>市占比</th><th>8日出单</th><th>上期市场状态</th><th>本期运作判断</th><th>本期市场状态</th><th>广告分类</th></tr></thead><tbody>';
    filtered.forEach(function(d) {
      h += '<tr><td>' + d.SKU + '</td><td>' + d.launchDate + '</td><td>' + d.analyst + '</td><td>' + d.category + '</td>';
      h += '<td>' + d.curSalesQty + '</td><td>' + hbSign(d.salesQtyChange) + '</td><td>' + fmtMoney(d.curRevenue) + '</td>';
      h += '<td>' + d.curRivalQty + '</td><td>' + d.curMarketShare + '%</td>';
      h += '<td>' + badge8d(d.cur8dStatus) + '</td>';
      h += '<td>' + badgeStatus(d.prevMarketStatus) + '</td><td>' + (d.curOperation || '-') + '</td><td>' + badgeStatus(d.curMarketStatus) + '</td>';
      h += '<td>' + badgeAdClass(d.adClass) + '</td></tr>';
      lsTotalSales += d.curSalesQty;
      lsTotalRev += (d.curRevenue || 0);
    });
    h += '</tbody><tfoot><tr><td colspan="2">合计（' + filtered.length + '条）</td><td></td><td></td><td>' + lsTotalSales + '</td><td></td><td>' + fmtMoney(lsTotalRev) + '</td><td colspan="7"></td></tr></tfoot></table></div>';
    document.getElementById('t4-lowshare-table').innerHTML = h;
    document.getElementById('t4-ls-count').textContent = '筛选结果:' + filtered.length + ' / ' + lowShareData.length + ' 条';
  };
  renderLowShareTable();
})();

(function() {
  var pt = plpTotal;
  var pp = plpPrevTotal;

  document.getElementById('t5-plp-kpi').innerHTML =
    '<div class="kpi-card primary"><div class="label">广告活动数</div><div class="val">' + pt.campaignCount + '</div><div class="hb">上周 ' + pp.campaignCount + '</div></div>' +
    '<div class="kpi-card info"><div class="label">投放链接数</div><div class="val">' + pt.linkCount + '</div><div class="hb">上周 ' + pp.linkCount + '</div></div>' +
    '<div class="kpi-card primary"><div class="label">曝光量</div><div class="val">' + fmtNum(pt.impression) + '</div></div>' +
    '<div class="kpi-card info"><div class="label">点击量</div><div class="val">' + fmtNum(pt.click) + '</div></div>' +
    '<div class="kpi-card success"><div class="label">售出数</div><div class="val">' + pt.sold + '</div></div>' +
    '<div class="kpi-card purple"><div class="label">广告销售额</div><div class="val">' + fmtMoney(pt.revenue) + '</div></div>';

  document.getElementById('t5-plp-core').innerHTML =
    '<div class="kpi-card primary"><div class="label">ROAS</div><div class="val">' + pt.roas + '</div><div class="hb">上周 ' + pp.roas + '</div></div>' +
    '<div class="kpi-card info"><div class="label">CVR</div><div class="val">' + pt.cvr + '</div><div class="hb">上周 ' + pp.cvr + '</div></div>' +
    '<div class="kpi-card primary"><div class="label">CTR</div><div class="val">' + pt.ctr + '</div><div class="hb">上周 ' + pp.ctr + '</div></div>' +
    '<div class="kpi-card info"><div class="label">CPC</div><div class="val">' + pt.cpc + '</div><div class="hb">上周 ' + pp.cpc + '</div></div>' +
    '<div class="kpi-card warning"><div class="label">CPA</div><div class="val">' + pt.cpa + '</div><div class="hb">上周 ' + pp.cpa + '</div></div>' +
    '<div class="kpi-card danger"><div class="label">ACOS</div><div class="val">' + pt.acos + '</div><div class="hb">上周 ' + pp.acos + '</div></div>' +
    '<div class="kpi-card danger"><div class="label">ACOAS</div><div class="val">' + pt.acoas + '</div><div class="hb">上周 ' + pp.acoas + '</div></div>';

  function renderPlpDim(data, labelKey) {
    var h = '<table class="data-table"><thead><tr><th>' + labelKey + '</th><th>活动数</th><th>链接数</th><th>曝光量</th><th>点击量</th><th>售出数</th><th>花费</th><th>广告销售额</th><th>ROAS</th><th>CVR</th><th>CTR</th><th>CPC</th><th>CPA</th><th>ACOS</th><th>ACOAS</th></tr></thead><tbody>';
    data.forEach(function(d) {
      h += '<tr><td>' + d.name + '</td><td>' + d.campaignCount + '</td><td>' + d.linkCount + '</td>';
      h += '<td>' + fmtNum(d.impression) + '</td><td>' + fmtNum(d.click) + '</td><td>' + d.sold + '</td>';
      h += '<td>' + fmtMoney(d.cost) + '</td><td>' + fmtMoney(d.revenue) + '</td>';
      h += '<td>' + d.roas + '</td><td>' + d.cvr + '</td><td>' + d.ctr + '</td>';
      h += '<td>' + d.cpc + '</td><td>' + d.cpa + '</td><td>' + d.acos + '</td><td>' + d.acoas + '</td></tr>';
    });
    h += '</tbody></table>';
    return h;
  }
  document.getElementById('t5-plp-an').innerHTML = renderPlpDim(plpAnalysts, '分析人');
  document.getElementById('t5-plp-cat').innerHTML = renderPlpDim(plpCategories, '品线');
  document.getElementById('t5-plp-exp').innerHTML = renderPlpDim(plpExpandTypes, '拓展类型');

  var pg = plgStats;
  var plgHtml = '<div class="kpi-grid" style="margin-bottom:12px">';
  plgHtml += '<div class="kpi-card primary"><div class="label">新品总数</div><div class="val">' + pg.totalNewProducts + '</div></div>';
  plgHtml += '<div class="kpi-card purple"><div class="label">PLP+PLG同开</div><div class="val">' + pg.plpAndPlgBothCount + '</div></div>';
  plgHtml += '<div class="kpi-card danger"><div class="label">单链接PLP+PLG同开</div><div class="val">' + pg.singleLinkPlpPlgCount + '</div></div>';
  plgHtml += '<div class="kpi-card info"><div class="label">单PLG</div><div class="val">' + pg.plgOnlyCount + '</div></div>';
  plgHtml += '<div class="kpi-card primary"><div class="label">单PLP</div><div class="val">' + pg.plpOnlyCount + '</div></div>';
  plgHtml += '<div class="kpi-card warning"><div class="label">无广告</div><div class="val">' + pg.noAdCount + '</div></div>';
  plgHtml += '<div class="kpi-card danger"><div class="label">单PLG且未出单</div><div class="val">' + (pg.plpDisabledNoSaleCount || 0) + '</div></div>';
  plgHtml += '</div>';
  plgHtml += '<table class="data-table"><thead><tr><th>分析人</th><th>总数</th><th>PLP+PLG同开</th><th>单链接PLP+PLG同开</th><th>单PLG</th><th>单PLP</th><th>无广告</th><th>PLP未开未出单</th></tr></thead><tbody>';
  pg.byAnalyst.forEach(function(d) {
    plgHtml += '<tr><td>' + d.analyst + '</td><td>' + d.total + '</td>';
    plgHtml += '<td>' + d.plpAndPlgBoth + '</td>';
    plgHtml += '<td style="color:#c0392b;font-weight:600">' + d.singleLinkPlpPlg + '</td>';
    plgHtml += '<td>' + d.plgOnly + '</td><td>' + d.plpOnly + '</td><td>' + d.noAd + '</td>';
    plgHtml += '<td style="color:#c0392b;font-weight:600">' + d.plpDisabledNoSale + '</td></tr>';
  });
  plgHtml += '</tbody></table>';
  document.getElementById('t5-plg').innerHTML = plgHtml;

  var detHtml = '<div class="table-scroll-wrap"><table class="data-table"><thead><tr><th>SKU</th><th>广告活动</th><th>分析人</th><th>品类</th><th>曝光</th><th>点击</th><th>售出</th><th>花费</th><th>广告销售额</th><th>总销售额</th><th>ROAS</th><th>ACOS</th><th>ACOAS</th><th>广告分类</th></tr></thead><tbody>';
  var detTotalImpr=0, detTotalClick=0, detTotalSold=0, detTotalCost=0, detTotalAdRev=0, detTotalRev=0;
  plpDetailData.forEach(function(d) {
    detHtml += '<tr><td>' + d.SKU + '</td><td>' + d.campaign + '</td><td>' + d.analyst + '</td><td>' + d.category + '</td>';
    detHtml += '<td>' + fmtNum(d.impressions) + '</td><td>' + fmtNum(d.clicks) + '</td><td>' + d.salesQty + '</td>';
    detHtml += '<td>' + fmtMoney(d.spend) + '</td><td>' + fmtMoney(d.adRevenue) + '</td><td>' + fmtMoney(d.totalRevenue) + '</td>';
    detHtml += '<td>' + (d.roas ? d.roas.toFixed(2) : '-') + '</td>';
    detHtml += '<td>' + (d.acos ? (d.acos*100).toFixed(2)+'%' : '0%') + '</td>';
    detHtml += '<td>' + acoasPct(d.acoas) + '</td>';
    detHtml += '<td>' + badgeAdClass(d.adClass) + '</td></tr>';
    detTotalImpr += (d.impressions || 0); detTotalClick += (d.clicks || 0); detTotalSold += (d.salesQty || 0);
    detTotalCost += (d.spend || 0); detTotalAdRev += (d.adRevenue || 0); detTotalRev += (d.totalRevenue || 0);
  });
  detHtml += '</tbody><tfoot><tr><td colspan="2">合计（' + plpDetailData.length + '条）</td><td></td><td></td><td>' + fmtNum(detTotalImpr) + '</td><td>' + fmtNum(detTotalClick) + '</td><td>' + detTotalSold + '</td><td>' + fmtMoney(detTotalCost) + '</td><td>' + fmtMoney(detTotalAdRev) + '</td><td>' + fmtMoney(detTotalRev) + '</td><td></td><td></td><td></td><td></td></tr></tfoot></table></div>';
  document.getElementById('t5-plp-detail').innerHTML = detHtml;
})();

(function() {
  var t = cum43Stats;
  document.getElementById('t6-kpi').innerHTML =
    '<div class="kpi-card primary"><div class="label">累计总SKU</div><div class="val">' + t.total + '</div></div>' +
    '<div class="kpi-card success"><div class="label">已出单(Y+N)</div><div class="val">' + (t.yCount + t.nCount) + '</div></div>' +
    '<div class="kpi-card warning"><div class="label">有对手未出单</div><div class="val">' + t.unCount + '</div></div>' +
    '<div class="kpi-card info"><div class="label">市场正常</div><div class="val">' + t.normalCount + '</div></div>' +
    '<div class="kpi-card warning"><div class="label">竞争无优势</div><div class="val">' + t.competitiveCount + '</div></div>' +
    '<div class="kpi-card danger"><div class="label">无市场</div><div class="val">' + t.noMarketCount + '</div></div>' +
    '<div class="kpi-card purple"><div class="label">站外出单</div><div class="val">' + (t.stationOutCount || 0) + '</div></div>';

  var uniqAnalysts = [];
  var seenAn = {};
  cum43Data.forEach(function(d) { if (!seenAn[d.analyst]) { seenAn[d.analyst]=1; uniqAnalysts.push(d.analyst); } });
  uniqAnalysts.sort();
  var uniqCats = [];
  var seenCat = {};
  cum43Data.forEach(function(d) { if (!seenCat[d.category]) { seenCat[d.category]=1; uniqCats.push(d.category); } });
  uniqCats.sort();

  var filHtml = '<span class="fg"><label>市场状态</label><select id="f-mkt" onchange="applyFilters()"><option value="">全部</option><option>正常</option><option>竞争无优势</option><option>无市场</option></select></span>';
  filHtml += '<span class="fg"><label>分析人</label><select id="f-an" onchange="applyFilters()"><option value="">全部</option>';
  uniqAnalysts.forEach(function(a) { filHtml += '<option>' + a + '</option>'; });
  filHtml += '</select></span>';
  filHtml += '<span class="fg"><label>品类</label><select id="f-cat" onchange="applyFilters()"><option value="">全部</option>';
  uniqCats.forEach(function(c) { filHtml += '<option>' + c + '</option>'; });
  filHtml += '</select></span>';
  filHtml += '<span class="fg"><label>拓展类型</label><select id="f-exp" onchange="applyFilters()"><option value="">全部</option><option>原开品</option><option>拓展品</option><option>组合件</option></select></span>';
  filHtml += '<span class="fg"><label>8日出单</label><select id="f-8d" onchange="applyFilters()"><option value="">全部</option><option>Y</option><option>N</option><option>未出单</option></select></span>';
  filHtml += '<span class="fg"><label>市占比</label><select id="f-share" onchange="applyFilters()"><option value="">全部</option><option value="high">75%及以上</option><option value="mid">50%-75%</option><option value="low">50%以下</option></select></span>';
  filHtml += '<span class="fg"><label>广告条件</label><select id="f-ad" onchange="applyFilters()"><option value="">全部</option><option>PLP+PLG同开</option><option>单链接PLP+PLG同开</option><option>单PLG</option><option>单PLP</option><option>单PLG且未出单</option><option>无广告</option></select></span>';
  filHtml += '<button class="reset-btn" onclick="resetFilters()">重置筛选</button>';
  filHtml += '<span class="count" id="t6-count"></span>';
  document.getElementById('t6-filters').innerHTML = filHtml;

	  window.renderT4Table = function(data) {
	    var t4TotalSales=0, t4TotalRev=0, t4TotalRival=0;
	    var h = '<div class="table-scroll-wrap"><table class="data-table"><thead><tr><th>SKU</th><th>上架日期</th><th>首次出单</th><th>分析人</th><th>品类</th><th>拓展类型</th><th>本周销量</th><th>本周销售额</th><th>对手量</th><th>市占比</th><th>PLG费率</th><th>市场状态</th><th>8日出单</th><th>广告分类</th></tr></thead><tbody>';
	    data.forEach(function(d) {
	      h += '<tr><td>' + d.SKU + '</td><td>' + d.listDate + '</td><td>' + d.firstOrderDate + '</td>';
	      h += '<td>' + d.analyst + '</td><td>' + d.category + '</td><td>' + d.expandType + '</td>';
	      h += '<td>' + d.curSalesQty + '</td><td>' + fmtMoney(d.curRevenue) + '</td>';
	      h += '<td>' + d.curRivalQty + '</td><td>' + d.curMarketShare + '%</td>';
	      h += '<td>' + (d.plgFee || '0%') + '</td>';
	      h += '<td>' + badgeStatus(d.curMarketStatus) + '</td><td>' + badge8d(d.cur8dStatus) + '</td>';
	      h += '<td>' + badgeAdClass(d.adClass) + '</td></tr>';
	      t4TotalSales += d.curSalesQty; t4TotalRev += (d.curRevenue || 0); t4TotalRival += d.curRivalQty;
	    });
	    h += '</tbody><tfoot><tr><td colspan="2">合计（' + data.length + '条）</td><td></td><td></td><td></td><td></td><td>' + t4TotalSales + '</td><td>' + fmtMoney(t4TotalRev) + '</td><td>' + t4TotalRival + '</td><td colspan="5"></td></tr></tfoot></table></div>';
	    document.getElementById('t6-table').innerHTML = h;
	    document.getElementById('t6-count').textContent = '筛选结果:' + data.length + ' / ' + cum43Data.length + ' 条';
	  };

  window.applyFilters = function() {
    var data = cum43Data.slice();
    var mkt = document.getElementById('f-mkt').value;
    var an = document.getElementById('f-an').value;
    var cat = document.getElementById('f-cat').value;
    var exp = document.getElementById('f-exp').value;
    var d8 = document.getElementById('f-8d').value;
    var share = document.getElementById('f-share').value;
    var ad = document.getElementById('f-ad').value;
    if (mkt) data = data.filter(function(d) { return d.curMarketStatus === mkt; });
    if (an) data = data.filter(function(d) { return d.analyst === an; });
    if (cat) data = data.filter(function(d) { return d.category === cat; });
    if (exp) data = data.filter(function(d) { return d.expandType === exp; });
    if (d8) data = data.filter(function(d) { return d.cur8dStatus === d8; });
    if (share === 'high') data = data.filter(function(d) { return d.curMarketShare >= 75; });
    else if (share === 'mid') data = data.filter(function(d) { return d.curMarketShare >= 50 && d.curMarketShare < 75; });
    else if (share === 'low') data = data.filter(function(d) { return d.curMarketShare < 50; });
    if (ad) data = data.filter(function(d) { return d.adClass === ad; });
    renderT4Table(data);
  };

  window.resetFilters = function() {
    document.querySelectorAll('#t4-filters select').forEach(function(s) { s.value = ''; });
    renderT4Table(cum43Data);
  };

  renderT4Table(cum43Data);
})();

(function() {
  var cohortHtml = '<table class="data-table pinxiao-table"><thead><tr><th>上架月份</th><th>SKU个数</th>';
  for (var m = 1; m <= 5; m++) cohortHtml += '<th>' + m + '月销售额</th>';
  for (var m = 1; m <= 5; m++) cohortHtml += '<th>' + m + '月品效($/SKU)</th>';
  cohortHtml += '<th>趋势</th></tr></thead><tbody>';

  pinxiaoData.forEach(function(d) {
    cohortHtml += '<tr><td><b>' + d.launchMonth + '月上架</b></td><td>' + d.skuCount + '</td>';
    d.monthlyRevenue.forEach(function(rev) { cohortHtml += '<td>' + fmtMoney(rev) + '</td>'; });
    d.monthlyPinxiao.forEach(function(px) {
      var cls = px >= 500 ? 'px-high' : (px >= 100 ? 'px-mid' : 'px-low');
      cohortHtml += '<td class="' + cls + '">' + (px > 0 ? '$' + px.toFixed(0) : '-') + '</td>';
    });
    var pxVals = d.monthlyPinxiao.filter(function(v) { return v > 0; });
    var trend = '-';
    if (pxVals.length >= 3) {
      if (pxVals[pxVals.length-1] > pxVals[pxVals.length-2] && pxVals[pxVals.length-2] > pxVals[pxVals.length-3]) trend = '<span class="trend-up">攀升</span>';
      else if (pxVals[pxVals.length-1] < pxVals[pxVals.length-2] && pxVals[pxVals.length-2] < pxVals[pxVals.length-3]) trend = '<span class="trend-down">衰减</span>';
      else trend = '<span class="trend-flat">平稳</span>';
    } else if (pxVals.length >= 2) {
      trend = pxVals[pxVals.length-1] > pxVals[pxVals.length-2] ? '<span class="trend-up">攀升</span>' : '<span class="trend-down">衰减</span>';
    }
    cohortHtml += '<td>' + trend + '</td></tr>';
  });
  cohortHtml += '</tbody></table>';
  document.getElementById('t3-cohort').innerHTML = cohortHtml;

  var catPxHtml = '<table class="data-table pinxiao-table"><thead><tr><th>品线</th>';
  launchMonths.forEach(function(m) { catPxHtml += '<th>' + m + '月上架品效</th>'; });
  catPxHtml += '</tr></thead><tbody>';
  pinxiaoCatData.forEach(function(d) {
    catPxHtml += '<tr><td>' + d.category + '</td>';
    d.pinxiao.forEach(function(px) {
      var cls = px >= 500 ? 'px-high' : (px >= 100 ? 'px-mid' : 'px-low');
      catPxHtml += '<td class="' + cls + '">' + (px > 0 ? '$' + px : '-') + '</td>';
    });
    catPxHtml += '</tr>';
  });
  catPxHtml += '</tbody></table>';
  document.getElementById('t3-cat').innerHTML = catPxHtml;

  var expPxHtml = '<table class="data-table pinxiao-table"><thead><tr><th>拓展类型</th>';
  launchMonths.forEach(function(m) { expPxHtml += '<th>' + m + '月上架品效</th>'; });
  expPxHtml += '</tr></thead><tbody>';
  pinxiaoExpData.forEach(function(d) {
    expPxHtml += '<tr><td>' + d.expandType + '</td>';
    d.pinxiao.forEach(function(px) {
      var cls = px >= 500 ? 'px-high' : (px >= 100 ? 'px-mid' : 'px-low');
      expPxHtml += '<td class="' + cls + '">' + (px > 0 ? '$' + px : '-') + '</td>';
    });
    expPxHtml += '</tr>';
  });
  expPxHtml += '</tbody></table>';
  document.getElementById('t3-exp').innerHTML = expPxHtml;
})();

window._charts3Init = false;
function initCharts3() {
  if (window._charts3Init) return;
  window._charts3Init = true;

  var colors = ['#0f3460', '#2980b9', '#8e44ad', '#e07b24', '#08845a'];
  var revDatasets = pinxiaoData.map(function(d, i) {
    return { label: d.launchMonth + '月上架', data: d.monthlyRevenue, borderColor: colors[i], backgroundColor: 'transparent', tension: 0.3 };
  });
  new Chart(document.getElementById('chart-px-rev'), {
    type: 'line', data: { labels: ['1月','2月','3月','4月','5月'], datasets: revDatasets },
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true } } }
  });

  var pxDatasets = pinxiaoData.map(function(d, i) {
    return { label: d.launchMonth + '月上架', data: d.monthlyPinxiao, borderColor: colors[i], backgroundColor: 'transparent', tension: 0.3 };
  });
  new Chart(document.getElementById('chart-px-trend'), {
    type: 'line', data: { labels: ['1月','2月','3月','4月','5月'], datasets: pxDatasets },
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true } } }
  });
}

(function() {
  var t = cum43Stats;
  var pk = prevWeekKpi;
  var saleRate = t.hasRivalCount ? (t.yCount + t.nCount) / t.hasRivalCount * 100 : 0;
  var timelyRate = parseFloat(timelinessData.total.timelyRate) || 0;

  document.getElementById('t7-kpi').innerHTML =
    '<div class="kpi-card primary"><div class="label">在售SKU</div><div class="val">' + t.total + '</div></div>' +
    '<div class="kpi-card success"><div class="label">总销量</div><div class="val">' + fmtNum(pk.prevTotalSalesQty) + '</div></div>' +
    '<div class="kpi-card purple"><div class="label">总销售额</div><div class="val">' + fmtMoney(pk.prevTotalRevenue) + '</div></div>' +
    '<div class="kpi-card info"><div class="label">出单率</div><div class="val">' + saleRate.toFixed(1) + '%</div></div>' +
    '<div class="kpi-card success"><div class="label">及时率</div><div class="val">' + timelinessData.total.timelyRate + '</div></div>' +
    '<div class="kpi-card warning"><div class="label">低占比新品</div><div class="val">' + lowShareData.length + '</div></div>';

  var risks = [];
  if (saleRate < 70) risks.push({level:'high', title:'出单率偏低', text:'出单率仅 ' + saleRate.toFixed(1) + '%，低于70%警戒线。有对手未出单新品' + hasCompetitorUnsold.total + '款，需重点排查市场状态和定价策略。'});
  if (unsoldNoCompetitor.total > 15) risks.push({level:'high', title:'无对手未出单新品过多', text:'无对手未出单新品达' + unsoldNoCompetitor.total + '款，占比' + (unsoldNoCompetitor.total/t.total*100).toFixed(1) + '%，需加速市场调研和Listing优化。'});
  if (parseFloat(plpTotal.roas) < 8) risks.push({level:'medium', title:'PLP广告ROAS偏低', text:'PLP广告ROAS为' + plpTotal.roas + '，较上周' + plpPrevTotal.roas + '有所下降，需优化广告投放策略。'});
  if (timelyRate < 50) {
    var worstAn = timelinessData.analysts.reduce(function(a,b){ return parseFloat(a.timelyRate)<parseFloat(b.timelyRate) ? a : b; });
    risks.push({level:'high', title:'分析及时率告急', text: worstAn.analyst + '及时率仅 ' + worstAn.timelyRate + '，严重低于平均水平，需立即完成补充分析。'});
  }
  if (t.competitiveCount > t.normalCount) risks.push({level:'medium', title:'竞争无优势SKU偏多', text:'竞争无优势SKU(' + t.competitiveCount + ')超过正常SKU(' + t.normalCount + ')，需加大差异化卖点挖掘和价格竞争分析。'});
  if (risks.length === 0) risks.push({level:'low', title:'整体平稳', text:'本周各项指标整体平稳，暂无重大风险。'});

  var riskHtml = '';
  risks.forEach(function(r) {
    var cls = r.level === 'high' ? 'risk-high' : (r.level === 'medium' ? 'risk-medium' : 'risk-low');
    riskHtml += '<div class="report-block ' + cls + '"><h4>' + (r.level==='high'?'🔴':(r.level==='medium'?'🟡':'🟢')) + ' ' + r.title + '</h4><pre>' + r.text + '</pre></div>';
  });
  document.getElementById('t7-risk').innerHTML = riskHtml;

  var findings = [
    {title: '新品销售占比 ' + pk.deptRatio, desc: '本周新品销售额占部门总销售额的' + pk.deptRatio + '，部门总销售额' + fmtMoney(pk.deptTotalRevenue) + '，新品贡献销售额' + fmtMoney(pk.prevTotalRevenue) + '。'},
    {title: '出单率 ' + saleRate.toFixed(1) + '%', desc: '有对手新品出单率' + saleRate.toFixed(1) + '%(Y:' + t.yCount + '个 N:' + t.nCount + '个 未出单:' + t.unCount + '个)，整体出单情况良好。'},
    {title: '品效Cohort趋势', desc: pinxiaoData.length + '个上架月份的新品品效跟踪中，需重点关注低品效上架月份的产品优化。'},
    {title: 'PLP广告ACOAS ' + plpTotal.acoas, desc: '本周PLP广告ACOAS为' + plpTotal.acoas + '，花费' + fmtMoney(plpTotal.cost) + '，ROAS ' + plpTotal.roas + '。单链接PLP+PLG同开' + plgStats.singleLinkPlpPlgCount + '个SKU需重点关注。'},
    {title: '低占比新品' + lowShareData.length + '款', desc: '市占比<75%的低占比新品共' + lowShareData.length + '款，占总SKU的' + (lowShareData.length/t.total*100).toFixed(1) + '%，需逐一排查原因。'},
  ];
  var findingsHtml = '';
  findings.forEach(function(f) {
    findingsHtml += '<div class="findings-card"><div class="title">' + f.title + '</div><div class="desc">' + f.desc + '</div></div>';
  });
  document.getElementById('t7-findings').innerHTML = findingsHtml;

  var actions = [
    {title: '低占比新品逐一排查', desc: '对' + lowShareData.length + '款低占比新品逐一分析市场状态，重点关注"竞争无优势"和"无市场"SKU，制定差异化优化方案。'},
    {title: '品效提升专项', desc: '针对低品效上架月份的产品，重点优化标题关键词、主图质量和定价策略，提升每SKU产出。'},
    {title: '单链接PLP+PLG同开SKU优化', desc: '关注' + plgStats.singleLinkPlpPlgCount + '个单链接PLP+PLG同开SKU的广告表现，评估是否需要扩展广告活动数量。'},
    {title: '分析及时率提升', desc: '督促分析及时率偏低的分析师，确保新品8日内完成首次分析，7日内完成低占比追踪分析。'},
    {title: '竞争价格监控', desc: '对有对手但低市占比例的SKU进行竞争价格对比，必要时调整定价或运费策略以提升竞争力。'},
  ];
  var actionsHtml = '';
  actions.forEach(function(a) {
    actionsHtml += '<div class="action-card"><div class="title">' + a.title + '</div><div class="desc">' + a.desc + '</div></div>';
  });
  document.getElementById('t7-actions').innerHTML = actionsHtml;

  var reportSections = [
    {title: '一、总盘概览（Tab1）', text: '【核心KPI】\n' +
      '累计SKU: ' + t.total + ' | 总销量: ' + fmtNum(pk.prevTotalSalesQty) + ' | 总销售额: ' + fmtMoney(pk.prevTotalRevenue) + '\n' +
      '新品销售占比: ' + pk.deptRatio + '（部门总销售额: ' + fmtMoney(pk.deptTotalRevenue) + '）\n' +
      '出单率: ' + saleRate.toFixed(1) + '%（Y:' + t.yCount + '个 / N:' + t.nCount + '个 / 未出单:' + t.unCount + '个）\n' +
      '分析及时率: ' + timelinessData.total.timelyRate + '（及时:' + timelinessData.total.timelyCount + '个 / 8日未分析:' + timelinessData.total.noAnalysis8dCount + '个 / 7日未分析:' + timelinessData.total.noAnalysis7dCount + '个）\n\n' +
      '【品线维度】\n' + categoryRevenueData.map(function(d){
        return d.category + ': ' + d.curSku + 'SKU（新上架' + d.curNewSku + '个），本周销量' + fmtNum(d.curSalesQty) + '（环比' + d.salesQtyChange + '），本周销售额' + fmtMoney(d.curRevenue) + '（环比' + d.revenueChange + '）';
      }).join('\n') + '\n\n' +
      '【分析人维度】\n' + analystRevenueData.map(function(d){
        return d.analyst + ': ' + d.curSku + 'SKU（新上架' + d.curNewSku + '个），本周销量' + fmtNum(d.curSalesQty) + '，本周销售额' + fmtMoney(d.curRevenue);
      }).join('\n') + '\n\n' +
      '【拓展类型维度】\n' + expandTypeData.map(function(d){
        return d.expandType + ': ' + d.curSku + 'SKU，出单率' + d.curSalesRate + '（上周' + d.prevSalesRate + '），本周销量' + fmtNum(d.curSalesQty) + '（环比' + d.salesQtyChange + '），本周销售额' + fmtMoney(d.curRevenue) + '（环比' + d.revenueChange + '）';
      }).join('\n')
    },
    {title: '二、低占比分析（Tab2）', text: '【有对手未出单新品：' + hasCompetitorUnsold.total + '款】\n' +
      '原因分布: ' + hasCompetitorUnsold.reasons.filter(function(r){return r.count>0;}).map(function(r){return r.name + '(' + r.count + '款)';}).join('、') + '\n' +
      '按分析人: ' + hasCompetitorUnsold.byAnalyst.filter(function(d){return d.total>0;}).map(function(d){return d.analyst + '(' + d.total + '款)';}).join('、') + '\n' +
      '按品线: ' + hasCompetitorUnsold.byCategory.filter(function(d){return d.total>0;}).map(function(d){return d.category + '(' + d.total + '款)';}).join('、') + '\n\n' +
      '【无对手未出单新品：' + unsoldNoCompetitor.total + '款】\n' +
      '原因分布: ' + unsoldNoCompetitor.reasons.filter(function(r){return r.count>0;}).map(function(r){return r.name + '(' + r.count + '款)';}).join('、') + '\n' +
      '按分析人: ' + unsoldNoCompetitor.byAnalyst.filter(function(d){return d.total>0;}).map(function(d){return d.analyst + '(' + d.total + '款)';}).join('、') + '\n' +
      '按品线: ' + unsoldNoCompetitor.byCategory.filter(function(d){return d.total>0;}).map(function(d){return d.category + '(' + d.total + '款)';}).join('、') + '\n\n' +
      '【低占比新品（市占比<75%）：' + lowShareData.length + '款】\n' +
      '占总SKU ' + (lowShareData.length/t.total*100).toFixed(1) + '%，涉及分析人: ' + (function(){var ans={};lowShareData.forEach(function(d){ans[d.analyst]=1;});return Object.keys(ans).join('、');})()}
    ,
    {title: '三、广告追踪（Tab3）', text: '【PLP广告核心指标】\n' +
      '广告活动数: ' + plpTotal.campaignCount + '（上周' + plpPrevTotal.campaignCount + '） | 投放链接: ' + plpTotal.linkCount + '（上周' + plpPrevTotal.linkCount + '）\n' +
      '曝光量: ' + fmtNum(plpTotal.impression) + ' | 点击量: ' + fmtNum(plpTotal.click) + ' | 售出: ' + plpTotal.sold + '单\n' +
      '广告花费: ' + fmtMoney(plpTotal.cost) + ' | 广告销售额: ' + fmtMoney(plpTotal.revenue) + ' | 总销售额: ' + fmtMoney(plpTotal.totalRevenue) + '\n' +
      'ROAS: ' + plpTotal.roas + '（上周' + plpPrevTotal.roas + '） | ACOS: ' + plpTotal.acos + ' | ACOAS: ' + plpTotal.acoas + '\n' +
      'CVR: ' + plpTotal.cvr + ' | CTR: ' + plpTotal.ctr + ' | CPC: ' + plpTotal.cpc + ' | CPA: ' + plpTotal.cpa + '\n\n' +
      '【PLG广告分类统计】\n' +
      '新品总数' + plgStats.totalNewProducts + '：PLP+PLG同开' + plgStats.plpAndPlgBothCount + '款、单链接PLP+PLG同开' + plgStats.singleLinkPlpPlgCount + '款、单PLG' + plgStats.plgOnlyCount + '款、单PLP' + plgStats.plpOnlyCount + '款、单PLG且未出单' + (plgStats.plpDisabledNoSaleCount||0) + '款、无广告' + plgStats.noAdCount + '款\n' +
      '按分析人: ' + plgStats.byAnalyst.map(function(d){
        return d.analyst + '(PLP+PLG:' + d.plpAndPlgBoth + '/单链:' + d.singleLinkPlpPlg + '/单PLG:' + d.plgOnly + '/单PLP:' + d.plpOnly + '/无广告:' + d.noAd + ')';
      }).join('、')
    },
    {title: '四、品效分析（Tab5）', text: '【品效Cohort（按上架月份）】\n' +
      pinxiaoData.map(function(d){
        var pxVals = d.monthlyPinxiao.filter(function(v){return v>0;});
        var lastPx = pxVals.length > 0 ? pxVals[pxVals.length-1] : 0;
        var trend = '-';
        if (pxVals.length >= 3) {
          if (pxVals[pxVals.length-1] > pxVals[pxVals.length-2] && pxVals[pxVals.length-2] > pxVals[pxVals.length-3]) trend = '↑攀升';
          else if (pxVals[pxVals.length-1] < pxVals[pxVals.length-2] && pxVals[pxVals.length-2] < pxVals[pxVals.length-3]) trend = '↓衰减';
          else trend = '→平稳';
        } else if (pxVals.length >= 2) {
          trend = pxVals[pxVals.length-1] > pxVals[pxVals.length-2] ? '↑攀升' : '↓衰减';
        }
        return d.launchMonth + '月上架: ' + d.skuCount + '个SKU，累计销售额' + fmtMoney(d.monthlyRevenue.reduce(function(a,b){return a+b;},0)) + '，最新月品效$' + (lastPx>0?lastPx.toFixed(0):'-') + '/SKU，趋势' + trend;
      }).join('\n') + '\n\n' +
      '【品线品效（5月上架）】\n' + pinxiaoCatData.map(function(d){
        var px = d.pinxiao.length > 0 ? d.pinxiao[d.pinxiao.length-1] : 0;
        return d.category + ': $' + (px>0?px:'-') + '/SKU';
      }).join(' | ') + '\n\n' +
      '【拓展类型品效（5月上架）】\n' + pinxiaoExpData.map(function(d){
        var px = d.pinxiao.length > 0 ? d.pinxiao[d.pinxiao.length-1] : 0;
        return d.expandType + ': $' + (px>0?px:'-') + '/SKU';
      }).join(' | ')
    },
    {title: '五、风险预警与下周动作', text: '【风险预警】\n' +
      risks.map(function(r){ return '[' + r.level.toUpperCase() + '] ' + r.title + ': ' + r.text; }).join('\n') + '\n\n' +
      '【主要发现】\n' + findings.map(function(f){ return '- ' + f.title + ': ' + f.desc; }).join('\n') + '\n\n' +
      '【下周重点动作】\n' + actions.map(function(a, i){ return (i+1) + '. ' + a.title + ': ' + a.desc; }).join('\n')
    },
  ];
  var reportHtml = '';
  reportSections.forEach(function(sec) {
    reportHtml += '<div class="report-block"><h4>' + sec.title + '</h4><pre>' + sec.text + '</pre><button class="copy-btn" onclick="copyReport(this)">复制</button></div>';
  });
  document.getElementById('t7-report').innerHTML = reportHtml;
})();

// ===== Tab2: 市场分布 =====
(function() {
  var mo = mktDistOverall;
  var normalItem = mo.distribution.find(function(d) { return d.status === '正常'; }) || {curCount:0, curPct:0, change:0};
  var competitiveItem = mo.distribution.find(function(d) { return d.status === '竞争无优势'; }) || {curCount:0, change:0};
  var noMarketItem = mo.distribution.find(function(d) { return d.status === '无市场'; }) || {curCount:0, change:0};
  var stationOutItem = mo.distribution.find(function(d) { return d.status === '站外出单'; }) || {curCount:0, change:0};

  document.getElementById('t2-kpi').innerHTML =
    '<div class="kpi-card primary"><div class="label">在售SKU总数</div><div class="val">' + mo.curTotal + '</div><div class="hb">上周 ' + mo.prevTotal + '</div></div>' +
    '<div class="kpi-card success"><div class="label">市场正常</div><div class="val">' + normalItem.curCount + '</div><div class="hb">占比 ' + normalItem.curPct + '%</div></div>' +
    '<div class="kpi-card warning"><div class="label">竞争无优势</div><div class="val">' + competitiveItem.curCount + '</div><div class="hb">' + hbSign((competitiveItem.change >= 0 ? '+' : '') + competitiveItem.change) + ' vs 上周</div></div>' +
    '<div class="kpi-card danger"><div class="label">无市场</div><div class="val">' + noMarketItem.curCount + '</div><div class="hb">' + hbSign((noMarketItem.change >= 0 ? '+' : '') + noMarketItem.change) + ' vs 上周</div></div>' +
    '<div class="kpi-card purple"><div class="label">站外出单</div><div class="val">' + stationOutItem.curCount + '</div><div class="hb">' + hbSign((stationOutItem.change >= 0 ? '+' : '') + stationOutItem.change) + ' vs 上周</div></div>';

  // 市场状态明细表
  var mh = '<div class="table-scroll-wrap"><table class="data-table"><thead><tr><th>市场状态</th><th>本周数量</th><th>本周占比</th><th>上周数量</th><th>上周占比</th><th>变化</th></tr></thead><tbody>';
  mo.distribution.forEach(function(d) {
    mh += '<tr><td>' + badgeStatus(d.status) + '</td><td>' + d.curCount + '</td><td>' + d.curPct + '%</td><td>' + d.prevCount + '</td><td>' + d.prevPct + '%</td><td>' + hbSign((d.change >= 0 ? '+' : '') + d.change) + '</td></tr>';
  });
  mh += '</tbody></table></div>';
  document.getElementById('t2-mkt-table').innerHTML = mh;

  // 货值明细表
  var po = priceOverview;
  var ph = '<div class="table-scroll-wrap"><table class="data-table"><thead><tr><th>价格区间</th><th>SKU数</th><th>占比</th></tr></thead><tbody>';
  po.distribution.forEach(function(d) {
    ph += '<tr><td>' + d.range + '</td><td>' + d.count + '</td><td>' + d.pct + '%</td></tr>';
  });
  ph += '</tbody></table></div>';
  document.getElementById('t2-price-table').innerHTML = ph;

  // 市占比分布明细表（品线 x 高中低）
  var sto = shareTierOverview;
  var sh = '<div class="table-scroll-wrap"><table class="data-table"><thead><tr><th>品线</th><th>总SKU</th><th>高市占比(>=75%)</th><th>中市占比(50-75%)</th><th>低市占比(<50%)</th></tr></thead><tbody>';
  sto.byCategory.forEach(function(d) {
    sh += '<tr><td><b>' + d.category + '</b></td><td>' + d.total + '</td><td>' + d.high + '</td><td>' + d.mid + '</td><td>' + d.low + '</td></tr>';
  });
  sh += '</tbody></table></div>';
  document.getElementById('t2-share-tier-table').innerHTML = sh;
})();

// ===== Tab2 图表（懒初始化）=====
window._charts2Init = false;
function initCharts2() {
  if (window._charts2Init) return;
  window._charts2Init = true;

  var mktColors = {
    '正常': '#08845a',
    '竞争无优势': '#e07b24',
    '无市场': '#c0392b',
    '站外出单': '#8e44ad',
    '站内无价格优势': '#f39c12',
    '#N/A': '#95a5a6',
    '未知': '#7f8c8d',
    '其他': '#bdc3c7'
  };

  var mo = mktDistOverall;

  // 图表1：环形图 - 本周市场状态占比
  var ringLabels = [], ringData = [], ringColors = [];
  mo.distribution.forEach(function(d) {
    if (d.curCount > 0) {
      ringLabels.push(d.status);
      ringData.push(d.curCount);
      ringColors.push(mktColors[d.status] || '#999');
    }
  });

  new Chart(document.getElementById('chart-mkt-ring'), {
    type: 'doughnut',
    data: { labels: ringLabels, datasets: [{ data: ringData, backgroundColor: ringColors, borderWidth: 1, borderColor: '#fff' }] },
    options: {
      responsive: true,
      plugins: {
        legend: { position: 'bottom', labels: { padding: 12, font: { size: 11 } } },
        tooltip: { callbacks: { label: function(ctx) { var total = ctx.dataset.data.reduce(function(a,b){return a+b;}, 0); var pct = total > 0 ? (ctx.parsed / total * 100).toFixed(1) : 0; return ctx.label + ': ' + ctx.parsed + ' SKU (' + pct + '%)'; } } }
      }
    }
  });

  // 图表2：分组柱状图 - 本周vs上周各状态SKU数
  var barLabels = [], curData = [], prevData = [];
  mo.distribution.forEach(function(d) {
    barLabels.push(d.status);
    curData.push(d.curCount);
    prevData.push(d.prevCount);
  });

  new Chart(document.getElementById('chart-mkt-bar'), {
    type: 'bar',
    data: { labels: barLabels, datasets: [
      { label: '本周', data: curData, backgroundColor: '#0f3460' },
      { label: '上周', data: prevData, backgroundColor: '#ccc' }
    ]},
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true, title: { display: true, text: 'SKU数量' } } } }
  });

  // 图表3：柱状图 - 价格区间SKU分布
  var po = priceOverview;
  new Chart(document.getElementById('chart-price-dist'), {
    type: 'bar',
    data: { labels: po.priceRanges, datasets: [
      { label: 'SKU数', data: po.distribution.map(function(d){return d.count;}), backgroundColor: '#0f3460' }
    ]},
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true, title: { display: true, text: 'SKU数量' } } } }
  });

  // 图表4：分组柱状图 - 按分析人各价格区间SKU数
  var anPriceLabels = po.byAnalyst.map(function(d){return d.analyst;});
  var priceRangeColors = ['#0f3460', '#08845a', '#e07b24', '#8e44ad', '#c0392b', '#f39c12'];
  var priceDatasets = po.priceRanges.map(function(r, i) {
    return { label: r, data: po.byAnalyst.map(function(d){return d[r] || 0;}), backgroundColor: priceRangeColors[i % priceRangeColors.length] };
  });

  new Chart(document.getElementById('chart-price-an'), {
    type: 'bar',
    data: { labels: anPriceLabels, datasets: priceDatasets },
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { x: { stacked: true }, y: { stacked: true, beginAtZero: true, title: { display: true, text: 'SKU数量' } } } }
  });

  // 图表5：堆叠柱状图 - 各品线高中低市占比
  var sto = shareTierOverview;
  var tierColors = ['#08845a', '#f39c12', '#c0392b'];
  new Chart(document.getElementById('chart-share-tier'), {
    type: 'bar',
    data: {
      labels: sto.byCategory.map(function(d){return d.category;}),
      datasets: [
        { label: '高市占比(>=75%)', data: sto.byCategory.map(function(d){return d.high;}), backgroundColor: tierColors[0] },
        { label: '中市占比(50-75%)', data: sto.byCategory.map(function(d){return d.mid;}), backgroundColor: tierColors[1] },
        { label: '低市占比(<50%)', data: sto.byCategory.map(function(d){return d.low;}), backgroundColor: tierColors[2] }
      ]
    },
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { x: { stacked: true }, y: { stacked: true, beginAtZero: true, title: { display: true, text: 'SKU数量' } } } }
  });
}
setTimeout(initCharts2, 100);

function copyReport(btn) {
  var pre = btn.parentElement.querySelector('pre');
  navigator.clipboard.writeText(pre.textContent).then(function() {
    btn.textContent = '已复制';
    setTimeout(function() { btn.textContent = '复制'; }, 1500);
  });
}
'''

full_html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>新品周报看板 5.14-5.20 v2</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
{CSS}
</head>
<body>
{HTML_BODY}
<script>
{js_vars_block}

{JS_CODE}
</script>
</body>
</html>'''

with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    f.write(full_html)

print(f"HTML已保存到: {OUTPUT_FILE}")
print(f"文件大小: {len(full_html)} 字符")
print("完成！")
