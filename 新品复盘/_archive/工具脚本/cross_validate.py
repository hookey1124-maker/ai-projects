"""
交叉验证脚本：对比源数据与HTML中嵌入的数据是否一致
"""
import re, json, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import defaultdict

SRC = 'C:/Users/Hardy/ai-projects/新品复盘/周报/新品检查周源数据和PLP数据.xlsx'
HTML = 'C:/Users/Hardy/ai-projects/新品复盘/新品板块_5.21-5.27.html'

# 读取HTML中的JS数据
with open(HTML, 'r', encoding='utf-8') as f:
    html = f.read()

# 提取所有const变量
var_pattern = re.findall(r'const (\w+) = (.*?);', html, re.DOTALL)
data_blocks = {}
for name, value in var_pattern:
    try:
        parsed = json.loads(value)
        data_blocks[name] = parsed
    except:
        pass  # skip non-JSON (functions etc)

print(f"提取到 {len(data_blocks)} 个数据块")

# 1. 验证cum43Stats
t = data_blocks.get('cum43Stats', {})
print(f"\n=== cum43Stats ===")
print(f"  total={t.get('total')}, yCount={t.get('yCount')}, nCount={t.get('nCount')}, unCount={t.get('unCount')}")
print(f"  noRivalSold={t.get('noRivalSold')}, noRivalUnsold={t.get('noRivalUnsold')}")
print(f"  normalCount={t.get('normalCount')}, competitiveCount={t.get('competitiveCount')}, noMarketCount={t.get('noMarketCount')}")
print(f"  totalMarketShare={t.get('totalMarketShare')}%, totalMarketSharePrev={t.get('totalMarketSharePrev')}%")

# 验证4段总和
total_4seg = t.get('yCount',0) + t.get('nCount',0) + t.get('unCount',0) + t.get('noRivalSold',0) + t.get('noRivalUnsold',0)
print(f"  4段总和={total_4seg}, total={t.get('total')}, 一致={'OK' if total_4seg == t.get('total',0) else 'MISMATCH!'}")

# 2. 验证市占比数据（源数据 vs HTML）
wb = load_workbook(SRC, data_only=True)
ws = ws_main = wb['四三数据累计']

# 计算总市占比
rows_raw = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1]: rows_raw.append(list(row))

from datetime import date, datetime
def get_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

def num(v, d=0):
    try: return float(v) if v else d
    except: return d

cutoff = date(2026, 5, 27)
rows_c = [r for r in rows_raw if get_date(r[2]) and get_date(r[2]) <= cutoff]
total_sales = sum(num(r[18]) for r in rows_c)
total_rival = sum(num(r[44]) for r in rows_c)
calc_share = round(total_sales / (total_sales + total_rival) * 100, 1) if (total_sales + total_rival) > 0 else 0
html_share = t.get('totalMarketShare', -1)
print(f"\n=== 总市占比交叉验证 ===")
print(f"  源数据: 新品销量={int(total_sales)}, 对手销量={int(total_rival)}, 市占比={calc_share}%")
print(f"  HTML: {html_share}%")
print(f"  一致: {'OK' if abs(calc_share - html_share) < 0.1 else 'MISMATCH!'}")

# 3. 验证品线市占比
print(f"\n=== 品线市占比 ===")
cat_curr = defaultdict(lambda: {'sales': 0, 'rival': 0})
for r in rows_c:
    cat = str(r[5] or '').strip() or '未分类'
    cat_curr[cat]['sales'] += num(r[18])
    cat_curr[cat]['rival'] += num(r[44])

cat_rev_data = {d['category']: d for d in data_blocks.get('categoryRevenueData', [])}
for cat in ['车门系统', '车身外扩件', '挡泥板', '机盖及附件', '其他', '饰条', '牌照板支架']:
    if cat in cat_curr:
        d = cat_curr[cat]
        s = d['sales']; r = d['rival']
        share = round(s / (s + r) * 100, 1) if (s + r) > 0 else 0
        html_d = cat_rev_data.get(cat, {})
        html_share2 = html_d.get('curMarketShare', -1)
        match = 'OK' if abs(share - html_share2) < 0.2 else f'MISMATCH({share} vs {html_share2})'
        print(f"  {cat}: src={share}%, html={html_share2}% [{match}]")

# 4. 验证PLG数据
print(f"\n=== PLG数据验证 ===")
plg_total_spend = sum(num(r[145]) for r in rows_c)
plg_total_ad_rev = sum(num(r[146]) for r in rows_c)
pg = data_blocks.get('plgStats', {})
html_spend = pg.get('totalSpend', -1)
html_ad_rev = pg.get('totalAdRev', -1)
print(f"  PLG花费: src={plg_total_spend:.2f}, html={html_spend}, match={'OK' if abs(plg_total_spend - html_spend) < 0.1 else 'MISMATCH'}")
print(f"  PLG广告销售额: src={plg_total_ad_rev:.2f}, html={html_ad_rev}, match={'OK' if abs(plg_total_ad_rev - html_ad_rev) < 0.1 else 'MISMATCH'}")

# 5. 验证PLP数据
print(f"\n=== PLP数据验证 ===")
plp = data_blocks.get('plpTotal', {})
print(f"  PLP活动数: {plp.get('campaignCount')}, 链接数: {plp.get('linkCount')}")
print(f"  PLP花费: {plp.get('cost')}, 广告销售额: {plp.get('revenue')}")
print(f"  PLP ROAS: {plp.get('roas')}, ACOS: {plp.get('acos')}, ACOAS: {plp.get('acoas')}")

# 6. 无对手SKU出单/未出单验证（按ord8历史出单状态）
no_rival_rows = [r for r in rows_c if num(r[44]) == 0]
no_rival_sold = sum(1 for r in no_rival_rows if str(r[80] or '').strip() in ('Y', 'N'))
no_rival_unsold = len(no_rival_rows) - no_rival_sold
print(f"\n=== 无对手SKU（按历史出单ord8）===")
print(f"  总数: {len(no_rival_rows)}, 已出单(ord8 Y/N): {no_rival_sold}, 未出单: {no_rival_unsold}")
print(f"  HTML: sold={t.get('noRivalSold')}, unsold={t.get('noRivalUnsold')}")
print(f"  一致: {'OK' if no_rival_sold == t.get('noRivalSold') and no_rival_unsold == t.get('noRivalUnsold') else 'MISMATCH!'}")

# 7. 检查关键百分比格式
print(f"\n=== 百分比格式检查 ===")
# 检查市占比是否都在0-100范围
all_shares = [d.get('curMarketShare', 0) for d in data_blocks.get('cum43Data', [])]
over_100 = [s for s in all_shares if s > 100]
under_0 = [s for s in all_shares if s < 0]
print(f"  市占比>100: {len(over_100)}个, <0: {len(under_0)}个")
if over_100:
    print(f"    异常值: {over_100[:5]}")

# 检查PLG费率格式
plg_fees = [d.get('plgFee', '') for d in data_blocks.get('cum43Data', [])]
print(f"  PLG费率样本: {plg_fees[:10]}")

print(f"\n=== 交叉验证完成 ===")
