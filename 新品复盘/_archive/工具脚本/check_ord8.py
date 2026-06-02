import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from datetime import date, datetime
from collections import Counter

wb = load_workbook('C:/Users/Hardy/ai-projects/新品复盘/周报/新品检查周源数据和PLP数据.xlsx', data_only=True)
ws = wb['四三数据累计']

def get_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

def num(v, d=0):
    try: return float(v) if v else d
    except: return d

cutoff = date(2026, 5, 27)
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1] and get_date(row[2]) and get_date(row[2]) <= cutoff:
        rows.append(list(row))

print(f'Total SKUs: {len(rows)}')
print(f'Has rival (rival>0): {sum(1 for r in rows if num(r[44])>0)}')
print(f'No rival (rival==0): {sum(1 for r in rows if num(r[44])==0)}')

# Ord8 value distribution for SKUs with rivals
rival_ord8 = Counter()
for r in rows:
    if num(r[44]) > 0:
        v = str(r[80] or '').strip()
        rival_ord8[v] += 1

print('\nOrd8 values for SKUs with rivals:')
for k, v in sorted(rival_ord8.items(), key=lambda x: -x[1]):
    print(f'  "{k}": {v}')

# Also check no-rival ord8
norival_ord8 = Counter()
for r in rows:
    if num(r[44]) == 0:
        v = str(r[80] or '').strip()
        norival_ord8[v] += 1

print('\nOrd8 values for SKUs without rivals:')
for k, v in sorted(norival_ord8.items(), key=lambda x: -x[1]):
    print(f'  "{k}": {v}')

# Find non-standard ord8 SKUs with rivals
for r in rows:
    if num(r[44]) > 0:
        v = str(r[80] or '').strip()
        if v not in ('Y', 'N', '未出单'):
            print(f'\nNon-standard ord8 SKU with rival: {r[1]} ord8="{v}" rival={num(r[44])} sales={num(r[18])} mkt={r[117]}')
