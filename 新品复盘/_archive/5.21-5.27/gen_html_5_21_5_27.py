"""
生成5.21-5.27周期新品周报HTML看板（5 Tab：总盘概览/低占比分析/广告追踪/四三累计/汇报输出）
关键变化：
- 删除品效Tab（移至月报）
- 出单分布甜甜圈4段（有对手已出单/有对手未出单/无对手已出单/无对手未出单）
- 所有图表添加环比数据
- 新品总市占比计算和环比
- 各分析人/品线市占比（表格+图表）
- PLG板块：花费、广告销售额、ACOS、ACOAS
- PLP使用自然周数据（5.18-5.24）
"""
import json, os, sys
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook
from collections import defaultdict, Counter
from datetime import date, datetime

WORKDIR = os.path.dirname(os.path.abspath(__file__)) + '/'
SOURCE_FILE = WORKDIR + '周报/新品检查周源数据和PLP数据.xlsx'
OUTPUT_FILE = WORKDIR + '新品板块_5.21-5.27.html'

# ===== 列索引（5.21-5.27周期）=====
C = {
    'sale_no': 0, 'sku': 1, 'list_date': 2, 'first_order': 3,
    'analyst': 4, 'category': 5, 'expand_type': 6,
    'sales_curr': 18, 'sales_prev': 17,   # 5.21-5.27 vs 5.14-5.20
    'rev_curr': 31, 'rev_prev': 30,       # 5.21-5.27 vs 5.14-5.20
    'rival_curr': 44, 'rival_prev': 43,   # 5.27 vs 5.20
    'share_curr': 56, 'share_prev': 55,   # 5.27 vs 5.20
    'ord8_curr': 80, 'ord8_prev': 79,     # 5.27 vs 5.20
    'freq7_curr': 92, 'freq7_prev': 91,   # 5.27 vs 5.20
    'nfreq7_curr': 104, 'nfreq7_prev': 103,  # 5.27 vs 5.20
    'mkt_curr': 117, 'mkt_prev': 116,     # 5.27 vs 5.20
    'op_curr': 129,                        # 5.21-5.27
    'plp_curr': 136,                       # 5.18-5.24开启PLP
    'plp_spend': 137,                      # 5.18-5.24 PLP广告花费
    'plp_ad_rev': 138,                     # 5.18-5.24 PLP广告销售额
    'plg_fee': 144,                        # 5.18-5.24 PLG最高费率
    'plg_spend': 145,                      # 5.18-5.24 PLG广告花费
    'plg_ad_rev': 146,                     # 5.18-5.24 PLG广告销售额
}

# PLP明细Sheet列索引（不变）
PC = {
    'period': 0, 'campaign': 1, 'sku': 2, 'id': 3, 'store': 4,
    'plp_start': 5, 'list_date': 6, 'first_order': 7, 'analyst': 8,
    'category': 9, 'expand_type': 10,
    'impr': 11, 'click': 12, 'sold': 13, 'cost': 14, 'ad_rev': 15,
    'total_rev': 16, 'roas': 17, 'cvr': 18, 'ctr': 19,
    'cpc': 20, 'cpa': 21, 'acos': 22, 'acoas': 23, 'plg_enabled': 24,
}

def safe_float(v, default=0):
    try: return float(v) if v else default
    except: return default

def get_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

def num(v, default=0):
    return safe_float(v, default)

def calc_rate(a, b):
    if not b: return 0
    return round(a/b, 4)

def ratio_str(a, b):
    """计算环比变化百分比字符串"""
    if not b: return '-'
    return f"{round((a-b)/abs(b)*100, 1)}%"

def ratio_num(a, b):
    """计算环比变化数值"""
    if not b: return 0
    return round((a-b)/abs(b)*100, 1)

def get_cat(r_data):
    c = str(r_data[5] or '').strip()
    return c if c else '未分类'

def get_an(r_data):
    a = str(r_data[4] or '').strip()
    return a if a else '未知'

def get_exp(r_data):
    e = str(r_data[6] or '').strip()
    return e if e else '其他'

ANALYSTS = ['俞东旭', '张潇', '朱培源', '王偲涵', '章鹏', '胡煜星']
CATEGORIES = ['车门系统', '车身外扩件', '挡泥板', '机盖及附件', '其他', '饰条', '牌照板支架', '未分类']
EXPAND_TYPES = ['原开品', '拓展品', '组合件']
ALL_MKT_STATUSES = ['正常', '竞争无优势', '无市场', '站外出单', '站内无价格优势', '#N/A', '未知', '其他']

# 自然周周期
PLP_CURR = '5.18-5.24'
PLP_PREV = '5.11-5.17'
# 截止日期
cutoff_curr = date(2026, 5, 27)
cutoff_prev = date(2026, 5, 20)

# ===== 读取源数据 =====
print("读取源数据...")
wb_src = load_workbook(SOURCE_FILE, data_only=True)
ws_main = wb_src['四三数据累计']
ws_plp = wb_src['PLP明细']
ws_dept = wb_src['四三销售数据']
ws_nat = wb_src['自然周销售数据']

# 读取主表数据
rows_raw = []
for row in ws_main.iter_rows(min_row=2, values_only=True):
    if row[C['sku']]:
        rows_raw.append(list(row))

rows_curr = [r for r in rows_raw if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) <= cutoff_curr]
rows_prev = [r for r in rows_raw if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) <= cutoff_prev]

total_sku = len(rows_curr)
total_sales_curr = sum(num(r[C['sales_curr']]) for r in rows_curr)
total_sales_prev = sum(num(r[C['sales_prev']]) for r in rows_prev)
total_rev_curr = sum(num(r[C['rev_curr']]) for r in rows_curr)
total_rev_prev = sum(num(r[C['rev_prev']]) for r in rows_prev)
has_rival_curr = sum(1 for r in rows_curr if num(r[C['rival_curr']]) > 0)
has_rival_prev = sum(1 for r in rows_prev if num(r[C['rival_prev']]) > 0)

# 总对手销量（用于计算新品总市占比）
total_rival_sales_curr = sum(num(r[C['rival_curr']]) for r in rows_curr)
total_rival_sales_prev = sum(num(r[C['rival_prev']]) for r in rows_prev)

# 新品总市占比 = 新品自身总销量 / (新品自身总销量 + 对手总销量)
total_market_share_curr = round(total_sales_curr / (total_sales_curr + total_rival_sales_curr) * 100, 1) if (total_sales_curr + total_rival_sales_curr) > 0 else 0
total_market_share_prev = round(total_sales_prev / (total_sales_prev + total_rival_sales_prev) * 100, 1) if (total_sales_prev + total_rival_sales_prev) > 0 else 0

# 部门销售数据
dept_total_sales = 0
dept_total_revenue = 0.0
for row in ws_dept.iter_rows(min_row=2, max_row=2, values_only=True):
    dept_total_sales = int(num(row[4])) if row[4] else 0
    dept_total_revenue = num(row[5])

dept_ratio = round(total_rev_curr / dept_total_revenue * 100, 1) if dept_total_revenue > 0 else 0

# 读取自然周销售数据（SKU级别的自然周销量和销售额）
nat_week_data = {}
for row in ws_nat.iter_rows(min_row=2, values_only=True):
    sku = str(row[0] or '').strip()
    if sku and sku != '#N/A':
        qty = num(row[2])
        rev = num(row[3])
        if sku in nat_week_data:
            nat_week_data[sku]['qty'] += qty
            nat_week_data[sku]['rev'] += rev
        else:
            nat_week_data[sku] = {'qty': qty, 'rev': rev}

print(f"总SKU: {total_sku}, 本周销量: {total_sales_curr}, 本周销售额: {total_rev_curr}")
print(f"有对手SKU: {has_rival_curr}, 总对手销量: {total_rival_sales_curr}")
print(f"新品总市占比: {total_market_share_curr}% (上周: {total_market_share_prev}%)")

# ===== PLP数据（自然周 5.18-5.24）=====
plp_curr_rows = []
plp_plg_enabled = set()
for row in ws_plp.iter_rows(min_row=2, values_only=True):
    period = str(row[PC['period']] or '').strip()
    sku = str(row[PC['sku']] or '').strip()
    if not sku or sku.startswith('广告') or sku.startswith('总数据'):
        continue
    list_d = get_date(row[PC['list_date']])
    if period == PLP_CURR and list_d and list_d <= cutoff_curr:
        plp_curr_rows.append(list(row))
        if str(row[PC['plg_enabled']] or '').strip() == 'Y':
            plp_plg_enabled.add(sku)

sku_campaign_count = Counter()
for row in plp_curr_rows:
    sku = str(row[PC['sku']] or '').strip()
    if sku in plp_plg_enabled:
        sku_campaign_count[sku] += 1

def agg_plp(rows):
    by_sku = defaultdict(lambda: {'impr': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'total_rev': 0, 'campaigns': set()})
    for r in rows:
        sku = str(r[PC['sku']] or '').strip()
        camp = str(r[PC['campaign']] or '').strip()
        by_sku[sku]['impr'] += num(r[PC['impr']])
        by_sku[sku]['click'] += num(r[PC['click']])
        by_sku[sku]['sold'] += num(r[PC['sold']])
        by_sku[sku]['cost'] += num(r[PC['cost']])
        by_sku[sku]['ad_rev'] += num(r[PC['ad_rev']])
        by_sku[sku]['total_rev'] += num(r[PC['total_rev']])
        if camp: by_sku[sku]['campaigns'].add(camp)
    return by_sku

def plp_totals(by_sku):
    t = {'imp': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'total_rev': 0, 'campaigns': 0}
    all_camps = set()
    for sku, d in by_sku.items():
        t['imp'] += d['impr']; t['click'] += d['click']; t['sold'] += d['sold']
        t['cost'] += d['cost']; t['ad_rev'] += d['ad_rev']
        t['total_rev'] += d['total_rev']
        all_camps.update(d['campaigns'])
    t['campaigns'] = len(all_camps)
    t['links'] = len(by_sku)
    t['roas'] = calc_rate(t['ad_rev'], t['cost'])
    t['cvr'] = calc_rate(t['sold'], t['click'])
    t['ctr'] = calc_rate(t['click'], t['imp'])
    t['cpc'] = round(t['cost'] / t['click'], 2) if t['click'] else 0
    t['cpa'] = round(t['cost'] / t['sold'], 2) if t['sold'] else 0
    t['acos'] = calc_rate(t['cost'], t['ad_rev'])
    t['acoas'] = calc_rate(t['cost'], t['total_rev'])
    return t

plp_curr_sku = agg_plp(plp_curr_rows)
plp_t = plp_totals(plp_curr_sku)

# 构建SKU->本周销售额map
sku_rev_curr = {}
for r in rows_curr:
    sku = str(r[C['sku']]).strip()
    if sku: sku_rev_curr[sku] = num(r[C['rev_curr']])

# ACOAS SKU去重计算
sku_total_rev_dedup = {}
for sku in plp_curr_sku:
    info_rows = [r for r in plp_curr_rows if str(r[PC['sku']] or '').strip() == sku]
    if info_rows:
        sku_total_rev_dedup[sku] = num(info_rows[0][PC['total_rev']])

dedup_total_rev = sum(sku_total_rev_dedup.values())
dedup_acoas = round(plp_t['cost'] / dedup_total_rev, 4) if dedup_total_rev > 0 else 0

# ===== 广告分类函数 =====
def get_ad_class(r_data):
    sku = str(r_data[C['sku']]).strip()
    plp_on = str(r_data[C['plp_curr']] or '').strip().upper() == 'Y'
    plg_on = num(r_data[C['plg_fee']]) > 0
    in_plg_detail = sku in plp_plg_enabled
    is_unsold = str(r_data[C['ord8_curr']] or '').strip() == '未出单'

    if in_plg_detail and sku_campaign_count.get(sku, 0) == 1:
        return '单链接PLP+PLG同开'
    elif plp_on and plg_on:
        return 'PLP+PLG同开'
    elif plg_on and not plp_on:
        return '单PLG且未出单' if is_unsold else '单PLG'
    elif plp_on and not plg_on:
        return '单PLP'
    return '无广告'

# ===== 构建数据块 =====
print("构建数据块...")

# 1. cum43Data
cum43Data = []
for r_data in rows_curr:
    sku = str(r_data[C['sku']]).strip()
    list_d = get_date(r_data[C['list_date']])
    first_d = get_date(r_data[C['first_order']])
    cum43Data.append({
        'saleNo': r_data[C['sale_no']] or '',
        'SKU': sku,
        'listDate': str(list_d)[:10] if list_d else '',
        'firstOrderDate': str(first_d)[:10] if first_d else '未出单',
        'analyst': get_an(r_data),
        'category': get_cat(r_data),
        'expandType': get_exp(r_data),
        'curSalesQty': int(num(r_data[C['sales_curr']])),
        'prevSalesQty': int(num(r_data[C['sales_prev']])),
        'curRevenue': round(num(r_data[C['rev_curr']]), 2),
        'prevRevenue': round(num(r_data[C['rev_prev']]), 2),
        'curRivalQty': int(num(r_data[C['rival_curr']])),
        'prevRivalQty': int(num(r_data[C['rival_prev']])),
        'curMarketShare': round(num(r_data[C['share_curr']]) * 100, 1),
        'prevMarketShare': round(num(r_data[C['share_prev']]) * 100, 1),
        'cur8dStatus': str(r_data[C['ord8_curr']] or '').strip(),
        'curFreq7': str(r_data[C['freq7_curr']] or '').strip(),
        'curNewFreq7': str(r_data[C['nfreq7_curr']] or '').strip(),
        'curMarketStatus': str(r_data[C['mkt_curr']] or '').strip(),
        'prevMarketStatus': str(r_data[C['mkt_prev']] or '').strip(),
        'curOperation': str(r_data[C['op_curr']] or '').strip(),
        'plpEnabled': 'Y' if str(r_data[C['plp_curr']] or '').strip().upper() == 'Y' else 'N',
        'plgFee': f"{round(num(r_data[C['plg_fee']]) * 100, 1)}%",
        'plgSpend': round(num(r_data[C['plg_spend']]), 2),
        'plgAdRev': round(num(r_data[C['plg_ad_rev']]), 2),
        'adClass': get_ad_class(r_data),
    })

# 2. cum43Stats — 含无对手已出单/未出单区分
rows_with_rival = [r for r in rows_curr if num(r[C['rival_curr']]) > 0]
rows_no_rival = [r for r in rows_curr if num(r[C['rival_curr']]) == 0]
y_curr = sum(1 for r in rows_with_rival if str(r[C['ord8_curr']] or '').strip() == 'Y')
n_curr = sum(1 for r in rows_with_rival if str(r[C['ord8_curr']] or '').strip() == 'N')
no_curr = sum(1 for r in rows_with_rival if str(r[C['ord8_curr']] or '').strip() == '未出单')
# 无对手已出单（有销量>0）
no_rival_sold = sum(1 for r in rows_no_rival if num(r[C['sales_curr']]) > 0)
no_rival_unsold = len(rows_no_rival) - no_rival_sold
normal_cnt = sum(1 for r in rows_curr if str(r[C['mkt_curr']] or '').strip() == '正常')
competitive_cnt = sum(1 for r in rows_curr if str(r[C['mkt_curr']] or '').strip() == '竞争无优势')
nomarket_cnt = sum(1 for r in rows_curr if str(r[C['mkt_curr']] or '').strip() == '无市场')
station_out_cnt = sum(1 for r in rows_curr if str(r[C['mkt_curr']] or '').strip() == '站外出单')

cum43Stats = {
    'total': total_sku, 'yCount': y_curr, 'nCount': n_curr, 'unCount': no_curr,
    'noRivalSold': no_rival_sold, 'noRivalUnsold': no_rival_unsold,
    'normalCount': normal_cnt, 'competitiveCount': competitive_cnt, 'noMarketCount': nomarket_cnt,
    'stationOutCount': station_out_cnt,
    'hasRivalCount': has_rival_curr, 'noRivalCount': total_sku - has_rival_curr,
    'totalMarketShare': total_market_share_curr,
    'totalMarketSharePrev': total_market_share_prev,
}

# 3. lowShareData
lowShareData = []
for r_data in rows_curr:
    share = num(r_data[C['share_curr']]) * 100
    rival = num(r_data[C['rival_curr']])
    if share < 75 and rival > 0:
        sku = str(r_data[C['sku']]).strip()
        lowShareData.append({
            'salesCode': str(r_data[C['sale_no']] or ''),
            'SKU': sku,
            'launchDate': str(get_date(r_data[C['list_date']]))[:10] if get_date(r_data[C['list_date']]) else '',
            'analyst': get_an(r_data),
            'category': get_cat(r_data),
            'expandType': get_exp(r_data),
            'curSalesQty': int(num(r_data[C['sales_curr']])),
            'salesQtyChange': ratio_str(num(r_data[C['sales_curr']]), num(r_data[C['sales_prev']])),
            'curRevenue': round(num(r_data[C['rev_curr']]), 2),
            'revenueChange': ratio_str(num(r_data[C['rev_curr']]), num(r_data[C['rev_prev']])),
            'prevRivalQty': int(num(r_data[C['rival_prev']])),
            'curRivalQty': int(num(r_data[C['rival_curr']])),
            'rivalQtyChange': ratio_str(num(r_data[C['rival_curr']]), num(r_data[C['rival_prev']])),
            'prevMarketShare': round(num(r_data[C['share_prev']]) * 100, 1),
            'curMarketShare': round(share, 1),
            'prevMarketStatus': str(r_data[C['mkt_prev']] or '').strip(),
            'curOperation': str(r_data[C['op_curr']] or '').strip(),
            'curMarketStatus': str(r_data[C['mkt_curr']] or '').strip(),
            'cur8dStatus': str(r_data[C['ord8_curr']] or '').strip(),
            'plpEnabled': 'Y' if str(r_data[C['plp_curr']] or '').strip().upper() == 'Y' else 'N',
            'plgFee': f"{round(num(r_data[C['plg_fee']]) * 100, 1)}%",
            'adClass': get_ad_class(r_data),
        })
lowShareData.sort(key=lambda x: x['curMarketShare'])

# 4. expandTypeData
exp_curr = defaultdict(lambda: {'sku': 0, 'sales_qty': 0, 'rev': 0, 'sold_sku': 0, 'has_rival': 0})
exp_prev = defaultdict(lambda: {'sku': 0, 'sales_qty': 0, 'rev': 0, 'sold_sku': 0, 'has_rival': 0})
for r_data in rows_curr:
    exp = get_exp(r_data)
    exp_curr[exp]['sku'] += 1
    exp_curr[exp]['sales_qty'] += num(r_data[C['sales_curr']])
    exp_curr[exp]['rev'] += num(r_data[C['rev_curr']])
    if num(r_data[C['rival_curr']]) > 0:
        exp_curr[exp]['has_rival'] += 1
        if str(r_data[C['ord8_curr']] or '').strip() in ('Y', 'N'):
            exp_curr[exp]['sold_sku'] += 1
for r_data in rows_prev:
    exp = get_exp(r_data)
    exp_prev[exp]['sku'] += 1
    exp_prev[exp]['sales_qty'] += num(r_data[C['sales_prev']])
    exp_prev[exp]['rev'] += num(r_data[C['rev_prev']])
    if num(r_data[C['rival_prev']]) > 0:
        exp_prev[exp]['has_rival'] += 1
        if str(r_data[C['ord8_prev']] or '').strip() in ('Y', 'N'):
            exp_prev[exp]['sold_sku'] += 1

expandTypeData = []
for exp in EXPAND_TYPES:
    if exp in exp_curr or exp in exp_prev:
        d_c = exp_curr.get(exp, {})
        d_p = exp_prev.get(exp, {})
        cur_sold = d_c.get('sold_sku', 0); has_r_c = d_c.get('has_rival', 0) or 1
        prev_sold = d_p.get('sold_sku', 0); has_r_p = d_p.get('has_rival', 0) or 1
        expandTypeData.append({
            'expandType': exp,
            'curSku': d_c.get('sku', 0), 'prevSku': d_p.get('sku', 0),
            'curSalesSku': cur_sold,
            'curSalesRate': f"{round(cur_sold/has_r_c*100, 1)}%",
            'prevSalesSku': prev_sold,
            'prevSalesRate': f"{round(prev_sold/has_r_p*100, 1)}%",
            'curSalesQty': int(d_c.get('sales_qty', 0)), 'prevSalesQty': int(d_p.get('sales_qty', 0)),
            'salesQtyChange': ratio_str(d_c.get('sales_qty', 0), d_p.get('sales_qty', 0)),
            'curRevenue': round(d_c.get('rev', 0), 2), 'prevRevenue': round(d_p.get('rev', 0), 2),
            'revenueChange': ratio_str(d_c.get('rev', 0), d_p.get('rev', 0)),
        })

# 5. timelinessData
an_time = defaultdict(lambda: {'timely': 0, 'no_8d': 0, 'no_7d': 0})
for r_data in rows_curr:
    an = get_an(r_data)
    freq = str(r_data[C['freq7_curr']] or '').strip()
    nfreq = str(r_data[C['nfreq7_curr']] or '').strip()
    if nfreq == '异常': an_time[an]['no_8d'] += 1
    elif freq == '异常': an_time[an]['no_7d'] += 1
    else: an_time[an]['timely'] += 1

an_time_prev = defaultdict(lambda: {'timely': 0, 'no_8d': 0, 'no_7d': 0})
for r_data in rows_prev:
    an = get_an(r_data)
    freq = str(r_data[C['freq7_prev']] or '').strip()
    nfreq = str(r_data[C['nfreq7_prev']] or '').strip()
    if nfreq == '异常': an_time_prev[an]['no_8d'] += 1
    elif freq == '异常': an_time_prev[an]['no_7d'] += 1
    else: an_time_prev[an]['timely'] += 1

ans_ordered = [a for a in ANALYSTS if a in an_time] + [a for a in sorted(an_time.keys()) if a not in ANALYSTS]
timeliness_total_curr = sum(1 for r in rows_curr if str(r[C['nfreq7_curr']] or '').strip() != '异常' and str(r[C['freq7_curr']] or '').strip() != '异常')
timeliness_total_prev = sum(1 for r in rows_prev if str(r[C['nfreq7_prev']] or '').strip() != '异常' and str(r[C['freq7_prev']] or '').strip() != '异常')

timelinessData = {'analysts': [], 'total': {}}
for an in ans_ordered:
    d_c = an_time.get(an, {'timely': 0, 'no_8d': 0, 'no_7d': 0})
    d_p = an_time_prev.get(an, {'timely': 0, 'no_8d': 0, 'no_7d': 0})
    total_c = d_c['timely'] + d_c['no_8d'] + d_c['no_7d']
    total_p = d_p['timely'] + d_p['no_8d'] + d_p['no_7d']
    rate_c = f"{round(d_c['timely']/total_c*100, 1)}%" if total_c else '0%'
    rate_p = f"{round(d_p['timely']/total_p*100, 1)}%" if total_p else '0%'
    timelinessData['analysts'].append({
        'analyst': an, 'curSku': total_c, 'prevSku': total_p,
        'timelyCount': d_c['timely'], 'prevTimelyCount': d_p['timely'],
        'noAnalysis8dCount': d_c['no_8d'], 'noAnalysis7dCount': d_c['no_7d'],
        'timelyRate': rate_c, 'prevTimelyRate': rate_p,
        'change': ratio_str(float(rate_c.replace('%','')), float(rate_p.replace('%',''))) if rate_c != '0%' and rate_p != '0%' else '-',
    })
tc_all = timeliness_total_curr
tp_all = timeliness_total_prev
tr_c = f"{round(tc_all/total_sku*100, 1)}%" if total_sku else '0%'
tr_p = f"{round(tp_all/len(rows_prev)*100, 1)}%" if rows_prev else '0%'
timelinessData['total'] = {
    'analyst': '合计', 'curSku': total_sku, 'prevSku': len(rows_prev),
    'timelyCount': tc_all, 'prevTimelyCount': tp_all,
    'noAnalysis8dCount': sum(1 for r in rows_curr if str(r[C['nfreq7_curr']] or '').strip() == '异常'),
    'noAnalysis7dCount': sum(1 for r in rows_curr if str(r[C['freq7_curr']] or '').strip() == '异常'),
    'timelyRate': tr_c, 'prevTimelyRate': tr_p,
    'change': ratio_str(float(tr_c.replace('%','')), float(tr_p.replace('%',''))) if tr_c != '0%' and tr_p != '0%' else '-',
}

# 6. hasCompetitorUnsold
has_rival_no_curr = [r for r in rows_curr if str(r[C['ord8_curr']] or '').strip() == '未出单' and num(r[C['rival_curr']]) > 0]
has_rival_no_prev = [r for r in rows_prev if str(r[C['ord8_prev']] or '').strip() == '未出单' and num(r[C['rival_prev']]) > 0]
mkt_has_order = ['竞争无优势', '无市场', '站内无价格优势', '站外出单', '正常', '#N/A', '未知']
mkt_no_order = ['无市场', '未知', '竞争无优势', '#N/A', '其他']

def build_unsold_analysis(items, mkt_order, col_mkt):
    reasons = []
    mkt_counter = Counter(str(r[col_mkt] or '未知') for r in items)
    for mkt in mkt_order:
        reasons.append({'name': mkt, 'count': mkt_counter.get(mkt, 0)})
    by_analyst = []
    for an in ans_ordered:
        row = {'analyst': an}
        an_items = [r for r in items if get_an(r) == an]
        an_mkt = Counter(str(r[col_mkt] or '未知') for r in an_items)
        for mkt in mkt_order:
            row[mkt] = an_mkt.get(mkt, 0)
        row['total'] = len(an_items)
        by_analyst.append(row)
    by_category = []
    cat_items_map = defaultdict(list)
    for r in items: cat_items_map[get_cat(r)].append(r)
    for cat in sorted(cat_items_map.keys()):
        row = {'category': cat}
        cat_mkt = Counter(str(r[col_mkt] or '未知') for r in cat_items_map[cat])
        for mkt in mkt_order:
            row[mkt] = cat_mkt.get(mkt, 0)
        row['total'] = len(cat_items_map[cat])
        by_category.append(row)
    return {
        'total': len(items), 'prevTotal': 0, 'reasons': reasons,
        'byAnalyst': by_analyst, 'byCategory': by_category,
    }

hasCompetitorUnsold = build_unsold_analysis(has_rival_no_curr, mkt_has_order, C['mkt_curr'])
hasCompetitorUnsold['prevTotal'] = len(has_rival_no_prev)
hasCompetitorUnsold['change'] = len(has_rival_no_curr) - len(has_rival_no_prev)

# 7-11. PLP 广告数据（自然周）
plpTotal = {
    'campaignCount': plp_t['campaigns'], 'linkCount': plp_t['links'],
    'impression': int(plp_t['imp']), 'click': int(plp_t['click']),
    'sold': int(plp_t['sold']), 'cost': round(plp_t['cost'], 2),
    'revenue': round(plp_t['ad_rev'], 2), 'totalRevenue': round(dedup_total_rev, 2),
    'roas': f"{plp_t['roas']:.2f}", 'cvr': f"{plp_t['cvr']*100:.2f}%",
    'ctr': f"{plp_t['ctr']*100:.2f}%", 'cpc': f"${plp_t['cpc']:.2f}",
    'cpa': f"${plp_t['cpa']:.2f}", 'acos': f"{plp_t['acos']*100:.2f}%",
    'acoas': f"{dedup_acoas*100:.2f}%",
}

# PLP by dimension
def plp_dim_data(by_sku, key_fn):
    groups = defaultdict(lambda: {'imp': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'total_rev': 0, 'campaigns': set(), 'skus': set()})
    for sku, d in by_sku.items():
        info = None
        for r in rows_curr:
            if str(r[C['sku']]).strip() == sku:
                info = r; break
        k = key_fn(sku, info) if info else key_fn(sku, None)
        g = groups[k]
        g['imp'] += d['impr']; g['click'] += d['click']; g['sold'] += d['sold']
        g['cost'] += d['cost']; g['ad_rev'] += d['ad_rev']
        g['campaigns'].update(d['campaigns']); g['skus'].add(sku)
    for k, g in groups.items():
        dedup_rev = 0
        for sku in g['skus']:
            dedup_rev += sku_total_rev_dedup.get(sku, 0)
        g['total_rev'] = dedup_rev
    result = []
    for k, g in sorted(groups.items()):
        roas_v = calc_rate(g['ad_rev'], g['cost'])
        cvr_v = calc_rate(g['sold'], g['click'])
        ctr_v = calc_rate(g['click'], g['imp'])
        acos_v = calc_rate(g['cost'], g['ad_rev'])
        acoas_v = calc_rate(g['cost'], g['total_rev'])
        result.append({
            'name': k, 'campaignCount': len(g['campaigns']), 'linkCount': len(g['skus']),
            'impression': int(g['imp']), 'click': int(g['click']), 'sold': int(g['sold']),
            'cost': round(g['cost'], 2), 'revenue': round(g['ad_rev'], 2),
            'totalRevenue': round(g['total_rev'], 2),
            'roas': f"{roas_v:.2f}", 'cvr': f"{cvr_v*100:.2f}%",
            'ctr': f"{ctr_v*100:.2f}%", 'cpc': f"${round(g['cost']/g['click'], 2)}" if g['click'] else '-',
            'cpa': f"${round(g['cost']/g['sold'], 2)}" if g['sold'] else '-',
            'acos': f"{acos_v*100:.2f}%", 'acoas': f"{acoas_v*100:.2f}%",
        })
    return result

plpCategories = plp_dim_data(plp_curr_sku, lambda sku, info: get_cat(info) if info else '未分类')
plpExpandTypes = plp_dim_data(plp_curr_sku, lambda sku, info: get_exp(info) if info else '其他')
plpAnalysts = plp_dim_data(plp_curr_sku, lambda sku, info: get_an(info) if info else '未知')

# PLP prev totals (自然周)
plp_prev_rows_list = []
for row in ws_plp.iter_rows(min_row=2, values_only=True):
    period = str(row[PC['period']] or '').strip()
    sku = str(row[PC['sku']] or '').strip()
    if not sku or sku.startswith('广告') or sku.startswith('总数据'): continue
    list_d = get_date(row[PC['list_date']])
    if period == PLP_PREV and list_d and list_d <= cutoff_prev:
        plp_prev_rows_list.append(list(row))
plp_prev_sku = agg_plp(plp_prev_rows_list)
plp_pt = plp_totals(plp_prev_sku)
plpPrevTotal = {
    'campaignCount': plp_pt['campaigns'], 'linkCount': plp_pt['links'],
    'impression': int(plp_pt['imp']), 'click': int(plp_pt['click']),
    'sold': int(plp_pt['sold']), 'cost': round(plp_pt['cost'], 2),
    'revenue': round(plp_pt['ad_rev'], 2), 'totalRevenue': round(plp_pt['total_rev'], 2),
    'roas': f"{plp_pt['roas']:.2f}", 'cvr': f"{plp_pt['cvr']*100:.2f}%",
    'ctr': f"{plp_pt['ctr']*100:.2f}%", 'cpc': f"${plp_pt['cpc']:.2f}",
    'cpa': f"${plp_pt['cpa']:.2f}", 'acos': f"{plp_pt['acos']*100:.2f}%",
    'acoas': f"{plp_pt['acoas']*100:.2f}%",
}

# 12. unsoldNoCompetitor
no_rival_no_curr = [r for r in rows_curr if str(r[C['ord8_curr']] or '').strip() == '未出单' and num(r[C['rival_curr']]) == 0]
no_rival_no_prev = [r for r in rows_prev if str(r[C['ord8_prev']] or '').strip() == '未出单' and num(r[C['rival_prev']]) == 0]

unsoldNoCompetitor = build_unsold_analysis(no_rival_no_curr, mkt_no_order, C['mkt_curr'])
unsoldNoCompetitor['prevTotal'] = len(no_rival_no_prev)
unsoldNoCompetitor['change'] = len(no_rival_no_curr) - len(no_rival_no_prev)

# 13. prevWeekKpi
prevWeekKpi = {
    'prevTotalSku': len(rows_prev),
    'prevTotalSalesQty': total_sales_prev,
    'prevTotalRevenue': round(total_rev_prev, 2),
    'salesQtyChange': ratio_str(total_sales_curr, total_sales_prev),
    'revenueChange': ratio_str(total_rev_curr, total_rev_prev),
    'skuChange': ratio_str(total_sku, len(rows_prev)),
    'prevTimelyRate': tr_p,
    'prevSoldRate': f"{round((y_curr+n_curr)/has_rival_curr*100, 1)}%" if has_rival_curr else '0%',
    'prevLowShareCount': len(lowShareData),
    'deptRatio': f"{dept_ratio:.1f}%",
    'deptTotalRevenue': round(dept_total_revenue, 2),
    'totalMarketShare': f"{total_market_share_curr}%",
    'totalMarketSharePrev': f"{total_market_share_prev}%",
    'marketShareChange': ratio_str(total_market_share_curr, total_market_share_prev),
}

# 14. plgStats（含PLG花费和ACOS/ACOAS）
# PLG汇总：从四三数据累计表聚合
plg_total_spend = sum(num(r[C['plg_spend']]) for r in rows_curr)
plg_total_ad_rev = sum(num(r[C['plg_ad_rev']]) for r in rows_curr)
# PLG ACOAS：PLG花费 / 对应SKU的总销售额（自然周）
plg_sku_nat_rev = {}
for r in rows_curr:
    sku = str(r[C['sku']]).strip()
    plg_spend_sku = num(r[C['plg_spend']])
    if plg_spend_sku > 0:
        nw = nat_week_data.get(sku, {'qty': 0, 'rev': 0})
        plg_sku_nat_rev[sku] = nw['rev']

plg_total_nat_rev = sum(plg_sku_nat_rev.values())
plg_acos = round(plg_total_spend / plg_total_ad_rev * 100, 2) if plg_total_ad_rev > 0 else 0
plg_acoas = round(plg_total_spend / plg_total_nat_rev * 100, 2) if plg_total_nat_rev > 0 else 0

plg_categories_map = {'PLP+PLG同开': [], '单链接PLP+PLG同开': [], '单PLG': [], '单PLP': [], '单PLG且未出单': [], '无广告': []}
for r_data in rows_curr:
    adc = get_ad_class(r_data)
    if adc in plg_categories_map:
        plg_categories_map[adc].append(r_data)

plgStats = {
    'plpAndPlgBothCount': len(plg_categories_map['PLP+PLG同开']),
    'singleLinkPlpPlgCount': len(plg_categories_map['单链接PLP+PLG同开']),
    'plgOnlyCount': len(plg_categories_map['单PLG']),
    'plpOnlyCount': len(plg_categories_map['单PLP']),
    'noAdCount': len(plg_categories_map['无广告']),
    'plpDisabledNoSaleCount': len(plg_categories_map['单PLG且未出单']),
    'totalNewProducts': total_sku,
    'totalSpend': round(plg_total_spend, 2),
    'totalAdRev': round(plg_total_ad_rev, 2),
    'totalNatRev': round(plg_total_nat_rev, 2),
    'acos': f"{plg_acos:.2f}%",
    'acoas': f"{plg_acoas:.2f}%",
    'byAnalyst': [],
}

# PLG by analyst
for an in ans_ordered:
    an_items = [r for r in rows_curr if get_an(r) == an]
    an_ad = Counter(get_ad_class(r) for r in an_items)
    an_plg_spend = sum(num(r[C['plg_spend']]) for r in an_items)
    an_plg_ad_rev = sum(num(r[C['plg_ad_rev']]) for r in an_items)
    an_plg_nat_rev = 0
    for r in an_items:
        sku = str(r[C['sku']]).strip()
        if num(r[C['plg_spend']]) > 0:
            an_plg_nat_rev += nat_week_data.get(sku, {'qty': 0, 'rev': 0})['rev']
    an_acos = round(an_plg_spend / an_plg_ad_rev * 100, 2) if an_plg_ad_rev > 0 else 0
    an_acoas = round(an_plg_spend / an_plg_nat_rev * 100, 2) if an_plg_nat_rev > 0 else 0
    plgStats['byAnalyst'].append({
        'analyst': an, 'total': len(an_items),
        'plpAndPlgBoth': an_ad.get('PLP+PLG同开', 0),
        'singleLinkPlpPlg': an_ad.get('单链接PLP+PLG同开', 0),
        'plgOnly': an_ad.get('单PLG', 0), 'plpOnly': an_ad.get('单PLP', 0),
        'noAd': an_ad.get('无广告', 0), 'plpDisabledNoSale': an_ad.get('单PLG且未出单', 0),
        'plgSpend': round(an_plg_spend, 2),
        'plgAdRev': round(an_plg_ad_rev, 2),
        'plgNatRev': round(an_plg_nat_rev, 2),
        'acos': f"{an_acos:.2f}%",
        'acoas': f"{an_acoas:.2f}%",
    })

# 15-16. categoryRevenueData / analystRevenueData（含市占比）
categoryRevenueData = []
cat_curr_data = defaultdict(lambda: {'sku': 0, 'new': 0, 'sales': 0, 'rev': 0, 'rival_sales': 0})
cat_prev_data = defaultdict(lambda: {'sku': 0, 'sales': 0, 'rev': 0, 'rival_sales': 0})
for r_data in rows_curr:
    cat = get_cat(r_data)
    list_d = get_date(r_data[C['list_date']])
    cat_curr_data[cat]['sku'] += 1
    if list_d and list_d > cutoff_prev: cat_curr_data[cat]['new'] += 1
    cat_curr_data[cat]['sales'] += num(r_data[C['sales_curr']])
    cat_curr_data[cat]['rev'] += num(r_data[C['rev_curr']])
    cat_curr_data[cat]['rival_sales'] += num(r_data[C['rival_curr']])
for r_data in rows_prev:
    cat = get_cat(r_data)
    cat_prev_data[cat]['sku'] += 1
    cat_prev_data[cat]['sales'] += num(r_data[C['sales_prev']])
    cat_prev_data[cat]['rev'] += num(r_data[C['rev_prev']])
    cat_prev_data[cat]['rival_sales'] += num(r_data[C['rival_prev']])

cats_ordered = [c for c in CATEGORIES if c in cat_curr_data] + [c for c in sorted(cat_curr_data.keys()) if c not in CATEGORIES]
for cat in cats_ordered:
    d_c = cat_curr_data.get(cat, {})
    d_p = cat_prev_data.get(cat, {})
    cat_sales = d_c.get('sales', 0)
    cat_rival = d_c.get('rival_sales', 0)
    cat_share = round(cat_sales / (cat_sales + cat_rival) * 100, 1) if (cat_sales + cat_rival) > 0 else 0
    cat_sales_p = d_p.get('sales', 0)
    cat_rival_p = d_p.get('rival_sales', 0)
    cat_share_p = round(cat_sales_p / (cat_sales_p + cat_rival_p) * 100, 1) if (cat_sales_p + cat_rival_p) > 0 else 0
    categoryRevenueData.append({
        'category': cat, 'curSku': d_c.get('sku', 0), 'curNewSku': d_c.get('new', 0),
        'curSalesQty': int(d_c.get('sales', 0)), 'prevSalesQty': int(d_p.get('sales', 0)),
        'salesQtyChange': ratio_str(d_c.get('sales', 0), d_p.get('sales', 0)),
        'curRevenue': round(d_c.get('rev', 0), 2), 'prevRevenue': round(d_p.get('rev', 0), 2),
        'revenueChange': ratio_str(d_c.get('rev', 0), d_p.get('rev', 0)),
        'curMarketShare': cat_share, 'prevMarketShare': cat_share_p,
        'marketShareChange': ratio_str(cat_share, cat_share_p),
    })

analystRevenueData = []
an_curr_data = defaultdict(lambda: {'sku': 0, 'new': 0, 'sales': 0, 'rev': 0, 'rival_sales': 0})
an_prev_data = defaultdict(lambda: {'sku': 0, 'sales': 0, 'rev': 0, 'rival_sales': 0})
for r_data in rows_curr:
    an = get_an(r_data)
    list_d = get_date(r_data[C['list_date']])
    an_curr_data[an]['sku'] += 1
    if list_d and list_d > cutoff_prev: an_curr_data[an]['new'] += 1
    an_curr_data[an]['sales'] += num(r_data[C['sales_curr']])
    an_curr_data[an]['rev'] += num(r_data[C['rev_curr']])
    an_curr_data[an]['rival_sales'] += num(r_data[C['rival_curr']])
for r_data in rows_prev:
    an = get_an(r_data)
    an_prev_data[an]['sku'] += 1
    an_prev_data[an]['sales'] += num(r_data[C['sales_prev']])
    an_prev_data[an]['rev'] += num(r_data[C['rev_prev']])
    an_prev_data[an]['rival_sales'] += num(r_data[C['rival_prev']])
for an in ans_ordered:
    d_c = an_curr_data.get(an, {})
    d_p = an_prev_data.get(an, {})
    an_sales = d_c.get('sales', 0)
    an_rival = d_c.get('rival_sales', 0)
    an_share = round(an_sales / (an_sales + an_rival) * 100, 1) if (an_sales + an_rival) > 0 else 0
    an_sales_p = d_p.get('sales', 0)
    an_rival_p = d_p.get('rival_sales', 0)
    an_share_p = round(an_sales_p / (an_sales_p + an_rival_p) * 100, 1) if (an_sales_p + an_rival_p) > 0 else 0
    analystRevenueData.append({
        'analyst': an, 'curSku': d_c.get('sku', 0), 'curNewSku': d_c.get('new', 0),
        'curSalesQty': int(d_c.get('sales', 0)), 'prevSalesQty': int(d_p.get('sales', 0)),
        'salesQtyChange': ratio_str(d_c.get('sales', 0), d_p.get('sales', 0)),
        'curRevenue': round(d_c.get('rev', 0), 2), 'prevRevenue': round(d_p.get('rev', 0), 2),
        'revenueChange': ratio_str(d_c.get('rev', 0), d_p.get('rev', 0)),
        'curMarketShare': an_share, 'prevMarketShare': an_share_p,
        'marketShareChange': ratio_str(an_share, an_share_p),
    })

# 17. plgRecords
plgRecords = []
for r_data in rows_curr:
    sku = str(r_data[C['sku']]).strip()
    list_d = get_date(r_data[C['list_date']])
    first_d = get_date(r_data[C['first_order']])
    ac = get_ad_class(r_data)
    plgRecords.append({
        'salesCode': str(r_data[C['sale_no']] or ''), 'SKU': sku,
        'launchDate': str(list_d)[:10] if list_d else '',
        'firstSaleDate': str(first_d)[:10] if first_d else '未出单',
        'analyst': get_an(r_data), 'category': get_cat(r_data),
        'expandType': get_exp(r_data),
        'cur8dStatus': str(r_data[C['ord8_curr']] or '').strip(),
        'curSalesQty': int(num(r_data[C['sales_curr']])),
        'curRevenue': round(num(r_data[C['rev_curr']]), 2),
        'curRivalQty': int(num(r_data[C['rival_curr']])),
        'curMarketShare': f"{round(num(r_data[C['share_curr']]) * 100, 1)}%",
        'curMarketStatus': str(r_data[C['mkt_curr']] or '').strip(),
        'curOperation': str(r_data[C['op_curr']] or '').strip(),
        'plpEnabled': 'Y' if str(r_data[C['plp_curr']] or '').strip().upper() == 'Y' else 'N',
        'plgFee': f"{round(num(r_data[C['plg_fee']]) * 100, 1)}%",
        'adClass': ac,
    })

# 18. plpSummaryData
plpSummaryData = []
for d in plpCategories:
    d['dimType'] = '品线'; plpSummaryData.append(d)
for d in plpExpandTypes:
    d['dimType'] = '拓展类型'; plpSummaryData.append(d)
for d in plpAnalysts:
    d['dimType'] = '分析人'; plpSummaryData.append(d)

# 19. plpDetailData
plpDetailData = []
for row in plp_curr_rows:
    sku = str(row[PC['sku']] or '').strip()
    info = None
    for r in rows_curr:
        if str(r[C['sku']]).strip() == sku:
            info = r; break
    cost = num(row[PC['cost']]); ad_rev = num(row[PC['ad_rev']])
    total_r = num(row[PC['total_rev']]); impr = num(row[PC['impr']])
    click = num(row[PC['click']]); sold = num(row[PC['sold']])
    plg_y = str(row[PC['plg_enabled']] or '').strip() == 'Y'
    roas_v = calc_rate(ad_rev, cost)
    cvr_v = calc_rate(sold, click)
    ctr_v = calc_rate(click, impr)
    acos_v = calc_rate(cost, ad_rev)
    acoas_v = calc_rate(cost, total_r)
    ad_class = get_ad_class(info) if info else '无广告'
    plpDetailData.append({
        'period': row[PC['period']] or '', 'campaign': row[PC['campaign']] or '',
        'SKU': sku, 'id': str(row[PC['id']] or ''), 'store': row[PC['store']] or '',
        'plpStartDate': str(get_date(row[PC['plp_start']]))[:10] if get_date(row[PC['plp_start']]) else '',
        'listDate': str(get_date(row[PC['list_date']]))[:10] if get_date(row[PC['list_date']]) else '',
        'firstOrderDate': str(get_date(row[PC['first_order']]))[:10] if get_date(row[PC['first_order']]) else '',
        'analyst': get_an(info) if info else (row[PC['analyst']] or ''),
        'category': get_cat(info) if info else (row[PC['category']] or '未分类'),
        'expandType': get_exp(info) if info else (row[PC['expand_type']] or ''),
        'impressions': int(impr), 'clicks': int(click), 'salesQty': int(sold),
        'spend': round(cost, 2), 'adRevenue': round(ad_rev, 2),
        'totalRevenue': round(total_r, 2),
        'roas': roas_v, 'cvr': cvr_v, 'ctr': ctr_v,
        'cpc': round(cost / click, 3) if click else 0,
        'cpa': round(cost / sold, 2) if sold else 0,
        'acos': acos_v, 'acoas': acoas_v,
        'plgEnabled': 'Y' if plg_y else 'N',
        'adClass': ad_class,
    })
plpDetailData.sort(key=lambda x: (x['analyst'], x['SKU']))

# ===== 市占比分布 by 品线 =====
shareTierByCat = []
for cat in cats_ordered:
    cat_rows = [r for r in rows_curr if get_cat(r) == cat]
    high = sum(1 for r in cat_rows if num(r[C["share_curr"]]) * 100 >= 75)
    mid = sum(1 for r in cat_rows if 50 <= num(r[C["share_curr"]]) * 100 < 75)
    low = sum(1 for r in cat_rows if num(r[C["share_curr"]]) * 100 < 50)
    shareTierByCat.append({"category": cat, "high": high, "mid": mid, "low": low, "total": len(cat_rows)})

shareTierOverview = {
    "tiers": ["高市占比(>=75%)", "中市占比(50-75%)", "低市占比(<50%)"],
    "byCategory": shareTierByCat,
}

# ===== 市场分布数据 =====
def normalize_mkt(r_data, col_idx):
    v = str(r_data[col_idx] or '').strip()
    return v if v in ALL_MKT_STATUSES else '其他'

mkt_curr_counter = Counter()
mkt_prev_counter = Counter()
for r_data in rows_curr:
    mkt_curr_counter[normalize_mkt(r_data, C['mkt_curr'])] += 1
for r_data in rows_prev:
    mkt_prev_counter[normalize_mkt(r_data, C['mkt_prev'])] += 1

mktDistOverall = {
    'curTotal': total_sku,
    'prevTotal': len(rows_prev),
    'distribution': []
}
for s in ALL_MKT_STATUSES:
    cur_c = mkt_curr_counter.get(s, 0)
    prev_c = mkt_prev_counter.get(s, 0)
    mktDistOverall['distribution'].append({
        'status': s,
        'curCount': cur_c,
        'prevCount': prev_c,
        'curPct': round(cur_c / total_sku * 100, 1) if total_sku else 0,
        'prevPct': round(prev_c / len(rows_prev) * 100, 1) if rows_prev else 0,
        'change': cur_c - prev_c,
    })

# ===== 序列化 JS 变量 =====
print("序列化数据...")
data_blocks = {
    'cum43Data': cum43Data, 'cum43Stats': cum43Stats, 'lowShareData': lowShareData,
    'expandTypeData': expandTypeData, 'timelinessData': timelinessData,
    'hasCompetitorUnsold': hasCompetitorUnsold, 'plpTotal': plpTotal,
    'plpPrevTotal': plpPrevTotal, 'plpCategories': plpCategories,
    'plpExpandTypes': plpExpandTypes, 'plpAnalysts': plpAnalysts,
    'unsoldNoCompetitor': unsoldNoCompetitor, 'prevWeekKpi': prevWeekKpi,
    'plgStats': plgStats, 'categoryRevenueData': categoryRevenueData,
    'analystRevenueData': analystRevenueData, 'plgRecords': plgRecords,
    'plpSummaryData': plpSummaryData, 'plpDetailData': plpDetailData,
    'mktDistOverall': mktDistOverall, 'shareTierOverview': shareTierOverview,
}

js_lines = []
for key, val in data_blocks.items():
    js_lines.append(f"const {key} = {json.dumps(val, ensure_ascii=False, indent=0)};")

js_vars_block = '\n'.join(js_lines)

print(f"数据块: {len(data_blocks)} 个")
print(f"JS变量总大小: {len(js_vars_block)} 字符")

# ===== HTML CSS 模板 =====
CSS = r'''
<style>
* { margin: 0; padding: 0; box-sizing: border-box; }
body { font-family: 'Segoe UI', '微软雅黑', sans-serif; background: #f0f2f5; color: #1a1a2e; display: flex; min-height: 100vh; }
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #f1f1f1; }
::-webkit-scrollbar-thumb { background: #ccc; border-radius: 3px; }

.sidebar { width: 230px; min-width: 230px; background: #fff; height: 100vh; position: sticky; top: 0; overflow-y: auto; box-shadow: 2px 0 8px rgba(0,0,0,0.06); z-index: 10; }
.sidebar h2 { font-size: 16px; color: #0f3460; padding: 20px 16px 12px; border-bottom: 2px solid #e8f0fe; }
.sidebar ul { list-style: none; padding: 8px 0; }
.sidebar li a { display: block; padding: 12px 16px; color: #555; text-decoration: none; font-size: 13px; border-left: 3px solid transparent; transition: all 0.2s; }
.sidebar li a:hover, .sidebar li a.active { background: #f0f6ff; color: #0f3460; border-left-color: #0f3460; font-weight: 600; }

.main { flex: 1; padding: 24px; overflow: auto; }
.header { background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%); background-size: 200% 200%; animation: gradientShift 8s ease infinite; color: #fff; padding: 28px 40px; border-radius: 10px; margin-bottom: 20px; }
.header h1 { font-size: 24px; font-weight: 700; letter-spacing: 2px; }
.header p { font-size: 13px; opacity: 0.75; margin-top: 6px; }

.tab-content { display: none; }
.tab-content.active { display: block; animation: fadeIn 0.4s ease-out; }

.kpi-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 12px; margin-bottom: 20px; }
.kpi-card { background: #fff; border-radius: 10px; padding: 16px 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); text-align: center; transition: transform 0.2s, box-shadow 0.2s; animation: fadeInUp 0.5s ease-out forwards; }
.kpi-card:hover { transform: translateY(-2px); box-shadow: 0 4px 16px rgba(0,0,0,0.12); }
.kpi-card .label { font-size: 11px; color: #888; margin-bottom: 6px; }
.kpi-card .val { font-size: 24px; font-weight: 700; }
.kpi-card .hb { font-size: 11px; margin-top: 4px; }
.kpi-card.primary .val { color: #0f3460; }
.kpi-card.success .val { color: #08845a; }
.kpi-card.warning .val { color: #e07b24; }
.kpi-card.danger .val { color: #c0392b; }
.kpi-card.info .val { color: #2980b9; }
.kpi-card.purple .val { color: #8e44ad; }
.kpi-card.animate-pulse { animation: pulse 2s ease-in-out infinite; }

.section { background: #fff; border-radius: 10px; padding: 20px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); animation: fadeInUp 0.55s ease-out forwards; }
.section h3 { font-size: 15px; color: #0f3460; margin-bottom: 16px; padding-bottom: 8px; border-bottom: 2px solid #e8f0fe; }
.sub-module { margin-bottom: 16px; }
.sub-module h4 { font-size: 13px; color: #0f3460; background: #f5f7ff; padding: 8px 12px; border-left: 3px solid #0f3460; margin-bottom: 10px; }

.chart-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 20px; }
.chart-box { background: #fff; border-radius: 10px; padding: 16px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); animation: fadeInUp 0.6s ease-out forwards; }
.chart-box h4 { font-size: 13px; color: #0f3460; margin-bottom: 12px; }
.chart-box canvas { max-height: 280px; }

.table-scroll-wrap { max-height: 68vh; overflow-y: auto; overflow-x: auto; border: 1px solid #e0e0e0; border-radius: 6px; }
.table-scroll-wrap thead th { position: sticky; top: 0; z-index: 10; }
.table-scroll-wrap tfoot td { position: sticky; bottom: 0; z-index: 10; background: #f0f6ff; font-weight: 700; }

.data-table { width: 100%; border-collapse: collapse; font-size: 12px; }
.data-table th { background: #0f3460; color: #fff; padding: 8px 10px; font-weight: 600; white-space: nowrap; text-align: center; }
.data-table td { padding: 6px 10px; border-bottom: 1px solid #eee; text-align: center; }
.data-table tr:hover td { background: #f5f7ff; }
.data-table .num { text-align: right; }
.data-table .total-row td { font-weight: 700; background: #e8f0fe; }

.hb-up { color: #08845a; font-weight: 600; }
.hb-down { color: #c0392b; font-weight: 600; }
.hb-flat { color: #888; }

.badge { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 10px; color: #fff; font-weight: 600; }
.badge-green { background: #08845a; }
.badge-orange { background: #e07b24; }
.badge-red { background: #c0392b; }
.badge-blue { background: #2980b9; }
.badge-purple { background: #8e44ad; }
.badge-gray { background: #6c757d; }

.filter-bar { background: #f5f7ff; padding: 12px 16px; border-radius: 8px; margin-bottom: 16px; display: flex; flex-wrap: wrap; gap: 10px; align-items: center; }
.filter-bar-sticky { position: sticky; top: -1px; z-index: 100; background: #f5f7ff; border-radius: 0; padding: 14px 18px; margin-bottom: 14px; display: flex; flex-wrap: wrap; gap: 10px; align-items: center; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }
.filter-bar select, .filter-bar .fg, .filter-bar-sticky .fg { display: flex; align-items: center; gap: 4px; font-size: 12px; }
.filter-bar .fg label, .filter-bar-sticky .fg label { color: #555; white-space: nowrap; }
.filter-bar-sticky select, .filter-bar select { padding: 6px 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 12px; background: #fff; }
.filter-bar-sticky select:focus, .filter-bar select:focus { border-color: #0f3460; outline: none; }
.filter-bar .reset-btn, .filter-bar-sticky .reset-btn { color: #c0392b; border: 1px solid #c0392b; background: #fff; padding: 6px 14px; border-radius: 6px; cursor: pointer; font-size: 12px; }
.filter-bar .reset-btn:hover, .filter-bar-sticky .reset-btn:hover { background: #c0392b; color: #fff; }
.filter-bar .count, .filter-bar-sticky .count { font-size: 12px; color: #888; margin-left: auto; }

@keyframes fadeInUp { from { opacity: 0; transform: translateY(24px); } to { opacity: 1; transform: translateY(0); } }
@keyframes fadeIn { from { opacity: 0; } to { opacity: 1; } }
@keyframes pulse { 0%,100% { transform: scale(1); } 50% { transform: scale(1.04); } }
@keyframes gradientShift { 0% { background-position: 0% 50%; } 50% { background-position: 100% 50%; } 100% { background-position: 0% 50%; } }

.report-block { background: #f9fafb; border-radius: 8px; padding: 14px 18px; margin-bottom: 12px; border-left: 4px solid #0f3460; position: relative; }
.report-block h4 { font-size: 13px; color: #0f3460; margin-bottom: 8px; }
.report-block pre { white-space: pre-wrap; font-size: 12px; color: #444; font-family: '微软雅黑', sans-serif; margin: 0; }
.report-block .copy-btn { position: absolute; top: 12px; right: 16px; background: #0f3460; color: #fff; border: none; padding: 4px 12px; border-radius: 4px; cursor: pointer; font-size: 11px; transition: background 0.2s; }
.report-block .copy-btn:hover { background: #16213e; }

.risk-high { background: #fff5f5; border-left-color: #c0392b; }
.risk-high h4 { color: #c0392b; }
.risk-medium { background: #fffdf5; border-left-color: #e07b24; }
.risk-medium h4 { color: #e07b24; }
.risk-low { background: #f5fff5; border-left-color: #08845a; }
.risk-low h4 { color: #08845a; }

.findings-card { background: #f5f9ff; border-radius: 8px; padding: 14px 18px; margin-bottom: 10px; border-left: 4px solid #2980b9; }
.findings-card .title { font-size: 13px; font-weight: 600; color: #0f3460; margin-bottom: 4px; }
.findings-card .desc { font-size: 12px; color: #555; }

.action-card { background: #fdf5ff; border-radius: 8px; padding: 14px 18px; margin-bottom: 10px; border-left: 4px solid #8e44ad; }
.action-card .title { font-size: 13px; font-weight: 600; color: #8e44ad; margin-bottom: 4px; }
.action-card .desc { font-size: 12px; color: #555; }

@media(max-width:900px) {
  .sidebar { display: none; }
  .chart-grid { grid-template-columns: 1fr; }
  .kpi-grid { grid-template-columns: repeat(2,1fr); }
}
</style>
'''

# ===== HTML Body 模板 (5 Tab) =====
HTML_BODY = r'''
<div class="sidebar">
  <h2>&#128202; 新品周报 5.21-5.27</h2>
  <ul>
    <li><a href="javascript:void(0)" class="active" onclick="switchTab('tab1',this)">&#128200; 总盘概览</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab2',this)">&#128065; 低占比分析</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab3',this)">&#128176; 广告追踪</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab4',this)">&#128202; 四三累计</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab5',this)">&#128221; 汇报输出</a></li>
  </ul>
</div>
<div class="main">
  <div class="header">
    <h1>&#128202; 新品周报看板 &middot; 5.21 - 5.27 &#127793;</h1>
    <p>数据截至 2026-05-27 &nbsp;|&nbsp; 三部新品 &nbsp;|&nbsp; 含PLP/PLG广告追踪（自然周5.18-5.24）</p>
  </div>

  <!-- Tab1: 总盘概览 -->
  <div class="tab-content active" id="tab1">
    <div class="kpi-grid" id="t1-kpi"></div>
    <div class="chart-grid">
      <div class="chart-box"><h4>&#128200; 出单分布（4段）</h4><canvas id="chart-ord8"></canvas></div>
      <div class="chart-box"><h4>&#128200; 新品总市占比</h4><canvas id="chart-total-share"></canvas></div>
      <div class="chart-box"><h4>&#128200; 品线市占比对比</h4><canvas id="chart-cat-share"></canvas></div>
      <div class="chart-box"><h4>&#128101; 分析人市占比对比</h4><canvas id="chart-an-share"></canvas></div>
      <div class="chart-box"><h4>&#128200; 品线销量对比</h4><canvas id="chart-cat-sales"></canvas></div>
      <div class="chart-box"><h4>&#128101; 分析人销量对比</h4><canvas id="chart-an-sales"></canvas></div>
      <div class="chart-box"><h4>&#128200; 品线销售额对比</h4><canvas id="chart-cat-rev"></canvas></div>
      <div class="chart-box"><h4>&#128101; 分析人销售额对比</h4><canvas id="chart-an-rev"></canvas></div>
    </div>
    <div class="section"><h3>&#128200; 新品出单情况</h3><div id="t1-ord8"></div></div>
    <div class="section"><h3>&#128269; 多维度分析（含市占比）</h3>
      <div class="sub-module"><h4>品线维度</h4><div id="t1-cat-table"></div></div>
      <div class="sub-module"><h4>分析人维度</h4><div id="t1-an-table"></div></div>
    </div>
    <div class="section"><h3>&#128202; 拓展类型 & 及时率</h3>
      <div class="sub-module"><h4>拓展类型</h4><div id="t1-exp-table"></div></div>
      <div class="sub-module"><h4>分析及时率</h4><div id="t1-time-table"></div></div>
    </div>
  </div>

  <!-- Tab2: 低占比分析 -->
  <div class="tab-content" id="tab2">
    <div class="kpi-grid" id="t2-kpi"></div>
    <div class="section"><h3>&#128308; A. 有对手未出单新品</h3>
      <div class="sub-module"><h4>按分析人</h4><div id="t2-has-an"></div></div>
      <div class="sub-module"><h4>按品线</h4><div id="t2-has-cat"></div></div>
    </div>
    <div class="section"><h3>&#128993; B. 无对手未出单新品</h3>
      <div class="sub-module"><h4>按分析人</h4><div id="t2-no-an"></div></div>
      <div class="sub-module"><h4>按品线</h4><div id="t2-no-cat"></div></div>
    </div>
    <div class="section"><h3>&#128065; 低占比新品明细</h3><div class="filter-bar-sticky" id="t2-lowshare-filters"></div><div id="t2-lowshare-table"></div></div>
  </div>

  <!-- Tab3: 广告追踪 -->
  <div class="tab-content" id="tab3">
    <div class="section"><h3>&#128176; PLP 广告概览（自然周5.18-5.24）</h3>
      <div class="kpi-grid" id="t3-plp-kpi"></div>
      <div class="kpi-grid" id="t3-plp-core"></div>
    </div>
    <div class="section"><h3>&#128269; PLP 维度分析</h3>
      <div class="sub-module"><h4>按分析人</h4><div id="t3-plp-an"></div></div>
      <div class="sub-module"><h4>按品线</h4><div id="t3-plp-cat"></div></div>
      <div class="sub-module"><h4>按拓展类型</h4><div id="t3-plp-exp"></div></div>
    </div>
    <div class="section"><h3>&#128200; PLG 广告概览（自然周5.18-5.24）</h3>
      <div class="kpi-grid" id="t3-plg-kpi"></div>
      <div class="sub-module"><h4>PLG费率分布</h4><div id="t3-plg"></div></div>
      <div class="sub-module"><h4>PLG按分析人（含花费/ACOS/ACOAS）</h4><div id="t3-plg-an"></div></div>
    </div>
    <div class="section"><h3>&#128221; PLP 广告明细</h3><div id="t3-plp-detail"></div></div>
  </div>

  <!-- Tab4: 四三累计 -->
  <div class="tab-content" id="tab4">
    <div class="kpi-grid" id="t4-kpi"></div>
    <div class="filter-bar-sticky" id="t4-filters"></div>
    <div class="section" style="padding-top:12px">
      <div id="t4-table"></div>
    </div>
  </div>

  <!-- Tab5: 汇报输出 -->
  <div class="tab-content" id="tab5">
    <div class="kpi-grid" id="t5-kpi"></div>
    <div class="section"><h3>&#9888;&#65039; 风险预警</h3><div id="t5-risk"></div></div>
    <div class="section"><h3>&#128270; 本周期主要发现</h3><div id="t5-findings"></div></div>
    <div class="section"><h3>&#127919; 下周重点动作</h3><div id="t5-actions"></div></div>
    <div class="section"><h3>&#128203; 可复制周报文案</h3><div id="t5-report"></div></div>
  </div>
</div>
'''

print("HTML第一部分完成，开始组装JS代码...")
print("脚本将在写入完第二部分后继续...")
