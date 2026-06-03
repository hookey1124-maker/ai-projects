"""
Generate 5.7-5.13 HTML dashboard from source data
"""
import openpyxl
import json
from datetime import date, datetime
from collections import defaultdict, Counter

# ===== Config =====
SOURCE_FILE = 'c:/Users/Hardy/ai-projects/新品复盘/周报/新品检查周源数据和PLP数据.xlsx'
OUTPUT_FILE = 'c:/Users/Hardy/ai-projects/新品复盘/新品周报_5.7-5.13_可视化.html'

C = {
    'sale_no': 0, 'sku': 1, 'list_date': 2, 'first_order': 3,
    'analyst': 4, 'category': 5, 'expand_type': 6,
    'sales_curr': 16, 'sales_prev': 15,
    'rev_curr': 27, 'rev_prev': 26,
    'rival_curr': 38, 'rival_prev': 37,
    'share_curr': 48, 'share_prev': 47,
    'ord8_curr': 68, 'ord8_prev': 67,
    'freq7_curr': 78, 'freq7_prev': 77,
    'nfreq7_curr': 88, 'nfreq7_prev': 87,
    'mkt_curr': 99, 'mkt_prev': 98,
    'op_curr': 109, 'plp_curr': 114, 'plg_curr': 118,
}

PC = {
    'period': 0, 'campaign': 1, 'sku': 2, 'id': 3, 'store': 4,
    'plp_start': 5, 'list_date': 6, 'first_order': 7, 'analyst': 8,
    'category': 9, 'expand_type': 10,
    'impr': 11, 'click': 12, 'sold': 13, 'cost': 14, 'ad_rev': 15,
    'total_rev': 16, 'roas': 17, 'cvr': 18, 'ctr': 19,
    'cpc': 20, 'cpa': 21, 'acos': 22, 'acoas': 23, 'plg_enabled': 24,
}

ANALYSTS = ['俞东旭', '张潇', '朱培源', '王偲涵', '章鹏', '胡煜星']
CATEGORIES = ['车门系统', '车身外扩件', '挡泥板', '机盖及附件', '其他', '饰条', '牌照板支架', '未分类']
EXPAND_TYPES = ['原开品', '拓展品', '组合件']

def safe_float(v, default=0):
    try: return float(v) if v is not None else default
    except: return default

def get_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

def num(v, default=0):
    return safe_float(v, default)

def pct_str(a, b):
    if not b: return '0%'
    return f"{round(a/b*100, 1)}%"

def ratio_str(a, b):
    if not b: return '-'
    return f"{round((a-b)/abs(b)*100, 1)}%"

def calc_rate(a, b):
    if not b: return 0
    return round(a/b, 4)

def get_cat(r_data):
    c = str(r_data[C['category']] or '').strip()
    return c if c else '未分类'

def get_an(r_data):
    a = str(r_data[C['analyst']] or '').strip()
    return a if a else '未知'

def get_exp(r_data):
    e = str(r_data[C['expand_type']] or '').strip()
    return e if e else '其他'

# ===== Read source data =====
print("Reading source data...")
wb_src = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
ws_main = wb_src['四三数据累计']
ws_plp = wb_src['PLP明细']

rows_raw = []
for row in ws_main.iter_rows(min_row=2, values_only=True):
    if row[C['sku']]:
        rows_raw.append(list(row))

cutoff_curr = date(2026, 5, 13)
cutoff_prev = date(2026, 5, 6)
rows_curr = [r for r in rows_raw if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) <= cutoff_curr]
rows_prev = [r for r in rows_raw if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) <= cutoff_prev]

PLP_CURR = '5.4-5.10'
PLP_PREV = '4.30-5.6'

# PLP data
plp_curr_rows = []
plp_prev_rows = []
plp_plg_enabled = set()

for row in ws_plp.iter_rows(min_row=2, values_only=True):
    period = str(row[PC['period']] or '').strip()
    sku = str(row[PC['sku']] or '').strip()
    if not sku or sku.startswith('广告') or sku.startswith('总数据'):
        continue
    list_d = get_date(row[PC['list_date']])
    if period == PLP_CURR and list_d and list_d <= cutoff_curr:
        plp_curr_rows.append(list(row))
        if str(row[PC['plg_enabled']] or '').strip() == 'Y':
            plp_plg_enabled.add(sku)
    elif period == PLP_PREV and list_d and list_d <= cutoff_prev:
        plp_prev_rows.append(list(row))

print(f"Rows curr: {len(rows_curr)}, prev: {len(rows_prev)}, PLP curr: {len(plp_curr_rows)}")

# ===== Compute all stats =====
total_sku = len(rows_curr)
new_list = sum(1 for r in rows_curr if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) > cutoff_prev)
total_sales_curr = sum(num(r[C['sales_curr']]) for r in rows_curr)
total_sales_prev = sum(num(r[C['sales_prev']]) for r in rows_prev)
total_rev_curr = sum(num(r[C['rev_curr']]) for r in rows_curr)
total_rev_prev = sum(num(r[C['rev_prev']]) for r in rows_prev)
has_rival_curr = sum(1 for r in rows_curr if num(r[C['rival_curr']]) > 0)
has_rival_prev = sum(1 for r in rows_prev if num(r[C['rival_prev']]) > 0)

# Order stats
def count_ord8(rows_list, col):
    y = n = no = 0
    for r in rows_list:
        v = str(r[col] or '').strip()
        if v == 'Y': y += 1
        elif v == 'N': n += 1
        elif v == '未出单': no += 1
    return y, n, no

y_curr, n_curr, no_curr = count_ord8(rows_curr, C['ord8_curr'])
y_prev, n_prev, no_prev = count_ord8(rows_prev, C['ord8_prev'])
has_rival_y, has_rival_n, has_rival_no = count_ord8([r for r in rows_curr if num(r[C['rival_curr']]) > 0], C['ord8_curr'])
sale_rate_curr = pct_str(has_rival_y + has_rival_n, has_rival_curr)

# Timeliness
def calc_timeliness(rows_list, col_freq, col_nfreq):
    timely = no_8d = no_7d = 0
    for r in rows_list:
        freq = str(r[col_freq] or '').strip()
        nfreq = str(r[col_nfreq] or '').strip()
        if nfreq == '异常': no_8d += 1
        elif freq == '异常': no_7d += 1
        else: timely += 1
    return timely, no_8d, no_7d

timely_curr, no_8d_curr, no_7d_curr = calc_timeliness(rows_curr, C['freq7_curr'], C['nfreq7_curr'])
timely_prev, no_8d_prev, no_7d_prev = calc_timeliness(rows_prev, C['freq7_prev'], C['nfreq7_prev'])
total_timeliness = timely_curr + no_8d_curr + no_7d_curr
timely_rate = f"{round(timely_curr/total_timeliness*100, 1)}%" if total_timeliness else '0%'

# Low share
low_share = [r for r in rows_curr if num(r[C['share_curr']]) < 75 and num(r[C['rival_curr']]) > 0]

# PLP aggregation
sku_rev_curr = {}
for r in rows_curr:
    sku = str(r[C['sku']]).strip()
    if sku: sku_rev_curr[sku] = num(r[C['rev_curr']])

def agg_plp_rows(rows_list):
    by_sku = defaultdict(lambda: {'impr': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'total_rev': 0, 'campaigns': set()})
    for r in rows_list:
        sku = str(r[PC['sku']] or '').strip()
        camp = str(r[PC['campaign']] or '').strip()
        by_sku[sku]['impr'] += num(r[PC['impr']])
        by_sku[sku]['click'] += num(r[PC['click']])
        by_sku[sku]['sold'] += num(r[PC['sold']])
        by_sku[sku]['cost'] += num(r[PC['cost']])
        by_sku[sku]['ad_rev'] += num(r[PC['ad_rev']])
        by_sku[sku]['total_rev'] += num(r[PC['total_rev']])
        if camp: by_sku[sku]['campaigns'].add(camp)
    return by_sku

plp_curr_sku = agg_plp_rows(plp_curr_rows)
plp_prev_sku = agg_plp_rows(plp_prev_rows)

def plp_totals(by_sku):
    all_camps = set()
    t = {'imp': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'total_rev': 0}
    for sku, d in by_sku.items():
        t['imp'] += d['impr']; t['click'] += d['click']; t['sold'] += d['sold']
        t['cost'] += d['cost']; t['ad_rev'] += d['ad_rev']
        t['total_rev'] += sku_rev_curr.get(sku, 0)
        all_camps.update(d['campaigns'])
    t['campaigns'] = len(all_camps)
    t['links'] = len(by_sku)
    t['roas'] = calc_rate(t['ad_rev'], t['cost'])
    t['cvr'] = calc_rate(t['sold'], t['click'])
    t['ctr'] = calc_rate(t['click'], t['imp'])
    t['cpc'] = round(t['cost'] / t['click'], 2) if t['click'] else 0
    t['cpa'] = round(t['cost'] / t['sold'], 2) if t['sold'] else 0
    t['acos'] = calc_rate(t['cost'], t['ad_rev'])
    t['acoas'] = calc_rate(t['cost'], t['total_rev'])
    return t

plp_t = plp_totals(plp_curr_sku)
plp_pt = plp_totals(plp_prev_sku)

# PLP dimensions
sku_info = {}
for r in rows_curr:
    sku = str(r[C['sku']]).strip()
    if sku: sku_info[sku] = r

def plp_dim(by_sku, key_fn):
    groups = defaultdict(lambda: {'imp': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'total_rev': 0, 'campaigns': set(), 'skus': set()})
    for sku, d in by_sku.items():
        info = sku_info.get(sku)
        k = key_fn(sku, info) if info else key_fn(sku, None)
        g = groups[k]
        g['imp'] += d['impr']; g['click'] += d['click']; g['sold'] += d['sold']
        g['cost'] += d['cost']; g['ad_rev'] += d['ad_rev']
        g['total_rev'] += sku_rev_curr.get(sku, 0)
        g['campaigns'].update(d['campaigns'])
        g['skus'].add(sku)
    result = {}
    for k, g in groups.items():
        result[k] = {
            'campaigns': len(g['campaigns']), 'links': len(g['skus']),
            'imp': int(g['imp']), 'click': int(g['click']), 'sold': int(g['sold']),
            'cost': round(g['cost'], 2), 'ad_rev': round(g['ad_rev'], 2),
            'total_rev': round(g['total_rev'], 2),
            'roas': calc_rate(g['ad_rev'], g['cost']),
            'cvr': calc_rate(g['sold'], g['click']),
            'ctr': calc_rate(g['click'], g['imp']),
            'cpc': round(g['cost'] / g['click'], 2) if g['click'] else 0,
            'cpa': round(g['cost'] / g['sold'], 2) if g['sold'] else 0,
            'acos': calc_rate(g['cost'], g['ad_rev']),
            'acoas': calc_rate(g['cost'], g['total_rev']),
        }
    return result

plp_by_ana = plp_dim(plp_curr_sku, lambda sku, info: get_an(info) if info else '未知')
plp_by_cat = plp_dim(plp_curr_sku, lambda sku, info: get_cat(info) if info else '未分类')
plp_by_exp = plp_dim(plp_curr_sku, lambda sku, info: get_exp(info) if info else '其他')

# PLG classification
def get_ad_class(r_data):
    plp_on = str(r_data[C['plp_curr']] or '').strip().upper() == 'Y'
    plg_on = num(r_data[C['plg_curr']]) > 0
    is_unsold = str(r_data[C['ord8_curr']] or '').strip() == '未出单'
    if plp_on and plg_on: return 'PLP+PLG同开'
    elif plg_on and not plp_on:
        return '单PLG且未出单' if is_unsold else '单PLG'
    elif plp_on and not plg_on: return '单PLP'
    return '无广告'

plg_cats = {'PLP+PLG同开': 0, '单PLG': 0, '单PLP': 0, '单PLG且未出单': 0, '无广告': 0}
plg_by_ana_counts = defaultdict(lambda: {'PLP+PLG同开': 0, '单PLG': 0, '单PLP': 0, '单PLG且未出单': 0, '无广告': 0})
plg_by_cat_counts = defaultdict(lambda: {'PLP+PLG同开': 0, '单PLG': 0, '单PLP': 0, '单PLG且未出单': 0, '无广告': 0})

for r in rows_curr:
    adc = get_ad_class(r)
    plg_cats[adc] += 1
    an = get_an(r)
    cat = get_cat(r)
    plg_by_ana_counts[an][adc] += 1
    plg_by_cat_counts[cat][adc] += 1

# Unsold analysis
has_rival_no_curr = [r for r in rows_curr if str(r[C['ord8_curr']] or '').strip() == '未出单' and num(r[C['rival_curr']]) > 0]
no_rival_no_curr = [r for r in rows_curr if str(r[C['ord8_curr']] or '').strip() == '未出单' and num(r[C['rival_curr']]) == 0]

# Category data
cat_data_curr = defaultdict(lambda: {'sku': 0, 'new': 0, 'sales': 0, 'rev': 0, 'has_rival': 0, 'y': 0, 'n': 0})
cat_data_prev = defaultdict(lambda: {'sku': 0, 'sales': 0, 'rev': 0, 'has_rival': 0})
for r in rows_curr:
    cat = get_cat(r)
    ld = get_date(r[C['list_date']])
    cat_data_curr[cat]['sku'] += 1
    if ld and ld > cutoff_prev: cat_data_curr[cat]['new'] += 1
    cat_data_curr[cat]['sales'] += num(r[C['sales_curr']])
    cat_data_curr[cat]['rev'] += num(r[C['rev_curr']])
    if num(r[C['rival_curr']]) > 0:
        cat_data_curr[cat]['has_rival'] += 1
        v = str(r[C['ord8_curr']] or '').strip()
        if v == 'Y': cat_data_curr[cat]['y'] += 1
        elif v == 'N': cat_data_curr[cat]['n'] += 1
for r in rows_prev:
    cat = get_cat(r)
    cat_data_prev[cat]['sales'] += num(r[C['sales_prev']])
    cat_data_prev[cat]['rev'] += num(r[C['rev_prev']])
    if num(r[C['rival_prev']]) > 0: cat_data_prev[cat]['has_rival'] += 1

all_cats_data = sorted(set(list(cat_data_curr.keys()) + list(cat_data_prev.keys())))
cats_ordered = [c for c in CATEGORIES if c in all_cats_data] + [c for c in all_cats_data if c not in CATEGORIES]

# Analyst data
an_data_curr = defaultdict(lambda: {'sku': 0, 'new': 0, 'sales': 0, 'rev': 0, 'has_rival': 0, 'y': 0, 'n': 0, 'timely': 0, 'no_8d': 0, 'no_7d': 0})
an_data_prev = defaultdict(lambda: {'sku': 0, 'sales': 0, 'rev': 0, 'has_rival': 0, 'timely': 0, 'no_8d': 0, 'no_7d': 0})
for r in rows_curr:
    an = get_an(r)
    ld = get_date(r[C['list_date']])
    an_data_curr[an]['sku'] += 1
    if ld and ld > cutoff_prev: an_data_curr[an]['new'] += 1
    an_data_curr[an]['sales'] += num(r[C['sales_curr']])
    an_data_curr[an]['rev'] += num(r[C['rev_curr']])
    if num(r[C['rival_curr']]) > 0:
        an_data_curr[an]['has_rival'] += 1
        v = str(r[C['ord8_curr']] or '').strip()
        if v == 'Y': an_data_curr[an]['y'] += 1
        elif v == 'N': an_data_curr[an]['n'] += 1
    freq = str(r[C['freq7_curr']] or '').strip()
    nfreq = str(r[C['nfreq7_curr']] or '').strip()
    if nfreq == '异常': an_data_curr[an]['no_8d'] += 1
    elif freq == '异常': an_data_curr[an]['no_7d'] += 1
    else: an_data_curr[an]['timely'] += 1
for r in rows_prev:
    an = get_an(r)
    an_data_prev[an]['sales'] += num(r[C['sales_prev']])
    an_data_prev[an]['rev'] += num(r[C['rev_prev']])
    if num(r[C['rival_prev']]) > 0: an_data_prev[an]['has_rival'] += 1
    freq = str(r[C['freq7_prev']] or '').strip()
    nfreq = str(r[C['nfreq7_prev']] or '').strip()
    if nfreq == '异常': an_data_prev[an]['no_8d'] += 1
    elif freq == '异常': an_data_prev[an]['no_7d'] += 1
    else: an_data_prev[an]['timely'] += 1

all_ans = sorted(set(list(an_data_curr.keys()) + list(an_data_prev.keys())))
ans_ordered = [a for a in ANALYSTS if a in all_ans] + [a for a in all_ans if a not in ANALYSTS]

# Expand type data
exp_data_curr = defaultdict(lambda: {'sku': 0, 'new': 0, 'sales': 0, 'rev': 0, 'has_rival': 0, 'y': 0, 'n': 0})
exp_data_prev = defaultdict(lambda: {'sku': 0, 'sales': 0, 'rev': 0, 'has_rival': 0, 'y': 0, 'n': 0})
for r in rows_curr:
    exp = get_exp(r)
    ld = get_date(r[C['list_date']])
    exp_data_curr[exp]['sku'] += 1
    if ld and ld > cutoff_prev: exp_data_curr[exp]['new'] += 1
    exp_data_curr[exp]['sales'] += num(r[C['sales_curr']])
    exp_data_curr[exp]['rev'] += num(r[C['rev_curr']])
    if num(r[C['rival_curr']]) > 0:
        exp_data_curr[exp]['has_rival'] += 1
        v = str(r[C['ord8_curr']] or '').strip()
        if v == 'Y': exp_data_curr[exp]['y'] += 1
        elif v == 'N': exp_data_curr[exp]['n'] += 1
for r in rows_prev:
    exp = get_exp(r)
    exp_data_prev[exp]['sales'] += num(r[C['sales_prev']])
    exp_data_prev[exp]['rev'] += num(r[C['rev_prev']])
    if num(r[C['rival_prev']]) > 0:
        exp_data_prev[exp]['has_rival'] += 1
        v = str(r[C['ord8_prev']] or '').strip()
        if v == 'Y': exp_data_prev[exp]['y'] += 1
        elif v == 'N': exp_data_prev[exp]['n'] += 1
all_exps = sorted(set(list(exp_data_curr.keys()) + list(exp_data_prev.keys())))
exps_ordered = [e for e in EXPAND_TYPES if e in all_exps] + [e for e in all_exps if e not in EXPAND_TYPES]

# Market status counts for cum43Stats
mkt_curr_counter = Counter(str(r[C['mkt_curr']] or '未知') for r in rows_curr)
ord8_counter = Counter(str(r[C['ord8_curr']] or '') for r in rows_curr)
cum43_stats = {
    'total': total_sku,
    'yCount': y_curr,
    'nCount': n_curr,
    'unCount': no_curr,
    'normalCount': mkt_curr_counter.get('正常', 0),
    'competitiveCount': mkt_curr_counter.get('竞争无优势', 0),
    'noMarketCount': mkt_curr_counter.get('无市场', 0),
}

print(f"Total SKU: {total_sku}, Sales: {total_sales_curr}, Rev: ${total_rev_curr:,.2f}")
print(f"Has rival: {has_rival_curr}, Order rate: {sale_rate_curr}")
print(f"Timely rate: {timely_rate}, Low share: {len(low_share)}")
print(f"PLG: {plg_cats}")

# ===== Build JSON data for HTML embedding =====
def j(v):
    return json.dumps(v, ensure_ascii=False, default=str)

# Build cum43Data
cum43_data = []
for r in rows_curr:
    sku = str(r[C['sku']]).strip()
    ld = get_date(r[C['list_date']])
    fo = get_date(r[C['first_order']])
    cum43_data.append({
        'SKU': sku,
        '实际上架日期': str(ld) if ld else '',
        '首次出单日期': str(fo) if fo else '未出单',
        '4月分析人': get_an(r),
        '品类': get_cat(r),
        '产品拓展': get_exp(r),
        '5.7-5.13销量': num(r[C['sales_curr']]),
        '4.30-5.6销量': num(r[C['sales_prev']]),
        '5.7-5.13销售额': num(r[C['rev_curr']]),
        '4.30-5.6销售额': num(r[C['rev_prev']]),
        '5.13对手销量': num(r[C['rival_curr']]),
        '5.6对手销量': num(r[C['rival_prev']]),
        '5.13市占比': num(r[C['share_curr']]),
        '5.6市占比': num(r[C['share_prev']]),
        '5.13市场状态': str(r[C['mkt_curr']] or ''),
        '5.13 8日出单情况': str(r[C['ord8_curr']] or ''),
        '5.13 7日频次标签': str(r[C['freq7_curr']] or ''),
        '5.13 7日新品频次标签': str(r[C['nfreq7_curr']] or ''),
        '开启PLP': 'Y' if str(r[C['plp_curr']] or '').strip().upper() == 'Y' else 'N',
        'PLG费率': num(r[C['plg_curr']]),
        '广告分类': get_ad_class(r),
    })

# Build lowShareData
low_share_data = []
for r in sorted(low_share, key=lambda x: num(x[C['share_curr']])):
    sku = str(r[C['sku']]).strip()
    cost = plp_curr_sku.get(sku, {}).get('cost', 0)
    total_rev = sku_rev_curr.get(sku, 0)
    acoas_val = f"{cost/total_rev*100:.2f}%" if total_rev > 0 else 'N/A'
    low_share_data.append({
        'salesCode': str(r[C['sale_no']] or ''),
        'sku': sku,
        'launchDate': str(get_date(r[C['list_date']])) if get_date(r[C['list_date']]) else '',
        'analyst': get_an(r),
        'category': get_cat(r),
        'expandType': get_exp(r),
        'curSalesQty': num(r[C['sales_curr']]),
        'salesQtyChange': ratio_str(num(r[C['sales_curr']]), num(r[C['sales_prev']])),
        'curRevenue': num(r[C['rev_curr']]),
        'revenueChange': ratio_str(num(r[C['rev_curr']]), num(r[C['rev_prev']])),
        'prevCompetitorQty': num(r[C['rival_prev']]),
        'curCompetitorQty': num(r[C['rival_curr']]),
        'competitorQtyChange': ratio_str(num(r[C['rival_curr']]), num(r[C['rival_prev']])),
        'curMarketShare': f"{num(r[C['share_curr']])}%",
        'prevMarketStatus': str(r[C['mkt_prev']] or ''),
        'curMarketStatus': str(r[C['mkt_curr']] or ''),
        'cur8dStatus': str(r[C['ord8_curr']] or ''),
        'plpEnabled': 'Y' if str(r[C['plp_curr']] or '').strip().upper() == 'Y' else 'N',
        'plgFee': f"{num(r[C['plg_curr']])}%",
        'acoas': acoas_val,
    })

# Build expandTypeData
expand_type_data = []
for exp in exps_ordered:
    d_curr = exp_data_curr.get(exp, {})
    d_prev = exp_data_prev.get(exp, {})
    cur_rate = (d_curr.get('y',0)+d_curr.get('n',0))/d_curr.get('has_rival',1)*100 if d_curr.get('has_rival') else 0
    prev_rate = (d_prev.get('y',0)+d_prev.get('n',0))/d_prev.get('has_rival',1)*100 if d_prev.get('has_rival') else 0
    expand_type_data.append({
        'expandType': exp,
        'curSku': d_curr.get('sku', 0),
        'prevSku': len([r for r in rows_prev if get_exp(r) == exp]),
        'curSalesCount': d_curr.get('y',0)+d_curr.get('n',0),
        'curSalesRate': f"{round(cur_rate,1)}%",
        'prevSalesCount': d_prev.get('y',0)+d_prev.get('n',0),
        'prevSalesRate': f"{round(prev_rate,1)}%",
        'salesRateChange': ratio_str(cur_rate, prev_rate),
        'curSalesQty': d_curr.get('sales', 0),
        'prevSalesQty': d_prev.get('sales', 0),
        'salesQtyChange': ratio_str(d_curr.get('sales', 0), d_prev.get('sales', 0)),
        'curRevenue': d_curr.get('rev', 0),
        'prevRevenue': d_prev.get('rev', 0),
        'revenueChange': ratio_str(d_curr.get('rev', 0), d_prev.get('rev', 0)),
    })

# Build timelinessData
total_prev_timeliness = timely_prev + no_8d_prev + no_7d_prev
prev_timely_rate = f"{round(timely_prev/total_prev_timeliness*100, 1)}%" if total_prev_timeliness else '0%'
timeliness_data = {'analysts': [], 'total': {}}
for an in ans_ordered:
    d = an_data_curr.get(an, {})
    d_prev = an_data_prev.get(an, {})
    total_an = d.get('timely', 0) + d.get('no_8d', 0) + d.get('no_7d', 0)
    total_an_prev = d_prev.get('timely', 0) + d_prev.get('no_8d', 0) + d_prev.get('no_7d', 0)
    rate_an = f"{round(d.get('timely', 0)/total_an*100, 1)}%" if total_an else '0%'
    rate_an_prev = f"{round(d_prev.get('timely', 0)/total_an_prev*100, 1)}%" if total_an_prev else '0%'
    timely_change = ratio_str(float(rate_an.replace('%','')), float(rate_an_prev.replace('%',''))) if rate_an != '0%' and rate_an_prev != '0%' else '-'
    timeliness_data['analysts'].append({
        'analyst': an, 'curSku': total_an,
        'timelyCount': d.get('timely', 0),
        'noAnalysis8dCount': d.get('no_8d', 0),
        'noAnalysis7dCount': d.get('no_7d', 0),
        'timelyRate': rate_an,
        'prevTimelyCount': d_prev.get('timely', 0),
        'prevTimelyRate': rate_an_prev,
        'timelyChange': timely_change,
    })
timeliness_data['total'] = {
    'analyst': '合计', 'curSku': total_timeliness,
    'timelyCount': timely_curr, 'noAnalysis8dCount': no_8d_curr,
    'noAnalysis7dCount': no_7d_curr, 'timelyRate': timely_rate,
    'prevTimelyCount': timely_prev, 'prevTimelyRate': prev_timely_rate,
    'timelyChange': ratio_str(float(timely_rate.replace('%','')), float(prev_timely_rate.replace('%',''))) if timely_rate != '0%' and prev_timely_rate != '0%' else '-',
}

# Build unsold analysis
mkt_order_has = ['竞争无优势', '无市场', '站内无价格优势', '站外出单', '正常', '#N/A', '未知']
mkt_order_no = ['无市场', '未知', '竞争无优势', '#N/A', '其他']

mkt_has_counter = Counter(str(r[C['mkt_curr']] or '未知') for r in has_rival_no_curr)
mkt_no_counter = Counter(str(r[C['mkt_curr']] or '未知') for r in no_rival_no_curr)

has_comp_unsold = {
    'total': len(has_rival_no_curr),
    'reasons': [{'reason': m, 'count': mkt_has_counter.get(m, 0)} for m in mkt_order_has],
    'byAnalyst': [],
    'byCategory': [],
}
for an in ans_ordered:
    entry = {'analyst': an}
    for m in mkt_order_has: entry[m] = 0
    entry['total'] = 0
    has_comp_unsold['byAnalyst'].append(entry)
for r in has_rival_no_curr:
    an = get_an(r); m = str(r[C['mkt_curr']] or '未知')
    for e in has_comp_unsold['byAnalyst']:
        if e['analyst'] == an:
            if m in e: e[m] += 1
            e['total'] += 1

unsold_no_comp = {
    'total': len(no_rival_no_curr),
    'byAnalyst': [],
    'byCategory': [],
}
for an in ans_ordered:
    entry = {'analyst': an}
    for m in mkt_order_no: entry[m] = 0
    entry['total'] = 0
    unsold_no_comp['byAnalyst'].append(entry)
for r in no_rival_no_curr:
    an = get_an(r); m = str(r[C['mkt_curr']] or '未知')
    for e in unsold_no_comp['byAnalyst']:
        if e['analyst'] == an:
            if m in e: e[m] += 1
            e['total'] += 1

# Build PLP data
plp_total = {
    'campaignCount': plp_t['campaigns'], 'linkCount': plp_t['links'],
    'impression': plp_t['imp'], 'click': plp_t['click'],
    'sold': plp_t['sold'], 'cost': round(plp_t['cost'], 2),
    'revenue': round(plp_t['ad_rev'], 2),
    'totalRevenue': round(plp_t['total_rev'], 2),
    'roas': f"{plp_t['roas']:.2f}", 'cvr': f"{plp_t['cvr']*100:.2f}%",
    'ctr': f"{plp_t['ctr']*100:.2f}%",
    'cpc': f"${plp_t['cpc']:.2f}", 'cpa': f"${plp_t['cpa']:.2f}",
    'acos': f"{plp_t['acos']*100:.2f}%", 'acoas': f"{plp_t['acoas']*100:.2f}%",
}
plp_prev_total = {
    'campaignCount': plp_pt['campaigns'], 'linkCount': plp_pt['links'],
    'impression': plp_pt['imp'], 'click': plp_pt['click'],
    'sold': plp_pt['sold'], 'cost': round(plp_pt['cost'], 2),
    'revenue': round(plp_pt['ad_rev'], 2),
    'totalRevenue': round(plp_pt['total_rev'], 2),
    'roas': f"{plp_pt['roas']:.2f}", 'cvr': f"{plp_pt['cvr']*100:.2f}%",
    'ctr': f"{plp_pt['ctr']*100:.2f}%",
    'cpc': f"${plp_pt['cpc']:.2f}", 'cpa': f"${plp_pt['cpa']:.2f}",
    'acos': f"{plp_pt['acos']*100:.2f}%", 'acoas': f"{plp_pt['acoas']*100:.2f}%",
}

# PLP dimension arrays
plp_analysts = []
for an in ans_ordered:
    d = plp_by_ana.get(an, {})
    if d:
        plp_analysts.append({
            'analyst': an, 'campaigns': d.get('campaigns', 0), 'links': d.get('links', 0),
            'imp': d.get('imp', 0), 'click': d.get('click', 0), 'sold': d.get('sold', 0),
            'cost': d.get('cost', 0), 'ad_rev': d.get('ad_rev', 0), 'total_rev': d.get('total_rev', 0),
            'roas': f"{d.get('roas', 0):.2f}", 'cvr': f"{d.get('cvr', 0)*100:.2f}%",
            'ctr': f"{d.get('ctr', 0)*100:.2f}%",
            'cpc': f"${d.get('cpc', 0):.2f}", 'cpa': f"${d.get('cpa', 0):.2f}",
            'acos': f"{d.get('acos', 0)*100:.2f}%", 'acoas': f"{d.get('acoas', 0)*100:.2f}%",
        })

plp_categories = []
for cat in cats_ordered:
    d = plp_by_cat.get(cat, {})
    if d:
        plp_categories.append({
            'category': cat, 'campaigns': d.get('campaigns', 0), 'links': d.get('links', 0),
            'imp': d.get('imp', 0), 'click': d.get('click', 0), 'sold': d.get('sold', 0),
            'cost': d.get('cost', 0), 'ad_rev': d.get('ad_rev', 0), 'total_rev': d.get('total_rev', 0),
            'roas': f"{d.get('roas', 0):.2f}", 'cvr': f"{d.get('cvr', 0)*100:.2f}%",
            'ctr': f"{d.get('ctr', 0)*100:.2f}%",
            'cpc': f"${d.get('cpc', 0):.2f}", 'cpa': f"${d.get('cpa', 0):.2f}",
            'acos': f"{d.get('acos', 0)*100:.2f}%", 'acoas': f"{d.get('acoas', 0)*100:.2f}%",
        })

plp_expand_types = []
for exp in exps_ordered:
    d = plp_by_exp.get(exp, {})
    if d:
        plp_expand_types.append({
            'expandType': exp, 'campaigns': d.get('campaigns', 0), 'links': d.get('links', 0),
            'imp': d.get('imp', 0), 'click': d.get('click', 0), 'sold': d.get('sold', 0),
            'cost': d.get('cost', 0), 'ad_rev': d.get('ad_rev', 0), 'total_rev': d.get('total_rev', 0),
            'roas': f"{d.get('roas', 0):.2f}", 'cvr': f"{d.get('cvr', 0)*100:.2f}%",
            'ctr': f"{d.get('ctr', 0)*100:.2f}%",
            'cpc': f"${d.get('cpc', 0):.2f}", 'cpa': f"${d.get('cpa', 0):.2f}",
            'acos': f"{d.get('acos', 0)*100:.2f}%", 'acoas': f"{d.get('acoas', 0)*100:.2f}%",
        })

# PLG stats
plg_stats = {
    'plpAndPlgBothCount': plg_cats['PLP+PLG同开'],
    'plgOnlyCount': plg_cats['单PLG'],
    'plpOnlyCount': plg_cats['单PLP'],
    'noAdCount': plg_cats['无广告'],
    'plgUnsoldCount': plg_cats['单PLG且未出单'],
    'plpEnabledCount': sum(1 for r in rows_curr if str(r[C['plp_curr']] or '').strip().upper() == 'Y'),
    'plpDisabledCount': sum(1 for r in rows_curr if str(r[C['plp_curr']] or '').strip().upper() != 'Y'),
    'totalNewProducts': total_sku,
    'byAnalyst': [],
}
for an in ans_ordered:
    d = plg_by_ana_counts.get(an, {})
    plg_stats['byAnalyst'].append({
        'analyst': an, 'total': sum(d.values()),
        'plpAndPlgBoth': d.get('PLP+PLG同开', 0),
        'plgOnly': d.get('单PLG', 0),
        'plpOnly': d.get('单PLP', 0),
        'noAd': d.get('无广告', 0),
        'plgUnsold': d.get('单PLG且未出单', 0),
    })

# PLG records for filter matching
plg_records = []
for r in rows_curr:
    sku = str(r[C['sku']]).strip()
    plg_records.append({
        'sku': sku,
        'plpEnabled': 'Y' if str(r[C['plp_curr']] or '').strip().upper() == 'Y' else 'N',
        'plgFee': f"{num(r[C['plg_curr']])}%",
        'adClass': get_ad_class(r),
    })

# Category/analyst revenue data
category_revenue_data = []
for cat in cats_ordered:
    d_curr = cat_data_curr.get(cat, {})
    d_prev = cat_data_prev.get(cat, {})
    category_revenue_data.append({
        'category': cat, 'curSku': d_curr.get('sku', 0),
        'curNewSku': d_curr.get('new', 0),
        'curSalesQty': d_curr.get('sales', 0),
        'prevSalesQty': d_prev.get('sales', 0),
        'salesQtyChange': ratio_str(d_curr.get('sales', 0), d_prev.get('sales', 0)),
        'curRevenue': d_curr.get('rev', 0),
        'prevRevenue': d_prev.get('rev', 0),
        'revenueChange': ratio_str(d_curr.get('rev', 0), d_prev.get('rev', 0)),
    })

analyst_revenue_data = []
for an in ans_ordered:
    d_curr = an_data_curr.get(an, {})
    d_prev = an_data_prev.get(an, {})
    analyst_revenue_data.append({
        'analyst': an, 'curSku': d_curr.get('sku', 0),
        'curNewSku': d_curr.get('new', 0),
        'curSalesQty': d_curr.get('sales', 0),
        'prevSalesQty': d_prev.get('sales', 0),
        'salesQtyChange': ratio_str(d_curr.get('sales', 0), d_prev.get('sales', 0)),
        'curRevenue': d_curr.get('rev', 0),
        'prevRevenue': d_prev.get('rev', 0),
        'revenueChange': ratio_str(d_curr.get('rev', 0), d_prev.get('rev', 0)),
    })

# prevWeekKpi
prev_week_kpi = {
    'prevTotalSku': len(rows_prev),
    'prevTotalSalesQty': total_sales_prev,
    'prevTotalRevenue': round(total_rev_prev, 2),
    'salesQtyChange': ratio_str(total_sales_curr, total_sales_prev),
    'revenueChange': ratio_str(total_rev_curr, total_rev_prev),
    'skuChange': ratio_str(total_sku, len(rows_prev)),
    'prevTimelyCount': timely_prev,
    'prevTimelyRate': f"{round(timely_prev/(timely_prev+no_8d_prev+no_7d_prev)*100,1)}%" if (timely_prev+no_8d_prev+no_7d_prev) else '0%',
    'prevTimelySku': timely_prev+no_8d_prev+no_7d_prev,
    'timelyChange': ratio_str(float(timely_rate.replace('%','')), float(f"{round(timely_prev/(timely_prev+no_8d_prev+no_7d_prev)*100,1)}" if (timely_prev+no_8d_prev+no_7d_prev) else '0')),
    'prevSoldCount': has_rival_y + has_rival_n,
    'prevSoldRate': pct_str(has_rival_y+has_rival_n, has_rival_curr),
}

# prevWeekCategories / prevWeekAnalysts
prev_week_categories = {}
for cat in cats_ordered:
    d = cat_data_prev.get(cat, {'sku': 0, 'sales': 0, 'rev': 0, 'has_rival': 0})
    prev_week_categories[cat] = {'sku': d.get('sku', 0), 'sales': d.get('sales', 0), 'rev': d.get('rev', 0)}

prev_week_analysts = {}
for an in ans_ordered:
    d = an_data_prev.get(an, {'sku': 0, 'sales': 0, 'rev': 0, 'has_rival': 0})
    prev_week_analysts[an] = {'sku': d.get('sku', 0), 'sales': d.get('sales', 0), 'rev': d.get('rev', 0)}

# PLP detail data
plp_detail_data = []
for row in sorted(plp_curr_rows, key=lambda x: (-num(x[PC['sold']]), -num(x[PC['click']]), -num(x[PC['impr']]))):
    sku = str(row[PC['sku']] or '').strip()
    info = sku_info.get(sku, {})
    cost = num(row[PC['cost']]); ad_rev = num(row[PC['ad_rev']])
    total_r = num(row[PC['total_rev']]); impr = num(row[PC['impr']])
    click = num(row[PC['click']]); sold = num(row[PC['sold']])
    plp_detail_data.append({
        'period': str(row[PC['period']] or ''),
        'campaign': str(row[PC['campaign']] or ''),
        'sku': sku, 'id': str(row[PC['id']] or ''),
        'store': str(row[PC['store']] or ''),
        'plpStartDate': str(row[PC['plp_start']] or ''),
        'actualListDate': str(row[PC['list_date']] or ''),
        'firstOrderDate': str(row[PC['first_order']] or ''),
        'analyst': get_an(info) if info else str(row[PC['analyst']] or ''),
        'category': get_cat(info) if info else str(row[PC['category']] or '未分类'),
        'productExpansion': get_exp(info) if info else str(row[PC['expand_type']] or ''),
        'impressions': int(impr), 'clicks': int(click), 'salesQty': int(sold),
        'spend': round(cost, 2), 'adRevenue': round(ad_rev, 2),
        'totalRevenue': round(total_r, 2),
        'roas': calc_rate(ad_rev, cost),
        'cvr': calc_rate(sold, click),
        'ctr': calc_rate(click, impr),
        'cpc': round(cost/click, 2) if click else 0,
        'cpa': round(cost/sold, 2) if sold else 0,
        'acos': calc_rate(cost, ad_rev),
        'acoas': calc_rate(cost, sku_rev_curr.get(sku, 0)),
        'plgEnabled': str(row[PC['plg_enabled']] or ''),
        'adClass': get_ad_class(info) if info else '无广告',
    })

# PLP summary data (for dimension tables)
plp_summary_data = []
for d in plp_analysts:
    plp_summary_data.append({'dimType': 'analyst', 'name': d['analyst'], **d})
for d in plp_categories:
    plp_summary_data.append({'dimType': 'category', 'name': d['category'], **d})
for d in plp_expand_types:
    plp_summary_data.append({'dimType': 'expandType', 'name': d['expandType'], **d})

# ===== Build data blocks dict =====
data_blocks = {
    'cum43Data': cum43_data,
    'cum43Stats': cum43_stats,
    'lowShareData': low_share_data,
    'expandTypeData': expand_type_data,
    'timelinessData': timeliness_data,
    'hasCompetitorUnsold': has_comp_unsold,
    'plpTotal': plp_total,
    'plpPrevTotal': plp_prev_total,
    'plpCategories': plp_categories,
    'plpExpandTypes': plp_expand_types,
    'plpAnalysts': plp_analysts,
    'unsoldNoCompetitor': unsold_no_comp,
    'prevWeekKpi': prev_week_kpi,
    'plgStats': plg_stats,
    'prevWeekCategories': prev_week_categories,
    'prevWeekAnalysts': prev_week_analysts,
    'plgRecords': plg_records,
    'categoryRevenueData': category_revenue_data,
    'analystRevenueData': analyst_revenue_data,
    'plpSummaryData': plp_summary_data,
    'plpDetailData': plp_detail_data,
}

# Verify all 21 keys
assert len(data_blocks) == 21, f"Expected 21 data blocks, got {len(data_blocks)}"
print(f"All 21 data blocks built successfully.")

# Save intermediate JSON (useful for debugging)
json_path = 'c:/Users/Hardy/ai-projects/新品复盘/sheets_5_7_5_13.json'
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(data_blocks, f, ensure_ascii=False, default=str)
print(f"JSON saved to {json_path}")

# ===== Generate HTML =====
print("Generating HTML...")

# Serialize each data block as JS const
js_lines = []
for k, v in data_blocks.items():
    js_lines.append(f"const {k} = {j(v)};")
js_data = "\n".join(js_lines)

html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>新品周报 · 5.7-5.13</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: "Microsoft YaHei", "PingFang SC", Arial, sans-serif; background: #f0f2f5; color: #1a1a2e; }}
.header {{ background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%); background-size: 200% 200%; animation: gradientShift 8s ease infinite; color: white; padding: 28px 40px; }}
.header h1 {{ font-size: 26px; font-weight: 700; letter-spacing: 2px; }}
.header .subtitle {{ font-size: 13px; opacity: 0.75; margin-top: 6px; }}
.container {{ display: flex; min-height: calc(100vh - 80px); }}
.sidebar {{ width: 230px; background: #fff; border-right: 1px solid #e8e8e8; padding: 20px 0; position: sticky; top: 0; height: 100vh; overflow-y: auto; flex-shrink: 0; }}
.sidebar ul {{ list-style: none; }}
.sidebar li a {{ display: block; padding: 10px 20px; font-size: 13px; color: #555; text-decoration: none; border-left: 3px solid transparent; transition: all 0.2s; }}
.sidebar li a:hover, .sidebar li a.active {{ background: #f0f6ff; color: #0f3460; border-left-color: #0f3460; font-weight: 600; }}
.main-content {{ flex: 1; padding: 24px; overflow: auto; }}
.tab-content {{ display: none; }}
.tab-content.active {{ display: block; }}
.kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 200px)); gap: 14px; margin-bottom: 24px; }}
.kpi-card {{ background: white; border-radius: 10px; padding: 18px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); text-align: center; transition: transform 0.2s, box-shadow 0.2s; animation: fadeInUp 0.5s ease-out forwards; }}
.kpi-card:hover {{ transform: translateY(-2px); box-shadow: 0 4px 16px rgba(0,0,0,0.12); }}
.kpi-card .val {{ font-size: 26px; font-weight: 700; color: #0f3460; }}
.kpi-card .label {{ font-size: 12px; color: #888; margin-top: 6px; }}
.kpi-card .hb {{ font-size: 11px; margin-top: 4px; font-weight: 600; }}
.kpi-card.green .val {{ color: #08845a; }}
.kpi-card.orange .val {{ color: #e07b24; }}
.kpi-card.red .val {{ color: #c0392b; }}
.kpi-card.purple .val {{ color: #8e44ad; }}
/* === v3.2 Animations === */
@keyframes fadeInUp {{ from {{ opacity: 0; transform: translateY(24px); }} to {{ opacity: 1; transform: translateY(0); }} }}
@keyframes fadeIn {{ from {{ opacity: 0; }} to {{ opacity: 1; }} }}
@keyframes pulse {{ 0%,100% {{ transform: scale(1); }} 50% {{ transform: scale(1.04); }} }}
@keyframes shimmer {{ 0% {{ background-position: -200px 0; }} 100% {{ background-position: calc(200px + 100%) 0; }} }}
@keyframes slideInLeft {{ from {{ opacity: 0; transform: translateX(-30px); }} to {{ opacity: 1; transform: translateX(0); }} }}
@keyframes slideInRight {{ from {{ opacity: 0; transform: translateX(30px); }} to {{ opacity: 1; transform: translateX(0); }} }}
@keyframes bounceIn {{ 0% {{ opacity: 0; transform: scale(0.3); }} 50% {{ opacity: 1; transform: scale(1.05); }} 70% {{ transform: scale(0.9); }} 100% {{ transform: scale(1); }} }}
@keyframes gradientShift {{ 0% {{ background-position: 0% 50%; }} 50% {{ background-position: 100% 50%; }} 100% {{ background-position: 0% 50%; }} }}
.anim-fade-in-up {{ animation: fadeInUp 0.6s ease-out forwards; }}
.anim-fade-in {{ animation: fadeIn 0.5s ease-out forwards; }}
.anim-slide-left {{ animation: slideInLeft 0.5s ease-out forwards; }}
.anim-slide-right {{ animation: slideInRight 0.5s ease-out forwards; }}
.anim-bounce {{ animation: bounceIn 0.7s ease-out forwards; }}
.kpi-card.animate-pulse {{ animation: pulse 2s ease-in-out infinite; }}
.section {{ animation: fadeInUp 0.55s ease-out forwards; background: white; border-radius: 10px; padding: 20px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
.section h3 {{ font-size: 16px; font-weight: 700; color: #0f3460; padding-bottom: 12px; border-bottom: 2px solid #e8f0fe; margin-bottom: 16px; }}
.sub-module {{ margin-bottom: 20px; }}
.sub-module h4 {{ font-size: 13px; font-weight: 600; color: #444; margin-bottom: 10px; padding: 6px 12px; background: #f5f7ff; border-radius: 4px; border-left: 3px solid #0f3460; }}
.chart-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(400px, 1fr)); gap: 20px; margin-bottom: 20px; }}
.chart-card {{ background: white; border-radius: 10px; padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); animation: fadeInUp 0.6s ease-out forwards; transition: transform 0.25s, box-shadow 0.25s; }}
.chart-card:hover {{ transform: translateY(-3px); box-shadow: 0 6px 20px rgba(0,0,0,0.12); }}
.chart-card h4 {{ font-size: 13px; font-weight: 600; color: #0f3460; margin-bottom: 14px; }}
.chart-card canvas {{ max-height: 260px; }}
.chart-card-wide {{ background: white; border-radius: 10px; padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); grid-column: 1 / -1; }}
.chart-card-wide h4 {{ font-size: 13px; font-weight: 600; color: #0f3460; margin-bottom: 14px; }}
.table-wrap {{ overflow-x: auto; }}
.table-scroll-wrap {{ max-height: 68vh; min-height: 360px; overflow-y: auto; overflow-x: auto; border: 1px solid #e0e0e0; border-radius: 6px; }}
.table-scroll-wrap thead th {{ position: sticky; top: 0; z-index: 10; }}
.table-scroll-wrap tfoot td {{ position: sticky; bottom: 0; z-index: 10; background: #f0f6ff; font-weight: 700; }}
.data-table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
.data-table th {{ background: #0f3460; color: white; padding: 8px 8px; text-align: center; white-space: nowrap; font-weight: 600; }}
.data-table th.p1 {{ background: #6c757d; }}
.data-table th.p2 {{ background: #667eea; }}
.data-table th.p3 {{ background: #2980b9; }}
.data-table th.p4 {{ background: #c0392b; }}
.data-table th.hb {{ background: #e07b24; }}
.data-table th.green {{ background: #08845a; }}
.data-table td {{ padding: 6px 8px; text-align: center; border-bottom: 1px solid #f0f0f0; white-space: nowrap; }}
.data-table-compact {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
.data-table-compact th {{ background: #0f3460; color: white; padding: 6px 8px; text-align: center; white-space: nowrap; font-weight: 600; }}
.data-table-compact td {{ padding: 4px 8px; text-align: center; border-bottom: 1px solid #f0f0f0; white-space: nowrap; }}
.data-table tr:hover td, .data-table-compact tr:hover td {{ background: #f5f7ff; }}
.hb-up {{ color: #c0392b; font-weight: 700; }}
.hb-down {{ color: #08845a; font-weight: 700; }}
.hb-flat {{ color: #888; }}
.badge {{ display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 10px; color: white; font-weight: 600; }}
.badge-y {{ background: #08845a; }}
.badge-n {{ background: #e07b24; }}
.badge-un {{ background: #c0392b; }}
.badge-normal {{ background: #2980b9; }}
select {{ padding: 6px 12px; border-radius: 6px; border: 1px solid #ddd; font-size: 13px; background: white; cursor: pointer; }}
select:focus {{ outline: none; border-color: #0f3460; }}
.filter-bar {{ background: #f5f7ff; border-radius: 8px; padding: 14px 18px; margin-bottom: 14px; display: flex; flex-wrap: wrap; gap: 10px; align-items: center; }}
.filter-bar-sticky {{ position: sticky; top: -1px; z-index: 100; background: #f5f7ff; border-radius: 0; padding: 14px 18px; margin-bottom: 14px; display: flex; flex-wrap: wrap; gap: 10px; align-items: center; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
.filter-bar .fg {{ display: flex; align-items: center; gap: 6px; font-size: 13px; }}
.filter-bar .fg label {{ color: #555; white-space: nowrap; }}
.filter-bar .reset-btn {{ border: 1px solid #c0392b; color: #c0392b; background: white; border-radius: 6px; padding: 6px 14px; cursor: pointer; font-size: 13px; }}
.filter-bar .reset-btn:hover {{ background: #c0392b; color: white; }}
.filter-info {{ font-size: 13px; color: #0f3460; font-weight: 600; margin-bottom: 10px; }}
.plp-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 16px; margin-bottom: 20px; }}
.plp-card {{ background: white; border-radius: 10px; padding: 18px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); border-top: 3px solid #8e44ad; }}
.plp-card h4 {{ font-size: 13px; font-weight: 600; color: #0f3460; margin-bottom: 12px; }}
.plp-metric {{ display: flex; justify-content: space-between; padding: 4px 0; border-bottom: 1px solid #f0f0f0; font-size: 12px; }}
.plp-metric:last-child {{ border-bottom: none; }}
.plp-metric .lbl {{ color: #666; }}
.plp-metric .val {{ font-weight: 600; color: #1a1a2e; }}
.plp-highlight {{ background: #f5f0ff; border-radius: 6px; padding: 10px; margin-top: 10px; }}
.plp-highlight .val {{ color: #8e44ad; font-weight: 700; }}
.risk-card {{ padding: 14px 18px; border-radius: 8px; margin-bottom: 10px; border-left: 4px solid; }}
.risk-high {{ background: #fdeaea; border-left-color: #c0392b; color: #721c24; }}
.risk-medium {{ background: #fef5e7; border-left-color: #e07b24; color: #7d5e29; }}
.risk-low {{ background: #e8f8f0; border-left-color: #08845a; color: #155724; }}
.report-block {{ background: #f9fafb; border-radius: 8px; padding: 14px 18px; margin-bottom: 10px; border-left: 4px solid #0f3460; }}
.report-block h4 {{ font-size: 14px; font-weight: 600; color: #0f3460; margin-bottom: 8px; display: flex; justify-content: space-between; align-items: center; }}
.report-block pre {{ font-size: 13px; color: #333; white-space: pre-wrap; line-height: 1.6; }}
.copy-btn {{ background: #0f3460; color: white; border: none; border-radius: 6px; padding: 4px 14px; cursor: pointer; font-size: 12px; transition: background 0.2s; }}
.copy-btn:hover {{ background: #16213e; }}
.note {{ font-size: 12px; color: #888; margin-bottom: 10px; }}
::-webkit-scrollbar {{ width: 6px; height: 6px; }}
::-webkit-scrollbar-track {{ background: #f1f1f1; }}
::-webkit-scrollbar-thumb {{ background: #ccc; border-radius: 3px; }}
@media (max-width: 900px) {{ .sidebar {{ display: none; }} .chart-grid {{ grid-template-columns: 1fr; }} .kpi-grid {{ grid-template-columns: repeat(2, 1fr); }} }}
</style>
</head>
<body>
<div class="header">
  <h1>&#128202; 新品周报 &middot; 5.7-5.13 &nbsp;&#127793;</h1>
  <div class="subtitle">统计周期：2026年5月7日 - 5月13日 &nbsp;|&nbsp; 在跟SKU：{total_sku} &nbsp;|&nbsp; 生成：2026-05-18</div>
</div>
<div class="container">
<nav class="sidebar">
  <ul>
    <li><a href="javascript:void(0)" onclick="switchTab('tab1',this)" class="active">&#128202; 总盘概览</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab2',this)">&#128201; 低占比分析</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab3',this)">&#128176; 广告追踪</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab4',this)">&#128203; 四三累计</a></li>
    <li><a href="javascript:void(0)" onclick="switchTab('tab5',this)">&#128172; 汇报输出</a></li>
  </ul>
</nav>
<div class="main-content">

<!-- ===== TAB 1: 总盘概览 ===== -->
<div id="tab1" class="tab-content active">
<div class="kpi-grid" id="kpiOverview"></div>
<div class="chart-grid">
  <div class="chart-card"><h4>&#127922; 出单分布（有对手口径）</h4><canvas id="chartOrderPie"></canvas></div>
  <div class="chart-card"><h4>&#127991; 品线销量对比（本周 vs 上周）</h4><canvas id="chartCatSales"></canvas></div>
  <div class="chart-card"><h4>&#128101; 分析人销量对比（本周 vs 上周）</h4><canvas id="chartAnSales"></canvas></div>
  <div class="chart-card"><h4>&#9201; 分析及时率（各分析人）</h4><canvas id="chartTimely"></canvas></div>
  <div class="chart-card"><h4>&#128200; 拓展类型出单率对比（本周 vs 上周）</h4><canvas id="chartExpandRate"></canvas></div>
</div>
<div id="tab1Tables"></div>
</div>

<!-- ===== TAB 2: 低占比分析 ===== -->
<div id="tab2" class="tab-content">
<div class="chart-grid">
  <div class="chart-card"><h4>&#127991; 低占比品类分布</h4><canvas id="chartLowShareCat"></canvas></div>
  <div class="chart-card"><h4>&#128308; 未出单原因分布（有对手）</h4><canvas id="chartUnsoldReason"></canvas></div>
</div>
<div id="tab2Content"></div>
</div>

<!-- ===== TAB 3: 广告追踪 ===== -->
<div id="tab3" class="tab-content">
<div class="chart-grid">
  <div class="chart-card"><h4>&#127919; PLG广告分类分布</h4><canvas id="chartPlgClass"></canvas></div>
  <div class="chart-card"><h4>&#128200; 广告核心指标本周vs上周</h4><canvas id="chartAdMetrics"></canvas></div>
</div>
<div id="tab3Content"></div>
</div>

<!-- ===== TAB 4: 四三累计 ===== -->
<div id="tab4" class="tab-content">
<div id="tab4Content"></div>
</div>

<!-- ===== TAB 5: 汇报输出 ===== -->
<div id="tab5" class="tab-content">
<div id="tab5Content"></div>
</div>

</div></div>

<script>
// ===== Data =====
{js_data}

// ===== Utility Functions =====
function fmtNum(n) {{ return n == null || n === '' ? '-' : Number(n).toLocaleString('zh-CN'); }}
function fmtMoney(n) {{ return n == null || n === '' ? '-' : '$' + Number(n).toFixed(2); }}
function pct(v) {{ return v == null || v === '' ? '-' : v; }}
function acoasPct(v) {{
  if (v == null || v === '' || v === 0) return '0.00%';
  if (typeof v === 'string') {{ const n = parseFloat(v); return isNaN(n) ? v : n.toFixed(2) + '%'; }}
  return (v * 100).toFixed(2) + '%';
}}
function hbSign(v) {{
  if (typeof v !== 'string' || !v) return '<span class="hb-flat">-</span>';
  if (v === '-' || v === '0%' || v === '0.0%' || v === '+0%') return '<span class="hb-flat">-</span>';
  if (v.startsWith('+')) return '<span class="hb-up">' + v + '</span>';
  if (v.startsWith('-')) return '<span class="hb-down">' + v + '</span>';
  return '<span class="hb-flat">' + v + '</span>';
}}
function badge8d(v) {{
  if (v === 'Y') return '<span class="badge badge-y">Y</span>';
  if (v === 'N') return '<span class="badge badge-n">N</span>';
  return '<span class="badge badge-un">' + v + '</span>';
}}
function badgeStatus(v) {{
  if (!v) return '-';
  if (v === '正常') return '<span class="badge badge-normal">正常</span>';
  if (v === '竞争无优势') return '<span class="badge badge-n">竞争无优势</span>';
  if (v === '无市场') return '<span class="badge badge-un">无市场</span>';
  return v;
}}
function badgePLP(v) {{ return v === 'Y' ? '<span class="badge badge-y">Y</span>' : '<span class="badge badge-n">N</span>'; }}
function badgeAdClass(v) {{
  if (v === '单PLG且未出单') return '<span class="badge badge-un" style="background:#c0392b">' + v + '</span>';
  if (v === 'PLP+PLG同开') return '<span class="badge badge-y">' + v + '</span>';
  if (v === '单PLG') return '<span class="badge badge-n">' + v + '</span>';
  if (v === '单PLP') return '<span class="badge badge-normal">' + v + '</span>';
  return '<span class="badge" style="background:#888">' + v + '</span>';
}}

// ===== Tab Switching =====
function switchTab(tabId, el) {{
  document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
  document.getElementById(tabId).classList.add('active');
  document.querySelectorAll('.sidebar li a').forEach(a => a.classList.remove('active'));
  if (el) el.classList.add('active');
  if (tabId === 'tab1' && !window._charts1) {{ initCharts1(); window._charts1 = true; }}
  if (tabId === 'tab2' && !window._charts2) {{ setTimeout(function(){{ initCharts2(); }}, 80); window._charts2 = true; }}
  if (tabId === 'tab3' && !window._charts3) {{ setTimeout(function(){{ initCharts3(); }}, 80); window._charts3 = true; }}
}}

// ===== TAB 1: 总盘概览 =====
(function() {{
  // KPI cards
  var kpi = document.getElementById('kpiOverview');
  var hbSale = prevWeekKpi.salesQtyChange || '-';
  var hbRev = prevWeekKpi.revenueChange || '-';
  var hbTime = prevWeekKpi.timelyChange || '-';
  kpi.innerHTML =
    '<div class="kpi-card"><div class="val">' + cum43Stats.total + '</div><div class="label">&#128230; 在跟SKU总数</div><div class="hb" style="color:#888">本周新上架 +' + {new_list} + '</div></div>' +
    '<div class="kpi-card green"><div class="val">' + fmtNum({total_sales_curr}) + '</div><div class="label">&#128722; 总销量</div><div class="hb">' + hbSign(hbSale) + '</div></div>' +
    '<div class="kpi-card purple"><div class="val">' + fmtMoney({total_rev_curr}) + '</div><div class="label">&#128176; 总销售额(USD)</div><div class="hb">' + hbSign(hbRev) + '</div></div>' +
    '<div class="kpi-card green"><div class="val">{sale_rate_curr}</div><div class="label">&#127919; 出单率(有对手)</div><div class="hb" style="color:#888">已出单 ' + ({has_rival_y}+{has_rival_n}) + '/' + {has_rival_curr} + '</div></div>' +
    '<div class="kpi-card orange"><div class="val">{timely_rate}</div><div class="label">&#9201; 分析及时率</div><div class="hb">' + hbSign(hbTime) + '</div></div>';

  // Dimension tables
  var tablesHTML = '';

  // Order by analyst
  var ordAnY = {{}}, ordAnN = {{}}, ordAnNo = {{}};
  cum43Data.forEach(function(r) {{
    var an = r['4月分析人'];
    ordAnY[an] = (ordAnY[an]||0); ordAnN[an] = (ordAnN[an]||0); ordAnNo[an] = (ordAnNo[an]||0);
    if (r['5.13 8日出单情况'] === 'Y') ordAnY[an]++;
    else if (r['5.13 8日出单情况'] === 'N') ordAnN[an]++;
    else ordAnNo[an]++;
  }});

  tablesHTML += '<div class="section"><h3>&#128230; 新品出单情况（有对手口径）</h3>';
  tablesHTML += '<div class="note">出单率 = (Y+N)/有对手SKU &times; 100%</div>';
  tablesHTML += '<div class="kpi-grid" style="margin-bottom:16px">' +
    '<div class="kpi-card"><div class="val">{has_rival_curr}</div><div class="label">有对手SKU</div></div>' +
    '<div class="kpi-card green"><div class="val">' + ({has_rival_y}+{has_rival_n}) + '</div><div class="label">已出单(Y+N)</div></div>' +
    '<div class="kpi-card green"><div class="val">{sale_rate_curr}</div><div class="label">出单率</div></div></div>';

  tablesHTML += '<div class="sub-module"><h4>按分析人</h4><div class="table-wrap"><table class="data-table">' +
    '<tr><th>分析人</th><th>Y</th><th>N</th><th>未出单</th><th>出单率</th></tr>';
  var aOrd = {json.dumps(ans_ordered, ensure_ascii=False)};
  aOrd.forEach(function(an) {{
    var y = ordAnY[an]||0, n = ordAnN[an]||0, no = ordAnNo[an]||0, total = y+n+no;
    tablesHTML += '<tr><td>' + an + '</td><td>' + y + '</td><td>' + n + '</td><td>' + no + '</td><td>' + (total ? (y+n)/total*100 : 0).toFixed(1) + '%</td></tr>';
  }});
  tablesHTML += '</table></div></div>';

  // Category dimension table
  tablesHTML += '<div class="sub-module"><h4>按品线</h4><div class="table-wrap"><table class="data-table">' +
    '<tr><th>品线</th><th>SKU</th><th>本周新上架</th><th>销量</th><th>上周销量</th><th>销量环比</th><th>销售额</th><th>上周销售额</th><th>有对手SKU</th></tr>';
  categoryRevenueData.forEach(function(d) {{
    tablesHTML += '<tr><td>' + d.category + '</td><td>' + d.curSku + '</td><td>' + d.curNewSku + '</td><td>' + fmtNum(d.curSalesQty) + '</td><td>' + fmtNum(d.prevSalesQty) + '</td><td>' + hbSign(d.salesQtyChange) + '</td><td>' + fmtMoney(d.curRevenue) + '</td><td>' + fmtMoney(d.prevRevenue) + '</td><td>' + hbSign(d.revenueChange) + '</td></tr>';
  }});
  tablesHTML += '</table></div></div>';

  // Analyst dimension table
  tablesHTML += '<div class="sub-module"><h4>按分析人</h4><div class="table-wrap"><table class="data-table">' +
    '<tr><th>分析人</th><th>SKU</th><th>本周新上架</th><th>销量</th><th>上周销量</th><th>销量环比</th><th>销售额</th><th>上周销售额</th><th>销售额环比</th></tr>';
  analystRevenueData.forEach(function(d) {{
    tablesHTML += '<tr><td>' + d.analyst + '</td><td>' + d.curSku + '</td><td>' + d.curNewSku + '</td><td>' + fmtNum(d.curSalesQty) + '</td><td>' + fmtNum(d.prevSalesQty) + '</td><td>' + hbSign(d.salesQtyChange) + '</td><td>' + fmtMoney(d.curRevenue) + '</td><td>' + fmtMoney(d.prevRevenue) + '</td><td>' + hbSign(d.revenueChange) + '</td></tr>';
  }});
  tablesHTML += '</table></div></div>';

  // Expand type table
  tablesHTML += '<div class="sub-module"><h4>拓展类型</h4><div class="table-wrap"><table class="data-table">' +
    '<tr><th>拓展类型</th><th>SKU</th><th>出单SKU</th><th>出单率</th><th>上周出单率</th><th>销量</th><th>销量环比</th><th>销售额</th><th>销售额环比</th></tr>';
  expandTypeData.forEach(function(d) {{
    tablesHTML += '<tr><td>' + d.expandType + '</td><td>' + d.curSku + '</td><td>' + d.curSalesCount + '</td><td>' + d.curSalesRate + '</td><td>' + d.prevSalesRate + '</td><td>' + fmtNum(d.curSalesQty) + '</td><td>' + hbSign(d.salesQtyChange) + '</td><td>' + fmtMoney(d.curRevenue) + '</td><td>' + hbSign(d.revenueChange) + '</td></tr>';
  }});
  tablesHTML += '</table></div></div>';

  // Timeliness table
  tablesHTML += '<div class="sub-module"><h4>分析及时率</h4><div class="table-wrap"><table class="data-table">' +
    '<tr><th>分析人</th><th>本周及时分析</th><th>上周及时分析</th><th>8日无分析</th><th>超7日未分析</th><th>SKU总数</th><th>本周及时率</th><th>上周及时率</th><th>环比变化</th></tr>';
  timelinessData.analysts.forEach(function(d) {{
    tablesHTML += '<tr><td>' + d.analyst + '</td><td>' + d.timelyCount + '</td><td>' + d.prevTimelyCount + '</td><td>' + d.noAnalysis8dCount + '</td><td>' + d.noAnalysis7dCount + '</td><td>' + d.curSku + '</td><td>' + d.timelyRate + '</td><td>' + d.prevTimelyRate + '</td><td>' + hbSign(d.timelyChange) + '</td></tr>';
  }});
  var t = timelinessData.total;
  tablesHTML += '<tr style="font-weight:700;background:#f0f6ff"><td>合计</td><td>' + t.timelyCount + '</td><td>' + t.prevTimelyCount + '</td><td>' + t.noAnalysis8dCount + '</td><td>' + t.noAnalysis7dCount + '</td><td>' + t.curSku + '</td><td>' + t.timelyRate + '</td><td>' + t.prevTimelyRate + '</td><td>' + hbSign(t.timelyChange) + '</td></tr>';
  tablesHTML += '</table></div></div>';

  tablesHTML += '</div>';
  document.getElementById('tab1Tables').innerHTML = tablesHTML;
}})();

// ===== TAB 1 Charts =====
window._charts1 = false;
function initCharts1() {{
  // Order pie
  new Chart(document.getElementById('chartOrderPie'), {{
    type: 'doughnut',
    data: {{ labels: ['8日内出单(Y)','8日外出单(N)','未出单'], datasets: [{{ data: [{y_curr},{n_curr},{no_curr}], backgroundColor: ['#08845a','#e07b24','#c0392b'], borderWidth: 2, borderColor: '#fff' }}] }},
    options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom' }} }}, animation: {{ animateRotate: true, animateScale: true, duration: 800 }} }}
  }});

  // Category sales bar
  var catLabels = categoryRevenueData.map(d => d.category);
  new Chart(document.getElementById('chartCatSales'), {{
    type: 'bar',
    data: {{ labels: catLabels, datasets: [
      {{ label: '本周销量', data: categoryRevenueData.map(d => d.curSalesQty), backgroundColor: '#667eea' }},
      {{ label: '上周销量', data: categoryRevenueData.map(d => d.prevSalesQty), backgroundColor: '#ccc' }}
    ] }},
    options: {{ responsive: true, plugins: {{ legend: {{ position: 'top' }} }}, scales: {{ y: {{ beginAtZero: true }} }} }}
  }});

  // Analyst sales bar
  var anLabels = analystRevenueData.map(d => d.analyst);
  new Chart(document.getElementById('chartAnSales'), {{
    type: 'bar',
    data: {{ labels: anLabels, datasets: [
      {{ label: '本周销量', data: analystRevenueData.map(d => d.curSalesQty), backgroundColor: '#2980b9' }},
      {{ label: '上周销量', data: analystRevenueData.map(d => d.prevSalesQty), backgroundColor: '#ccc' }}
    ] }},
    options: {{ responsive: true, plugins: {{ legend: {{ position: 'top' }} }}, scales: {{ y: {{ beginAtZero: true }} }} }}
  }});

  // Timely bar
  new Chart(document.getElementById('chartTimely'), {{
    type: 'bar',
    data: {{ labels: timelinessData.analysts.map(d => d.analyst), datasets: [
      {{ label: '及时率(%)', data: timelinessData.analysts.map(d => parseFloat(d.timelyRate)), backgroundColor: '#e07b24' }}
    ] }},
    options: {{ responsive: true, plugins: {{ legend: {{ display: false }} }}, scales: {{ y: {{ beginAtZero: true, max: 100 }} }} }}
  }});

  // Expand type order rate comparison
  new Chart(document.getElementById('chartExpandRate'), {{
    type: 'bar',
    data: {{ labels: expandTypeData.map(d => d.expandType), datasets: [
      {{ label: '本周出单率(%)', data: expandTypeData.map(d => parseFloat(d.curSalesRate)), backgroundColor: '#667eea' }},
      {{ label: '上周出单率(%)', data: expandTypeData.map(d => parseFloat(d.prevSalesRate)), backgroundColor: '#ccc' }}
    ] }},
    options: {{ responsive: true, plugins: {{ legend: {{ position: 'top' }} }}, scales: {{ y: {{ beginAtZero: true, max: 100 }} }} }}
  }});
}}

// ===== TAB 2 Charts (lazy) =====
window._charts2 = false;
function initCharts2() {{
  // Low share by category doughnut
  var lowCatMap = {{}};
  lowShareData.forEach(function(d) {{ lowCatMap[d.category] = (lowCatMap[d.category]||0) + 1; }});
  var lowCatLabels = Object.keys(lowCatMap);
  var lowCatData = Object.values(lowCatMap);
  var el1 = document.getElementById('chartLowShareCat');
  if (el1) new Chart(el1, {{
    type: 'doughnut',
    data: {{ labels: lowCatLabels, datasets: [{{ data: lowCatData, backgroundColor: ['#667eea','#08845a','#e07b24','#c0392b','#8e44ad','#3498db','#f39c12','#1abc9c'], borderWidth: 2, borderColor: '#fff' }}] }},
    options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom' }} }}, animation: {{ animateRotate: true, animateScale: true, duration: 800 }} }}
  }});

  // Unsold reason distribution bar
  var reasonLabels = hasCompetitorUnsold.reasons.map(function(d){{ return d.reason; }});
  var reasonData = hasCompetitorUnsold.reasons.map(function(d){{ return d.count; }});
  var el2 = document.getElementById('chartUnsoldReason');
  if (el2) new Chart(el2, {{
    type: 'bar',
    data: {{ labels: reasonLabels, datasets: [
      {{ label: 'SKU数', data: reasonData, backgroundColor: ['#c0392b','#e07b24','#f39c12','#3498db','#08845a','#95a5a6','#7f8c8d'] }}
    ] }},
    options: {{ responsive: true, plugins: {{ legend: {{ display: false }} }}, scales: {{ y: {{ beginAtZero: true }} }} }}
  }});
}}

// ===== TAB 3 Charts (lazy) =====
window._charts3 = false;
function initCharts3() {{
  // PLG class distribution doughnut
  var plgLabels = ['PLP+PLG同开','单PLG','单PLP','单PLG且未出单','无广告'];
  var plgData = [plgStats.plpAndPlgBothCount, plgStats.plgOnlyCount, plgStats.plpOnlyCount, plgStats.plgUnsoldCount, plgStats.noAdCount];
  var el3 = document.getElementById('chartPlgClass');
  if (el3) new Chart(el3, {{
    type: 'doughnut',
    data: {{ labels: plgLabels, datasets: [{{ data: plgData, backgroundColor: ['#08845a','#e07b24','#2980b9','#c0392b','#95a5a6'], borderWidth: 2, borderColor: '#fff' }}] }},
    options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom' }} }}, animation: {{ animateRotate: true, animateScale: true, duration: 800 }} }}
  }});

  // Ad metrics comparison bar (ROAS/CVR/CTR)
  var metricLabels = ['ROAS','CVR(%)','CTR(%)','ACOS(%)'];
  var curMetrics = [parseFloat(plpTotal.roas)||0, parseFloat(plpTotal.cvr)||0, parseFloat(plpTotal.ctr)||0, parseFloat(plpTotal.acos)||0];
  var prevMetrics = [parseFloat(plpPrevTotal.roas)||0, parseFloat(plpPrevTotal.cvr)||0, parseFloat(plpPrevTotal.ctr)||0, parseFloat(plpPrevTotal.acos)||0];
  var el4 = document.getElementById('chartAdMetrics');
  if (el4) new Chart(el4, {{
    type: 'bar',
    data: {{ labels: metricLabels, datasets: [
      {{ label: '本周', data: curMetrics, backgroundColor: '#667eea' }},
      {{ label: '上周', data: prevMetrics, backgroundColor: '#ccc' }}
    ] }},
    options: {{ responsive: true, plugins: {{ legend: {{ position: 'top' }} }}, scales: {{ y: {{ beginAtZero: true }} }} }}
  }});
}}

// ===== TAB 2: 低占比分析 =====
(function() {{
  var html = '';
  html += '<div class="kpi-grid">' +
    '<div class="kpi-card orange"><div class="val">' + hasCompetitorUnsold.total + '</div><div class="label">&#9888; 有对手未出单</div></div>' +
    '<div class="kpi-card red"><div class="val">' + unsoldNoCompetitor.total + '</div><div class="label">&#10060; 无对手未出单</div></div>' +
    '<div class="kpi-card"><div class="val">' + cum43Stats.total + '</div><div class="label">&#128269; 有对手SKU总数</div></div>' +
    '<div class="kpi-card orange"><div class="val">' + lowShareData.length + '</div><div class="label">&#128201; 低占比新品</div></div>' +
  '</div>';

  // Low share detail table
  html += '<div class="section"><h3>&#128201; 低占比新品明细 (' + lowShareData.length + '个)</h3>';
  html += '<div class="note">条件：市占比 < 75% 且 对手销量 > 0</div>';
  html += '<div class="table-wrap"><table class="data-table"><tr>' +
    '<th>销售编号</th><th>SKU</th><th>上架日期</th><th>分析人</th><th>品类</th><th>拓展类型</th>' +
    '<th>销量</th><th>销量环比</th><th>销售额</th><th>销售额环比</th>' +
    '<th>对手销量</th><th>市占比</th><th>上周市场状态</th><th>本周市场状态</th><th>8日出单</th><th>PLP</th><th>ACOAS</th></tr>';
  lowShareData.forEach(function(d) {{
    html += '<tr><td>' + (d.salesCode||'') + '</td><td>' + d.sku + '</td><td>' + (d.launchDate||'') + '</td><td>' + d.analyst + '</td><td>' + d.category + '</td><td>' + d.expandType + '</td>' +
      '<td>' + d.curSalesQty + '</td><td>' + hbSign(d.salesQtyChange) + '</td><td>' + fmtMoney(d.curRevenue) + '</td><td>' + hbSign(d.revenueChange) + '</td>' +
      '<td>' + d.curCompetitorQty + '</td><td>' + d.curMarketShare + '</td><td>' + badgeStatus(d.prevMarketStatus) + '</td><td>' + badgeStatus(d.curMarketStatus) + '</td><td>' + badge8d(d.cur8dStatus) + '</td><td>' + badgePLP(d.plpEnabled) + '</td>' +
      '<td style="color:' + (d.acoas && parseFloat(d.acoas) > 10 ? '#c0392b' : '#333') + '">' + (d.acoas||'N/A') + '</td></tr>';
  }});
  html += '</table></div></div>';

  // Unsold analysis A
  html += '<div class="section"><h3>&#128308; A. 有对手未出单新品 (' + hasCompetitorUnsold.total + '个)</h3>';
  html += '<div class="table-wrap"><table class="data-table"><tr>' +
    '<th>分析人</th><th>竞争无优势</th><th>无市场</th><th>无价格优势</th><th>站外出单</th><th>正常</th><th>N/A</th><th>未知</th><th>合计</th></tr>';
  hasCompetitorUnsold.byAnalyst.forEach(function(d) {{
    html += '<tr><td>' + d.analyst + '</td><td>' + (d['竞争无优势']||0) + '</td><td>' + (d['无市场']||0) + '</td><td>' + (d['站内无价格优势']||0) + '</td><td>' + (d['站外出单']||0) + '</td><td>' + (d['正常']||0) + '</td><td>' + (d['#N/A']||0) + '</td><td>' + (d['未知']||0) + '</td><td>' + d.total + '</td></tr>';
  }});
  html += '</table></div></div>';

  // Unsold analysis B
  html += '<div class="section"><h3>&#10060; B. 无对手未出单新品 (' + unsoldNoCompetitor.total + '个)</h3>';
  html += '<div class="table-wrap"><table class="data-table"><tr>' +
    '<th>分析人</th><th>无市场</th><th>未知</th><th>竞争无优势</th><th>N/A</th><th>其他</th><th>合计</th></tr>';
  unsoldNoCompetitor.byAnalyst.forEach(function(d) {{
    html += '<tr><td>' + d.analyst + '</td><td>' + (d['无市场']||0) + '</td><td>' + (d['未知']||0) + '</td><td>' + (d['竞争无优势']||0) + '</td><td>' + (d['#N/A']||0) + '</td><td>' + (d['其他']||0) + '</td><td>' + d.total + '</td></tr>';
  }});
  html += '</table></div></div>';

  document.getElementById('tab2Content').innerHTML = html;
}})();

// ===== TAB 3: 广告追踪 =====
(function() {{
  var html = '';
  // PLP KPI cards
  html += '<div class="kpi-grid">' +
    '<div class="kpi-card"><div class="val">' + plpTotal.campaignCount + '</div><div class="label">&#128205; 广告活动数</div></div>' +
    '<div class="kpi-card"><div class="val">' + plpTotal.linkCount + '</div><div class="label">&#128279; 投放链接数</div></div>' +
    '<div class="kpi-card"><div class="val">' + fmtNum(plpTotal.impression) + '</div><div class="label">&#128065; 曝光量</div></div>' +
    '<div class="kpi-card"><div class="val">' + fmtNum(plpTotal.click) + '</div><div class="label">&#128433; 点击量</div></div>' +
    '<div class="kpi-card green"><div class="val">' + plpTotal.sold + '</div><div class="label">&#128722; 售出数</div></div>' +
    '<div class="kpi-card purple"><div class="val">' + fmtMoney(plpTotal.revenue) + '</div><div class="label">&#128176; 广告销售额</div></div>' +
  '</div>';

  // PLP core metrics with prev comparison
  html += '<div class="section"><h3>&#128200; PLP核心指标</h3>';
  html += '<div class="plp-grid">';
  var plpMetrics = [
    {{ label: 'ROAS', curr: plpTotal.roas, prev: plpPrevTotal.roas }},
    {{ label: 'CVR', curr: plpTotal.cvr, prev: plpPrevTotal.cvr }},
    {{ label: 'CTR', curr: plpTotal.ctr, prev: plpPrevTotal.ctr }},
    {{ label: 'CPC', curr: plpTotal.cpc, prev: plpPrevTotal.cpc }},
    {{ label: 'CPA', curr: plpTotal.cpa, prev: plpPrevTotal.cpa }},
    {{ label: 'ACOS', curr: plpTotal.acos, prev: plpPrevTotal.acos }},
    {{ label: 'ACOAS', curr: plpTotal.acoas, prev: plpPrevTotal.acoas }},
  ];
  plpMetrics.forEach(function(m) {{
    html += '<div class="plp-card"><h4>' + m.label + '</h4>';
    html += '<div class="plp-metric"><span class="lbl">本周</span><span class="val">' + m.curr + '</span></div>';
    html += '<div class="plp-metric"><span class="lbl">上周</span><span class="val">' + m.prev + '</span></div>';
    html += '<div class="plp-highlight"><span class="val">花费: ' + fmtMoney(plpTotal.cost) + ' | 总销售额: ' + fmtMoney(plpTotal.totalRevenue) + '</span></div>';
    html += '</div>';
  }});
  html += '</div></div>';

  // PLP analyst dimension
  html += '<div class="section"><h3>&#128101; 分析人维度PLP</h3><div class="table-wrap"><table class="data-table"><tr>' +
    '<th>分析人</th><th>活动数</th><th>链接数</th><th>曝光</th><th>点击</th><th>售出</th><th>花费</th><th>广告销售额</th><th>ROAS</th><th>CVR</th><th>CTR</th><th>CPC</th><th>CPA</th><th>ACOS</th><th>ACOAS</th></tr>';
  plpAnalysts.forEach(function(d) {{
    html += '<tr><td>' + d.analyst + '</td><td>' + d.campaigns + '</td><td>' + d.links + '</td><td>' + fmtNum(d.imp) + '</td><td>' + fmtNum(d.click) + '</td><td>' + d.sold + '</td><td>' + fmtMoney(d.cost) + '</td><td>' + fmtMoney(d.ad_rev) + '</td><td>' + d.roas + '</td><td>' + d.cvr + '</td><td>' + d.ctr + '</td><td>' + d.cpc + '</td><td>' + d.cpa + '</td><td>' + d.acos + '</td><td>' + d.acoas + '</td></tr>';
  }});
  html += '</table></div></div>';

  // PLP category dimension
  html += '<div class="section"><h3>&#127991; 品线维度 PLP</h3><div class="table-wrap"><table class="data-table"><tr>' +
    '<th>品线</th><th>活动数</th><th>链接数</th><th>曝光</th><th>点击</th><th>售出</th><th>花费</th><th>广告销售额</th><th>ROAS</th><th>CVR</th><th>CTR</th><th>CPC</th><th>CPA</th><th>ACOS</th><th>ACOAS</th></tr>';
  plpCategories.forEach(function(d) {{
    html += '<tr><td>' + d.category + '</td><td>' + d.campaigns + '</td><td>' + d.links + '</td><td>' + fmtNum(d.imp) + '</td><td>' + fmtNum(d.click) + '</td><td>' + d.sold + '</td><td>' + fmtMoney(d.cost) + '</td><td>' + fmtMoney(d.ad_rev) + '</td><td>' + d.roas + '</td><td>' + d.cvr + '</td><td>' + d.ctr + '</td><td>' + d.cpc + '</td><td>' + d.cpa + '</td><td>' + d.acos + '</td><td>' + d.acoas + '</td></tr>';
  }});
  html += '</table></div></div>';

  // PLG stats
  html += '<div class="section"><h3>&#128293; PLG费率统计</h3>';
  html += '<div class="kpi-grid" style="margin-bottom:16px">' +
    '<div class="kpi-card"><div class="val">' + plgStats.totalNewProducts + '</div><div class="label">&#128230; 新品总数</div></div>' +
    '<div class="kpi-card green"><div class="val">' + plgStats.plpAndPlgBothCount + '</div><div class="label">&#9989; PLP+PLG同开</div></div>' +
    '<div class="kpi-card orange"><div class="val">' + plgStats.plgOnlyCount + '</div><div class="label">&#128993; 单PLG</div></div>' +
    '<div class="kpi-card"><div class="val">' + plgStats.plpOnlyCount + '</div><div class="label">&#128309; 单PLP</div></div>' +
    '<div class="kpi-card red"><div class="val">' + plgStats.plgUnsoldCount + '</div><div class="label">&#128308; 单PLG且未出单</div></div>' +
    '<div class="kpi-card"><div class="val">' + plgStats.noAdCount + '</div><div class="label">&#9898; 无广告</div></div>' +
  '</div>';

  html += '<div class="table-wrap"><table class="data-table"><tr>' +
    '<th>分析人</th><th>总数</th><th>PLP+PLG同开</th><th>单PLG</th><th>单PLP</th><th>单PLG且未出单</th><th>无广告</th></tr>';
  plgStats.byAnalyst.forEach(function(d) {{
    html += '<tr><td>' + d.analyst + '</td><td>' + d.total + '</td><td>' + d.plpAndPlgBoth + '</td><td>' + d.plgOnly + '</td><td>' + d.plpOnly + '</td><td style="color:#c0392b;font-weight:700">' + d.plgUnsold + '</td><td>' + d.noAd + '</td></tr>';
  }});
  html += '</table></div></div>';

  // PLP detail table
  // PLP detail with sort + filter
  plpDetailData.sort(function(a,b) {{ return (b.salesQty||0)-(a.salesQty||0) || (b.clicks||0)-(a.clicks||0) || (b.impressions||0)-(a.impressions||0); }});
  var adClasses = [...new Set(plpDetailData.map(d => d.adClass).filter(Boolean))].sort();
  html += '<div class="section"><h3>&#128203; PLP广告明细 (' + plpDetailData.length + '条)</h3>';
  html += '<div class="filter-bar" style="margin-bottom:10px"><div class="fg"><label>广告分类</label><select id="fPlpAdClass" onchange="renderPlpDetail()"><option value="">全部</option>' + adClasses.map(c => '<option>' + c + '</option>').join('') + '</select></div><span style="font-size:13px;color:#555">排序: 售出↓ 点击↓ 曝光↓</span></div>';
  html += '<div class="table-scroll-wrap"><table class="data-table-compact" id="plpDetailTable"><thead><tr>' +
    '<th>SKU</th><th>广告活动</th><th>分析人</th><th>品类</th><th>拓展类型</th>' +
    '<th>曝光</th><th>点击</th><th>售出</th><th>花费</th><th>广告销售额</th><th>总销售额</th>' +
    '<th>ROAS</th><th>CVR</th><th>CTR</th><th>CPC</th><th>CPA</th><th>ACOS</th><th>ACOAS</th><th>广告分类</th></tr></thead><tbody></tbody></table></div></div>';

  window.renderPlpDetail = function() {{
    var sel = document.getElementById('fPlpAdClass');
    var f = sel ? sel.value : '';
    var tbody = document.querySelector('#plpDetailTable tbody');
    if (!tbody) return;
    var rows = '';
    plpDetailData.forEach(function(d) {{
      if (f && d.adClass !== f) return;
      rows += '<tr><td>' + d.sku + '</td><td style="font-size:10px">' + d.campaign + '</td><td>' + d.analyst + '</td><td>' + d.category + '</td><td>' + d.productExpansion + '</td>' +
        '<td>' + fmtNum(d.impressions) + '</td><td>' + fmtNum(d.clicks) + '</td><td>' + d.salesQty + '</td><td>' + fmtMoney(d.spend) + '</td><td>' + fmtMoney(d.adRevenue) + '</td><td>' + fmtMoney(d.totalRevenue) + '</td>' +
        '<td>' + (d.roas||0).toFixed(2) + '</td><td>' + ((d.cvr||0)*100).toFixed(2) + '%</td><td>' + ((d.ctr||0)*100).toFixed(2) + '%</td><td>$' + (d.cpc||0).toFixed(2) + '</td><td>$' + (d.cpa||0).toFixed(2) + '</td><td>' + ((d.acos||0)*100).toFixed(2) + '%</td><td style="color:' + (d.acoas>0.1?'#c0392b':'#333') + '">' + ((d.acoas||0)*100).toFixed(2) + '%</td>' +
        '<td>' + badgeAdClass(d.adClass) + '</td></tr>';
    }});
    tbody.innerHTML = rows;
  }};
  document.getElementById('tab3Content').innerHTML = html;
  window.renderPlpDetail();
}})();

// ===== TAB 4: 四三累计 =====
(function() {{
  // Build filter options from data
  var categories = [...new Set(cum43Data.map(r => r['品类']))].sort();
  var analysts = [...new Set(cum43Data.map(r => r['4月分析人']))].sort();
  var expands = [...new Set(cum43Data.map(r => r['产品拓展']))].sort();

  var html = '';
  html += '<div class="kpi-grid">' +
    '<div class="kpi-card"><div class="val">' + cum43Stats.total + '</div><div class="label">&#128230; 累计总SKU</div></div>' +
    '<div class="kpi-card green"><div class="val">' + (cum43Stats.yCount + cum43Stats.nCount) + '</div><div class="label">&#9989; 已出单</div></div>' +
    '<div class="kpi-card orange"><div class="val">' + hasCompetitorUnsold.total + '</div><div class="label">&#9888; 有竞争未出单</div></div>' +
    '<div class="kpi-card red"><div class="val">' + cum43Stats.noMarketCount + '</div><div class="label">&#10060; 无市场</div></div>' +
    '<div class="kpi-card"><div class="val">' + cum43Stats.normalCount + '</div><div class="label">&#9989; 市场正常</div></div>' +
    '<div class="kpi-card orange"><div class="val">' + cum43Stats.competitiveCount + '</div><div class="label">&#128308; 竞争无优势</div></div>' +
  '</div>';

  // Filters - sticky
  var adClassOptions = ['PLP+PLG同开','单PLG','单PLP','单PLG且未出单','无广告'];
  html += '<div class="filter-bar-sticky">' +
    '<div class="fg"><label>市场状态</label><select id="fMkt"><option value="">全部</option><option>正常</option><option>竞争无优势</option><option>无市场</option></select></div>' +
    '<div class="fg"><label>分析人</label><select id="fAn"><option value="">全部</option>' + analysts.map(a => '<option>' + a + '</option>').join('') + '</select></div>' +
    '<div class="fg"><label>品类</label><select id="fCat"><option value="">全部</option>' + categories.map(c => '<option>' + c + '</option>').join('') + '</select></div>' +
    '<div class="fg"><label>拓展类型</label><select id="fExp"><option value="">全部</option>' + expands.map(e => '<option>' + e + '</option>').join('') + '</select></div>' +
    '<div class="fg"><label>8日出单</label><select id="fOrd"><option value="">全部</option><option>Y</option><option>N</option><option>未出单</option></select></div>' +
    '<div class="fg"><label>市占比</label><select id="fShare"><option value="">全部</option><option value="high">≥75%</option><option value="mid">50%~75%</option><option value="low"><50%</option></select></div>' +
    '<div class="fg"><label>竞争</label><select id="fRival"><option value="">全部</option><option value="yes">有竞争</option><option value="no">无竞争</option></select></div>' +
    '<div class="fg"><label>广告分类</label><select id="fAdClass"><option value="">全部</option>' + adClassOptions.map(c => '<option>' + c + '</option>').join('') + '</select></div>' +
    '<button class="reset-btn" onclick="resetFilters()">取消筛选</button>' +
  '</div>';
  html += '<div class="filter-info" id="filterInfo">筛选结果：' + cum43Data.length + ' / ' + cum43Data.length + ' 条</div>';

  // Table
  html += '<div class="table-scroll-wrap"><table class="data-table" id="tab4Table"><thead><tr>' +
    '<th>SKU</th><th>实际上架</th><th>首次出单</th><th>分析人</th><th>品类</th><th>拓展类型</th>' +
    '<th>本周销量</th><th>本周销售额</th><th>对手销量</th><th>市占比</th>' +
    '<th>市场状态</th><th>8日出单</th><th>PLP</th><th>PLG费率</th><th>广告分类</th></tr></thead><tbody></tbody><tfoot></tfoot></table></div>';

  document.getElementById('tab4Content').innerHTML = html;

  // Render table
  window.renderTab4 = function() {{
    var getVal = function(id) {{ var el = document.getElementById(id); return el ? el.value : ''; }};
    var mkt = getVal('fMkt');
    var an = getVal('fAn');
    var cat = getVal('fCat');
    var exp = getVal('fExp');
    var ord = getVal('fOrd');
    var share = getVal('fShare');
    var rival = getVal('fRival');
    var adClass = getVal('fAdClass');

    // Build PLG map
    var plgMap = {{}};
    plgRecords.forEach(function(r) {{ plgMap[r.sku] = r; }});

    var filtered = cum43Data.filter(function(r) {{
      if (mkt && r['5.13市场状态'] !== mkt) return false;
      if (an && r['4月分析人'] !== an) return false;
      if (cat && r['品类'] !== cat) return false;
      if (exp && r['产品拓展'] !== exp) return false;
      if (ord && r['5.13 8日出单情况'] !== ord) return false;
      if (share === 'high' && r['5.13市占比'] < 75) return false;
      if (share === 'mid' && (r['5.13市占比'] < 50 || r['5.13市占比'] >= 75)) return false;
      if (share === 'low' && r['5.13市占比'] >= 50) return false;
      if (rival === 'yes' && r['5.13对手销量'] <= 0) return false;
      if (rival === 'no' && r['5.13对手销量'] > 0) return false;
      if (adClass && r['广告分类'] !== adClass) return false;
      return true;
    }});

    var infoEl = document.getElementById('filterInfo');
    if (infoEl) infoEl.textContent = '筛选结果：' + filtered.length + ' / ' + cum43Data.length + ' 条';

    var tbody = '';
    var totalSales = 0, totalRev = 0;
    filtered.forEach(function(r) {{
      var pg = plgMap[r.SKU] || {{}};
      totalSales += r['5.7-5.13销量']||0;
      totalRev += r['5.7-5.13销售额']||0;
      tbody += '<tr><td>' + r.SKU + '</td><td>' + (r['实际上架日期']||'') + '</td><td>' + (r['首次出单日期']||'') + '</td><td>' + r['4月分析人'] + '</td><td>' + r['品类'] + '</td><td>' + r['产品拓展'] + '</td>' +
        '<td>' + (r['5.7-5.13销量']||0) + '</td><td>' + fmtMoney(r['5.7-5.13销售额']) + '</td><td>' + (r['5.13对手销量']||0) + '</td><td>' + (r['5.13市占比']||0) + '%</td>' +
        '<td>' + badgeStatus(r['5.13市场状态']) + '</td><td>' + badge8d(r['5.13 8日出单情况']) + '</td><td>' + badgePLP(r['开启PLP']) + '</td><td>' + (pg.plgFee||'0%') + '</td><td>' + badgeAdClass(r['广告分类']) + '</td></tr>';
    }});
    var tfoot = '<tr><td colspan="6">合计 (' + filtered.length + '条)</td><td>' + totalSales + '</td><td>' + fmtMoney(totalRev) + '</td><td colspan="7"></td></tr>';
    var tab4Body = document.querySelector('#tab4Table tbody');
    if (tab4Body) tab4Body.innerHTML = tbody;
    var tab4Foot = document.querySelector('#tab4Table tfoot');
    if (tab4Foot) tab4Foot.innerHTML = tfoot;
  }};

  // Bind filter events
  ['fMkt','fAn','fCat','fExp','fOrd','fShare','fRival','fAdClass'].forEach(function(id) {{
    var el = document.getElementById(id);
    if (el) el.addEventListener('change', window.renderTab4);
  }});
  window.renderTab4();
}})();

function resetFilters() {{
  ['fMkt','fAn','fCat','fExp','fOrd','fShare','fRival','fAdClass'].forEach(function(id) {{
    var el = document.getElementById(id);
    if (el) el.value = '';
  }});
  window.renderTab4();
}}

// ===== TAB 5: 汇报输出 =====
(function() {{
  var html = '';
  html += '<div class="kpi-grid">' +
    '<div class="kpi-card"><div class="val">' + cum43Stats.total + '</div><div class="label">&#128230; 在售SKU</div></div>' +
    '<div class="kpi-card green"><div class="val">' + fmtNum({total_sales_curr}) + '</div><div class="label">&#128722; 总销量</div></div>' +
    '<div class="kpi-card purple"><div class="val">' + fmtMoney({total_rev_curr}) + '</div><div class="label">&#128176; 总销售额</div></div>' +
    '<div class="kpi-card green"><div class="val">{sale_rate_curr}</div><div class="label">&#127919; 出单率</div></div>' +
    '<div class="kpi-card orange"><div class="val">{timely_rate}</div><div class="label">&#9201; 及时率</div></div>' +
    '<div class="kpi-card red"><div class="val">' + lowShareData.length + '</div><div class="label">&#128201; 低占比新品</div></div>' +
  '</div>';

  // Risk detection
  var risks = [];
  var soldRate = parseFloat('{sale_rate_curr}');
  if (soldRate < 70) risks.push({{ level: 'high', text: '出单率仅' + soldRate + '%，低于70%警戒线，需重点关注未出单新品' }});
  if (unsoldNoCompetitor.total > 15) risks.push({{ level: 'high', text: '无对手未出单新品达' + unsoldNoCompetitor.total + '款，占比过高，建议排查上架时间及市场需求' }});
  var plpRoas = parseFloat(plpTotal.roas);
  if (plpRoas < 8) risks.push({{ level: 'medium', text: 'PLP广告ROAS为' + plpTotal.roas + '，投放效率需优化' }});
  if (timelinessData.analysts.some(function(d) {{ return parseFloat(d.timelyRate) < 50; }})) risks.push({{ level: 'high', text: '部分分析人及时率低于50%，需重点跟进' }});
  if (cum43Stats.competitiveCount > cum43Stats.normalCount) risks.push({{ level: 'medium', text: '竞争无优势SKU(' + cum43Stats.competitiveCount + ')超过市场正常SKU(' + cum43Stats.normalCount + ')，竞争环境需关注' }});
  if (risks.length === 0) risks.push({{ level: 'low', text: '本期各项指标整体平稳，暂无重大风险。' }});

  html += '<div class="section"><h3>&#9888; 风险预警</h3>';
  risks.forEach(function(r) {{
    var cls = r.level === 'high' ? 'risk-high' : r.level === 'medium' ? 'risk-medium' : 'risk-low';
    var icon = r.level === 'high' ? '&#128308;' : r.level === 'medium' ? '&#128993;' : '&#128994;';
    html += '<div class="risk-card ' + cls + '">' + icon + ' ' + r.text + '</div>';
  }});
  html += '</div>';

  // Report copy blocks
  var reportTexts = [
    {{ title: '核心指标', text: '【5.7-5.13新品周报核心指标】\\n在跟SKU：' + cum43Stats.total + '款（+{new_list}新上架）\\n总销量：{total_sales_curr}单（环比' + prevWeekKpi.salesQtyChange + '）\\n总销售额：${total_rev_curr:,.2f}（环比' + prevWeekKpi.revenueChange + '）\\n出单率：{sale_rate_curr}（有对手口径，Y:{has_rival_y} N:{has_rival_n} 未:{has_rival_no}）\\n分析及时率：{timely_rate}\\n低占比新品：' + lowShareData.length + '款' }},
    {{ title: 'PLP广告', text: '【PLP广告(' + plpTotal.campaignCount + '活动/' + plpTotal.linkCount + '链接)】\\n曝光：' + fmtNum(plpTotal.impression) + ' | 点击：' + fmtNum(plpTotal.click) + ' | 售出：' + plpTotal.sold + '\\n花费：' + fmtMoney(plpTotal.cost) + ' | 广告销售额：' + fmtMoney(plpTotal.revenue) + '\\nROAS：' + plpTotal.roas + ' | CVR：' + plpTotal.cvr + ' | ACOS：' + plpTotal.acos + ' | ACOAS：' + plpTotal.acoas }},
    {{ title: 'PLG分类', text: '【PLG广告分类】\\nPLP+PLG同开：' + plgStats.plpAndPlgBothCount + ' | 单PLG：' + plgStats.plgOnlyCount + ' | 单PLP：' + plgStats.plpOnlyCount + '\\n单PLG且未出单：' + plgStats.plgUnsoldCount + ' | 无广告：' + plgStats.noAdCount }}
  ];
  var plpRoasNum = parseFloat(plpTotal.roas);

  html += '<div class="section"><h3>&#128221; 可复制周报文案</h3>';
  reportTexts.forEach(function(rt) {{
    html += '<div class="report-block"><h4>' + rt.title + ' <button class="copy-btn" onclick="copyReport(this,\\'' + rt.text.replace(/'/g, "\\'").replace(/\\n/g, '\\\\n') + '\\')">复制</button></h4><pre>' + rt.text + '</pre></div>';
  }});
  html += '</div>';

  document.getElementById('tab5Content').innerHTML = html;
}})();

// ===== Copy function =====
function copyReport(btn, text) {{
  navigator.clipboard.writeText(text.replace(/\\\\n/g, '\\n')).then(function() {{
    btn.textContent = '已复制';
    setTimeout(function() {{ btn.textContent = '复制'; }}, 1500);
  }});
}}

// ===== INIT =====
setTimeout(function() {{ initCharts1(); window._charts1 = true; }}, 100);
</script>
</body>
</html>'''

# Write output
print(f"Writing HTML to {OUTPUT_FILE}...")
with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    f.write(html)

import os
size_kb = os.path.getsize(OUTPUT_FILE) / 1024
print(f"Done! File: {OUTPUT_FILE} ({size_kb:.0f} KB)")
