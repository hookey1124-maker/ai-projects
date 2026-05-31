"""
生成5.14-5.20周期新品周报数据汇总（含PLP ACOAS + 品效分析 + 三部占比）
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict, Counter
from datetime import date, datetime

# ===== 配置 =====
import os
WORKDIR = os.path.dirname(os.path.abspath(__file__)) + '/'
SOURCE_FILE = WORKDIR + '周报/新品检查周源数据和PLP数据.xlsx'
OUTPUT_FILE = WORKDIR + '新品周报数据表_5.14-5.20.xlsx'

# 列索引（基于四三数据累计 sheet, 0-based）— 适配5.14-5.20周期
C = {
    'sale_no': 0, 'sku': 1, 'list_date': 2, 'first_order': 3,
    'analyst': 4, 'category': 5, 'expand_type': 6,
    # 销量: C8-C18 (3.5-3.11 ~ 5.14-5.20)
    'sales_curr': 17, 'sales_prev': 16,      # 5.14-5.20, 5.7-5.13
    'sales_jan': 7, 'sales_feb': 8, 'sales_mar': 9,  # 3月各周
    # 销售额: C20-C30
    'rev_curr': 29, 'rev_prev': 28,          # 5.14-5.20, 5.7-5.13
    # 对手销量: C32-C42
    'rival_curr': 41, 'rival_prev': 40,      # 5.20, 5.13
    # 市占比: C43-C53
    'share_curr': 52, 'share_prev': 51,      # 5.20, 5.13
    # 8日出单: C65-C75
    'ord8_curr': 74, 'ord8_prev': 73,        # 5.20, 5.13
    # 7日频次: C76-C86
    'freq7_curr': 85, 'freq7_prev': 84,      # 5.20, 5.13
    # 上架8日新品频次: C87-C97
    'nfreq7_curr': 96, 'nfreq7_prev': 95,    # 5.20, 5.13
    # 市场状态: C99-C109
    'mkt_curr': 108, 'mkt_prev': 107,        # 5.20, 5.13
    # 操作判断: C110-C120
    'op_curr': 119,                          # 5.14-5.20
    # PLP: C121-C126
    'plp_curr': 125,                         # 5.11-5.17
    # PLG: C127-C131
    'plg_curr': 130,                         # 5.11-5.17
}

# PLP明细列索引 (0-based)
PC = {
    'period': 0, 'campaign': 1, 'sku': 2, 'id': 3, 'store': 4,
    'plp_start': 5, 'list_date': 6, 'first_order': 7, 'analyst': 8,
    'category': 9, 'expand_type': 10,
    'impr': 11, 'click': 12, 'sold': 13, 'cost': 14, 'ad_rev': 15,
    'total_rev': 16, 'roas': 17, 'cvr': 18, 'ctr': 19,
    'cpc': 20, 'cpa': 21, 'acos': 22, 'acoas': 23, 'plg_enabled': 24,
}

# ===== 辅助函数 =====
def safe_float(v, default=0):
    try: return float(v) if v else default
    except: return default

def get_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

def num(v, default=0):
    return safe_float(v, default)

def pct_str(a, b):
    if not b: return '0%'
    return f"{round(a/b*100, 1)}%"

def ratio_str(a, b):
    if not b: return '-'
    return f"{round((a-b)/abs(b)*100, 1)}%"

def calc_rate(a, b):
    if not b: return 0
    return round(a/b, 4)

def hdr_style(ws, r, c, val, bg='4472C4', fc='FFFFFF', bold=True, align='center'):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.font = Font(bold=bold, color=fc, name='微软雅黑', size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    return cell

def data_cell(ws, r, c, val, bg='FFFFFF', bold=False, align='center', fc='000000'):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.font = Font(bold=bold, color=fc, name='微软雅黑', size=9)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    return cell

def set_border(ws, start_r, end_r, start_c, end_c):
    thin = Side(style='thin')
    for r in range(start_r, end_r+1):
        for c in range(start_c, end_c+1):
            ws.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ===== 常量 =====
ANALYSTS = ['俞东旭', '张潇', '朱培源', '王偲涵', '章鹏', '胡煜星']
CATEGORIES = ['车门系统', '车身外扩件', '挡泥板', '机盖及附件', '其他', '饰条', '牌照板支架', '未分类']
EXPAND_TYPES = ['原开品', '拓展品', '组合件']

COLOR_HEADER = '4472C4'
COLOR_ODD = 'EEF2FF'
COLOR_EVEN = 'FFFFFF'
COLOR_TOTAL = 'E8F0FE'

# ===== 读取源数据 =====
print("读取源数据...")
wb_src = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
ws_main = wb_src['四三数据累计']
ws_plp = wb_src['PLP明细']
ws_dept = wb_src['周销售数据']  # 三部总销量/销售额明细

rows_raw = []
for row in ws_main.iter_rows(min_row=2, values_only=True):
    if row[C['sku']]:
        rows_raw.append(list(row))

cutoff_curr = date(2026, 5, 20)
cutoff_prev = date(2026, 5, 13)
rows_curr = [r for r in rows_raw if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) <= cutoff_curr]
rows_prev = [r for r in rows_raw if get_date(r[C['list_date']]) and get_date(r[C['list_date']]) <= cutoff_prev]

print(f"当前周期SKU: {len(rows_curr)}, 上周期SKU: {len(rows_prev)}")

# ===== 读取三部总销量 =====
dept_total_sales = 0
dept_total_revenue = 0.0
for row in ws_dept.iter_rows(min_row=2, max_row=2, values_only=True):
    dept_total_sales = int(num(row[4])) if row[4] else 0
    dept_total_revenue = num(row[5])
print(f"三部总销量: {dept_total_sales}, 三部总销售额: ${dept_total_revenue:,.2f}")

def get_cat(r_data):
    c = str(r_data[C['category']] or '').strip()
    return c if c else '未分类'

def get_an(r_data):
    a = str(r_data[C['analyst']] or '').strip()
    return a if a else '未知'

def get_exp(r_data):
    e = str(r_data[C['expand_type']] or '').strip()
    return e if e else '其他'

# SKU映射
sku_rev_curr = {}
sku_info = {}
for r in rows_curr:
    sku = str(r[C['sku']]).strip()
    if sku:
        sku_rev_curr[sku] = num(r[C['rev_curr']])
        sku_info[sku] = r

# ===== 品效分析：按上架月份聚合 =====
print("\n构建品效数据...")
# 周 → 月映射
WEEK_SALES_COLS = [7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17]  # C8-C18 (0-based)
WEEK_REV_COLS  = [19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29]  # C20-C30 (0-based)

# 每个周的结束日期（用于判断属于哪个月）
WEEK_PERIODS = [
    ('3.5-3.11', 3), ('3.12-3.18', 3), ('3.19-3.25', 3), ('3.26-4.1', 3),
    ('4.2-4.8', 4), ('4.9-4.15', 4), ('4.16-4.22', 4), ('4.23-4.29', 4),
    ('4.30-5.6', 5), ('5.7-5.13', 5), ('5.14-5.20', 5),
]

# 按上架月份分组
cohort_data = defaultdict(lambda: {
    'sku_count': 0, 'skus': set(),
    'monthly_sales': defaultdict(float),   # {month_num: total_sales}
    'monthly_revenue': defaultdict(float), # {month_num: total_revenue}
    'monthly_sku_sold': defaultdict(set),  # {month_num: set(skus_that_sold)}
})

for r_data in rows_curr:
    list_d = get_date(r_data[C['list_date']])
    if not list_d:
        continue
    launch_month = list_d.month
    sku = str(r_data[C['sku']]).strip()

    cohort = cohort_data[launch_month]
    cohort['sku_count'] += 1
    cohort['skus'].add(sku)

    for wi, (period, month_num) in enumerate(WEEK_PERIODS):
        sales = num(r_data[WEEK_SALES_COLS[wi]])
        rev = num(r_data[WEEK_REV_COLS[wi]])
        if sales > 0:
            cohort['monthly_sales'][month_num] += sales
            cohort['monthly_sku_sold'][month_num].add(sku)
        if rev > 0:
            cohort['monthly_revenue'][month_num] += rev

# 排序
launch_months = sorted(cohort_data.keys())
MONTH_LABELS = {1: '1月', 2: '2月', 3: '3月', 4: '4月', 5: '5月'}

print("品效Cohort数据:")
for m in launch_months:
    d = cohort_data[m]
    print(f"  {m}月上架: {d['sku_count']} SKU, 月销售额: {dict(d['monthly_revenue'])}")

# ===== 读取PLP数据 =====
print("\n读取PLP数据...")
PLP_CURR = '5.11-5.17'
PLP_PREV = '5.4-5.10'

# 先扫描PLP明细中实际有哪些period
plp_periods_in_data = set()
for row in ws_plp.iter_rows(min_row=2, values_only=True):
    p = str(row[PC['period']] or '').strip()
    if p:
        plp_periods_in_data.add(p)
print(f"PLP明细中的周期: {sorted(plp_periods_in_data)}")

plp_curr_rows = []
plp_prev_rows = []
plp_plg_enabled = set()

for row in ws_plp.iter_rows(min_row=2, values_only=True):
    period = str(row[PC['period']] or '').strip()
    sku = str(row[PC['sku']] or '').strip()
    if not sku or sku.startswith('广告') or sku.startswith('总数据'):
        continue
    list_d = get_date(row[PC['list_date']])
    if period == PLP_CURR and list_d and list_d <= cutoff_curr:
        plp_curr_rows.append(list(row))
        if str(row[PC['plg_enabled']] or '').strip() == 'Y':
            plp_plg_enabled.add(sku)
    elif period == PLP_PREV and list_d and list_d <= cutoff_prev:
        plp_prev_rows.append(list(row))

print(f"PLP {PLP_CURR}: {len(plp_curr_rows)} rows, PLP {PLP_PREV}: {len(plp_prev_rows)} rows")

# 计算每个PLG enabled SKU的campaign数量（用于区分单链接/多链接PLP+PLG同开）
sku_campaign_count = Counter()
for row in plp_curr_rows:
    sku = str(row[PC['sku']] or '').strip()
    if sku in plp_plg_enabled:
        sku_campaign_count[sku] += 1
print(f"PLG enabled SKU: {len(plp_plg_enabled)}, 单链接: {sum(1 for c in sku_campaign_count.values() if c == 1)}")

def agg_plp_rows(rows):
    by_sku = defaultdict(lambda: {'impr': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'total_rev': 0, 'campaigns': set()})
    for r in rows:
        sku = str(r[PC['sku']] or '').strip()
        camp = str(r[PC['campaign']] or '').strip()
        by_sku[sku]['impr'] += num(r[PC['impr']])
        by_sku[sku]['click'] += num(r[PC['click']])
        by_sku[sku]['sold'] += num(r[PC['sold']])
        by_sku[sku]['cost'] += num(r[PC['cost']])
        by_sku[sku]['ad_rev'] += num(r[PC['ad_rev']])
        by_sku[sku]['total_rev'] += num(r[PC['total_rev']])
        if camp: by_sku[sku]['campaigns'].add(camp)
    return by_sku

plp_curr_sku = agg_plp_rows(plp_curr_rows)
plp_prev_sku = agg_plp_rows(plp_prev_rows)

def plp_totals(by_sku):
    all_camps = set()
    t = {'imp': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'total_rev': 0}
    for sku, d in by_sku.items():
        t['imp'] += d['impr']; t['click'] += d['click']; t['sold'] += d['sold']
        t['cost'] += d['cost']; t['ad_rev'] += d['ad_rev']
        t['total_rev'] += sku_rev_curr.get(sku, 0)
        all_camps.update(d['campaigns'])
    t['campaigns'] = len(all_camps)
    t['links'] = len(by_sku)
    t['roas'] = calc_rate(t['ad_rev'], t['cost'])
    t['cvr'] = calc_rate(t['sold'], t['click'])
    t['ctr'] = calc_rate(t['click'], t['imp'])
    t['cpc'] = round(t['cost'] / t['click'], 2) if t['click'] else 0
    t['cpa'] = round(t['cost'] / t['sold'], 2) if t['sold'] else 0
    t['acos'] = calc_rate(t['cost'], t['ad_rev'])
    t['acoas'] = calc_rate(t['cost'], t['total_rev'])
    return t

plp_t = plp_totals(plp_curr_sku)
plp_pt = plp_totals(plp_prev_sku)

def plp_dim(by_sku, key_fn):
    groups = defaultdict(lambda: {'imp': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0, 'total_rev': 0, 'campaigns': set(), 'skus': set()})
    for sku, d in by_sku.items():
        info = sku_info.get(sku)
        k = key_fn(sku, info) if info else key_fn(sku, None)
        g = groups[k]
        g['imp'] += d['impr']; g['click'] += d['click']; g['sold'] += d['sold']
        g['cost'] += d['cost']; g['ad_rev'] += d['ad_rev']
        g['total_rev'] += sku_rev_curr.get(sku, 0)
        g['campaigns'].update(d['campaigns'])
        g['skus'].add(sku)
    result = {}
    for k, g in groups.items():
        result[k] = {
            'campaigns': len(g['campaigns']), 'links': len(g['skus']),
            'imp': int(g['imp']), 'click': int(g['click']), 'sold': int(g['sold']),
            'cost': round(g['cost'], 2), 'ad_rev': round(g['ad_rev'], 2),
            'total_rev': round(g['total_rev'], 2),
            'roas': calc_rate(g['ad_rev'], g['cost']),
            'cvr': calc_rate(g['sold'], g['click']),
            'ctr': calc_rate(g['click'], g['imp']),
            'cpc': round(g['cost'] / g['click'], 2) if g['click'] else 0,
            'cpa': round(g['cost'] / g['sold'], 2) if g['sold'] else 0,
            'acos': calc_rate(g['cost'], g['ad_rev']),
            'acoas': calc_rate(g['cost'], g['total_rev']),
        }
    return result

plp_by_ana = plp_dim(plp_curr_sku, lambda sku, info: get_an(info) if info else '未知')
plp_by_cat = plp_dim(plp_curr_sku, lambda sku, info: get_cat(info) if info else '未分类')
plp_by_exp = plp_dim(plp_curr_sku, lambda sku, info: get_exp(info) if info else '其他')

# ===== 创建新Workbook =====
wb = openpyxl.Workbook()

# ===== Sheet 1: 总体数据 =====
ws1 = wb.active
ws1.title = '总体数据'
ws1.sheet_properties.tabColor = '4472C4'

r = 1
hdr_style(ws1, r, 1, '新品周报汇总 - 5.14-5.20（含PA+品效）', COLOR_HEADER, 'FFFFFF', True, 'left')
ws1.merge_cells(f'A{r}:H{r}')
ws1.row_dimensions[r].height = 28

r += 2
hdr_style(ws1, r, 1, '一、总体概况', COLOR_HEADER, 'FFFFFF', True)
ws1.merge_cells(f'A{r}:D{r}')

r += 1
for ci, h in enumerate(['指标', '5.14-5.20', '5.7-5.13', '环比'], 1):
    hdr_style(ws1, r, ci, h)

total_sku = len(rows_curr)
new_list = sum(1 for r_data in rows_curr if get_date(r_data[C['list_date']]) and get_date(r_data[C['list_date']]) > cutoff_prev)
total_sales_curr = sum(num(r_data[C['sales_curr']]) for r_data in rows_curr)
total_sales_prev = sum(num(r_data[C['sales_prev']]) for r_data in rows_prev)
total_rev_curr = sum(num(r_data[C['rev_curr']]) for r_data in rows_curr)
total_rev_prev = sum(num(r_data[C['rev_prev']]) for r_data in rows_prev)
has_rival_curr = sum(1 for r_data in rows_curr if num(r_data[C['rival_curr']]) > 0)
has_rival_prev = sum(1 for r_data in rows_prev if num(r_data[C['rival_prev']]) > 0)

# 新品销售额占三部总销售额比
dept_ratio = total_rev_curr / dept_total_revenue * 100 if dept_total_revenue > 0 else 0
dept_ratio_str = f"{dept_ratio:.1f}%"

rows_1_data = [
    ('累计SKU数', total_sku, len(rows_prev), ratio_str(total_sku, len(rows_prev))),
    ('本周新上架SKU', new_list, '-', '-'),
    ('总销量', total_sales_curr, total_sales_prev, ratio_str(total_sales_curr, total_sales_prev)),
    ('总销售额(USD)', total_rev_curr, total_rev_prev, ratio_str(total_rev_curr, total_rev_prev)),
    ('新品销售占比', dept_ratio_str, '-', '-'),
    ('三部总销售额(USD)', dept_total_revenue, '-', '-'),
    ('有对手SKU数', has_rival_curr, has_rival_prev, '-'),
    ('无对手SKU数', total_sku - has_rival_curr, len(rows_prev) - has_rival_prev, '-'),
]

for i, (name, v1, v2, v3) in enumerate(rows_1_data):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    data_cell(ws1, r, 1, name, bg, align='left')
    data_cell(ws1, r, 2, v1, bg)
    data_cell(ws1, r, 3, v2, bg)
    data_cell(ws1, r, 4, v3, bg)

# 分析及时率
r += 2
hdr_style(ws1, r, 1, '二、分析及时率', COLOR_HEADER, 'FFFFFF', True)
ws1.merge_cells(f'A{r}:D{r}')

r += 1
for ci, h in enumerate(['指标', '5.20（本周）', '5.13（上周）', '变化'], 1):
    hdr_style(ws1, r, ci, h)

def calc_timeliness(rows, col_freq, col_nfreq):
    timely = no_8d = no_7d = 0
    for r_data in rows:
        freq = str(r_data[col_freq] or '').strip()
        nfreq = str(r_data[col_nfreq] or '').strip()
        if nfreq == '异常': no_8d += 1
        elif freq == '异常': no_7d += 1
        else: timely += 1
    return timely, no_8d, no_7d

timely_curr, no_8d_curr, no_7d_curr = calc_timeliness(rows_curr, C['freq7_curr'], C['nfreq7_curr'])
timely_prev, no_8d_prev, no_7d_prev = calc_timeliness(rows_prev, C['freq7_prev'], C['nfreq7_prev'])
total_timeliness = timely_curr + no_8d_curr + no_7d_curr
rate_curr = f"{round(timely_curr/total_timeliness*100, 1)}%" if total_timeliness else '0%'

timeliness_data = [
    ('及时分析产品数', timely_curr, timely_prev, timely_curr - timely_prev),
    ('8日内新品无分析', no_8d_curr, no_8d_prev, no_8d_curr - no_8d_prev),
    ('超7日低占比未分析', no_7d_curr, no_7d_prev, no_7d_curr - no_7d_prev),
    ('统计总数', total_timeliness, timely_prev + no_8d_prev + no_7d_prev, '-'),
    ('及时分析率', rate_curr, '-', '-'),
]

for i, (name, v1, v2, v3) in enumerate(timeliness_data):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    data_cell(ws1, r, 1, name, bg, align='left')
    data_cell(ws1, r, 2, v1, bg)
    data_cell(ws1, r, 3, v2, bg)
    data_cell(ws1, r, 4, v3, bg)

# 新品出单情况
r += 2
hdr_style(ws1, r, 1, '三、新品出单情况（有对手口径）', COLOR_HEADER, 'FFFFFF', True)
ws1.merge_cells(f'A{r}:D{r}')

r += 1
for ci, h in enumerate(['指标', '5.20（本周）', '5.13（上周）', '变化'], 1):
    hdr_style(ws1, r, ci, h)

def count_ord8(rows, col):
    y = n = no = 0
    for r_data in rows:
        v = str(r_data[col] or '').strip()
        if v == 'Y': y += 1
        elif v == 'N': n += 1
        elif v == '未出单': no += 1
    return y, n, no

y_curr, n_curr, no_curr = count_ord8([r_data for r_data in rows_curr if num(r_data[C['rival_curr']]) > 0], C['ord8_curr'])
y_prev, n_prev, no_prev = count_ord8([r_data for r_data in rows_prev if num(r_data[C['rival_prev']]) > 0], C['ord8_prev'])
sale_rate_curr = f"{round((y_curr+n_curr)/has_rival_curr*100, 1)}%" if has_rival_curr else '0%'

ord8_data = [
    ('有对手总SKU', has_rival_curr, has_rival_prev, has_rival_curr - has_rival_prev),
    ('8日内出单（Y）', y_curr, y_prev, y_curr - y_prev),
    ('8日外出单（N）', n_curr, n_prev, n_curr - n_prev),
    ('真正未出单', no_curr, no_prev, no_curr - no_prev),
    ('已出单合计(Y+N)', y_curr+n_curr, y_prev+n_prev, (y_curr+n_curr)-(y_prev+n_prev)),
    ('出单率', sale_rate_curr, '-', '-'),
]

for i, (name, v1, v2, v3) in enumerate(ord8_data):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    data_cell(ws1, r, 1, name, bg, align='left')
    data_cell(ws1, r, 2, v1, bg)
    data_cell(ws1, r, 3, v2, bg)
    data_cell(ws1, r, 4, v3, bg)

set_border(ws1, 2, r, 1, 4)
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 12

# ===== Sheet 2: 品线维度 =====
ws2 = wb.create_sheet('品线维度')
ws2.sheet_properties.tabColor = '70AD47'

r = 1
hdr_style(ws2, r, 1, '品线维度 - 5.14-5.20', COLOR_HEADER, 'FFFFFF', True, 'left')
ws2.merge_cells(f'A{r}:J{r}')
ws2.row_dimensions[r].height = 28

r += 1
cols = ['品线', 'SKU数', '本周新上架', '本周销量', '上周销量', '销量环比', '本周销售额', '上周销售额', '销售额环比', '有对手SKU']
for ci, h in enumerate(cols, 1):
    hdr_style(ws2, r, ci, h)

cat_data_curr = defaultdict(lambda: {'sku': 0, 'new': 0, 'sales': 0, 'rev': 0, 'has_rival': 0})
cat_data_prev = defaultdict(lambda: {'sku': 0, 'sales': 0, 'rev': 0, 'has_rival': 0})

for r_data in rows_curr:
    cat = get_cat(r_data)
    list_d = get_date(r_data[C['list_date']])
    cat_data_curr[cat]['sku'] += 1
    if list_d and list_d > cutoff_prev: cat_data_curr[cat]['new'] += 1
    cat_data_curr[cat]['sales'] += num(r_data[C['sales_curr']])
    cat_data_curr[cat]['rev'] += num(r_data[C['rev_curr']])
    if num(r_data[C['rival_curr']]) > 0: cat_data_curr[cat]['has_rival'] += 1

for r_data in rows_prev:
    cat = get_cat(r_data)
    cat_data_prev[cat]['sku'] += 1
    cat_data_prev[cat]['sales'] += num(r_data[C['sales_prev']])
    cat_data_prev[cat]['rev'] += num(r_data[C['rev_prev']])
    if num(r_data[C['rival_prev']]) > 0: cat_data_prev[cat]['has_rival'] += 1

all_cats_in_data = sorted(set(list(cat_data_curr.keys()) + list(cat_data_prev.keys())))
cats_ordered = [c for c in CATEGORIES if c in all_cats_in_data] + [c for c in all_cats_in_data if c not in CATEGORIES]

for i, cat in enumerate(cats_ordered):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d_curr = cat_data_curr.get(cat, {})
    d_prev = cat_data_prev.get(cat, {})
    data_cell(ws2, r, 1, cat, bg, align='left')
    data_cell(ws2, r, 2, d_curr.get('sku', 0), bg)
    data_cell(ws2, r, 3, d_curr.get('new', 0), bg)
    data_cell(ws2, r, 4, d_curr.get('sales', 0), bg)
    data_cell(ws2, r, 5, d_prev.get('sales', 0), bg)
    data_cell(ws2, r, 6, ratio_str(d_curr.get('sales', 0), d_prev.get('sales', 0)), bg)
    data_cell(ws2, r, 7, d_curr.get('rev', 0), bg)
    data_cell(ws2, r, 8, d_prev.get('rev', 0), bg)
    data_cell(ws2, r, 9, ratio_str(d_curr.get('rev', 0), d_prev.get('rev', 0)), bg)
    data_cell(ws2, r, 10, d_curr.get('has_rival', 0), bg)

r += 1
data_cell(ws2, r, 1, '合计', COLOR_TOTAL, bold=True, align='left')
data_cell(ws2, r, 2, total_sku, COLOR_TOTAL, bold=True)
data_cell(ws2, r, 3, new_list, COLOR_TOTAL, bold=True)
data_cell(ws2, r, 4, total_sales_curr, COLOR_TOTAL, bold=True)
data_cell(ws2, r, 5, total_sales_prev, COLOR_TOTAL, bold=True)
data_cell(ws2, r, 6, ratio_str(total_sales_curr, total_sales_prev), COLOR_TOTAL, bold=True)
data_cell(ws2, r, 7, total_rev_curr, COLOR_TOTAL, bold=True)
data_cell(ws2, r, 8, total_rev_prev, COLOR_TOTAL, bold=True)
data_cell(ws2, r, 9, ratio_str(total_rev_curr, total_rev_prev), COLOR_TOTAL, bold=True)
data_cell(ws2, r, 10, has_rival_curr, COLOR_TOTAL, bold=True)

set_border(ws2, 2, r, 1, 10)
for ci, w in enumerate([12, 8, 10, 10, 10, 10, 12, 12, 10, 10], 1):
    ws2.column_dimensions[chr(64+ci)].width = w

# ===== Sheet 3: 分析人维度 =====
ws3 = wb.create_sheet('分析人维度')
ws3.sheet_properties.tabColor = 'ED7D31'

r = 1
hdr_style(ws3, r, 1, '分析人维度 - 5.14-5.20', COLOR_HEADER, 'FFFFFF', True, 'left')
ws3.merge_cells(f'A{r}:J{r}')

r += 1
for ci, h in enumerate(cols, 1):
    hdr_style(ws3, r, ci, h)

an_data_curr = defaultdict(lambda: {'sku': 0, 'new': 0, 'sales': 0, 'rev': 0, 'has_rival': 0})
an_data_prev = defaultdict(lambda: {'sku': 0, 'sales': 0, 'rev': 0, 'has_rival': 0})

for r_data in rows_curr:
    an = get_an(r_data)
    list_d = get_date(r_data[C['list_date']])
    an_data_curr[an]['sku'] += 1
    if list_d and list_d > cutoff_prev: an_data_curr[an]['new'] += 1
    an_data_curr[an]['sales'] += num(r_data[C['sales_curr']])
    an_data_curr[an]['rev'] += num(r_data[C['rev_curr']])
    if num(r_data[C['rival_curr']]) > 0: an_data_curr[an]['has_rival'] += 1

for r_data in rows_prev:
    an = get_an(r_data)
    an_data_prev[an]['sku'] += 1
    an_data_prev[an]['sales'] += num(r_data[C['sales_prev']])
    an_data_prev[an]['rev'] += num(r_data[C['rev_prev']])
    if num(r_data[C['rival_prev']]) > 0: an_data_prev[an]['has_rival'] += 1

all_ans = sorted(set(list(an_data_curr.keys()) + list(an_data_prev.keys())))
ans_ordered = [a for a in ANALYSTS if a in all_ans] + [a for a in all_ans if a not in ANALYSTS]

for i, an in enumerate(ans_ordered):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d_curr = an_data_curr.get(an, {})
    d_prev = an_data_prev.get(an, {})
    data_cell(ws3, r, 1, an, bg, align='left')
    data_cell(ws3, r, 2, d_curr.get('sku', 0), bg)
    data_cell(ws3, r, 3, d_curr.get('new', 0), bg)
    data_cell(ws3, r, 4, d_curr.get('sales', 0), bg)
    data_cell(ws3, r, 5, d_prev.get('sales', 0), bg)
    data_cell(ws3, r, 6, ratio_str(d_curr.get('sales', 0), d_prev.get('sales', 0)), bg)
    data_cell(ws3, r, 7, d_curr.get('rev', 0), bg)
    data_cell(ws3, r, 8, d_prev.get('rev', 0), bg)
    data_cell(ws3, r, 9, ratio_str(d_curr.get('rev', 0), d_prev.get('rev', 0)), bg)
    data_cell(ws3, r, 10, d_curr.get('has_rival', 0), bg)

r += 1
data_cell(ws3, r, 1, '合计', COLOR_TOTAL, bold=True, align='left')
data_cell(ws3, r, 2, total_sku, COLOR_TOTAL, bold=True)
data_cell(ws3, r, 3, new_list, COLOR_TOTAL, bold=True)
data_cell(ws3, r, 4, total_sales_curr, COLOR_TOTAL, bold=True)
data_cell(ws3, r, 5, total_sales_prev, COLOR_TOTAL, bold=True)
data_cell(ws3, r, 6, ratio_str(total_sales_curr, total_sales_prev), COLOR_TOTAL, bold=True)
data_cell(ws3, r, 7, total_rev_curr, COLOR_TOTAL, bold=True)
data_cell(ws3, r, 8, total_rev_prev, COLOR_TOTAL, bold=True)
data_cell(ws3, r, 9, ratio_str(total_rev_curr, total_rev_prev), COLOR_TOTAL, bold=True)
data_cell(ws3, r, 10, has_rival_curr, COLOR_TOTAL, bold=True)

set_border(ws3, 2, r, 1, 10)
for ci, w in enumerate([12, 8, 10, 10, 10, 10, 12, 12, 10, 10], 1):
    ws3.column_dimensions[chr(64+ci)].width = w

# ===== Sheet 4: 拓展类型 =====
ws4 = wb.create_sheet('拓展类型')
ws4.sheet_properties.tabColor = '9E67AB'

r = 1
hdr_style(ws4, r, 1, '拓展类型 - 5.14-5.20', COLOR_HEADER, 'FFFFFF', True, 'left')
ws4.merge_cells(f'A{r}:K{r}')
ws4.row_dimensions[r].height = 28

r += 1
cols4 = ['拓展类型', 'SKU数', '本周新上架', '本周销量', '上周销量', '销量环比', '本周销售额', '上周销售额', '有对手SKU', '有对手出单率', '出单率环比']
for ci, h in enumerate(cols4, 1):
    hdr_style(ws4, r, ci, h)

exp_data_curr = defaultdict(lambda: {'sku': 0, 'new': 0, 'sales': 0, 'rev': 0, 'has_rival': 0, 'y': 0, 'n': 0})
exp_data_prev = defaultdict(lambda: {'sku': 0, 'sales': 0, 'rev': 0, 'has_rival': 0, 'y': 0, 'n': 0})

for r_data in rows_curr:
    exp = get_exp(r_data)
    list_d = get_date(r_data[C['list_date']])
    exp_data_curr[exp]['sku'] += 1
    if list_d and list_d > cutoff_prev: exp_data_curr[exp]['new'] += 1
    exp_data_curr[exp]['sales'] += num(r_data[C['sales_curr']])
    exp_data_curr[exp]['rev'] += num(r_data[C['rev_curr']])
    if num(r_data[C['rival_curr']]) > 0:
        exp_data_curr[exp]['has_rival'] += 1
        v = str(r_data[C['ord8_curr']] or '').strip()
        if v == 'Y': exp_data_curr[exp]['y'] += 1
        elif v == 'N': exp_data_curr[exp]['n'] += 1

for r_data in rows_prev:
    exp = get_exp(r_data)
    exp_data_prev[exp]['sku'] += 1
    exp_data_prev[exp]['sales'] += num(r_data[C['sales_prev']])
    exp_data_prev[exp]['rev'] += num(r_data[C['rev_prev']])
    if num(r_data[C['rival_prev']]) > 0:
        exp_data_prev[exp]['has_rival'] += 1
        v = str(r_data[C['ord8_prev']] or '').strip()
        if v == 'Y': exp_data_prev[exp]['y'] += 1
        elif v == 'N': exp_data_prev[exp]['n'] += 1

all_exps = sorted(set(list(exp_data_curr.keys()) + list(exp_data_prev.keys())))
exps_ordered = [e for e in EXPAND_TYPES if e in all_exps] + [e for e in all_exps if e not in EXPAND_TYPES]

for i, exp in enumerate(exps_ordered):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d_curr = exp_data_curr.get(exp, {})
    d_prev = exp_data_prev.get(exp, {})
    rate_c_val = (d_curr.get('y',0)+d_curr.get('n',0))/d_curr.get('has_rival',1)*100 if d_curr.get('has_rival') else 0
    rate_c = f"{round(rate_c_val, 1)}%"
    rate_p = f"{round((d_prev.get('y',0)+d_prev.get('n',0))/d_prev.get('has_rival',1)*100, 1)}%" if d_prev.get('has_rival') else '0%'
    data_cell(ws4, r, 1, exp, bg, align='left')
    data_cell(ws4, r, 2, d_curr.get('sku', 0), bg)
    data_cell(ws4, r, 3, d_curr.get('new', 0), bg)
    data_cell(ws4, r, 4, d_curr.get('sales', 0), bg)
    data_cell(ws4, r, 5, d_prev.get('sales', 0), bg)
    data_cell(ws4, r, 6, ratio_str(d_curr.get('sales', 0), d_prev.get('sales', 0)), bg)
    data_cell(ws4, r, 7, d_curr.get('rev', 0), bg)
    data_cell(ws4, r, 8, d_prev.get('rev', 0), bg)
    data_cell(ws4, r, 9, d_curr.get('has_rival', 0), bg)
    data_cell(ws4, r, 10, rate_c, bg)
    data_cell(ws4, r, 11, ratio_str(rate_c_val, float(rate_p.replace('%','')) if rate_p != '0%' else 0), bg)

r += 1
data_cell(ws4, r, 1, '合计', COLOR_TOTAL, bold=True, align='left')
data_cell(ws4, r, 2, total_sku, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 3, new_list, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 4, total_sales_curr, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 5, total_sales_prev, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 6, ratio_str(total_sales_curr, total_sales_prev), COLOR_TOTAL, bold=True)
data_cell(ws4, r, 7, total_rev_curr, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 8, total_rev_prev, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 9, has_rival_curr, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 10, sale_rate_curr, COLOR_TOTAL, bold=True)
data_cell(ws4, r, 11, '-', COLOR_TOTAL, bold=True)

set_border(ws4, 2, r, 1, 11)
for ci, w in enumerate([12, 8, 10, 10, 10, 10, 12, 12, 10, 10, 10], 1):
    ws4.column_dimensions[chr(64+ci)].width = w

# ===== Sheet 5: 分析及时率 =====
ws5 = wb.create_sheet('分析及时率')
ws5.sheet_properties.tabColor = 'FFC000'

r = 1
hdr_style(ws5, r, 1, '分析及时率 - 5.14-5.20', COLOR_HEADER, 'FFFFFF', True, 'left')
ws5.merge_cells(f'A{r}:F{r}')

r += 1
cols5 = ['分析人', '及时分析产品数', '8日内新品无分析', '超7日低占比未分析', '统计总数', '及时分析率']
for ci, h in enumerate(cols5, 1):
    hdr_style(ws5, r, ci, h)

an_time_curr = defaultdict(lambda: {'timely': 0, 'no_8d': 0, 'no_7d': 0})
for r_data in rows_curr:
    an = get_an(r_data)
    freq = str(r_data[C['freq7_curr']] or '').strip()
    nfreq = str(r_data[C['nfreq7_curr']] or '').strip()
    if nfreq == '异常': an_time_curr[an]['no_8d'] += 1
    elif freq == '异常': an_time_curr[an]['no_7d'] += 1
    else: an_time_curr[an]['timely'] += 1

for i, an in enumerate(ans_ordered):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = an_time_curr.get(an, {})
    total = d.get('timely', 0) + d.get('no_8d', 0) + d.get('no_7d', 0)
    rate = f"{round(d.get('timely', 0)/total*100, 1)}%" if total else '0%'
    data_cell(ws5, r, 1, an, bg, align='left')
    data_cell(ws5, r, 2, d.get('timely', 0), bg)
    data_cell(ws5, r, 3, d.get('no_8d', 0), bg)
    data_cell(ws5, r, 4, d.get('no_7d', 0), bg)
    data_cell(ws5, r, 5, total, bg)
    data_cell(ws5, r, 6, rate, bg)

r += 1
data_cell(ws5, r, 1, '合计', COLOR_TOTAL, bold=True, align='left')
data_cell(ws5, r, 2, timely_curr, COLOR_TOTAL, bold=True)
data_cell(ws5, r, 3, no_8d_curr, COLOR_TOTAL, bold=True)
data_cell(ws5, r, 4, no_7d_curr, COLOR_TOTAL, bold=True)
data_cell(ws5, r, 5, total_timeliness, COLOR_TOTAL, bold=True)
data_cell(ws5, r, 6, rate_curr, COLOR_TOTAL, bold=True)

set_border(ws5, 2, r, 1, 6)
for ci, w in enumerate([12, 15, 15, 18, 10, 12], 1):
    ws5.column_dimensions[chr(64+ci)].width = w

# ===== Sheet 6: 低占比新品 =====
ws6 = wb.create_sheet('低占比新品')
ws6.sheet_properties.tabColor = 'C55A11'

r = 1
hdr_style(ws6, r, 1, '低占比新品明细 - 5.14-5.20（含ACOAS）', COLOR_HEADER, 'FFFFFF', True, 'left')
ws6.merge_cells(f'A{r}:Z{r}')

r += 1
cols6 = ['销售编号', 'SKU', '上架日期', '分析人', '品类', '拓展类型',
          '本周销量', '销量环比', '本周销售额', '销售额环比',
          '上期末对手销量', '本期末对手销量', '对手销量环比',
          '上期末市占比', '本期末市占比', '市占比环比',
          '8日出单情况', '7日频次标签', '7日新品频次标签',
          '上期市场状态', '本期操作判断', '本期市场状态', '开启PLP', 'PLG最高费率', 'ACOAS']
for ci, h in enumerate(cols6, 1):
    hdr_style(ws6, r, ci, h)

low_share_skus = []
for r_data in rows_curr:
    share = num(r_data[C['share_curr']]) * 100
    rival = num(r_data[C['rival_curr']])
    if share < 75 and rival > 0:
        low_share_skus.append(r_data)
low_share_skus.sort(key=lambda x: num(x[C['share_curr']]))

for i, r_data in enumerate(low_share_skus):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    sku = str(r_data[C['sku']]).strip()
    rev_curr = num(r_data[C['rev_curr']])
    rev_prev = num(r_data[C['rev_prev']])
    share_curr = num(r_data[C['share_curr']]) * 100
    share_prev = num(r_data[C['share_prev']]) * 100
    cost = plp_curr_sku.get(sku, {}).get('cost', 0)
    total_rev = sku_rev_curr.get(sku, 0)
    acoas_val = cost / total_rev if total_rev > 0 else None
    acoas_str = f"{acoas_val:.4f}" if acoas_val is not None else 'N/A'

    data_cell(ws6, r, 1, r_data[C['sale_no']], bg)
    data_cell(ws6, r, 2, sku, bg)
    data_cell(ws6, r, 3, str(get_date(r_data[C['list_date']]))[:10] if get_date(r_data[C['list_date']]) else '', bg)
    data_cell(ws6, r, 4, get_an(r_data), bg)
    data_cell(ws6, r, 5, get_cat(r_data), bg)
    data_cell(ws6, r, 6, get_exp(r_data), bg)
    data_cell(ws6, r, 7, num(r_data[C['sales_curr']]), bg)
    data_cell(ws6, r, 8, ratio_str(num(r_data[C['sales_curr']]), num(r_data[C['sales_prev']])), bg)
    data_cell(ws6, r, 9, rev_curr, bg)
    data_cell(ws6, r, 10, ratio_str(rev_curr, rev_prev), bg)
    data_cell(ws6, r, 11, num(r_data[C['rival_prev']]), bg)
    data_cell(ws6, r, 12, num(r_data[C['rival_curr']]), bg)
    data_cell(ws6, r, 13, ratio_str(num(r_data[C['rival_curr']]), num(r_data[C['rival_prev']])), bg)
    data_cell(ws6, r, 14, share_prev, bg)
    data_cell(ws6, r, 15, share_curr, bg)
    data_cell(ws6, r, 16, ratio_str(share_curr, share_prev), bg)
    data_cell(ws6, r, 17, r_data[C['ord8_curr']] or '', bg)
    data_cell(ws6, r, 18, r_data[C['freq7_curr']] or '', bg)
    data_cell(ws6, r, 19, r_data[C['nfreq7_curr']] or '', bg)
    data_cell(ws6, r, 20, r_data[C['mkt_prev']] or '', bg)    # 上期市场状态
    data_cell(ws6, r, 21, r_data[C['op_curr']] or '', bg)      # 本期操作判断
    data_cell(ws6, r, 22, r_data[C['mkt_curr']] or '', bg)     # 本期市场状态
    data_cell(ws6, r, 23, 'Y' if str(r_data[C['plp_curr']] or '').strip().upper() == 'Y' else 'N', bg)
    data_cell(ws6, r, 24, num(r_data[C['plg_curr']]) * 100, bg)
    data_cell(ws6, r, 25, acoas_str, bg, fc='FF0000' if acoas_val and acoas_val > 0.1 else '000000')

set_border(ws6, 2, r, 1, 25)
for ci in range(1, 26):
    col_letter = chr(64+ci) if ci <= 26 else 'A' + chr(64+ci-26)
    ws6.column_dimensions[col_letter].width = 10

# ===== Sheet 7: 新品PLP =====
ws7 = wb.create_sheet('新品PLP')
ws7.sheet_properties.tabColor = '00B0F0'

r = 1
hdr_style(ws7, r, 1, f'新品PLP - {PLP_CURR}（含ACOAS）', COLOR_HEADER, 'FFFFFF', True, 'left')
ws7.merge_cells(f'A{r}:P{r}')

r += 2
hdr_style(ws7, r, 1, '【总数据】', '00B0F0', 'FFFFFF', True)
ws7.merge_cells(f'A{r}:P{r}')

r += 1
cols7 = ['周期', '广告活动数', '广告链接数', '曝光量', '点击量', '售出数', '广告花费', '广告销售额', '总销售额', 'ROAS', 'CVR', 'CTR', 'CPC', 'CPA', 'ACOS', 'ACOAS']
for ci, h in enumerate(cols7, 1):
    hdr_style(ws7, r, ci, h)

def write_plp_row(ws, r, period, t, bg):
    data_cell(ws, r, 1, period, bg)
    data_cell(ws, r, 2, t['campaigns'], bg)
    data_cell(ws, r, 3, t['links'], bg)
    data_cell(ws, r, 4, t['imp'], bg)
    data_cell(ws, r, 5, t['click'], bg)
    data_cell(ws, r, 6, t['sold'], bg)
    data_cell(ws, r, 7, round(t['cost'], 2), bg)
    data_cell(ws, r, 8, round(t['ad_rev'], 2), bg)
    data_cell(ws, r, 9, round(t['total_rev'], 2), bg)
    data_cell(ws, r, 10, f"{t['roas']:.2f}" if t['roas'] else '-', bg)
    data_cell(ws, r, 11, f"{t['cvr']*100:.4f}%" if t['cvr'] else '0%', bg)
    data_cell(ws, r, 12, f"{t['ctr']*100:.4f}%" if t['ctr'] else '0%', bg)
    data_cell(ws, r, 13, f"{t['cpc']:.2f}" if t['cpc'] else '-', bg)
    data_cell(ws, r, 14, f"{t['cpa']:.2f}" if t['cpa'] else '-', bg)
    data_cell(ws, r, 15, f"{t['acos']*100:.4f}%" if t['acos'] else '0%', bg)
    data_cell(ws, r, 16, f"{t['acoas']*100:.4f}%" if t['acoas'] else '0%', bg, fc='FF0000')

r += 1
write_plp_row(ws7, r, PLP_CURR, plp_t, COLOR_ODD)
r += 1
write_plp_row(ws7, r, PLP_PREV, plp_pt, COLOR_EVEN)

# 环比行
r += 1
bg_hb = COLOR_TOTAL
delta = lambda a, b: a - b
data_cell(ws7, r, 1, '环比', bg_hb, bold=True)
data_cell(ws7, r, 2, delta(plp_t['campaigns'], plp_pt['campaigns']), bg_hb, bold=True)
data_cell(ws7, r, 3, '-', bg_hb, bold=True)
data_cell(ws7, r, 4, ratio_str(plp_t['imp'], plp_pt['imp']), bg_hb, bold=True)
data_cell(ws7, r, 5, ratio_str(plp_t['click'], plp_pt['click']), bg_hb, bold=True)
data_cell(ws7, r, 6, ratio_str(plp_t['sold'], plp_pt['sold']), bg_hb, bold=True)
data_cell(ws7, r, 7, ratio_str(plp_t['cost'], plp_pt['cost']), bg_hb, bold=True)
data_cell(ws7, r, 8, ratio_str(plp_t['ad_rev'], plp_pt['ad_rev']), bg_hb, bold=True)
data_cell(ws7, r, 9, ratio_str(plp_t['total_rev'], plp_pt['total_rev']), bg_hb, bold=True)
data_cell(ws7, r, 10, ratio_str(plp_t['roas'], plp_pt['roas']), bg_hb, bold=True)
data_cell(ws7, r, 11, ratio_str(plp_t['cvr'], plp_pt['cvr']), bg_hb, bold=True)
data_cell(ws7, r, 12, ratio_str(plp_t['ctr'], plp_pt['ctr']), bg_hb, bold=True)
data_cell(ws7, r, 13, ratio_str(plp_t['cpc'], plp_pt['cpc']), bg_hb, bold=True)
data_cell(ws7, r, 14, ratio_str(plp_t['cpa'], plp_pt['cpa']), bg_hb, bold=True)
data_cell(ws7, r, 15, ratio_str(plp_t['acos'], plp_pt['acos']), bg_hb, bold=True)
data_cell(ws7, r, 16, ratio_str(plp_t['acoas'], plp_pt['acoas']), bg_hb, bold=True)

set_border(ws7, 3, r, 1, 16)

def write_plp_dim_section(ws, start_r, title, data_dict, sort_keys, ncols=16):
    r = start_r + 1
    hdr_style(ws, r, 1, title, '8e44ad', 'FFFFFF', True)
    ws.merge_cells(f'A{r}:P{r}')
    r += 1
    dim_hdrs = ['维度', '活动数', '链接数', '曝光量', '点击量', '售出数', '花费', '广告销售额', '总销售额', 'ROAS', 'CVR', 'CTR', 'CPC', 'CPA', 'ACOS', 'ACOAS']
    for ci, h in enumerate(dim_hdrs, 1):
        hdr_style(ws, r, ci, h, '9b59b6')
    r += 1
    for i, k in enumerate(sort_keys):
        d = data_dict.get(k)
        if not d: continue
        bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
        data_cell(ws, r, 1, k, bg, align='left')
        data_cell(ws, r, 2, d['campaigns'], bg)
        data_cell(ws, r, 3, d['links'], bg)
        data_cell(ws, r, 4, d['imp'], bg)
        data_cell(ws, r, 5, d['click'], bg)
        data_cell(ws, r, 6, d['sold'], bg)
        data_cell(ws, r, 7, d['cost'], bg)
        data_cell(ws, r, 8, d['ad_rev'], bg)
        data_cell(ws, r, 9, d['total_rev'], bg)
        data_cell(ws, r, 10, f"{d['roas']:.2f}" if d['roas'] else '-', bg)
        data_cell(ws, r, 11, f"{d['cvr']*100:.4f}%" if d['cvr'] else '0%', bg)
        data_cell(ws, r, 12, f"{d['ctr']*100:.4f}%" if d['ctr'] else '0%', bg)
        data_cell(ws, r, 13, f"{d['cpc']:.2f}" if d['cpc'] else '-', bg)
        data_cell(ws, r, 14, f"{d['cpa']:.2f}" if d['cpa'] else '-', bg)
        data_cell(ws, r, 15, f"{d['acos']*100:.4f}%" if d['acos'] else '0%', bg)
        data_cell(ws, r, 16, f"{d['acoas']*100:.4f}%" if d['acoas'] else '0%', bg, fc='FF0000' if d['acoas'] > 0.1 else '000000')
        r += 1
    set_border(ws, start_r+2, r-1, 1, ncols)
    return r

plp_ana_all = sorted(set(list(plp_by_ana.keys())))
plp_ana_ordered = [a for a in ANALYSTS if a in plp_ana_all] + [a for a in plp_ana_all if a not in ANALYSTS]
plp_cat_all = sorted(set(list(plp_by_cat.keys())))
plp_cat_ordered = [c for c in CATEGORIES if c in plp_cat_all] + [c for c in plp_cat_all if c not in CATEGORIES]
plp_exp_all = sorted(set(list(plp_by_exp.keys())))
plp_exp_ordered = [e for e in EXPAND_TYPES if e in plp_exp_all] + [e for e in plp_exp_all if e not in EXPAND_TYPES]

r_plp = r + 2
r_plp = write_plp_dim_section(ws7, r_plp, '【按分析人】', plp_by_ana, plp_ana_ordered) + 1
r_plp = write_plp_dim_section(ws7, r_plp, '【按品线】', plp_by_cat, plp_cat_ordered) + 1
r_plp = write_plp_dim_section(ws7, r_plp, '【按拓展类型】', plp_by_exp, plp_exp_ordered)

for ci, w in enumerate([12, 8, 8, 10, 10, 8, 10, 12, 12, 8, 8, 8, 8, 8, 8, 8], 1):
    col_letter = chr(64+ci) if ci <= 26 else 'A' + chr(64+ci-26)
    ws7.column_dimensions[col_letter].width = w

# ===== Sheet 8: 新品出单情况 =====
ws8 = wb.create_sheet('新品出单情况')
ws8.sheet_properties.tabColor = '7030A0'

r = 1
hdr_style(ws8, r, 1, '新品出单情况 - 5.14-5.20', COLOR_HEADER, 'FFFFFF', True, 'left')
ws8.merge_cells(f'A{r}:G{r}')

r += 2
hdr_style(ws8, r, 1, '【有对手口径】', '7030A0', 'FFFFFF', True)
ws8.merge_cells(f'A{r}:G{r}')

r += 1
cols8 = ['指标', '本周(5.20)', '上周(5.13)', '变化', '', '', '']
for ci, h in enumerate(cols8, 1):
    hdr_style(ws8, r, ci, h)

ord8_rows = [
    ('有对手总SKU', has_rival_curr, has_rival_prev, has_rival_curr - has_rival_prev, COLOR_ODD),
    ('8日内出单（Y）', y_curr, y_prev, y_curr - y_prev, COLOR_EVEN),
    ('8日外出单（N）', n_curr, n_prev, n_curr - n_prev, COLOR_ODD),
    ('真正未出单', no_curr, no_prev, no_curr - no_prev, COLOR_EVEN),
    ('已出单合计(Y+N)', y_curr+n_curr, y_prev+n_prev, (y_curr+n_curr)-(y_prev+n_prev), COLOR_ODD),
]
for name, v1, v2, v3, bg in ord8_rows:
    r += 1
    data_cell(ws8, r, 1, name, bg, align='left')
    data_cell(ws8, r, 2, v1, bg)
    data_cell(ws8, r, 3, v2, bg)
    data_cell(ws8, r, 4, v3, bg)

r += 1
data_cell(ws8, r, 1, '出单率', COLOR_TOTAL, bold=True, align='left')
data_cell(ws8, r, 2, sale_rate_curr, COLOR_TOTAL, bold=True)
data_cell(ws8, r, 3, '-', COLOR_TOTAL, bold=True)
data_cell(ws8, r, 4, '-', COLOR_TOTAL, bold=True)
set_border(ws8, 3, r, 1, 4)
ws8.column_dimensions['A'].width = 18
for cc in ['B', 'C', 'D']: ws8.column_dimensions[cc].width = 12

# ===== Sheet 9: 新品未出单原因 =====
ws9 = wb.create_sheet('新品未出单原因')
ws9.sheet_properties.tabColor = 'FFA726'

r = 1
hdr_style(ws9, r, 1, '新品未出单原因分析 - 5.14-5.20', COLOR_HEADER, 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:H{r}')
ws9.row_dimensions[r].height = 28

r += 1
hdr_style(ws9, r, 1, '【说明】Y=8日内出单 | N=8日外出单 | 未出单=从未出单', 'F3E5F5', '4A148C', False, 'left')
ws9.merge_cells(f'A{r}:F{r}')

has_rival_no_curr = [x for x in rows_curr if str(x[C['ord8_curr']] or '').strip() == '未出单' and num(x[C['rival_curr']]) > 0]
no_rival_no_curr = [x for x in rows_curr if str(x[C['ord8_curr']] or '').strip() == '未出单' and num(x[C['rival_curr']]) == 0]
has_rival_no_prev = [x for x in rows_prev if str(x[C['ord8_prev']] or '').strip() == '未出单' and num(x[C['rival_prev']]) > 0]
no_rival_no_prev = [x for x in rows_prev if str(x[C['ord8_prev']] or '').strip() == '未出单' and num(x[C['rival_prev']]) == 0]

mkt_order_has = ['竞争无优势', '无市场', '站内无价格优势', '站外出单', '正常', '#N/A', '未知']
mkt_order_no = ['无市场', '未知', '竞争无优势', '#N/A', '其他']

# ── A板块 ──
r += 2
sep_cell = ws9.cell(row=r, column=1, value='━' * 60)
sep_cell.fill = PatternFill('solid', fgColor='E65100')
sep_cell.font = Font(bold=True, color='FFFFFF', name='微软雅黑', size=10)
ws9.merge_cells(f'A{r}:H{r}')

r += 1
hdr_style(ws9, r, 1, f'【A. 有对手未出单新品】  本周: {len(has_rival_no_curr)}个  上周: {len(has_rival_no_prev)}个', 'BF360C', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:F{r}')
ws9.row_dimensions[r].height = 24

# A1
r += 1
hdr_style(ws9, r, 1, '【A1】未出单原因分布', 'E65100', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:F{r}')
r += 1
for ci, h in enumerate(['市场状态', '本周SKU', '占比', '上周SKU', '上周占比', '变化'], 1):
    hdr_style(ws9, r, ci, h, 'FFCCBC')

mkt_has_curr = Counter(str(x[C['mkt_curr']] or '未知') for x in has_rival_no_curr)
mkt_has_prev = Counter(str(x[C['mkt_prev']] or '未知') for x in has_rival_no_prev)

for i, mkt in enumerate(mkt_order_has):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    cnt_c = mkt_has_curr.get(mkt, 0); cnt_p = mkt_has_prev.get(mkt, 0)
    tot_c = len(has_rival_no_curr) or 1; tot_p = len(has_rival_no_prev) or 1
    data_cell(ws9, r, 1, mkt, bg, align='left')
    data_cell(ws9, r, 2, cnt_c, bg)
    data_cell(ws9, r, 3, f"{round(cnt_c/tot_c*100, 1)}%", bg)
    data_cell(ws9, r, 4, cnt_p, bg)
    data_cell(ws9, r, 5, f"{round(cnt_p/tot_p*100, 1)}%", bg)
    data_cell(ws9, r, 6, cnt_c - cnt_p, bg)

r += 1
data_cell(ws9, r, 1, '合计', COLOR_TOTAL, bold=True, align='left')
data_cell(ws9, r, 2, len(has_rival_no_curr), COLOR_TOTAL, bold=True)
data_cell(ws9, r, 3, '100%', COLOR_TOTAL, bold=True)
data_cell(ws9, r, 4, len(has_rival_no_prev), COLOR_TOTAL, bold=True)
data_cell(ws9, r, 5, '100%', COLOR_TOTAL, bold=True)
data_cell(ws9, r, 6, len(has_rival_no_curr) - len(has_rival_no_prev), COLOR_TOTAL, bold=True)

# A2
r += 2
hdr_style(ws9, r, 1, '【A2】按分析人', 'E65100', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:I{r}')
r += 1
cols_an_a = ['分析人'] + mkt_order_has + ['未出单总数']
for ci, h in enumerate(cols_an_a, 1): hdr_style(ws9, r, ci, h, 'FFCCBC')

an_has_data = defaultdict(lambda: {m: 0 for m in mkt_order_has})
for x in has_rival_no_curr:
    an = get_an(x); m = str(x[C['mkt_curr']] or '未知')
    if m in an_has_data[an]: an_has_data[an][m] += 1

for i, an in enumerate(ans_ordered):
    r += 1; bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = an_has_data.get(an, {m: 0 for m in mkt_order_has}); total = sum(d.values())
    data_cell(ws9, r, 1, an, bg, align='left')
    for ci, m in enumerate(mkt_order_has, 2): data_cell(ws9, r, ci, d.get(m, 0), bg)
    data_cell(ws9, r, len(cols_an_a), total, bg)

r += 1
total_a2 = ['合计'] + [sum(an_has_data[a].get(m, 0) for a in ans_ordered) for m in mkt_order_has] + [len(has_rival_no_curr)]
for ci, v in enumerate(total_a2, 1): data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci==1 else 'center')

# A3
r += 2
hdr_style(ws9, r, 1, '【A3】按品线', 'E65100', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:I{r}')
r += 1
cols_cat_a = ['品线'] + mkt_order_has + ['未出单总数']
for ci, h in enumerate(cols_cat_a, 1): hdr_style(ws9, r, ci, h, 'FFCCBC')

cat_has_data = defaultdict(lambda: {m: 0 for m in mkt_order_has})
for x in has_rival_no_curr:
    cat = get_cat(x); m = str(x[C['mkt_curr']] or '未知')
    if m in cat_has_data[cat]: cat_has_data[cat][m] += 1
all_cats_a = sorted(cat_has_data.keys())

for i, cat in enumerate(all_cats_a):
    r += 1; bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = cat_has_data[cat]; total = sum(d.values())
    data_cell(ws9, r, 1, cat, bg, align='left')
    for ci, m in enumerate(mkt_order_has, 2): data_cell(ws9, r, ci, d.get(m, 0), bg)
    data_cell(ws9, r, len(cols_cat_a), total, bg)

r += 1
total_a3 = ['合计'] + [sum(cat_has_data[c].get(m, 0) for c in all_cats_a) for m in mkt_order_has] + [len(has_rival_no_curr)]
for ci, v in enumerate(total_a3, 1): data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci==1 else 'center')

# ── B板块 ──
r += 2
sep_cell_b = ws9.cell(row=r, column=1, value='━' * 60)
sep_cell_b.fill = PatternFill('solid', fgColor='1B5E20')
sep_cell_b.font = Font(bold=True, color='FFFFFF', name='微软雅黑', size=10)
ws9.merge_cells(f'A{r}:H{r}')

r += 1
hdr_style(ws9, r, 1, f'【B. 无对手未出单新品】  本周: {len(no_rival_no_curr)}个  上周: {len(no_rival_no_prev)}个', '2E7D32', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:F{r}')
ws9.row_dimensions[r].height = 24

# B1
r += 1
hdr_style(ws9, r, 1, '【B1】未出单原因分布', '388E3C', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:F{r}')
r += 1
for ci, h in enumerate(['市场状态', '本周SKU', '占比', '上周SKU', '上周占比', '变化'], 1):
    hdr_style(ws9, r, ci, h, 'C8E6C9')

mkt_no_curr = Counter(str(x[C['mkt_curr']] or '未知') for x in no_rival_no_curr)
mkt_no_prev = Counter(str(x[C['mkt_prev']] or '未知') for x in no_rival_no_prev)

for i, mkt in enumerate(mkt_order_no):
    r += 1; bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    cnt_c = mkt_no_curr.get(mkt, 0); cnt_p = mkt_no_prev.get(mkt, 0)
    tot_c = len(no_rival_no_curr) or 1; tot_p = len(no_rival_no_prev) or 1
    data_cell(ws9, r, 1, mkt, bg, align='left')
    data_cell(ws9, r, 2, cnt_c, bg)
    data_cell(ws9, r, 3, f"{round(cnt_c/tot_c*100, 1)}%", bg)
    data_cell(ws9, r, 4, cnt_p, bg)
    data_cell(ws9, r, 5, f"{round(cnt_p/tot_p*100, 1)}%", bg)
    data_cell(ws9, r, 6, cnt_c - cnt_p, bg)

r += 1
data_cell(ws9, r, 1, '合计', COLOR_TOTAL, bold=True, align='left')
data_cell(ws9, r, 2, len(no_rival_no_curr), COLOR_TOTAL, bold=True)
data_cell(ws9, r, 3, '100%', COLOR_TOTAL, bold=True)
data_cell(ws9, r, 4, len(no_rival_no_prev), COLOR_TOTAL, bold=True)
data_cell(ws9, r, 5, '100%', COLOR_TOTAL, bold=True)
data_cell(ws9, r, 6, len(no_rival_no_curr) - len(no_rival_no_prev), COLOR_TOTAL, bold=True)

# B2
r += 2
hdr_style(ws9, r, 1, '【B2】按分析人', '388E3C', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:G{r}')
r += 1
cols_an_b = ['分析人'] + mkt_order_no + ['未出单总数']
for ci, h in enumerate(cols_an_b, 1): hdr_style(ws9, r, ci, h, 'C8E6C9')

an_no_data = defaultdict(lambda: {m: 0 for m in mkt_order_no})
for x in no_rival_no_curr:
    an = get_an(x); m = str(x[C['mkt_curr']] or '未知')
    if m in an_no_data[an]: an_no_data[an][m] += 1

for i, an in enumerate(ans_ordered):
    r += 1; bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = an_no_data.get(an, {m: 0 for m in mkt_order_no}); total = sum(d.values())
    data_cell(ws9, r, 1, an, bg, align='left')
    for ci, m in enumerate(mkt_order_no, 2): data_cell(ws9, r, ci, d.get(m, 0), bg)
    data_cell(ws9, r, len(cols_an_b), total, bg)

r += 1
total_b2 = ['合计'] + [sum(an_no_data[a].get(m, 0) for a in ans_ordered) for m in mkt_order_no] + [len(no_rival_no_curr)]
for ci, v in enumerate(total_b2, 1): data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci==1 else 'center')

# B3
r += 2
hdr_style(ws9, r, 1, '【B3】按品线', '388E3C', 'FFFFFF', True, 'left')
ws9.merge_cells(f'A{r}:G{r}')
r += 1
cols_cat_b = ['品线'] + mkt_order_no + ['未出单总数']
for ci, h in enumerate(cols_cat_b, 1): hdr_style(ws9, r, ci, h, 'C8E6C9')

cat_no_data = defaultdict(lambda: {m: 0 for m in mkt_order_no})
for x in no_rival_no_curr:
    cat = get_cat(x); m = str(x[C['mkt_curr']] or '未知')
    if m in cat_no_data[cat]: cat_no_data[cat][m] += 1
all_cats_b = sorted(cat_no_data.keys())

for i, cat in enumerate(all_cats_b):
    r += 1; bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = cat_no_data[cat]; total = sum(d.values())
    data_cell(ws9, r, 1, cat, bg, align='left')
    for ci, m in enumerate(mkt_order_no, 2): data_cell(ws9, r, ci, d.get(m, 0), bg)
    data_cell(ws9, r, len(cols_cat_b), total, bg)

r += 1
total_b3 = ['合计'] + [sum(cat_no_data[c].get(m, 0) for c in all_cats_b) for m in mkt_order_no] + [len(no_rival_no_curr)]
for ci, v in enumerate(total_b3, 1): data_cell(ws9, r, ci, v, COLOR_TOTAL, bold=True, align='left' if ci==1 else 'center')

ws9.column_dimensions['A'].width = 18
for cc_i in range(2, 10):
    col_letter = chr(64+cc_i) if cc_i <= 26 else 'A' + chr(64+cc_i-26)
    ws9.column_dimensions[col_letter].width = 12

# ===== Sheet 10: 新品PLG维度 =====
ws10 = wb.create_sheet('新品PLG维度')
ws10.sheet_properties.tabColor = '00B050'

plg_categories = {'PLP+PLG同开': [], '单链接PLP+PLG同开': [], '单PLG': [], '单PLP': [], '单PLG且未出单': [], '无广告': []}

for r_data in rows_curr:
    sku = str(r_data[C['sku']]).strip()
    plp_on = str(r_data[C['plp_curr']] or '').strip().upper() == 'Y'
    plg_on = num(r_data[C['plg_curr']]) > 0
    in_plg_detail = sku in plp_plg_enabled     # PLP明细中是否开启PLG=Y
    is_unsold = str(r_data[C['ord8_curr']] or '').strip() == '未出单'

    # 单链接PLP+PLG同开: PLP明细中开启PLG=Y 且仅1个campaign
    if in_plg_detail and sku_campaign_count.get(sku, 0) == 1:
        plg_categories['单链接PLP+PLG同开'].append(r_data)
    elif plp_on and plg_on:
        plg_categories['PLP+PLG同开'].append(r_data)
    elif plg_on and not plp_on:
        if is_unsold:
            plg_categories['单PLG且未出单'].append(r_data)
        else:
            plg_categories['单PLG'].append(r_data)
    elif plp_on and not plg_on:
        plg_categories['单PLP'].append(r_data)
    else:
        plg_categories['无广告'].append(r_data)

r = 1
hdr_style(ws10, r, 1, '新品PLG维度 - 5.14-5.20', COLOR_HEADER, 'FFFFFF', True, 'left')
ws10.merge_cells(f'A{r}:G{r}')

r += 2
hdr_style(ws10, r, 1, '分类汇总', COLOR_HEADER, 'FFFFFF', True)
ws10.merge_cells(f'A{r}:K{r}')

r += 1
cats10 = ['类型', 'SKU数', 'PLP+PLG同开', '单链接PLP+PLG同开', '单PLG', '单PLP', '单PLG且未出单', '无广告']
for ci, h in enumerate(cats10, 1):
    hdr_style(ws10, r, ci, h)

r += 1
data_cell(ws10, r, 1, '合计', COLOR_ODD, bold=True, align='left')
data_cell(ws10, r, 2, total_sku, COLOR_ODD, bold=True)
data_cell(ws10, r, 3, len(plg_categories['PLP+PLG同开']), COLOR_ODD, bold=True)
data_cell(ws10, r, 4, len(plg_categories['单链接PLP+PLG同开']), COLOR_ODD, bold=True, fc='FF0000')
data_cell(ws10, r, 5, len(plg_categories['单PLG']), COLOR_ODD, bold=True)
data_cell(ws10, r, 6, len(plg_categories['单PLP']), COLOR_ODD, bold=True)
data_cell(ws10, r, 7, len(plg_categories['单PLG且未出单']), COLOR_ODD, bold=True, fc='FF0000')
data_cell(ws10, r, 8, len(plg_categories['无广告']), COLOR_ODD, bold=True)

# 按分析人
r += 2
hdr_style(ws10, r, 1, '【按分析人】', '00B050', 'FFFFFF', True)
ws10.merge_cells(f'A{r}:G{r}')
r += 1
for ci, h in enumerate(cats10, 1):
    hdr_style(ws10, r, ci, h)

def count_plg_dim(rlist, key_fn):
    result = defaultdict(lambda: {'PLP+PLG同开': 0, '单链接PLP+PLG同开': 0, '单PLG': 0, '单PLP': 0, '单PLG且未出单': 0, '无广告': 0})
    for r_data in rlist:
        k = key_fn(r_data)
        sku = str(r_data[C['sku']]).strip()
        plp_on = str(r_data[C['plp_curr']] or '').strip().upper() == 'Y'
        plg_on = num(r_data[C['plg_curr']]) > 0
        in_plg_detail = sku in plp_plg_enabled
        is_unsold = str(r_data[C['ord8_curr']] or '').strip() == '未出单'

        if in_plg_detail and sku_campaign_count.get(sku, 0) == 1:
            result[k]['单链接PLP+PLG同开'] += 1
        elif plp_on and plg_on:
            result[k]['PLP+PLG同开'] += 1
        elif plg_on and not plp_on:
            if is_unsold: result[k]['单PLG且未出单'] += 1
            else: result[k]['单PLG'] += 1
        elif plp_on and not plg_on: result[k]['单PLP'] += 1
        else: result[k]['无广告'] += 1
    return result

plg_by_ana = count_plg_dim(rows_curr, lambda r: get_an(r))
for i, an in enumerate(ans_ordered):
    r += 1; bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = plg_by_ana.get(an, {})
    data_cell(ws10, r, 1, an, bg, align='left')
    data_cell(ws10, r, 2, sum(d.values()), bg)
    data_cell(ws10, r, 3, d.get('PLP+PLG同开', 0), bg)
    data_cell(ws10, r, 4, d.get('单链接PLP+PLG同开', 0), bg, fc='FF0000' if d.get('单链接PLP+PLG同开', 0) > 0 else '000000')
    data_cell(ws10, r, 5, d.get('单PLG', 0), bg)
    data_cell(ws10, r, 6, d.get('单PLP', 0), bg)
    data_cell(ws10, r, 7, d.get('单PLG且未出单', 0), bg, fc='FF0000' if d.get('单PLG且未出单', 0) > 0 else '000000')
    data_cell(ws10, r, 8, d.get('无广告', 0), bg)

# 按品线
r += 2
hdr_style(ws10, r, 1, '【按品线】', '00B050', 'FFFFFF', True)
ws10.merge_cells(f'A{r}:G{r}')
r += 1
for ci, h in enumerate(cats10, 1):
    hdr_style(ws10, r, ci, h)

plg_by_cat = count_plg_dim(rows_curr, lambda r: get_cat(r))
for i, cat in enumerate(cats_ordered):
    r += 1; bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = plg_by_cat.get(cat, {})
    data_cell(ws10, r, 1, cat, bg, align='left')
    data_cell(ws10, r, 2, sum(d.values()), bg)
    data_cell(ws10, r, 3, d.get('PLP+PLG同开', 0), bg)
    data_cell(ws10, r, 4, d.get('单链接PLP+PLG同开', 0), bg, fc='FF0000' if d.get('单链接PLP+PLG同开', 0) > 0 else '000000')
    data_cell(ws10, r, 5, d.get('单PLG', 0), bg)
    data_cell(ws10, r, 6, d.get('单PLP', 0), bg)
    data_cell(ws10, r, 7, d.get('单PLG且未出单', 0), bg, fc='FF0000' if d.get('单PLG且未出单', 0) > 0 else '000000')
    data_cell(ws10, r, 8, d.get('无广告', 0), bg)

set_border(ws10, 2, r, 1, 8)
for ci, w in enumerate([14, 8, 14, 14, 8, 8, 14, 8], 1):
    ws10.column_dimensions[chr(64+ci)].width = w

# ===== Sheet 11: 新品明细 =====
ws11 = wb.create_sheet('新品明细')
ws11.sheet_properties.tabColor = '4472C4'

r = 1
hdr_style(ws11, r, 1, '新品明细 - 5.14-5.20', COLOR_HEADER, 'FFFFFF', True, 'left')
ws11.merge_cells(f'A{r}:W{r}')
ws11.row_dimensions[r].height = 28

r += 1
cols11 = [
    '销售编号', 'SKU', '上架日期', '首次出单', '分析人', '品类', '产品拓展',
    '5.14-5.20销量', '5.7-5.13销量', '5.14-5.20销售额', '5.7-5.13销售额',
    '5.20对手销量', '5.13对手销量', '5.20市占比', '5.13市占比',
    '5.20 8日出单', '5.20 7日频次', '5.20 7日新品频次',
    '5.20市场状态', '5.14-5.20操作判断',
    '开启PLP(5.11-5.17)', 'PLG最高费率(5.11-5.17)',
    '广告分类',
]
for ci, h in enumerate(cols11, 1):
    hdr_style(ws11, r, ci, h)

def get_ad_class(r_data):
    sku = str(r_data[C['sku']]).strip()
    plp_on = str(r_data[C['plp_curr']] or '').strip().upper() == 'Y'
    plg_on = num(r_data[C['plg_curr']]) > 0
    in_plg_detail = sku in plp_plg_enabled     # PLP明细中是否开启PLG=Y
    is_unsold = str(r_data[C['ord8_curr']] or '').strip() == '未出单'

    # 单链接PLP+PLG同开: PLP明细中开启PLG=Y 且仅1个campaign
    if in_plg_detail and sku_campaign_count.get(sku, 0) == 1:
        return '单链接PLP+PLG同开'
    elif plp_on and plg_on:
        return 'PLP+PLG同开'
    elif plg_on and not plp_on:
        return '单PLG且未出单' if is_unsold else '单PLG'
    elif plp_on and not plg_on: return '单PLP'
    return '无广告'

for i, r_data in enumerate(rows_curr):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    sku = str(r_data[C['sku']]).strip()
    first_order = get_date(r_data[C['first_order']])

    data_cell(ws11, r, 1, r_data[C['sale_no']], bg)
    data_cell(ws11, r, 2, sku, bg)
    data_cell(ws11, r, 3, str(get_date(r_data[C['list_date']]))[:10] if get_date(r_data[C['list_date']]) else '', bg)
    data_cell(ws11, r, 4, str(first_order)[:10] if first_order else '未出单', bg)
    data_cell(ws11, r, 5, get_an(r_data), bg)
    data_cell(ws11, r, 6, get_cat(r_data), bg)
    data_cell(ws11, r, 7, get_exp(r_data), bg)
    data_cell(ws11, r, 8, num(r_data[C['sales_curr']]), bg)
    data_cell(ws11, r, 9, num(r_data[C['sales_prev']]), bg)
    data_cell(ws11, r, 10, num(r_data[C['rev_curr']]), bg)
    data_cell(ws11, r, 11, num(r_data[C['rev_prev']]), bg)
    data_cell(ws11, r, 12, num(r_data[C['rival_curr']]), bg)
    data_cell(ws11, r, 13, num(r_data[C['rival_prev']]), bg)
    data_cell(ws11, r, 14, num(r_data[C['share_curr']]) * 100, bg)
    data_cell(ws11, r, 15, num(r_data[C['share_prev']]) * 100, bg)
    data_cell(ws11, r, 16, r_data[C['ord8_curr']] or '', bg)
    data_cell(ws11, r, 17, r_data[C['freq7_curr']] or '', bg)
    data_cell(ws11, r, 18, r_data[C['nfreq7_curr']] or '', bg)
    data_cell(ws11, r, 19, r_data[C['mkt_curr']] or '', bg)
    data_cell(ws11, r, 20, r_data[C['op_curr']] or '', bg)
    plp_val = str(r_data[C['plp_curr']] or '').strip().upper()
    data_cell(ws11, r, 21, 'Y' if plp_val == 'Y' else 'N', bg)
    data_cell(ws11, r, 22, num(r_data[C['plg_curr']]) * 100, bg)
    ad_class = get_ad_class(r_data)
    ad_fc = 'FF0000' if ad_class in ('单PLG且未出单', '单链接PLP+PLG同开') else '000000'
    data_cell(ws11, r, 23, ad_class, bg, fc=ad_fc)

set_border(ws11, 2, r, 1, len(cols11))

for ci in range(1, len(cols11)+1):
    col_letter = chr(64+ci) if ci <= 26 else 'A' + chr(64+ci-26)
    ws11.column_dimensions[col_letter].width = 8 if ci > 7 else 10
ws11.column_dimensions['B'].width = 18
ws11.column_dimensions['C'].width = 12
ws11.column_dimensions['D'].width = 12

# ===== Sheet 12: 新品广告明细 =====
ws12 = wb.create_sheet('新品广告明细')
ws12.sheet_properties.tabColor = 'FF6347'

r = 1
hdr_style(ws12, r, 1, f'新品广告明细 - {PLP_CURR}', COLOR_HEADER, 'FFFFFF', True, 'left')
ws12.merge_cells(f'A{r}:Y{r}')
ws12.row_dimensions[r].height = 28

r += 1
cols12 = [
    '周期', '广告活动', 'SKU', '店铺', 'PLP开始', '上架日期', '首次出单',
    '分析人', '品类', '拓展类型',
    '曝光量', '点击量', '售出数', '广告花费', '广告销售额', '总销售额',
    'ROAS', 'CVR', 'CTR', 'CPC', 'CPA', 'ACOS', 'ACOAS',
    '开启PLG', '广告分类',
]
for ci, h in enumerate(cols12, 1):
    hdr_style(ws12, r, ci, h)

plp_curr_sorted = sorted(plp_curr_rows, key=lambda x: (
    get_an(sku_info.get(str(x[PC['sku']]).strip(), {})) if str(x[PC['sku']]).strip() in sku_info else '未知',
    str(x[PC['sku']]).strip()
))

for i, row in enumerate(plp_curr_sorted):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    sku = str(row[PC['sku']] or '').strip()
    info = sku_info.get(sku, {})
    cost = num(row[PC['cost']])
    ad_rev = num(row[PC['ad_rev']])
    total_r = num(row[PC['total_rev']])
    impr = num(row[PC['impr']])
    click = num(row[PC['click']])
    sold = num(row[PC['sold']])
    plg_y = str(row[PC['plg_enabled']] or '').strip() == 'Y'

    data_cell(ws12, r, 1, row[PC['period']] or '', bg)
    data_cell(ws12, r, 2, row[PC['campaign']] or '', bg)
    data_cell(ws12, r, 3, sku, bg)
    data_cell(ws12, r, 4, row[PC['store']] or '', bg)
    data_cell(ws12, r, 5, str(get_date(row[PC['plp_start']]))[:10] if get_date(row[PC['plp_start']]) else '', bg)
    data_cell(ws12, r, 6, str(get_date(row[PC['list_date']]))[:10] if get_date(row[PC['list_date']]) else '', bg)
    data_cell(ws12, r, 7, str(get_date(row[PC['first_order']]))[:10] if get_date(row[PC['first_order']]) else '未出单', bg)
    data_cell(ws12, r, 8, get_an(info) if info else (row[PC['analyst']] or ''), bg)
    data_cell(ws12, r, 9, get_cat(info) if info else (row[PC['category']] or '未分类'), bg)
    data_cell(ws12, r, 10, get_exp(info) if info else (row[PC['expand_type']] or ''), bg)
    data_cell(ws12, r, 11, int(impr), bg)
    data_cell(ws12, r, 12, int(click), bg)
    data_cell(ws12, r, 13, int(sold), bg)
    data_cell(ws12, r, 14, round(cost, 2), bg)
    data_cell(ws12, r, 15, round(ad_rev, 2), bg)
    data_cell(ws12, r, 16, round(total_r, 2), bg)
    roas_v = calc_rate(ad_rev, cost)
    cvr_v = calc_rate(sold, click)
    ctr_v = calc_rate(click, impr)
    acos_v = calc_rate(cost, ad_rev)
    acoas_v = calc_rate(cost, sku_rev_curr.get(sku, 0))
    data_cell(ws12, r, 17, f"{roas_v:.2f}", bg)
    data_cell(ws12, r, 18, f"{cvr_v*100:.2f}%", bg)
    data_cell(ws12, r, 19, f"{ctr_v*100:.2f}%", bg)
    cpc_v = round(cost / click, 2) if click else '#DIV/0!'
    cpa_v = round(cost / sold, 2) if sold else '#DIV/0!'
    data_cell(ws12, r, 20, cpc_v, bg)
    data_cell(ws12, r, 21, cpa_v, bg)
    data_cell(ws12, r, 22, f"{acos_v*100:.2f}%", bg, fc='FF0000' if acos_v > 0.3 else '000000')
    data_cell(ws12, r, 23, f"{acoas_v*100:.2f}%", bg, fc='FF0000' if acoas_v > 0.1 else '000000')
    data_cell(ws12, r, 24, 'Y' if plg_y else 'N', bg)
    ad_class = get_ad_class(info) if info else '无广告'
    ad_fc_12 = 'FF0000' if ad_class in ('单PLG且未出单', '单链接PLP+PLG同开') else '000000'
    data_cell(ws12, r, 25, ad_class, bg, fc=ad_fc_12)

set_border(ws12, 2, r, 1, len(cols12))
for ci in range(1, len(cols12)+1):
    col_letter = chr(64+ci) if ci <= 26 else 'A' + chr(64+ci-26)
    ws12.column_dimensions[col_letter].width = 8
ws12.column_dimensions['B'].width = 18
ws12.column_dimensions['C'].width = 18
ws12.column_dimensions['D'].width = 10
ws12.column_dimensions['E'].width = 12
ws12.column_dimensions['F'].width = 12
ws12.column_dimensions['G'].width = 12

# ===== Sheet 13: 品效分析 =====
ws13 = wb.create_sheet('品效分析')
ws13.sheet_properties.tabColor = 'FF4081'

r = 1
hdr_style(ws13, r, 1, '品效分析 - 按上架月份Cohort（截至5.20）', COLOR_HEADER, 'FFFFFF', True, 'left')
ws13.merge_cells(f'A{r}:K{r}')
ws13.row_dimensions[r].height = 28

# 主Cohort表：上架月份 × 自然月销售额 + 品效
r += 2
hdr_style(ws13, r, 1, '【销售额Cohort表】', 'FF4081', 'FFFFFF', True)
ws13.merge_cells(f'A{r}:Q{r}')

r += 1
cohort_cols = ['上架月份', 'SKU个数']
for m in range(1, 6):
    cohort_cols.append(f'26-{m}月销售额')
for m in range(1, 6):
    cohort_cols.append(f'{m}月品效($/SKU)')
cohort_cols.append('最新品效趋势')

for ci, h in enumerate(cohort_cols, 1):
    hdr_style(ws13, r, ci, h)

for i, m in enumerate(launch_months):
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    d = cohort_data[m]
    sku_count = d['sku_count']

    data_cell(ws13, r, 1, f'{m}月上架', bg, align='left')
    data_cell(ws13, r, 2, sku_count, bg)

    # 销售额列
    for mon in range(1, 6):
        rev = d['monthly_revenue'].get(mon, 0)
        data_cell(ws13, r, 2 + mon, round(rev, 2), bg)

    # 品效列（销售额/SKU数）
    pinxiao_values = []
    for mon in range(1, 6):
        rev = d['monthly_revenue'].get(mon, 0)
        px = rev / sku_count if sku_count > 0 else 0
        pinxiao_values.append(px)
        # 品效分级着色
        if px >= 500:
            fc = '08845a'  # 高效绿色
        elif px >= 100:
            fc = '000000'  # 一般黑色
        else:
            fc = 'c0392b'  # 低效红色
        data_cell(ws13, r, 2 + 5 + mon, round(px, 2), bg, fc=fc)

    # 趋势简写
    recent = [v for v in pinxiao_values if v > 0]
    if len(recent) >= 3:
        if recent[-1] > recent[-2] > recent[-3]:
            trend = '📈爬坡'
        elif recent[-1] < recent[-2] < recent[-3]:
            trend = '📉衰退'
        else:
            trend = '➡平稳'
    elif len(recent) >= 2:
        trend = '📈爬坡' if recent[-1] > recent[-2] else '📉衰退'
    else:
        trend = '数据不足'
    data_cell(ws13, r, len(cohort_cols), trend, bg, bold=True)

set_border(ws13, 3, r, 1, len(cohort_cols))
for ci in range(1, len(cohort_cols)+1):
    col_letter = chr(64+ci) if ci <= 26 else 'A' + chr(64+ci-26)
    ws13.column_dimensions[col_letter].width = 12
ws13.column_dimensions['A'].width = 14

# 品效维度交叉
r += 2
hdr_style(ws13, r, 1, '【品效 × 品线】最新月品效（$/SKU）', 'FF4081', 'FFFFFF', True)
ws13.merge_cells(f'A{r}:G{r}')

# 按品线×上架月份计算品效
r += 1
pinxiao_cat_hdrs = ['品线'] + [f'{m}月上架' for m in launch_months]
for ci, h in enumerate(pinxiao_cat_hdrs, 1):
    hdr_style(ws13, r, ci, h)

# 需要按品线+上架月分组计算
cat_cohort = defaultdict(lambda: defaultdict(lambda: {'sku_count': 0, 'latest_month_rev': 0}))
for r_data in rows_curr:
    list_d = get_date(r_data[C['list_date']])
    if not list_d: continue
    launch_month = list_d.month
    cat = get_cat(r_data)
    sku = str(r_data[C['sku']]).strip()
    cat_cohort[cat][launch_month]['sku_count'] += 1
    # 取最新月（5月）的销售额
    for wi, (period, month_num) in enumerate(WEEK_PERIODS):
        if month_num == 5:
            cat_cohort[cat][launch_month]['latest_month_rev'] += num(r_data[WEEK_REV_COLS[wi]])

for i, cat in enumerate(cats_ordered if cats_ordered else CATEGORIES):
    if cat not in cat_cohort: continue
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    data_cell(ws13, r, 1, cat, bg, align='left')
    for j, m in enumerate(launch_months):
        d = cat_cohort[cat].get(m, {'sku_count': 0, 'latest_month_rev': 0})
        sku_n = d['sku_count']
        rev = d['latest_month_rev']
        px = rev / sku_n if sku_n > 0 else 0
        cell_val = f"{px:.0f}" if px > 0 else '-'
        data_cell(ws13, r, 2 + j, cell_val, bg)

set_border(ws13, r - len([c for c in (cats_ordered if cats_ordered else CATEGORIES) if c in cat_cohort]) + 1, r, 1, len(pinxiao_cat_hdrs))

# 品效 × 拓展类型
r += 2
hdr_style(ws13, r, 1, '【品效 × 拓展类型】最新月品效（$/SKU）', 'FF4081', 'FFFFFF', True)
ws13.merge_cells(f'A{r}:G{r}')

r += 1
pinxiao_exp_hdrs = ['拓展类型'] + [f'{m}月上架' for m in launch_months]
for ci, h in enumerate(pinxiao_exp_hdrs, 1):
    hdr_style(ws13, r, ci, h)

exp_cohort = defaultdict(lambda: defaultdict(lambda: {'sku_count': 0, 'latest_month_rev': 0}))
for r_data in rows_curr:
    list_d = get_date(r_data[C['list_date']])
    if not list_d: continue
    launch_month = list_d.month
    exp = get_exp(r_data)
    exp_cohort[exp][launch_month]['sku_count'] += 1
    for wi, (period, month_num) in enumerate(WEEK_PERIODS):
        if month_num == 5:
            exp_cohort[exp][launch_month]['latest_month_rev'] += num(r_data[WEEK_REV_COLS[wi]])

for i, exp in enumerate(exps_ordered):
    if exp not in exp_cohort: continue
    r += 1
    bg = COLOR_ODD if i % 2 == 0 else COLOR_EVEN
    data_cell(ws13, r, 1, exp, bg, align='left')
    for j, m in enumerate(launch_months):
        d = exp_cohort[exp].get(m, {'sku_count': 0, 'latest_month_rev': 0})
        sku_n = d['sku_count']
        rev = d['latest_month_rev']
        px = rev / sku_n if sku_n > 0 else 0
        cell_val = f"{px:.0f}" if px > 0 else '-'
        data_cell(ws13, r, 2 + j, cell_val, bg)

set_border(ws13, r - len(exps_ordered) + 1, r, 1, len(pinxiao_exp_hdrs))

# ===== 保存 =====
print(f"\n保存到: {OUTPUT_FILE}")
wb.save(OUTPUT_FILE)
print("完成！")

# Summary
print(f"\n=== 数据摘要 ===")
print(f"总SKU: {total_sku}, 新上架: {new_list}")
print(f"销量: {total_sales_curr} (上期: {total_sales_prev}), 销售额: ${total_rev_curr:,.2f}")
print(f"新品销售占比: {dept_ratio_str}")
print(f"有对手: {has_rival_curr} (Y:{y_curr} N:{n_curr} 未:{no_curr}), 出单率: {sale_rate_curr}")
print(f"及时率: {rate_curr}, 低占比: {len(low_share_skus)}")
print(f"PLP {PLP_CURR}: {plp_t['campaigns']}活动/{plp_t['links']}链接, 花费${plp_t['cost']:.2f}, ROAS={plp_t['roas']:.2f}, ACOAS={plp_t['acoas']*100:.2f}%")
print(f"PLG: ", {k: len(v) for k, v in plg_categories.items()})
print(f"A(有对手未出单): {len(has_rival_no_curr)}, B(无对手未出单): {len(no_rival_no_curr)}")
print(f"\n品效Cohort:")
for m in launch_months:
    d = cohort_data[m]
    px_vals = []
    for mon in range(1, 6):
        rev = d['monthly_revenue'].get(mon, 0)
        px_vals.append(f"${rev/d['sku_count']:.0f}" if d['sku_count'] > 0 else '-')
    print(f"  {m}月: {' | '.join(px_vals)}")
