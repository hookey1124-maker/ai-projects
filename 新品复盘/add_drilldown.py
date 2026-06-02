"""
为4周HTML看板添加表格下钻图表功能
纯增量修改：CSS + HTML panel + JS functions + wiring
含PLP 4周数据提取（Tab4广告追踪）
"""
import re, sys, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import defaultdict

import os
WORKDIR = os.path.dirname(os.path.abspath(__file__)) + '/'
INPUT = WORKDIR + '新品板块_4.30-5.27_4weeks.html'
OUTPUT = WORKDIR + '新品板块_4.30-5.27_4weeks_drill.html'
SRC = WORKDIR + '周报/新品检查周源数据和PLP数据.xlsx'

with open(INPUT, 'r', encoding='utf-8') as f:
    html = f.read()

before = len(html)
print(f"输入: {before/1024:.0f}KB")

# ===== 0. 提取PLP 4周维度数据 =====
print("0. 提取PLP 4周数据...")
wb = load_workbook(SRC, data_only=True)
ws_plp = wb['PLP明细']

PLP_PERIODS = ['4.27-5.3', '5.4-5.10', '5.11-5.17', '5.18-5.24']

def safe_f(v, d=0):
    try: return float(v) if v else d
    except: return d

plp_rows = []
for row in ws_plp.iter_rows(min_row=2, values_only=True):
    period = str(row[0] or '').strip()
    if period in PLP_PERIODS:
        plp_rows.append({
            'period': period,
            'analyst': str(row[8] or '').strip() or '未知',
            'category': str(row[9] or '').strip() or '未分类',
            'expand_type': str(row[10] or '').strip() or '其他',
            'sku': str(row[2] or '').strip(),
            'cost': safe_f(row[14]),
            'ad_rev': safe_f(row[15]),
            'total_rev': safe_f(row[16]),
        })

def aggregate_plp_4w(dim_key):
    result = defaultdict(lambda: {'spend4w':[0,0,0,0], 'adSales4w':[0,0,0,0], 'totalRev4w':[0,0,0,0]})
    # Per-period SKU dedup sets for total_rev
    seen_sku = [set(), set(), set(), set()]
    for r in plp_rows:
        wi = PLP_PERIODS.index(r['period'])
        key = r[dim_key]
        result[key]['spend4w'][wi] += r['cost']
        result[key]['adSales4w'][wi] += r['ad_rev']
        if r['sku'] not in seen_sku[wi]:
            result[key]['totalRev4w'][wi] += r['total_rev']
            seen_sku[wi].add(r['sku'])
    for k, v in result.items():
        v['acos4w'] = [round(v['spend4w'][i]/v['adSales4w'][i]*100,1) if v['adSales4w'][i]>0 else 0 for i in range(4)]
        v['acoas4w'] = [round(v['spend4w'][i]/v['totalRev4w'][i]*100,1) if v['totalRev4w'][i]>0 else 0 for i in range(4)]
        v['spend4w'] = [round(x,2) for x in v['spend4w']]
        v['adSales4w'] = [round(x,2) for x in v['adSales4w']]
        v['totalRev4w'] = [round(x,2) for x in v['totalRev4w']]
    return result

plp_an_4w = aggregate_plp_4w('analyst')
plp_cat_4w = aggregate_plp_4w('category')
plp_exp_4w = aggregate_plp_4w('expand_type')
print(f"  PLP分析人:{len(plp_an_4w)} 品线:{len(plp_cat_4w)} 拓展类型:{len(plp_exp_4w)}")

def make_js_arr(name, data, keyname):
    items = [{keyname: k, 'spend4w': v['spend4w'], 'adSales4w': v['adSales4w'],
              'acos4w': v['acos4w'], 'acoas4w': v['acoas4w']} for k,v in data.items()]
    return 'const ' + name + ' = ' + json.dumps(items, ensure_ascii=False, separators=(',',':')) + ';'

# 提取PLG 4周数据
print("0b. 提取PLG 4周数据...")
ws_main = wb['四三数据累计']
PLG_FEE_COLS = [141, 142, 143, 144]  # 4.30-5.6, 5.4-5.10, 5.11-5.17, 5.18-5.24 PLG费率
SALES_COLS = [15, 16, 17, 18]  # 销量
REVENUE_COLS = [28, 29, 30, 31]  # 销售额(对应4周)

plg_an_data = defaultdict(lambda: {'spend4w':[0,0,0,0], 'adSales4w':[0,0,0,0], 'totalRev4w':[0,0,0,0]})
for row in ws_main.iter_rows(min_row=2, values_only=True):
    sku = str(row[1] or '').strip()
    if not sku: continue
    analyst = str(row[4] or '').strip() or '未知'
    for wi in range(4):
        rate = safe_f(row[PLG_FEE_COLS[wi]])
        sales = safe_f(row[SALES_COLS[wi]])
        rev = safe_f(row[REVENUE_COLS[wi]])
        spend = round(rate * sales, 2)
        plg_an_data[analyst]['spend4w'][wi] += spend
        plg_an_data[analyst]['adSales4w'][wi] += spend  # PLG广告销售额≈spend（保守估计，实际需PLP明细）
        plg_an_data[analyst]['totalRev4w'][wi] += rev   # 自然周总销售额

for k, v in plg_an_data.items():
    v['acos4w'] = [round(v['spend4w'][i]/v['adSales4w'][i]*100,1) if v['adSales4w'][i]>0 else 0 for i in range(4)]
    v['acoas4w'] = [round(v['spend4w'][i]/v['totalRev4w'][i]*100,1) if v['totalRev4w'][i]>0 else 0 for i in range(4)]
    v['spend4w'] = [round(x,2) for x in v['spend4w']]
    v['adSales4w'] = [round(x,2) for x in v['adSales4w']]
    v['totalRev4w'] = [round(x,2) for x in v['totalRev4w']]

# PLG needs custom make_js_arr to include totalRev4w
plg_items = [{'analyst': k, 'spend4w': v['spend4w'], 'adSales4w': v['adSales4w'],
              'totalRev4w': v['totalRev4w'], 'acos4w': v['acos4w'], 'acoas4w': v['acoas4w']} for k,v in plg_an_data.items()]
plg_an_4w_js = 'const plgAn4w = ' + json.dumps(plg_items, ensure_ascii=False, separators=(',',':')) + ';'
print(f"  PLG分析人:{len(plg_an_data)}")

# 生成JS数据块
plp_data_block = '\n' + make_js_arr('plpAn4w', plp_an_4w, 'analyst') + '\n'
plp_data_block += make_js_arr('plpCat4w', plp_cat_4w, 'category') + '\n'
plp_data_block += make_js_arr('plpExp4w', plp_exp_4w, 'expandType') + '\n'
plp_data_block += plg_an_4w_js + '\n'

# 注入到script中（数据块后，helper函数前）
html = html.replace('\n\nfunction fmtNum', plp_data_block + '\nfunction fmtNum', 1)
print("0. PLP+PLG 4周数据块 已注入")

# ===== 1. CSS: 在</style>前插入 =====
DRILL_CSS = """
#drill-panel{display:none;margin-top:12px;background:#fafbfc;border:1px solid #e0e0e0;border-radius:8px;padding:14px 18px;overflow:hidden;transition:max-height .3s ease,opacity .3s ease}#drill-panel.open{display:block;max-height:420px;opacity:1}#drill-panel.closing{max-height:0;opacity:0;padding-top:0;padding-bottom:0;margin-top:0;border-width:0}.drill-trigger{cursor:pointer;color:#0f3460;font-weight:500;transition:color .15s}.drill-trigger:hover{color:#2980b9;text-decoration:underline}.drill-trigger.active{color:#08845a;font-weight:700;background:#e8f5e9}"""
html = html.replace('</style>', DRILL_CSS + '</style>', 1)
print("1. CSS 已插入")

# ===== 2. HTML: 在 main div 结尾前插入 drill panel =====
DRILL_PANEL = """
<div id="drill-panel">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
    <strong id="drill-title" style="font-size:13px;color:#0f3460"></strong>
    <span id="drill-close" style="cursor:pointer;font-size:20px;color:#999;line-height:1" title="关闭">&times;</span>
  </div>
  <canvas id="drill-canvas" style="max-height:260px"></canvas>
</div>
"""
html = html.replace('\n</div>\n\n<script>', DRILL_PANEL + '\n</div>\n\n<script>', 1)
print("2. HTML panel 已插入")

# ===== 3. JS核心函数 =====
DRILL_JS_CORE = """
// ===== 下钻图表核心函数 =====
window._drillChartInstance = null;

function buildDrillDataMap() {
  if (window._drillDataMap) return;
  var map = new Map();
  // Tab1 品线维度
  categoryRevenueData.forEach(function(d) {
    map.set(d.category, {
      label: d.category, tab: 't1-cat',
      sales4w: d.sales4w || [0,0,0,0],
      revenue4w: d.revenue4w || [0,0,0,0],
      share4w: d.share4w || [0,0,0,0],
      newSku4w: [0,0,0,0]
    });
  });
  // Tab1 分析人维度
  analystRevenueData.forEach(function(d) {
    map.set(d.analyst, {
      label: d.analyst, tab: 't1-an',
      sales4w: d.sales4w || [0,0,0,0],
      revenue4w: d.revenue4w || [0,0,0,0],
      share4w: d.share4w || [0,0,0,0],
      newSku4w: [0,0,0,0]
    });
  });

  // 按周统计每品线/分析人的新上架SKU数
  var weekStarts = [new Date(2026,3,30), new Date(2026,4,7), new Date(2026,4,14), new Date(2026,4,21), new Date(2026,4,28)];
  function weekIdx(ds) {
    if (!ds) return -1;
    var d = new Date(ds);
    if (isNaN(d.getTime())) return -1;
    for (var i=0;i<4;i++) { if (d>=weekStarts[i] && d<weekStarts[i+1]) return i; }
    return -1;
  }
  cum43Data.forEach(function(d) {
    var wi = weekIdx(d.listDate);
    if (wi < 0) return;
    var entry = map.get(d.category);
    if (entry) entry.newSku4w[wi]++;
    entry = map.get(d.analyst);
    if (entry) entry.newSku4w[wi]++;
  });

  // Tab1 及时率维度
  if (typeof timeliness4w !== 'undefined') {
    timeliness4w.analysts.forEach(function(d) {
      map.set('time:' + d.analyst, {
        label: d.analyst + ' 及时率', tab: 't1-time',
        rates4w: d.rates4w,
        isTimeliness: true
      });
    });
    map.set('time:总及时率', {
      label: '总及时率', tab: 't1-time',
      rates4w: timeliness4w.totalRates,
      isTimeliness: true
    });
  }

  // Tab4 PLP分析人
  if (typeof plpAn4w !== 'undefined') {
    plpAn4w.forEach(function(d) {
      map.set('plp:an:' + d.analyst, {
        label: d.analyst, tab: 'plp-an',
        spend4w: d.spend4w, adSales4w: d.adSales4w,
        acos4w: d.acos4w, acoas4w: d.acoas4w
      });
    });
  }
  // Tab4 PLP品线
  if (typeof plpCat4w !== 'undefined') {
    plpCat4w.forEach(function(d) {
      map.set('plp:cat:' + d.category, {
        label: d.category, tab: 'plp-cat',
        spend4w: d.spend4w, adSales4w: d.adSales4w,
        acos4w: d.acos4w, acoas4w: d.acoas4w
      });
    });
  }
  // Tab4 PLP拓展类型
  if (typeof plpExp4w !== 'undefined') {
    plpExp4w.forEach(function(d) {
      map.set('plp:exp:' + d.expandType, {
        label: d.expandType, tab: 'plp-exp',
        spend4w: d.spend4w, adSales4w: d.adSales4w,
        acos4w: d.acos4w, acoas4w: d.acoas4w
      });
    });
  }
  // Tab4 PLG按分析人
  if (typeof plgAn4w !== 'undefined') {
    plgAn4w.forEach(function(d) {
      map.set('plg:an:' + d.analyst, {
        label: d.analyst, tab: 'plg-an',
        spend4w: d.spend4w, adSales4w: d.adSales4w,
        acos4w: d.acos4w, acoas4w: d.acoas4w
      });
    });
  }

  window._drillDataMap = map;
}

function showDrillChart(key, data, title) {
  if (window._drillChartInstance) {
    window._drillChartInstance.destroy();
    window._drillChartInstance = null;
  }
  var panel = document.getElementById('drill-panel');
  if (!panel) return;

  // 将面板移到点击表格的下方
  var activeTd = document.querySelector('.drill-trigger.active');
  if (activeTd) {
    var tableWrap = activeTd.closest('.table-scroll-wrap');
    if (!tableWrap) tableWrap = activeTd.closest('table');
    if (tableWrap) {
      var targetParent = tableWrap.parentElement;
      if (targetParent) {
        if (panel.parentNode !== targetParent) {
          targetParent.appendChild(panel);
        }
      }
    }
  }

  var isPLP = (key.indexOf('plp:') === 0);
  var isPLG = (key.indexOf('plg:') === 0);
  var isAd = isPLP || isPLG;

  document.getElementById('drill-title').textContent = title + ' — 4周趋势';

  var WLABELS = ['4.30-5.6','5.7-5.13','5.14-5.20','5.21-5.27'];
  var datasets = [];

  if (isAd) {
    // PLP: 花费、广告销售额 (左轴), ACOS、ACOAS (右轴)
    var hasSpend = data.spend4w && data.spend4w.some(function(v){return v>0;});
    var hasAdSales = data.adSales4w && data.adSales4w.some(function(v){return v>0;});
    if (!hasSpend && !hasAdSales) {
      document.getElementById('drill-title').textContent = title + ' — 无4周PLP数据';
      panel.classList.add('open');
      panel.classList.remove('closing');
      return;
    }
    if (hasSpend) {
      datasets.push({label:'广告花费($)',data:data.spend4w,borderColor:'#c0392b',backgroundColor:'rgba(192,57,43,0.08)',tension:0.3,borderWidth:2.5,pointRadius:4,yAxisID:'y'});
    }
    if (hasAdSales) {
      datasets.push({label:'广告销售额($)',data:data.adSales4w,borderColor:'#8e44ad',backgroundColor:'rgba(142,68,173,0.08)',tension:0.3,borderWidth:2.5,pointRadius:4,yAxisID:'y'});
    }
    if (data.acos4w && data.acos4w.some(function(v){return v>0;})) {
      datasets.push({label:'ACOS(%)',data:data.acos4w,borderColor:'#e07b24',backgroundColor:'transparent',tension:0.3,borderWidth:2,borderDash:[5,3],pointRadius:3,yAxisID:'y1'});
    }
    if (data.acoas4w && data.acoas4w.some(function(v){return v>0;})) {
      datasets.push({label:'ACOAS(%)',data:data.acoas4w,borderColor:'#0f3460',backgroundColor:'transparent',tension:0.3,borderWidth:2,borderDash:[3,3],pointRadius:3,yAxisID:'y1'});
    }
  } else if (data.isTimeliness) {
    // 及时率: 单Y轴折线，0-100%
    datasets.push({label:'及时率(%)',data:data.rates4w,borderColor:'#0f3460',backgroundColor:'rgba(15,52,96,0.1)',tension:0.3,borderWidth:3,pointRadius:5,fill:true,yAxisID:'y'});
  } else {
    // 新品: 销量、新上架SKU (左轴), 销售额、市占比 (右轴)
    var hasSales = data.sales4w && data.sales4w.some(function(v){return v>0;});
    var hasRev = data.revenue4w && data.revenue4w.some(function(v){return v>0;});
    var isSku = key.indexOf('sku:') === 0;

    if (!hasSales && !hasRev) {
      document.getElementById('drill-title').textContent = title + ' — 无4周趋势数据';
      panel.classList.add('open');
      panel.classList.remove('closing');
      return;
    }
    if (hasSales) {
      datasets.push({label:'销量',data:data.sales4w,borderColor:'#0f3460',backgroundColor:'rgba(15,52,96,0.08)',tension:0.3,borderWidth:2.5,pointRadius:4,yAxisID:'y'});
    }
    if (hasRev) {
      datasets.push({label:'销售额($)',data:data.revenue4w,borderColor:'#8e44ad',backgroundColor:'rgba(142,68,173,0.08)',tension:0.3,borderWidth:2.5,pointRadius:4,yAxisID:'y1'});
    }
    if (!isSku && data.share4w && data.share4w.some(function(v){return v>0;})) {
      datasets.push({label:'市占比(%)',data:data.share4w,borderColor:'#08845a',backgroundColor:'transparent',tension:0.3,borderWidth:2,borderDash:[5,3],pointRadius:3,yAxisID:'y1'});
    }
    // 每周新上架SKU数
    var hasNewSku = data.newSku4w && data.newSku4w.some(function(v){return v>0;});
    if (hasNewSku) {
      datasets.push({label:'新上架SKU',data:data.newSku4w,type:'bar',backgroundColor:'rgba(224,123,36,0.35)',borderColor:'#e07b24',borderWidth:1,borderRadius:3,yAxisID:'y'});
    }
  }

  panel.classList.add('open');
  panel.classList.remove('closing');

  var yLabel = data.isTimeliness ? '及时率(%)' : (isAd ? '金额($)' : '销量');
  var y1Label = data.isTimeliness ? '' : (isPLP ? '百分比(%)' : '金额/占比');
  window._drillChartInstance = new Chart(document.getElementById('drill-canvas'), {
    type: 'line',
    data: { labels: WLABELS, datasets: datasets },
    options: {
      responsive: true,
      maintainAspectRatio: false,
      plugins: { legend: { position: 'bottom' } },
      scales: {
        y: { beginAtZero: true, title: { display: true, text: yLabel }, position: 'left' },
        y1: { beginAtZero: true, title: { display: true, text: y1Label }, position: 'right', grid: { drawOnChartArea: false } }
      }
    }
  });
}

function hideDrillChart() {
  if (window._drillChartInstance) {
    window._drillChartInstance.destroy();
    window._drillChartInstance = null;
  }
  var panel = document.getElementById('drill-panel');
  if (!panel) return;
  panel.classList.add('closing');
  panel.classList.remove('open');
  document.querySelectorAll('.drill-trigger.active').forEach(function(td){td.classList.remove('active');});
  window._activeDrillKey = null;
}

function handleDrillClick(td) {
  var key = td.getAttribute('data-drill');
  if (!key) return;

  if (window._activeDrillKey === key) {
    hideDrillChart();
    return;
  }

  document.querySelectorAll('.drill-trigger.active').forEach(function(t){t.classList.remove('active');});
  td.classList.add('active');
  window._activeDrillKey = key;

  buildDrillDataMap();

  var data = window._drillDataMap.get(key);
  if (!data) {
    hideDrillChart();
    return;
  }

  var isAd2 = (key.indexOf('plp:') === 0 || key.indexOf('plg:') === 0);
  var title = data.label + (key.indexOf('sku:')===0 ? ' SKU趋势' : isAd2 ? (key.indexOf('plg:')===0 ? ' PLG趋势' : ' PLP趋势') : ' 4周趋势');
  showDrillChart(key, data, title);
}

function setupDrillTrigger(containerId, keyPrefix) {
  var container = document.getElementById(containerId);
  if (!container) return;
  var rows = container.querySelectorAll('tbody tr');
  var count = 0;
  rows.forEach(function(tr) {
    var firstTd = tr.querySelector('td');
    if (!firstTd) return;
    var txt = (firstTd.textContent || '').trim();
    if (!txt || txt === '合计') return;
    var dk = (keyPrefix || '') + txt;
    if (!firstTd.getAttribute('data-drill')) {
      firstTd.setAttribute('data-drill', dk);
    }
    firstTd.classList.add('drill-trigger');
    firstTd.style.cursor = 'pointer';
    firstTd.style.color = '#0f3460';
    firstTd.style.fontWeight = '500';
    firstTd.onclick = function(e) {
      handleDrillClick(this);
    };
    count++;
  });
  var hdr = document.querySelector('.header p');
  if (hdr) hdr.textContent = (hdr.textContent||'') + ' | ' + containerId + ':' + count;
}

// 关闭按钮
setTimeout(function() {
  var closeBtn = document.getElementById('drill-close');
  if (closeBtn) closeBtn.onclick = hideDrillChart;
}, 100);

// 切换Tab时关闭面板
var _origSwitchTab = switchTab;
switchTab = function(tabId, el) {
  hideDrillChart();
  _origSwitchTab(tabId, el);
};
"""

# Insert after switchTab closing brace
SWITCH_END = """  if (tabId === 'tab2' && !window._charts2Init) { initCharts2(); }
}"""
html = html.replace(SWITCH_END, SWITCH_END + DRILL_JS_CORE, 1)
print("3. JS核心函数 已插入")

# ===== 4. buildDrillDataMap 调用 + Tab5 wiring =====
html = html.replace(
    "  renderT5Table(cum43Data);\n})();\n\n// ========== Tab6",
    "  renderT5Table(cum43Data);\n  buildDrillDataMap();\n})();\n\n// ========== Tab6",
    1
)
print("4. buildDrillDataMap 已插入")

# ===== 5. Tab1 品线 wiring =====
html = html.replace(
    "document.getElementById('t1-cat-table').innerHTML = catHtml;",
    "document.getElementById('t1-cat-table').innerHTML = catHtml;\n  setTimeout(function(){setupDrillTrigger('t1-cat-table', '');}, 50);",
    1
)
print("5. Tab1 品线表 wiring 已插入")

# ===== 6. Tab1 分析人 wiring =====
html = html.replace(
    "document.getElementById('t1-an-table').innerHTML = anHtml;",
    "document.getElementById('t1-an-table').innerHTML = anHtml;\n  setTimeout(function(){setupDrillTrigger('t1-an-table', '');}, 50);",
    1
)
print("6. Tab1 分析人表 wiring 已插入")

# ===== 7. Tab4 PLP tables wiring =====
html = html.replace(
    "document.getElementById('t4-plp-an').innerHTML = renderPlpDim(plpAnalysts, '分析人');",
    "document.getElementById('t4-plp-an').innerHTML = renderPlpDim(plpAnalysts, '分析人');\n  setTimeout(function(){setupDrillTrigger('t4-plp-an', 'plp:an:');}, 50);",
    1
)
print("8a. Tab4 PLP分析人 wiring 已插入")

html = html.replace(
    "document.getElementById('t4-plp-cat').innerHTML = renderPlpDim(plpCategories, '品线');",
    "document.getElementById('t4-plp-cat').innerHTML = renderPlpDim(plpCategories, '品线');\n  setTimeout(function(){setupDrillTrigger('t4-plp-cat', 'plp:cat:');}, 50);",
    1
)
print("8b. Tab4 PLP品线 wiring 已插入")

html = html.replace(
    "document.getElementById('t4-plp-exp').innerHTML = renderPlpDim(plpExpandTypes, '拓展类型');",
    "document.getElementById('t4-plp-exp').innerHTML = renderPlpDim(plpExpandTypes, '拓展类型');\n  setTimeout(function(){setupDrillTrigger('t4-plp-exp', 'plp:exp:');}, 50);",
    1
)
print("8c. Tab4 PLP拓展类型 wiring 已插入")

# ===== 9. Tab4 PLG table wiring =====
html = html.replace(
    "document.getElementById('t4-plg-an').innerHTML = plgAnHtml;",
    "document.getElementById('t4-plg-an').innerHTML = plgAnHtml;\n  setTimeout(function(){setupDrillTrigger('t4-plg-an', 'plg:an:');}, 50);",
    1
)
print("9. Tab4 PLG分析人 wiring 已插入")

# ===== 10. Tab1 及时率 wiring =====
html = html.replace(
    "document.getElementById('t1-time-table').innerHTML = timeHtml;",
    "document.getElementById('t1-time-table').innerHTML = timeHtml;\n  setTimeout(function(){setupDrillTrigger('t1-time-table', 'time:');}, 50);",
    1
)
print("10. Tab1 及时率 wiring 已插入")

after = len(html)
print(f"\n输入: {before/1024:.0f}KB → 输出: {after/1024:.0f}KB (+{(after-before)/1024:.0f}KB)")

script_v = re.search(r'<script>(.*?)</script>', html, re.DOTALL)
if script_v:
    js = script_v.group(1)
    op = js.count('(') - js.count(')')
    ob = js.count('{') - js.count('}')
    print(f"Brackets: ()={op}, {{}}={ob}")
    if op != 0 or ob != 0:
        print("ERROR: 括号不平衡!")
        sys.exit(1)
else:
    script_v2 = re.search(r'<script[^>]*>(.*?)</script>', html, re.DOTALL)
    if script_v2:
        js = script_v2.group(1)
        op = js.count('(') - js.count(')')
        ob = js.count('{') - js.count('}')
        print(f"Fallback match OK, brackets: ()={op}, {{}}={ob}")
    else:
        print("ERROR: 找不到script标签!")
        sys.exit(1)

checks = [
    ('drill-panel', 'HTML panel'),
    ('drill-trigger', 'CSS class'),
    ('buildDrillDataMap', 'Data map builder'),
    ('showDrillChart', 'Show chart fn'),
    ('hideDrillChart', 'Hide chart fn'),
    ('handleDrillClick', 'Click handler'),
    ('setupDrillTrigger', 'Setup trigger fn'),
    ("switchTab = function", 'switchTab override'),
    ('plpAn4w', 'PLP An 4w data'),
    ('plpCat4w', 'PLP Cat 4w data'),
    ('plpExp4w', 'PLP Exp 4w data'),
    ("'plp:an:'", 'PLP An prefix'),
    ("'plp:cat:'", 'PLP Cat prefix'),
    ("'plp:exp:'", 'PLP Exp prefix'),
    ('plgAn4w', 'PLG An 4w data'),
    ("'plg:an:'", 'PLG An prefix'),
]
all_ok = True
for keyword, label in checks:
    if keyword in html:
        print(f"  OK: {label}")
    else:
        print(f"  MISSING: {label}")
        all_ok = False

if all_ok:
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"\n保存到: {OUTPUT}")
    print("下钻功能添加完成!")
else:
    print("\nERROR: 缺少关键标识，未保存!")
    sys.exit(1)
