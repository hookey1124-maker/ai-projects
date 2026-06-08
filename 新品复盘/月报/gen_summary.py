import openpyxl, sys
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.styles import Font, PatternFill, Alignment

SRC = r'C:\Users\Hardy\ai-projects\新品复盘\月报\新品月报数据.xlsx'
DST = r'C:\Users\Hardy\ai-projects\新品复盘\月报\数据汇总_5月_v2.xlsx'
wb_src = openpyxl.load_workbook(SRC, data_only=True)
wb = openpyxl.Workbook()

def hdr(ws, row, n):
    for c in range(1, n+1):
        cl = ws.cell(row, c); cl.font = Font(bold=True, color='FFFFFF', size=11)
        cl.fill = PatternFill(start_color='0F3460', end_color='0F3460', fill_type='solid')
        cl.alignment = Alignment(horizontal='center')

def sf(v):
    try: return float(v or 0)
    except: return 0.0

names = wb_src.sheetnames
ws_new = wb_src[names[0]]; ws_plp = wb_src[names[1]]
ws_hist = wb_src[names[2]]; ws_pw = wb_src[names[3]]

rows = [r for r in ws_new.iter_rows(min_row=2, values_only=True) if r[1]]
hist_rows = list(ws_hist.iter_rows(min_row=2, values_only=True))
plp_rows = [r for r in ws_plp.iter_rows(min_row=3, values_only=True) if r[2]]

# PW maps
pw_all = list(ws_pw.iter_rows(min_row=2, values_only=True))
pw_rival = {}; pw_share = {}
for r in pw_all:
    code = str(r[1] or '').strip()
    if code:
        try: pw_rival[code] = float(r[2] or 0)
        except: pw_rival[code] = 0
        try: pw_share[code] = float(r[4] or 0)
        except: pw_share[code] = 0

def has_rival(code): return pw_rival.get(str(code or '').strip(), 0) > 0
def pw_pct(code): return pw_share.get(str(code or '').strip(), 0)

N = len(rows)
s4 = sum(sf(r[14]) for r in rows); s5 = sum(sf(r[15]) for r in rows)
r4 = sum(sf(r[16]) for r in rows); r5 = sum(sf(r[17]) for r in rows)
y5 = sum(1 for r in rows if str(r[11] or '').strip() == 'Y')
n5 = sum(1 for r in rows if str(r[11] or '').strip() == 'N')

def sh1():  # 月度总览
    ws = wb.active; ws.title = '月度总览'
    ws.append(['KPI指标','4月','5月'])
    for k,v4,v5 in [('SKU总数',N,N),('销量',int(s4),int(s5)),('销售额',round(r4),round(r5)),
        ('出单Y',y5,y5),('出单N',n5,n5),('未出单',N-y5-n5,N-y5-n5),
        ('出单率','',f'{(y5+n5)/N*100:.1f}%')]: ws.append([k,v4,v5])
    hdr(ws,1,3)
    for c in ws.columns: ws.column_dimensions[c[0].column_letter].width = 16

def sh2():  # SKU月度明细
    ws = wb.create_sheet('SKU月度明细')
    ws.append(['销售编号','SKU','上架日期','首次出单','4月分析人','5月分析人','品类','拓展类型',
        '4月销量','5月销量','4月销售额','5月销售额','4月对手出单','5月对手出单',
        'PW市占比(5.31)','4月市场状态','5月市场状态','4月8日出单','5月8日出单','4月分析频次','5月分析频次'])
    for r in rows:
        ws.append([r[0],r[1],str(r[2]or'')[:10],str(r[3]or'')[:10],r[4],r[5],r[6],r[7],
            sf(r[14]),sf(r[15]),sf(r[16]),sf(r[17]),sf(r[18]),sf(r[19]),
            f'{pw_pct(r[0])*100:.1f}%',r[12],r[13],r[10],r[11],r[8],r[9]])
    hdr(ws,1,21)
    for c in ws.columns: ws.column_dimensions[c[0].column_letter].width = 14
    ws.freeze_panes = 'A2'

def sh3():  # 销售维度汇总 (上架批次+品线+分析人 合并)
    ws = wb.create_sheet('销售维度汇总')

    def wtab(title, headers, rows_data):
        ws.append([]); ws.append([title])
        r0=ws.max_row+1; ws.append(headers)
        for rd in rows_data: ws.append(rd)
        hdr(ws,r0,len(headers))

    hd=['维度','SKU数','4月销量','5月销量','销量环比','4月销售额','5月销售额','销售额环比','月均$/SKU(5月)']

    # 上架批次
    b=defaultdict(lambda:{'n':0,'s4':0,'s5':0,'r4':0,'r5':0})
    for r in rows:
        ld=str(r[2]or'')[:7]; b[ld]['n']+=1
        b[ld]['s4']+=sf(r[14]);b[ld]['s5']+=sf(r[15]);b[ld]['r4']+=sf(r[16]);b[ld]['r5']+=sf(r[17])
    br=[]; [br.append([k or'无日期',d['n'],int(d['s4']),int(d['s5']),f'{((d["s5"]-d["s4"])/d["s4"]*100)if d["s4"]else 0:+.1f}%',round(d['r4']),round(d['r5']),round(d['r5']/d['n'])if d['n']else 0])for k in sorted(b)for d in[b[k]]]
    wtab('上架批次',hd,br)

    # 品线
    c=defaultdict(lambda:{'n':0,'s4':0,'s5':0,'r4':0,'r5':0})
    for r in rows:
        ct=str(r[6]or'未分类'); c[ct]['n']+=1;c[ct]['s4']+=sf(r[14]);c[ct]['s5']+=sf(r[15]);c[ct]['r4']+=sf(r[16]);c[ct]['r5']+=sf(r[17])
    hd2=['品线','SKU数','4月销量','5月销量','销量环比','4月销售额','5月销售额','销售额环比']
    crs=[]; [crs.append([k,d['n'],int(d['s4']),int(d['s5']),f'{((d["s5"]-d["s4"])/d["s4"]*100)if d["s4"]else 0:+.1f}%',round(d['r4']),round(d['r5']),f'{((d["r5"]-d["r4"])/d["r4"]*100)if d["r4"]else 0:+.1f}%'])for k in sorted(c)for d in[c[k]]]
    wtab('品线维度',hd2,crs)

    # 分析人
    a=defaultdict(lambda:{'n':0,'s4':0,'s5':0,'r4':0,'r5':0})
    for r in rows:
        an=str(r[5]or'未知'); a[an]['n']+=1;a[an]['s4']+=sf(r[14]);a[an]['s5']+=sf(r[15]);a[an]['r4']+=sf(r[16]);a[an]['r5']+=sf(r[17])
    hd3=['分析人','SKU数','4月销量','5月销量','销量环比','4月销售额','5月销售额','销售额环比']
    ars=[]; [ars.append([k,d['n'],int(d['s4']),int(d['s5']),f'{((d["s5"]-d["s4"])/d["s4"]*100)if d["s4"]else 0:+.1f}%',round(d['r4']),round(d['r5']),f'{((d["r5"]-d["r4"])/d["r4"]*100)if d["r4"]else 0:+.1f}%'])for k in sorted(a)for d in[a[k]]]
    wtab('分析人维度',hd3,ars)

    for cl in ws.columns: ws.column_dimensions[cl[0].column_letter].width = 16

def sh6():  # 市场状态分布
    ws = wb.create_sheet('市场状态分布')
    labels = ['正常','竞争无优势','无市场','站外出单']
    m4=defaultdict(int); m5=defaultdict(int)
    for r in rows: m4[str(r[12]or'').strip()]+=1; m5[str(r[13]or'').strip()]+=1
    ws.append(['市场状态','4月SKU数','4月占比','5月SKU数','5月占比','变化'])
    for ml in labels:
        c4=m4.get(ml,0); c5=m5.get(ml,0)
        ws.append([ml,c4,f'{c4/N*100:.1f}%',c5,f'{c5/N*100:.1f}%',f'{c5-c4:+d}'])
    hdr(ws,1,6)
    for cl in ws.columns: ws.column_dimensions[cl[0].column_letter].width = 16

def sh7():  # 出单分析 (四象限+分析人+品线 合并)
    ws = wb.create_sheet('出单分析')
    def wtab(title,hd,rd):
        ws.append([]); ws.append([title]); r0=ws.max_row+1; ws.append(hd)
        for x in rd: ws.append(x); hdr(ws,r0,len(hd))

    def q4(data,kf,df):
        r=defaultdict(lambda:{'t':0,'hy':0,'hn':0,'ny':0,'nn':0})
        for d in data:
            k=df(d); r[k]['t']+=1; hr=kf(d); od=str(d[11]or'').strip() in ('Y','N')
            if hr and od: r[k]['hy']+=1
            elif hr: r[k]['hn']+=1
            elif od: r[k]['ny']+=1
            else: r[k]['nn']+=1
        return r

    hy=[r for r in rows if has_rival(r[0]) and str(r[11]or'').strip() in ('Y','N')]
    hn=[r for r in rows if has_rival(r[0]) and str(r[11]or'').strip() not in ('Y','N')]
    ny=[r for r in rows if not has_rival(r[0]) and str(r[11]or'').strip() in ('Y','N')]
    nn=[r for r in rows if not has_rival(r[0]) and str(r[11]or'').strip() not in ('Y','N')]
    wtab('出单分析(依据5.31PW)',['分类','SKU数','占比'],[
        ['有对手出单',len(hy),f'{len(hy)/N*100:.1f}%'],['有对手未出单',len(hn),f'{len(hn)/N*100:.1f}%'],
        ['无对手出单',len(ny),f'{len(ny)/N*100:.1f}%'],['无对手未出单',len(nn),f'{len(nn)/N*100:.1f}%'],
        ['合计',N,'100%']])

    hd=['维度','总数','有对手出单','有对手未出单','无对手出单','无对手未出单']
    an_q=q4(rows,lambda r:has_rival(r[0]),lambda r:str(r[5]or'未知'))
    wtab('按分析人',hd,[[k,d['t'],d['hy'],d['hn'],d['ny'],d['nn']]for k in sorted(an_q)for d in[an_q[k]]])
    cat_q=q4(rows,lambda r:has_rival(r[0]),lambda r:str(r[6]or'未分类'))
    wtab('按品线',hd,[[k,d['t'],d['hy'],d['hn'],d['ny'],d['nn']]for k in sorted(cat_q)for d in[cat_q[k]]])
    for cl in ws.columns: ws.column_dimensions[cl[0].column_letter].width = 18

def sh8():  # 分析频次 (汇总+分析人合并)
    ws = wb.create_sheet('分析频次')
    f4 = sum(sf(r[8]) for r in rows); f5 = sum(sf(r[9]) for r in rows)
    af = defaultdict(lambda: {'n':0,'f4':0,'f5':0})
    for r in rows:
        an=str(r[5]or'未知'); af[an]['n']+=1; af[an]['f4']+=sf(r[8]); af[an]['f5']+=sf(r[9])
    ws.append(['分析人','SKU数','4月总频次','5月总频次','4月均次/SKU','5月均次/SKU'])
    ws.append(['合计',N,int(f4),int(f5),f'{f4/N:.1f}',f'{f5/N:.1f}'])
    for an in sorted(af):
        d=af[an]; ws.append([an,d['n'],int(d['f4']),int(d['f5']),f'{d["f4"]/d["n"]:.1f}',f'{d["f5"]/d["n"]:.1f}'])
    hdr(ws,1,6)
    for cl in ws.columns: ws.column_dimensions[cl[0].column_letter].width = 16

def sh9():  # 广告汇总 (从近三月新品明细 Col 25-44 取PLP/PLG数据)
    ws = wb.create_sheet('广告汇总')
    # Col indices (0-based): PLP 5月=29(sales)30(rev)31(cost); PLG 5月=39(cost)40(rev)
    # PLP 4月=24(sales)25(rev)26(cost); PLG 4月=34(cost)35(rev)

    plp_cost5=sum(sf(r[31])for r in rows); plp_rev5=sum(sf(r[30])for r in rows); plp_sales5=sum(sf(r[29])for r in rows)
    plp_cost4=sum(sf(r[26])for r in rows); plp_rev4=sum(sf(r[25])for r in rows); plp_sales4=sum(sf(r[24])for r in rows)
    plp_n=sum(1 for r in rows if sf(r[31])>0 or sf(r[30])>0)

    plg_cost5=sum(sf(r[39])for r in rows); plg_rev5=sum(sf(r[40])for r in rows)
    plg_cost4=sum(sf(r[34])for r in rows); plg_rev4=sum(sf(r[35])for r in rows)
    plg_n=sum(1 for r in rows if sf(r[39])>0 or sf(r[40])>0)

    def wtab(title, headers, rows_data):
        ws.append([]); ws.append([title])
        r0=ws.max_row+1; ws.append(headers)
        for rd in rows_data: ws.append(rd)
        hdr(ws,r0,len(headers))

    def acoas(c,total_rev): return f'{c/total_rev*100:.1f}%'if total_rev>0 else'-'

    # PLP汇总: ACOAS = PLP花费 / 总销售额(r5)
    wtab('PLP广告汇总(5月)',['指标','4月','5月'], [
        ['SKU数',plp_n,plp_n],['广告销量',int(plp_sales4),int(plp_sales5)],
        ['广告销售额',round(plp_rev4,2),round(plp_rev5,2)],['广告花费',round(plp_cost4,2),round(plp_cost5,2)],
        ['ACOAS',acoas(plp_cost4,r4),acoas(plp_cost5,r5)]])

    # PLG汇总
    wtab('PLG广告汇总(5月)',['指标','4月','5月'], [
        ['SKU数',plg_n,plg_n],['广告销售额',round(plg_rev4,2),round(plg_rev5,2)],
        ['广告花费',round(plg_cost4,2),round(plg_cost5,2)],
        ['ACOAS',acoas(plg_cost4,r4),acoas(plg_cost5,r5)]])

    # PLP按分析人(5月): ACOAS = 花费 / 该分析人总销售额
    pa=defaultdict(lambda:{'n':0,'s':0,'c':0,'r':0,'tr':0})  # tr = total revenue
    for r in rows:
        an=str(r[5]or'未知');c=sf(r[31]);rv=sf(r[30]);tot=sf(r[17])
        pa[an]['tr']+=tot
        if c>0 or rv>0: pa[an]['n']+=1;pa[an]['c']+=c;pa[an]['r']+=rv;pa[an]['s']+=sf(r[29])
    hd=['维度','SKU数','广告销量','广告销售额','花费','ACOAS']
    anrs=[]; [anrs.append([an,d['n'],int(d['s']),round(d['r'],2),round(d['c'],2),acoas(d['c'],d['tr'])])for an in sorted(pa)for d in[pa[an]]]
    wtab('PLP按分析人(5月)',hd,anrs)

    # PLP按品线(5月)
    pc=defaultdict(lambda:{'n':0,'s':0,'c':0,'r':0,'tr':0})
    for r in rows:
        cat=str(r[6]or'未知');c=sf(r[31]);rv=sf(r[30]);tot=sf(r[17])
        pc[cat]['tr']+=tot
        if c>0 or rv>0: pc[cat]['n']+=1;pc[cat]['c']+=c;pc[cat]['r']+=rv;pc[cat]['s']+=sf(r[29])
    crs=[]; [crs.append([cat,d['n'],int(d['s']),round(d['r'],2),round(d['c'],2),acoas(d['c'],d['tr'])])for cat in sorted(pc)for d in[pc[cat]]]
    wtab('PLP按品线(5月)',hd,crs)

    # PLG按分析人(5月)
    pga=defaultdict(lambda:{'n':0,'c':0,'r':0,'tr':0})
    for r in rows:
        an=str(r[5]or'未知');c=sf(r[39]);rv=sf(r[40]);tot=sf(r[17])
        pga[an]['tr']+=tot
        if c>0 or rv>0: pga[an]['n']+=1;pga[an]['c']+=c;pga[an]['r']+=rv
    hg=['维度','SKU数','广告销售额','花费','ACOAS']
    agr=[]; [agr.append([an,d['n'],round(d['r'],2),round(d['c'],2),acoas(d['c'],d['tr'])])for an in sorted(pga)for d in[pga[an]]]
    wtab('PLG按分析人(5月)',hg,agr)

    # PLG按品线(5月)
    pgc2=defaultdict(lambda:{'n':0,'c':0,'r':0,'tr':0})
    for r in rows:
        cat=str(r[6]or'未知');c=sf(r[39]);rv=sf(r[40]);tot=sf(r[17])
        pgc2[cat]['tr']+=tot
        if c>0 or rv>0: pgc2[cat]['n']+=1;pgc2[cat]['c']+=c;pgc2[cat]['r']+=rv
    cgr=[]; [cgr.append([cat,d['n'],round(d['r'],2),round(d['c'],2),acoas(d['c'],d['tr'])])for cat in sorted(pgc2)for d in[pgc2[cat]]]
    wtab('PLG按品线(5月)',hg,cgr)

    # 广告构成
    both=sum(1 for r in rows if(sf(r[31])>0 or sf(r[30])>0)and(sf(r[39])>0 or sf(r[40])>0))
    po=sum(1 for r in rows if(sf(r[31])>0 or sf(r[30])>0)and not(sf(r[39])>0 or sf(r[40])>0))
    go=sum(1 for r in rows if not(sf(r[31])>0 or sf(r[30])>0)and(sf(r[39])>0 or sf(r[40])>0))
    na=N-both-po-go
    wtab('广告构成',['广告分类','SKU数','占比'], [
        ['PLP+PLG同开',both,f'{both/N*100:.1f}%'],['单PLP',po,f'{po/N*100:.1f}%'],
        ['仅PLG',go,f'{go/N*100:.1f}%'],['无广告',na,f'{na/N*100:.1f}%']])

    for cl in ws.columns: ws.column_dimensions[cl[0].column_letter].width = 18

def sh10():  # 低占比新品 (PW市占)
    ws = wb.create_sheet('低占比新品')
    ws.append(['SKU','上架日期','分析人','品类','5月销量','5月销售额','PW对手销量','PW市占比(5.31)','8日出单','市场状态'])
    cnt = 0
    for r in rows:
        pp = pw_pct(r[0]); rv = pw_rival.get(str(r[0]or'').strip(),0)
        if pp < 0.75 and rv > 0:
            ws.append([r[1],str(r[2]or'')[:10],r[5],r[6],sf(r[15]),round(sf(r[17]),2),int(rv),f'{pp*100:.1f}%',r[11],r[13]]); cnt += 1
    hdr(ws,1,10)
    for cl in ws.columns: ws.column_dimensions[cl[0].column_letter].width = 14
    ws.freeze_panes = 'A2'; print(f'低占比新品: {cnt} SKUs')

def sh11():  # PLP广告明细
    ws = wb.create_sheet('PLP广告明细')
    ws.append(['广告活动','SKU','ID','分析人','品类','曝光','点击','售出','花费','广告销售额','总销售额','ROAS','ACOS','ACOAS','是否PLG'])
    for r in plp_rows:
        ws.append([r[1],r[2],r[3],r[8],r[9],sf(r[11]),sf(r[12]),sf(r[13]),round(sf(r[14]),2),round(sf(r[15]),2),round(sf(r[16]),2),
            f'{sf(r[17]):.2f}'if sf(r[17])else'-',f'{sf(r[22])*100:.1f}%'if sf(r[22])else'-',f'{sf(r[23])*100:.1f}%'if sf(r[23])else'-',str(r[24]or'').strip()])
    hdr(ws,1,15)
    for cl in ws.columns: ws.column_dimensions[cl[0].column_letter].width = 14
    ws.freeze_panes = 'A2'

def sh12():  # 品效Cohort
    ws = wb.create_sheet('品效Cohort')
    hm = defaultdict(lambda: {'n':0,'s':[0]*5,'r':[0]*5})
    for r in hist_rows:
        ld = str(r[2]or'')[:7]
        if not ld or '-' not in ld: continue
        m = int(ld.split('-')[1])
        if m<1 or m>5: continue
        hm[m]['n']+=1
        for mi in range(5): hm[m]['s'][mi]+=sf(r[8+mi]); hm[m]['r'][mi]+=sf(r[13+mi])
    ws.append(['上架月份','SKU数']+[f'{i+1}月销售额'for i in range(5)]+[f'{i+1}月品效$/SKU'for i in range(5)]+['趋势'])
    for m in sorted(hm):
        d=hm[m]; rd=[f'{m}月上架',d['n']]; rd+=[round(d['r'][i])for i in range(5)]
        px=[round(d['r'][i]/d['n'])if d['n']else 0 for i in range(5)]; rd+=px
        nz=[x for x in px if x>0][-3:]
        trend='攀升'if len(nz)>=2 and nz[-1]>nz[0]else('衰减'if len(nz)>=2 and nz[-1]<nz[0]else'平稳')
        rd.append(trend); ws.append(rd)
    hdr(ws,1,13)
    for cl in ws.columns: ws.column_dimensions[cl[0].column_letter].width = 15

def sh13():  # 5.31PW
    ws = wb.create_sheet('5.31PW快照')
    ws.append(['销售编号','近7日对手销量','近7日自店销量','近7日自店市占比'])
    for r in pw_all: ws.append([r[1],sf(r[2]),sf(r[3]),f'{sf(r[4])*100:.1f}%'if r[4]is not None else'-'])
    hdr(ws,1,4)
    for cl in ws.columns: ws.column_dimensions[cl[0].column_letter].width = 18
    ws.freeze_panes = 'A2'

# Run all
for fn in [sh1,sh2,sh3,sh6,sh7,sh8,sh9,sh10,sh11,sh12,sh13]: fn()

wb.save(DST)
print(f'Done: {DST}')
print(f'Sheets ({len(wb.sheetnames)}): {wb.sheetnames}')
