import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from datetime import date, datetime
from collections import Counter

wb = load_workbook('C:/Users/Hardy/ai-projects/新品复盘/周报/新品检查周源数据和PLP数据.xlsx', data_only=True)
ws = wb['四三数据累计']

def get_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

def num(v, d=0):
    try: return float(v) if v else d
    except: return d

cutoff = date(2026, 5, 27)
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1] and get_date(row[2]) and get_date(row[2]) <= cutoff:
        rows.append(list(row))

# Check all unique market statuses
mkt_counter = Counter()
for r in rows:
    v = str(r[117] or '').strip()
    mkt_counter[v] += 1

print('=== Market Status Distribution (5.27) ===')
for k, v in sorted(mkt_counter.items(), key=lambda x: -x[1]):
    print(f'  [{k}]: {v} SKUs')

print('\n=== SKUs with [unknown] status ===')
for r in rows:
    v = str(r[117] or '').strip()
    if v == '未知':
        sku = r[1]
        analyst = r[4]
        cat = r[5]
        rival = num(r[44])
        sales = num(r[18])
        print(f'  {sku} | analyst={analyst} | cat={cat} | rival={rival} | sales={sales}')

print('\n=== SKUs with [other] status ===')
for r in rows:
    v = str(r[117] or '').strip()
    if v == '其他':
        sku = r[1]
        analyst = r[4]
        cat = r[5]
        rival = num(r[44])
        sales = num(r[18])
        print(f'  {sku} | analyst={analyst} | cat={cat} | rival={rival} | sales={sales}')

# Check if there are any other unusual statuses
print('\n=== ALL unique statuses ===')
for k, v in sorted(mkt_counter.items(), key=lambda x: -x[1]):
    print(f'  {k}: {v}')
