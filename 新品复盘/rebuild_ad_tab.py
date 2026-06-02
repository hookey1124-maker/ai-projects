"""
重构广告追踪Tab — 替换为统一的广告结构板块
新增：部门销售占比 + PW市占对比 + PLG费率四档 + PLP链接开启PLG
"""
import re, sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from collections import defaultdict

INPUT  = 'c:/Users/Hardy/ai-projects/新品复盘/新品板块_4.30-5.27_4weeks_drill.html'
OUTPUT = 'c:/Users/Hardy/ai-projects/新品复盘/新品板块_4.30-5.27_4weeks_drill_v2.html'
SRC    = 'c:/Users/Hardy/ai-projects/新品复盘/周报/新品检查周源数据和PLP数据（新）xlsx.xlsx'

# ============================================================
# 0. 读取数据
# ============================================================
print("读取数据...")
with open(INPUT, 'r', encoding='utf-8') as f:
    html = f.read()

wb = load_workbook(SRC, data_only=True)
ws_main = wb['四三数据累计']
ws_sales = wb['四三销售数据']
ws_pw = wb['PW表']
ws_plp_detail = wb['PLP明细']

def safe_f(v, d=0):
    try: return float(v) if v else d
    except: return d

def safe_str(v):
    return str(v or '').strip()

# ===== 0a. 提取主表 SKU 数据 =====
print("0a. 提取主表SKU数据...")
sku_rows = []
for row in ws_main.iter_rows(min_row=2, values_only=True):
    sku = safe_str(row[1])
    if not sku: continue
    sku_rows.append({
        'sku': sku,
        'analyst': safe_str(row[4]) or '未知',
        'category': safe_str(row[5]) or '未分类',
        'expandType': safe_str(row[6]) or '其他',
        'salesCurr': safe_f(row[18]),       # 5.21-5.27 销量
        'revenueCurr': safe_f(row[31]),     # 5.21-5.27 销售额
        'rivalCurr': safe_f(row[44]),       # 5.27 对手销量
        'shareCurr': safe_f(row[56]),       # 5.27 市占比 (0-1)
        'plpEnabled': safe_str(row[136]),   # 5.21-5.27 开启PLP
        'plgFee': safe_f(row[142]),         # 5.21-5.27 PLG最高费率 (0-1)
    })

# 新品总销量/总销售额
new_sales = sum(r['salesCurr'] for r in sku_rows)
new_revenue = sum(r['revenueCurr'] for r in sku_rows)
print(f"  新品SKU: {len(sku_rows)}, 销量: {new_sales}, 销售额: ${new_revenue:,.2f}")

# ===== 0b. 部门总销售额（四三销售数据 Row 2）=====
print("0b. 提取部门总数据...")
row2 = list(ws_sales.iter_rows(min_row=2, max_row=2, values_only=True))[0]
# Row 2: 统计 | 统计 | 总销量 | 总销售额 | 新品销售额? | 占比?
dept_sales = safe_f(row2[2])      # 3877
dept_revenue = safe_f(row2[3])    # 275290.47
print(f"  部门总销量: {dept_sales}, 总销售额: ${dept_revenue:,.2f}")

# 新品占部门比
new_sales_pct = round(new_sales / dept_sales * 100, 1) if dept_sales else 0
new_rev_pct = round(new_revenue / dept_revenue * 100, 1) if dept_revenue else 0
print(f"  新品销量占部门: {new_sales_pct}%, 销售额占部门: {new_rev_pct}%")

# ===== 0c. PW表数据 =====
print("0c. 提取PW数据...")
pw_rows = []
pw_our_share_sum = 0.0
pw_our_share_count = 0
pw_total_sales_sum = 0.0
pw_weighted_share_sum = 0.0
pw_weighted_total = 0.0

for row in ws_pw.iter_rows(min_row=4, values_only=True):  # Row 1-3 are headers
    code = safe_str(row[0])
    if not code: continue
    rival_sales = safe_f(row[1])
    rival_share_str = safe_str(row[2])
    total_sales = safe_f(row[3])

    # Use Col C (competitor share) as primary metric
    # It's more reliable than Col B (competitor sales) which has different time window
    our_share = None
    if rival_share_str and rival_share_str != '市场无出单':
        try:
            rival_share = float(rival_share_str)
            our_share = round((1 - rival_share) * 100, 1)
        except:
            pass
    elif rival_share_str == '市场无出单':
        our_share = 100.0  # No market = no competitor, 100% ours
    else:
        # Fallback: if total > rival, use direct calculation
        if total_sales > rival_sales and total_sales > 0:
            our_share = round((total_sales - rival_sales) / total_sales * 100, 1)

    if our_share is not None:
        pw_our_share_sum += our_share
        pw_our_share_count += 1
        pw_total_sales_sum += total_sales
        # Weighted by total sales
        pw_weighted_share_sum += our_share * total_sales
        pw_weighted_total += total_sales

    pw_rows.append({
        'code': code,
        'rivalSales': rival_sales,
        'rivalShare': rival_share_str if rival_share_str != '市场无出单' else '市场无出单',
        'totalSales': total_sales,
        'ourShare': our_share,
    })

# PW aggregate metrics
pw_avg_share = round(pw_our_share_sum / pw_our_share_count, 1) if pw_our_share_count > 0 else 0
pw_weighted_share = round(pw_weighted_share_sum / pw_weighted_total, 1) if pw_weighted_total > 0 else 0

# 新品维度总市占
sku_with_rival = [r for r in sku_rows if r['rivalCurr'] > 0]
w_new_sales = sum(r['salesCurr'] for r in sku_with_rival)
w_rival_sales = sum(r['rivalCurr'] for r in sku_with_rival)
w_share = round(w_new_sales / (w_new_sales + w_rival_sales) * 100, 1) if (w_new_sales + w_rival_sales) > 0 else 0

print(f"  PW: {pw_our_share_count}有效链接, 加权市占={pw_weighted_share}% (算术平均={pw_avg_share}%)")
print(f"  新品: 有对手SKU={len(sku_with_rival)}, 加权市占={w_share}%")

# ===== 0d. PLG费率四档 + 广告构成 + PLP开启PLG =====
print("0d. 计算广告构成和PLG费率分档...")

# PLG费率四档 (百分数)
def plg_tier(rate_pct):
    """rate_pct is percentage (0-100 scale), e.g. 2.5 means 2.5%"""
    if rate_pct == 0: return '无广告'
    if rate_pct <= 2: return '低费率'
    if rate_pct <= 4: return '中费率'
    return '高费率'

# 广告构成 per SKU
ad_comp = defaultdict(int)  # key: composition label
ad_comp_by_an = defaultdict(lambda: defaultdict(int))
ad_comp_by_cat = defaultdict(lambda: defaultdict(int))

# PLG费率分布 per SKU
plg_tier_dist = defaultdict(int)
plg_tier_by_an = defaultdict(lambda: defaultdict(int))

# PLP链接开启PLG: aggregate from PLP明细 (ID level)
# PLP明细: col4=是否开启PLG (Y/N), col2=SKU
plp_plg_by_sku = defaultdict(lambda: {'total': 0, 'plgY': 0})
for row in ws_plp_detail.iter_rows(min_row=2, values_only=True):
    sku = safe_str(row[2])
    plg_flag = safe_str(row[4])
    if not sku: continue
    plp_plg_by_sku[sku]['total'] += 1
    if plg_flag == 'Y':
        plp_plg_by_sku[sku]['plgY'] += 1

# Now per SKU stats
for r in sku_rows:
    sku = r['sku']
    plp = r['plpEnabled']
    fee_pct = round(r['plgFee'] * 100, 2)  # convert 0-1 to percentage

    # PLP status
    has_plp = (plp == 'Y' or plp == '是')

    # PLG status: has PLG if fee > 0
    has_plg = (fee_pct > 0)

    # Ad composition label
    if has_plp and has_plg:
        comp = 'PLP+PLG'
    elif has_plp and not has_plg:
        comp = '单PLP'
    elif not has_plp and has_plg:
        comp = '仅PLG'
    else:
        comp = '无广告'

    ad_comp[comp] += 1
    ad_comp_by_an[r['analyst']][comp] += 1
    ad_comp_by_cat[r['category']][comp] += 1

    # PLG tier
    tier = plg_tier(fee_pct)
    plg_tier_dist[tier] += 1
    plg_tier_by_an[r['analyst']][tier] += 1

    # Store tier on row
    r['plgTier'] = tier
    r['plgFeePct'] = fee_pct
    r['hasPlp'] = has_plp
    r['hasPlg'] = has_plg
    r['adComp'] = comp

# PLP链接开启PLG 统计 (ID维度)
plp_link_plg_y = sum(1 for v in plp_plg_by_sku.values() if v['plgY'] > 0)
plp_link_plg_n = sum(1 for v in plp_plg_by_sku.values() if v['total'] > 0 and v['plgY'] == 0)
plp_link_total = len(plp_plg_by_sku)

print(f"  广告构成: {dict(ad_comp)}")
print(f"  PLG费率分档: {dict(plg_tier_dist)}")
print(f"  PLP链接: {plp_link_total} SKU, 开启PLG: {plp_link_plg_y}, 未开启: {plp_link_plg_n}")

# ===== 0e. 分析人维度详细数据 =====
analysts_order = ['俞东旭', '张潇', '朱培源', '王偲涵', '章鹏', '胡煜星']
categories_order = ['车门系统', '车身外扩件', '挡泥板', '机盖及附件', '其他', '饰条', '牌照板支架']

# 按分析人 - 广告构成 + PLG费率
an_ad_detail = []
for an in analysts_order:
    comps = ad_comp_by_an.get(an, {})
    tiers = plg_tier_by_an.get(an, {})
    total = sum(comps.values())
    an_ad_detail.append({
        'analyst': an,
        'total': total,
        'plpPlg': comps.get('PLP+PLG', 0),
        'plpOnly': comps.get('单PLP', 0),
        'plgOnly': comps.get('仅PLG', 0),
        'noAd': comps.get('无广告', 0),
        'tierNone': tiers.get('无广告', 0),
        'tierLow': tiers.get('低费率', 0),
        'tierMid': tiers.get('中费率', 0),
        'tierHigh': tiers.get('高费率', 0),
    })

# 按品线 - 广告构成
cat_ad_detail = []
for cat in categories_order:
    comps = ad_comp_by_cat.get(cat, {})
    total = sum(comps.values())
    cat_ad_detail.append({
        'category': cat,
        'total': total,
        'plpPlg': comps.get('PLP+PLG', 0),
        'plpOnly': comps.get('单PLP', 0),
        'plgOnly': comps.get('仅PLG', 0),
        'noAd': comps.get('无广告', 0),
    })

# ===== 1. 构建新 JS 数据块 =====
print("\n1. 构建新数据块...")

js_data = """\
// ===== 广告结构数据（v2 重构）=====
var adDeptPct = __AD_DEPT_PCT__;
var adPwVsNew = __AD_PW_VS_NEW__;
var adCompDist = __AD_COMP_DIST__;
var adPlgTierDist = __AD_PLG_TIER__;
var adPlpPlgLink = __AD_PLP_PLG_LINK__;
var adAnDetail = __AD_AN_DETAIL__;
var adCatDetail = __AD_CAT_DETAIL__;
var adCompLabels = __AD_COMP_LABELS__;
var adPlgTierLabels = __AD_PLG_TIER_LABELS__;
"""

js_data = js_data.replace('__AD_DEPT_PCT__', json.dumps({
    'salesPct': new_sales_pct, 'revPct': new_rev_pct,
    'newSales': new_sales, 'newRevenue': round(new_revenue, 2),
    'deptSales': dept_sales, 'deptRevenue': round(dept_revenue, 2)
}, ensure_ascii=False))
js_data = js_data.replace('__AD_PW_VS_NEW__', json.dumps({
    'pwShare': pw_weighted_share, 'newShare': w_share,
    'pwTotalLinks': pw_our_share_count, 'pwAvgShare': pw_avg_share,
    'newTotalSales': w_new_sales, 'newRivalSales': w_rival_sales,
    'pwLinks': len(pw_rows), 'newSkuCount': len(sku_with_rival)
}, ensure_ascii=False))
js_data = js_data.replace('__AD_COMP_DIST__', json.dumps(dict(ad_comp), ensure_ascii=False))
js_data = js_data.replace('__AD_PLG_TIER__', json.dumps(dict(plg_tier_dist), ensure_ascii=False))
js_data = js_data.replace('__AD_PLP_PLG_LINK__', json.dumps({
    'totalSku': plp_link_total, 'plgY': plp_link_plg_y, 'plgN': plp_link_plg_n
}, ensure_ascii=False))
js_data = js_data.replace('__AD_AN_DETAIL__', json.dumps(an_ad_detail, ensure_ascii=False))
js_data = js_data.replace('__AD_CAT_DETAIL__', json.dumps(cat_ad_detail, ensure_ascii=False))
js_data = js_data.replace('__AD_COMP_LABELS__', json.dumps(['PLP+PLG','单PLP','仅PLG','无广告'], ensure_ascii=False))
js_data = js_data.replace('__AD_PLG_TIER_LABELS__', json.dumps(['无广告','低费率','中费率','高费率'], ensure_ascii=False))

# ===== 2. 构建新 Tab4 HTML body =====
print("2. 构建新Tab4 HTML...")

new_tab4_html = '''  <!-- Tab4: 广告结构 -->
  <div class="tab-content" id="tab4">
    <div class="section"><h3>&#128200; 部门占比 &amp; 市占对比</h3>
      <div class="kpi-grid" id="t4-dept-kpi"></div>
    </div>
    <div class="section"><h3>&#127976; PW爬虫市占 vs 新品市占</h3>
      <div class="kpi-grid" id="t4-pw-kpi"></div>
    </div>
    <div class="section"><h3>&#128176; 广告构成分布</h3>
      <div class="kpi-grid" id="t4-comp-kpi"></div>
      <div class="sub-module"><h4>按分析人</h4><div id="t4-comp-an"></div></div>
      <div class="sub-module"><h4>按品线</h4><div id="t4-comp-cat"></div></div>
    </div>
    <div class="section"><h3>&#127991; PLG费率分档（四档）</h3>
      <div class="kpi-grid" id="t4-plg-tier-kpi"></div>
      <div class="sub-module"><h4>按分析人</h4><div id="t4-plg-tier-an"></div></div>
    </div>
    <div class="section"><h3>&#128279; PLP链接开启PLG（ID维度）</h3>
      <div class="kpi-grid" id="t4-plp-plg-link"></div>
    </div>
  </div>'''

# ===== 3. 构建新 Tab4 JS 渲染代码 =====
print("3. 构建新Tab4 JS...")

diff_pct = round(abs(pw_weighted_share - w_share), 1)

new_tab4_js = '''// ========== Tab4: 广告结构（v2 重构）==========
(function() {
  var d = adDeptPct;
  var p = adPwVsNew;
  var c = adCompDist;
  var t = adPlgTierDist;
  var l = adPlpPlgLink;

  // 部门占比 KPI
  document.getElementById('t4-dept-kpi').innerHTML =
    '<div class="kpi-card primary"><div class="label">新品销量占部门比</div><div class="val">__NS_PCT__%</div><div class="hb">新品' + d.newSales + ' / 部门' + d.deptSales + '</div></div>' +
    '<div class="kpi-card info"><div class="label">新品销售额占部门比</div><div class="val">__NR_PCT__%</div><div class="hb">新品$' + fmtMoney(d.newRevenue) + ' / 部门$' + fmtMoney(d.deptRevenue) + '</div></div>';

  // PW vs 新品市占 KPI
  document.getElementById('t4-pw-kpi').innerHTML =
    '<div class="kpi-card purple"><div class="label">PW爬虫市占</div><div class="val">__PW_SHARE__%</div><div class="hb">' + p.pwTotalLinks + '个有效链接 | 加权均价（算术均值' + p.pwAvgShare + '%）</div></div>' +
    '<div class="kpi-card success"><div class="label">新品加权市占</div><div class="val">__NEW_SHARE__%</div><div class="hb">__NSKU__个有对手SKU | 我方__NSALES__ / 对手__RSALES__</div></div>' +
    '<div class="kpi-card info"><div class="label">差值</div><div class="val">__DIFF__%</div><div class="hb">PW vs 新品（维度不同，平行参考）</div></div>';

  // 广告构成 KPI
  var compColors = {'PLP+PLG':'#c0392b','单PLP':'#2980b9','仅PLG':'#e67e22','无广告':'#95a5a6'};
  var compHtml = '';
  adCompLabels.forEach(function(k) {
    compHtml += '<div class="kpi-card" style="border-left:4px solid ' + (compColors[k]||'#999') + '"><div class="label">' + k + '</div><div class="val">' + (c[k]||0) + '</div></div>';
  });
  document.getElementById('t4-comp-kpi').innerHTML = compHtml;

  // 广告构成按分析人
  var anHtml = '<table class="data-table"><thead><tr><th>分析人</th><th>总数</th><th>PLP+PLG</th><th>单PLP</th><th>仅PLG</th><th>无广告</th></tr></thead><tbody>';
  var anTotals = {};
  adAnDetail.forEach(function(d) {
    anHtml += '<tr><td>' + d.analyst + '</td><td>' + d.total + '</td>';
    anHtml += '<td style="color:#c0392b;font-weight:600">' + d.plpPlg + '</td>';
    anHtml += '<td>' + d.plpOnly + '</td><td>' + d.plgOnly + '</td><td>' + d.noAd + '</td></tr>';
    for (var k in d) { if (k !== 'analyst') anTotals[k] = (anTotals[k]||0) + d[k]; }
  });
  anHtml += '<tfoot><tr class="total-row"><td><b>合计</b></td><td><b>' + (anTotals.total||0) + '</b></td><td><b>' + (anTotals.plpPlg||0) + '</b></td><td><b>' + (anTotals.plpOnly||0) + '</b></td><td><b>' + (anTotals.plgOnly||0) + '</b></td><td><b>' + (anTotals.noAd||0) + '</b></td></tr></tfoot></tbody></table>';
  document.getElementById('t4-comp-an').innerHTML = anHtml;

  // 广告构成按品线
  var catHtml = '<table class="data-table"><thead><tr><th>品线</th><th>总数</th><th>PLP+PLG</th><th>单PLP</th><th>仅PLG</th><th>无广告</th></tr></thead><tbody>';
  var catTotals = {};
  adCatDetail.forEach(function(d) {
    catHtml += '<tr><td>' + d.category + '</td><td>' + d.total + '</td>';
    catHtml += '<td style="color:#c0392b;font-weight:600">' + d.plpPlg + '</td>';
    catHtml += '<td>' + d.plpOnly + '</td><td>' + d.plgOnly + '</td><td>' + d.noAd + '</td></tr>';
    for (var k in d) { if (k !== 'category') catTotals[k] = (catTotals[k]||0) + d[k]; }
  });
  catHtml += '<tfoot><tr class="total-row"><td><b>合计</b></td><td><b>' + (catTotals.total||0) + '</b></td><td><b>' + (catTotals.plpPlg||0) + '</b></td><td><b>' + (catTotals.plpOnly||0) + '</b></td><td><b>' + (catTotals.plgOnly||0) + '</b></td><td><b>' + (catTotals.noAd||0) + '</b></td></tr></tfoot></tbody></table>';
  document.getElementById('t4-comp-cat').innerHTML = catHtml;

  // PLG费率分档 KPI
  var tierColors = {'高费率':'#c0392b','中费率':'#e67e22','低费率':'#27ae60','无广告':'#95a5a6'};
  var tierHtml = '';
  adPlgTierLabels.forEach(function(k) {
    tierHtml += '<div class="kpi-card" style="border-left:4px solid ' + (tierColors[k]||'#999') + '"><div class="label">' + k + '</div><div class="val">' + (t[k]||0) + '</div></div>';
  });
  document.getElementById('t4-plg-tier-kpi').innerHTML = tierHtml;

  // PLG费率分档按分析人
  var tierAnHtml = '<table class="data-table"><thead><tr><th>分析人</th><th>总数</th><th>无广告</th><th>低费率(&le;2%)</th><th>中费率(2-4%)</th><th>高费率(>4%)</th></tr></thead><tbody>';
  var tierTotals = {};
  adAnDetail.forEach(function(d) {
    tierAnHtml += '<tr><td>' + d.analyst + '</td><td>' + d.total + '</td>';
    tierAnHtml += '<td>' + d.tierNone + '</td><td>' + d.tierLow + '</td><td>' + d.tierMid + '</td>';
    tierAnHtml += '<td style="color:#c0392b;font-weight:600">' + d.tierHigh + '</td></tr>';
    for (var k in d) { if (k !== 'analyst') tierTotals[k] = (tierTotals[k]||0) + d[k]; }
  });
  tierAnHtml += '<tfoot><tr class="total-row"><td><b>合计</b></td><td><b>' + (tierTotals.total||0) + '</b></td><td><b>' + (tierTotals.tierNone||0) + '</b></td><td><b>' + (tierTotals.tierLow||0) + '</b></td><td><b>' + (tierTotals.tierMid||0) + '</b></td><td><b>' + (tierTotals.tierHigh||0) + '</b></td></tr></tfoot></tbody></table>';
  document.getElementById('t4-plg-tier-an').innerHTML = tierAnHtml;

  // PLP链接开启PLG (ID维度)
  document.getElementById('t4-plp-plg-link').innerHTML =
    '<div class="kpi-card info"><div class="label">SKU总数</div><div class="val">' + l.totalSku + '</div></div>' +
    '<div class="kpi-card success"><div class="label">开启PLG</div><div class="val">' + l.plgY + '</div></div>' +
    '<div class="kpi-card warning"><div class="label">未开启PLG</div><div class="val">' + l.plgN + '</div></div>' +
    '<div class="kpi-card primary"><div class="label">PLG开启率</div><div class="val">' + (l.totalSku ? (l.plgY/l.totalSku*100).toFixed(1) + '%' : '-') + '</div></div>';
})();
'''

# Substitute Python values into JS template
new_tab4_js = new_tab4_js.replace('__NS_PCT__', str(new_sales_pct))
new_tab4_js = new_tab4_js.replace('__NR_PCT__', str(new_rev_pct))
new_tab4_js = new_tab4_js.replace('__PW_SHARE__', str(pw_weighted_share))
new_tab4_js = new_tab4_js.replace('__NEW_SHARE__', str(w_share))
new_tab4_js = new_tab4_js.replace('__NSKU__', str(len(sku_with_rival)))
new_tab4_js = new_tab4_js.replace('__NSALES__', str(w_new_sales))
new_tab4_js = new_tab4_js.replace('__RSALES__', str(w_rival_sales))
new_tab4_js = new_tab4_js.replace('__DIFF__', str(diff_pct))

# ===== 4. 替换 HTML body 中的 Tab4 =====
print("4. 执行HTML body替换...")

# Find and replace the Tab4 HTML section
old_tab4_start = html.find('<!-- Tab4: 广告追踪 -->')
old_tab4_end = html.find('<!-- Tab5:', old_tab4_start)

if old_tab4_start == -1 or old_tab4_end == -1:
    print("ERROR: Cannot find Tab4 HTML boundaries")
    sys.exit(1)

html = html[:old_tab4_start] + new_tab4_html + '\n\n' + html[old_tab4_end:]
print(f"  Tab4 HTML replaced ({old_tab4_end - old_tab4_start} -> {len(new_tab4_html)} chars)")

# ===== 5. 替换 JS 中的 Tab4 逻辑 =====
print("5. 执行JS替换...")

# Find the Tab4 JS section
old_js_start = html.find('// ========== Tab4: 广告追踪')
old_js_end = html.find('// ========== Tab5:', old_js_start)

if old_js_start == -1 or old_js_end == -1:
    print("ERROR: Cannot find Tab4 JS boundaries")
    sys.exit(1)

html = html[:old_js_start] + new_tab4_js + '\n\n' + html[old_js_end:]
print(f"  Tab4 JS replaced ({old_js_end - old_js_start} -> {len(new_tab4_js)} chars)")

# Find the JS data insertion point — just before "// ========== Tab1:"
# We need to insert our new ad data blocks
data_insert = html.find('// ========== Tab1:')
if data_insert == -1:
    print("ERROR: Cannot find Tab1 JS start")
    sys.exit(1)

html = html[:data_insert] + js_data + '\n' + html[data_insert:]
print(f"  New ad data blocks inserted before Tab1 JS")

# ===== 6. 更新汇报输出（移除PLP/PLG引用）=====
print("6. 更新汇报输出...")

# Update tab labels in nav
html = html.replace(
    '<li><a href="javascript:void(0)" onclick="switchTab(\'tab4\',this)">&#128176; 广告追踪</a></li>',
    '<li><a href="javascript:void(0)" onclick="switchTab(\'tab4\',this)">&#128176; 广告结构</a></li>'
)

# Update subtitle
html = html.replace(
    '含PLP/PLG广告追踪（自然周5.18-5.24）',
    '含广告结构分析 + PW市占对比'
)

# ===== 7. 写入输出 =====
print("\n7. 写入输出...")
with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write(html)

# ===== 8. 验证 =====
print("8. 验证...")
# Check braces balance
open_braces = html.count('{')
close_braces = html.count('}')
open_parens = html.count('(')
close_parens = html.count(')')
print(f"  Braces: {{ {open_braces}, }} {close_braces} {'OK' if open_braces == close_braces else 'MISMATCH!'}")
print(f"  Parens: ( {open_parens}, ) {close_parens} {'OK' if open_parens == close_parens else 'MISMATCH!'}")

# Check key elements exist
checks = ['t4-dept-kpi', 't4-pw-kpi', 't4-comp-kpi', 't4-plg-tier-kpi', 't4-plp-plg-link',
          't4-comp-an', 't4-comp-cat', 't4-plg-tier-an', 'adDeptPct', 'adPwVsNew',
          'adCompDist', 'adPlgTierDist', 'adPlpPlgLink']
for c in checks:
    assert c in html, f"MISSING: {c}"
print(f"  All {len(checks)} key identifiers present ✓")

in_size = len(open(INPUT, 'r', encoding='utf-8').read())
out_size = len(html)
print(f"\n✅ 完成! {INPUT.split('/')[-1]} ({in_size//1024}KB) -> {OUTPUT.split('/')[-1]} ({out_size//1024}KB)")
print(f"   变化: {out_size - in_size:+d} bytes")
