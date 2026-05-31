"""
4周升级补丁：在现有2周HTML基础上注入额外2周数据 + 升级图表为4周趋势
"""
import re, json, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from datetime import date, datetime
from collections import defaultdict

HTML = 'C:/Users/Hardy/ai-projects/新品复盘/新品板块_5.21-5.27.html'
OUTPUT = 'C:/Users/Hardy/ai-projects/新品复盘/新品板块_4.30-5.27_4weeks.html'
SRC = 'C:/Users/Hardy/ai-projects/新品复盘/周报/新品检查周源数据和PLP数据.xlsx'

with open(HTML, 'r', encoding='utf-8') as f:
    html = f.read()
before = len(html)

# ===== 读取源数据额外2周 =====
wb = load_workbook(SRC, data_only=True)
ws = wb['四三数据累计']

def get_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None
def num(v, d=0):
    try: return float(v) if v else d
    except: return d

# 4周列索引: [本周, 上周, 上上周, 上上上周]
WEEKS = {
    'sales':   [18, 17, 16, 15],  # 5.21, 5.14, 5.7, 4.30
    'revenue': [31, 30, 29, 28],
    'rival':   [44, 43, 42, 41],
    'share':   [56, 55, 54, 53],
    'ord8':    [80, 79, 78, 77],
    'freq7':   [92, 91, 90, 89],
    'nfreq7':  [104,103,102,101],
    'mkt':     [117,116,115,114],
    'op':      [129,128,127,126],
}
WEEK_LABELS = ['5.21-5.27', '5.14-5.20', '5.7-5.13', '4.30-5.6']
WEEK_DATES = ['5.27', '5.20', '5.13', '5.6']

cutoff = date(2026, 5, 27)
rows_raw = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1] and get_date(row[2]) and get_date(row[2]) <= cutoff:
        rows_raw.append(list(row))

print(f"SKU总数: {len(rows_raw)}")

# ===== 1. 更新 cum43Data — 添加额外2周字段 =====
print("1. 更新cum43Data...")
script_m = re.search(r'<script>(.*?)</script>', html, re.DOTALL)
js = script_m.group(1)

# 找到cum43Data并解析
cum_match = re.search(r'const cum43Data = (\[.*?\]);', js, re.DOTALL)
cum43 = json.loads(cum_match.group(1))
print(f"  现有 {len(cum43)} 条")

# 构建SKU→源数据行映射
sku_map = {}
for r in rows_raw:
    sku = str(r[1] or '').strip()
    if sku:
        sku_map[sku] = r

# 给每条记录添加额外2周字段
added_fields = 0
for d in cum43:
    sku = d.get('SKU', '')
    if sku in sku_map:
        r = sku_map[sku]
        # 添加上上周和上上上周的字段
        d['salesW2'] = int(num(r[WEEKS['sales'][2]]))   # 5.7-5.13
        d['salesW3'] = int(num(r[WEEKS['sales'][3]]))   # 4.30-5.6
        d['revenueW2'] = round(num(r[WEEKS['revenue'][2]]), 2)
        d['revenueW3'] = round(num(r[WEEKS['revenue'][3]]), 2)
        d['rivalW2'] = int(num(r[WEEKS['rival'][2]]))
        d['rivalW3'] = int(num(r[WEEKS['rival'][3]]))
        d['shareW2'] = round(num(r[WEEKS['share'][2]]) * 100, 1)
        d['shareW3'] = round(num(r[WEEKS['share'][3]]) * 100, 1)
        d['mktW2'] = str(r[WEEKS['mkt'][2]] or '').strip()
        d['mktW3'] = str(r[WEEKS['mkt'][3]] or '').strip()
        added_fields += 1

print(f"  已添加W2/W3字段: {added_fields} 条")

# 替换cum43Data
new_cum43 = json.dumps(cum43, ensure_ascii=False, separators=(',', ':'))
js = js[:cum_match.start(1)] + new_cum43 + js[cum_match.end(1):]

# ===== 2. 更新 categoryRevenueData / analystRevenueData — 添加4周趋势 =====
print("2. 更新维度数据...")

def upgrade_dim_data(js, var_name, key_field):
    """给维度数据添加4周销量/销售额/市占比"""
    pattern = rf'const {var_name} = (\[.*?\]);'
    m = re.search(pattern, js, re.DOTALL)
    if not m:
        print(f"  WARNING: {var_name} not found")
        return js
    data = json.loads(m.group(1))

    # 按维度聚合4周数据
    for d in data:
        key = d.get(key_field, '')
        sales_w = [0, 0, 0, 0]
        rev_w = [0, 0, 0, 0]
        rival_w = [0, 0, 0, 0]
        for cd in cum43:
            if cd.get('analyst' if key_field == 'analyst' else 'category') == key:
                sales_w[0] += cd.get('curSalesQty', 0)
                sales_w[1] += cd.get('prevSalesQty', 0)
                sales_w[2] += cd.get('salesW2', 0)
                sales_w[3] += cd.get('salesW3', 0)
                rev_w[0] += cd.get('curRevenue', 0)
                rev_w[1] += cd.get('prevRevenue', 0)
                rev_w[2] += cd.get('revenueW2', 0)
                rev_w[3] += cd.get('revenueW3', 0)
                rival_w[0] += cd.get('curRivalQty', 0)
                rival_w[1] += cd.get('prevRivalQty', 0)
                rival_w[2] += cd.get('rivalW2', 0)
                rival_w[3] += cd.get('rivalW3', 0)
        d['sales4w'] = sales_w
        d['revenue4w'] = [round(v, 2) for v in rev_w]
        # 计算4周市占比
        shares = []
        for i in range(4):
            s = sales_w[i]; r = rival_w[i]
            shares.append(round(s/(s+r)*100, 1) if (s+r) > 0 else 0)
        d['share4w'] = shares

    new_data = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
    js = js[:m.start(1)] + new_data + js[m.end(1):]
    return js

js = upgrade_dim_data(js, 'categoryRevenueData', 'category')
js = upgrade_dim_data(js, 'analystRevenueData', 'analyst')

# ===== 3. 更新 cum43Stats — 添加4周统计 =====
print("3. 更新cum43Stats...")
stats_m = re.search(r'const cum43Stats = ({.*?});', js, re.DOTALL)
stats = json.loads(stats_m.group(1))

# 添加4周总销量/销售额/对手/市占比
totals_4w = {'sales': [0]*4, 'revenue': [0]*4, 'rival': [0]*4, 'share': [0]*4}
for cd in cum43:
    totals_4w['sales'][0] += cd.get('curSalesQty', 0)
    totals_4w['sales'][1] += cd.get('prevSalesQty', 0)
    totals_4w['sales'][2] += cd.get('salesW2', 0)
    totals_4w['sales'][3] += cd.get('salesW3', 0)
    totals_4w['revenue'][0] += cd.get('curRevenue', 0)
    totals_4w['revenue'][1] += cd.get('prevRevenue', 0)
    totals_4w['revenue'][2] += cd.get('revenueW2', 0)
    totals_4w['revenue'][3] += cd.get('revenueW3', 0)
    totals_4w['rival'][0] += cd.get('curRivalQty', 0)
    totals_4w['rival'][1] += cd.get('prevRivalQty', 0)
    totals_4w['rival'][2] += cd.get('rivalW2', 0)
    totals_4w['rival'][3] += cd.get('rivalW3', 0)

for i in range(4):
    s = totals_4w['sales'][i]; r = totals_4w['rival'][i]
    totals_4w['share'][i] = round(s/(s+r)*100, 1) if (s+r) > 0 else 0

stats['sales4w'] = totals_4w['sales']
stats['revenue4w'] = [round(v, 2) for v in totals_4w['revenue']]
stats['share4w'] = totals_4w['share']
stats['weekLabels'] = WEEK_LABELS
stats['weekDates'] = WEEK_DATES

new_stats = json.dumps(stats, ensure_ascii=False, separators=(',', ':'))
js = js[:stats_m.start(1)] + new_stats + js[stats_m.end(1):]

# ===== 4. 更新 prevWeekKpi =====
print("4. 更新prevWeekKpi...")
kpi_m = re.search(r'const prevWeekKpi = ({.*?});', js, re.DOTALL)
kpi = json.loads(kpi_m.group(1))
kpi['sales4w'] = totals_4w['sales']
kpi['revenue4w'] = [round(v, 2) for v in totals_4w['revenue']]
kpi['share4w'] = totals_4w['share']
new_kpi = json.dumps(kpi, ensure_ascii=False, separators=(',', ':'))
js = js[:kpi_m.start(1)] + new_kpi + js[kpi_m.end(1):]

# ===== 5. 升级Tab1图表：品线/分析人销量/销售额从2-bar改为4-week折线 =====
print("5. 升级图表为4周趋势...")

# 5a. 升级 chart-cat-sales → 4周折线图
old_cat_sales = r"""new Chart(document.getElementById('chart-cat-sales'), {
    type: 'bar', data: { labels: catLabels, datasets: [
      { label: '本周销量', data: categoryRevenueData.map(function(d){return d.curSalesQty;}), backgroundColor: '#0f3460', yAxisID: 'y' },
      { label: '上周销量', data: categoryRevenueData.map(function(d){return d.prevSalesQty;}), backgroundColor: '#ccc', yAxisID: 'y' },
      { label: '环比变化(%)', data: categoryRevenueData.map(function(d){ var v=parseFloat(d.salesQtyChange); return isNaN(v)?0:v; }), borderColor: '#c0392b', backgroundColor: 'transparent', type: 'line', yAxisID: 'y1', tension: 0.3, borderWidth: 2, pointRadius: 4, pointBackgroundColor: '#c0392b' }
    ]},
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true, title: { display: true, text: '销量' } }, y1: { position: 'right', title: { display: true, text: '环比变化(%)' }, grid: { drawOnChartArea: false } } } }
  });"""

new_cat_sales = r"""new Chart(document.getElementById('chart-cat-sales'), {
    type: 'line', data: { labels: ['4.30-5.6','5.7-5.13','5.14-5.20','5.21-5.27'], datasets: categoryRevenueData.map(function(d,i){ var colors=['#0f3460','#2980b9','#8e44ad','#e07b24','#08845a','#c0392b']; return { label: d.category, data: d.sales4w || [d.curSalesQty,d.prevSalesQty,d.salesW2||0,d.salesW3||0], borderColor: colors[i%colors.length], backgroundColor: 'transparent', tension: 0.3, borderWidth: 2, pointRadius: 3 }; }) },
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true, title: { display: true, text: '销量' } } } }
  });"""

js = js.replace(old_cat_sales, new_cat_sales)

# 5b. 升级 chart-an-sales → 4周折线图
old_an_sales = r"""new Chart(document.getElementById('chart-an-sales'), {
    type: 'bar', data: { labels: anLabels, datasets: [
      { label: '本周销量', data: analystRevenueData.map(function(d){return d.curSalesQty;}), backgroundColor: '#0f3460', yAxisID: 'y' },
      { label: '上周销量', data: analystRevenueData.map(function(d){return d.prevSalesQty;}), backgroundColor: '#ccc', yAxisID: 'y' },
      { label: '环比变化(%)', data: analystRevenueData.map(function(d){ var v=parseFloat(d.salesQtyChange); return isNaN(v)?0:v; }), borderColor: '#c0392b', backgroundColor: 'transparent', type: 'line', yAxisID: 'y1', tension: 0.3, borderWidth: 2, pointRadius: 4, pointBackgroundColor: '#c0392b' }
    ]},
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true, title: { display: true, text: '销量' } }, y1: { position: 'right', title: { display: true, text: '环比变化(%)' }, grid: { drawOnChartArea: false } } } }
  });"""

new_an_sales = r"""new Chart(document.getElementById('chart-an-sales'), {
    type: 'line', data: { labels: ['4.30-5.6','5.7-5.13','5.14-5.20','5.21-5.27'], datasets: analystRevenueData.map(function(d,i){ var colors=['#0f3460','#2980b9','#8e44ad','#e07b24','#08845a','#c0392b']; return { label: d.analyst, data: d.sales4w || [d.curSalesQty,d.prevSalesQty,d.salesW2||0,d.salesW3||0], borderColor: colors[i%colors.length], backgroundColor: 'transparent', tension: 0.3, borderWidth: 2, pointRadius: 3 }; }) },
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true, title: { display: true, text: '销量' } } } }
  });"""

js = js.replace(old_an_sales, new_an_sales)

# 5c. 升级 chart-cat-rev → 4周折线图
old_cat_rev = r"""new Chart(document.getElementById('chart-cat-rev'), {
    type: 'bar', data: { labels: catLabels, datasets: [
      { label: '本周销售额($)', data: categoryRevenueData.map(function(d){return d.curRevenue;}), backgroundColor: '#8e44ad' },
      { label: '上周销售额($)', data: categoryRevenueData.map(function(d){return d.prevRevenue;}), backgroundColor: '#ddd' }
    ]},
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true } } }
  });"""

new_cat_rev = r"""new Chart(document.getElementById('chart-cat-rev'), {
    type: 'line', data: { labels: ['4.30-5.6','5.7-5.13','5.14-5.20','5.21-5.27'], datasets: categoryRevenueData.map(function(d,i){ var colors=['#8e44ad','#2980b9','#e07b24','#08845a','#c0392b','#0f3460']; return { label: d.category, data: d.revenue4w || [d.curRevenue,d.prevRevenue,d.revenueW2||0,d.revenueW3||0], borderColor: colors[i%colors.length], backgroundColor: 'transparent', tension: 0.3, borderWidth: 2, pointRadius: 3 }; }) },
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true, title: { display: true, text: '销售额($)' } } } }
  });"""

js = js.replace(old_cat_rev, new_cat_rev)

# 5d. 升级 chart-an-rev → 4周折线图
old_an_rev = r"""new Chart(document.getElementById('chart-an-rev'), {
    type: 'bar', data: { labels: anLabels, datasets: [
      { label: '本周销售额($)', data: analystRevenueData.map(function(d){return d.curRevenue;}), backgroundColor: '#8e44ad' },
      { label: '上周销售额($)', data: analystRevenueData.map(function(d){return d.prevRevenue;}), backgroundColor: '#ddd' }
    ]},
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true } } }
  });"""

new_an_rev = r"""new Chart(document.getElementById('chart-an-rev'), {
    type: 'line', data: { labels: ['4.30-5.6','5.7-5.13','5.14-5.20','5.21-5.27'], datasets: analystRevenueData.map(function(d,i){ var colors=['#8e44ad','#2980b9','#e07b24','#08845a','#c0392b','#0f3460']; return { label: d.analyst, data: d.revenue4w || [d.curRevenue,d.prevRevenue,d.revenueW2||0,d.revenueW3||0], borderColor: colors[i%colors.length], backgroundColor: 'transparent', tension: 0.3, borderWidth: 2, pointRadius: 3 }; }) },
    options: { responsive: true, plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true, title: { display: true, text: '销售额($)' } } } }
  });"""

js = js.replace(old_an_rev, new_an_rev)

# 5e. 在initCharts1末尾添加4周趋势图
charts1_end = "setTimeout(initCharts1, 100);"
new_4w_charts_js = r"""
  // 6. 4周总销量趋势
  new Chart(document.getElementById('chart-4w-sales'), {
    type: 'line', data: { labels: ['4.30-5.6','5.7-5.13','5.14-5.20','5.21-5.27'], datasets: [
      { label: '总销量', data: cum43Stats.sales4w || [0,0,0,0], borderColor: '#0f3460', backgroundColor: 'rgba(15,52,96,0.1)', tension: 0.3, borderWidth: 3, pointRadius: 5, fill: true }
    ]},
    options: { responsive: true, plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true, title: { display: true, text: '销量' } } } }
  });
  // 7. 4周总销售额趋势
  new Chart(document.getElementById('chart-4w-rev'), {
    type: 'line', data: { labels: ['4.30-5.6','5.7-5.13','5.14-5.20','5.21-5.27'], datasets: [
      { label: '总销售额($)', data: cum43Stats.revenue4w || [0,0,0,0], borderColor: '#8e44ad', backgroundColor: 'rgba(142,68,173,0.1)', tension: 0.3, borderWidth: 3, pointRadius: 5, fill: true }
    ]},
    options: { responsive: true, plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true, title: { display: true, text: '销售额($)' } } } }
  });
  // 8. 4周总市占比趋势
  new Chart(document.getElementById('chart-4w-share'), {
    type: 'line', data: { labels: ['4.30-5.6','5.7-5.13','5.14-5.20','5.21-5.27'], datasets: [
      { label: '总市占比(%)', data: cum43Stats.share4w || [0,0,0,0], borderColor: '#08845a', backgroundColor: 'rgba(8,132,90,0.1)', tension: 0.3, borderWidth: 3, pointRadius: 5, fill: true }
    ]},
    options: { responsive: true, plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true, max: 100, title: { display: true, text: '市占比(%)' } } } }
  });
"""

js = js.replace(charts1_end, charts1_end + new_4w_charts_js)

# ===== 6. 更新图表注释 =====
js = js.replace("// 2. 品线销量对比", "// 2. 品线4周销量趋势")
js = js.replace("// 3. 分析人销量对比", "// 3. 分析人4周销量趋势")
js = js.replace("// 4. 品线销售额对比", "// 4. 品线4周销售额趋势")
js = js.replace("// 5. 分析人销售额对比", "// 5. 分析人4周销售额趋势")

# ===== 写入 =====
html = html[:script_m.start(1)] + js + html[script_m.end(1):]

# 5f. Tab1 body中插入4周趋势图canvas（在重组后操作，避免位移）
new_4w_chart = '<div class="chart-box"><h4>&#128200; 4周销量趋势</h4><canvas id="chart-4w-sales"></canvas></div>\n      <div class="chart-box"><h4>&#128200; 4周销售额趋势</h4><canvas id="chart-4w-rev"></canvas></div>\n      <div class="chart-box"><h4>&#128200; 4周市占比趋势</h4><canvas id="chart-4w-share"></canvas></div>'
html_body_part = re.search(r'(<div class="chart-box"><h4>&#128101; 分析人销售额对比</h4><canvas id="chart-an-rev"></canvas></div>)\s*</div>\s*<div class="section"><h3>&#128200; 新品出单情况</h3>', html, re.DOTALL)
if html_body_part:
    html = html[:html_body_part.end(1)] + '\n      ' + new_4w_chart + html[html_body_part.end(1):]
    print("5f. 已插入4周趋势canvas到body")
else:
    print("WARNING: 未找到chart-grid结束位置，跳过canvas插入")

after = len(html)
print(f"\n原始: {before/1024:.0f}KB → 升级后: {after/1024:.0f}KB (+{(after-before)/1024:.0f}KB)")

# 验证
script_v = re.search(r'<script>(.*?)</script>', html, re.DOTALL)
if script_v:
    new_js = script_v.group(1)
    op = new_js.count('(') - new_js.count(')')
    ob = new_js.count('{') - new_js.count('}')
    print(f"Brackets: ()={op}, {{}}={ob}")
    if op != 0 or ob != 0:
        print("ERROR: 括号不平衡!")
        sys.exit(1)
else:
    # Fallback: try matching any script tag
    script_v2 = re.search(r'<script[^>]*>(.*?)</script>', html, re.DOTALL)
    if script_v2:
        new_js = script_v2.group(1)
        op = new_js.count('(') - new_js.count(')')
        ob = new_js.count('{') - new_js.count('}')
        print(f"Fallback match OK, brackets: ()={op}, {{}}={ob}")
    else:
        print("WARNING: Could not find script tag for validation, writing anyway")

with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"保存到: {OUTPUT}")
print("4周升级完成!")
