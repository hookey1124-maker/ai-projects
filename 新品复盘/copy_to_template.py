"""
将新品检查周源数据抽取约10个SKU，按胡煜星模板表头复制数据
"""
import openpyxl
from copy import copy
from collections import defaultdict

SRC = '周报/新品检查周源数据和PLP数据.xlsx'
TPL = '周报/胡煜星-周报统一Workbook表头模板.xlsx'
OUT = '周报/测试-模板数据填充_10SKU.xlsx'

def safe_float(v, default=0):
    """安全转float"""
    try: return float(v)
    except: return default

def safe_int(v, default=0):
    """安全转int"""
    try: return int(float(v))
    except: return default

def safe_str(v, default=''):
    """安全转str"""
    if v is None: return default
    return str(v).strip()

# ===== 1. 读取源数据 =====
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_main = wb_src['四三数据累计']
ws_plp = wb_src['PLP明细']

# 解析四三数据累计
rows_main = []
for row in ws_main.iter_rows(min_row=2, values_only=True):
    if row[0] is not None:
        rows_main.append(list(row))

# 解析PLP明细（仅本期5.4-5.10）
plp_rows = []
for row in ws_plp.iter_rows(min_row=2, values_only=True):
    period = str(row[0] or '').strip()
    if '5.4' in period or '5.10' in period:
        sku = str(row[2] or '').strip()
        if sku and sku.startswith('LYAP'):
            plp_rows.append(list(row))

# 按SKU聚合PLP数据
plp_by_sku = defaultdict(lambda: {'impr': 0, 'click': 0, 'sold': 0, 'cost': 0, 'ad_rev': 0})
for r in plp_rows:
    sku = str(r[2] or '').strip()
    if not sku:
        continue
    d = plp_by_sku[sku]
    try: d['impr'] += int(r[11] or 0)
    except: pass
    try: d['click'] += int(r[12] or 0)
    except: pass
    try: d['sold'] += int(r[13] or 0)
    except: pass
    try: d['cost'] += float(r[14] or 0)
    except: pass
    try: d['ad_rev'] += float(r[15] or 0)
    except: pass

# ===== 2. 选取10个有代表性的SKU =====
# 优先选有PLP数据、不同品类、不同分析人的SKU
selected = []
seen_analysts = set()
seen_cats = set()

for r in rows_main:
    sku = str(r[1] or '').strip()
    if not sku or not sku.startswith('LYAP'):
        continue
    an = str(r[4] or '').strip()
    cat = str(r[5] or '').strip()

    # 优先选不同分析人和品类
    score = 0
    if an not in seen_analysts:
        score += 3
    if cat not in seen_cats:
        score += 2
    if sku in plp_by_sku:
        score += 2  # 有PLP数据加分
    # 有销量的加分
    try:
        if safe_int(r[16]) > 0:
            score += 1
    except:
        pass

    selected.append((score, r))
    seen_analysts.add(an)
    seen_cats.add(cat)

# 按分数排序取前12（确保有12个候选）
selected.sort(key=lambda x: -x[0])
selected = [r for _, r in selected[:12]]

# 确保至少有10个
print(f"选取 {len(selected)} 个SKU:")
for r in selected:
    sku = r[1]
    an = r[4]
    cat = r[5]
    has_plp = 'Y' if sku in plp_by_sku else 'N'
    print(f"  {sku} | {an} | {cat} | PLP:{has_plp}")

# ===== 3. 打开模板，复制样式 =====
wb_tpl = openpyxl.load_workbook(TPL)
wb_out = openpyxl.Workbook()
# 删除默认sheet
wb_out.remove(wb_out.active)

# 处理每个sheet
for sn in wb_tpl.sheetnames:
    ws_tpl = wb_tpl[sn]
    ws_out = wb_out.create_sheet(title=sn)

    # 复制列宽
    for col_idx in range(1, ws_tpl.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if ws_tpl.column_dimensions[col_letter].width:
            ws_out.column_dimensions[col_letter].width = ws_tpl.column_dimensions[col_letter].width

    # 获取模板表头
    headers = []
    for c in ws_tpl[1]:
        headers.append(str(c.value) if c.value is not None else '')

    # 复制表头样式和内容
    for col_idx, h in enumerate(headers, 1):
        cell_tpl = ws_tpl.cell(row=1, column=col_idx)
        cell_out = ws_out.cell(row=1, column=col_idx, value=h)
        if cell_tpl.font:
            cell_out.font = copy(cell_tpl.font)
        if cell_tpl.fill:
            cell_out.fill = copy(cell_tpl.fill)
        if cell_tpl.alignment:
            cell_out.alignment = copy(cell_tpl.alignment)

    # 如果不是数据sheet，只复制表头（保留模板说明）
    if sn in ('99_字段说明', '账号流量转移'):
        for row_idx in range(2, ws_tpl.max_row + 1):
            for col_idx in range(1, ws_tpl.max_column + 1):
                v = ws_tpl.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    ws_out.cell(row=row_idx, column=col_idx, value=v)
        continue

    # ===== 4. 按sheet映射数据 =====
    if sn == '01_产品维度':
        # 映射: SKU(1) 销售编码(2) 产品名称(3) 产品分类(4) 三级类目(5)
        #        分析人(6) 产品等级(7) 账号(8) 站点(9) 实际上架日期(10)
        #        库存(11) 可售周(12) 是否亏本(13) 是否淘汰(14) 备注(15)
        for i, r in enumerate(selected):
            row_out = i + 2
            ws_out.cell(row=row_out, column=1, value=str(r[1] or ''))   # SKU
            ws_out.cell(row=row_out, column=2, value=str(r[0] or ''))   # 销售编码
            ws_out.cell(row=row_out, column=3, value='')                 # 产品名称
            ws_out.cell(row=row_out, column=4, value=str(r[5] or ''))   # 产品分类(品类)
            ws_out.cell(row=row_out, column=5, value='')                 # 三级类目
            ws_out.cell(row=row_out, column=6, value=str(r[4] or ''))   # 分析人
            ws_out.cell(row=row_out, column=7, value='')                 # 产品等级
            ws_out.cell(row=row_out, column=8, value='')                 # 账号
            ws_out.cell(row=row_out, column=9, value='')                 # 站点
            ws_out.cell(row=row_out, column=10, value=str(r[2] or '')[:10] if r[2] else '')  # 上架日期
            ws_out.cell(row=row_out, column=11, value='')                # 库存
            ws_out.cell(row=row_out, column=12, value='')                # 可售周
            ws_out.cell(row=row_out, column=13, value='')                # 是否亏本
            ws_out.cell(row=row_out, column=14, value='')                # 是否淘汰
            ws_out.cell(row=row_out, column=15, value='')                # 备注

    elif sn == '02_销售明细':
        # 日期(1) SKU(2) 销售编码(3) 账号(4) 销量(5) 销售额(6) 订单数(7) 退款量(8) 退款金额(9) 备注(10)
        # 源数据只有周汇总，这里放一行周数据
        for i, r in enumerate(selected):
            row_out = i + 2
            ws_out.cell(row=row_out, column=1, value='5.7-5.13')         # 日期
            ws_out.cell(row=row_out, column=2, value=str(r[1] or ''))   # SKU
            ws_out.cell(row=row_out, column=3, value=str(r[0] or ''))   # 销售编码
            ws_out.cell(row=row_out, column=4, value='')                 # 账号
            ws_out.cell(row=row_out, column=5, value=safe_int(r[16]))   # 销量(5.7-5.13)
            ws_out.cell(row=row_out, column=6, value=round(safe_float(r[27]), 2))  # 销售额
            ws_out.cell(row=row_out, column=7, value='')                 # 订单数
            ws_out.cell(row=row_out, column=8, value='')                 # 退款量
            ws_out.cell(row=row_out, column=9, value='')                 # 退款金额
            ws_out.cell(row=row_out, column=10, value='')                # 备注

    elif sn == '03_价格明细':
        # 日期/SKU/销售编码/在售价格...源数据无价格明细，留空
        for i, r in enumerate(selected):
            row_out = i + 2
            ws_out.cell(row=row_out, column=1, value='5.7-5.13')
            ws_out.cell(row=row_out, column=2, value=str(r[1] or ''))
            ws_out.cell(row=row_out, column=3, value=str(r[0] or ''))

    elif sn == '04_新品明细':
        # 这是最关键的sheet，27列
        # 1:SKU 2:销售编码 3:产品名称 4:产品分类 5:分析人 6:产品拓展类型
        # 7:实际上架日期 8:首次出单日期 9:销量 10:销售额 11:市占比 12:追踪间隔
        # 13:8日出单情况 14:7日频次标签 15:上架8日内新品频次标签
        # 16:PLP是否开启 17:PLP曝光 18:PLP点击 19:PLP花费 20:PLP出单
        # 21:PLP销售额 22:PLP_ACOS 23:PLP_ACOAS 24:PLG最高费率
        # 25:市场状态 26:操作判断 27:备注
        for i, r in enumerate(selected):
            row_out = i + 2
            sku = str(r[1] or '').strip()
            plp = plp_by_sku.get(sku, {})

            # 计算ACOS/ACOAS
            cost = plp.get('cost', 0)
            ad_rev = plp.get('ad_rev', 0)
            total_rev = safe_float(r[27])
            acos = round(cost / ad_rev * 100, 2) if ad_rev else 0
            acoas = round(cost / total_rev * 100, 2) if total_rev else 0

            ws_out.cell(row=row_out, column=1, value=sku)                      # SKU
            ws_out.cell(row=row_out, column=2, value=str(r[0] or ''))          # 销售编码
            ws_out.cell(row=row_out, column=3, value='')                        # 产品名称
            ws_out.cell(row=row_out, column=4, value=str(r[5] or ''))          # 产品分类
            ws_out.cell(row=row_out, column=5, value=str(r[4] or ''))          # 分析人
            ws_out.cell(row=row_out, column=6, value=str(r[6] or ''))          # 产品拓展类型
            ws_out.cell(row=row_out, column=7, value=str(r[2] or '')[:10] if r[2] else '')  # 上架日期
            ws_out.cell(row=row_out, column=8, value=str(r[3] or '')[:10] if r[3] else '')  # 首次出单
            ws_out.cell(row=row_out, column=9, value=safe_int(r[16]))          # 销量(5.7-5.13)
            ws_out.cell(row=row_out, column=10, value=round(safe_float(r[27]), 2))  # 销售额
            share_val = safe_float(r[48])
            ws_out.cell(row=row_out, column=11, value=f"{share_val*100:.1f}%" if share_val else str(r[48] or ''))  # 市占比
            ws_out.cell(row=row_out, column=12, value=str(r[58] or ''))        # 追踪间隔(5.13)
            ws_out.cell(row=row_out, column=13, value=str(r[68] or ''))        # 8日出单情况
            ws_out.cell(row=row_out, column=14, value=str(r[78] or ''))        # 7日频次标签
            ws_out.cell(row=row_out, column=15, value=str(r[88] or ''))        # 上架8日内新品频次标签
            ws_out.cell(row=row_out, column=16, value=str(r[114] or ''))       # PLP是否开启
            ws_out.cell(row=row_out, column=17, value=plp.get('impr', 0))     # PLP曝光
            ws_out.cell(row=row_out, column=18, value=plp.get('click', 0))    # PLP点击
            ws_out.cell(row=row_out, column=19, value=plp.get('cost', 0))     # PLP花费
            ws_out.cell(row=row_out, column=20, value=plp.get('sold', 0))     # PLP出单
            ws_out.cell(row=row_out, column=21, value=plp.get('ad_rev', 0))   # PLP销售额
            ws_out.cell(row=row_out, column=22, value=f"{acos:.2f}%")          # PLP_ACOS
            ws_out.cell(row=row_out, column=23, value=f"{acoas:.2f}%")         # PLP_ACOAS
            ws_out.cell(row=row_out, column=24, value=str(r[118] or ''))       # PLG最高费率
            ws_out.cell(row=row_out, column=25, value=str(r[99] or ''))        # 市场状态
            ws_out.cell(row=row_out, column=26, value=str(r[109] or ''))       # 操作判断
            ws_out.cell(row=row_out, column=27, value='')                      # 备注

    elif sn == '05_账号流量':
        # 日期/账号/SKU/销售编码/曝光/点击...
        for i, r in enumerate(selected):
            row_out = i + 2
            sku = str(r[1] or '').strip()
            plp = plp_by_sku.get(sku, {})
            ws_out.cell(row=row_out, column=1, value='5.7-5.13')
            ws_out.cell(row=row_out, column=2, value='')                # 账号
            ws_out.cell(row=row_out, column=3, value=sku)
            ws_out.cell(row=row_out, column=4, value=str(r[0] or ''))
            ws_out.cell(row=row_out, column=5, value=plp.get('impr', 0))  # 曝光(用PLP曝光)
            ws_out.cell(row=row_out, column=6, value=plp.get('click', 0)) # 点击
            ws_out.cell(row=row_out, column=7, value='')                # 自然曝光
            ws_out.cell(row=row_out, column=8, value='')                # 自然点击
            ws_out.cell(row=row_out, column=9, value=plp.get('impr', 0))  # 广告曝光
            ws_out.cell(row=row_out, column=10, value=plp.get('click', 0)) # 广告点击
            ws_out.cell(row=row_out, column=11, value=plp.get('cost', 0))  # 广告花费
            ws_out.cell(row=row_out, column=12, value=plp.get('ad_rev', 0)) # 广告销售额
            ws_out.cell(row=row_out, column=13, value=plp.get('sold', 0))  # 广告出单
            ws_out.cell(row=row_out, column=14, value=safe_int(r[16]))   # 账号销量
            ws_out.cell(row=row_out, column=15, value=round(safe_float(r[27]), 2)) # 账号销售额

    elif sn == '06_销量预估':
        for i, r in enumerate(selected):
            row_out = i + 2
            ws_out.cell(row=row_out, column=1, value=str(r[1] or ''))
            ws_out.cell(row=row_out, column=2, value=str(r[0] or ''))

# 保存
wb_out.save(OUT)
print(f"\nDone! 输出文件: {OUT}")

# ===== 5. 验证输出 =====
print("\n===== 验证 =====")
wb_check = openpyxl.load_workbook(OUT, data_only=True)
for sn in ['01_产品维度', '04_新品明细']:
    ws = wb_check[sn]
    print(f"\n--- {sn} ---")
    # 打印表头
    h = [str(ws.cell(row=1, column=c).value or '') for c in range(1, min(ws.max_column+1, 28))]
    print(f"表头: {' | '.join(h[:8])}...")
    # 打印第一行数据
    if ws.max_row >= 2:
        d = [str(ws.cell(row=2, column=c).value or '') for c in range(1, min(ws.max_column+1, 28))]
        print(f"数据1: {' | '.join(d[:8])}...")
        print(f"数据列数: {ws.max_column}")
