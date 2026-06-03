import openpyxl, sys
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout.reconfigure(encoding='utf-8')

DST = r'C:\Users\Hardy\Desktop\三部周报V1_导出\源数据_新品检查周源数据和PLP数据.xlsx'
wb = openpyxl.load_workbook(DST)

if '字段详解' in wb.sheetnames:
    del wb['字段详解']

ws = wb.create_sheet('字段详解', 0)

rows = [
    ['数据域', '源Excel列', 'JSON字段', 'Adapter函数/用途', '类型', '说明'],
    ['SKU基础', '销售编号', 'skuAll[].saleNo', 'getRawRows() -> cum43Data', 'string', '唯一销售编码'],
    ['SKU基础', 'SKU', 'skuAll[].sku', 'getRawRows() -> cum43Data', 'string', '产品SKU编码，关联PLP明细的键'],
    ['SKU基础', '实际上架日期', 'skuAll[].listDate', 'countActiveSku / countNewSku', 'string(YYYY-MM-DD)', '剔除周期后上架SKU，统计新上架数'],
    ['SKU基础', '首次出单日期', 'skuAll[].firstOrderDate', 'getRawRows() -> cum43Data', 'string', '未出单则为空'],
    ['SKU基础', '分析人', 'skuAll[].analyst', 'getCategoryMetrics/getAnalystMetrics', 'string', '6位分析人之一'],
    ['SKU基础', '品类', 'skuAll[].category', 'getCategoryMetrics', 'string', '7个品线之一'],
    ['SKU基础', '产品拓展', 'skuAll[].expandType', 'getRawRows() -> cum43Data', 'string', '原开品/拓展品/组合件'],
    ['每周-销量', 'M.D-M.D销量 (12列)', 'skuAll[].salesAll[12]', 'curSalesQty / prevSalesQty', 'number', '本周销量，索引0最早3.5-3.11，11最新5.21-5.27'],
    ['每周-销量', '(聚合)', 'totalSalesAll[12]', 'getData() -> totalSales4w', 'number[]', '全SKU每周销量合计'],
    ['每周-销量', '(聚合)', 'catSalesAll[].sales[12]', 'getCategoryMetrics()', 'number[]', '按品类每周销量合计'],
    ['每周-销量', '(聚合)', 'anSalesAll[].sales[12]', 'getAnalystMetrics()', 'number[]', '按分析人每周销量合计'],
    ['每周-销售额', 'M.D-M.D销售额 (12列)', 'skuAll[].revenueAll[12]', 'curRevenue / prevRevenue', 'number', '本周销售额($)'],
    ['每周-销售额', '(聚合)', 'totalRevAll[12]', 'getData() -> totalRev4w', 'number[]', '全SKU每周销售额合计'],
    ['每周-对手', 'M.D对手销量 (12列)', 'skuAll[].rivalAll[12]', 'curRivalQty / prevRivalQty', 'number', '直接竞品对手周出单量'],
    ['每周-对手', '(聚合)', '(实时计算)', 'hasRivalCount', 'number', 'curRivalQty>0的SKU数'],
    ['每周-市占比', 'M.D市占比 (12列)', 'skuAll[].shareAll[12]', 'curMarketShare / prevMarketShare', 'number(%)', '已x100，用于getShareTier分档'],
    ['每周-市占比', '(聚合)', 'totalShareAll[12]', 'getData() -> totalShare4w', 'number[]', '加权市占比=总销量/(总销量+总对手)'],
    ['每周-市占比', '(计算)', '-', 'shareTier', 'getShareTier(share,rival)', 'string', '高>=75/中50-75/低<50或对手>0/无市场=0且对手=0'],
    ['每周-状态', 'M.D 8日出单情况 (12列)', 'skuAll[].ord8All[12]', 'cur8dStatus', 'string', 'Y=8日内出单 N=8日外出单 未出单=无订单'],
    ['每周-状态', '(计算)', '-', 'hasOrder', 'getHasOrder(status,sales)', 'string', '上架后出过单=是: 8日出单Y/N或salesQty>0'],
    ['每周-状态', 'M.D市场状态 (12列)', 'skuAll[].mktAll[12]', 'curMarketStatus', 'string', '正常/竞争无优势/无市场/站外出单'],
    ['每周-状态', '(聚合)', '(实时计算)', 'mktDistOverall.distribution', 'object[]', '4种状态各SKU计数,getData()实时算'],
    ['每周-广告', 'M.D-M.D开启PLP (6列)', 'skuAll[].plpAll[12]', 'plpEnabled', 'string', 'Y=开启PLP N=未开启'],
    ['每周-广告', 'M.D-M.D PLG最高费率 (6列)', 'skuAll[].plgFeeAll[12]', 'plgFee', 'number(%)', '用于getAdClassLabel/getPlgFeeTier'],
    ['每周-广告', '(计算)', '-', 'adClassLabel', 'getAdClassLabel(plp,fee)', 'string', 'PLP+PLG/单PLP/仅PLG/无广告'],
    ['每周-广告', '(计算)', '-', 'plgFeeTier', 'getPlgFeeTier(fee)', 'string', '无广告/低<=2%/中2-4%/高>4%'],
    ['分析及时率', 'M.D 7日频次标签 (12列)', 'skuAll[].freq7All[12]', 'computeTimeliness(cum43Data)', 'string', 'freq7!=异常的SKU为及时，timelyRate=及时SKU/总数'],
    ['分析及时率', 'M.D 上架8日内新品频次标签 (12列)', 'skuAll[].nfreq7All[12]', 'computeTimeliness(cum43Data)', 'string', 'nfreq7=异常的SKU计入noAnalysis7dCount'],
    ['分析及时率', '(计算)', 'timelinessData.total', 'getData()实时计算', 'object', 'curSku/timelyCount/timelyRate/noAnalysis8dCount/noAnalysis7dCount'],
    ['分析及时率', '(聚合)', 'timelinessData.analysts[]', 'getData()实时计算', 'object[]', '按分析人的及时率/未分析数明细，随周期变化'],
    ['分析及时率', '(聚合)', 'timeliness4w', '(未接入周期选择)', 'object', '4周及时率趋势，暂未接入'],
    ['部门对比', '四三销售数据 Row2 Col3', 'deptTotalSales', 'getDeptKpi()', 'number', '部门总销量，算新品占部门比'],
    ['部门对比', '四三销售数据 Row2 Col4', 'deptTotalRevenue', 'getDeptKpi()', 'number', '部门总销售额($)，算新品占部门比'],
    ['部门对比', '(计算)', 'salesPct / revPct', 'getDeptKpi()', 'string(%)', '新品销量/销售额占部门百分比'],
    ['PW爬虫', 'PW表 全表', 'pwShare', 'getDeptKpi()', 'number(%)', 'PW爬虫加权市占，987个有效链接加权平均'],
    ['PW爬虫', 'PW表 全表', 'pwTotalLinks', 'getDeptKpi()', 'number', 'PW爬虫有竞品数据的链接数'],
    ['PW爬虫', '(计算)', 'newShareW', 'getDeptKpi()', 'number(%)', '新品加权市占，有对手SKU的销量/(销量+对手)'],
    ['PLP明细', '统计周期', 'plpDetailData[].period', 'getLinkDetail()', 'string', 'PLP广告统计周期(5.18-5.24)'],
    ['PLP明细', '广告活动', 'plpDetailData[].campaign', 'getLinkDetail() -> campaign', 'string', '广告系列名称'],
    ['PLP明细', 'SKU', 'plpDetailData[].SKU', 'getLinkDetail() -> sku', 'string', '关联SKU，与skuMap关联取市占销量等'],
    ['PLP明细', 'ID', 'plpDetailData[].id', 'getLinkDetail() -> linkId', 'string', '广告链接ID'],
    ['PLP明细', '是否开启PLG', 'plpDetailData[].plgEnabled', 'getLinkDetail()过滤条件', 'string', 'Y=PLP+PLG同开链接，计入链接明细表'],
    ['PLP明细', '广告分类', 'plpDetailData[].adClass', 'getLinkDetail()过滤条件', 'string', 'PLP+PLG同开/单链接PLP+PLG同开/单PLP/仅PLG'],
    ['周期选择', '(计算)', 'allWeekLabels[12]', 'ALL_WEEK_LABELS', 'string[]', '全部12周标签，索引0=3.5-3.11'],
    ['周期选择', '(计算)', 'allPeriods[9]', 'ALL_PERIODS', '{label,index}[]', '可选周期(需>=3前序周)，默认最新'],
    ['周期选择', '(计算)', '-', 'slice4w(arr,endIdx)', 'function', '从12周长数组切4周窗口'],
    ['周期选择', '(计算)', '-', 'toDateKey(d)', 'function', '日期转YYYYMMDD整数，支持2026-03-04和5.21'],
    ['Tab3-低占比', '(实时计算)', 'lowShareData', 'getData()->cum43Data过滤', 'any[]', 'curMarketShare<75且curRivalQty>0的SKU'],
    ['Tab3-未出单', '(实时计算)', 'hasCompetitorUnsold', 'getData()->buildUnsoldDetail', 'object', 'total+byAnalyst[]+byCategory[]'],
    ['Tab3-未出单', '(实时计算)', 'unsoldNoCompetitor', 'getData()->buildUnsoldDetail', 'object', '同上，无对手的未出单SKU'],
    ['Tab4-广告构成', '(实时计算)', 'compKpis/tierKpis', 'getAdCompData(idx)', 'object', '按选定周期SKU统计广告分类和费率分档'],
    ['Tab4-链接明细', '(关联计算)', 'plpPlgLinks', 'getLinkDetail(idx)', 'object[]', '11列:SKU/ID/活动/分析人/品类/是否出单/市占比/对手/销量/销售额/PLG费率'],
    ['Tab5-四三累计', '(实时计算)', 'cum43Data', 'getRawRows(idx)', 'any[]', '按周期重建140行SKU(剔除未来SKU)'],
]

for row in rows:
    ws.append(row)

ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 32
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 52
ws.freeze_panes = 'A2'

for cell in ws[1]:
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill(start_color='0F3460', end_color='0F3460', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=6):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)

wb.save(DST)
print(f'Done: {ws.max_row-1} field mappings written')
