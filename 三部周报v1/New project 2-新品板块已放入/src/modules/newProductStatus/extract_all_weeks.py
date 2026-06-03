"""
从源Excel提取全部12周数据，输出多周期JSON
覆盖字段: 销量/销售额/对手销量/市占比/市场状态/8日出单
"""
import openpyxl, json, re, sys
from collections import defaultdict

SRC = r'C:\Users\Hardy\ai-projects\新品复盘\周报\新品检查周源数据和PLP数据（新）xlsx.xlsx'
TGT = r'C:\Users\Hardy\ai-projects\三部周报v1\New project 2-新品板块已放入\src\modules\newProductStatus\corrected_data.json'

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['四三数据累计']
headers = [str(ws.cell(1, c).value or '') for c in range(1, ws.max_column+1)]

# --- 1. Map week columns ---
week_cols = {}  # metric -> [(col_1based, date_str), ...]
for i, h in enumerate(headers):
    if not h: continue
    m = re.match(r'^(\d{1,2}\.\d{1,2}(?:-\d{1,2}\.\d{1,2})?)(.+)', h)
    if m:
        metric = m.group(2).strip()
        if metric not in week_cols: week_cols[metric] = []
        week_cols[metric].append((i+1, m.group(1)))

# Sort each metric by date
def sort_key(x):
    d = x[1].split('-')[0].split('.')
    return int(d[0])*100 + int(d[1])
for m in week_cols:
    week_cols[m].sort(key=sort_key)

# Use sales weeks as canonical periods
sales_weeks = week_cols.get('销量', [])
if len(sales_weeks) < 4:
    print('ERROR: less than 4 weeks of sales data')
    sys.exit(1)

NWEEKS = len(sales_weeks)
all_week_labels = [e[1] for e in sales_weeks]
print(f'Total weeks: {NWEEKS}, labels: {all_week_labels[0]} ... {all_week_labels[-1]}')

# Build column index lookup: for each metric, map week_index -> col_number
def build_col_map(metric_name):
    entries = week_cols.get(metric_name, [])
    # Match entries to sales_weeks by date
    result = [None] * NWEEKS
    for col, date_str in entries:
        # Try to find matching week in sales_weeks
        for wi, wl in enumerate(all_week_labels):
            # The date formats differ: sales uses "M.D-M.D", others use "M.D"
            # Match by end date
            if date_str in wl or wl.endswith(date_str) or date_str.endswith(wl.split('-')[-1]):
                result[wi] = col
                break
        # Fallback: use index position
        if all(v is None for v in result):
            for wi in range(NWEEKS):
                if result[wi] is None:
                    result[wi] = col
                    break
    return result

sales_cols = build_col_map('销量')
revenue_cols = build_col_map('销售额')
rival_cols = build_col_map('对手销量')
share_cols = build_col_map('市占比')
mkt_cols = build_col_map('市场状态')
ord8_cols = build_col_map('8日出单情况')
plp_cols = build_col_map('开启PLP')
plg_fee_cols = build_col_map('PLG最高费率')

# --- Helper ---
def safe_float(val, default=0.0):
    if val is None: return default
    try: return float(val)
    except (ValueError, TypeError): return default

def safe_int(val, default=0):
    if val is None: return default
    try: return int(float(val))
    except (ValueError, TypeError): return default

# --- 2. Extract SKU data ---
# Fixed metadata columns
META_COLS = {
    'saleNo': 1, 'sku': 2, 'listDate': 3, 'firstOrderDate': 4,
    'analyst': 5, 'category': 6, 'expandType': 7
}

sku_data = []
total_sales = [0.0] * NWEEKS
total_revenue = [0.0] * NWEEKS
total_rival = [0.0] * NWEEKS
# For share: weighted average

for row_idx in range(2, ws.max_row + 1):
    sku = str(ws.cell(row_idx, META_COLS['sku']).value or '').strip()
    if not sku: continue

    rec = {
        'sku': sku,
        'saleNo': str(ws.cell(row_idx, META_COLS['saleNo']).value or '').strip(),
        'listDate': str(ws.cell(row_idx, META_COLS['listDate']).value or '').strip()[:10],
        'firstOrderDate': str(ws.cell(row_idx, META_COLS['firstOrderDate']).value or '').strip()[:10],
        'analyst': str(ws.cell(row_idx, META_COLS['analyst']).value or '').strip() or '未知',
        'category': str(ws.cell(row_idx, META_COLS['category']).value or '').strip() or '未分类',
        'expandType': str(ws.cell(row_idx, META_COLS['expandType']).value or '').strip() or '其他',
        'salesAll': [],
        'revenueAll': [],
        'rivalAll': [],
        'shareAll': [],
        'mktAll': [],
        'ord8All': [],
        'plpAll': [],
        'plgFeeAll': [],
    }

    for wi in range(NWEEKS):
        # Sales
        sc = sales_cols[wi]
        sv = safe_float(ws.cell(row_idx, sc).value) if sc else 0
        rec['salesAll'].append(sv)
        total_sales[wi] += sv

        # Revenue
        rc = revenue_cols[wi]
        rv = safe_float(ws.cell(row_idx, rc).value) if rc else 0
        rec['revenueAll'].append(round(rv, 2))
        total_revenue[wi] += rv

        # Rival
        ric = rival_cols[wi]
        riv = safe_int(ws.cell(row_idx, ric).value) if ric else 0
        rec['rivalAll'].append(riv)
        total_rival[wi] += riv

        # Share (raw decimal from Excel, multiply by 100 for %)
        shc = share_cols[wi]
        shv = safe_float(ws.cell(row_idx, shc).value) if shc else 0
        rec['shareAll'].append(round(shv * 100, 1))

        # Market status
        mc = mkt_cols[wi]
        mkt = str(ws.cell(row_idx, mc).value or '').strip() if mc else ''
        rec['mktAll'].append(mkt)

        # 8-day status
        o8c = ord8_cols[wi]
        o8 = str(ws.cell(row_idx, o8c).value or '').strip() if o8c else ''
        rec['ord8All'].append(o8)

        # PLP
        pc = plp_cols[wi] if wi < len(plp_cols) else None
        plp = str(ws.cell(row_idx, pc).value or '').strip() if pc else ''
        rec['plpAll'].append(plp)

        # PLG fee
        pfc = plg_fee_cols[wi] if wi < len(plg_fee_cols) else None
        pfv = safe_float(ws.cell(row_idx, pfc).value) if pfc else 0
        rec['plgFeeAll'].append(round(pfv * 100, 1) if pfv < 1 else pfv)

    sku_data.append(rec)

print(f'Extracted {len(sku_data)} SKUs')

# --- 3. Compute category/analyst totals per week ---
cat_names = sorted(set(r['category'] for r in sku_data))
an_names = sorted(set(r['analyst'] for r in sku_data))

cat_sales_all = []
for cat in cat_names:
    cat_rows = [r for r in sku_data if r['category'] == cat]
    entry = {'category': cat, 'sales': [0.0]*NWEEKS, 'revenue': [0.0]*NWEEKS, 'share': [0.0]*NWEEKS}
    for wi in range(NWEEKS):
        for r in cat_rows:
            entry['sales'][wi] += r['salesAll'][wi]
            entry['revenue'][wi] += r['revenueAll'][wi]
        # Share = cat_sales / (cat_sales + cat_rival)
        cat_s = sum(r['salesAll'][wi] for r in cat_rows)
        cat_r = sum(r['rivalAll'][wi] for r in cat_rows)
        entry['share'][wi] = round(cat_s / (cat_s + cat_r) * 100, 1) if (cat_s + cat_r) > 0 else 0
    cat_sales_all.append(entry)

an_sales_all = []
for an in an_names:
    an_rows = [r for r in sku_data if r['analyst'] == an]
    entry = {'analyst': an, 'sales': [0.0]*NWEEKS, 'revenue': [0.0]*NWEEKS, 'share': [0.0]*NWEEKS}
    for wi in range(NWEEKS):
        for r in an_rows:
            entry['sales'][wi] += r['salesAll'][wi]
            entry['revenue'][wi] += r['revenueAll'][wi]
        an_s = sum(r['salesAll'][wi] for r in an_rows)
        an_r = sum(r['rivalAll'][wi] for r in an_rows)
        entry['share'][wi] = round(an_s / (an_s + an_r) * 100, 1) if (an_s + an_r) > 0 else 0
    an_sales_all.append(entry)

# --- 4. Compute total share per week ---
total_share_all = [0.0] * NWEEKS
for wi in range(NWEEKS):
    ts = total_sales[wi]
    tr = total_rival[wi]
    total_share_all[wi] = round(ts / (ts + tr) * 100, 1) if (ts + tr) > 0 else 0

# Round totals
total_sales_all = [round(x, 1) for x in total_sales]
total_rev_all = [round(x, 2) for x in total_revenue]

# --- 5. Build allPeriods (weeks with >=3 prior weeks) ---
all_periods = []
for i in range(3, NWEEKS):
    all_periods.append({'label': all_week_labels[i], 'index': i})

# --- 6. Read existing JSON and merge ---
with open(TGT, 'r', encoding='utf-8') as f:
    existing = json.load(f)

# Add multi-period data
existing['allWeekLabels'] = all_week_labels
existing['allPeriods'] = all_periods
existing['totalSalesAll'] = total_sales_all
existing['totalRevAll'] = total_rev_all
existing['totalShareAll'] = total_share_all
existing['catSalesAll'] = cat_sales_all
existing['catRevAll'] = []  # placeholder
existing['catShareAll'] = []  # placeholder
existing['anSalesAll'] = an_sales_all
existing['anShareAll'] = []  # placeholder
existing['skuAll'] = sku_data  # All SKU data with per-week arrays

with open(TGT, 'w', encoding='utf-8') as f:
    json.dump(existing, f, ensure_ascii=False, indent=2)

first_label = all_periods[0]['label'] if all_periods else 'N/A'
last_label = all_periods[-1]['label'] if all_periods else 'N/A'
print(f'Updated {TGT}')
print(f'allPeriods: {len(all_periods)} options ({first_label} .. {last_label})')
print(f'totalSalesAll: {total_sales_all}')
