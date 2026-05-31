"""
更新胡煜星-周报统一Workbook表头模板 — 根据源数据表头同步
"""
import os, sys
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook
from copy import copy

THIS_DIR = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(THIS_DIR, '..', '新品复盘', '周报', '新品检查周源数据和PLP数据.xlsx')
TPL = os.path.join(THIS_DIR, '胡煜星-周报统一Workbook表头模板.xlsx')

# 加载源数据和模板
wb_src = load_workbook(SRC, data_only=True)
wb_tpl = load_workbook(TPL)

ws_src = wb_src['四三数据累计']
ws_plp = wb_src['PLP明细']

# 获取源数据表头
src_headers = []
for col in range(1, ws_src.max_column + 1):
    v = ws_src.cell(row=1, column=col).value
    if v:
        src_headers.append((col, str(v).strip()))

print("=== 源数据表头（四三数据累计）===")
for col, h in src_headers:
    print(f"  col {col}: {h}")

print(f"\n=== PLP明细表头 ===")
plp_headers = []
for col in range(1, ws_plp.max_column + 1):
    v = ws_plp.cell(row=1, column=col).value
    if v:
        plp_headers.append((col, str(v).strip()))
        print(f"  col {col}: {v}")

print(f"\n=== 模板现有 Sheet ===")
for sn in wb_tpl.sheetnames:
    ws = wb_tpl[sn]
    print(f"  {sn}: {ws.max_row}r x {ws.max_column}c")

# --- 更新 99_字段说明 Sheet ---
# 这个 Sheet 应该反映最新的字段定义
print(f"\n=== 更新 99_字段说明 ===")
ws_desc = wb_tpl['99_字段说明']

# 清空旧数据（保留表头）
current_rows = ws_desc.max_row
for r in range(2, current_rows + 1):
    for c in range(1, 7):
        ws_desc.cell(row=r, column=c).value = None

# 写入新的字段说明
row = 2
field_defs = [
    # (Sheet, 字段, 必填, 说明, 示例, 对应模块)
    ('四三数据累计', '销售编号', '是', '唯一销售标识', 'LYAP-X1001', '新品复盘'),
    ('四三数据累计', 'SKU', '是', '产品SKU编码', 'LYAP-X1001', '新品复盘'),
    ('四三数据累计', '实际上架日期', '是', '产品实际上架日期 YYYY-MM-DD', '2026-03-04', '新品复盘'),
    ('四三数据累计', '首次出单日期', '否', '首次产生订单日期', '2026-03-12', '新品复盘'),
    ('四三数据累计', '分析人', '是', '负责分析人员（如5月分析人）', '俞东旭', '新品复盘'),
    ('四三数据累计', '品类', '是', '产品品类：车门系统/车身外扩件/挡泥板/机盖及附件/其他/饰条/牌照板支架', '车门系统', '新品复盘'),
    ('四三数据累计', '产品拓展', '是', '拓展类型：原开品/拓展品/组合件', '原开品', '新品复盘'),
    ('四三数据累计', '各周销量 (col 8-18)', '是', '按周统计销量，周期为周四-周三，如 3.5-3.11销量', '5', '新品复盘/三部周报v1'),
    ('四三数据累计', '各周销售额 (col 20-30)', '是', '按周统计销售额(USD)', '420.00', '新品复盘/三部周报v1'),
    ('四三数据累计', '各期对手量 (col 32-42)', '是', '截止各日期的对手出单量', '7', '新品复盘'),
    ('四三数据累计', '各期市占比 (col 43-53)', '是', '截止各日期的市场占有率(0-1小数)', '0.75', '新品复盘'),
    ('四三数据累计', '各期追溯价 (col 54-64)', '否', '截止各日期的追溯价格', '89.99', '新品复盘'),
    ('四三数据累计', '各期8日出单情况 (col 65-75)', '是', 'Y=8日内出单, N=8日外出单, 未出单=真正未出单', 'Y', '新品复盘/三部周报v1'),
    ('四三数据累计', '各期7日频次标签 (col 76-86)', '是', '7日内分析频次标签，异常=超7日未分析', '正常', '新品复盘'),
    ('四三数据累计', '各期新品频次标签 (col 87-97)', '是', '8日内新品分析频次标签，异常=8日内无分析', '正常', '新品复盘'),
    ('四三数据累计', '各期市场状态 (col 98-109)', '否', '正常/竞争无优势/无市场/站外出单/站内无价格优势', '正常', '新品复盘'),
    ('四三数据累计', '各期操作判断 (col 110-120)', '否', '周期性操作判断结论', '继续跟进', '新品复盘'),
    ('四三数据累计', '各期开启PLP (col 121-126)', '否', '是否开启PLP广告', '是', '新品复盘'),
    ('四三数据累计', '各期PLG最高费率 (col 127-131)', '否', 'PLG最高费率(0-1小数)', '0.05', '新品复盘'),
    ('PLP明细', '周期', '是', '广告周期，如 5.11-5.17', '5.11-5.17', '新品复盘'),
    ('PLP明细', '广告活动', '是', '广告活动名称', 'SP-001', '新品复盘'),
    ('PLP明细', 'SKU', '是', '关联SKU', 'LYAP-X1001', '新品复盘'),
    ('PLP明细', '曝光量', '是', '广告曝光量', '5000', '新品复盘'),
    ('PLP明细', '点击量', '是', '广告点击量', '50', '新品复盘'),
    ('PLP明细', '售出数', '是', '广告售出数', '3', '新品复盘'),
    ('PLP明细', '花费', '是', '广告花费(USD)', '15.50', '新品复盘'),
    ('PLP明细', '广告销售额', '是', '广告带来的销售额(USD)', '180.00', '新品复盘'),
    ('PLP明细', '总销售额', '是', 'SKU总销售额(USD)，用于ACOAS计算', '420.00', '新品复盘'),
    ('PLP明细', 'ROAS', '否', '广告支出回报率', '11.61', '新品复盘'),
    ('PLP明细', 'CVR', '否', '转化率', '6.0%', '新品复盘'),
    ('PLP明细', 'CTR', '否', '点击率', '1.0%', '新品复盘'),
    ('PLP明细', 'CPC', '否', '单次点击成本', '0.31', '新品复盘'),
    ('PLP明细', 'CPA', '否', '单次获客成本', '5.17', '新品复盘'),
    ('PLP明细', 'ACOS', '否', '广告销售成本比', '8.61%', '新品复盘'),
    ('PLP明细', 'ACOAS', '否', '广告花费/总销售额', '3.69%', '新品复盘'),
]

for fd in field_defs:
    ws_desc.cell(row=row, column=1, value=fd[0])
    ws_desc.cell(row=row, column=2, value=fd[1])
    ws_desc.cell(row=row, column=3, value=fd[2])
    ws_desc.cell(row=row, column=4, value=fd[3])
    ws_desc.cell(row=row, column=5, value=fd[4])
    ws_desc.cell(row=row, column=6, value=fd[5])
    row += 1

# --- 更新 04_产品明细 表头，增加缺失的关键字段 ---
print(f"\n=== 更新 04_产品明细 表头 ===")
ws_04 = wb_tpl['04_新品明细']
existing_h = {}
for col in range(1, ws_04.max_column + 1):
    v = ws_04.cell(row=1, column=col).value
    if v:
        existing_h[col] = str(v).strip()

print(f"现有字段: {list(existing_h.values())}")

# 需要新增的字段（在源数据中有，但模板缺少的）
new_fields = [
    '市场状态',       # 竞争无优势/正常/无市场等
    'PLG最高费率',    # PLG费率
    '单品广告销售额',  # 广告带来的销售额
    '总销售额(汇总)',  # SKU总销售额
    'ROAS', 'CVR', 'CTR', 'CPC', 'CPA', 'ACOS', 'ACOAS',
]

# 在现有字段后面追加
next_col = max(existing_h.keys()) + 1 if existing_h else 1
for f in new_fields:
    if f not in existing_h.values():
        ws_04.cell(row=1, column=next_col, value=f)
        print(f"  新增 col {next_col}: {f}")
        next_col += 1

# --- 保存 ---
out_path = os.path.join(THIS_DIR, '胡煜星-周报统一Workbook表头模板_已更新.xlsx')
wb_tpl.save(out_path)
print(f"\n✅ 已保存: {out_path}")
print(f"   原文件未修改，新文件: 胡煜星-周报统一Workbook表头模板_已更新.xlsx")
