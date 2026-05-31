# -*- coding: utf-8 -*-
"""生成新品周报 XLSX（4.30-5.6 周期）"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── 加载数据 ──
with open(r'C:\Users\Administrator\Desktop\三部周报v1\New project 2-新品板块已放入\src\modules\newProductStatus\corrected_data.json', 'r', encoding='utf-8') as f:
    D = json.load(f)

wb = Workbook()

# ── 样式常量 ──
HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
SUB_HEADER_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
SUB_HEADER_FONT = Font(name='微软雅黑', bold=True, size=10)
DATA_FONT = Font(name='微软雅黑', size=10)
TOTAL_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
TOTAL_FONT = Font(name='微软雅黑', bold=True, size=10)
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
AMBER_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
TITLE_FONT = Font(name='微软雅黑', bold=True, size=14, color='333333')
SUBTITLE_FONT = Font(name='微软雅黑', size=11, color='666666')


def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, value, fmt=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = align or CENTER
    if fmt:
        cell.number_format = fmt
    return cell


def style_total_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER


def auto_width(ws, max_col, min_w=10, max_w=30):
    for c in range(1, max_col + 1):
        max_len = min_w
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    val = str(cell.value)
                    # CJK chars take ~2 width
                    clen = sum(2 if ord(ch) > 127 else 1 for ch in val)
                    max_len = max(max_len, min(clen + 2, max_w))
        ws.column_dimensions[get_column_letter(c)].width = max_len


def write_table(ws, start_row, headers, rows, col_formats=None, total_row=None):
    """通用表格写入"""
    max_col = len(headers)
    # 标题行
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header(ws, start_row, max_col)

    # 数据行
    for r_idx, row_data in enumerate(rows):
        r = start_row + 1 + r_idx
        for c_idx, val in enumerate(row_data):
            fmt = col_formats[c_idx] if col_formats and c_idx < len(col_formats) else None
            al = RIGHT if isinstance(val, (int, float)) and val is not None else CENTER
            style_data_cell(ws, r, c_idx + 1, val, fmt=fmt, align=al)

    # 合计行
    if total_row:
        r = start_row + 1 + len(rows)
        for c_idx, val in enumerate(total_row):
            style_data_cell(ws, r, c_idx + 1, val)
        style_total_row(ws, r, max_col)

    return start_row + 1 + len(rows) + (1 if total_row else 0)


def add_title(ws, title, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
        return 4
    return 3


def add_section_title(ws, row, text, max_col=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name='微软雅黑', bold=True, size=12, color='4472C4')
    cell.alignment = LEFT
    return row + 1


# ================================================================
# Sheet 1: 新品总览 KPI
# ================================================================
ws1 = wb.active
ws1.title = '新品总览'
r = add_title(ws1, '新品周报总览', '数据周期：2026/4/30 - 5/6（周四至周三）')

# KPI 核心指标
s = D['cum43Stats']
prev = D['prevWeekKpi']
r = add_section_title(ws1, r, '核心 KPI 指标')

kpi_headers = ['指标', '本周', '上周', '环比变化', '说明']
kpi_rows = [
    ['在售新品 SKU', s['total'], prev['prevTotalSku'], f"+{prev['skuChange']}%" if prev.get('skuChange') else '--', '四三累计在售新品总数'],
    ['本周出单 SKU', s['yCount'], prev.get('prevSoldCount', '--'), '--', '8日内出单的新品'],
    ['8日后出单 SKU', s['nCount'], '--', '--', '超过8天才出单的新品'],
    ['未出单 SKU', s['unCount'], '--', '--', '上架后至今未出单'],
    ['总销量', f"{sum(d['4.30-5.6销量'] for d in D['cum43Data'])}", prev['prevTotalSalesQty'],
     f"+{prev['salesQtyChange']}%" if prev.get('salesQtyChange') else '--', '本周所有新品总销量'],
    ['总销售额 ($)', f"${sum(d['4.30-5.6销售额'] for d in D['cum43Data']):,.2f}",
     f"${prev['prevTotalRevenue']}",
     f"+{prev['revenueChange']}%" if prev.get('revenueChange') else '--', 'USD 原值'],
    ['低占比新品', len(D['lowShareData']), prev.get('prevLowShareCount', '--'), '--',
     '市场份额<0.75%且存在竞品订单'],
]

# 及时率
td = D['timelinessData']
timely_total = td.get('total', {})
kpi_rows.append(['及时率', f"{timely_total.get('timelyRate', '--')}%",
                 f"{prev.get('prevTimelyRate', '--')}%",
                 f"+{prev.get('timelyChange', '--')}%" if prev.get('timelyChange') else '--',
                 '8日内出单/总出单SKU'])

# PLG
plg = D['plgStats']
kpi_rows.append(['PLP开启数', plg.get('plpEnabledCount', '--'), '--', '--', 'PLP 广告开启中'])
kpi_rows.append(['PLG开启数', plg.get('plpAndPlgBothCount', '--'), '--', '--', 'PLP+PLG 同时开启'])

fmts = [None, None, None, None, None]
r = write_table(ws1, r, kpi_headers, kpi_rows, col_formats=fmts)
auto_width(ws1, len(kpi_headers))


# ================================================================
# Sheet 2: 新品明细
# ================================================================
ws2 = wb.create_sheet('新品明细')
r = add_title(ws2, '新品明细清单', '数据周期：2026/4/30 - 5/6 | 共101个SKU')

cum_headers = ['SKU', '实际上架日期', '首次出单日期', '分析人', '品类', '产品拓展',
               '本周销量', '本周销售额($)', '对手销量', '市占比(%)', '市场状态', '8日出单']
cum_rows = []
for d in D['cum43Data']:
    cum_rows.append([
        d['SKU'], d['实际上架日期'], d['首次出单日期'], d['4月分析人'], d['品类'],
        d['产品拓展'], d['4.30-5.6销量'], d['4.30-5.6销售额'],
        d['5.6对手销量'], d['5.6市占比'], d['5.6市场状态'], d['5.6 8日出单情况']
    ])

# 合计行
total_sales = sum(d['4.30-5.6销量'] for d in D['cum43Data'])
total_revenue = sum(d['4.30-5.6销售额'] for d in D['cum43Data'])
cum_total = ['合计', '', '', '', '', '', total_sales, f"${total_revenue:,.2f}", '', '', '', '']

cum_fmts = [None, None, None, None, None, None, '#,##0', '$#,##0.00', '#,##0', '0.0', None, None]
r = write_table(ws2, r, cum_headers, cum_rows, col_formats=cum_fmts, total_row=cum_total)

# 市场状态着色
for row in ws2.iter_rows(min_row=r - len(cum_rows) - 1, max_row=r - 2, min_col=11, max_col=11):
    for cell in row:
        if cell.value == '竞争无优势':
            cell.fill = RED_FILL
        elif cell.value == '正常':
            cell.fill = GREEN_FILL

auto_width(ws2, len(cum_headers))


# ================================================================
# Sheet 3: 低占比新品
# ================================================================
ws3 = wb.create_sheet('低占比新品')
r = add_title(ws3, '低占比新品详情', f'共{len(D["lowShareData"])}个 SKU | 市场份额<0.75%且存在竞品订单')

ls_headers = ['销售代码', 'SKU', '上架日期', '分析人', '品类', '拓展类型',
              '本周销量', '销量变化', '本周销售额($)', '销售额变化',
              '上周期手销量', '本周期手销量', '周期手变化',
              '上周市占比(%)', '本周市占比(%)', '市占比变化',
              '8日出单', '出单频率', '上周市场状态', '操作建议', '本周市场状态', 'PLP开启', 'PLG费率']
ls_rows = []
for d in D['lowShareData']:
    ls_rows.append([
        d['salesCode'], d['sku'], d['launchDate'], d['analyst'], d['category'],
        d['expandType'], d['curSalesQty'], d['salesQtyChange'],
        d['curRevenue'], d['revenueChange'],
        d['prevCompetitorQty'], d['curCompetitorQty'], d['competitorQtyChange'],
        d['prevMarketShare'], d['curMarketShare'], d['marketShareChange'],
        d['cur8dStatus'], d['cur7dFreqTag'], d['prevMarketStatus'],
        d['curOperation'], d['curMarketStatus'], d['plpEnabled'], d['plgFee']
    ])

ls_fmts = [None, None, None, None, None, None, '#,##0', None, '$#,##0.00', None,
           '#,##0', '#,##0', None, '0.0', '0.0', None, None, None, None, None, None, None, None]
r = write_table(ws3, r, ls_headers, ls_rows, col_formats=ls_fmts)
auto_width(ws3, len(ls_headers), min_w=10, max_w=18)


# ================================================================
# Sheet 4: 品类汇总
# ================================================================
ws4 = wb.create_sheet('品类汇总')
r = add_title(ws4, '品类维度汇总', '按品类汇总新品表现')

cat_headers = ['品类', '在售SKU', '新增SKU', '本周销量', '上周销量', '销量变化',
               '本周销售额($)', '上周销售额($)', '销售额变化', '有竞品SKU']
cat_rows = []
for d in D['categoryRevenueData']:
    cat_rows.append([
        d['category'], d['curSku'], d['curNewSku'],
        d['curSalesQty'], d['prevSalesQty'], d['salesQtyChange'],
        d['curRevenue'], d['prevRevenue'], d['revenueChange'],
        d.get('curHasCompetitor', '--')
    ])

cat_total = ['合计',
             sum(d['curSku'] for d in D['categoryRevenueData']),
             sum(d['curNewSku'] for d in D['categoryRevenueData']),
             sum(d['curSalesQty'] for d in D['categoryRevenueData']),
             sum(d['prevSalesQty'] for d in D['categoryRevenueData']),
             '--',
             sum(d['curRevenue'] for d in D['categoryRevenueData']),
             sum(d['prevRevenue'] for d in D['categoryRevenueData']),
             '--',
             sum(d.get('curHasCompetitor', 0) for d in D['categoryRevenueData'] if isinstance(d.get('curHasCompetitor'), (int, float)))]
cat_fmts = [None, '#,##0', '#,##0', '#,##0', '#,##0', None, '$#,##0.00', '$#,##0.00', None, '#,##0']
r = write_table(ws4, r, cat_headers, cat_rows, col_formats=cat_fmts, total_row=cat_total)
auto_width(ws4, len(cat_headers))


# ================================================================
# Sheet 5: 分析人汇总
# ================================================================
ws5 = wb.create_sheet('分析人汇总')
r = add_title(ws5, '分析人维度汇总', '按分析人汇总新品表现')

an_headers = ['分析人', '在售SKU', '新增SKU', '本周销量', '上周销量', '销量变化',
              '本周销售额($)', '上周销售额($)', '销售额变化']
an_rows = []
for d in D['analystRevenueData']:
    an_rows.append([
        d['analyst'], d['curSku'], d['curNewSku'],
        d['curSalesQty'], d['prevSalesQty'], d['salesQtyChange'],
        d['curRevenue'], d['prevRevenue'], d['revenueChange']
    ])

an_total = ['合计',
            sum(d['curSku'] for d in D['analystRevenueData']),
            sum(d['curNewSku'] for d in D['analystRevenueData']),
            sum(d['curSalesQty'] for d in D['analystRevenueData']),
            sum(d['prevSalesQty'] for d in D['analystRevenueData']),
            '--',
            sum(d['curRevenue'] for d in D['analystRevenueData']),
            sum(d['prevRevenue'] for d in D['analystRevenueData']),
            '--']
an_fmts = [None, '#,##0', '#,##0', '#,##0', '#,##0', None, '$#,##0.00', '$#,##0.00', None]
r = write_table(ws5, r, an_headers, an_rows, col_formats=an_fmts, total_row=an_total)
auto_width(ws5, len(an_headers))


# ================================================================
# Sheet 6: 广告概览
# ================================================================
ws6 = wb.create_sheet('广告概览')
r = add_title(ws6, '广告投放概览', '数据周期：2026/5/5 - 5/11（自然周）')

t = D['plpTotal']
p = D['plpPrevTotal']

r = add_section_title(ws6, r, '核心广告指标')
ad_headers = ['指标', '本周', '上周', '变化']

def pct_or_raw(v):
    """将百分比字符串转为数字，或原样返回"""
    if isinstance(v, str) and '%' in v:
        return float(v.replace('%', ''))
    return v

def pct_change(cur, prev):
    if isinstance(cur, str) and '%' in cur:
        cur = float(cur.replace('%', ''))
    if isinstance(prev, str) and '%' in prev:
        prev = float(prev.replace('%', ''))
    if isinstance(cur, (int, float)) and isinstance(prev, (int, float)) and prev != 0:
        return f"{((cur - prev) / abs(prev) * 100):+.1f}%"
    return '--'

def fmt_ad_val(v):
    """格式化广告指标值"""
    if isinstance(v, str) and '%' in v:
        return v
    if isinstance(v, float) and v > 1000:
        return f"${v:,.2f}"
    if isinstance(v, float):
        return f"{v:.2f}"
    return v

ad_map = [
    ('广告活动数', 'campaignCount'), ('广告链接数', 'linkCount'),
    ('曝光量', 'impression'), ('点击量', 'click'),
    ('销量(出单)', 'sold'), ('花费($)', 'cost'), ('广告销售额($)', 'revenue'),
    ('ROAS', 'roas'), ('转化率', 'cvr'), ('点击率', 'ctr'),
    ('CPC($)', 'cpc'), ('CPA($)', 'cpa'), ('ACOS', 'acos'), ('ACOAS', 'acoas'),
]

ad_rows = []
for label, key in ad_map:
    cv = t.get(key, '--')
    pv = p.get(key, '--')
    change = pct_change(cv, pv) if cv != '--' and pv != '--' else '--'
    ad_rows.append([label, fmt_ad_val(cv), fmt_ad_val(pv), change])

r = write_table(ws6, r, ad_headers, ad_rows)
auto_width(ws6, len(ad_headers))


# ================================================================
# Sheet 7: 广告分析人维度
# ================================================================
ws7 = wb.create_sheet('广告分析人')
r = add_title(ws7, '广告 - 分析人维度', '按分析人汇总广告投放数据')

anl_headers = ['分析人', '曝光量', '点击量', '销量', '花费($)', '广告销售额($)',
               'ROAS', '转化率', '点击率', 'CPC($)', 'CPA($)', 'ACOS', 'ACOAS']
anl_rows = []
for d in D['plpAnalysts']:
    anl_rows.append([
        d['name'], d['impression'], d['click'], d['sold'],
        d['cost'], d['revenue'], d['roas'], d['cvr'], d['ctr'],
        d['cpc'], d['cpa'], d['acos'], d['acoas']
    ])

# 合计
anl_total = ['合计',
             sum(d['impression'] for d in D['plpAnalysts']),
             sum(d['click'] for d in D['plpAnalysts']),
             sum(d['sold'] for d in D['plpAnalysts']),
             sum(d['cost'] for d in D['plpAnalysts']),
             sum(d['revenue'] for d in D['plpAnalysts']),
             '--', '--', '--', '--', '--', '--', '--']
anl_fmts = [None, '#,##0', '#,##0', '#,##0', '$#,##0.00', '$#,##0.00',
            None, None, None, '$#,##0.00', '$#,##0.00', None, None]
r = write_table(ws7, r, anl_headers, anl_rows, col_formats=anl_fmts, total_row=anl_total)
auto_width(ws7, len(anl_headers))


# ================================================================
# Sheet 8: 广告品线维度
# ================================================================
ws8 = wb.create_sheet('广告品线')
r = add_title(ws8, '广告 - 品线维度', '按品类汇总广告投放数据')

cat_headers = ['品类', '活动数', '链接数', '曝光量', '点击量', '销量', '花费($)',
               '广告销售额($)', 'ROAS', '转化率', '点击率', 'CPC($)', 'CPA($)', 'ACOS', 'ACOAS']
cat_rows = []
for d in D['plpCategories']:
    cat_rows.append([
        d['category'], d['campaignCount'], d['linkCount'],
        d['impression'], d['click'], d['sold'],
        d['cost'], d['revenue'], d['roas'], d['cvr'], d['ctr'],
        d['cpc'], d['cpa'], d['acos'], d['acoas']
    ])

cat_total = ['合计',
             sum(d['campaignCount'] for d in D['plpCategories']),
             sum(d['linkCount'] for d in D['plpCategories']),
             sum(d['impression'] for d in D['plpCategories']),
             sum(d['click'] for d in D['plpCategories']),
             sum(d['sold'] for d in D['plpCategories']),
             sum(d['cost'] for d in D['plpCategories']),
             sum(d['revenue'] for d in D['plpCategories']),
             '--', '--', '--', '--', '--', '--', '--']
cat_fmts = [None, '#,##0', '#,##0', '#,##0', '#,##0', '#,##0', '$#,##0.00',
            '$#,##0.00', None, None, None, '$#,##0.00', '$#,##0.00', None, None]
r = write_table(ws8, r, cat_headers, cat_rows, col_formats=cat_fmts, total_row=cat_total)
auto_width(ws8, len(cat_headers))


# ================================================================
# Sheet 9: 广告明细
# ================================================================
ws9 = wb.create_sheet('广告明细')
r = add_title(ws9, '广告投放明细', f'共{len(D["plpDetailData"])}条广告活动记录')

# SKU 去重计算 ACOAS
sku_rev_map = {}
for d in D['plpDetailData']:
    if d['sku'] not in sku_rev_map:
        sku_rev_map[d['sku']] = d.get('totalRevenue', 0)
dedup_total_revenue = sum(sku_rev_map.values())

det_headers = ['周期', '广告活动', 'SKU', 'ID', '店铺', 'PLP开始日期', '实际上架日期',
               '首单日期', '分析人', '品类', '产品拓展',
               '曝光量', '点击量', '销量', '花费($)', '广告销售额($)', '总销售额($)',
               'ROAS', '转化率', '点击率', 'CPC($)', 'CPA($)', 'ACOS', 'ACOAS']
det_rows = []
for d in D['plpDetailData']:
    acoas_val = d.get('acoas', '')
    # acoas 可能是小数(0.0423)或字符串("4.23%")
    if isinstance(acoas_val, (int, float)) and acoas_val < 1:
        acoas_display = f"{acoas_val * 100:.2f}%"
    elif isinstance(acoas_val, str):
        acoas_display = acoas_val
    else:
        acoas_display = acoas_val

    acos_val = d.get('acos', '')
    if isinstance(acos_val, (int, float)) and acos_val < 1:
        acos_display = f"{acos_val * 100:.2f}%"
    elif isinstance(acos_val, str):
        acos_display = acos_val
    else:
        acos_display = acos_val

    det_rows.append([
        d['period'], d['campaign'], d['sku'], d['id'], d['store'],
        d['plpStartDate'], d['actualListDate'], d['firstOrderDate'],
        d['analyst'], d['category'], d['productExpansion'],
        d['impressions'], d['clicks'], d['salesQty'],
        d['spend'], d['adRevenue'], d['totalRevenue'],
        d['roas'], d['cvr'], d['ctr'], d['cpc'], d['cpa'],
        acos_display, acoas_display
    ])

det_fmts = [None, None, None, None, None, None, None, None, None, None, None,
            '#,##0', '#,##0', '#,##0', '$#,##0.00', '$#,##0.00', '$#,##0.00',
            None, None, None, '$#,##0.00', '$#,##0.00', None, None]
r = write_table(ws9, r, det_headers, det_rows, col_formats=det_fmts)
auto_width(ws9, len(det_headers), min_w=10, max_w=18)


# ================================================================
# Sheet 10: 拓展类型
# ================================================================
ws10 = wb.create_sheet('拓展类型')
r = add_title(ws10, '拓展类型分析', '新品 vs 拓展品表现对比')

# 新品维度
r = add_section_title(ws10, r, '新品维度 - 拓展类型对比')
et_headers = ['拓展类型', '在售SKU', '上周SKU', '出单SKU数', '出单率', '上周出单SKU', '上周出单率',
              '出单率变化', '本周销量', '上周销量', '销量变化', '本周销售额($)', '上周销售额($)', '销售额变化']
et_rows = []
for d in D['expandTypeData']:
    et_rows.append([
        d['expandType'], d['curSku'], d['prevSku'],
        d['curSalesCount'], d['curSalesRate'], d['prevSalesCount'], d['prevSalesRate'],
        d['salesRateChange'], d['curSalesQty'], d['prevSalesQty'], d['salesQtyChange'],
        d['curRevenue'], d['prevRevenue'], d['revenueChange']
    ])
r = write_table(ws10, r, et_headers, et_rows)

# 广告维度
r = add_section_title(ws10, r + 1, '广告维度 - 拓展类型对比')
pet_headers = ['拓展类型', '活动数', '链接数', '曝光量', '点击量', '销量', '花费($)',
               '广告销售额($)', 'ROAS', '转化率', '点击率', 'CPC($)', 'CPA($)', 'ACOS', 'ACOAS']
pet_rows = []
for d in D['plpExpandTypes']:
    pet_rows.append([
        d['expandType'], d['campaignCount'], d['linkCount'],
        d['impression'], d['click'], d['sold'],
        d['cost'], d['revenue'], d['roas'], d['cvr'], d['ctr'],
        d['cpc'], d['cpa'], d['acos'], d['acoas']
    ])
r = write_table(ws10, r, pet_headers, pet_rows)
auto_width(ws10, max(len(et_headers), len(pet_headers)))


# ================================================================
# Sheet 11: PLG 统计
# ================================================================
ws11 = wb.create_sheet('PLG统计')
r = add_title(ws11, 'PLG 费率统计', 'PLP/PLG 开启状态分布')

plg = D['plgStats']
r = add_section_title(ws11, r, 'PLP/PLG 开启概况')

pg_headers = ['指标', '数量']
pg_rows = [
    ['PLP+PLG 同时开启', plg.get('plpAndPlgBothCount', '--')],
    ['仅 PLG 开启', plg.get('plgOnlyCount', '--')],
    ['仅 PLP 开启', plg.get('plpOnlyCount', '--')],
    ['无广告', plg.get('noAdCount', '--')],
    ['PLP 开启总数', plg.get('plpEnabledCount', '--')],
    ['PLP 关闭无出单', plg.get('plpDisabledNoSaleCount', '--')],
    ['新品总数', plg.get('totalNewProducts', '--')],
]
r = write_table(ws11, r, pg_headers, pg_rows)

# PLG 明细
r = add_section_title(ws11, r + 1, 'PLG 明细记录')
plg_det_headers = ['销售代码', 'SKU', '上架日期', '首单日期', '分析人', '品类', '拓展类型',
                   '8日出单', '本周销量', '本周销售额($)', '周期手销量', '市占比(%)',
                   '市场状态', '操作建议', 'PLP开启', 'PLG费率']
plg_det_rows = []
for d in D['plgRecords']:
    plg_det_rows.append([
        d['salesCode'], d['sku'], d['launchDate'], d['firstSaleDate'],
        d['analyst'], d['category'], d['expandType'],
        d['cur8dStatus'], d['curSalesQty'], d['curRevenue'],
        d['curCompetitorQty'], d['curMarketShare'],
        d['curMarketStatus'], d['curOperation'], d['plpEnabled'], d['plgFee']
    ])
plg_fmts = [None, None, None, None, None, None, None, None, '#,##0', '$#,##0.00',
            '#,##0', '0.0', None, None, None, None]
r = write_table(ws11, r, plg_det_headers, plg_det_rows, col_formats=plg_fmts)
auto_width(ws11, len(plg_det_headers), min_w=10, max_w=18)


# ================================================================
# Sheet 12: 及时率
# ================================================================
ws12 = wb.create_sheet('及时率')
r = add_title(ws12, '出单及时率分析', '按分析人统计8日内出单率')

td = D['timelinessData']
tm_headers = ['分析人', '在售SKU', '及时出单SKU', '未分析8天', '未分析7天', '及时率(%)', '上周及时率(%)', '环比变化']
tm_rows = []
for d in td.get('analysts', []):
    tm_rows.append([
        d['analyst'], d.get('curSku', '--'),
        d.get('timelyCount', '--'), d.get('noAnalysis8dCount', '--'), d.get('noAnalysis7dCount', '--'),
        d.get('timelyRate', '--'), d.get('prevTimelyRate', '--'), d.get('change', '--')
    ])

# 总计
tt = td.get('total', {})
tm_rows.append([
    '合计', tt.get('curSku', '--'),
    tt.get('timelyCount', '--'), tt.get('noAnalysis8dCount', '--'), tt.get('noAnalysis7dCount', '--'),
    tt.get('timelyRate', '--'), tt.get('prevTimelyRate', '--'), tt.get('change', '--')
])

r = write_table(ws12, r, tm_headers, tm_rows)
# 给总计行加粗
style_total_row(ws12, r - 1, len(tm_headers))
auto_width(ws12, len(tm_headers))


# ================================================================
# 保存
# ================================================================
output_path = r'C:\Users\Administrator\Desktop\三部周报v1\新品周报_4.30-5.6.xlsx'
wb.save(output_path)
print(f'XLSX 已保存: {output_path}')
print(f'Sheet 数量: {len(wb.sheetnames)}')
print(f'Sheet 列表: {wb.sheetnames}')
